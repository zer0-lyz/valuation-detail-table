"""
文件检测器
==========
自动识别：1) 文档类型（科目余额表/序时账/资产负债表...）
         2) Sheet 名称
         3) 表头行位置
         4) 数据起始行
         5) 列名 → 标准字段的候选映射
"""

import re
import json
import pandas as pd
import openpyxl
from pathlib import Path
from difflib import SequenceMatcher
from core import config as cfg


def _load_column_mappings():
    """加载列名变体映射表"""
    mappings_path = Path(__file__).parent / "column_mappings.json"
    with open(mappings_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _find_sheet_name(excel_path: str, doc_type_keywords: list) -> str:
    """根据关键词匹配最适合的 Sheet"""
    xls = pd.ExcelFile(excel_path)
    for sheet in xls.sheet_names:
        for kw in doc_type_keywords:
            if kw.lower() in sheet.lower():
                return sheet
    return xls.sheet_names[0]


def _detect_header_row(df: pd.DataFrame) -> int:
    """智能检测表头行位置"""
    all_col_mappings = _load_column_mappings()
    all_variants = set()
    for dtype_key, doc_type in all_col_mappings.items():
        if dtype_key.startswith('_'):
            continue
        for field_config in doc_type.values():
            for v in field_config.get("variants", []):
                all_variants.add(v.lower())

    best_row = 0
    best_score = 0
    for i in range(min(15, len(df))):
        row = df.iloc[i].astype(str).tolist()
        match_count = sum(
            1 for cell in row
            if any(v in cell.lower().strip() for v in all_variants)
        )
        digit_count = sum(1 for cell in row if re.match(r'^[\d,.\-]+$', cell.strip()))
        score = match_count - digit_count * 0.5
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _is_two_row_header(excel_path: str, sheet_name: str, header_row: int) -> bool:
    """检测是否为双行表头（如 Row5='期初余额', Row6='借方|贷方'）

    用 openpyxl 直接读两行，检查第二行是否包含方向关键词。
    """
    DIRECTION_KEYWORDS = {'借方', '贷方', '借', '贷', 'debit', 'credit'}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    child_row = header_row + 2  # openpyxl is 1-based, header_row is 0-based

    if child_row > ws.max_row:
        wb.close()
        return False

    direction_count = 0
    total_cells = 0
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=child_row, column=col).value
        if val:
            val_str = str(val).strip()
            total_cells += 1
            if val_str in DIRECTION_KEYWORDS or any(kw in val_str for kw in DIRECTION_KEYWORDS):
                direction_count += 1

    wb.close()
    # 至少有两个方向关键词才算双行表头
    return direction_count >= 2


def _merge_two_row_header(excel_path: str, sheet_name: str, header_row: int) -> list:
    """合并双行表头：将父行（期初余额/本期发生/期末余额）与子行（借方/贷方）合并。

    例如：
      父行: ['会计期间', '科目', '', '币种名称', '期初余额', '', '本期发生', '', '期末余额', '']
      子行: ['', '编码', '名称', '', '借方', '贷方', '借方', '贷方', '借方', '贷方']
      →
      合并: ['会计期间', '科目编码', '科目名称', '币种名称', '期初余额借方', '期初余额贷方',
             '本期发生借方', '本期发生贷方', '期末余额借方', '期末余额贷方']
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    parent_row_num = header_row + 1  # 0-based → 1-based openpyxl
    child_row_num = parent_row_num + 1

    max_col = ws.max_column

    # 读取两行
    parent_vals = []
    child_vals = []
    for col in range(1, max_col + 1):
        pv = ws.cell(row=parent_row_num, column=col).value
        cv = ws.cell(row=child_row_num, column=col).value
        parent_vals.append(str(pv).strip() if pv else '')
        child_vals.append(str(cv).strip() if cv else '')

    wb.close()

    # 扩展合并单元格：如果父行某列为空，用左侧非空值填充
    expanded_parent = []
    last_parent = ''
    for pv in parent_vals:
        if pv:
            last_parent = pv
        expanded_parent.append(last_parent)

    # 合并
    DIRECTION_WORDS = {'借方', '贷方', '借', '贷', 'debit', 'credit'}
    merged = []
    for i in range(len(parent_vals)):
        p = expanded_parent[i]
        c = child_vals[i]

        if not p and not c:
            merged.append('')
        elif not p:
            merged.append(c)
        elif not c:
            merged.append(p)
        elif c in DIRECTION_WORDS or any(kw in c for kw in DIRECTION_WORDS):
            # 子行是方向词 → 合并为 "父行+子行" (如"期初余额借方")
            merged.append(p + c)
        else:
            # 子行不是方向词 → 子行可能是独立的列名（如"编码"/"名称"）
            # 如果父行和子行不同，优先用子行（更具体）
            if p != c and c:
                merged.append(c)
            else:
                merged.append(p)

    return merged


def _match_column(source_col: str, col_mappings: dict) -> str:
    """
    将源列名匹配到标准字段名。
    优先级：完全匹配 > 最长包含匹配 > 相似度匹配。
    避免短关键词（如"借方"）错误覆盖长关键词（如"期末借方"）。
    """
    source_clean = source_col.strip().lower()

    # 收集所有候选项
    candidates = []  # [(field_name, variant, match_type, score)]

    for field_name, field_config in col_mappings.items():
        for variant in field_config.get("variants", []):
            v = variant.lower()
            
            # 1. 完全匹配
            if source_clean == v:
                candidates.append((field_name, v, "exact", 1000 + len(v)))
                continue
            
            # 2. 包含匹配（用较长字符串包含较短字符串）
            if len(source_clean) >= len(v):
                if v in source_clean:
                    candidates.append((field_name, v, "contains", len(v)))
            else:
                if source_clean in v:
                    candidates.append((field_name, v, "contained_by", len(source_clean)))

    # 如果有精确或包含匹配，选择得分最高的（最长变体优先）
    if candidates:
        # 按分数降序排列
        candidates.sort(key=lambda x: -x[3])
        best = candidates[0]
        return best[0]

    # 3. 相似度匹配（无文本匹配时才用）
    best_field = None
    best_ratio = 0.6
    for field_name, field_config in col_mappings.items():
        for variant in field_config.get("variants", []):
            ratio = SequenceMatcher(None, source_clean, variant.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_field = field_name

    return best_field


def _detect_doc_type(file_name: str, headers: list, all_sheets: list) -> str:
    """检测文档类型"""
    from core import schemas
    file_lower = file_name.lower()
    best_type = None
    best_score = 0

    for dt_key, dt_info in schemas.DOCUMENT_TYPES.items():
        score = 0
        for kw in dt_info.get("keywords", []):
            if kw.lower() in file_lower:
                score += 3
        for sheet in all_sheets:
            for pat in dt_info.get("sheet_pattern", []):
                if pat.lower() in sheet.lower():
                    score += 2
        header_text = " ".join(str(h).lower() for h in headers)
        for f_info in dt_info["schema"].values():
            if f_info.get("label", "").lower() in header_text:
                score += 1
        if score > best_score:
            best_score = score
            best_type = dt_key
    return best_type


def guess(excel_path: str) -> dict:
    """对 Excel 文件进行 Guess，生成映射配置。

    支持双行表头自动合并（用友/金蝶导出格式）。
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {excel_path}")
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx/.xls")

    all_col_mappings = _load_column_mappings()
    xls = pd.ExcelFile(excel_path)

    doc_type = _detect_doc_type(path.name, [], xls.sheet_names)
    if not doc_type:
        doc_type = "trial_balance"

    from core import schemas
    doc_info = schemas.DOCUMENT_TYPES.get(doc_type, {})
    sheet_name = _find_sheet_name(excel_path, doc_info.get("sheet_pattern", []))
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    header_row = _detect_header_row(df)

    # 检测双行表头并合并
    # 尝试两种模式：
    #   A) header_row 是父行（期初余额/本期发生），header_row+1 是子行（借方/贷方）
    #   B) header_row 是子行，header_row-1 是父行（常见于 detector 误判）
    data_start_row = header_row + 1
    is_two_row = False

    if _is_two_row_header(excel_path, sheet_name, header_row):
        # 模式A：header_row是父行
        is_two_row = True
        parent_row = header_row
        headers = _merge_two_row_header(excel_path, sheet_name, header_row)
        data_start_row = header_row + 2
        print(f"  📐 检测到双行表头 (父行={parent_row+1})，已合并")
    elif header_row > 0 and _is_two_row_header(excel_path, sheet_name, header_row - 1):
        # 模式B：header_row-1是父行，header_row被误选为子行
        is_two_row = True
        parent_row = header_row - 1
        headers = _merge_two_row_header(excel_path, sheet_name, header_row - 1)
        header_row = parent_row
        data_start_row = parent_row + 2
        print(f"  📐 检测到双行表头 (父行={parent_row+1}, 子行={parent_row+2})，已纠正并合并")

    if not is_two_row:
        headers = df.iloc[header_row].astype(str).tolist()
        headers = [h.strip() for h in headers]

    # 重新检测 doc_type（以文件名检测为准，表头检测仅作为补充）
    header_doc_type = _detect_doc_type(path.name, headers, xls.sheet_names)
    if header_doc_type and not doc_type:
        doc_type = header_doc_type
    # 如果文件名已明确识别（初始 doc_type 非空），不覆盖
    col_mappings_for_type = all_col_mappings.get(doc_type, {})

    column_mapping = {}
    for i, header in enumerate(headers):
        if not header or header.lower() in ("nan", "none", ""):
            continue
        matched_field = _match_column(header, col_mappings_for_type)
        if matched_field:
            column_mapping[header] = matched_field
        else:
            column_mapping[header] = f"__unknown_col_{i}__"

    mapping = cfg.create_blank_mapping(str(path), doc_type)
    mapping["sheet_name"] = sheet_name
    mapping["header_row"] = int(header_row)
    mapping["data_start_row"] = int(data_start_row)
    mapping["column_mapping"] = column_mapping

    # 双行表头：额外存储 column_index_map 供 mapper 直接使用
    if is_two_row:
        mapping["_two_row_header"] = True
        # 构建 {标准字段名: 列索引} 的映射
        col_index_map = {}
        for i, header in enumerate(headers):
            if header and header.lower() not in ("nan", "none", ""):
                matched_field = column_mapping.get(header, '')
                if matched_field and not matched_field.startswith("__unknown"):
                    col_index_map[matched_field] = i
        mapping["_column_index_map"] = col_index_map
    else:
        mapping["_two_row_header"] = False

    return mapping
