#!/usr/bin/env python3
"""
normalize_input.py — 财务资料标准化桥接模块 (DT-NORM)

整合 financial-normalizer 与 valuation-detail-table 流程。

功能:
  1. 自动发现项目文件夹中的财务文件（科目余额表、资产负债表等）
  2. 调用 financial-normalizer 的 guess+apply 管道进行标准化
  3. 将标准化输出转换为 valuation-detail-table 的 Phase 0 缓存格式
  4. 保存到 _dt_cache/ 目录，供 Phase 0 优先使用

用法:
    python normalize_input.py --project <项目文件夹路径> [--force]

设计原则:
  - 标准化数据作为 Phase 0 主数据源，原始解析作为降级备用
  - 遵循 DT-130 中间数据持久化规范
  - 输出格式与 source_header_parser.py 兼容

v1.0 (2026-05-29): 初始版本
"""

import argparse
import glob
import json
import os
import re
import sys
import math
import traceback
from datetime import datetime
from pathlib import Path


# ════════════════════════════════════════════════════════════
# 路径配置
# ════════════════════════════════════════════════════════════

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent

# financial-normalizer 路径（相对于项目根目录的兄弟路径）
# Go up: scripts/ -> valuation-detail-table/ -> .../评估明细表填写/ -> financial-normalizer/
NORMALIZER_PATH = Path(__file__).resolve().parent.parent.parent.parent / 'financial-normalizer'
if NORMALIZER_PATH.exists():
    sys.path.insert(0, str(NORMALIZER_PATH))
else:
    print(f'[NORM] ⚠️ financial-normalizer 未找到: {NORMALIZER_PATH}')
    print(f'[NORM] 将使用 valuation-detail-table 内置解析器作为降级方案')


# ════════════════════════════════════════════════════════════
# 文件发现与类型识别
# ════════════════════════════════════════════════════════════

# 文件名关键词 → 文档类型映射
FILE_TYPE_PATTERNS = {
    'trial_balance': [
        '科目余额', '余额表', '试算平衡', '科目汇总', 'trial balance',
    ],
    'balance_sheet': [
        '资产负债表', '财务报表', 'balance sheet',
    ],
    'journal': [
        '序时账', '明细账', '凭证一览表', '日记账', 'journal',
    ],
    'fixed_asset': [
        '固定资产', '资产台账', '资产卡片', 'fixed asset',
    ],
    'income_statement': [
        '利润表', '损益表', 'income statement', '利润及利润分配',
    ],
}


def detect_file_type(filename: str):
    """根据文件名关键词检测文档类型"""
    name_lower = filename.lower().replace(' ', '')
    for doc_type, patterns in FILE_TYPE_PATTERNS.items():
        for pat in patterns:
            if pat.lower().replace(' ', '') in name_lower:
                return doc_type
    return None


def discover_financial_files(project_dir: str):
    """扫描项目文件夹，发现各类财务文件

    Args:
        project_dir: 项目文件夹路径

    Returns:
        dict: {doc_type: [filepath, ...]}
    """
    found = {}
    exclude_keywords = [
        '评估明细表', '评估说明', '底稿', '抽凭',
        '_dt_cache', 'output', 'mapping_configs', '标准化',
    ]

    # 搜索所有 xlsx 文件
    xlsx_files = glob.glob(os.path.join(project_dir, '**', '*.xlsx'), recursive=True)
    xlsx_files += glob.glob(os.path.join(project_dir, '**', '*.xls'), recursive=True)

    for fp in xlsx_files:
        basename = os.path.basename(fp)
        # 跳过排除文件
        if any(kw in basename for kw in exclude_keywords):
            continue
        doc_type = detect_file_type(basename)
        if doc_type:
            found.setdefault(doc_type, []).append(fp)
            print(f'  [DISCOVER] {doc_type:20s} ← {basename}')

    return found


# ════════════════════════════════════════════════════════════
# 标准化执行器
# ════════════════════════════════════════════════════════════

def _ensure_normalizer_imports():
    """确保 financial-normalizer 核心模块可导入"""
    try:
        from core import detector, mapper, config, schemas
        return True
    except ImportError:
        return False


def normalize_trial_balance(filepath: str, cache_dir: str, force: bool = False) -> dict:
    """标准化科目余额表

    Args:
        filepath: 源文件路径
        cache_dir: _dt_cache/ 目录
        force: 是否强制重新执行

    Returns:
        dict: 标准化结果（已转换为 valuation-detail-table 格式），
              None 表示失败（已降级到内置解析器）
    """
    cache_path = os.path.join(cache_dir, 'subjects_normalized.json')
    if not force and os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    if not _ensure_normalizer_imports():
        print(f'  [NORM] ⚠️ financial-normalizer 不可用，降级到内置解析')
        return None

    from core import detector, mapper, config as nm_config, schemas

    print(f'  [NORM] 🔍 标准化科目余额表: {os.path.basename(filepath)}')
    try:
        # Step 1: Guess — 自动检测结构
        mapping = detector.guess(filepath)
        print(f'  [NORM]   detected: {mapping["_detected_type_name"]}')
        print(f'  [NORM]   sheet: {mapping["sheet_name"]}, header_row: {mapping["header_row"]}')

        # Step 2: 自动确认映射（不要求用户手动编辑）
        mapping['confirmed'] = True

        # Step 3: Apply — 执行标准化
        df = mapper.apply(mapping)

        # Step 4: 转换为 valuation-detail-table 格式
        subjects = _convert_tb_to_vdt_format(df)

        # Step 4.5: 数据质量检查（双行表头等异常检测）
        ok, reason = _check_normalization_quality(subjects)
        if not ok:
            raise ValueError(reason)

        # Step 5: 保存缓存
        result = {
            '_meta': {
                'rule': 'DT-NORM',
                'created_at': datetime.now().isoformat(),
                'source_file': filepath,
                'normalizer_version': '0.1.0',
            },
            'source_type': 'subject_balance',
            'filepath': filepath,
            'status': 'parsed',
            'header_row': mapping.get('header_row', 0),
            'col_map': mapping.get('column_mapping', {}),
            'subjects': subjects,
            'warnings': [],
        }

        os.makedirs(cache_dir, exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'  [NORM] ✅ 标准化完成: {len(subjects)} 科目 → {cache_path}')
        return result

    except Exception as e:
        print(f'  [NORM] ❌ 标准化失败: {e}')
        traceback.print_exc()
        return None


def _extract_bs_right_column(filepath: str, mapping: dict, unknown_text_cols: list) -> list:
    """提取双栏资产负债表的右侧（负债+所有者权益）列数据。

    标准中国资产负债表格式：左侧列=资产，右侧列=负债及所有者权益。
    本函数读取Excel，从右侧列提取项目。
    """
    import openpyxl
    header_row_0 = mapping.get('header_row', 3)  # detector 0-indexed
    # openpyxl 1-indexed, header实际在第 header_row_0+1 行
    header_row = header_row_0 + 1
    data_start_row = header_row + 1

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 从 col_map 中找到右侧列的索引
    # 右侧列通常位置更大
    right_col_info = None
    for src_col, target in mapping['column_mapping'].items():
        if target.startswith('__unknown') and src_col.strip() not in ('行次', '行数', '序号'):
            if right_col_info is None:
                right_col_info = src_col
            # 保留位置更靠后的列（即右侧列）
            # 但无法直接获取列索引，需要从表头匹配

    # 通过表头行获取列索引
    header_cells = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            header_cells[str(val).strip()] = col

    # 找右侧 item_name 列（column_mapping 中第二个 unknown 列名对应的 col 索引）
    sorted_unknown = sorted(
        [(k, header_cells.get(k.strip(), 999)) for k in unknown_text_cols],
        key=lambda x: x[1]
    )
    if len(sorted_unknown) < 2:
        return []

    right_src_col = sorted_unknown[-1][0]  # 列名
    right_item_col = sorted_unknown[-1][1]  # 列索引

    # 找右侧的期末余额列（通常在 item 列后第3列）
    # 资产负债表格式：项目 | 行次 | 期末余额 | 年初余额
    bs_ending_col = None
    bs_opening_col = None
    for col in range(right_item_col + 1, min(right_item_col + 5, ws.max_column + 1)):
        val = ws.cell(row=header_row, column=col).value
        if val:
            val_s = str(val).strip()
            if '期末' in val_s:
                bs_ending_col = col
            elif '年初' in val_s:
                bs_opening_col = col

    if bs_ending_col is None:
        return []

    items = []
    for row in range(data_start_row, ws.max_row + 1):
        name_val = ws.cell(row=row, column=right_item_col).value
        if name_val is None:
            continue
        name = str(name_val).strip()
        # 跳过空行、分隔线、纯粹的数字行
        if not name or name in ('None', 'nan', ''):
            continue

        ending_val = ws.cell(row=row, column=bs_ending_col).value
        opening_val = ws.cell(row=row, column=bs_opening_col).value if bs_opening_col else None

        ending = _safe_float(ending_val)
        opening = _safe_float(opening_val)

        items.append({
            'label': name,
            'ending_balance': ending,
            'beginning_balance': opening,
        })

    return items



def normalize_balance_sheet(filepath: str, cache_dir: str, force: bool = False) -> dict:
    """标准化资产负债表

    科目余额表标准化后可以反向推导 BS 数据，
    但更可靠的方案是直接用 normalizer 标准化 BS 文件。

    Args:
        filepath: BS 文件路径
        cache_dir: _dt_cache/ 目录
        force: 是否强制重新执行

    Returns:
        dict: 标准化结果，None 表示失败
    """
    cache_path = os.path.join(cache_dir, 'bs_normalized.json')
    if not force and os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    if not _ensure_normalizer_imports():
        print(f'  [NORM] ⚠️ financial-normalizer 不可用，降级到内置解析')
        return None

    from core import detector, mapper, config as nm_config, schemas

    print(f'  [NORM] 🔍 标准化资产负债表: {os.path.basename(filepath)}')
    try:
        mapping = detector.guess(filepath)
        # DT-NORM-FIX: 自动修补 item_name 字段映射
        # 资产负债表常有左右两列项目名，detector可能无法自动识别
        col_map = mapping.get('column_mapping', {})
        mapped_values = set(col_map.values())
        if 'item_name' not in mapped_values:
            # 找到第一个文本列（非行次/行数）映射到 item_name
            for src_col, target in col_map.items():
                if target.startswith('__unknown') and src_col.strip() not in ('行次', '行数', '序号'):
                    col_map[src_col] = 'item_name'
                    print(f'  [NORM] 🔧 自动修补: "{src_col.strip()}" → item_name')
                    break
            else:
                # fallback: 取第一个 unknown 列
                for src_col, target in col_map.items():
                    if target.startswith('__unknown'):
                        col_map[src_col] = 'item_name'
                        print(f'  [NORM] 🔧 自动修补(fallback): "{src_col.strip()}" → item_name')
                        break
        mapping['confirmed'] = True
        df = mapper.apply(mapping)

        # 转换为 valuation-detail-table 的 bs_balances 格式
        items = _convert_bs_to_vdt_format(df)

        # DT-NORM-FIX: 检测并处理双栏资产负债表
        # 在修补前保存原始 unknown 列用于判断
        _orig_unknown = [
            src for src, tgt in mapping.get('column_mapping', {}).items()
            if tgt.startswith('__unknown') and src.strip() not in ('行次', '行数', '序号')
        ]
        # 同时检查原始的 col_map（在修补前已修改，需计算）
        # 重新获取原始检测结果中的 unknown 列
        _raw_mapping = detector.guess(filepath)
        _raw_col_map = _raw_mapping.get('column_mapping', {})
        _raw_unknown = [
            src for src, tgt in _raw_col_map.items()
            if tgt.startswith('__unknown') and src.strip() not in ('行次', '行数', '序号')
        ]
        if len(_raw_unknown) >= 2:
            try:
                right_items = _extract_bs_right_column(filepath, _raw_mapping, _raw_unknown)
                if right_items:
                    items = items + right_items
                    print(f'  [NORM] 🔧 双栏BS: 合并右侧 {len(right_items)} 项, 总计 {len(items)} 项')
            except Exception as e:
                print(f'  [NORM] ⚠️ 双栏BS右侧提取失败: {e}')

        result = {
            '_meta': {
                'rule': 'DT-NORM',
                'created_at': datetime.now().isoformat(),
                'source_file': filepath,
                'normalizer_version': '0.1.0',
            },
            'source_type': 'balance_sheet',
            'filepath': filepath,
            'status': 'parsed',
            'header_row': mapping.get('header_row', 0),
            'col_map': mapping.get('column_mapping', {}),
            'items': items,
            'warnings': [],
        }

        # 计算合计
        total_assets = next(
            (i['ending_balance'] for i in items if '资产总计' in i['label'] or '资产合计' in i['label']),
            0
        )
        result['total_assets'] = total_assets
        # 负债和权益合计
        total_liab_equity = next(
            (i['ending_balance'] for i in items
             if '负债' in i['label'] and '所有者权益' in i['label'] and '总计' in i['label']),
            0
        )
        result['total_liabilities'] = total_liab_equity
        result['total_equity'] = 0
        result['filepath'] = filepath

        os.makedirs(cache_dir, exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'  [NORM] ✅ BS标准化完成: {len(items)} 项 → {cache_path}')
        return result

    except Exception as e:
        print(f'  [NORM] ❌ BS标准化失败: {e}')
        traceback.print_exc()
        return None


# ════════════════════════════════════════════════════════════
# 格式转换器
# ════════════════════════════════════════════════════════════



def _check_normalization_quality(subjects, doc_type='trial_balance'):
    """检查标准化结果的数据质量，防止零余额/空数据污染缓存。

    Returns:
        (bool, str): (是否通过, 失败原因)
    """
    if not subjects:
        return False, '标准化结果为空'

    if doc_type == 'trial_balance':
        total = len(subjects)
        zero_balance = sum(1 for s in subjects if s.get('balance', 0) == 0)
        zero_ending = sum(
            1 for s in subjects
            if s.get('ending_debit', 0) == 0 and s.get('ending_credit', 0) == 0
        )
        zero_rate = zero_balance / total if total > 0 else 1.0

        # 如果超过80%的科目余额为零，且所有ending_debit/ending_credit都为零，
        # 说明normalizer没识别到期末余额列（双行表头场景）
        if zero_rate > 0.8 and zero_ending == total:
            return False, (
                f'数据质量异常: {zero_balance}/{total} 科目零余额 ({zero_rate:.0%}), '
                f'所有ending_debit/ending_credit均为零 → 疑似双行表头未识别, '
                f'应降级到source_header_parser'
            )

        # 全部科目都零余额 → 肯定有问题
        if zero_rate > 0.99:
            return False, (
                f'数据质量异常: {zero_balance}/{total} 科目零余额 ({zero_rate:.0%}) → 标准化结果无效'
            )

    return True, 'OK'



def _safe_float(val):
    """安全转换为 float，NaN 转为 0.0"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if math.isnan(val) else float(val)
    try:
        return float(str(val).replace(',', '').replace(' ', ''))
    except (ValueError, TypeError):
        return 0.0


def _load_field_mapping():
    """从 field_mapping.json 加载字段映射规则"""
    global _field_mapping_cache
    if '_field_mapping_cache' not in globals():
        _field_mapping_cache = {}
    if _field_mapping_cache:
        return _field_mapping_cache
    try:
        mapping_path = Path(__file__).resolve().parent.parent / 'assets' / 'field_mapping.json'
        if mapping_path.exists():
            with open(mapping_path, 'r', encoding='utf-8') as f:
                _field_mapping_cache = json.load(f)
                return _field_mapping_cache
    except Exception:
        pass
    return {}


def _convert_tb_to_vdt_format(df_rows):
    """将 normalizer 的科目余额表输出转换为 valuation-detail-table 格式

    字段映射规则来自 field_mapping.json 的 trial_balance.field_map。
    修复字段映射只需改 field_mapping.json，无需改代码。
    """
    fm = _load_field_mapping()
    tb_map = fm.get('trial_balance', {}).get('field_map', {})
    
    # 降级：如果 JSON 不可用，使用内置 fallback
    if not tb_map:
        tb_map = {
            'account_code': 'code',
            'account_name': 'name',
            'opening_debit': 'beginning_debit',
            'opening_credit': 'beginning_credit',
            'current_debit': 'current_debit',
            'current_credit': 'current_credit',
            'closing_debit': 'ending_debit',
            'closing_credit': 'ending_credit',
        }

    if hasattr(df_rows, 'to_dict'):
        rows = df_rows.to_dict('records')
    elif isinstance(df_rows, list):
        rows = df_rows
    else:
        rows = []

    subjects = []
    for row in rows:
        if not isinstance(row, dict):
            continue

        code = str(row.get('account_code', '') or '')
        name = str(row.get('account_name', '') or '')
        if not code or not name:
            continue

        # 按 field_mapping 转换字段
        subject = {}
        for src_field, tgt_field in tb_map.items():
            if tgt_field and src_field in row:
                val = row[src_field]
                if tgt_field in ('beginning_debit', 'beginning_credit', 'current_debit',
                                 'current_credit', 'ending_debit', 'ending_credit',
                                 'balance'):
                    val = _safe_float(val)
                if tgt_field == 'code':
                    val = str(val)
                if tgt_field == 'name':
                    val = str(val)
                subject[tgt_field] = val

        # 计算 balance 和 direction（field_mapping 中的 compute_fields）
        ed = subject.get('ending_debit', 0)
        ec = subject.get('ending_credit', 0)
        if ed and not ec:
            subject['balance'] = ed
            subject['direction'] = '借'
        elif ec and not ed:
            subject['balance'] = ec
            subject['direction'] = '贷'
        elif ed and ec:
            subject['balance'] = ed - ec
            subject['direction'] = '借' if ed >= ec else '贷'
        else:
            subject['balance'] = 0.0
            subject['direction'] = ''

        # 级次
        subject['level'] = max(len(code) // 2, 1) if len(code) <= 8 else 1

        subjects.append(subject)

    return subjects


def _convert_bs_to_vdt_format(df_rows):
    """将 normalizer 的资产负债表输出转换为 valuation-detail-table 格式

    字段映射规则来自 field_mapping.json 的 balance_sheet.field_map。
    """
    fm = _load_field_mapping()
    bs_map = fm.get('balance_sheet', {}).get('field_map', {})
    
    if not bs_map:
        bs_map = {
            'item_name': 'label',
            'closing_balance': 'ending_balance',
            'opening_balance': 'beginning_balance',
        }

    if hasattr(df_rows, 'to_dict'):
        rows = df_rows.to_dict('records')
    elif isinstance(df_rows, list):
        rows = df_rows
    else:
        rows = []

    items = []
    for row in rows:
        if not isinstance(row, dict):
            continue

        name = str(row.get('item_name', '') or '')
        if not name:
            continue

        # 按 field_mapping 转换
        item = {}
        for src_field, tgt_field in bs_map.items():
            if tgt_field and src_field in row:
                val = row[src_field]
                if 'balance' in tgt_field:
                    val = _safe_float(val)
                item[tgt_field] = val

        # side 判断
        category = str(row.get('item_category', '') or '')
        if '负债' in category or '所有者权益' in category:
            item['side'] = '负债及权益'
        elif '资产' in category:
            item['side'] = '资产'
        else:
            item['side'] = ''

        items.append(item)

    return items


# ════════════════════════════════════════════════════════════
# 主流程
# ════════════════════════════════════════════════════════════

def _convert_fa_to_vdt_format(df_rows):
    """将 normalizer 的固定资产台账输出转换为 valuation-detail-table 格式

    mapper.apply() 的输出已经是标准字段名（asset_code, original_value 等），
    本函数将其转换为评估明细表所需的字段格式。
    输出格式兼容评估明细表设备类Sheet（4-8-x）。
    """
    # mapper 输出的标准字段 → 评估明细表字段的直接映射
    # 字段名基本一致，只需确保有 seq/manufacturer/spec 等补充字段

    if hasattr(df_rows, 'to_dict'):
        rows = df_rows.to_dict('records')
    elif isinstance(df_rows, list):
        rows = df_rows
    else:
        rows = []

    items = []
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue

        name = str(row.get('asset_name', '') or '')
        if not name:
            continue

        # mapper 输出已是标准字段名，直接提取
        item = {
            'seq': i + 1,
            'asset_name': str(row.get('asset_name', '') or ''),
            'spec': str(row.get('specification', '') or ''),
            'manufacturer': str(row.get('supplier', '') or ''),
            'department': str(row.get('department', '') or ''),
            'location': str(row.get('location', '') or ''),
            'status': str(row.get('status', '') or ''),
            'original_value': _safe_float(row.get('original_value')),
            'accumulated_depreciation': _safe_float(row.get('accumulated_depreciation')),
            'net_value': _safe_float(row.get('net_value')),
            'impairment_amount': _safe_float(row.get('impairment_amount')),
            'quantity': _safe_float(row.get('quantity')),
            'depreciation_life': _safe_float(row.get('depreciation_life')),
            'residual_rate': _safe_float(row.get('residual_rate')),
            'monthly_depreciation': _safe_float(row.get('monthly_depreciation')),
            'start_date': str(row.get('start_date', '') or ''),
            'acquisition_date': str(row.get('acquisition_date', '') or ''),
            'asset_code': str(row.get('asset_code', '') or ''),
            'asset_category': str(row.get('asset_category', '') or ''),
            'depreciation_method': str(row.get('depreciation_method', '') or ''),
            'unit': str(row.get('unit', '') or ''),
            'currency': str(row.get('currency', '') or ''),
        }

        # 净值回退计算
        if not item['net_value'] and item['original_value']:
            item['net_value'] = item['original_value'] - item['accumulated_depreciation']

        items.append(item)

    return items



def normalize_fixed_asset(filepath: str, cache_dir: str, force: bool = False) -> dict:
    """标准化固定资产台账

    Args:
        filepath: 源文件路径
        cache_dir: _dt_cache/ 目录
        force: 是否强制重新执行

    Returns:
        dict: 标准化结果，None 表示失败
    """
    cache_path = os.path.join(cache_dir, 'fixed_assets_normalized.json')
    if not force and os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    if not _ensure_normalizer_imports():
        print(f'  [NORM] ⚠️ financial-normalizer 不可用，降级到内置解析')
        return None

    from core import detector, mapper, config as nm_config, schemas

    print(f'  [NORM] 🔍 标准化固定资产台账: {os.path.basename(filepath)}')
    try:
        # Step 1: Guess — 自动检测结构
        mapping = detector.guess(filepath)
        print(f'  [NORM]   detected: {mapping["_detected_type_name"]}')
        print(f'  [NORM]   sheet: {mapping["sheet_name"]}, header_row: {mapping["header_row"]}')

        # Step 2: 应用 column_priority 优化映射（去重，选最优列）
        fa_config = _load_field_mapping().get('fixed_asset', {})
        column_priority = fa_config.get('column_priority', {})
        col_map = mapping.get('column_mapping', {})

        if column_priority:
            # 构建 {标准字段: [源列名列表]} 的反向索引
            field_sources = {}
            for src, tgt in col_map.items():
                src_clean = src.strip()
                if not tgt.startswith('__unknown'):
                    if tgt not in field_sources:
                        field_sources[tgt] = []
                    field_sources[tgt].append(src_clean)

            # 对每个冲突字段，按优先级选择最优列
            for std_field, priority_list in column_priority.items():
                if std_field in field_sources and len(field_sources[std_field]) > 1:
                    candidates = field_sources[std_field]
                    # 按优先级排序选最优
                    best = None
                    for preferred in priority_list:
                        # 精确匹配优先，再fallback到子串匹配
                        exact = [c for c in candidates if c == preferred]
                        if exact:
                            best = exact[0]
                            break
                        fuzzy = [c for c in candidates if preferred in c]
                        if fuzzy:
                            best = fuzzy[0]
                            break
                    if best:
                        # 把其他冲突列移除（mapper会跳过缺失的列）
                        removed = []
                        for c in candidates:
                            if c != best:
                                del col_map[c]
                                removed.append(c)
                        if removed:
                            print(f'  [NORM]   column_priority: {std_field} → "{best}" (移除冲突: {removed})')

        mapping['confirmed'] = True

        # Step 3: Apply — 执行标准化
        df = mapper.apply(mapping)

        # Step 4: 转换为 valuation-detail-table 格式
        items = _convert_fa_to_vdt_format(df)

        # Step 5: 保存缓存
        result = {
            '_meta': {
                'rule': 'DT-NORM',
                'created_at': datetime.now().isoformat(),
                'source_file': filepath,
                'normalizer_version': '0.1.0',
            },
            'source_type': 'fixed_asset',
            'filepath': filepath,
            'status': 'parsed',
            'header_row': mapping.get('header_row', 0),
            'col_map': mapping.get('column_mapping', {}),
            'items': items,
            'warnings': [],
        }

        # 统计
        total_ov = sum(i.get('original_value', 0) or 0 for i in items)
        total_ad = sum(i.get('accumulated_depreciation', 0) or 0 for i in items)
        result['total_original_value'] = total_ov
        result['total_accumulated_depreciation'] = total_ad
        result['total_net_value'] = total_ov - total_ad

        os.makedirs(cache_dir, exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'  [NORM] ✅ 固定资产标准化完成: {len(items)} 项 → {cache_path}')
        print(f'  [NORM]   原值合计={total_ov:,.2f}, 累计折旧={total_ad:,.2f}, 净值={total_ov-total_ad:,.2f}')
        return result

    except Exception as e:
        print(f'  [NORM] ❌ 固定资产标准化失败: {e}')
        traceback.print_exc()
        return None



def _convert_journal_to_vdt_format(df_rows):
    """将 normalizer 的序时账输出转换为 valuation-detail-table 格式。

    mapper.apply() 的输出已是标准字段名（date, account_code 等）。
    """
    if hasattr(df_rows, 'to_dict'):
        rows = df_rows.to_dict('records')
    elif isinstance(df_rows, list):
        rows = df_rows
    else:
        rows = []

    entries = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        code = str(row.get('account_code', '') or '')
        if not code:
            continue

        entry = {
            'date': str(row.get('date', '') or ''),
            'voucher_no': str(row.get('voucher_no', '') or ''),
            'account_code': code,
            'account_name': str(row.get('account_name', '') or ''),
            'summary': str(row.get('summary', '') or ''),
            'currency': str(row.get('currency', '') or ''),
            'debit_amount': _safe_float(row.get('debit_amount')),
            'credit_amount': _safe_float(row.get('credit_amount')),
            'customer_supplier': str(row.get('customer_supplier', '') or ''),
            'department': str(row.get('department', '') or ''),
            'project_name': str(row.get('project_name', '') or ''),
            'personnel': str(row.get('personnel', '') or ''),
        }
        entries.append(entry)

    return entries


def normalize_journal(filepath: str, cache_dir: str, force: bool = False) -> dict:
    """标准化序时账

    Args:
        filepath: 源文件路径
        cache_dir: _dt_cache/ 目录
        force: 是否强制重新执行

    Returns:
        dict: 标准化结果，None 表示失败
    """
    cache_path = os.path.join(cache_dir, 'journal_normalized.json')
    if not force and os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    if not _ensure_normalizer_imports():
        print(f'  [NORM] ⚠️ financial-normalizer 不可用，降级到内置解析')
        return None

    from core import detector, mapper, config as nm_config, schemas

    print(f'  [NORM] 🔍 标准化序时账: {os.path.basename(filepath)}')
    try:
        mapping = detector.guess(filepath)
        print(f'  [NORM]   detected: {mapping["_detected_type_name"]}')
        print(f'  [NORM]   sheet: {mapping["sheet_name"]}, header_row: {mapping["header_row"]}')

        mapping['confirmed'] = True
        df = mapper.apply(mapping)
        entries = _convert_journal_to_vdt_format(df)
        print(f'  [NORM]   解析: {len(entries)} 条分录')

        # 只保留往来科目相关的分录（减少体积，加速Phase 3查询）
        # 非往来科目的分录在Phase 3中用不到
        counter_party_codes = {'1122', '1123', '1221', '2202', '2203', '2241',
                               '1121', '2201', '1231', '1241'}
        filtered = [e for e in entries
                    if any(e['account_code'].startswith(c) for c in counter_party_codes)]

        # 统计
        unique_codes = set(e['account_code'] for e in entries)
        date_range = ''
        dates = sorted(set(e['date'] for e in entries if e['date']))
        if dates:
            date_range = f'{dates[0]} ~ {dates[-1]}'

        result = {
            '_meta': {
                'rule': 'DT-NORM',
                'created_at': datetime.now().isoformat(),
                'source_file': filepath,
                'normalizer_version': '0.1.0',
            },
            'source_type': 'journal',
            'filepath': filepath,
            'status': 'parsed',
            'header_row': mapping.get('header_row', 0),
            'col_map': mapping.get('column_mapping', {}),
            'entries': entries,
            'filtered_entries': filtered,
            'total_entries': len(entries),
            'filtered_count': len(filtered),
            'unique_subjects': len(unique_codes),
            'date_range': date_range,
            'warnings': [],
        }

        os.makedirs(cache_dir, exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f'  [NORM] ✅ 序时账标准化完成: {len(entries)} 条分录 (往来相关{len(filtered)}条) → {cache_path}')
        print(f'  [NORM]   日期范围: {date_range}, 涉及{len(unique_codes)}个科目')
        return result

    except Exception as e:
        print(f'  [NORM] ❌ 序时账标准化失败: {e}')
        traceback.print_exc()
        return None



def _export_standardized_workbook(project_dir: str, report: dict):
    """生成统一标准化工作簿（Excel），与评估明细表并列输出。

    包含Sheet:
      1. 科目余额表（标准化） — 含映射到的评估明细表Sheet
      2. 资产负债表（标准化）
      3. 固定资产台账（标准化）
      4. 映射关系总表 — 科目编码→Sheet→列位完整链路
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  [EXPORT] ⚠️ openpyxl 未安装，跳过标准化工作簿导出')
        return

    cache_dir = os.path.join(project_dir, '_dt_cache')
    output_path = os.path.join(project_dir, '标准化财务数据汇总.xlsx')

    print(f'\n[Export] 生成标准化工作簿 → {os.path.basename(output_path)}')

    wb = openpyxl.Workbook()
    # 删默认sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    num_fmt = '#,##0.00'

    code_to_sheet = {}
    # 搜索 field_mapping.json（支持多种目录结构）
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    fm_path = os.path.join(_script_dir, 'assets', 'field_mapping.json')
    if not os.path.exists(fm_path):
        fm_path = os.path.join(os.path.dirname(_script_dir), 'assets', 'field_mapping.json')
    try:
        with open(fm_path, 'r', encoding='utf-8') as f:
            fm_data = json.load(f)
        code_to_sheet = fm_data.get('code_to_sheet', {}).get('mappings', {})
    except Exception:
        pass

    # ── Sheet 1: 科目余额表（标准化） ──
    ws1 = wb.create_sheet('科目余额表（标准化）')
    headers1 = ['科目编码', '科目名称', '期末余额', '方向', '科目级次',
                '映射Sheet编码', '映射Sheet名称']
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    subjects_path = os.path.join(cache_dir, 'subjects_normalized.json')
    subjects = []
    if os.path.exists(subjects_path):
        with open(subjects_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        subjects = data.get('subjects', []) if isinstance(data, dict) else data

    # 加载 subject_schema 用于Sheet名称查找
    ss_data = {}
    _ss_path = os.path.join(os.path.dirname(fm_path), 'subject_schema.json')
    if os.path.exists(_ss_path):
        try:
            with open(_ss_path, 'r', encoding='utf-8') as f:
                ss_data = json.load(f)
        except Exception:
            pass

    row = 2
    for s in subjects:
        code = str(s.get('code', ''))
        name = str(s.get('name', ''))
        bal = s.get('balance', 0) or 0
        direction = s.get('direction', '')
        level = s.get('level', 1)

        # 查找映射
        code4 = code[:4]
        sheet_code = code_to_sheet.get(code4) or code_to_sheet.get(code, '')
        sheet_label = ''

        ws1.cell(row=row, column=1, value=code).border = thin_border
        ws1.cell(row=row, column=2, value=name).border = thin_border
        c3 = ws1.cell(row=row, column=3, value=bal)
        c3.number_format = num_fmt
        c3.border = thin_border
        ws1.cell(row=row, column=4, value=direction).border = thin_border
        ws1.cell(row=row, column=5, value=level).border = thin_border
        c6 = ws1.cell(row=row, column=6, value=sheet_code)
        c6.border = thin_border
        if not sheet_code and bal != 0:
            c6.font = Font(color='FF0000')
        # 尝试从 subject_schema 获取Sheet完整名称
        ss_name = ''
        if sheet_code:
            for sk in ss_data.get('subjects', {}):
                import re as _re
                pm = _re.match(r'^([3-6]-[\d\-]+)', sk)
                if pm and pm.group(1).rstrip('-') == sheet_code:
                    ss_name = sk
                    break
                if sk == sheet_code:
                    ss_name = sk
                    break
        ws1.cell(row=row, column=7, value=ss_name or sheet_label).border = thin_border
        row += 1

    ws1.auto_filter.ref = f'A1:G{row-1}'
    for c in range(1, 8):
        ws1.column_dimensions[get_column_letter(c)].width = [12, 30, 16, 6, 8, 14, 25][c-1]

    # ── Sheet 2: 资产负债表（标准化） ──
    ws2 = wb.create_sheet('资产负债表（标准化）')
    headers2 = ['项目名称', '期末余额', '年初余额', '资产/负债']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    bs_path = os.path.join(cache_dir, 'bs_normalized.json')
    bs_items = []
    if os.path.exists(bs_path):
        with open(bs_path, 'r', encoding='utf-8') as f:
            bs_data = json.load(f)
        bs_items = bs_data.get('items', [])

    # 判断资产/负债侧（简单推断：在资产总计之前=资产，之后=负债）
    asset_end = 0
    for i, item in enumerate(bs_items):
        if '资产总计' in (item.get('label', '') or ''):
            asset_end = i
            break

    row = 2
    for i, item in enumerate(bs_items):
        label = item.get('label', '') or ''
        ending = item.get('ending_balance', 0) or 0
        beginning = item.get('beginning_balance', 0) or 0
        side = '资产' if i <= asset_end else '负债及权益'

        ws2.cell(row=row, column=1, value=label.strip()).border = thin_border
        c2 = ws2.cell(row=row, column=2, value=ending)
        c2.number_format = num_fmt
        c2.border = thin_border
        c3 = ws2.cell(row=row, column=3, value=beginning)
        c3.number_format = num_fmt
        c3.border = thin_border
        ws2.cell(row=row, column=4, value=side).border = thin_border
        row += 1

    ws2.auto_filter.ref = f'A1:D{row-1}'
    for c, w in enumerate([30, 16, 16, 12], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ── Sheet 3: 固定资产台账（标准化） ──
    ws3 = wb.create_sheet('固定资产台账（标准化）')
    fa_headers = ['序号', '资产名称', '规格型号', '原值', '累计折旧', '净值',
                  '使用部门', '存放地点', '开始使用日期', '折旧方法', '使用年限(月)', '状态']
    for c, h in enumerate(fa_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    fa_path = os.path.join(cache_dir, 'fixed_assets_normalized.json')
    fa_items = []
    if os.path.exists(fa_path):
        with open(fa_path, 'r', encoding='utf-8') as f:
            fa_data = json.load(f)
        fa_items = fa_data.get('items', [])

    row = 2
    for item in fa_items:
        vals = [
            item.get('seq', ''),
            item.get('asset_name', ''),
            item.get('spec', ''),
            item.get('original_value', 0) or 0,
            item.get('accumulated_depreciation', 0) or 0,
            item.get('net_value', 0) or 0,
            item.get('department', ''),
            item.get('location', ''),
            item.get('start_date', ''),
            item.get('depreciation_method', ''),
            item.get('depreciation_life', 0) or 0,
            item.get('status', ''),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.border = thin_border
            if c in (4, 5, 6):
                cell.number_format = num_fmt
        row += 1

    ws3.auto_filter.ref = f'A1:L{row-1}'
    for c, w in enumerate([6, 28, 22, 14, 14, 14, 14, 14, 14, 14, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # 行合计
    if fa_items:
        total_row = row
        ws3.cell(row=total_row, column=1, value='合计').font = header_font
        fa_key_map = {4: 'original_value', 5: 'accumulated_depreciation', 6: 'net_value'}
        for c in (4, 5, 6):
            key = fa_key_map.get(c, fa_headers[c-1])
            cell = ws3.cell(row=total_row, column=c,
                           value=sum((i.get(key, 0) or 0) for i in fa_items))
            cell.number_format = num_fmt
            cell.font = header_font
            cell.border = thin_border

    # ── Sheet 3.5: 序时账（标准化·往来科目摘要） ──
    ws35 = wb.create_sheet('序时账（往来科目摘要）')
    j_headers = ['日期', '凭证号', '科目编码', '科目名称', '摘要', '借方金额', '贷方金额']
    for c, h in enumerate(j_headers, 1):
        cell = ws35.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    j_path = os.path.join(cache_dir, 'journal_normalized.json')
    j_entries = []
    if os.path.exists(j_path):
        try:
            with open(j_path, 'r', encoding='utf-8') as f:
                j_data = json.load(f)
            # 优先展示往来科目（filtered），如无则用前500条
            j_entries = j_data.get('filtered_entries', [])
            if not j_entries:
                j_entries = j_data.get('entries', [])[:500]
        except Exception:
            pass

    row = 2
    for e in j_entries[:2000]:  # 最多2000行
        ws35.cell(row=row, column=1, value=str(e.get('date', ''))[:10]).border = thin_border
        ws35.cell(row=row, column=2, value=str(e.get('voucher_no', ''))).border = thin_border
        ws35.cell(row=row, column=3, value=str(e.get('account_code', ''))).border = thin_border
        ws35.cell(row=row, column=4, value=str(e.get('account_name', ''))).border = thin_border
        ws35.cell(row=row, column=5, value=str(e.get('summary', ''))[:80]).border = thin_border
        c6 = ws35.cell(row=row, column=6, value=e.get('debit_amount', 0) or 0)
        c6.number_format = num_fmt
        c6.border = thin_border
        c7 = ws35.cell(row=row, column=7, value=e.get('credit_amount', 0) or 0)
        c7.number_format = num_fmt
        c7.border = thin_border
        row += 1

    ws35.auto_filter.ref = f'A1:G{row-1}'
    for c, w in enumerate([12, 14, 12, 25, 50, 14, 14], 1):
        ws35.column_dimensions[get_column_letter(c)].width = w
    ws35.freeze_panes = 'A2'

    # ── Sheet 4: 映射关系总表 ──
    ws4 = wb.create_sheet('映射关系总表')
    map_headers = ['科目编码', '科目名称', '映射Sheet编码', '数据命中', '说明']
    for c, h in enumerate(map_headers, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 构建 code→name 映射
    code_name = {}
    for s in subjects:
        code_name[str(s.get('code', ''))] = str(s.get('name', ''))

    row = 2
    for code, sheet_code in sorted(code_to_sheet.items()):
        name = code_name.get(code, '')
        has_data = any(
            str(s.get('code', '')).startswith(code) and (s.get('balance', 0) or 0) != 0
            for s in subjects
        )
        desc = ''
        if code.startswith('400'):
            desc = '所有者权益→净资产汇总'
        elif code.startswith('160'):
            desc = '固定资产→设备类Sheet'
        elif code.startswith('500') or code.startswith('520'):
            desc = '成本类→存货Sheet'

        ws4.cell(row=row, column=1, value=code).border = thin_border
        ws4.cell(row=row, column=2, value=name).border = thin_border
        ws4.cell(row=row, column=3, value=sheet_code).border = thin_border
        c4 = ws4.cell(row=row, column=4, value='✓' if has_data else '')
        c4.border = thin_border
        c4.font = Font(color='008000') if has_data else Font(color='C0C0C0')
        ws4.cell(row=row, column=5, value=desc).border = thin_border
        row += 1

    ws4.auto_filter.ref = f'A1:E{row-1}'
    for c, w in enumerate([12, 30, 16, 10, 30], 1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    # 冻结首行
    for ws in [ws1, ws2, ws3, ws4]:
        ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f'  [EXPORT] ✅ 标准化工作簿已生成: {os.path.basename(output_path)}')
    print(f'  [EXPORT]   Sheet1: 科目余额表 ({len(subjects)}条)')
    print(f'  [EXPORT]   Sheet2: 资产负债表 ({len(bs_items)}项)')
    print(f'  [EXPORT]   Sheet3: 固定资产台账 ({len(fa_items)}项)')
    print(f'  [EXPORT]   Sheet4: 映射关系总表 ({len(code_to_sheet)}条)')

    return output_path



def run_normalization(project_dir: str, force: bool = False):
    """执行财务资料标准化全流程

    Args:
        project_dir: 项目文件夹路径
        force: 是否强制重新执行（忽略缓存）

    Returns:
        dict: 标准化执行报告
    """
    cache_dir = os.path.join(project_dir, '_dt_cache')
    os.makedirs(cache_dir, exist_ok=True)

    report = {
        'phase': '-1.5',
        'project_dir': project_dir,
        'status': 'completed',
        'normalized': [],
        'failed': [],
        'skipped': [],
        'warnings': [],
    }

    print(f'\n{"="*60}')
    print('Phase -1.5: 财务资料标准化 (normalize_input.py)')
    print(f'{"="*60}')

    # Step 1: 发现财务文件
    print('\n[Step -1.5.1] 发现财务文件')
    found_files = discover_financial_files(project_dir)
    if not found_files:
        print('  [NORM] ⚠️ 未发现可标准化的财务文件')
        report['status'] = 'skipped'
        report['warnings'].append('未发现财务文件，跳过标准化')
        return report

    print(f'  [NORM] 发现 {sum(len(v) for v in found_files.values())} 个财务文件')

    # Step 2: 检查 normalizer 可用性
    print('\n[Step -1.5.2] 检查 normalizer 可用性')
    normalizer_available = _ensure_normalizer_imports()
    if normalizer_available:
        print('  [NORM] ✅ financial-normalizer 可用')
    else:
        print('  [NORM] ⚠️ financial-normalizer 不可用，Phase -1.5 将跳过')

    # Step 3: 标准化科目余额表（最高优先级，支持多文件尝试）
    print('\n[Step -1.5.3] 标准化科目余额表')
    tb_files = found_files.get('trial_balance', [])
    if tb_files:
        # 优先尝试文件名含"更新"/"最新"的文件
        preferred = [f for f in tb_files if '更新' in f or '最新' in f]
        remaining = [f for f in tb_files if f not in preferred]
        ordered_files = preferred + remaining

        tb_result = None
        for tb_file in ordered_files:
            tb_result = normalize_trial_balance(tb_file, cache_dir, force)
            if tb_result:
                break
            print(f'  [NORM] ⚠️ {os.path.basename(tb_file)} 标准化失败，尝试下一个...')

        if tb_result:
            report['normalized'].append({
                'type': 'trial_balance',
                'file': tb_file,
                'count': len(tb_result.get('subjects', [])),
            })
            # 同时创建 subjects.json 别名（兼容 Phase 0 缓存检测）
            cache_path = os.path.join(cache_dir, 'subjects.json')
            if not os.path.exists(cache_path) or force:
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(tb_result['subjects'], f, ensure_ascii=False, indent=2)
                print(f'  [NORM] 🔗 同时写入 subjects.json 供 Phase 0 使用')
        else:
            report['failed'].append({'type': 'trial_balance', 'file': tb_files[-1]})
            report['warnings'].append('科目余额表标准化失败，Phase 0 将使用内置解析器')
    else:
        print('  [NORM] ⚠️ 未找到科目余额表文件')
        report['skipped'].append('trial_balance')

    # Step 4: 标准化资产负债表
    print('\n[Step -1.5.4] 标准化资产负债表')
    bs_files = found_files.get('balance_sheet', [])
    if bs_files:
        bs_result = normalize_balance_sheet(bs_files[0], cache_dir, force)
        if bs_result:
            report['normalized'].append({
                'type': 'balance_sheet',
                'file': bs_files[0],
                'count': len(bs_result.get('items', [])),
            })
            # 同时创建 bs_balances.json 别名
            cache_path = os.path.join(cache_dir, 'bs_balances.json')
            if not os.path.exists(cache_path) or force:
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(bs_result, f, ensure_ascii=False, indent=2)
                print(f'  [NORM] 🔗 同时写入 bs_balances.json 供 Phase 0 使用')
        else:
            report['failed'].append({'type': 'balance_sheet', 'file': bs_files[0]})
            report['warnings'].append('资产负债表标准化失败，Phase 0 将使用内置解析器')
    else:
        print('  [NORM] ⚠️ 未找到资产负债表文件')
        report['skipped'].append('balance_sheet')

    # Step 4.5: 标准化固定资产台账
    print('\n[Step -1.5.4b] 标准化固定资产台账')
    fa_files = found_files.get('fixed_asset', [])
    if fa_files:
        # 优先非standardized文件（原始文件）
        original_fa = [f for f in fa_files if 'standardized' not in f and '标准化' not in f]
        fa_target = original_fa[0] if original_fa else fa_files[0]
        fa_result = normalize_fixed_asset(fa_target, cache_dir, force)
        if fa_result:
            report['normalized'].append({
                'type': 'fixed_asset',
                'file': fa_target,
                'count': len(fa_result.get('items', [])),
            })
            # 同时写入 fixed_assets.json 别名
            cache_path = os.path.join(cache_dir, 'fixed_assets.json')
            if not os.path.exists(cache_path) or force:
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(fa_result['items'], f, ensure_ascii=False, indent=2)
                print(f'  [NORM] 🔗 同时写入 fixed_assets.json 供后续阶段使用')
        else:
            report['failed'].append({'type': 'fixed_asset', 'file': fa_target})
            report['warnings'].append('固定资产台账标准化失败')
    else:
        print('  [NORM] ℹ️ 未发现固定资产台账文件')
        report['skipped'].append('fixed_asset')

    # Step 5: 标准化序时账
    print('\n[Step -1.5.5] 标准化序时账')
    journal_files = found_files.get('journal', [])
    if journal_files:
        # 优先选当年原始序时账（排除已标准化的）
        original_j = [f for f in journal_files if 'standardized' not in f and '标准化' not in f]
        current_year = datetime.now().year
        year_files = [f for f in original_j if str(current_year) in f]
        j_target = year_files[0] if year_files else (original_j[0] if original_j else journal_files[0])
        j_result = normalize_journal(j_target, cache_dir, force)
        if j_result:
            report['normalized'].append({
                'type': 'journal',
                'file': j_target,
                'count': j_result.get('total_entries', 0),
            })
            # 同时写入 journal.json 别名
            cache_path = os.path.join(cache_dir, 'journal.json')
            if not os.path.exists(cache_path) or force:
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(j_result['entries'], f, ensure_ascii=False, indent=2)
                print(f'  [NORM] 🔗 同时写入 journal.json 供 Phase 3 使用')
        else:
            report['failed'].append({'type': 'journal', 'file': j_target})
            report['warnings'].append('序时账标准化失败，Phase 3 将使用内置解析器')
    else:
        print(f'  [NORM] ℹ️ 未发现序时账文件')
        report['skipped'].append('journal')

    # 输出报告
    print(f'\n{"="*60}')
    print(f'Phase -1.5 执行报告:')
    print(f'  标准化成功: {len(report["normalized"])} 项')
    for n in report['normalized']:
        print(f'    ✅ {n["type"]}: {os.path.basename(n["file"])} ({n["count"]} 条)')
    print(f'  标准化失败: {len(report["failed"])} 项')
    for f in report['failed']:
        print(f'    ❌ {f["type"]}: {os.path.basename(f["file"])}')
    if report['warnings']:
        print(f'  警告:')
        for w in report['warnings']:
            print(f'    ⚠️ {w}')
    print(f'{"="*60}')

    # DT-NORM: 导出统一标准化工作簿（与评估明细表并列）
    try:
        _export_standardized_workbook(project_dir, report)
    except Exception as e:
        print(f'  [EXPORT] ⚠️ 标准化工作簿导出失败: {e}')

    return report


# ════════════════════════════════════════════════════════════
# CLI入口
# ════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='normalize_input.py — 财务资料标准化桥接模块 (DT-NORM)'
    )
    parser.add_argument('--project', required=True,
                       help='项目文件夹路径')
    parser.add_argument('--force', action='store_true',
                       help='强制重新标准化（忽略缓存）')
    parser.add_argument('--dry-run', action='store_true',
                       help='仅检测文件，不执行标准化')

    args = parser.parse_args()

    project_dir = os.path.abspath(args.project)
    if not os.path.isdir(project_dir):
        print(f'❌ 项目目录不存在: {project_dir}')
        sys.exit(1)

    if args.dry_run:
        print(f'🔍 文件发现预览 (dry-run): {project_dir}')
        found = discover_financial_files(project_dir)
        print(f'\n发现 {sum(len(v) for v in found.values())} 个财务文件:')
        for doc_type, files in found.items():
            for f in files:
                print(f'  [{doc_type:20s}] {os.path.relpath(f, project_dir)}')
        return

    run_normalization(project_dir, force=args.force)


if __name__ == '__main__':
    main()
