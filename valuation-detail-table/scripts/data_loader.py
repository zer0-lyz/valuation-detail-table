"""
data_loader.py — 统一数据加载器 v1.0

设计目的:
  消灭"数据源混乱"问题（河南平绿项目3大根因之一）。
  每个科目的数据从哪来、怎么去重、怎么校验，全部在subject_schema.json中定义。
  脚本加载Schema而非自行推断，Agent不自行决定数据源。

核心原则:
  - subject_schema.json为唯一数据源规格，Agent不可绕过
  - 所有科目数据加载MUST通过本模块执行
  - 内置去重逻辑（按dedup_key去重，按dedup_strategy决定保留策略）
  - 内置校验：加载后自动与reconcile_to比对，差异>阈值则告警
  - 幂等保护：每次加载生成数据指纹，重复加载不重复写入

加载流程:
  1. 从subject_schema.json读取科目规格
  2. 按primary_source加载数据
  3. 按dedup_key去重
  4. 如需补充，按secondary_source加载+去重
  5. 与reconcile_to比对校验
  6. 输出data_rows + reconcile_target + load_report

新增规则:
  DT-155: 幂等保护（fill前先clear目标区域）
  DT-156: 数据源去重（辅助余额表跨表去重）
  DT-158: 每sheet填后即时勾稽
"""

import json
import os
import sys
import hashlib
from pathlib import Path

# 路径配置
_SCRIPT_DIR = Path(__file__).parent
_SCHEMA_PATH = _SCRIPT_DIR.parent / 'assets' / 'subject_schema.json'


def load_subject_schema(schema_path=None):
    """加载subject_schema.json

    Args:
        schema_path: Schema文件路径（默认使用Skill内置路径）

    Returns:
        dict: 完整Schema数据
    """
    path = Path(schema_path) if schema_path else _SCHEMA_PATH
    if not path.exists():
        raise FileNotFoundError(
            f'DT-156 CRITICAL: subject_schema.json未找到({path})！'
            f'每个科目的数据源规格必须在此文件中定义。'
        )
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_subject_config(schema, sheet_name):
    """获取指定科目的配置

    Args:
        schema: load_subject_schema()返回的完整数据
        sheet_name: Sheet名称（如'5-5应付账款'）

    Returns:
        dict or None: 该科目的配置
    """
    subjects = schema.get('subjects', {})
    # 精确匹配
    if sheet_name in subjects:
        return subjects[sheet_name]
    # 模糊匹配（去掉括号内容）
    import re
    clean_name = re.sub(r'[（(].+?[）)]', '', sheet_name).strip()
    for key, config in subjects.items():
        clean_key = re.sub(r'[（(].+?[）)]', '', key).strip()
        if clean_name == clean_key:
            return config
    # DT-FR2: 容错匹配——处理"5-8职工薪酬" vs "5-8应付职工薪酬"等命名差异
    # 依次去掉常见修饰词（应付/应收/其他/减值准备等）后再比较
    _strip_words = ['应付', '应收', '其他', '减值准备', '跌价准备', '坏账准备']
    _name_stripped = clean_name
    for w in _strip_words:
        _name_stripped = _name_stripped.replace(w, '')
    for key, config in subjects.items():
        clean_key = re.sub(r'[（(].+?[）)]', '', key).strip()
        _key_stripped = clean_key
        for w in _strip_words:
            _key_stripped = _key_stripped.replace(w, '')
        if _name_stripped and _key_stripped and _name_stripped == _key_stripped:
            return config
    return None


def load_cache_data(cache_dir, filename):
    """加载_dt_cache/中的JSON数据

    Args:
        cache_dir: _dt_cache/目录路径
        filename: JSON文件名

    Returns:
        dict/list or None: 加载的数据
    """
    path = os.path.join(cache_dir, filename)
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_reconcile_target(schema, config, cache_dir):
    """从reconcile_to字段获取勾稽目标值

    reconcile_to格式: "subjects.json:2202:balance"
    或: "bs_balances.json:存货:ending_balance"
    字段名已统一：subjects -> balance, bs -> ending_balance

    Args:
        schema: 完整Schema数据
        config: 科目配置
        cache_dir: _dt_cache/目录路径

    Returns:
        float or None: 勾稽目标值
    """
    reconcile_spec = config.get('reconcile_to')
    if not reconcile_spec:
        return None

    parts = reconcile_spec.split(':')
    if len(parts) < 3:
        return None

    cache_file, key, field = parts[0], parts[1], parts[2]
    data = load_cache_data(cache_dir, cache_file)
    if data is None:
        return None

    # 根据数据结构查找目标值
    if isinstance(data, dict):
        # subjects.json: {subjects: [{code, balance, ...}]}
        if 'subjects' in data:
            for subj in data['subjects']:
                if subj.get('code', '').startswith(key):
                    return subj.get(field, subj.get('balance', 0))
        # bs_balances.json: {items: [{label, ending_balance, ...}]}
        elif 'items' in data:
            for item in data['items']:
                if key in item.get('label', ''):
                    return item.get(field, 0)
    elif isinstance(data, list):
        for item in data:
            if item.get('code', '').startswith(key):
                return item.get(field, item.get('balance', 0))

    return None


def dedup_data(data_rows, dedup_key, strategy='keep_first'):
    """按dedup_key去重

    P7修复核心原则：
    - dedup_key=["code"]时：严格按科目编码去重，不同编码的同名资产是不同资产
    - dedup_key=["name","balance"]时：按名称+金额组合去重（往来科目场景）
    - 永远不做纯name去重（会把同名不同编码的资产误合并）

    Args:
        data_rows: 数据行列表
        dedup_key: 去重键列表（如['name', 'balance']或['code']）
        strategy: 去重策略（keep_first/merge_by_key）

    Returns:
        list: 去重后的数据行
    """
    if not dedup_key:
        return data_rows

    # P7修复: 若数据行有'code'字段但去重键中不含'code'，自动纳入
    # 防止不同子科目下同名同余额的明细被错误去重
    if dedup_key and data_rows:
        first = data_rows[0]
        if 'code' in first and 'code' not in dedup_key:
            dedup_key = ['code'] + list(dedup_key)

    seen = set()
    deduped = []
    duplicates = []

    for row in data_rows:
        # 构建去重键值
        key_parts = []
        for k in dedup_key:
            val = row.get(k, '')
            if isinstance(val, float):
                val = round(val, 2)  # 浮点数精度处理
            key_parts.append(str(val))
        key_str = '|'.join(key_parts)

        if key_str in seen:
            duplicates.append(row)
            if strategy == 'merge_by_key':
                # 合并金额
                for dk in dedup_key:
                    if dk == 'balance' and dk in row:
                        existing = next(r for r in deduped if '|'.join(
                            str(r.get(k2, '')) for k2 in dedup_key if k2 != 'balance'
                        ) == '|'.join(str(row.get(k2, '')) for k2 in dedup_key if k2 != 'balance'))
                        existing[dk] = existing.get(dk, 0) + row.get(dk, 0)
            # keep_first: 跳过重复
            continue

        seen.add(key_str)
        deduped.append(row)

    removed = len(data_rows) - len(deduped)
    if removed > 0:
        print(f'  [DT-156] 去重: {len(data_rows)}行→{len(deduped)}行(移除{removed}行重复, dedup_key={dedup_key})')

    return deduped


def _adapt_dedup_keys(data_rows, dedup_key):
    """适配dedup_key：当data_rows中缺少dedup_key指定的字段时，从已有字段映射。

    典型场景：subjects.json的secondary数据有code/name/balance字段，
    但schema要求的dedup_key是bank_name/account_no。

    映射规则：
    - bank_name ← name (银行/单位名称)
    - account_no ← code (账号/科目编码)
    - name ← code (结算对象名称)
    """
    if not data_rows or not dedup_key:
        return

    first_row = data_rows[0]
    missing_keys = [k for k in dedup_key if k not in first_row]

    if not missing_keys:
        return

    # 建立映射
    key_mapping = {
        'bank_name': 'name',
        'account_no': 'code',
    }

    for row in data_rows:
        for missing_key in missing_keys:
            if missing_key not in row:
                source_key = key_mapping.get(missing_key)
                if source_key and source_key in row:
                    row[missing_key] = row[source_key]


def load_subject_data(sheet_name, cache_dir, schema_path=None):
    """统一数据加载入口

    从subject_schema.json读取科目配置，按配置加载、去重、校验数据。

    Args:
        sheet_name: Sheet名称（如'5-5应付账款'）
        cache_dir: _dt_cache/目录路径
        schema_path: Schema文件路径（可选）

    Returns:
        dict: {
            'data_rows': list,         # 去重后的数据行
            'reconcile_target': float,  # 勾稽目标值
            'config': dict,            # 科目配置
            'load_report': dict,       # 加载报告
        }
    """
    # Step 1: 加载Schema
    schema = load_subject_schema(schema_path)
    config = get_subject_config(schema, sheet_name)

    if not config:
        return {
            'data_rows': [],
            'reconcile_target': None,
            'config': None,
            'load_report': {
                'status': 'no_schema',
                'message': f'Sheet {sheet_name} 未在subject_schema.json中定义'
            }
        }

    load_report = {
        'status': 'loaded',
        'sheet_name': sheet_name,
        'primary_source': config.get('primary_source'),
        'dedup_removed': 0,
        'reconcile_diff': None,
        'warnings': []
    }

    # Step 2: 加载primary_source数据
    primary_file = config.get('primary_source')
    primary_data = load_cache_data(cache_dir, primary_file)

    if primary_data is None:
        load_report['status'] = 'primary_missing'
        load_report['warnings'].append(
            f'DT-0: 主数据源{primary_file}未找到'
        )
        # 不直接返回空，继续尝试加载secondary_source
        # 当primary_source缺失但有secondary_source时，用secondary数据填充
        if not config.get('secondary_source'):
            return {
                'data_rows': [],
                'reconcile_target': None,
                'config': config,
                'load_report': load_report
            }
        # primary缺失，尝试从secondary加载
        secondary_data = load_cache_data(cache_dir, config['secondary_source'])
        if secondary_data:
            data_rows = _filter_secondary(secondary_data, config)
            if data_rows and config.get('dedup_key'):
                _adapt_dedup_keys(data_rows, config['dedup_key'])
                data_rows = dedup_data(data_rows, config['dedup_key'], config.get('dedup_strategy', 'keep_first'))
            if data_rows:
                reconcile_target = get_reconcile_target(schema, config, cache_dir)
                load_report['status'] = 'secondary_only'
                load_report['warnings'].append(
                    f'ℹ️ 主数据源缺失，已从secondary_source({config["secondary_source"]})加载{len(data_rows)}行数据'
                )
                return {
                    'data_rows': data_rows,
                    'reconcile_target': reconcile_target,
                    'config': config,
                    'load_report': load_report
                }
        return {
            'data_rows': [],
            'reconcile_target': None,
            'config': config,
            'load_report': load_report
        }

    # Step 3: 按source_code_prefix或source_filter筛选数据
    data_rows = _filter_data(primary_data, config, primary_file, cache_dir)


    # ── DT-PROTECT: 备抵科目硬保护 ──
    # 防止坏账准备(1231)/累计折旧(1602)/减值准备(1603)/累计摊销(1702)
    # 被当作普通数据行写入明细表。这些科目应由 has_bad_debt / contra_account 处理。
    CONTRA_CODES_BLACKLIST = {'1231', '1602', '1603', '1702'}
    before_protect = len(data_rows)
    data_rows = [
        r for r in data_rows
        if str(r.get('code', r.get('subject_code', '')))[:4] not in CONTRA_CODES_BLACKLIST
    ]
    removed = before_protect - len(data_rows)
    if removed > 0:
        load_report['warnings'].append(
            f'🛡️ 备抵科目保护: 已从数据行中移除{removed}行'
            f'(codes={CONTRA_CODES_BLACKLIST})，这些科目由has_bad_debt/contra_account单独处理'
        )

    # ── DT-PROTECT: 空数据源检测 ──
    # 如果source_code_prefix在科目余额表中找不到数据，记录警告
    prefix = config.get('source_code_prefix')
    if prefix and len(data_rows) == 0:
        if isinstance(prefix, str):
            prefix = [prefix]
        load_report['warnings'].append(
            f'ℹ️ source_code_prefix={prefix} 在科目余额表中无数据，'
            f'Sheet将保持空白（E列为0）'
        )

    # Step 4: 去重（DT-156）
    dedup_key = config.get('dedup_key')
    if dedup_key:
        before_count = len(data_rows)
        data_rows = dedup_data(data_rows, dedup_key, config.get('dedup_strategy', 'keep_first'))
        load_report['dedup_removed'] = before_count - len(data_rows)

    # Step 5: 补充secondary_source（如需）
    # 修复：当primary_source缺失时应仍加载secondary_source数据
    # DT-134新增: 当data_template=bank_deposit且有pdf_extractions数据时，
    # 用PDF逐行数据替换subjects.json汇总行（DT-104: 逐行展开禁止汇总）
    is_bank_deposit_replace = (
        config.get('data_template') == 'bank_deposit'
        and config.get('secondary_source') == 'pdf_extractions.json'
    )

    if config.get('secondary_source') and config.get('dedup_strategy') != 'primary_only_if_reconciled':
        # 仅当策略允许时补充
        secondary_data = load_cache_data(cache_dir, config['secondary_source'])
        if secondary_data:
            secondary_rows = _filter_secondary(secondary_data, config)
            if secondary_rows:
                # DT-134: 银行存款——PDF逐行数据替换subjects.json汇总行
                if is_bank_deposit_replace and len(secondary_rows) > 1:
                    # 用PDF逐行数据替换primary汇总行（避免汇总+明细混排）
                    primary_total = sum(abs(r.get('balance', 0)) for r in data_rows)
                    secondary_total = sum(abs(r.get('balance', 0)) for r in secondary_rows)
                    print(f'  [DT-134] 银行存款: primary汇总{len(data_rows)}行({primary_total:,.2f}) '
                          f'→ PDF逐行{len(secondary_rows)}行({secondary_total:,.2f})')
                    data_rows = secondary_rows
                    load_report['status'] = 'secondary_replaced'
                    load_report['warnings'].append(
                        f'ℹ️ 银行存款已从pdf_extractions.json加载{len(secondary_rows)}个账户逐行数据(DT-104/DT-134)'
                    )
                else:
                    # 当primary_source缺失时，secondary_rows直接作为数据源
                    # 需要将subjects.json格式的数据适配为dedup_key可识别的格式
                    if config.get('dedup_key') and secondary_rows:
                        # 检查secondary_rows是否包含dedup_key指定的字段
                        first_row = secondary_rows[0]
                        missing_dedup_keys = [k for k in config['dedup_key'] if k not in first_row]
                        if missing_dedup_keys:
                            # dedup_key字段缺失，将subjects.json的code+name映射为dedup_key格式
                            # 例如：bank_name → name中的银行名, account_no → code中的账号
                            for row in secondary_rows:
                                if 'bank_name' in missing_dedup_keys and 'name' in row:
                                    row['bank_name'] = row['name']
                                if 'account_no' in missing_dedup_keys and 'code' in row:
                                    row['account_no'] = row['code']
                    before_count = len(data_rows)
                    data_rows.extend(secondary_rows)
                    data_rows = dedup_data(data_rows, dedup_key, config.get('dedup_strategy', 'keep_first'))
                    load_report['dedup_removed'] += before_count + len(secondary_rows) - len(data_rows)

    # Step 6: 获取勾稽目标值（DT-158: 即时勾稽基础）
    reconcile_target = get_reconcile_target(schema, config, cache_dir)

    if reconcile_target is not None and data_rows:
        # 计算数据行合计
        book_value_key = 'balance'
        if data_rows and 'balance' in data_rows[0]:
            book_value_key = 'balance'
        elif data_rows and 'balance' in data_rows[0] and 'code' in data_rows[0]:
            book_value_key = 'balance'
        elif data_rows and '期末余额' in data_rows[0]:
            book_value_key = '期末余额'

        total = sum(abs(r.get(book_value_key, 0)) for r in data_rows)
        diff = abs(total - abs(reconcile_target))
        load_report['reconcile_diff'] = diff

        if diff > 1.0:
            load_report['warnings'].append(
                f'DT-158: Sheet {sheet_name} 数据合计={total:,.2f} '
                f'vs 勾稽目标={reconcile_target:,.2f}，差额={diff:,.2f}'
            )

    # Step 7: 生成数据指纹（DT-155: 幂等保护基础）
    fingerprint = _compute_data_fingerprint(data_rows)
    load_report['fingerprint'] = fingerprint

    return {
        'data_rows': data_rows,
        'reconcile_target': reconcile_target,
        'config': config,
        'load_report': load_report
    }


def _match_filter_rule(filter_rule, name):
    """执行filter_rule分类规则匹配（DT-FR1实现）

    filter_rule格式:
    - "名称含'股票'" → 名称包含'股票'
    - "名称含'房屋'或'建筑物'或'厂房'" → 名称包含任一关键词
    - "默认(无法分类时填入此Sheet)" → 始终匹配（兜底Sheet）
    - 空字符串 → 无filter_rule，始终匹配（向后兼容）

    Args:
        filter_rule: filter_rule规则字符串
        name: 科目名称

    Returns:
        bool: 是否匹配该规则
    """
    if not filter_rule:
        return True  # 无filter_rule时始终匹配（向后兼容）

    rule = filter_rule.strip()

    # "默认"规则：兜底Sheet，始终匹配
    if rule.startswith('默认'):
        return True

    # "名称含'XX'" 或 "名称含'XX'或'YY'或'ZZ'" 格式解析
    import re
    # DT-contra: 检测"(默认)"后缀——当关键词不匹配时，含"(默认)"标记的Sheet也应匹配（兜底）
    is_default_suffix = '(默认)' in rule or '（默认）' in rule

    # 提取"名称含"后面的所有单引号内的关键词
    m = re.match(r'名称含(.+)', rule)
    if m:
        keywords_part = m.group(1)
        # 提取所有单引号内的关键词
        keywords = re.findall(r"'([^']+)'", keywords_part)
        if keywords:
            # 任一关键词匹配即通过
            if any(kw in name for kw in keywords):
                return True
            # DT-contra: 关键词不匹配但有"(默认)"标记时，也匹配（兜底Sheet）
            if is_default_suffix:
                return True
            return False
        # 也支持无引号的关键词（如"名称含房屋或建筑物"）
        bare_keywords = re.split(r'或', keywords_part.strip())
        matched = any(kw.strip() in name for kw in bare_keywords if kw.strip() and kw.strip() not in ('(默认)', '（默认）'))
        if matched:
            return True
        if is_default_suffix:
            return True
        return False

    # 未识别的filter_rule格式，默认匹配（向后兼容）
    return True


def _filter_data(raw_data, config, source_file, cache_dir):
    """按科目配置筛选数据

    Args:
        raw_data: 原始JSON数据
        config: 科目配置
        source_file: 数据源文件名
        cache_dir: 缓存目录

    Returns:
        list: 筛选后的数据行
    """
    rows = []

    # subjects.json格式: {subjects: [{code, name, balance, ...}]} 或 [{code, name, balance, ...}] (顶层list)
    if source_file == 'subjects.json':
        subjects = raw_data
        if isinstance(raw_data, dict):
            # dict格式: 优先取'subjects'键，否则取'data'键，否则把dict值当list
            subjects = raw_data.get('subjects', raw_data.get('data', []))
            if isinstance(subjects, dict):
                # 嵌套dict: 再试一层
                subjects = subjects.get('subjects', subjects.get('data', []))
        # 此时subjects应该是list
        if not isinstance(subjects, list):
            print(f'  [WARN] subjects.json格式异常: type={type(subjects).__name__}, 期望list')
            return rows

        prefix = config.get('source_code_prefix')
        if isinstance(prefix, str):
            prefix = [prefix]
        elif prefix is None:
            prefix = []

        # BUG-5修复：排除所有非末级行（汇总行和中间节点）
        # 规则1：子编码检测 — 如果有更长的子code以当前code开头，则为父级
        # 规则2：同一code多行检测 — 同一code出现多次时，第一次出现是汇总行（辅助核算格式）
        all_codes = sorted(str(s.get('code', '')) for s in subjects)
        codes_with_children = set()
        
        # 规则1: 子编码检测（如 1122 → 112201）
        for i, code in enumerate(all_codes):
            for j in range(i + 1, len(all_codes)):
                if all_codes[j].startswith(code) and all_codes[j] != code:
                    codes_with_children.add(code)
                    break
        
        # 规则2: 同一code多行检测（辅助核算格式：code相同但名称不同）
        # 同一code多行时，跳过名称不含'*'或'　'的行（汇总行），保留明细行
        # 这些汇总行的余额=子行余额之和，不应重复计入
        # 注：汇总行特征是名称较短且不含'*'前缀/首字符缩进
        from collections import defaultdict as _dd
        code_summary_names = _dd(set)  # code → 需要跳过的汇总行名称
        code_rows = _dd(list)
        for s in subjects:
            c = str(s.get('code', ''))
            if c and c[0].isdigit():
                code_rows[c].append(s.get('name', '').strip())
        for c, names in code_rows.items():
            unique_names = set(n for n in names if n)
            if len(unique_names) > 1:
                # 第一行（不含*前缀且名称最短的）通常是汇总行
                summary_names = [n for n in unique_names if not n.startswith('*') and not n.startswith('　')]
                if summary_names:
                    if len(summary_names) == len(unique_names):
                        # DT-212: 所有名称都无*前缀时，只取最短名称为汇总行（如银行存款→银行明细户）
                        code_summary_names[c] = {sorted(summary_names, key=len)[0]}
                    else:
                        # 取所有无*前缀的名称作为汇总行
                        code_summary_names[c] = set(summary_names)
                else:
                    # 全部都带*，取第一个作为汇总（不应出现这种情况）
                    code_summary_names[c] = {sorted(unique_names, key=len)[0]}

        # DT-FR1: 实现filter_rule分类逻辑
        # filter_rule格式: "名称含'股票'" / "名称含'房屋'或'建筑物'或'厂房'" / "默认(无法分类时填入此Sheet)"
        filter_rule = config.get('filter_rule', '')

        # DT-FR1: 对"默认"规则，需收集同prefix组其他sheet的具体filter_rule，
        # 以便排除已被其他规则匹配的数据行（互斥分配）
        is_default_rule = filter_rule and filter_rule.strip().startswith('默认')
        sibling_rules = []  # 同prefix组其他sheet的非默认filter_rule列表
        if is_default_rule:
            schema = load_subject_schema()
            if schema:
                subjects_schema = schema.get('subjects', schema)
                my_prefix = config.get('source_code_prefix', '')
                if isinstance(my_prefix, list):
                    my_prefix = my_prefix[0] if my_prefix else ''
                for _sname, _scfg in subjects_schema.items():
                    _sp = _scfg.get('source_code_prefix', '')
                    if isinstance(_sp, list):
                        _sp = _sp[0] if _sp else ''
                    if _sp == my_prefix and _scfg.get('filter_rule', '') != filter_rule:
                        _fr = _scfg.get('filter_rule', '')
                        if _fr and not _fr.strip().startswith('默认'):
                            sibling_rules.append(_fr)

        for s in subjects:
            code = str(s.get('code', ''))
            if any(code.startswith(str(p)) for p in prefix):
                # 排除非末级行：有子code的行都是汇总/中间节点
                # DT-contra例外：如果科目有contra_account配置（如固定资产1601→1602），
                # 即使只有一级汇总行也要保留，因为contra_account联动需要用汇总行原值
                has_contra = bool(config.get('contra_account'))
                # 规则1：子编码检测（code在codes_with_children中）
                # DT-FIX: 即使有contra_account配置，当父级和子级同时存在时也排除父级，
                # 否则会导致数据行重复计算(如1701+1701.01同时写入，1701已包含1701.01的余额)
                if code in codes_with_children:
                    continue
                if not has_contra:
                    # 规则2：同一code多行检测（汇总行名称匹配）
                    if code in code_summary_names:
                        name = str(s.get('name', '')).strip()
                        if name in code_summary_names[code]:
                            continue
                # DT-FR1: 执行filter_rule过滤
                if filter_rule:
                    name = s.get('name', '')
                    matched = _match_filter_rule(filter_rule, name)
                    if not matched:
                        continue
                    # DT-FR1: 对"默认"规则，排除已被兄弟规则匹配的行（互斥分配）
                    if is_default_rule and sibling_rules:
                        claimed = False
                        for sib_rule in sibling_rules:
                            if _match_filter_rule(sib_rule, name):
                                claimed = True
                                break
                        if claimed:
                            continue
                rows.append({
                    'code': code,
                    'name': s.get('name', ''),
                    'balance': s.get('balance', 0),
                    'direction': s.get('direction', s.get('closing_direction', '')),
                })

    # fixed_assets.json格式: {items: [{asset_basics, financial_records}]}
    elif source_file == 'fixed_assets.json' and isinstance(raw_data, dict):
        items = raw_data.get('items', [])
        category_filter = config.get('source_filter', {})
        category = category_filter.get('category', '')

        for item in items:
            basics = item.get('asset_basics', {})
            if category and category not in basics.get('category', ''):
                continue
            records = item.get('financial_records', {})
            rows.append({
                'asset_name': basics.get('name', ''),
                'category': basics.get('category', ''),
                'spec': basics.get('spec', ''),
                'manufacturer': basics.get('manufacturer', ''),
                'quantity': basics.get('quantity', 1),
                'original_value': records.get('original_value', 0),
                'net_value': records.get('net_value', 0),
                'accumulated_depreciation': records.get('accumulated_depreciation', 0),
                'code': basics.get('code', ''),
            })

    # bank_statement_extractions.json格式
    elif source_file == 'bank_statement_extractions.json' and isinstance(raw_data, dict):
        extractions = raw_data.get('extraction_results', [])
        for ext in extractions:
            rows.append({
                'bank_name': ext.get('bank_name', ''),
                'account_no': ext.get('account_no', ''),
                'balance': ext.get('ending_balance', 0),
                'name': f"{ext.get('bank_name', '')}-{ext.get('account_no', '')}",
            })

    # pdf_extractions.json格式 (DT-134: 银行对账单结构化提取)
    # 结构: {bank_statements: {records: [{bank_name, account_no, balance, type, ...}]}}
    elif source_file == 'pdf_extractions.json' and isinstance(raw_data, dict):
        bank_stmts = raw_data.get('bank_statements', {})
        records = bank_stmts.get('records', []) if isinstance(bank_stmts, dict) else []
        for rec in records:
            rows.append({
                'bank_name': rec.get('bank_name', ''),
                'account_no': rec.get('account_no', ''),
                'account_name': rec.get('account_name', ''),
                'balance': rec.get('balance', 0),
                'type': rec.get('type', ''),
                'name': f"{rec.get('bank_name', '')}-{rec.get('account_no', '')}",
                'code': rec.get('account_no', ''),
                'direction': '借',  # 银行存款为资产类，正常方向为借方
            })

    # auxiliary_balance.json / auxiliary_balance_all.json
    elif 'auxiliary' in source_file and isinstance(raw_data, dict):
        filter_kw = config.get('secondary_filter', '')
        for key, objects in raw_data.items():
            if filter_kw and filter_kw not in key:
                continue
            if isinstance(objects, list):
                rows.extend(objects)

    return rows


def _filter_secondary(secondary_data, config):
    """筛选secondary_source数据

    修复：当secondary_source为subjects.json（list格式）时，
    按secondary_filter（科目编码前缀）筛选末级子科目数据。

    DT-134新增: 当secondary_source为pdf_extractions.json时，
    提取bank_statements.records并转为标准行格式。
    """
    if not secondary_data:
        return []

    filter_kw = config.get('secondary_filter', '')
    rows = []

    # DT-134: pdf_extractions.json格式（银行对账单提取结果）
    if isinstance(secondary_data, dict) and 'bank_statements' in secondary_data:
        bank_stmts = secondary_data.get('bank_statements', {})
        if isinstance(bank_stmts, dict):
            records = bank_stmts.get('records', [])
            for rec in records:
                rows.append({
                    'bank_name': rec.get('bank_name', ''),
                    'account_no': rec.get('account_no', ''),
                    'account_name': rec.get('account_name', ''),
                    'balance': rec.get('balance', 0),
                    'type': rec.get('type', ''),
                    'name': f"{rec.get('bank_name', '')}-{rec.get('account_no', '')}",
                    'code': rec.get('account_no', ''),
                    'direction': '借',  # 银行存款为资产类
                })
        return rows

    if isinstance(secondary_data, dict):
        for key, objects in secondary_data.items():
            if filter_kw and filter_kw not in key:
                continue
            if isinstance(objects, list):
                rows.extend(objects)
    elif isinstance(secondary_data, list):
        # subjects.json格式: [{code, name, balance, direction, ...}]
        if filter_kw:
            # 按科目编码前缀筛选末级子科目（排除汇总行）
            for s in secondary_data:
                code = str(s.get('code', ''))
                name = str(s.get('name', ''))
                # 匹配前缀但排除汇总行（汇总行code长度等于前缀长度或含有_分隔的上级编码）
                if code.startswith(str(filter_kw)):
                    # 判断是否为末级科目：code比filter_kw长，或name不含下划线分隔的子科目标记
                    is_leaf = len(code) > len(str(filter_kw))
                    # 对于像'100201'这种6位编码，比'1002'长，说明是末级
                    # 对于'1122_01.004'这种辅助核算编码，比'1122'长，说明是末级
                    if is_leaf or '_' in code:
                        balance = s.get('balance', 0)
                        direction = s.get('direction', s.get('closing_direction', ''))
                        rows.append({
                            'code': code,
                            'name': name,
                            'balance': balance,
                            'direction': direction,
                        })
        else:
            rows = list(secondary_data)

    return rows


def _compute_data_fingerprint(data_rows):
    """计算数据行指纹（用于幂等保护检测）

    Args:
        data_rows: 数据行列表

    Returns:
        str: MD5指纹
    """
    content = json.dumps(data_rows, ensure_ascii=False, sort_keys=True)
    return hashlib.md5(content.encode()).hexdigest()


def check_idempotent(cache_dir, sheet_name, new_fingerprint):
    """检查幂等性——如果指纹一致则说明数据未变化，跳过重复写入

    Args:
        cache_dir: 缓存目录
        sheet_name: Sheet名称
        new_fingerprint: 新数据指纹

    Returns:
        bool: True=数据已变化需写入, False=数据未变化可跳过
    """
    fp_path = os.path.join(cache_dir, '_data_fingerprints.json')
    fingerprints = {}
    if os.path.exists(fp_path):
        with open(fp_path, 'r', encoding='utf-8') as f:
            fingerprints = json.load(f)

    old_fp = fingerprints.get(sheet_name)
    if old_fp == new_fingerprint:
        print(f'  [DT-155] 幂等检测: {sheet_name} 数据指纹未变化，跳过重复写入')
        return False

    # 更新指纹
    fingerprints[sheet_name] = new_fingerprint
    with open(fp_path, 'w', encoding='utf-8') as f:
        json.dump(fingerprints, f, ensure_ascii=False, indent=2)

    return True


# ============================================================
# 序时账数据加载 (DT-161/L1保障: 有序时账时MUST加载并传入fill_sheet)
# ============================================================

def load_journal_data(journal_path, project_dir=None):
    """加载序时账数据，提取每个结算对象的发生日期和业务内容。

    DT-161保障：有序时账时，Phase 2e MUST执行。本函数是L1保障的数据加载层，
    将序时账数据解析为 journal_data dict，供 fill_sheet(journal_data=...) 使用。

    ⚠️ 不同项目的序时账导出格式不同，列号/列名可能完全不同！
    因此本函数采用"表头自动检测"策略而非硬编码列号：
    1. 读取序时账第1行（表头行），按列名关键词匹配关键列
    2. 匹配失败则报错，不静默跳过

    Args:
        journal_path: 序时账Excel文件路径
        project_dir: 项目目录（用于查找_dt_cache/存储解析结果）

    Returns:
        dict: {
            'column_map': {关键字段: 列序号},  # 检测到的列映射
            'counterparty_data': {结算对象名: {'date': datetime, 'business': str}},
            'raw_count': int,  # 总行数
            'matched_count': int,  # 匹配到往来科目的行数
            'warnings': list,
        }

    Raises:
        FileNotFoundError: 序时账文件不存在
        ValueError: 关键列无法识别（日期/科目名称/摘要/借方/贷方）
    """
    import openpyxl
    from datetime import datetime, timedelta
    from collections import defaultdict

    if not os.path.exists(journal_path):
        raise FileNotFoundError(f'DT-161: 序时账文件不存在({journal_path})')

    wb = openpyxl.load_workbook(journal_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # ---- Step 1: 自动检测列映射（DT-208: 优先source_header_parser动态定位表头行） ----
    # DT-208: 优先使用source_header_parser（支持多行扫描+双行表头+全角归一化）
    header_row_num = 1
    col_map = {}

    try:
        # 动态补充sys.path（dt_runner已添加，独立调用时可能缺失）
        import sys as _sys
        _shp_dir = os.path.normpath(os.path.join(
            os.path.dirname(os.path.abspath(__file__)), '..', '..', 'valuation-common', 'scripts'
        ))
        if os.path.isdir(_shp_dir) and _shp_dir not in _sys.path:
            _sys.path.insert(0, _shp_dir)
        from source_header_parser import locate_header_row as _locate_header_row
        _hr, _cm = _locate_header_row(ws, 'journal')
        if _hr > 0 and _cm:
            header_row_num = _hr
            # source_header_parser键名→data_loader中文键名映射
            shp_to_dl = {
                'date': '日期',
                'code': '科目编码',
                'name': '科目名称',
                'summary': '摘要',
                'debit': '借方金额',
                'credit': '贷方金额',
                'settlement': '结算对象',
                'voucher_no': '凭证字号',
            }
            for shp_key, dl_key in shp_to_dl.items():
                if shp_key in _cm:
                    col_map[dl_key] = _cm[shp_key]
    except ImportError:
        pass

    # 兜底：旧逻辑（假设第1行是表头+关键词子串匹配）
    if not col_map:
        HEADER_KEYWORDS = {
            '日期': ['日期', 'date', '凭证日期', '记账日期', '业务日期'],
            '摘要': ['摘要', 'summary', '业务摘要', '凭证摘要'],
            '科目编码': ['科目编码', '科目代码', '科目号', 'account code'],
            '科目名称': ['科目名称', '科目', 'account name'],
            '借方金额': ['借方金额', '借方', '借方本币', 'debit'],
            '贷方金额': ['贷方金额', '贷方', '贷方本币', 'credit'],
        }

        for col_idx in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=1, column=col_idx).value
            if not val:
                continue
            val_str = str(val).strip()
            for field, keywords in HEADER_KEYWORDS.items():
                if field in col_map:
                    continue
                for kw in keywords:
                    if kw.lower() in val_str.lower():
                        col_map[field] = col_idx
                        break

    # 必须检测到5个关键列
    required_fields = ['日期', '摘要', '科目名称', '借方金额', '贷方金额']
    missing = [f for f in required_fields if f not in col_map]
    if missing:
        # 输出表头信息帮助诊断
        header_info = {}
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=1, column=c).value
            if v:
                header_info[c] = str(v)[:30]
        raise ValueError(
            f'DT-51① CRITICAL: 序时账列映射检测失败！缺少关键列: {missing}\n'
            f'检测到的列映射: {col_map}\n'
            f'表头行内容: {header_info}\n'
            f'请确认序时账导出格式，或手动指定列映射。'
        )

    # ---- Step 2: 读取序时账数据（DT-51② + DT-54日期多格式） ----
    # 往来科目编码前缀
    RE_SUBJECT_PREFIXES = ['1122', '1123', '1221', '1461', '2202', '2203', '2241']

    date_col = col_map['日期']
    summary_col = col_map['摘要']
    subj_name_col = col_map['科目名称']
    debit_col = col_map['借方金额']
    credit_col = col_map['贷方金额']

    # 科目编码列可能不存在，从科目名称中提取
    subj_code_col = col_map.get('科目编码')

    # 存储每个结算对象的全部序时账记录
    # key: 结算对象名(从科目名称中提取), value: list of records
    counterparty_records = defaultdict(list)
    raw_count = 0
    matched_count = 0
    warnings = []

    # DT-208: 使用source_header_parser返回的header_row_num，不再逐行扫描
    # 兜底：如果source_header_parser未命中，仍用旧逻辑检测数据起始行
    if header_row_num > 1:
        data_start_row = header_row_num + 1
    else:
        data_start_row = 2  # 默认从第2行开始
        for r in range(1, min(ws.max_row + 1, 10)):
            val = ws.cell(row=r, column=subj_name_col).value
            if val and ('科目名称' in str(val) or '科目' in str(val)):
                data_start_row = r + 1
                break
            # 如果第1行就是数据（非表头），则从第1行开始
            if val and str(val).strip() not in ('', 'None') and '科目' not in str(val):
                data_start_row = r
                break

    for r in range(data_start_row, ws.max_row + 1):
        raw_count += 1

        # 读取科目名称
        subj_name = ws.cell(row=r, column=subj_name_col).value
        if not subj_name:
            continue
        subj_name = str(subj_name).strip()

        # 判断是否往来科目（从科目编码或科目名称判断）
        is_reciprocity = False
        if subj_code_col:
            code = str(ws.cell(row=r, column=subj_code_col).value or '').strip()
            for prefix in RE_SUBJECT_PREFIXES:
                if code.startswith(prefix):
                    is_reciprocity = True
                    break
        else:
            # 无科目编码列时，从科目名称关键词判断
            re_keywords = ['应收账款', '预付账款', '其他应收',
                          '合同资产', '应付账款', '预收账款', '其他应付']
            for kw in re_keywords:
                if kw in subj_name:
                    is_reciprocity = True
                    break

        if not is_reciprocity:
            continue

        # 提取结算对象名（从科目名称中，分隔符通常为_或\\）
        cp_name = subj_name
        for sep in ['_', '\\\\', '/']:
            if sep in cp_name:
                parts = cp_name.split(sep)
                cp_name = parts[-1].strip() if len(parts) >= 2 else cp_name
                break

        # 读取日期（DT-54: 多格式兼容）
        date_val = ws.cell(row=r, column=date_col).value
        dt = None
        if isinstance(date_val, datetime):
            dt = date_val
        elif isinstance(date_val, str):
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                try:
                    dt = datetime.strptime(date_val.strip()[:10], fmt[:8] if '年' in fmt else fmt)
                    break
                except ValueError:
                    continue
        elif isinstance(date_val, (int, float)):
            try:
                dt = datetime(1899, 12, 30) + timedelta(days=int(date_val))
            except:
                pass

        # 读取摘要
        summary = str(ws.cell(row=r, column=summary_col).value or '').strip()

        # 读取金额
        try:
            debit = float(ws.cell(row=r, column=debit_col).value or 0)
        except (TypeError, ValueError):
            debit = 0
        try:
            credit = float(ws.cell(row=r, column=credit_col).value or 0)
        except (TypeError, ValueError):
            credit = 0

        matched_count += 1
        counterparty_records[cp_name].append({
            'date': dt,
            'summary': summary,
            'debit': debit,
            'credit': credit,
            'subject_name': subj_name,
        })

    # ---- Step 3: 归纳每个结算对象的发生日期和业务内容 ----
    # 发生日期: 取该结算对象最旧的一笔日期（DT-51④: 资产类取借方末笔/负债类取贷方末笔，
    #          此处取最旧日期作为近似，精确匹配需Phase 2e关键词匹配）
    # 业务内容: 从摘要中归纳（DT-60: 取高频业务关键词）
    BIZ_KEYWORDS = {
        '货款': ['货款', '收货款', '发货', '出货', '采购', '进货'],
        '服务费': ['服务费', '技术服务', '咨询费', '管理费', '开发费'],
        '工程款': ['工程款', '工程', '施工', '建设', '安装', '装修'],
        '租金': ['租金', '租赁', '房租', '场地费'],
        '保证金': ['保证金', '押金', '投标保证金'],
        '报销款': ['报销', '差旅', '办公费', '通讯费'],
        '社保': ['社保', '公积金', '五险一金', '养老', '医疗'],
        '税费': ['增值税', '所得税', '税', '附加'],
        '借款': ['借款', '贷款', '融资', '利息'],
        '往来款': ['往来', '划款', '调拨', '内部'],
    }

    counterparty_data = {}
    for cp_name, records in counterparty_records.items():
        # 发生日期：取有效日期中最旧的
        valid_dates = [r['date'] for r in records if r['date'] is not None]
        earliest_date = min(valid_dates) if valid_dates else None

        # 业务内容：从摘要中归纳（DT-60）
        summaries = [r['summary'] for r in records if r['summary']]
        inferred_biz = None
        if summaries:
            # 统计每个业务关键词的匹配次数
            keyword_counts = defaultdict(int)
            for summary in summaries:
                for biz_type, patterns in BIZ_KEYWORDS.items():
                    for p in patterns:
                        if p in summary:
                            keyword_counts[biz_type] += 1
                            break
            if keyword_counts:
                inferred_biz = max(keyword_counts, key=keyword_counts.get)
            else:
                # 取最短摘要的前6字
                shortest = min(summaries, key=len)
                inferred_biz = shortest[:6]

        if earliest_date or inferred_biz:
            counterparty_data[cp_name] = {
                'date': earliest_date,
                'business': inferred_biz or '',
                'record_count': len(records),
            }

    wb.close()

    # ---- Step 4: 缓存结果（DT-130） ----
    if project_dir:
        cache_dir = os.path.join(project_dir, '_dt_cache')
        os.makedirs(cache_dir, exist_ok=True)
        cache_data = {
            '_meta': {
                'rule': 'DT-161',
                'created_at': str(datetime.now()),
                'source': 'load_journal_data',
                'journal_path': str(journal_path),
                'column_map': col_map,
            },
            'counterparty_data': {
                k: {
                    'date': v['date'].strftime('%Y-%m-%d') if isinstance(v['date'], datetime) else str(v['date']),
                    'business': v['business'],
                    'record_count': v['record_count'],
                }
                for k, v in counterparty_data.items()
            },
            'raw_count': raw_count,
            'matched_count': matched_count,
            'warnings': warnings,
        }
        cache_path = os.path.join(cache_dir, 'journal_data.json')
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)

    return {
        'column_map': col_map,
        'counterparty_data': counterparty_data,
        'raw_count': raw_count,
        'matched_count': matched_count,
        'warnings': warnings,
    }


# ============================================================
# CLI入口
# ============================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='data_loader.py — 统一数据加载器')
    parser.add_argument('--sheet', required=True, help='Sheet名称')
    parser.add_argument('--cache-dir', required=True, help='_dt_cache/目录路径')
    parser.add_argument('--schema', default=None, help='subject_schema.json路径')

    args = parser.parse_args()

    result = load_subject_data(args.sheet, args.cache_dir, args.schema)

    print(f'\n=== 数据加载报告: {args.sheet} ===')
    print(f'状态: {result["load_report"]["status"]}')
    print(f'数据行数: {len(result["data_rows"])}')
    print(f'勾稽目标: {result["reconcile_target"]}')
    print(f'去重移除: {result["load_report"]["dedup_removed"]}行')
    if result['load_report']['warnings']:
        print('警告:')
        for w in result['load_report']['warnings']:
            print(f'  ⚠️ {w}')
    if result['load_report'].get('reconcile_diff'):
        print(f'勾稽差额: {result["load_report"]["reconcile_diff"]:,.2f}')
