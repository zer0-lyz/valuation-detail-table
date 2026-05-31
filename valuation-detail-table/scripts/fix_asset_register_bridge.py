#!/usr/bin/env python3
"""
fix_asset_register_bridge.py — 固定资产明细→asset_register_by_sheet.json 桥接

绕过 financial-normalizer 缺失的标准化管线，直接从原始固定资产 Excel 提取
逐行资产明细并按类别分拆到各评估明细表 Sheet。

映射规则:
  运输工具                               → 4-8-5车辆
  与生产经营活动有关的器具、工具、家具     → 4-8-4机器设备
  电子设备                               → 4-8-6电子设备
  房屋建筑物                             → 4-8-1房屋建筑物
  构筑物                                 → 4-8-2构筑物
  管道沟槽                               → 4-8-3管道沟槽
  其他/空                                → 4-8-4机器设备（默认）
"""

import json
import os
import openpyxl
import re
from datetime import datetime

# 类别名称 → Sheet 名称映射
CATEGORY_TO_SHEET = {
    # 汤浦水库等水务/基建企业特有类别
    '房屋': '4-8-1房屋建筑物',
    '建筑物': '4-8-2构筑物',
    '办公家具': '4-8-6电子设备',
    '其他办公设备': '4-8-6电子设备',
    '工具类设备': '4-8-4机器设备',
    '成套自控设备': '4-8-4机器设备',
    '其他生产设备': '4-8-4机器设备',
    '水工机械': '4-8-4机器设备',
    '水文设备': '4-8-4机器设备',
    '管线设施': '4-8-3管道沟槽',
    '船只': '4-8-5车辆',
    '其他运输工具': '4-8-5车辆',
    '运输工具': '4-8-5车辆',
    '与生产经营活动有关的器具、工具、家具': '4-8-4机器设备',
    '电子设备': '4-8-6电子设备',
    '房屋建筑物': '4-8-1房屋建筑物',
    '构筑物': '4-8-2构筑物',
    '管道沟槽': '4-8-3管道沟槽',
    '机器设备': '4-8-4机器设备',
    '车辆': '4-8-5车辆',
    '电子设备及其他': '4-8-6电子设备',
    '运输设备': '4-8-5车辆',
    '办公设备': '4-8-6电子设备',
    '房屋及建筑物': '4-8-1房屋建筑物',
    '项目设备': '4-8-4机器设备',
    # 折旧表格式：'（一）房屋建筑物' → 房屋建筑物
    '（一）房屋建筑物': '4-8-1房屋建筑物',
    '（二）机器设备': '4-8-4机器设备',
    '（三）运输设备': '4-8-5车辆',
    '（四）电子设备': '4-8-6电子设备',
    '（三）与生产经营活动有关的器具工具家具）': '4-8-4机器设备',
    '（四）飞机火车轮船以外的运输工具': '4-8-5车辆',
    '（五）电子设备': '4-8-6电子设备',
}


def _safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if str(val) == 'nan' else float(val)
    try:
        return float(str(val).replace(',', '').replace('￥', '').strip())
    except (ValueError, TypeError):
        return 0.0


def build_asset_register_by_sheet(fa_filepath: str, cache_dir: str) -> dict:
    """解析固定资产 Excel → 按 Sheet 分组的资产明细"""
    wb = openpyxl.load_workbook(fa_filepath, data_only=True)
    ws = wb.active

    # 定位表头
    header_row = None
    headers = []
    for r in range(1, min(ws.max_row + 1, 10)):
        vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else '' for c in range(1, ws.max_column + 1)]
        if '固定资产名称' in vals or '卡片编号' in vals:
            header_row = r
            headers = vals
            break

    # 如果标准表头未找到，尝试其他常见格式
    if header_row is None:
        # 格式B: 资产编码/资产名称/类别/型号/原值本币/累计折旧/净值
        for r in range(1, min(ws.max_row + 1, 5)):
            vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else '' for c in range(1, min(ws.max_column + 1, 30))]
            if '资产名称' in vals or '资产编码' in vals:
                header_row = r
                headers = vals
                print(f'  格式B: 表头行={header_row}')
                break

    # 格式C: 折旧表格式——资产编号/类别/资产描述/购置价值/至本年累计折旧/帐面价价值
    if header_row is None:
        for r in range(1, min(ws.max_row + 1, 5)):
            vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else '' for c in range(1, min(ws.max_column + 1, 30))]
            if '资产描述' in vals or ('资产编号' in vals and '类别' in vals):
                header_row = r
                headers = vals
                print(f'  格式C(折旧表): 表头行={header_row}')
                break

    if header_row is None:
        print('  ❌ 无法定位固定资产表头行')
        wb.close()
        return {}

    # 列索引
    # 列索引（精确匹配优先，排除易混淆列）
    def _find_col(exact=None, contains=None, exclude=None):
        for i, h in enumerate(headers):
            if not h:
                continue
            h_clean = h.strip()
            if exact and h_clean == exact:
                return i
            if contains:
                matched = any(kw in h_clean for kw in (contains if isinstance(contains, list) else [contains]))
                if matched and exclude:
                    if any(ex in h_clean for ex in (exclude if isinstance(exclude, list) else [exclude])):
                        continue
                if matched:
                    return i
        return None

    col_idx = {
        'asset_name': _find_col(contains='名称', exclude='类别') or _find_col(contains=['资产描述']),
        'category': _find_col(contains=['类别']),
        'spec': _find_col(contains=['规格', '型号']),
        # 优先原值本币（local currency），其次原值原币（original currency），最后原值
        'original_value': _find_col(exact='原值本币') or _find_col(exact='原值', contains='原值', exclude=['年初', '减值', '原币'])
                          or _find_col(contains=['购置价值', '购置原值']),
        'accum_dep': _find_col(contains=['本年累计折旧']) or _find_col(exact='累计折旧') or _find_col(contains=['累计折旧']),
        # 优先净额（净值-减值），其次净值（原值-折旧），最后帐面价值
        'net_value': _find_col(exact='净额') or _find_col(exact='净值') or _find_col(contains=['帐面价值', '账面价值', '净值', '帐面价']),
        'start_date': _find_col(contains=['开始使用', '使用日', '启用日', '资本化']),
        'department': _find_col(exact='使用部门'),
        'location': _find_col(exact='存放地点'),
        'dep_method': _find_col(contains=['折旧方法', '折旧法']),
        'life_month': _find_col(contains=['使用年限', '使用寿命', '折旧年限']),
        'status': _find_col(contains=['使用状况', '使用情况']),
        'asset_code': _find_col(contains=['资产编码', '固定资产编号', '卡片编号', '资产编号']),
    }

    print(f'  表头行: {header_row}')
    print(f'  列映射: {json.dumps({k: headers[v] if v is not None else None for k, v in col_idx.items()}, ensure_ascii=False)}')

    # 按类别分类
    by_sheet = {}
    for r in range(header_row + 1, ws.max_row + 1):
        # 跳过合计行（卡片编号列含"合计"标识）
        first_col_val = str(ws.cell(r, 1).value or '').strip()
        if '合计' in first_col_val:
            continue

        raw_name = ws.cell(r, col_idx['asset_name'] + 1).value if col_idx['asset_name'] is not None else None
        name = str(raw_name).strip() if raw_name else ''
        if not name or '合计' in name:
            continue

        category_str = ''
        if col_idx['category'] is not None:
            cat_val = ws.cell(r, col_idx['category'] + 1).value
            category_str = str(cat_val).strip() if cat_val else ''

        # 精确匹配优先；失败则关键词子串匹配
        sheet_name = CATEGORY_TO_SHEET.get(category_str)
        if sheet_name is None:
            for key, val in CATEGORY_TO_SHEET.items():
                if key in category_str or category_str in key:
                    sheet_name = val
                    break
        if sheet_name is None:
            sheet_name = '4-8-4机器设备'

        # 字段名必须与 dt_runner.py Phase 2 读取的 key 一致
        # dt_runner 期望: name, spec, cost(原值), depreciation(累计折旧), net_value,
        #   start_date, dept, location, dep_method, life_months, status, asset_code
        item = {
            'name': name,
            'spec': str(ws.cell(r, col_idx['spec'] + 1).value).strip() if col_idx['spec'] is not None and ws.cell(r, col_idx['spec'] + 1).value else '',
            'cost': _safe_float(ws.cell(r, col_idx['original_value'] + 1).value if col_idx['original_value'] is not None else 0),
            'depreciation': _safe_float(ws.cell(r, col_idx['accum_dep'] + 1).value if col_idx['accum_dep'] is not None else 0),
            'net_value': _safe_float(ws.cell(r, col_idx['net_value'] + 1).value if col_idx['net_value'] is not None else 0),
            'start_date': str(ws.cell(r, col_idx['start_date'] + 1).value).strip() if col_idx['start_date'] is not None and ws.cell(r, col_idx['start_date'] + 1).value else '',
            'dept': str(ws.cell(r, col_idx['department'] + 1).value).strip() if col_idx['department'] is not None and ws.cell(r, col_idx['department'] + 1).value else '',
            'location': str(ws.cell(r, col_idx['location'] + 1).value).strip() if col_idx['location'] is not None and ws.cell(r, col_idx['location'] + 1).value else '',
            'category': category_str,
            'status': str(ws.cell(r, col_idx['status'] + 1).value).strip() if col_idx['status'] is not None and ws.cell(r, col_idx['status'] + 1).value else '',
            'dep_method': str(ws.cell(r, col_idx['dep_method'] + 1).value).strip() if col_idx['dep_method'] is not None and ws.cell(r, col_idx['dep_method'] + 1).value else '',
            'life_months': str(int(_safe_float(ws.cell(r, col_idx['life_month'] + 1).value)) if col_idx['life_month'] is not None and ws.cell(r, col_idx['life_month'] + 1).value else ''),
            'asset_code': str(ws.cell(r, col_idx['asset_code'] + 1).value).strip() if col_idx['asset_code'] is not None and ws.cell(r, col_idx['asset_code'] + 1).value else '',
        }

        # 净值回退计算（字段名已改为 cost/depreciation 以匹配 dt_runner）
        if not item.get('net_value') and item.get('cost'):
            item['net_value'] = item['cost'] - item.get('depreciation', 0)

        # 检查是否为合计行（名称、编码均为空时跳过）
        if not name and not item.get('asset_code'):
            continue

        by_sheet.setdefault(sheet_name, []).append(item)

    wb.close()

    # 写入缓存
    cache_path = os.path.join(cache_dir, 'asset_register_by_sheet.json')
    os.makedirs(cache_dir, exist_ok=True)
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(by_sheet, f, ensure_ascii=False, indent=2)

    total = sum(len(v) for v in by_sheet.values())
    print(f'  ✅ asset_register_by_sheet.json: {total}项 → {cache_path}')
    for k, v in by_sheet.items():
        print(f'    {k}: {len(v)}项, 净值={sum(i["net_value"] for i in v):,.2f}')

    return by_sheet


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print('用法: python fix_asset_register_bridge.py <固定资产Excel路径> <_dt_cache目录>')
        sys.exit(1)
    build_asset_register_by_sheet(sys.argv[1], sys.argv[2])
