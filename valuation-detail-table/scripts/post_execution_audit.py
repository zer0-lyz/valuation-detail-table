#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
post_execution_audit.py — DT-160强制自检脚本（P1层防护）

在Phase 2-6执行完毕后MUST运行，检测6类已知问题。
发现CRITICAL问题→输出audit_result.json→Phase 5门控读它→CRITICAL>0则警告（不阻断交付，但MUST记录）

6类检测项：
1. 序号列非空（DT-153）
2. 坏账准备行为正数（DT-18）
3. 名称/结算对象列非空（DT-166/167）
4. 发生日期列格式正确（DT-46）
5. 边框完整性（DT-82）
6. SUM公式覆盖范围（DT-2）

版本: v1.0
创建: 2026-05-25
原因: P0全流程闭环后兜底检测，防止绕过管线后6类问题未被捕获
"""

import json
import os
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell


def run_audit(detail_file_path, cache_dir=None):
    """执行6类强制自检

    Args:
        detail_file_path: 评估明细表Excel文件路径
        cache_dir: _dt_cache目录路径（可选，用于输出audit_result.json）

    Returns:
        dict: 审计结果，含critical_count/warning_count/issues列表
    """
    if not os.path.exists(detail_file_path):
        return {'critical_count': 1, 'warning_count': 0, 'issues': [{
            'check': 'file_exists', 'severity': 'CRITICAL',
            'message': f'评估明细表不存在: {detail_file_path}'
        }]}

    wb = openpyxl.load_workbook(detail_file_path, data_only=True)

    issues = []
    total_checks = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state == 'hidden':
            continue
        # 只检查数据Sheet（3-X/4-X/5-X/6-X格式）
        import re
        if not re.match(r'^[3-6]-\d', sn):
            continue

        sheet_issues = _audit_sheet(ws, sn)
        issues.extend(sheet_issues)
        total_checks += 1

    wb.close()

    critical_count = sum(1 for i in issues if i['severity'] == 'CRITICAL')
    warning_count = sum(1 for i in issues if i['severity'] == 'WARNING')

    result = {
        'total_sheets_checked': total_checks,
        'critical_count': critical_count,
        'warning_count': warning_count,
        'total_issues': len(issues),
        'issues': issues,
    }

    # 保存审计结果
    if cache_dir:
        audit_path = os.path.join(cache_dir, 'audit_result.json')
        with open(audit_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        print(f'  审计结果已保存: {audit_path}')

    return result


def _audit_sheet(ws, sheet_name):
    """对单个Sheet执行6类检测"""
    issues = []

    # 定位关键行
    data_start = _find_data_start(ws)
    total1_row = _find_total1_row(ws)
    bd_row = _find_bad_debt_row(ws)

    if not data_start or not total1_row:
        return issues  # 非标准Sheet，跳过

    # 读取列位映射
    col_map = _detect_col_map(ws)

    # ===== 检查1: 序号列非空 (DT-153) =====
    seq_col = col_map.get('seq')
    if seq_col:
        for r in range(data_start, total1_row):
            cell = ws.cell(row=r, column=seq_col)
            if isinstance(cell, MergedCell):
                continue
            # 检查是否有名称但无序号
            name_col = col_map.get('name') or col_map.get('settlement_object')
            if name_col:
                name_val = ws.cell(row=r, column=name_col).value
                if name_val and str(name_val).strip():
                    seq_val = cell.value
                    if seq_val is None or seq_val == '':
                        issues.append({
                            'check': 'seq_not_empty',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'row': r,
                            'message': f'行{r}有结算对象但序号为空 (DT-153)'
                        })
                        break  # 每个Sheet只报一次

    # ===== 检查2: 坏账准备行为正数 (DT-18) =====
    if bd_row:
        bv_col = col_map.get('bv')
        if bv_col:
            bd_val = ws.cell(row=bd_row, column=bv_col).value
            if bd_val is not None and isinstance(bd_val, (int, float)):
                if bd_val < 0:
                    issues.append({
                        'check': 'bad_debt_positive',
                        'severity': 'CRITICAL',
                        'sheet': sheet_name,
                        'row': bd_row,
                        'message': f'坏账准备行金额为负数({bd_val:,.2f})，应为正数 (DT-18)'
                    })

    # ===== 检查3: 名称/结算对象列非空 (DT-166/167) =====
    name_col = col_map.get('name') or col_map.get('settlement_object')
    bv_col = col_map.get('bv')
    if name_col and bv_col:
        for r in range(data_start, total1_row):
            name_val = ws.cell(row=r, column=name_col).value
            bv_val = ws.cell(row=r, column=bv_col).value
            # 有金额但无名称
            if bv_val and isinstance(bv_val, (int, float)) and abs(bv_val) > 0.01:
                if not name_val or not str(name_val).strip():
                    issues.append({
                        'check': 'name_not_empty',
                        'severity': 'CRITICAL',
                        'sheet': sheet_name,
                        'row': r,
                        'message': f'行{r}有账面价值({bv_val:,.2f})但名称列为空 (DT-166/167)'
                    })

    # ===== 检查4: 发生日期列格式 (DT-46) =====
    date_col = col_map.get('occurrence_date')
    if date_col:
        for r in range(data_start, total1_row):
            date_val = ws.cell(row=r, column=date_col).value
            if date_val is not None and date_val != '':
                # 发生日期不应是纯文字（如"其他"），也不应是序号
                if isinstance(date_val, str):
                    # 检测是否为数字字符串（序号被误填到日期列）
                    if date_val.strip().isdigit():
                        issues.append({
                            'check': 'date_format',
                            'severity': 'WARNING',
                            'sheet': sheet_name,
                            'row': r,
                            'message': f'行{r}发生日期列含数字字符串"{date_val}"，疑似序号误填 (DT-46)'
                        })

    # ===== 检查5: 边框完整性 (DT-82) =====
    border_missing_count = 0
    last_print_col = _find_last_print_col(ws)
    if last_print_col:
        for r in range(data_start, min(total1_row, data_start + 50)):
            for c in range(2, last_print_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None and cell.value != '':
                    if not cell.border or not cell.border.left or cell.border.left.style is None:
                        border_missing_count += 1

        if border_missing_count > 5:  # 容忍少量缺失
            issues.append({
                'check': 'border_integrity',
                'severity': 'WARNING',
                'sheet': sheet_name,
                'message': f'数据区{border_missing_count}个单元格缺边框 (DT-82)'
            })

    # ===== 检查6: SUM公式覆盖范围 (DT-2) =====
    if total1_row:
        for c in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=total1_row, column=c)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if isinstance(val, str) and val.startswith('=SUM('):
                # 解析SUM范围
                import re
                m = re.match(r'=SUM\(([A-Z])(\d+):([A-Z])(\d+)\)', val.upper())
                if m:
                    sum_start = int(m.group(2))
                    sum_end = int(m.group(4))
                    # SUM范围应覆盖data_start到total1_row-1
                    if sum_start > data_start or sum_end < total1_row - 1:
                        issues.append({
                            'check': 'sum_range',
                            'severity': 'WARNING',
                            'sheet': sheet_name,
                            'row': total1_row,
                            'col': c,
                            'message': f'SUM范围{sum_start}:{sum_end}未覆盖数据区{data_start}:{total1_row-1} (DT-2)'
                        })

    return issues


def _find_data_start(ws):
    """定位数据起始行（检索表头1下一行）"""
    for r in range(1, min(ws.max_row + 1, 20)):
        a = ws.cell(row=r, column=1).value
        if a and '检索表头' in str(a):
            # 检查是否有检索表头2
            next_a = ws.cell(row=r + 1, column=1).value
            if next_a and '检索表头' in str(next_a):
                return r + 2
            return r + 1
    return 7  # 兜底


def _find_total1_row(ws):
    """定位合计1行"""
    for r in range(1, min(ws.max_row + 1, 80)):
        a = ws.cell(row=r, column=1).value
        if a and isinstance(a, str) and a.strip() == '合计1':
            return r
    return None


def _find_bad_debt_row(ws):
    """定位坏账准备行"""
    for r in range(1, min(ws.max_row + 1, 80)):
        a = ws.cell(row=r, column=1).value
        if a and isinstance(a, str) and '坏账准备' in a:
            return r
    return None


def _detect_col_map(ws):
    """从检索表头行检测列位映射"""
    col_map = {}
    # 读取检索表头1（Row5）
    for r in range(4, 8):
        for c in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=r, column=c).value
            if not val:
                continue
            h = str(val).strip()
            if '序号' in h:
                col_map['seq'] = c
            elif '账面价值' in h:
                col_map['bv'] = c
            elif '评估价值' in h:
                col_map['ev'] = c
            elif '结算对象' in h or '户名' in h:
                col_map['settlement_object'] = c
                col_map['name'] = c
            elif '业务内容' in h or '结算内容' in h:
                col_map['business_content'] = c
            elif '发生日期' in h:
                col_map['occurrence_date'] = c
    return col_map


def _find_last_print_col(ws):
    """找到打印范围右边界列"""
    for c in range(ws.max_column, 0, -1):
        for r in range(1, min(ws.max_row + 1, 10)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                return c
    return ws.max_column


def print_audit_report(result):
    """打印审计报告"""
    print('\n' + '='*60)
    print('P1自检报告 (DT-160强制审计)')
    print('='*60)
    print(f'检查Sheet数: {result.get("total_sheets_checked", 0)}')
    print(f'CRITICAL: {result.get("critical_count", 0)}')
    print(f'WARNING: {result.get("warning_count", 0)}')
    print(f'总问题数: {result.get("total_issues", 0)}')

    if result.get('issues'):
        print('\n问题清单:')
        for i, issue in enumerate(result['issues'], 1):
            sev = issue.get('severity', 'UNKNOWN')
            icon = '🚨' if sev == 'CRITICAL' else '⚠️'
            print(f'  {icon} [{i}] {issue.get("check", "")} | {issue.get("sheet", "")} | {issue.get("message", "")}')
    else:
        print('\n✅ 无问题')

    print('='*60)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='DT-160强制自检')
    parser.add_argument('xlsx_path', help='评估明细表Excel文件路径')
    parser.add_argument('--cache-dir', help='_dt_cache目录路径')
    args = parser.parse_args()

    result = run_audit(args.xlsx_path, args.cache_dir)
    print_audit_report(result)

    if result.get('critical_count', 0) > 0:
        print('\n⚠️ 发现CRITICAL问题！建议修复后重新运行。')
        sys.exit(1)
    else:
        print('\n✅ 自检通过。')
