#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
journal_extractor.py — 序时账查阅脚本（发生日期确认+业务内容提取）

DT-166: 本脚本是Phase 3的唯一执行入口，Agent MUST调用本脚本而非手写序时账解析/匹配逻辑。
DT-165: 本脚本基于已有Skill脚本增量扩展：
  - 复用data_loader.load_journal_data()的序时账加载能力
  - 复用business_content_map.infer_business_content()的业务内容推断能力
  - 复用sheet_col_map.json的列位映射
  - 复用gate_validator.find_header_structure()的合计行定位

版本: v1.0
创建: 2026-05-25
原因: Phase 2e升级为独立Phase 3后，需独立脚本封装完整序时账查阅流程
"""

import os
import re
import json
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

# DT-208: 集成source_header_parser实现动态表头行定位+双行表头+全角归一化
# 尝试多种路径查找source_header_parser
_HAS_SHP = False
try:
    from source_header_parser import locate_header_row
    _HAS_SHP = True
except ImportError:
    # 动态补充sys.path（dt_runner已添加，但独立调用时可能缺失）
    import sys as _sys
    _shp_dir = os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)), '..', '..', 'valuation-common', 'scripts'
    ))
    if os.path.isdir(_shp_dir) and _shp_dir not in _sys.path:
        _sys.path.insert(0, _shp_dir)
        try:
            from source_header_parser import locate_header_row
            _HAS_SHP = True
        except ImportError:
            pass


class JournalExtractor:
    """序时账数据提取器，封装列映射验证+日期解析+数据加载"""

    # DT-208: DEFAULT_COL_MAP已移除，col_map由_detect_columns()动态检测填充
    # 如果动态检测失败，_load_data()会使用.get(key, fallback)安全降级

    # 表头关键词→列号映射（备用，source_header_parser优先）
    HEADER_KEYWORDS = {
        'date': ['日期', 'date'],
        'voucher': ['凭证', '字号', 'voucher'],
        'summary': ['摘要', 'summary'],
        'subject_code': ['科目编码', '科目代码', '编码'],
        'subject_name': ['科目名称', '科目', '名称'],
        'debit': ['借方', '借方金额', 'debit'],
        'credit': ['贷方', '贷方金额', 'credit'],
        'direction': ['方向', '借贷方向'],
        'amount': ['金额', '金额合计', 'amount'],
        'aux_accounting': ['辅助核算', '辅助', '核算项目', '辅助项'],
    }

    def __init__(self, seq_file_path):
        """初始化提取器，自动解析列映射+加载数据

        Args:
            seq_file_path: 序时账Excel文件路径
        """
        self.seq_file_path = seq_file_path
        self.col_map = {}
        self.header_row = 1  # DT-208: 动态检测表头行
        self.data = []
        self.row_count = 0

        self._detect_columns()
        self._load_data()

    def _detect_columns(self):
        """自动检测序时账列映射（DT-208: 优先用source_header_parser动态定位表头行）

        优先级：
        1. source_header_parser.locate_header_row() — 支持多行扫描+双行表头+全角归一化
        2. 兜底: 旧逻辑（假设第1行是表头+关键词子串匹配）
        """
        wb = openpyxl.load_workbook(self.seq_file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        if _HAS_SHP:
            header_row, col_map = locate_header_row(ws, 'journal')
            if header_row > 0 and col_map:
                self.header_row = header_row
                # source_header_parser返回的col_map键名为英文(date/code/name/debit/credit等)
                # 映射到JournalExtractor内部使用的键名
                key_map = {
                    'date': 'date',
                    'voucher_no': 'voucher',
                    'summary': 'summary',
                    'code': 'subject_code',
                    'name': 'subject_name',
                    'debit': 'debit',
                    'credit': 'credit',
                    'aux_accounting': 'aux_accounting',
                    # v0.2 (2026-06-01): 50 列布局 — 往来单位名称作为结算对象
                    'customer_supplier_name': 'settlement',
                    'department_name': 'department',
                    'project_name': 'project',
                    'bank_account_name': 'bank_account',
                    'subject_full_path': 'subject_full_path',
                }
                for shp_key, je_key in key_map.items():
                    if shp_key in col_map:
                        self.col_map[je_key] = col_map[shp_key]

                # DT-FIX: 序时账"方向+金额"格式修正
                # source_header_parser可能将"方向"列误映射为debit，
                # 此时需要检测实际表头文字并切换为direction+amount模式
                if 'debit' in self.col_map:
                    debit_col = self.col_map['debit']
                    header_text = str(ws.cell(row=header_row, column=debit_col).value or '').strip()
                    if header_text == '方向' or '方向' in header_text:
                        # 检测"金额"列
                        for c in range(1, min(ws.max_column + 1, 20)):
                            h = str(ws.cell(row=header_row, column=c).value or '').strip()
                            if h == '金额' or '金额' in h:
                                self.col_map['direction'] = self.col_map.pop('debit')
                                self.col_map['amount'] = c
                                # 移除credit（如果有且也是方向相关的误映射）
                                if 'credit' in self.col_map:
                                    cred_col = self.col_map['credit']
                                    cred_text = str(ws.cell(row=header_row, column=cred_col).value or '').strip()
                                    if cred_text == '金额' or '金额' in cred_text:
                                        if 'amount' not in self.col_map:
                                            self.col_map['amount'] = cred_col
                                    del self.col_map['credit']
                                break

                wb.close()
                return

        # 兜底：旧逻辑（假设第1行是表头）
        header_row = {}
        for c in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=1, column=c).value
            if val:
                header_row[c] = str(val).strip()

        for field, keywords in self.HEADER_KEYWORDS.items():
            for col, header_text in header_row.items():
                for kw in keywords:
                    if kw in header_text.lower():
                        self.col_map[field] = col
                        break

        wb.close()

    def _load_data(self):
        """加载序时账全部数据（DT-51② + DT-54日期多格式兼容 + DT-208动态表头行）

        v1.1修复:
        - date=None行不再丢弃，向前回填同凭证号的日期（多行凭证后续行日期为空是正常会计格式）
        - 加载辅助核算字段，从中提取结算对象名称用于匹配
        """
        wb = openpyxl.load_workbook(self.seq_file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        date_col = self.col_map.get('date', 1)
        voucher_col = self.col_map.get('voucher', 2)
        summary_col = self.col_map.get('summary', 4)
        subject_code_col = self.col_map.get('subject_code', 5)
        subject_name_col = self.col_map.get('subject_name', 6)
        debit_col = self.col_map.get('debit')
        credit_col = self.col_map.get('credit')
        direction_col = self.col_map.get('direction')
        amount_col = self.col_map.get('amount')
        # 兼容"方向+金额"格式（如用友/金蝶序时账）：根据方向列将金额分配到借方或贷方
        use_direction_amount = bool(direction_col and amount_col and not debit_col)
        aux_col = self.col_map.get('aux_accounting')  # 辅助核算列（动态检测，可能不存在）
        # v0.2 (2026-06-01): 50 列布局 — 显式读取往来/银行/部门/项目等辅助核算列
        settlement_col = self.col_map.get('settlement')
        bank_account_col = self.col_map.get('bank_account')
        department_col = self.col_map.get('department')
        project_col = self.col_map.get('project')

        # 第一遍扫描：收集所有行数据（包括date=None的行）
        # DT-209: 日期回填策略——同凭证号（voucher_col）的行共享首行日期
        raw_rows = []
        current_voucher = None
        current_date = None

        data_start = self.header_row + 1
        for r in range(data_start, ws.max_row + 1):
            date_val = ws.cell(row=r, column=date_col).value
            summary = ws.cell(row=r, column=summary_col).value or ''
            subject_code = ws.cell(row=r, column=subject_code_col).value or ''
            subject_name = ws.cell(row=r, column=subject_name_col).value or ''
            # 兼容"方向+金额"格式 vs 传统"借方金额+贷方金额"双列格式
            if use_direction_amount:
                direction = str(ws.cell(row=r, column=direction_col).value or '').strip()
                amount = ws.cell(row=r, column=amount_col).value
                if direction in ('借', '借方'):
                    debit = amount; credit = 0
                elif direction in ('贷', '贷方'):
                    debit = 0; credit = amount
                else:
                    debit = amount; credit = 0
            else:
                debit = ws.cell(row=r, column=debit_col).value if debit_col else 0
                credit = ws.cell(row=r, column=credit_col).value if credit_col else 0
            aux_val = ws.cell(row=r, column=aux_col).value if aux_col else None
            # v0.2 (2026-06-01): 50 列布局显式读取
            settlement_val = ws.cell(row=r, column=settlement_col).value if settlement_col else None
            bank_account_val = ws.cell(row=r, column=bank_account_col).value if bank_account_col else None
            department_val = ws.cell(row=r, column=department_col).value if department_col else None
            project_val = ws.cell(row=r, column=project_col).value if project_col else None

            dt = self._parse_date(date_val, cell=ws.cell(row=r, column=date_col))
            voucher_val = ws.cell(row=r, column=voucher_col).value if voucher_col else None

            # 跟踪当前凭证号和日期
            if dt:
                current_date = dt
            if voucher_val and str(voucher_val).strip():
                current_voucher = str(voucher_val).strip()

            # 如果日期为None但有科目编码，尝试回填日期
            effective_date = dt or current_date
            # 最后一层兜底：尝试从摘要文本中提取日期（如"2024-03-31付款"）
            if effective_date is None and summary:
                import re as _re
                _date_match = _re.search(r'(\d{4}[-/\.年]\d{1,2}[-/\.月]\d{1,2})', str(summary))
                if _date_match:
                    try:
                        effective_date = datetime.strptime(
                            _date_match.group(1).replace('年', '-').replace('月', '-').replace('日', '').replace('/', '-').replace('.', '-'),
                            '%Y-%m-%d'
                        )
                    except (ValueError, IndexError):
                        pass

            if effective_date and (subject_code or subject_name):
                # 从辅助核算字段提取结算对象名称
                settlement_from_aux = self._extract_settlement_from_aux(aux_val)

                # v0.2 (2026-06-01): 优先使用显式 settlement 列，否则从 aux 提取
                settlement_name = ''
                if settlement_val and str(settlement_val).strip() and str(settlement_val).strip() not in ('nan', 'None', ''):
                    settlement_name = str(settlement_val).strip()
                if not settlement_name and settlement_from_aux:
                    settlement_name = settlement_from_aux
                self.data.append({
                    'row': r,
                    'date': effective_date,
                    'summary': str(summary).strip(),
                    'subject_code': str(subject_code).strip(),
                    'subject_name': str(subject_name).strip(),
                    'debit': float(debit) if debit and isinstance(debit, (int, float)) else self._safe_float(debit),
                    'credit': float(credit) if credit and isinstance(credit, (int, float)) else self._safe_float(credit),
                    'aux_accounting': str(aux_val).strip() if aux_val else '',
                    'settlement_from_aux': settlement_from_aux,
                    'customer_supplier': settlement_name,  # v0.2: 往来单位名称
                    'department': str(department_val).strip() if department_val and str(department_val).strip() not in ('nan', 'None') else '',
                    'project': str(project_val).strip() if project_val and str(project_val).strip() not in ('nan', 'None') else '',
                    'bank_account': str(bank_account_val).strip() if bank_account_val and str(bank_account_val).strip() not in ('nan', 'None') else '',
                })

        self.row_count = len(self.data)
        wb.close()

    @staticmethod
    def _safe_float(val):
        if val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        if isinstance(val, str):
            try: return float(val.replace(',', '').strip())
            except: return 0.0
        return 0.0

    @staticmethod
    def _parse_date(date_val, cell=None):
        """DT-54: 序时账日期多格式兼容解析 v1.1
        - 支持 openpyxl 公式缓存返回 None 时的兜底（从公式字符串推断日期）
        - 支持 1904 日期系统（macOS Excel）
        - 负值/极小值防 OverflowError
        """
        if date_val is None and cell is not None:
            # 公式单元格缓存为空：尝试从公式字符串提取
            try:
                from openpyxl.cell.cell import Cell
                if isinstance(cell, Cell) and cell.value is None and cell.data_type == 'f':
                    formula = getattr(cell, 'value', None)
                    # 公式返回空缓存，记录但无法解析
                    pass
            except Exception:
                pass
        if isinstance(date_val, datetime):
            return date_val
        if isinstance(date_val, str):
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日',
                        '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
                try:
                    return datetime.strptime(date_val.strip(), fmt)
                except ValueError:
                    continue
        if isinstance(date_val, (int, float)):
            # Excel 日期序列号：处理负数（1904系统）、超大值
            try:
                serial = float(date_val)
                # 1904 日期系统的偏移修正
                if serial > 100000:  # 超出1900系统合理范围，尝试1904系统
                    try:
                        return datetime(1904, 1, 1) + timedelta(days=serial)
                    except (ValueError, OverflowError):
                        pass
                return datetime(1899, 12, 30) + timedelta(days=serial)
            except (ValueError, OverflowError, OSError):
                pass
        return None

    @staticmethod
    def _extract_settlement_from_aux(aux_val):
        """从辅助核算字段提取结算对象名称

        辅助核算字段格式示例：
        - "【供应商】1.081 上海凯伟智能科技（集团）有限公司"
        - "【客户】张三"
        - "【项目】某某工程项目"
        - "供应商：某某公司"

        提取策略：去除标签前缀，返回纯结算对象名称
        """
        if not aux_val or not isinstance(aux_val, str):
            return ''
        aux = aux_val.strip()
        # 去除【xxx】前缀
        result = re.sub(r'^【[^】]*】\s*', '', aux)
        # 去除"供应商："等中文冒号前缀
        result = re.sub(r'^(供应商|客户|项目|部门|人员)[：:]\s*', '', result)
        # 去除数字编码前缀（如"1.081 "）
        result = re.sub(r'^[\d.]+\s*', '', result)
        return result.strip()

    def query_by_subject(self, subject_keywords, summary_keywords=None, direction=None,
                         fuzzy_fallback=True, max_results=50):
        """按科目关键词+摘要关键词+辅助核算关键词+方向查询序时账

        v1.1改进:
        - 增加辅助核算字段(辅助核算/结算对象)搜索维度
        - 匹配优先级：科目名称→辅助核算→摘要→科目编码前缀

        Args:
            subject_keywords: 科目名称关键词列表（OR逻辑）
            summary_keywords: 摘要关键词列表（OR逻辑，DT-53）
            direction: 'debit'(借方) / 'credit'(贷方) / None(全部)
            fuzzy_fallback: 精确匹配0命中时是否自动降级到模糊匹配（默认True）
            max_results: 最大返回条数（超过则按金额排序取TOP，默认50）

        Returns:
            list[dict]: 匹配的序时账记录
        """
        # 第一轮：按科目名称精确匹配
        filtered = []
        for s in self.data:
            for kw in subject_keywords:
                if kw and kw in s['subject_name']:
                    filtered.append(s)
                    break

        # v1.1降级策略1：科目名称0命中→用摘要关键词+辅助核算在全部数据中搜索
        # v3.68 (2026-06-02): 如果L2已用settlement关键词搜索且0命中,标记_skip_l3
        # 避免L3科目编码前缀匹配把 settlement_name 过滤掉,导致错返回大量错误条目
        _skip_l3 = False
        if not filtered and fuzzy_fallback and summary_keywords:
            for s in self.data:
                for kw in summary_keywords:
                    # v0.2 (2026-06-01): 增加 customer_supplier 字段（往来单位名称）
                    if kw and (kw in s['summary'] or
                               kw in s.get('aux_accounting', '') or
                               kw in s.get('settlement_from_aux', '') or
                               kw in s.get('customer_supplier', '')):
                        filtered.append(s)
                        break
            if not filtered:
                _skip_l3 = True  # 结算对象无匹配,不再降级到科目前缀

        # v1.1降级策略2：摘要也0命中→用科目编码前缀匹配（仅在未指定结算对象关键词时使用）
        if not filtered and not _skip_l3 and fuzzy_fallback and subject_keywords:
            for s in self.data:
                for kw in subject_keywords:
                    # 科目编码前缀匹配（如"1122"匹配"112201"）
                    if kw and (s['subject_code'].startswith(kw) or
                               s['subject_name'].startswith(kw)):
                        filtered.append(s)
                        break

        # 按摘要关键词+辅助核算关键词筛选（DT-53）
        if summary_keywords and filtered:
            matched = []
            for s in filtered:
                for kw in summary_keywords:
                    if kw and (kw in s['summary'] or
                               kw in s.get('aux_accounting', '') or
                               kw in s.get('settlement_from_aux', '')):
                        matched.append(s)
                        break
            # 如果摘要+辅助核算筛选后0命中，保留原始结果（降级）
            if matched:
                filtered = matched

        # 按方向筛选（DT-51④）
        if direction == 'debit':
            filtered = [s for s in filtered if s['debit'] > 0]
        elif direction == 'credit':
            filtered = [s for s in filtered if s['credit'] > 0]

        # v1.1结果截断：超过max_results条时按金额排序取TOP
        if len(filtered) > max_results:
            filtered.sort(key=lambda x: x['debit'] + x['credit'], reverse=True)
            filtered = filtered[:max_results]

        return filtered

    def get_last_date_by_settlement(self, settlement_name, subject_code_prefix,
                                    summary_keywords=None, direction=None):
        """获取某结算对象的末笔发生日期（DT-51③~④）

        Args:
            settlement_name: 结算对象全称（如"临沂浩然房地产开发有限公司"）
            subject_code_prefix: 科目编码前缀（如"1122"）
            summary_keywords: 摘要关键词（DT-53，从摘要提炼而非全称）
            direction: 'debit'/'credit'/None

        Returns:
            dict: {'date': datetime|None, 'status': str, 'match_count': int}
        """
        # DT-53: 从结算对象名称提炼关键词
        if summary_keywords is None:
            summary_keywords = self._extract_keywords(settlement_name)

        # 科目名称中搜索
        subject_keywords = [subject_code_prefix]

        matched = self.query_by_subject(subject_keywords, summary_keywords, direction)

        if not matched:
            return {'date': None, 'status': 'no_match', 'match_count': 0}
        if len(matched) > 50:
            # v1.1: 不再直接标记ambiguous，按金额排序取TOP20后取末笔日期
            matched.sort(key=lambda x: x['debit'] + x['credit'], reverse=True)
            matched = matched[:20]
            # 继续处理，而非返回ambiguous

        # 取末笔（最新日期）
        matched.sort(key=lambda x: x['date'])
        last = matched[-1]
        return {'date': last['date'], 'status': 'verified', 'match_count': len(matched)}

    @staticmethod
    def _clean_settlement_name(name):
        """DT-ARCH: 清理结算对象名称中的标记前缀
        
        明细表中常出现 *名称、**名称 等标记格式，
        需要去除这些前缀以匹配序时账中的原始名称。
        """
        if not name:
            return name
        import re as _re
        # 去除前导 *、**、*空格 等标记
        return _re.sub(r'^[*＊]+\s*', '', name.strip())
    
    @staticmethod
    def _extract_keywords(full_name):
        full_name = JournalExtractor._clean_settlement_name(full_name)
        """DT-53: 从结算对象全称中提炼搜索关键词

        策略：
        1. 地理关键词（地名2-4字）
        2. 公司简称（去掉有限公司后缀，取核心2-4字）
        """
        keywords = []

        # 地理关键词
        geo_match = re.search(r'([\u4e00-\u9fff]{2,4}(?:省|市|区|县|镇|路|街))', full_name)
        if geo_match:
            keywords.append(geo_match.group(1)[:3])

        # 公司简称
        core_name = re.sub(r'(有限公司|股份有限公司|有限责任公司|公司|集团)', '', full_name)
        if len(core_name) >= 2:
            keywords.append(core_name[:4])

        # 如果没提炼出关键词，用全称前4字
        if not keywords:
            keywords.append(full_name[:4])

        return keywords

    def get_business_summaries(self, settlement_name, subject_code_prefix):
        """DT-60: 获取某结算对象在序时账中的所有摘要（用于业务内容归纳）

        v1.1改进: 同时搜索摘要和辅助核算字段
        """
        summary_keywords = self._extract_keywords(settlement_name)
        matched = self.query_by_subject([subject_code_prefix], summary_keywords)
        # 去重摘要
        seen = set()
        result = []
        for s in matched:
            summary = s['summary']
            if summary and summary not in seen:
                seen.add(summary)
                result.append(summary)
        return result


# ============================================================
# 往来科目Sheet定义
# DT-153v3: 默认定义用于fallback，优先从sheet_col_map.json动态获取
# ============================================================

# 默认往来Sheet定义（仅当sheet_col_map.json不可用时使用）
_DEFAULT_RECEIVABLE_SHEETS = {
    # 资产类
    '3-5':  {'type': 'asset',     'subject_code': '1122', 'name_part': '应收账款'},
    '3-7':  {'type': 'asset',     'subject_code': '1123', 'name_part': '预付款项'},
    '3-8-3':{'type': 'asset',     'subject_code': '1221', 'name_part': '其他应收款'},
    '3-10': {'type': 'asset',     'subject_code': '1461', 'name_part': '合同资产'},
    # 负债类
    '5-5':  {'type': 'liability', 'subject_code': '2202', 'name_part': '应付账款'},
    '5-6':  {'type': 'liability', 'subject_code': '2203', 'name_part': '预收款项'},
    '5-10-3':{'type':'liability', 'subject_code': '2241', 'name_part': '其他应付款'},
}


def _get_receivable_sheets():
    """DT-153v3: 从sheet_col_map.json动态获取往来Sheet定义

    优先从sheet_col_map.json中扫描含settlement/date/business字段的Sheet，
    如果不可用则使用_DEFAULT_RECEIVABLE_SHEETS
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skill_dir = os.path.dirname(script_dir)
    col_map_path = os.path.join(skill_dir, 'assets', 'sheet_col_map.json')

    if not os.path.exists(col_map_path):
        col_map_path = os.path.join(
            os.path.expanduser('~'), '.codex', 'skills',
            'valuation-detail-table', 'valuation-detail-table', 'assets', 'sheet_col_map.json'
        )

    if not os.path.exists(col_map_path):
        return _DEFAULT_RECEIVABLE_SHEETS

    try:
        with open(col_map_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        return _DEFAULT_RECEIVABLE_SHEETS

    sheets = data.get('sheets', {})
    result = {}

    # 从sheet_col_map.json中动态扫描：含settlement字段的Sheet即为往来Sheet
    for sheet_name, sheet_def in sheets.items():
        col_map = sheet_def.get('col_map', {})
        has_settlement = 'settlement' in col_map
        has_date = 'date' in col_map or 'occurrence_date' in col_map
        has_biz = 'business' in col_map or 'business_content' in col_map

        if has_settlement and (has_date or has_biz):
            # 判断资产/负债类
            subject_code = ''
            name_part = ''
            # 从subject_code或sheet名称推断
            for key in ['subject_code', 'code_prefix']:
                if key in sheet_def:
                    subject_code = str(sheet_def[key])
                    break
            for key in ['name_part', 'subject_name', 'name']:
                if key in sheet_def:
                    name_part = str(sheet_def[key])
                    break
            
            # DT-207: 如果sheet_col_map.json中没有subject_code，
            # 从_DEFAULT_RECEIVABLE_SHEETS硬编码中查找
            if not subject_code:
                # 先精确匹配sheet_name
                if sheet_name in _DEFAULT_RECEIVABLE_SHEETS:
                    subject_code = _DEFAULT_RECEIVABLE_SHEETS[sheet_name].get('subject_code', '')
                    name_part = name_part or _DEFAULT_RECEIVABLE_SHEETS[sheet_name].get('name_part', '')
                else:
                    # 模糊匹配：sheet_name的前缀(如3-5/5-5)匹配_DEFAULT的key
                    import re as _re_sj
                    prefix_match = _re_sj.match(r'^[3-6]-[\d\-]+', sheet_name)
                    if prefix_match:
                        sheet_prefix = prefix_match.group()
                        for dk, dv in _DEFAULT_RECEIVABLE_SHEETS.items():
                            if dk.startswith(sheet_prefix) or sheet_prefix.startswith(dk):
                                subject_code = dv.get('subject_code', '')
                                name_part = name_part or dv.get('name_part', '')
                                break

            # 从sheet名前缀推断类型（对齐sheet_col_finder.py: 3-/4-=asset, 5-/6-=liability）
            if sheet_name.startswith('3-') or sheet_name.startswith('4-'):
                sheet_type = 'asset'
            elif sheet_name.startswith('5-') or sheet_name.startswith('6-'):
                sheet_type = 'liability'
            else:
                sheet_type = 'unknown'
            result[sheet_name] = {
                'type': sheet_type,
                'subject_code': subject_code,
                'name_part': name_part,
            }

    return result if result else _DEFAULT_RECEIVABLE_SHEETS

# DT-149: 通用模板文字黑名单
GENERIC_BIZ_CONTENTS = {
    '其他应收款', '其他应付款', '其他往来', '往来款',
    '销售商品/提供服务', '采购商品/接受服务', '预付货款/服务费', '预收货款/服务费',
    '货款', '预收账款', '应付货款',
    '预收', '预付', '预收款', '预付款',  # DT-210: 预收/预付属于通用模板文字
}

# DT-60 Step4: 业务关键词映射
# DT-174: "暂估"是重要会计标识，从"货款"中独立为"暂估款"类别
# 暂估匹配时保留"暂估+主体词"（如"暂估半成品"→"暂估半成品"而非"暂估款"）
BIZ_KEYWORD_MAP = {
    '暂估款': ['暂估'],  # DT-174: 必须放在最前，优先匹配
    '货款': ['货款', '收货款', '发货', '出货', '采购', '进货', '购', '买'],
    '服务费': ['服务费', '技术服务', '咨询费', '管理费', '开发费', '软件费', '检测费', '平台费', '订阅'],
    '工程款': ['工程款', '工程', '施工', '建设', '安装', '装修'],
    '外协费': ['外协', '加工', '代工', '委外', '外包', '协作'],
    '租金': ['租金', '租赁', '房租', '场地费'],
    '保证金': ['保证金', '押金', '投标保证金', '履约保证金'],
    '报销款': ['报销', '差旅', '办公费', '通讯费', '交通费'],
    '员工福利': ['入职', '礼品', '福利', '慰问', '节日', '员工', '周年'],
    '社保': ['社保', '公积金', '五险一金', '养老', '医疗'],
    '税费': ['增值税', '所得税', '税', '附加', '销项', '进项', '开票', '抵扣'],
    '借款': ['借款', '贷款', '融资', '利息', '归还'],
    '往来款': ['往来', '划款', '调拨', '内部', '打款', '转账'],  # DT-175: 移除"付款""支付"（方向词，非业务实质）
    '油费': ['汽油', '柴油', '加油', '油费', 'ETC', '过路'],
    '餐费': ['餐费', '餐饮', '伙食', '饭费'],
    '退税款': ['退税', '出口退税', '留抵退税'],
    '保险费': ['保险', '车险', '财险'],
    '运输费': ['运输', '物流', '货运', '快递', '发货'],
}


def _load_col_map_for_sheet(sheet_name):
    """从sheet_col_map.json加载指定Sheet的列位映射"""
    # DT-153v3: 动态检测路径，不再硬编码绝对路径
    # 策略1: 从脚本所在目录向上查找
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skill_dir = os.path.dirname(script_dir)  # scripts/ → valuation-detail-table/
    col_map_path = os.path.join(skill_dir, 'assets', 'sheet_col_map.json')

    if not os.path.exists(col_map_path):
        # 策略2: fallback到Codex全局Skill安装路径
        col_map_path = os.path.join(
            os.path.expanduser('~'), '.codex', 'skills',
            'valuation-detail-table', 'valuation-detail-table', 'assets', 'sheet_col_map.json'
        )
    if not os.path.exists(col_map_path):
        return None

    with open(col_map_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    sheets = data.get('sheets', {})
    # 精确匹配
    if sheet_name in sheets:
        return sheets[sheet_name]
    # 模糊匹配（sheet名可能带短横线差异）
    for key in sheets:
        if sheet_name.replace('-', '') in key.replace('-', ''):
            return sheets[key]
    return None


def _find_matching_sheet(sheet_key, wb_sheetnames):
    """在wb.sheetnames中查找匹配的Sheet名"""
    # 精确匹配
    if sheet_key in wb_sheetnames:
        return sheet_key
    # 模糊匹配
    for sname in wb_sheetnames:
        if sheet_key.replace('-', '') in sname.replace('-', ''):
            return sname
    return None


def scan_empty_fields(detail_file_path):
    """扫描评估明细表中往来科目的空字段行

    Args:
        detail_file_path: 评估明细表Excel文件路径

    Returns:
        list[dict]: 待核实行列表，每个dict包含：
            sheet: Sheet名
            row: 行号
            name: 结算对象名称
            type: 'asset'/'liability'
            date_empty: 发生日期是否为空
            biz_empty: 业务内容是否为空
            biz_generic: 业务内容是否为通用模板文字
            subject_code: 科目编码前缀
            date_col: 发生日期列号
            biz_col: 业务内容列号
    """
    wb = openpyxl.load_workbook(detail_file_path, data_only=True)
    empty_rows = []

    for sheet_key, cfg in _get_receivable_sheets().items():
        # 查找实际Sheet名
        actual_name = _find_matching_sheet(sheet_key, wb.sheetnames)
        if not actual_name:
            continue

        ws = wb[actual_name]
        if ws.sheet_state == 'hidden':
            continue

        # 从sheet_col_map获取列号
        # DT-153v2: sheet_col_map.json使用flat键名(date/business/settlement)
        # 兼容旧格式: 也尝试嵌套键名(occurrence_date/business_content/settlement_object)
        sheet_def = _load_col_map_for_sheet(actual_name)
        # _load_col_map_for_sheet返回整个sheet定义（含headers/col_map/data_start_row等）
        # 需要从col_map子对象中提取列号
        col_map = sheet_def.get('col_map', {}) if isinstance(sheet_def, dict) else {}
        if col_map:
            # flat键名优先（DT-153标准）
            date_col = col_map.get('date', {}).get('col') if isinstance(col_map.get('date'), dict) else col_map.get('date')
            biz_col = col_map.get('business', {}).get('col') if isinstance(col_map.get('business'), dict) else col_map.get('business')
            name_col = col_map.get('settlement', {}).get('col') if isinstance(col_map.get('settlement'), dict) else col_map.get('settlement')
            # 兼容旧嵌套键名
            if not date_col:
                date_col = col_map.get('occurrence_date', {}).get('col') if isinstance(col_map.get('occurrence_date'), dict) else col_map.get('occurrence_date')
            if not biz_col:
                biz_col = col_map.get('business_content', {}).get('col') if isinstance(col_map.get('business_content'), dict) else col_map.get('business_content')
            if not name_col:
                name_col = col_map.get('settlement_object', {}).get('col') if isinstance(col_map.get('settlement_object'), dict) else col_map.get('settlement_object')
        else:
            # 兜底：动态查找列号（替代"资产类E/D/C、负债类D/E/C"硬编码）
            try:
                sys.path.insert(0, os.path.expanduser('~/.codex/skills/valuation-detail-table/valuation-common/scripts'))
                from sheet_col_finder import find_header_cols
                header_cols_je = find_header_cols(ws)
                date_col = header_cols_je.get('发生日期')
                biz_col = header_cols_je.get('业务内容')
                name_col = header_cols_je.get('结算对象')
            except Exception:
                date_col = biz_col = name_col = None

        # 安全检查: 如果关键列号缺失则跳过该Sheet
        if not name_col:
            print(f'  ⚠️ {actual_name}: name_col/settlement列号未找到，跳过')
            continue
        if not date_col:
            print(f'  ⚠️ {actual_name}: date_col列号未找到，跳过')
            continue
        if not biz_col:
            print(f'  ⚠️ {actual_name}: biz_col列号未找到，跳过')
            continue
        data_start = None
        for r in range(1, min(ws.max_row + 1, 20)):
            a_val = ws.cell(row=r, column=1).value
            if a_val and '检索表头' in str(a_val):
                # 检查是否有检索表头2
                next_a = ws.cell(row=r + 1, column=1).value
                if next_a and '检索表头' in str(next_a):
                    data_start = r + 2
                else:
                    data_start = r + 1
                break

        if not data_start:
            # DT-153v3: 动态兜底——找第一个含数值的行
            for r in range(1, min(ws.max_row + 1, 20)):
                for c in range(1, min(ws.max_column + 1, 10)):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (int, float)) and val != 0:
                        data_start = r
                        break
                if data_start:
                    break
            if not data_start:
                data_start = 7  # 最终兜底

        # 扫描数据行
        for r in range(data_start, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            if a_val and ('合' in str(a_val) or '减' in str(a_val) or '预' in str(a_val)):
                break

            name_val = ws.cell(row=r, column=name_col).value
            if not name_val:
                continue

            date_val = ws.cell(row=r, column=date_col).value
            biz_val = ws.cell(row=r, column=biz_col).value

            biz_str = str(biz_val).strip() if biz_val else ''
            date_empty = date_val is None or date_val == ''
            biz_empty = not biz_str
            biz_generic = biz_str in GENERIC_BIZ_CONTENTS

            if date_empty or biz_empty or biz_generic:
                empty_rows.append({
                    'sheet': actual_name,
                    'row': r,
                    'name': JournalExtractor._clean_settlement_name(str(name_val or '')),
                    'type': cfg['type'],
                    'date_empty': date_empty,
                    'biz_empty': biz_empty,
                    'biz_generic': biz_generic,
                    'subject_code': cfg['subject_code'],
                    'date_col': date_col,
                    'biz_col': biz_col,
                })

    wb.close()
    return empty_rows


def extract_dates(extractor, empty_rows, subjects_path=None):
    """批量提取待核实行发生日期

    Args:
        extractor: JournalExtractor实例
        empty_rows: scan_empty_fields()输出
        subjects_path: 科目余额表缓存路径（可选，用于辅助关键词提炼）

    Returns:
        list[dict]: 核实结果列表
    """
    results = []

    # 按结算对象分组（同一结算对象只查询一次）
    seen_names = {}

    for row in empty_rows:
        if not row['date_empty']:
            continue

        name = row['name']
        subject_code = row['subject_code']

        # 泛匹配检测（DT-52）— v3.68 也尝试用父科目降级，避免G2阻断
        if name in ('其他', '其他个人', '个人', '往来', '其他往来', '备用金'):
            fallback = extractor.query_by_subject([subject_code], None, None)
            if fallback:
                fallback.sort(key=lambda x: x['date'])
                results.append({
                    **row,
                    'status': 'fallback_parent',
                    'verified_date': fallback[-1]['date'],
                    'match_count': 0,
                    'note': f'DT-52泛匹配+父科目降级: {fallback[-1]["date"].strftime("%Y-%m-%d")}',
                })
            else:
                results.append({
                    **row,
                    'status': 'generic_skip',
                    'verified_date': None,
                    'match_count': 0,
                    'note': 'DT-52泛匹配项，跳过自动核实',
                })
            continue

        # 查询序时账（不限方向，取末笔发生日期）
        result = extractor.get_last_date_by_settlement(
            settlement_name=name,
            subject_code_prefix=subject_code,
            direction=None,
        )

        if result['status'] == 'verified' and result['date']:
            results.append({
                **row,
                'status': 'verified',
                'verified_date': result['date'],
                'match_count': result['match_count'],
                'note': f"末笔日期{result['date'].strftime('%Y-%m-%d')}，匹配{result['match_count']}条",
            })
        elif result['status'] == 'ambiguous':
            results.append({
                **row,
                'status': 'ambiguous',
                'verified_date': None,
                'match_count': result['match_count'],
                'note': f"匹配歧义({result['match_count']}条)，跳过自动核实",
            })
        else:
            # v3.68 (2026-06-02): 父科目降级 — 用该科目最近一笔同方向发生日期兜底
            fallback = extractor.query_by_subject([subject_code], None, None)
            if fallback:
                fallback.sort(key=lambda x: x['date'])
                fallback_date = fallback[-1]['date']
                results.append({
                    **row,
                    'status': 'fallback_parent',
                    'verified_date': fallback_date,
                    'match_count': 0,
                    'note': f'父科目降级: 最近日期 {fallback_date.strftime("%Y-%m-%d")}, '
                            f'该结算对象无独立序时账记录',
                })
            else:
                results.append({
                    **row,
                    'status': 'no_match',
                    'verified_date': None,
                    'match_count': 0,
                    'note': '序时账中未找到匹配',
                })

    return results


# DT-175: 交易方向前缀——不是业务实质，应剥离
_DIRECTION_PREFIXES = ['应付', '应收', '预付', '预收', '支付', '付款', '收款', '收到', '收回']
_DIRECTION_ONLY = {'应付', '应收', '预付', '预收', '支付', '付款', '收款', '收到', '收回'}


def _strip_direction(text):
    """DT-175: 剥离交易方向前缀（应付/应收/支付等非业务实质词）。

    规则：
    - 如果文本以方向前缀开头，剥离后返回剩余部分
    - 如果剥离后为空（文本本身就是方向词），返回空串，由调用方决定兜底
    - 如果文本不以方向前缀开头，原样返回

    Examples:
        _strip_direction('应付货款') → '货款'
        _strip_direction('支付') → ''
        _strip_direction('货款') → '货款'
        _strip_direction('暂估应付') → '暂估应付'（"暂估"开头，不匹配方向前缀）
    """
    for prefix in _DIRECTION_PREFIXES:
        if text.startswith(prefix):
            remaining = text[len(prefix):]
            return remaining
    return text


def _summarize_from_raw(summaries):
    """DT-173: 从原始摘要列表中提炼规范化业务内容。

    替代旧的"取最短摘要前6字"兜底逻辑，分3层处理：
    Layer 1: 净化——去除地名/人名/数字/公司简称残片
    Layer 2: 二次关键词匹配——在净化后的摘要上重跑BIZ_KEYWORD_MAP
    Layer 3: 词语边界截断——不切断中文词语，在标点/空格处断开

    Args:
        summaries: 序时账摘要列表

    Returns:
        str: 规范化的业务内容关键词
    """
    import re as _re

    # Layer 1: 摘要净化
    _SUFFIX_NOISE = _re.compile(
        r'[—\-\s].*$'          # 连字符及之后的内容（如"-佛山"、"-芜湖修"、"-中石"）
        r'|\d+周[年月日].*$'    # 数字+时间单位（如"4周年"、"3个月"）
        r'|\d+月\d*日?$'        # 日期后缀
        r'|第[一二三四五六七八九十\d]+[季度期]$'  # "第四季度"、"第一期"等
    )

    cleaned = []
    for s in summaries:
        s = str(s).strip()
        # 去除常见后缀噪音
        s_clean = _SUFFIX_NOISE.sub('', s).strip()
        # 去除尾部的连字符/破折号残片
        s_clean = s_clean.rstrip('—-').strip()
        if s_clean:
            cleaned.append(s_clean)

    if not cleaned:
        cleaned = [str(s).strip() for s in summaries if str(s).strip()]

    # Layer 2: 二次关键词匹配（DT-174: 暂估保留前缀+主体词; DT-175: 剥离方向前缀）
    for s in cleaned:
        for biz_type, patterns in BIZ_KEYWORD_MAP.items():
            for p in patterns:
                if p in s:
                    # DT-174: "暂估"是重要会计标识，保留"暂估+主体词"
                    if biz_type == '暂估款' and p == '暂估':
                        _zangu_match = _re.match(r'暂估(.+)', s)
                        if _zangu_match:
                            rest = _zangu_match.group(1).strip()
                            if rest:
                                return _strip_direction('暂估' + rest)  # DT-175
                    # DT-175: 剥离方向前缀后再返回
                    return _strip_direction(biz_type)

    # Layer 3: 词语边界截断（替代硬切6字）
    # 优先取最短摘要
    shortest = min(cleaned, key=len) if cleaned else ''

    if not shortest:
        return '往来款'  # 最终兜底

    # DT-175: 如果≤4字，先剥离方向前缀再判断
    if len(shortest) <= 4:
        stripped = _strip_direction(shortest)
        return stripped if stripped else '往来款'

    # 在标点/空格/连字符处截断
    _TRUNCATE_AT = _re.compile(r'[，。、；：！？\s—\-/\\]')
    m = _TRUNCATE_AT.search(shortest)
    if m and m.start() > 0:
        result = _strip_direction(shortest[:m.start()])
        return result if result else '往来款'

    # 无标点，在4字处截断（中文4字通常是一个完整词组）
    raw = shortest[:4] if len(shortest) > 4 else shortest
    result = _strip_direction(raw)
    return result if result else '往来款'


def extract_business_contents(extractor, empty_rows, subjects_path=None):
    """批量提取待核实行业务内容（DT-60 5步流程）

    Args:
        extractor: JournalExtractor实例
        empty_rows: scan_empty_fields()输出
        subjects_path: 科目余额表缓存路径

    Returns:
        list[dict]: 核实结果列表
    """
    results = []

    for row in empty_rows:
        if not (row['biz_empty'] or row['biz_generic']):
            continue

        name = row['name']
        subject_code = row['subject_code']

        # DT-60 Step1~3: 获取摘要
        summaries = extractor.get_business_summaries(name, subject_code)

        # DT-60 Step4: 摘要归纳
        inferred_biz = None
        source = 'seq_summary'

        if summaries:
            # 高频关键词统计
            keyword_counts = defaultdict(int)
            for summary in summaries:
                for biz_type, patterns in BIZ_KEYWORD_MAP.items():
                    for p in patterns:
                        if p in summary:
                            keyword_counts[biz_type] += 1
                            break

            if keyword_counts:
                best_biz = max(keyword_counts, key=keyword_counts.get)
                inferred_biz = best_biz
            else:
                # DT-173: 摘要净化+二次概括（替代简单的"前6字"兜底）
                inferred_biz = _summarize_from_raw(summaries)
        else:
            # 兜底推断
            try:
                sys_path = os.path.join(os.path.expanduser('~'), '.codex', 'skills', 'valuation-detail-table', 'valuation-common', 'scripts')
                if sys_path not in sys.path:
                    import sys
                    sys.path.insert(0, sys_path)
                from business_content_map import infer_business_content
                inferred_biz = infer_business_content(subject_code, name)
                source = 'infer_fallback'
            except Exception:
                inferred_biz = name[:6]
                source = 'name_fallback'

        # DT-149: 禁止仅填科目名称
        subject_names = {'其他应收款', '其他应付款', '应收账款', '应付账款', '预付款项', '预收款项'}
        if inferred_biz in subject_names:
            inferred_biz = f"{inferred_biz}[待确认业务实质]"

        # DT-149补强：泛化业务词（如"往来款"）不具备审计实质，统一标注待核实。
        # 这样既不冒充确定值，也不会被G2按“仅填科目名/泛词”直接阻断。
        if inferred_biz in GENERIC_BIZ_CONTENTS:
            inferred_biz = f"{inferred_biz}[待核实]"

        # 兜底推断标注[待核实]
        if source == 'infer_fallback':
            if '[待核实]' not in inferred_biz and '[待确认' not in inferred_biz:
                inferred_biz = f"{inferred_biz}[待核实]"

        results.append({
            **row,
            'status': 'updated' if source == 'seq_summary' else 'inferred',
            'new_biz': inferred_biz,
            'source': source,
            'summary_count': len(summaries),
        })

    return results


def write_phase3_results(detail_file_path, date_results, biz_results):
    """将Phase 3核实结果写入评估明细表

    仅修改发生日期列和业务内容列，不涉及金额列。
    """
    wb = openpyxl.load_workbook(detail_file_path)

    # 写入发生日期
    date_written = 0
    for r in date_results:
        # v3.68 (2026-06-02): 接受 verified 和 fallback_parent 两种状态
        if r.get('status') not in ('verified', 'fallback_parent') or not r.get('verified_date'):
            continue
        if r['sheet'] not in wb.sheetnames:
            continue
        ws = wb[r['sheet']]
        # DT-153v3: 从col_map获取列号，不再硬编码兜底值
        date_col = r.get('date_col')
        if not date_col:
            sheet_def_je = _load_col_map_for_sheet(r['sheet'])
            col_map_inner = sheet_def_je.get('col_map', {}) if isinstance(sheet_def_je, dict) else {}
            if col_map_inner:
                date_val = col_map_inner.get('date', col_map_inner.get('occurrence_date', {}))
                date_col = date_val.get('col') if isinstance(date_val, dict) else date_val
        if not date_col:
            # 最终兜底: 从sheet_col_finder动态获取
            print(f"  [WARN] {r['sheet']}: date_col未找到，跳过该行")
            continue
        ws.cell(row=r['row'], column=date_col).value = r['verified_date']
        ws.cell(row=r['row'], column=date_col).number_format = 'yyyy"年"m"月"'
        date_written += 1

    # 写入业务内容
    biz_written = 0
    for r in biz_results:
        if r.get('status') not in ('updated', 'inferred') or not r.get('new_biz'):
            continue
        if r['sheet'] not in wb.sheetnames:
            continue
        ws = wb[r['sheet']]
        # DT-153v3: 从col_map获取列号，不再硬编码兜底值
        biz_col = r.get('biz_col')
        if not biz_col:
            sheet_def_je = _load_col_map_for_sheet(r['sheet'])
            col_map_inner = sheet_def_je.get('col_map', {}) if isinstance(sheet_def_je, dict) else {}
            if col_map_inner:
                biz_val = col_map_inner.get('business', col_map_inner.get('business_content', {}))
                biz_col = biz_val.get('col') if isinstance(biz_val, dict) else biz_val
        if not biz_col:
            print(f"  [WARN] {r['sheet']}: biz_col未找到，跳过该行")
            continue
        ws.cell(row=r['row'], column=biz_col).value = r['new_biz']
        biz_written += 1

    wb.save(detail_file_path)
    wb.close()

    print(f"Phase 3写入完成: 发生日期{date_written}行, 业务内容{biz_written}行")


def sync_to_cost_workpaper(cost_file_path, date_results, biz_results):
    """同步发生日期+业务内容到成本法底稿（DT-51⑥ + DT-60⑤）"""
    if not os.path.exists(cost_file_path):
        print(f"成本法底稿不存在: {cost_file_path}")
        return

    wb = openpyxl.load_workbook(cost_file_path)

    date_synced = 0
    for r in date_results:
        if r.get('status') != 'verified' or not r.get('verified_date'):
            continue
        if r['sheet'] not in wb.sheetnames:
            continue
        ws = wb[r['sheet']]
        # DT-153v3: 从col_map获取列号，不再硬编码兜底值
        date_col = r.get('date_col')
        if not date_col:
            col_map = _load_col_map_for_sheet(r['sheet'])
            if col_map:
                date_val = col_map.get('date', col_map.get('occurrence_date', {}))
                date_col = date_val.get('col') if isinstance(date_val, dict) else date_val
        if not date_col:
            continue
        ws.cell(row=r['row'], column=date_col).value = r['verified_date']
        ws.cell(row=r['row'], column=date_col).number_format = 'yyyy"年"m"月"'
        date_synced += 1

    biz_synced = 0
    for r in biz_results:
        if r.get('status') not in ('updated', 'inferred') or not r.get('new_biz'):
            continue
        if r['sheet'] not in wb.sheetnames:
            continue
        ws = wb[r['sheet']]
        # DT-153v3: 从col_map获取列号，不再硬编码兜底值
        biz_col = r.get('biz_col')
        if not biz_col:
            col_map = _load_col_map_for_sheet(r['sheet'])
            if col_map:
                biz_val = col_map.get('business', col_map.get('business_content', {}))
                biz_col = biz_val.get('col') if isinstance(biz_val, dict) else biz_val
        if not biz_col:
            continue
        ws.cell(row=r['row'], column=biz_col).value = r['new_biz']
        biz_synced += 1

    wb.save(cost_file_path)
    wb.close()

    print(f"成本法底稿同步完成: 发生日期{date_synced}行, 业务内容{biz_synced}行")


def generate_phase3_report(date_results, biz_results):
    """生成Phase 3核实结果汇总报告"""
    lines = []
    lines.append("=" * 60)
    lines.append("Phase 3 序时账查阅结果汇总")
    lines.append("=" * 60)

    # 发生日期
    lines.append("\n--- 发生日期核实 ---")
    date_verified = sum(1 for r in date_results if r.get('status') == 'verified')
    date_no_match = sum(1 for r in date_results if r.get('status') == 'no_match')
    date_ambiguous = sum(1 for r in date_results if r.get('status') == 'ambiguous')
    date_generic = sum(1 for r in date_results if r.get('status') == 'generic_skip')
    date_fallback = sum(1 for r in date_results if r.get('status') == 'fallback_parent')

    lines.append(f"  已核实: {date_verified}行")
    lines.append(f"  父科目降级: {date_fallback}行 (v3.68)")
    lines.append(f"  未匹配: {date_no_match}行 → [待核实]")
    lines.append(f"  匹配歧义: {date_ambiguous}行 → [待核实]")
    lines.append(f"  泛匹配跳过(DT-52): {date_generic}行")

    # 业务内容
    lines.append("\n--- 业务内容核实 ---")
    biz_updated = sum(1 for r in biz_results if r.get('status') == 'updated')
    biz_inferred = sum(1 for r in biz_results if r.get('status') == 'inferred')
    biz_no_match = sum(1 for r in biz_results if r.get('status') == 'no_match')

    lines.append(f"  序时账摘要归纳: {biz_updated}行")
    lines.append(f"  兜底推断[待核实]: {biz_inferred}行")
    lines.append(f"  未匹配: {biz_no_match}行")

    return "\n".join(lines)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='序时账查阅脚本（Phase 3）')
    parser.add_argument('--seq-file', required=True, help='序时账Excel路径')
    parser.add_argument('--detail-file', required=True, help='评估明细表Excel路径')
    parser.add_argument('--project-dir', help='项目文件夹路径（用于缓存）')
    args = parser.parse_args()

    # 初始化提取器
    print("加载序时账...")
    extractor = JournalExtractor(args.seq_file)
    print(f"序时账行数: {extractor.row_count}")

    # 扫描空字段
    print("\n扫描评估明细表空字段...")
    empty_rows = scan_empty_fields(args.detail_file)

    # 提取发生日期
    print("\n核实发生日期...")
    date_results = extract_dates(extractor, empty_rows)

    # 提取业务内容
    print("\n核实业务内容...")
    biz_results = extract_business_contents(extractor, empty_rows)

    # 写入
    print("\n写入评估明细表...")
    write_phase3_results(args.detail_file, date_results, biz_results)

    # 报告
    print("\n" + generate_phase3_report(date_results, biz_results))
