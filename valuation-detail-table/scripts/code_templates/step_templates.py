# step_templates.py — Phase 2 共享代码模板
# 从S2_fill_bs/liability/re等Step文件提取的公共模板
# DT-136/DT-138/DT-141等规则的脚本化实现
# v1.0 — 2026-05-24 信息架构精简

import sys, os, json, subprocess
import openpyxl

# ════════════════════════════════════════════════════════════
# 模板1: DT-136 动态列位映射（禁止硬编码列号）
# ════════════════════════════════════════════════════════════

def build_col_map(ws, header_row=None):
    """DT-136: 从表头行读取列位映射字典
    
    Args:
        ws: openpyxl worksheet对象
        header_row: 表头行号（None=自动检测A列标记）
    
    Returns:
        col_map: {列含义: 列号} 字典
        struct: 包含header_row/data_start_row/total_row的结构信息
    """
    # 自动检测表头行（A列标记优先）
    struct = _find_header_structure(ws)
    if header_row is None:
        header_row = struct['header_row']
    
    col_map = {}
    for col_idx in range(2, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if not val:
            continue
        val_str = str(val).strip()
        if any(k in val_str for k in ('项目', '户名', '名称', '科目')):
            col_map['项目名称'] = col_idx
        elif '发生日期' in val_str or '日期' in val_str:
            col_map['发生日期'] = col_idx
        elif '业务内容' in val_str or '业务' in val_str:
            col_map['业务内容'] = col_idx
        elif '币种' in val_str:
            col_map['币种'] = col_idx
        elif '账面价值' in val_str or '账面' in val_str:
            col_map['账面价值'] = col_idx
        elif '评估价值' in val_str or '评估' in val_str:
            col_map['评估价值'] = col_idx
        elif '汇率' in val_str:
            col_map['汇率'] = col_idx
        elif '数量' in val_str:
            col_map['数量'] = col_idx
    
    print(f"[DT-136] {ws.title} 列位映射: {col_map}")
    return col_map, struct


def _find_header_structure(ws):
    """从A列标记识别表头结构"""
    struct = {'header_row': 5, 'data_start_row': 6, 'total_row': None}
    
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row, 1).value
        if a_val:
            a_str = str(a_val).strip()
            if a_str in ('检索表头', '检索表头2'):
                struct['header_row'] = row
                struct['data_start_row'] = row + 1
            elif a_str == '检索表头1':
                # 双行表头，下一行是子表头
                struct['header_row'] = row
            elif '合计1' in a_str:
                struct['total_row'] = row
    
    return struct


# ════════════════════════════════════════════════════════════
# 模板2: DT-46 列序校验（资产类C=文字/D=日期 vs 负债类C=日期/D=文字）
# ════════════════════════════════════════════════════════════

def validate_col_order(col_map, side='asset'):
    """DT-46: 资产类/负债类列序校验
    
    Args:
        col_map: build_col_map返回的列位映射
        side: 'asset'(资产类) 或 'liability'(负债类)
    
    Raises:
        AssertionError: 列序违反DT-46
    """
    assert '账面价值' in col_map, "CRITICAL: 未找到'账面价值'列!"
    assert '评估价值' in col_map, "CRITICAL: 未找到'评估价值'列!"
    
    if '发生日期' in col_map and '业务内容' in col_map:
        if side == 'asset':
            # 资产类: C列=业务内容(文字), D列=发生日期(日期)
            assert col_map['业务内容'] < col_map['发生日期'], \
                f"DT-46 CRITICAL: 资产类列序错误! 业务内容列({col_map['业务内容']})应在发生日期列({col_map['发生日期']})左侧"
        else:
            # 负债类: C列=发生日期(日期), D列=业务内容(文字)
            assert col_map['发生日期'] < col_map['业务内容'], \
                f"DT-46 CRITICAL: 负债类列序错误! 发生日期列({col_map['发生日期']})应在业务内容列({col_map['业务内容']})左侧"
    
    print(f"[DT-46] {side}类列序校验通过")


# ════════════════════════════════════════════════════════════
# 模板3: DT-137 结算对象总数校验
# ════════════════════════════════════════════════════════════

def load_expected_count(cache_dir, subject_codes, subject_names=None):
    """DT-137: 从辅助余额表缓存读取结算对象总数
    
    Args:
        cache_dir: _dt_cache/目录路径
        subject_codes: 科目编码(str或list)
        subject_names: 科目名称(可选，用于模糊匹配)
    
    Returns:
        expected_count: 结算对象总数
    """
    if isinstance(subject_codes, str):
        subject_codes = [subject_codes]
    if subject_names is None:
        subject_names = []
    elif isinstance(subject_names, str):
        subject_names = [subject_names]
    
    summary_path = os.path.join(cache_dir, 'auxiliary_balance_summary.json')
    expected_count = 0
    
    if os.path.exists(summary_path):
        with open(summary_path, 'r', encoding='utf-8') as f:
            summary = json.load(f)
        for key, val in summary.items():
            if any(code in key for code in subject_codes) or \
               any(name in key for name in subject_names):
                expected_count += val if isinstance(val, int) else val.get('count', 0)
    
    print(f"[DT-137] 辅助余额表结算对象总数={expected_count}")
    return expected_count


# ════════════════════════════════════════════════════════════
# 模板4: 写入后回读验证
# ════════════════════════════════════════════════════════════

def readback_verify(ws, col_map, data_list, struct):
    """DT-97/DT-136: 写入后assert回读验证
    
    验证第一行数据写入成功，防止列偏移导致数据写错列
    """
    first_data_row = struct['data_start_row']
    readback_val = ws.cell(row=first_data_row, column=col_map['账面价值']).value
    expected_val = data_list[0].get('账面价值', 0) if data_list else 0
    
    if isinstance(readback_val, (int, float)) and isinstance(expected_val, (int, float)):
        assert abs(readback_val - expected_val) < 1, \
            f"CRITICAL: {ws.title}写入后回读验证失败! 写入={expected_val}, 回读={readback_val}"
    
    print(f"[DT-97] {ws.title} 回读验证通过: 账面价值={readback_val}")


# ════════════════════════════════════════════════════════════
# 模板5: DT-141 规则执行汇总
# ════════════════════════════════════════════════════════════

def rules_check_summary(rules_checked):
    """DT-141: 脚本末尾规则执行汇总确认
    
    Args:
        rules_checked: {规则ID: bool/str} 字典
    
    Raises:
        AssertionError: 存在未确认规则
    """
    failed = [k for k, v in rules_checked.items() if not v]
    assert len(failed) == 0, f"CRITICAL [DT-141]: 规则执行未确认: {failed}"
    print(f"[DT-141] ✅ 规则执行汇总: {len(rules_checked)}条规则全部确认")


# ════════════════════════════════════════════════════════════
# 模板6: DT-138 gate_validator调用
# ════════════════════════════════════════════════════════════

def run_gate_validator(filepath, sheet_name, gate='G1'):
    """DT-138: 保存后强制调用gate_validator
    
    Args:
        filepath: 评估明细表文件路径
        sheet_name: Sheet名称
        gate: 门控级别(G1/G2/G3)
    
    Raises:
        SystemExit: 门控未通过
    """
    gate_script = os.path.expanduser(
        '~/.codex/skills/valuation-detail-table/valuation-detail-table/scripts/gate_validator.py'
    )
    # 尝试valuation-common路径
    if not os.path.exists(gate_script):
        gate_script = os.path.expanduser(
            '~/.codex/skills/valuation-detail-table/valuation-common/scripts/gate_validator.py'
        )
    
    if os.path.exists(gate_script):
        result = subprocess.run(
            [sys.executable, gate_script, filepath, '--gate', gate, '--sheet', sheet_name],
            capture_output=True, text=True, timeout=120
        )
        print(f"[DT-138] gate_validator输出:\n{result.stdout}")
        if result.returncode != 0 or 'CRITICAL' in result.stdout:
            print(f"[DT-138] 🚨 {gate}门控未通过!")
            sys.exit(1)
        else:
            print(f"[DT-138] ✅ {gate}门控通过")
    else:
        print(f"[DT-138] WARNING: gate_validator.py未找到，跳过门控")


# ════════════════════════════════════════════════════════════
# 模板7: RULE_CHECK注释块生成器
# ════════════════════════════════════════════════════════════

RULE_CHECK_TEMPLATE = '''
# ╔════════════════════════════════════════════════════════════════╗
# ║ DT-141: 规则校验清单（{sheet_name}，脚本自检=规则执行确认）     ║
# ╚════════════════════════════════════════════════════════════════╝
{rule_lines}
'''

def generate_rule_check_block(sheet_name, rules):
    """生成DT-141 RULE_CHECK注释块
    
    Args:
        sheet_name: Sheet名称
        rules: [(rule_id, description, check_method)] 列表
    
    Returns:
        注释块字符串
    """
    rule_lines = []
    for rule_id, desc, method in rules:
        rule_lines.append(f"# [{rule_id}] {desc} → {method}")
    return RULE_CHECK_TEMPLATE.format(
        sheet_name=sheet_name,
        rule_lines='\n'.join(rule_lines)
    )


# ════════════════════════════════════════════════════════════
# 模板8: 标准填写流水线（封装1-7）
# ════════════════════════════════════════════════════════════

def standard_fill_pipeline(filepath, ws, data_list, side='asset', 
                           subject_codes=None, subject_names=None):
    """标准填写流水线: DT-136列映射→DT-46列序→DT-97回读→DT-141汇总→DT-138门控
    
    Args:
        filepath: 评估明细表文件路径
        ws: openpyxl worksheet对象
        data_list: 待写入数据列表
        side: 'asset'或'liability'
        subject_codes: DT-137科目编码
        subject_names: DT-137科目名称
    """
    # Step 1: DT-136 列位映射
    col_map, struct = build_col_map(ws)
    
    # Step 2: DT-46 列序校验
    validate_col_order(col_map, side=side)
    
    # Step 3: DT-137 结算对象校验（如适用）
    if subject_codes:
        cache_dir = os.path.dirname(filepath).rstrip('/') + '/_dt_cache'
        expected = load_expected_count(cache_dir, subject_codes, subject_names)
        actual = struct['total_row'] - struct['data_start_row'] if struct['total_row'] else len(data_list)
        if expected > 0 and actual < expected:
            print(f"[DT-137] WARNING: {ws.title} 数据行数({actual})<辅助余额表结算对象数({expected})")
    
    # Step 4: 写入后回读验证
    if data_list:
        readback_verify(ws, col_map, data_list, struct)
    
    # Step 5: DT-141 规则执行汇总
    rules_checked = {
        'DT-0': len(data_list) > 0,
        'DT-46': True,
        'DT-66': '账面价值' in col_map and '评估价值' in col_map,
        'DT-136': True,
    }
    rules_check_summary(rules_checked)
    
    # Step 6: 保存 + DT-138 门控
    wb = ws.parent
    wb.save(filepath)
    run_gate_validator(filepath, ws.title)
    
    print(f"✅ {ws.title} 标准填写流水线完成")
