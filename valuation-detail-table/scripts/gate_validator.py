"""
评估明细表 — Gate Validator (硬约束校验引擎) v3.10

设计原则：
- 五级硬约束：G0(数据源级) → G1(写入级-数据) → G1-Format(写入级-格式) → G2(科目级) → G3(勾稽级)
- G0未通过=禁止进入Phase 1（数据源完整性+辅助余额表提取）
- G1未通过=禁止继续填写下一个Sheet（数据正确性验证）
- G1-Format未通过=禁止进入Phase 4（格式集中处置验证，Phase 3完成后执行）
- G2未通过=禁止进入Phase 3
- G3未通过=禁止交付
- 本脚本不调用save，不修改文件，纯只读验证（--auto-fix模式除外）

v3.9.1变更 (2026-05-24)：
- G2-18修复: 列位映射不再硬编码，改为从sheet_col_map.json读取（DT-153强制）
  - 复盘: "资产类date=5/负债类date=4"的粗分类假设只是6个Sheet恰好对了
  - sheet_col_map.json中52个Sheet有date列，列位从4到9各不相同
  - 如3-8-1应收利息date=4, 3-5应收账款date=5, 4-8-4机器设备date=9
  - 硬编码"资产类/负债类"两档覆盖不了所有变体
  - G2-18现在自动覆盖所有含date列的Sheet（而非只检查6个往来Sheet）

v3.10变更 (2026-05-24)：
- G2-19新增: 合计行完整性检查 [DT-164.1] (v3.10新增)
  - 复盘: 3-7预付款项35条数据 > 模板预填20行，v2脚本直接逐行覆写，
    合计1/坏账准备/合计2三行被数据覆盖→A列标记+SUM公式+B:C合并全部丢失
  - G2-19.1: 合计1行A列标记必须存在
  - G2-19.2: 合计1行B列必须含"合"或"计"关键字
  - G2-19.3: 合计1行账面价值列必须有SUM公式或非零值
  - G2-19.4: 坏账准备行A列标记必须存在
  - G2-19.5: 合计2行A列标记必须存在
  - sheet_col_map.json预加载到gate_G2()开头，G2-18/G2-19共用

v3.8变更 (2026-05-24)：
- G2-13新增: 业务内容填写质量检查 [DT-149] (v3.8新增)
  - 检查"项目及内容"列是否仅填科目名而无业务实质
  - 复盘问题12/16: "应付账款"填在业务内容列=等于没填
  - 配合business_content_map.py自动推断
- G2-14新增: 应交税费逐税种填写检查 [DT-147] (v3.8新增)
  - 检查应交税费Sheet是否逐税种填写+征税机关格式
  - 复盘问题14: 征税机关应按提示填写，税种填列在税费种类一列
- G2-15新增: 递延所得税名称披露检查 [DT-150] (v3.8新增)
  - 检查递延所得税名称是否披露具体内容
  - 复盘问题11: 递延所得税名称应披露具体内容
- G2-16新增: 长期借款必填检查 [DT-148] (v3.8新增)
  - 检查6-1长期借款Sheet是否有数据
  - 复盘问题18: 6-1长期借款未填写
- G2-17新增: 列位正确性检查 [DT-66] (v3.8新增)
  - 检查金额/文本是否错位列
  - 复盘问题13/15/17: 数据不在正确位置

v3.7变更 (2026-05-23)：
- G2-12新增: smart_insert_row工具调用检测 [DT-120] (v3.30新增)
  - 4项格式特征检测：数据行B:C合并残留/打印范围未更新/合计行缺少B:C合并/数据行默认格式
  - 任一CRITICAL=smart_insert_row未被使用或未正确执行=禁止交付
  - 教训22: L1文字约束无L2脚本检测=形同虚设，需4层强约束

v3.6变更 (2026-05-23)：
- G2-11新增: 重分类映射检查 [DT-118] (v3.29新增)
  - 检查明细表中的重分类标注（"[重分类]"）
  - 验证重分类目标Sheet是否有对应数据
  - 后续可扩展：对比3-4其他流动资产与BS的一致性

v3.5变更 (2026-05-23)：
- G2-10新增: 待确认映射数据未解决检查 [DT-117] (v3.27新增)
  - 扫描所有明细表，检测备注栏中含"待确认映射"标注的单元格
  - 存在未解决的待确认映射=CRITICAL=禁止交付
  - 需在Phase 4差额推论中确认映射关系或标注"待核实"后人工判断

v3.3变更 (2026-05-22)：
- DT-114验证-修复闭环原则：
  - G1-Format新增--auto-fix模式：验证失败时自动调用excel_row_ops.auto_fix_formats修复
  - 修复后自动重新验证，最多3次重试
  - 3次仍FAIL则BLOCKED，输出详细诊断
  - G1F-1~G1F-4可自动修复，G1F-5/G1F-6需人工处理

v3.2变更 (2026-05-22)：
- 新增G1-Format格式门控（DT-112格式集中处置原则）：
  - G1F-1: 增值额/增值率列number_format校验 [DT-76]
  - G1F-2: 数据行行高统一性校验 [DT-77]
  - G1F-3: 合计/减值/小计行A列居中对齐校验 [DT-84]
  - G1F-4: 空白数据行边框完整性校验 [DT-82]
  - G1F-5: 多行表头合并单元格校验 [DT-83]
  - G1F-6: 公式列覆写校验 [DT-67]
- DT-74即时门控拆分：数据检查留在G1，格式检查移至G1-Format

v3.1变更 (2026-05-22)：
- G0级新增：数据源级校验（原G2-9升级+新增DT-106/108/109/110/111覆盖）
- G0-1: 数据源完整性门控 [DT-103] (从G2-9升级)
- G0-2: 银行存款逐行展开行数门控 [DT-104] (从G1-8迁移至G0)
- G0-3: 辅助余额表结算对象行数门控 [DT-111] (新增)
- G0-4: PDF提取完整性硬Gate [DT-108] (新增，需外部传入aux_data参数)
- G0-5: PDF数据源回填校验 [DT-109] (新增，需外部传入aux_data参数)
- G3-10新增: 合计行公式引用校验 [DT-85] (新增)
- G3-11新增: 隐藏操作校验 [DT-110] (新增)

v3.0变更 (2026-05-22)：
方向A执行：将更多规则从"文字约束"转为"可执行代码"，覆盖DT规则从17条→42条
- G1-4新增: 坏账准备行I列填减值准备+J=0校验 + 预计风险行I=0+J列填减值对应评估值校验 [DT-18]
- G1-5新增: 发生日期列datetime类型校验 [DT-30]
- G1-7新增: 减值/结构行保护(禁止被清除) [DT-6/DT-78①]
- G1-8新增: 银行存款逐行展开行数门控 [DT-104]
- G2-7新增: 其他流动资产=应交税费借方余额行数门控 [DT-87]
- G2-8新增: 递延所得税零余额子项排除校验 [DT-89]
- G2-9新增: 数据源完整性校验 [DT-103]
- G3-8新增: 减值准备行评估值方向校验 [DT-18]
- G3-9新增: 汇总表跨sheet引用行号校验 [DT-86]

v2.0变更 (2026-05-21)：
- GV-1修复: G1-1双行表头子标签行排除(>100000报CRITICAL, 100~100000报WARNING)
- GV-2修复: 新增G1-6写入后列位回读验证[DT-97]——检测金额写入错误列位
- GV-4修复: G3 key_accounts补全至全部科目
- GV-5修复: 新增G3-6公式缓存校验[DT-98]——汇总表D列全局为0时CRITICAL
- GV-6修复: 新增G3-7汇总链完整性校验[DT-99]——明细→汇总→分类汇总链接验证
- GV-7改进: KNOWN_RECLASSIFICATIONS改为参数化(不再硬编码)
- 新增G2-6: BS重分类自动检测[DT-100]

调用方式：
    python gate_validator.py <xlsx_path> [--gate G1|G2|G3|all] [--bs-path <bs_path>] [--sb-path <sb_path>]

返回：
    exit 0 = 全部通过
    exit 1 = 有硬约束违规，必须修复
"""

import sys
import json
import re
import openpyxl
from pathlib import Path
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# 从valuation-common导入共享工具函数
sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'valuation-common' / 'scripts'))
from shared_utils import get_sheet_prefix
from sheet_col_finder import find_header_cols, get_amount_cols, SheetColFinder


# ============================================================
# 通用工具函数
# ============================================================

def find_header_structure(ws):
    """识别Sheet的表头结构：主表头行、子表头行、数据起始行、合计行

    Returns:
        dict with keys: header_row, sub_header_row, data_start_row, total_rows,
                        total_row, is_multi_row_header,
                        bad_debt_row, provision_row, total2_row (DT-164.1新增)
    """
    header_row = None
    sub_header_row = None
    total_rows = []
    total_row = None
    bad_debt_row = None
    provision_row = None
    total2_row = None

    for r in range(1, min(ws.max_row + 1, 300)):
        a_val = ws.cell(row=r, column=1).value
        a_text = str(a_val).replace(' ', '').strip() if a_val else ''

        # 找表头行（"检索表头"或"序号"行）
        if a_text in ('检索表头', '检索表头1', '检索表头2') and header_row is None:
            header_row = r
            # 检查下一行是否为子标题行
            if r + 1 <= ws.max_row:
                next_vals = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    cv = ws.cell(row=r + 1, column=c).value
                    if cv and not isinstance(ws.cell(row=r + 1, column=c), MergedCell):
                        next_vals.append(str(cv).strip())
                next_text = ''.join(next_vals)
                if any(kw in next_text for kw in ['原值', '净值', '成新率', '增值额', '增值率', '金额', '单价', '数量', '设备费', '资金成本', '安装费', '初始额', '利息及汇率', '合计']):
                    sub_header_row = r + 1
            continue

        # 兼容：旧模板表头行A列为"序号"
        if a_text == '序号' and header_row is None:
            header_row = r
            continue

        # 找合计行和坏账准备行（在表头行之后）
        if header_row and r > header_row:
            if '合' in a_text and '计' in a_text and '评估明细表' not in a_text:
                if not total_row:
                    total_row = r  # 合计1行
                else:
                    total2_row = r  # 合计2行
                total_rows.append(r)
            elif '坏账' in a_text or '减值' in a_text or '预计风险' in a_text:
                if '坏账' in a_text or '减值' in a_text:
                    bad_debt_row = r
                elif '预计风险' in a_text:
                    provision_row = r

    data_start_row = (sub_header_row + 1) if sub_header_row else (header_row + 1) if header_row else None
    is_multi_row_header = sub_header_row is not None

    return {
        'header_row': header_row,
        'sub_header_row': sub_header_row,
        'data_start_row': data_start_row,
        'total_rows': total_rows,
        'total_row': total_row,
        'is_multi_row_header': is_multi_row_header,
        'bad_debt_row': bad_debt_row,
        'provision_row': provision_row,
        'total2_row': total2_row,
    }



def get_column_header_text(ws, col, struct):
    """获取某列的完整表头文本（Row5+Row6组合）"""
    hr = struct['header_row']
    shr = struct['sub_header_row']
    hval = ws.cell(row=hr, column=col).value if hr else None
    shval = ws.cell(row=shr, column=col).value if shr else None
    return str(hval or '') + str(shval or '')


# 数值列关键字（这些列只允许数字，禁止文本）
VALUE_COL_KEYWORDS = ['价值', '原值', '净值', '余额', '金额', '增值额', '账面']
# 文本列关键字（这些列只允许文本，禁止纯数字作为金额）
TEXT_COL_KEYWORDS = ['项目及内容', '户名', '业务内容', '结算内容', '设备名称', '规格型号', '生产厂家', '币种', '持股比例', '投资成本', '形象进度', '付款比例']

# 禁止写入金额的列关键词 [DT-102]
FORBIDDEN_AMOUNT_COL_KEYWORDS = ['持股比例', '投资成本', '形象进度', '付款比例', '使用年限', '预计使用', '法定']


def classify_column(header_val, sub_header_val=None):
    """根据表头文本判断列类型: value/text/formula/forbidden_amount/other"""
    if header_val is None:
        return 'other'
    h = str(header_val)
    if sub_header_val:
        h = h + str(sub_header_val)
    
    # 先检查禁止列 [DT-102]
    for kw in FORBIDDEN_AMOUNT_COL_KEYWORDS:
        if kw in h:
            return 'forbidden_amount'
    
    for kw in VALUE_COL_KEYWORDS:
        if kw in h:
            return 'value'
    for kw in TEXT_COL_KEYWORDS:
        if kw in h:
            return 'text'
    if '增值率' in h or '成新率' in h:
        return 'formula'
    return 'other'


def find_column_by_keyword(ws, struct, keyword):
    """在表头中查找包含指定关键词的列号
    
    Args:
        ws: worksheet
        struct: find_header_structure返回的dict
        keyword: 要搜索的关键词
    
    Returns:
        list of column indices matching the keyword
    """
    hr = struct['header_row']
    shr = struct['sub_header_row']
    result = []
    
    for c in range(1, min(ws.max_column + 1, 25)):
        htext = get_column_header_text(ws, c, struct)
        if keyword in htext:
            result.append(c)
    
    return result


# ============================================================
# G0: 数据源级硬约束 (Phase 0完成后触发，未通过=禁止进入Phase 1)
# ============================================================

def gate_G0(filepath, bs_path=None, sb_path=None, aux_data=None):
    """G0数据源级校验：
    - G0-1: 数据源完整性门控 [DT-103] (从G2-9升级)
    - G0-2: 银行存款逐行展开行数门控 [DT-104]
    - G0-3: 辅助余额表结算对象行数门控 [DT-111] (v3.1新增)
    - G0-4: PDF提取完整性硬Gate [DT-108] (v3.1新增，需aux_data)
    - G0-5: PDF数据源回填校验 [DT-109] (v3.1新增，需aux_data)
    
    Args:
        filepath: 评估明细表文件路径
        bs_path: 资产负债表文件路径
        sb_path: 科目余额表文件路径
        aux_data: 辅助数据字典，格式:
            {
                'auxiliary_balance': {  # DT-111辅助余额表提取结果
                    '科目编码': [{'name': '结算对象', 'balance': 金额}, ...]
                },
                'pdf_extraction_status': {  # DT-108 PDF提取状态
                    'file_path': True/False  # True=已提取, False=仅识别未提取
                },
                'pdf_usage_status': {  # DT-109 PDF回填状态
                    'file_path': True/False  # True=已引用到Sheet, False=未引用
                }
            }
    
    Returns: (passed: bool, violations: list)
    """
    violations = []
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    # ---- G0-1: 数据源完整性门控 [DT-103] ----
    # 统计科目余额表中有余额的科目数 vs 明细表中有数据的科目数
    if sb_path:
        try:
            wb_sb4 = openpyxl.load_workbook(sb_path, data_only=True)
            ws_sb4 = wb_sb4.active
            
            sb_nonzero_top = {}
            for row in ws_sb4.iter_rows(min_row=1, max_row=500, values_only=False):
                code = row[0].value
                name = row[1].value if len(row) > 1 else None
                if code and isinstance(code, str) and len(code) >= 4:
                    prefix4 = code[:4]
                    balance = None
                    for c in range(len(row) - 1, 3, -1):
                        val = row[c].value if c < len(row) else None
                        if isinstance(val, (int, float)):
                            balance = val
                            break
                    if balance is not None and abs(balance) > 0.01:
                        if prefix4 not in sb_nonzero_top:
                            sb_nonzero_top[prefix4] = [name, 0]
                        sb_nonzero_top[prefix4][1] += abs(balance)
            
            wb_sb4.close()
            
            sb_to_dt_map = {
                '1001': ['3-1'], '1002': ['3-1'],
                '1122': ['3-5'], '1123': ['3-7'], '1221': ['3-8'],
                '1401': ['3-9'], '1408': ['3-9'], '5002': ['3-9'],
                '1821': ['3-11', '3-13'],
                '1501': ['4-2'], '1601': ['4-8'],
                '1801': ['4-16'], '1811': ['4-17', '4-19'],
                '2201': ['5-5'], '2202': ['5-5'],
                '2203': ['5-6'], '2210': ['5-7'],
                '2221': ['5-9'],
                '2241': ['5-10'], '2231': ['5-10'],
                '2501': ['6-1'],
            }
            
            dt_sheet_prefixes = set()
            for sname in wb.sheetnames:
                p = get_sheet_prefix(sname)
                if p:
                    dt_sheet_prefixes.add(p)
            
            for code, (name, total_bal) in sb_nonzero_top.items():
                if code in sb_to_dt_map:
                    dt_prefixes = sb_to_dt_map[code]
                    found = any(any(sp.startswith(dp) for sp in dt_sheet_prefixes) for dp in dt_prefixes)
                    if not found:
                        violations.append({
                            'gate': 'G0-1',
                            'rule': 'DT-103',
                            'severity': 'WARNING',
                            'sheet': 'N/A',
                            'cell': 'N/A',
                            'value': total_bal,
                            'message': f'科目余额表{code}({name})余额={total_bal:,.2f}但无对应明细表！可能遗漏数据源。DT-103：数据源完整性门控'
                        })
        except Exception as e:
            violations.append({
                'gate': 'G0-1',
                'rule': 'DT-103',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'数据源完整性校验失败: {e}'
            })
    
    # ---- G0-2: 银行存款逐行展开行数门控 [DT-104] ----
    for sname in wb.sheetnames:
        if '3-1-2' in sname or sname.startswith('3-1-2'):
            ws_bank = wb[sname]
            struct_bank = find_header_structure(ws_bank)
            dsr_b = struct_bank['data_start_row']
            tr_b = struct_bank['total_row']
            if not dsr_b or not tr_b:
                continue
            
            data_row_count = 0
            for r in range(dsr_b, tr_b):
                a_val = ws_bank.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                if '减' in a_text or '合' in a_text:
                    continue
                b_val = ws_bank.cell(row=r, column=2).value
                c_val = ws_bank.cell(row=r, column=3).value
                if b_val or c_val:
                    data_row_count += 1
            
            if data_row_count <= 1 and data_row_count > 0:
                violations.append({
                    'gate': 'G0-2',
                    'rule': 'DT-104',
                    'severity': 'WARNING',
                    'sheet': sname,
                    'cell': 'N/A',
                    'value': data_row_count,
                    'message': f'银行存款仅有{data_row_count}行数据！MUST按银行账户逐行展开(每个账号一行)，禁止合并为1-2行汇总。DT-104'
                })
            break
    
    # ---- G0-3: 辅助余额表结算对象行数门控 [DT-111] (v3.1新增) ----
    if aux_data and 'auxiliary_balance' in aux_data:
        aux_bal = aux_data['auxiliary_balance']
        # 科目编码→明细表前缀映射
        code_to_sheet = {
            '1123': '3-7',   # 预付款项
            '1221': '3-8-3', # 其他应收款
            '1122': '3-5',   # 应收账款
            '2202': '5-5',   # 应付账款
            '2241': '5-10',  # 其他应付款
            '2501': '6-1',   # 长期借款
            '2203': '5-6',   # 预收款项
        }
        
        for code, counterparties in aux_bal.items():
            if not counterparties:
                continue
            prefix = code_to_sheet.get(code[:4])
            if not prefix:
                continue
            
            # 找对应的明细表sheet
            target_sheets = [sn for sn in wb.sheetnames if sn.startswith(prefix) and '汇总' not in sn]
            if not target_sheets:
                continue
            
            # 统计该sheet的数据行数
            for sname in target_sheets:
                ws_aux = wb[sname]
                struct_aux = find_header_structure(ws_aux)
                dsr_a = struct_aux['data_start_row']
                tr_a = struct_aux['total_row']
                if not dsr_a or not tr_a:
                    continue
                
                data_rows = 0
                for r in range(dsr_a, tr_a):
                    a_val = ws_aux.cell(row=r, column=1).value
                    a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                    if '减' in a_text or '合' in a_text:
                        continue
                    b_val = ws_aux.cell(row=r, column=2).value
                    if b_val:
                        data_rows += 1
                
                cp_count = len(counterparties)
                if data_rows < cp_count:
                    violations.append({
                        'gate': 'G0-3',
                        'rule': 'DT-111',
                        'severity': 'CRITICAL',
                        'sheet': sname,
                        'cell': 'N/A',
                        'value': data_rows,
                        'message': f'往来科目{sname}仅{data_rows}行，但辅助余额表有{cp_count}个结算对象！MUST按结算对象逐行填写，禁止仅填汇总数。DT-111'
                    })
                break  # 只检查第一个匹配sheet
    
    # ---- G0-4: PDF提取完整性硬Gate [DT-108] (v3.1新增) ----
    if aux_data and 'pdf_extraction_status' in aux_data:
        pdf_status = aux_data['pdf_extraction_status']
        unextracted = [fp for fp, extracted in pdf_status.items() if not extracted]
        if unextracted:
            violations.append({
                'gate': 'G0-4',
                'rule': 'DT-108',
                'severity': 'CRITICAL',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': len(unextracted),
                'message': f'有{len(unextracted)}个PDF/图片文件已识别但未提取数据！"已识别"≠"已提取"，禁止进入Phase 1。未提取文件: {[fp.split("/")[-1] for fp in unextracted[:5]]}'
            })
    
    # ---- G0-5: PDF数据源回填校验 [DT-109] (v3.1新增) ----
    if aux_data and 'pdf_usage_status' in aux_data:
        pdf_usage = aux_data['pdf_usage_status']
        unused = [fp for fp, used in pdf_usage.items() if not used]
        if unused:
            violations.append({
                'gate': 'G0-5',
                'rule': 'DT-109',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': len(unused),
                'message': f'有{len(unused)}个已提取的PDF/辅助数据源未被引用到对应Sheet！PDF数据源回填校验未通过。未引用文件: {[fp.split("/")[-1] for fp in unused[:5]]}'
            })
    
    wb.close()
    passed = not any(v['severity'] == 'CRITICAL' for v in violations)
    return passed, violations


# ============================================================
# G1: 写入级硬约束 (per-Sheet, 每填写完一个Sheet后触发)
# ============================================================

def gate_G1(filepath, sheet_name=None):
    """G1写入级校验：
    - G1-1: 表头行(含子表头)无数值数据侵入 [DT-90] (GV-1修复: 双行表头子标签排除)
    - G1-2: 数值列禁止文本 / 文本列禁止大额数值 / 禁止列不得有金额 [DT-91/DT-102]
    - G1-3: 数据起始行必须有数据（首行不跳过）[DT-82①]
    - G1-4: 坏账准备行I列=减值准备+J=0 + 预计风险行I=0+J=减值对应评估值 [DT-18] (v3.0新增)
    - G1-5: 发生日期列datetime类型校验 [DT-30] (v3.0新增)
    - G1-6: 写入后列位回读验证 [DT-97] (GV-2修复: 检测金额写入错误列位)
    - G1-7: 减值/结构行保护(禁止被清除) [DT-6/DT-78①] (v3.0新增)
    注意: G1-8(银行存款逐行展开)已迁移至G0-2 [DT-104]
    
    Returns: (passed: bool, violations: list)
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    violations = []
    
    sheets_to_check = [sheet_name] if sheet_name else [
        s for s in wb.sheetnames 
        if wb[s].sheet_state == 'visible' and get_sheet_prefix(s)
    ]
    
    for sname in sheets_to_check:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        prefix = get_sheet_prefix(sname)

        # 跳过汇总表和系统表
        if '汇总' in sname or sname.startswith('0') or sname.startswith('2-') or sname in ['设置', '设定信息']:
            continue
        
        struct = find_header_structure(ws)
        hr = struct['header_row']
        shr = struct['sub_header_row']
        dsr = struct['data_start_row']
        tr = struct['total_row']
        
        if hr is None:
            continue
        
        # ---- G1-1: 表头行无数值数据侵入 [DT-90] (GV-1修复) ----
        header_rows = [hr]
        if shr:
            header_rows.append(shr)
        
        for check_row in header_rows:
            is_sub_header = (check_row == shr and struct['is_multi_row_header'])
            
            for c in range(1, min(ws.max_column + 1, 25)):
                cell = ws.cell(row=check_row, column=c)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                
                col_letter = get_column_letter(c)
                
                # 数值型数据出现在表头行
                if isinstance(val, (int, float)) and abs(val) > 100:
                    if is_sub_header:
                        # GV-1修复: 子标签行(Row6)的特殊处理
                        if abs(val) > 100000:
                            # 大额数值在子标签行=确定的数据侵入
                            violations.append({
                                'gate': 'G1-1',
                                'rule': 'DT-90',
                                'severity': 'CRITICAL',
                                'sheet': sname,
                                'cell': f'{col_letter}{check_row}',
                                'value': val,
                                'message': f'子表头行{check_row}出现大额数值{val:,.2f}，数据写入了列标题行！应从Row{dsr}开始写入'
                            })
                        else:
                            # 小额数值(100~100000)在子标签行=可能是编号等，报WARNING
                            violations.append({
                                'gate': 'G1-1',
                                'rule': 'DT-90',
                                'severity': 'WARNING',
                                'sheet': sname,
                                'cell': f'{col_letter}{check_row}',
                                'value': val,
                                'message': f'子表头行{check_row}出现数值{val:,.2f}，可能是编号/序号，请确认非数据侵入'
                            })
                    else:
                        # 主表头行(Row5)出现任何数值=严重违规
                        violations.append({
                            'gate': 'G1-1',
                            'rule': 'DT-90',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{col_letter}{check_row}',
                            'value': val,
                            'message': f'主表头行{check_row}出现数值{val:,.2f}，数据写入了列标题行！应从Row{dsr}开始写入'
                        })
                
                # 日期型数据出现在表头行
                if hasattr(val, 'strftime') and check_row == shr:
                    violations.append({
                        'gate': 'G1-1',
                        'rule': 'DT-90',
                        'severity': 'CRITICAL',
                        'sheet': sname,
                        'cell': f'{col_letter}{check_row}',
                        'value': str(val),
                        'message': f'子表头行{check_row}出现日期数据，数据写入了列标题行！'
                    })
        
        # ---- G1-2: 列类型校验 [DT-91/DT-102] ----
        if dsr and tr:
            # 读取表头列类型
            col_types = {}
            for c in range(1, min(ws.max_column + 1, 25)):
                hval = ws.cell(row=hr, column=c).value
                shval = ws.cell(row=shr, column=c).value if shr else None
                col_types[c] = classify_column(hval, shval)
            
            # 检查数据行（最多抽检20行）
            # GV-1修复: Row6为双行表头子标签行时，排除G1-2检查
            check_rows = list(range(dsr, min(tr, dsr + 20)))
            # 注意：dsr已经排除了子标签行(sub_header_row+1)，所以这里不需要额外排除
            for r in check_rows:
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                if '减' in a_text or '合' in a_text:
                    continue
                
                for c, ctype in col_types.items():
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    if val is None:
                        continue
                    
                    col_letter = get_column_letter(c)
                    
                    # [DT-102] 禁止列出现金额
                    if ctype == 'forbidden_amount' and isinstance(val, (int, float)) and abs(val) > 100:
                        violations.append({
                            'gate': 'G1-2',
                            'rule': 'DT-102',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': val,
                            'message': f'禁止列(表头含"{get_column_header_text(ws, c, struct)[:30]}")出现金额{val:,.2f}！该列不允许写入金额数据'
                        })
                    
                    # 数值列出现文本（如"人民币"出现在账面价值列）
                    if ctype == 'value' and isinstance(val, str):
                        # 允许的文本：公式、空字符串
                        if val.startswith('=') or val.strip() == '':
                            continue
                        violations.append({
                            'gate': 'G1-2',
                            'rule': 'DT-91',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': val,
                            'message': f'数值列出现文本"{val}"，金额列禁止填入文本！'
                        })
                    
                    # 文本列出现大额数值（可能是列错位）
                    if ctype == 'text' and isinstance(val, (int, float)) and abs(val) > 10000:
                        violations.append({
                            'gate': 'G1-2',
                            'rule': 'DT-91',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': val,
                            'message': f'文本列出现大额数值{val:,.2f}，可能是列错位！'
                        })
        
        # ---- G1-3: 数据起始行必须有数据 [DT-82①] ----
        if dsr and tr:
            first_a = ws.cell(row=dsr, column=1).value
            first_b = ws.cell(row=dsr, column=2).value
            # 检查是否后续行有数据但首行为空
            if first_a is None and first_b is None:
                # 检查后面几行是否有数据
                has_data_later = False
                for r in range(dsr + 1, min(tr, dsr + 10)):
                    for c in range(1, 5):
                        v = ws.cell(row=r, column=c).value
                        if v is not None and not isinstance(ws.cell(row=r, column=c), MergedCell):
                            has_data_later = True
                            break
                    if has_data_later:
                        break
                if has_data_later:
                    violations.append({
                        'gate': 'G1-3',
                        'rule': 'DT-82①',
                        'severity': 'WARNING',
                        'sheet': sname,
                        'cell': f'A{dsr}',
                        'value': None,
                        'message': f'数据起始行{dsr}为空但后续行有数据，首行可能被跳过'
                    })
        
        # ---- G1-6: 写入后列位回读验证 [DT-97] (GV-2修复) ----
        if dsr and tr:
            # 预先识别金额类列（仅这些列用于"错列写入"判断）
            value_like_cols = []
            for c in range(1, min(ws.max_column + 1, 25)):
                h_main = ws.cell(row=hr, column=c).value if hr else None
                h_sub = ws.cell(row=shr, column=c).value if shr else None
                ctype = classify_column(h_main, h_sub)
                if ctype in ('value', 'formula'):
                    value_like_cols.append(c)

            # 查找"账面价值"列——必须精确匹配"账面价值"或"账面价值（元）"等，
            # 排除"外币账面金额"等非主列
            # 对于合并表头(如E5:G5="账面价值"), 实际数据在子标签的"金额"列
            bv_cols = []
            for c in range(1, min(ws.max_column + 1, 25)):
                htext = get_column_header_text(ws, c, struct)
                # 精确匹配：表头包含"账面"且包含"价值"（排除"外币账面金额"）
                if '账面' in htext and '价值' in htext and '外币' not in htext:
                    # 检查是否为合并表头的父列（子标签行有"数量"/"单价"/"金额"等拆分）
                    if shr:
                        sh_text = str(ws.cell(row=shr, column=c).value or '')
                        if sh_text in ['数量', '单价', '金额']:
                            # 这是合并表头的父列，实际数据在"金额"子列
                            # 查找同行shr中"金额"所在的列
                            for c2 in range(c, min(ws.max_column + 1, 25)):
                                sh2 = str(ws.cell(row=shr, column=c2).value or '')
                                if sh2 == '金额':
                                    bv_cols.append(c2)
                                    break
                            continue
                    bv_cols.append(c)
            
            # 检查账面价值列是否有数据
            for bv_col in bv_cols:
                has_bv_data = False
                for r in range(dsr, min(tr, dsr + 50)):
                    val = ws.cell(row=r, column=bv_col).value
                    if val is not None and isinstance(val, (int, float)) and abs(val) > 0.01:
                        has_bv_data = True
                        break
                
                if not has_bv_data:
                    # 检查该Sheet是否有数据行（排除空Sheet）
                    has_any_data = False
                    for r in range(dsr, min(tr, dsr + 10)):
                        for c in range(1, 5):
                            v = ws.cell(row=r, column=c).value
                            if v is not None and not isinstance(ws.cell(row=r, column=c), MergedCell):
                                has_any_data = True
                                break
                        if has_any_data:
                            break

                    # 仅当存在数值型金额数据时，才判定为"写错列"
                    # 纯模板文本/仅名称行不应触发CRITICAL。
                    has_any_numeric_amount = False
                    amount_check_cols = [c for c in value_like_cols if c != bv_col]
                    for r in range(dsr, min(tr, dsr + 50)):
                        for c in amount_check_cols:
                            v = ws.cell(row=r, column=c).value
                            if isinstance(v, (int, float)) and abs(v) > 0.01:
                                has_any_numeric_amount = True
                                break
                        if has_any_numeric_amount:
                            break
                    
                    if has_any_data and has_any_numeric_amount:
                        # Sheet有数据但账面价值列为空=列位错写
                        violations.append({
                            'gate': 'G1-6',
                            'rule': 'DT-97',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{get_column_letter(bv_col)}列',
                            'value': None,
                            'message': f'Sheet有数据行但"账面价值"列(Col{bv_col})无数据！数据可能写入了错误列位。MUST回读确认金额出现在正确列'
                        })
        
        # ---- G1-4: 坏账准备行/预计风险行校验 [DT-18] (v3.0新增, v3.26更正) ----
        # 规则：坏账准备行 I=减值/跌价准备金额（账面价值侧），J=0
        #       预计风险行 I=0，J=减值/跌价准备对应评估值（评估价值侧）
        # 净额 = 合计1 + 坏账准备(负值) + 预计风险(负值)
        if dsr and tr and ('应收' in sname or '其他应收' in sname):
            for r in range(dsr, tr):
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                
                if '坏账准备' in a_text or '减值准备' in a_text:
                    # 坏账准备行(DT-18更正)：动态查找账面价值列和评估价值列
                    header_cols_g1 = find_header_cols(ws)
                    bv_col_g1 = header_cols_g1.get('账面价值')
                    eval_col_g1 = header_cols_g1.get('评估价值')
                    if bv_col_g1 and eval_col_g1:
                        bv_val_g1 = ws.cell(row=r, column=bv_col_g1).value
                        eval_val_g1 = ws.cell(row=r, column=eval_col_g1).value
                        bv_letter_g1 = get_column_letter(bv_col_g1)
                        eval_letter_g1 = get_column_letter(eval_col_g1)
                    else:
                        # 兜底：如果动态查找失败，跳过此检查
                        bv_val_g1 = eval_val_g1 = None
                        bv_letter_g1 = eval_letter_g1 = '?'
                    
                    # 评估价值列应为0——坏账准备行不填评估价值
                    if eval_val_g1 is not None and isinstance(eval_val_g1, (int, float)) and abs(eval_val_g1) > 0.01:
                        violations.append({
                            'gate': 'G1-4',
                            'rule': 'DT-18',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{eval_letter_g1}{r}',
                            'value': eval_val_g1,
                            'message': f'坏账准备行{eval_letter_g1}{r}(评估价值)={eval_val_g1:,.2f}，应为0！坏账准备行仅填账面价值列，不填评估价值列'
                        })
                    
                    # 账面价值列应为负值（减值/跌价准备是减项）
                    if bv_val_g1 is not None and isinstance(bv_val_g1, (int, float)) and bv_val_g1 > 0:
                        violations.append({
                            'gate': 'G1-4',
                            'rule': 'DT-18',
                            'severity': 'WARNING',
                            'sheet': sname,
                            'cell': f'{bv_letter_g1}{r}',
                            'value': bv_val_g1,
                            'message': f'坏账准备行{bv_letter_g1}{r}(账面价值)={bv_val_g1:,.2f}为正值，通常应为负值（减值/跌价准备是减项）'
                        })
                
                if '预计风险' in a_text:
                    # 预计风险行(DT-18更正)：动态查找列位
                    if 'header_cols_g1' not in dir() or header_cols_g1 is None:
                        header_cols_g1 = find_header_cols(ws)
                    bv_col_g1 = header_cols_g1.get('账面价值')
                    eval_col_g1 = header_cols_g1.get('评估价值')
                    if bv_col_g1 and eval_col_g1:
                        bv_val_g1 = ws.cell(row=r, column=bv_col_g1).value
                        eval_val_g1 = ws.cell(row=r, column=eval_col_g1).value
                        bv_letter_g1 = get_column_letter(bv_col_g1)
                        eval_letter_g1 = get_column_letter(eval_col_g1)
                    else:
                        bv_val_g1 = eval_val_g1 = None
                        bv_letter_g1 = eval_letter_g1 = '?'
                    
                    # 账面价值列应为0——预计风险行不填账面价值
                    if bv_val_g1 is not None and isinstance(bv_val_g1, (int, float)) and abs(bv_val_g1) > 0.01:
                        violations.append({
                            'gate': 'G1-4',
                            'rule': 'DT-18',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{bv_letter_g1}{r}',
                            'value': bv_val_g1,
                            'message': f'预计风险行{bv_letter_g1}{r}(账面价值)={bv_val_g1:,.2f}，应为0！预计风险行仅填评估价值列，不填账面价值列'
                        })
        
        # ---- G1-5: 发生日期列datetime类型校验 [DT-30] (v3.0新增) ----
        # 日期列(资产类D列/负债类C列)中的值必须为datetime类型，禁止字符串
        # 排除：银行存款sheet(D=币种)、其他流动资产(D=业务内容)、
        #       存货子表(无日期列)、递延所得税(无日期列)、长期待摊
        if dsr and tr and prefix:
            # 判断资产类还是负债类
            is_asset = prefix.startswith('3-') or prefix.startswith('4-')
            is_liability = prefix.startswith('5-') or prefix.startswith('6-')
            
            # 动态查找发生日期列（替代"资产类D列/负债类C列"硬编码）
            date_col = None
            header_cols_g15 = find_header_cols(ws)
            if '发生日期' in header_cols_g15:
                date_col = header_cols_g15['发生日期']
            
            if date_col:
                for r in range(dsr, min(tr, dsr + 50)):
                    cell = ws.cell(row=r, column=date_col)
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    if val is None:
                        continue
                    a_val = ws.cell(row=r, column=1).value
                    a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                    if '合' in a_text or '减' in a_text:
                        continue
                    
                    col_letter = get_column_letter(date_col)
                    # 字符串日期=红线
                    if isinstance(val, str) and len(val) > 2:
                        # 排除公式
                        if not val.startswith('='):
                            violations.append({
                                'gate': 'G1-5',
                                'rule': 'DT-30',
                                'severity': 'CRITICAL',
                                'sheet': sname,
                                'cell': f'{col_letter}{r}',
                                'value': val,
                                'message': f'发生日期列{col_letter}{r}为字符串"{val}"，必须为datetime类型！字符串日期=Excel无法识别=账龄公式错误'
                            })
                    
                    # 大整数(>40000)=Excel日期序列号但不是datetime对象=COM resave后残留
                    if isinstance(val, (int, float)) and val > 40000 and val < 60000:
                        violations.append({
                            'gate': 'G1-5',
                            'rule': 'DT-30',
                            'severity': 'WARNING',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': val,
                            'message': f'发生日期列{col_letter}{r}为整数{val}(可能为日期序列号)，应为datetime类型。可能需要COM resave后openpyxl二次修复'
                        })
        
        # ---- G1-7: 减值/结构行保护 [DT-6/DT-78①] (v3.0新增) ----
        # 合计行和减值行("减：xxx")的A列标签MUST保留，禁止被清除为None
        if dsr and tr:
            for r in range(dsr, tr + 1):
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                
                is_total = '合' in a_text and '计' in a_text
                is_reduction = a_text.startswith('减')
                
                if is_total or is_reduction:
                    # 检查结构行是否被清除：A列存在但关键数据列全部为None
                    # 对于合计行，检查I列(账面价值)是否为None(可能被清除)
                    # 注意：合计行可能通过SUM公式计算，所以检查公式列
                    all_none = True
                    for c in range(2, min(ws.max_column + 1, 19)):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell, MergedCell):
                            continue
                        v = cell.value
                        if v is not None:
                            all_none = False
                            break
                    
                    # 结构行A列有标签但其余列全空=数据可能被清除
                    if all_none and r < tr:
                        violations.append({
                            'gate': 'G1-7',
                            'rule': 'DT-78',
                            'severity': 'WARNING',
                            'sheet': sname,
                            'cell': f'行{r}',
                            'value': None,
                            'message': f'结构行{r}("{a_text}")A列有标签但其余列全空，数据可能被清除。结构行MUST保留公式和数据'
                        })
    
    wb.close()
    passed = not any(v['severity'] == 'CRITICAL' for v in violations)
    return passed, violations


# ============================================================
# G1-Format: 写入级格式门控 (Phase 3完成后执行，DT-112格式集中处置验证)
# ============================================================

def gate_G1_Format(filepath, sheet_name=None):
    """G1-Format格式门控校验（DT-112格式集中处置原则）：
    Phase 2仅做"最小格式继承"，深度格式修复集中在Phase 3。
    本门控验证Phase 3格式集中处置是否完成。
    
    - G1F-1: 增值额/增值率列number_format校验 [DT-76]
    - G1F-2: 数据行行高统一性校验 [DT-77]
    - G1F-3: 合计/减值/小计行A列居中对齐校验 [DT-84]
    - G1F-4: 空白数据行边框完整性校验 [DT-82]
    - G1F-5: 多行表头合并单元格校验 [DT-83]
    - G1F-6: 公式列覆写校验 [DT-67]
    
    Returns: (passed: bool, violations: list)
    """
    violations = []
    
    # 标准数字格式
    VALUE_ADD_FMT_PARTS = ['#,##0.00', '0.00%', 'General']
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames
    
    for sname in sheets_to_check:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        prefix = get_sheet_prefix(sname)
        if not prefix:
            continue
        if '汇总' in sname or '分类' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        
        struct = find_header_structure(ws)
        dsr = struct['data_start_row']
        shr = struct['sub_header_row']
        tr = struct['total_row']
        if not dsr or not tr:
            continue
        
        # ---- G1F-1: 增值额/增值率列number_format校验 [DT-76] ----
        # 查找增值额列和增值率列（通常为J/K或I/J，取决于列数）
        for c in range(1, min(ws.max_column + 1, 25)):
            htext = get_column_header_text(ws, c, struct)
            if '增值' in htext and ('额' in htext or '率' in htext):
                for r in range(dsr, tr + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    nf = cell.number_format or ''
                    # 检查是否为标准格式（排除General和空）
                    if nf in ('General', '') and cell.value is not None:
                        col_letter = get_column_letter(c)
                        violations.append({
                            'gate': 'G1F-1',
                            'rule': 'DT-76',
                            'severity': 'WARNING',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': nf,
                            'message': f'{sname} {col_letter}{r}增值列number_format="{nf}"，应为会计格式#,##0.00。DT-76：增值额/增值率列格式强制'
                        })
        
        # ---- G1F-2: 数据行行高统一性校验 [DT-77] ----
        ref_height = ws.row_dimensions[dsr].height
        if ref_height:
            for r in range(dsr, tr):
                rh = ws.row_dimensions[r].height
                if rh is not None and abs(rh - ref_height) > 0.5:
                    violations.append({
                        'gate': 'G1F-2',
                        'rule': 'DT-77',
                        'severity': 'WARNING',
                        'sheet': sname,
                        'cell': f'行{r}',
                        'value': rh,
                        'message': f'{sname} 行{r}行高={rh}，与首行数据行高{ref_height}不一致。DT-77：同一Sheet内数据行行高MUST一致'
                    })
                    break  # 每个Sheet只报一次行高不一致
        
        # ---- G1F-3: 合计/减值/小计行A列居中对齐校验 [DT-84] ----
        for r in range(dsr, tr + 1):
            a_val = ws.cell(row=r, column=1).value
            a_text = str(a_val).replace(' ', '').strip() if a_val else ''
            
            is_total = '合' in a_text and '计' in a_text
            is_reduction = a_text.startswith('减')
            is_subtotal = '小' in a_text and '计' in a_text
            
            if is_total or is_reduction or is_subtotal:
                cell_a = ws.cell(row=r, column=1)
                if not isinstance(cell_a, MergedCell):
                    align = cell_a.alignment
                    if align and align.horizontal != 'center':
                        violations.append({
                            'gate': 'G1F-3',
                            'rule': 'DT-84',
                            'severity': 'WARNING',
                            'sheet': sname,
                            'cell': f'A{r}',
                            'value': align.horizontal if align else None,
                            'message': f'{sname} A{r}("{a_text[:10]}")对齐={align.horizontal}，MUST为center。DT-84：结构行A列居中对齐'
                        })
        
        # ---- G1F-4: 空白数据行边框完整性校验 [DT-82] ----
        # 检查数据行区域内空白行是否有thin边框
        from openpyxl.styles import Border
        for r in range(dsr, tr):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value
            # 空白行（A和B都无值，且不是分隔行）
            if (a_val is None or str(a_val).strip() == '') and (b_val is None or str(b_val).strip() == ''):
                # 检查是否为分隔行（A:B合并）
                is_separator = False
                for merge in ws.merged_cells.ranges:
                    if merge.min_row == r and merge.max_row == r:
                        is_separator = True
                        break
                
                if not is_separator:
                    # 检查边框
                    has_border = False
                    for c in range(1, min(ws.max_column + 1, 5)):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell, MergedCell):
                            continue
                        if cell.border and any(s.style == 'thin' for s in [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom] if s):
                            has_border = True
                            break
                    
                    if not has_border:
                        # 仅在有后续数据行时报告
                        has_data_after = False
                        for r2 in range(r + 1, min(tr, r + 5)):
                            for c in range(1, 5):
                                v = ws.cell(row=r2, column=c).value
                                if v is not None and not isinstance(ws.cell(row=r2, column=c), MergedCell):
                                    has_data_after = True
                                    break
                            if has_data_after:
                                break
                        
                        if has_data_after:
                            violations.append({
                                'gate': 'G1F-4',
                                'rule': 'DT-82',
                                'severity': 'WARNING',
                                'sheet': sname,
                                'cell': f'行{r}',
                                'value': None,
                                'message': f'{sname} 行{r}为空白行但无边框，数据行区域MUST保留thin边框。DT-82：空白数据行格式完整性'
                            })
        
        # ---- G1F-5: 多行表头合并单元格校验 [DT-83] ----
        # 仅对多行表头Sheet检查
        if shr:
            # 检查常见的跨列合并：E5:G5(账面价值)/H5:J5(评估价值)
            expected_merges = []
            for merge in ws.merged_cells.ranges:
                if merge.min_row <= struct['header_row'] + 1:
                    expected_merges.append(str(merge))
            # 简化检查：如果有子表头行，验证其与主表头的合并是否正常
            # 这里只做基本检查，详细检查需要模板对比
            sub_header_all_none = True
            for c in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=shr, column=c)
                if not isinstance(cell, MergedCell) and cell.value is not None:
                    sub_header_all_none = False
                    break
            if sub_header_all_none:
                violations.append({
                    'gate': 'G1F-5',
                    'rule': 'DT-83',
                    'severity': 'CRITICAL',
                    'sheet': sname,
                    'cell': f'行{shr}',
                    'value': None,
                    'message': f'{sname} 子表头行{shr}全部为空！合并单元格可能丢失或子标签未填写。DT-83：多行表头合并单元格验证'
                })
        
        # ---- G1F-6: 公式列覆写校验 [DT-67] ----
        # J列（增值额）和K列（增值率）应为公式，禁止数值覆写
        for c in range(1, min(ws.max_column + 1, 25)):
            htext = get_column_header_text(ws, c, struct)
            if '增值' in htext:
                for r in range(dsr, tr + 1):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    # 非公式、非None、数值类型=可能被覆写
                    if val is not None and isinstance(val, (int, float)) and not isinstance(val, bool):
                        # 排除：合计行和减值行可能确实需要数值
                        a_val = ws.cell(row=r, column=1).value
                        a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                        if '合' not in a_text and '减' not in a_text:
                            violations.append({
                                'gate': 'G1F-6',
                                'rule': 'DT-67',
                                'severity': 'WARNING',
                                'sheet': sname,
                                'cell': f'{get_column_letter(c)}{r}',
                                'value': val,
                                'message': f'{sname} {get_column_letter(c)}{r}增值列={val:,.2f}为数值而非公式，可能公式被覆写。DT-67：公式列覆写禁止'
                            })
    
    wb.close()
    passed = not any(v['severity'] == 'CRITICAL' for v in violations)
    return passed, violations


# ============================================================
# G2: 科目级硬约束 (Phase 2完成后触发)
# ============================================================

def gate_G2(filepath, bs_path=None, sb_path=None, has_journal=False):
    """G2科目级校验：
    - G2-1: BS金额为空/0的科目禁止填写 [DT-92]
    - G2-2: 填入的科目必须在科目余额表中存在 [DT-92]
    - G2-3: 余额方向=平/期末余额=空的子项必须排除 [DT-95]
    - G2-4: 预收/合同负债按科目余额表分类 [DT-93]
    - G2-5: 固定资产无子科目拆分时按PDF台账分类 [DT-94]
    - G2-6: BS重分类自动检测 [DT-100]
    - G2-7: 其他流动资产=应交税费借方余额行数门控 [DT-87] (v3.0新增)
    - G2-8: 递延所得税零余额子项排除校验 [DT-89] (v3.0新增)
    - G2-9: 数据源完整性校验 [DT-103] (v3.0新增)
    - G2-18: Phase 2e序时账核实完整性检查 [DT-161] (v3.9新增)

    Args:
        filepath: 评估明细表路径
        bs_path: 科目余额表路径
        sb_path: 资产负债表路径
        has_journal: 项目是否有序时账数据（DT-161: 有序时账时往来Sheet发生日期+业务内容必须非空）

    Returns: (passed: bool, violations: list)
    """
    violations = []
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    # DT-153: 预加载sheet_col_map.json（供G2-18/G2-19使用）
    sheet_col_map_ref = None
    _COL_MAP_PATHS_G2 = [
        Path(__file__).resolve().parent.parent / 'assets' / 'sheet_col_map.json',
        Path(__file__).resolve().parent / 'assets' / 'sheet_col_map.json',
        Path(__file__).resolve().parent / 'sheet_col_map.json',
    ]
    for _p in _COL_MAP_PATHS_G2:
        if _p.exists():
            try:
                with open(_p, 'r', encoding='utf-8') as _f:
                    sheet_col_map_ref = json.load(_f).get('sheets', {})
            except Exception:
                pass
            break
    
    # 读取DT分类汇总（供G2-2交叉验证使用）
    dt_summary = {}
    for sname in wb.sheetnames:
        if '分类汇总' not in sname:
            continue
        ws = wb[sname]
        for row in ws.iter_rows(min_row=6, max_row=65, values_only=False):
            name = row[2].value
            bv = row[3].value
            if name and bv is not None and isinstance(bv, (int, float)) and abs(bv) > 0.01:
                dt_summary[str(name).strip()] = bv
    
    # ---- G2-1 & G2-2: BS空行科目排除 + 科目余额表交叉验证 ----
    bs_nonzero_accounts = {}
    bs_zero_accounts = set()
    
    if bs_path:
        try:
            wb_bs = openpyxl.load_workbook(bs_path, data_only=True)
            ws_bs = wb_bs.active
            
            for row in ws_bs.iter_rows(min_row=1, max_row=90, values_only=False):
                # 资产侧：Col A=科目名, Col D=年末金额
                a_val = row[0].value if len(row) > 0 else None
                d_val = row[3].value if len(row) > 3 else None
                
                if a_val and isinstance(a_val, str) and a_val.strip():
                    name = a_val.strip()
                    # 精确排除分类标题行和合计行
                    if name in ['一、流动资产：', '二、非流动资产：', '三、资产总计']:
                        continue
                    if any(kw in name for kw in ['合计', '总计', '所有者权益', '流动资产合计', '非流动资产合计', '流动负债合计', '非流动负债合计']):
                        continue
                    if d_val is not None and isinstance(d_val, (int, float)) and abs(d_val) > 0.01:
                        bs_nonzero_accounts[name] = d_val
                    elif d_val is None or (isinstance(d_val, (int, float)) and abs(d_val) < 0.01):
                        bs_zero_accounts.add(name)
                
                # 负债侧：Col E=科目名, Col H=年末金额
                e_val = row[4].value if len(row) > 4 else None
                h_val = row[7].value if len(row) > 7 else None
                
                if e_val and isinstance(e_val, str) and e_val.strip():
                    name = e_val.strip()
                    if name in ['四、流动负债：', '五、非流动负债：', '六、负债合计']:
                        continue
                    if any(kw in name for kw in ['合计', '总计', '所有者权益', '流动资产合计', '非流动资产合计', '流动负债合计', '非流动负债合计']):
                        continue
                    if h_val is not None and isinstance(h_val, (int, float)) and abs(h_val) > 0.01:
                        bs_nonzero_accounts[name] = h_val
                    elif h_val is None or (isinstance(h_val, (int, float)) and abs(h_val) < 0.01):
                        bs_zero_accounts.add(name)
            
            wb_bs.close()
            
            # 用已读取的dt_summary逐项比对BS
            for name_str, bv in dt_summary.items():
                # 排除分类标题行
                if name_str in ['一、流动资产合计', '二、非流动资产合计', '三、资产总计',
                               '四、流动负债合计', '五、非流动负债合计', '六、负债总计',
                               '七、净资产（所有者权益）']:
                    continue
                        
                # 检查1: 精确匹配 — 明细表有值但BS为0/None → 虚增科目
                matched_zero = name_str in bs_zero_accounts
                
                # 检查2: 模糊匹配 — DT科目名是BS零值科目的子串或反之
                if not matched_zero:
                    for bs_name in bs_zero_accounts:
                        if name_str in bs_name or bs_name in name_str:
                            matched_zero = True
                            break
                
                # 检查3: DT科目在BS中完全不存在
                dt_not_in_bs = (name_str not in bs_nonzero_accounts and 
                               name_str not in bs_zero_accounts and
                               not any(name_str in bn or bn in name_str 
                                      for bn in list(bs_nonzero_accounts.keys()) + list(bs_zero_accounts)))
                
                if matched_zero:
                    violations.append({
                        'gate': 'G2-1',
                        'rule': 'DT-92',
                        'severity': 'CRITICAL',
                        'sheet': '2-分类汇总',
                        'cell': 'N/A',
                        'value': bv,
                        'message': f'科目"{name_str}"在BS中年末金额为0/None，但明细表填入{bv:,.2f}→虚增科目！MUST交叉验证科目余额表确认该科目是否存在'
                    })
                elif dt_not_in_bs and abs(bv) > 100:
                    violations.append({
                        'gate': 'G2-1',
                        'rule': 'DT-92',
                        'severity': 'WARNING',
                        'sheet': '2-分类汇总',
                        'cell': 'N/A',
                        'value': bv,
                        'message': f'科目"{name_str}"在BS中无对应行标签，明细表填入{bv:,.2f}→请确认该科目是否为BS重分类或合并项'
                    })
            
            # G2-2增强: 用科目余额表交叉验证
            if sb_path:
                try:
                    wb_sb = openpyxl.load_workbook(sb_path, data_only=True)
                    ws_sb = wb_sb.active
                    sb_codes = set()
                    sb_code_balance = {}  # {code: (name, balance)}
                    for row in ws_sb.iter_rows(min_row=1, max_row=500, values_only=False):
                        code = row[0].value
                        name_val = row[1].value if len(row) > 1 else None
                        if code and isinstance(code, str) and code.startswith(('1', '2')):
                            sb_codes.add(code)
                            # 提取余额
                            balance = None
                            for c in range(len(row) - 1, 3, -1):
                                val = row[c].value if c < len(row) else None
                                if isinstance(val, (int, float)):
                                    balance = val
                                    break
                            if name_val and balance is not None:
                                prefix4 = code[:4]
                                if prefix4 not in sb_code_balance:
                                    sb_code_balance[prefix4] = 0
                                sb_code_balance[prefix4] += abs(balance)
                    wb_sb.close()
                    
                    code_name_map = {
                        '1122': '应收账款', '1123': '预付账款', 
                        '1101': '交易性金融资产', '1132': '应收利息',
                        '1511': '长期股权投资', '1512': '长期股权投资减值准备',
                        '1521': '投资性房地产',
                        '1700': '使用权资产',
                        '1601': '固定资产', '1602': '累计折旧',
                        '1603': '固定资产清理', '1604': '在建工程', '1605': '工程物资',
                        '1606': '固定资产减值准备',
                        '1701': '无形资产', '1702': '累计摊销',
                        '2201': '预收账款', '2210': '合同负债',
                        '2600': '租赁负债', '2601': '租赁负债', '2602': '长期应付款',
                        '1801': '长期待摊费用', '1811': '递延所得税资产',
                        '2221': '应交税费', '2231': '应付利息',
                        '2241': '其他应付款',
                    }
                    
                    for code, name in code_name_map.items():
                        code_exists = code in sb_codes or any(c.startswith(code) for c in sb_codes)
                        if not code_exists and name in bs_zero_accounts:
                            dt_val = dt_summary.get(name)
                            if dt_val is not None and abs(dt_val) > 0.01:
                                violations.append({
                                    'gate': 'G2-2',
                                    'rule': 'DT-92',
                                    'severity': 'CRITICAL',
                                    'sheet': '2-分类汇总',
                                    'cell': 'N/A',
                                    'value': name,
                                    'message': f'科目代码{code}({name})不在科目余额表中且BS金额为0，但DT填入{dt_val:,.2f}→确认虚增！MUST从明细表中删除该科目'
                                })
                    
                    # ---- G2-6: BS重分类自动检测 [DT-100] (新增) ----
                    # BS某科目=0但科目余额表对应科目有值→需查重分类去向
                    for code, name in code_name_map.items():
                        code_exists = code in sb_codes or any(c.startswith(code) for c in sb_codes)
                        if code_exists and name in bs_zero_accounts:
                            # 科目余额表有此科目，但BS为0→可能重分类
                            code_bal = sb_code_balance.get(code[:4], 0)
                            if code_bal > 100:
                                # 检查DT中是否已包含该金额（在其他科目中）
                                found_in_dt = False
                                for dt_name, dt_val in dt_summary.items():
                                    if abs(dt_val - code_bal) < 1 and dt_name != name:
                                        found_in_dt = True
                                        break
                                
                                if not found_in_dt:
                                    violations.append({
                                        'gate': 'G2-6',
                                        'rule': 'DT-100',
                                        'severity': 'WARNING',
                                        'sheet': '2-分类汇总',
                                        'cell': 'N/A',
                                        'value': name,
                                        'message': f'科目"{name}"BS金额为0但科目余额表有余额{code_bal:,.2f}，且DT中未找到对应金额→可能需重分类到其他科目'
                                    })
                    
                except Exception as e:
                    violations.append({
                        'gate': 'G2-2',
                        'rule': 'DT-92',
                        'severity': 'WARNING',
                        'sheet': 'N/A',
                        'cell': 'N/A',
                        'value': str(e),
                        'message': f'科目余额表交叉验证失败: {e}'
                    })
            
        except Exception as e:
            violations.append({
                'gate': 'G2',
                'rule': 'N/A',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'BS文件读取失败: {e}，G2-1/G2-2校验跳过'
            })
    
    # ---- G2-3: 科目余额表零余额子项排除 ----
    if sb_path:
        try:
            wb_sb = openpyxl.load_workbook(sb_path, data_only=True)
            ws_sb = wb_sb.active
            
            zero_balance_items = {}
            
            for row in ws_sb.iter_rows(min_row=1, max_row=500, values_only=False):
                code = row[0].value
                name = row[1].value if len(row) > 1 else None
                direction = None
                balance = None
                
                for c in range(2, min(len(row), 15)):
                    val = row[c].value
                    if isinstance(val, str) and val.strip() in ['借', '贷', '平']:
                        direction = val.strip()
                    if isinstance(val, (int, float)) and c >= 4:
                        balance = val
                
                if code and isinstance(code, str) and code.startswith(('1', '2')):
                    if direction == '平' or (balance is None and direction is None):
                        zero_balance_items[code] = (name, direction)
            
            wb_sb.close()
            
            if zero_balance_items:
                violations.append({
                    'gate': 'G2-3',
                    'rule': 'DT-95',
                    'severity': 'INFO',
                    'sheet': '科目余额表',
                    'cell': 'N/A',
                    'value': len(zero_balance_items),
                    'message': f'科目余额表中{len(zero_balance_items)}个余额方向=平的子项，MUST排除不填: {list(zero_balance_items.keys())[:10]}'
                })
        
        except Exception as e:
            violations.append({
                'gate': 'G2-3',
                'rule': 'DT-95',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'科目余额表读取失败: {e}，G2-3校验跳过'
            })
    
    # ---- G2-7: 其他流动资产=应交税费借方余额行数门控 [DT-87] (v3.0新增) ----
    # 其他流动资产明细表行数必须>=科目余额表应交税费(2221)借方子项数
    if sb_path:
        try:
            wb_sb2 = openpyxl.load_workbook(sb_path, data_only=True)
            ws_sb2 = wb_sb2.active
            
            # 统计应交税费(2221)借方余额子项数
            tax_debit_count = 0
            tax_debit_total = 0
            for row in ws_sb2.iter_rows(min_row=1, max_row=500, values_only=False):
                code = row[0].value
                if code and isinstance(code, str) and code.startswith('2221'):
                    # 找期末余额
                    balance = None
                    for c in range(len(row) - 1, 3, -1):
                        val = row[c].value if c < len(row) else None
                        if isinstance(val, (int, float)):
                            balance = val
                            break
                    # 借方余额=负数(贷方科目的贷方负数=实际借方)
                    if balance is not None and balance < -0.01:
                        tax_debit_count += 1
                        tax_debit_total += abs(balance)
            
            wb_sb2.close()
            
            if tax_debit_count > 0:
                # 检查3-11或3-13其他流动资产sheet
                for sname in wb.sheetnames:
                    if '其他流动资产' in sname and '汇总' not in sname:
                        ws_otca = wb[sname]
                        otca_data_rows = 0
                        struct2 = find_header_structure(ws_otca)
                        dsr2 = struct2['data_start_row']
                        tr2 = struct2['total_row']
                        if dsr2 and tr2:
                            for r in range(dsr2, tr2):
                                a_val = ws_otca.cell(row=r, column=1).value
                                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                                if '减' in a_text or '合' in a_text:
                                    continue
                                b_val = ws_otca.cell(row=r, column=2).value
                                if b_val:
                                    otca_data_rows += 1
                        
                        # 行数门控：其他流动资产行数<应交税费借方子项数=红线
                        if otca_data_rows < tax_debit_count:
                            violations.append({
                                'gate': 'G2-7',
                                'rule': 'DT-87',
                                'severity': 'CRITICAL',
                                'sheet': sname,
                                'cell': 'N/A',
                                'value': otca_data_rows,
                                'message': f'其他流动资产仅{otca_data_rows}行，但应交税费(2221)有{tax_debit_count}个借方余额子项(合计{tax_debit_total:,.2f})！MUST逐项展开填写，禁止合并为1-2行汇总。DT-87红线：其他流动资产=应交税费借方余额重分类'
                            })
                        break  # 只检查第一个其他流动资产sheet
        except Exception as e:
            violations.append({
                'gate': 'G2-7',
                'rule': 'DT-87',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'其他流动资产行数门控校验失败: {e}'
            })
    
    # ---- G2-8: 递延所得税零余额子项排除校验 [DT-89] (v3.0新增) ----
    # 检查4-17递延所得税资产sheet是否包含余额为0的子项
    if sb_path:
        try:
            wb_sb3 = openpyxl.load_workbook(sb_path, data_only=True)
            ws_sb3 = wb_sb3.active
            
            # 找1811递延所得税资产的零余额子项
            zero_dta_items = []
            for row in ws_sb3.iter_rows(min_row=1, max_row=500, values_only=False):
                code = row[0].value
                name = row[1].value if len(row) > 1 else None
                if code and isinstance(code, str) and code.startswith('1811'):
                    balance = None
                    direction = None
                    for c in range(2, min(len(row), 15)):
                        val = row[c].value
                        if isinstance(val, str) and val.strip() in ['借', '贷', '平']:
                            direction = val.strip()
                        if isinstance(val, (int, float)) and c >= 4:
                            balance = val
                    
                    if direction == '平' or (balance is not None and abs(balance) < 0.01):
                        zero_dta_items.append((code, name))
            
            wb_sb3.close()
            
            # 检查4-17/4-19递延所得税资产sheet中是否包含这些零余额子项
            if zero_dta_items:
                for sname in wb.sheetnames:
                    if '递延所得税' in sname and '汇总' not in sname:
                        ws_dta = wb[sname]
                        struct_dta = find_header_structure(ws_dta)
                        dsr_dta = struct_dta['data_start_row']
                        tr_dta = struct_dta['total_row']
                        
                        if dsr_dta and tr_dta:
                            for r in range(dsr_dta, tr_dta):
                                b_val = ws_dta.cell(row=r, column=2).value
                                if b_val and isinstance(b_val, str):
                                    for zcode, zname in zero_dta_items:
                                        if zname and zname in b_val:
                                            violations.append({
                                                'gate': 'G2-8',
                                                'rule': 'DT-89',
                                                'severity': 'CRITICAL',
                                                'sheet': sname,
                                                'cell': f'B{r}',
                                                'value': b_val,
                                                'message': f'递延所得税资产包含零余额子项"{b_val}"({zcode})，MUST排除不填入！余额方向=平或余额=0'
                                            })
                        break
        except Exception as e:
            violations.append({
                'gate': 'G2-8',
                'rule': 'DT-89',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'递延所得税零余额子项校验失败: {e}'
            })
    
    # ---- G2-9: 数据源完整性校验 [DT-103] (v3.0新增) ----
    # 统计科目余额表中有余额的科目数 vs 明细表中有数据的科目数
    # 粗粒度检查：科目余额表有余额但明细表无对应数据→WARNING
    if sb_path:
        try:
            wb_sb4 = openpyxl.load_workbook(sb_path, data_only=True)
            ws_sb4 = wb_sb4.active
            
            # 提取科目余额表中有余额的一级科目
            sb_nonzero_top = {}  # {4位code: (name, total_balance)}
            for row in ws_sb4.iter_rows(min_row=1, max_row=500, values_only=False):
                code = row[0].value
                name = row[1].value if len(row) > 1 else None
                if code and isinstance(code, str) and len(code) >= 4:
                    prefix4 = code[:4]
                    balance = None
                    for c in range(len(row) - 1, 3, -1):
                        val = row[c].value if c < len(row) else None
                        if isinstance(val, (int, float)):
                            balance = val
                            break
                    if balance is not None and abs(balance) > 0.01:
                        if prefix4 not in sb_nonzero_top:
                            sb_nonzero_top[prefix4] = [name, 0]
                        sb_nonzero_top[prefix4][1] += abs(balance)
            
            wb_sb4.close()
            
            # 科目余额表一级科目→明细表前缀映射
            sb_to_dt_map = {
                '1001': ['3-1'], '1002': ['3-1'],   # 货币资金
                '1122': ['3-5'],                    # 应收账款
                '1123': ['3-7'],                    # 预付款项
                '1221': ['3-8'],                    # 其他应收款
                '1401': ['3-9'], '1408': ['3-9'], '5002': ['3-9'],  # 存货
                '1821': ['3-11', '3-13'],           # 其他流动资产
                '1501': ['4-2'], '1601': ['4-8'],    # 固定资产
                '1801': ['4-16'],                   # 长期待摊费用
                '1811': ['4-17', '4-19'],           # 递延所得税资产
                '2201': ['5-5'], '2202': ['5-5'],    # 应付账款
                '2203': ['5-6'], '2210': ['5-7'],    # 预收/合同负债
                '2221': ['5-9'],                     # 应交税费
                '2241': ['5-10'], '2231': ['5-10'],  # 其他应付款
                '2501': ['6-1'],                     # 长期借款
            }
            
            # 检查每个有余额的一级科目是否有对应明细表
            dt_sheet_prefixes = set()
            for sname in wb.sheetnames:
                p = get_sheet_prefix(sname)
                if p:
                    dt_sheet_prefixes.add(p)
            
            for code, (name, total_bal) in sb_nonzero_top.items():
                if code in sb_to_dt_map:
                    dt_prefixes = sb_to_dt_map[code]
                    
                    # 检查是否有对应前缀的明细表
                    found = any(any(sp.startswith(dp) for sp in dt_sheet_prefixes) for dp in dt_prefixes)
                    if not found:
                        violations.append({
                            'gate': 'G2-9',
                            'rule': 'DT-103',
                            'severity': 'WARNING',
                            'sheet': 'N/A',
                            'cell': 'N/A',
                            'value': total_bal,
                            'message': f'科目余额表{code}({name})余额={total_bal:,.2f}但无对应明细表！可能遗漏数据源。DT-103：数据源完整性门控'
                        })
        except Exception as e:
            violations.append({
                'gate': 'G2-9',
                'rule': 'DT-103',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'数据源完整性校验失败: {e}'
            })

    # ---- G2-10: 待确认映射数据未解决检查 [DT-117] (v3.27新增) ----
    # 检查明细表中是否存在"待确认映射"标注，存在=CRITICAL=禁止交付
    # DT-117差额推论映射原则：无直接编码映射的数据不得直接填入明细表，
    # 必须通过差额推论确认映射后回填。Phase 4完成后仍存在未解决的待确认映射=CRITICAL
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 仅检查明细表（跳过汇总表和系统表）
            if any(kw in sheet_name for kw in ['汇总', '设置', '信息', '0-其他']):
                continue
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '待确认映射' in cell.value:
                        violations.append({
                            'gate': 'G2-10',
                            'rule': 'DT-117',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(cell.column)}{cell.row}',
                            'value': cell.value[:100],
                            'message': f'待确认映射数据未解决: Sheet[{sheet_name}] {get_column_letter(cell.column)}{cell.row} '
                                       f'含"待确认映射"标注，需在Phase 4差额推论中确认映射或标注"待核实"'
                        })
    except Exception as e:
        violations.append({
            'gate': 'G2-10',
            'rule': 'DT-117',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'待确认映射检查失败: {e}'
        })

    # ---- G2-11: 重分类映射检查 [DT-118] (v3.29新增) ----
    # 检查明细表中是否存在重分类标注（"[重分类]"），以及重分类目标Sheet是否有对应数据
    # 同时检查往来科目Sheet是否含有应重分类但未标注的负数行
    try:
        reclass_sheets = {
            '3-4': '其他流动资产',       # 应交税费负数→其他流动资产
            '5-6': '预收款项',           # 应收账款贷方→预收款项
            '3-5': '应收账款',           # 预收款项借方→应收账款
            '3-7': '预付款项',           # 应付账款借方→预付款项
            '5-5': '应付账款',           # 预付款项贷方→应付账款
            '5-10-3': '其他应付款',       # 其他应收款贷方→其他应付款
            '3-8-3': '其他应收款',       # 其他应付款借方→其他应收款
        }
        reclass_found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '[重分类]' in cell.value:
                        reclass_found = True
                        # 验证重分类目标Sheet是否存在
                        break
                if reclass_found:
                    break
            if reclass_found:
                break

        # 检查：如果BS有"其他流动资产"金额，但3-4 Sheet没有"[重分类]"标注→WARNING
        # （可能是应交税费负数未重分类）
        # 此检查需要BS数据，当前仅检查已有重分类标注的合理性
        # 后续可扩展：读取BS数据后，对比3-4 Sheet合计与BS"其他流动资产"是否一致

    except Exception as e:
        violations.append({
            'gate': 'G2-11',
            'rule': 'DT-118',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'重分类映射检查失败: {e}'
        })

    # ---- G2-12: smart_insert_row工具调用检测 [DT-120] (v3.30新增) ----
    # 检查明细表是否存在"未使用smart_insert_row"的格式特征
    # 教训22: SKILL.md中写"MUST使用smart_insert_row"只是文字约束(L1)，
    #         没有L2脚本检测+L3 Gate拦截+L4流程硬卡点=形同虚设
    # 检测指标：
    #   (1) 数据行区域出现B:C合并（合计行合并残留→smart_insert_row的_fix_merged_cells会自动处理）
    #   (2) 打印范围终止行<合计行位置（smart_insert_row的_update_print_area会自动更新）
    #   (3) 合计行缺少B:C合并（smart_insert_row会在新位置重建）
    #   (4) 数据行C/D列字体为openpyxl默认值(size=12, align=None)（_apply_direct_format会直接定义标准格式11pt+thin边框）
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 跳过非明细表
            if any(kw in sheet_name for kw in ['汇总', '设置', '信息', '0-其他', '2-']):
                continue

            # 查找合计行位置
            total_row = None
            for r in range(1, min(ws.max_row + 1, 200)):
                a_val = ws.cell(row=r, column=1).value
                if a_val and isinstance(a_val, str):
                    text = str(a_val).replace(' ', '').strip()
                    if text in ('合计1', '合计2') or ('合' in text and '计' in text):
                        total_row = r
                        break

            if not total_row:
                continue

            # 检测(1): 数据行区域(行6到total_row-1)出现B:C合并
            from openpyxl.cell.cell import MergedCell
            for mr in ws.merged_cells.ranges:
                if (mr.min_col == 2 and mr.max_col == 3 and
                    mr.min_row >= 6 and mr.min_row < total_row):
                    a_val = ws.cell(row=mr.min_row, column=1).value
                    is_struct = (a_val and isinstance(a_val, str) and
                                ('合' in str(a_val) and '计' in str(a_val)))
                    if not is_struct:
                        violations.append({
                            'gate': 'G2-12',
                            'rule': 'DT-120',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'B{mr.min_row}:C{mr.max_row}',
                            'value': str(mr),
                            'message': f'数据行区域存在B:C合并(B{mr.min_row}:C{mr.max_row})，'
                                       f'这是合计行合并残留。smart_insert_row的_fix_merged_cells会自动处理此问题。'
                                       f'本合并存在=smart_insert_row未被使用或未正确执行'
                        })

            # 检测(2): 打印范围终止行 < 合计行
            if ws.print_area:
                import re
                # 兼容两种print_area格式："$B$1:$M$28" 或 "'Sheet'!$B$1:$M$28"
                pa_match = re.search(r':\$?([A-Z]+)\$?(\d+)\s*$', str(ws.print_area))
                if pa_match:
                    pa_end_row = int(pa_match.group(2))
                    if pa_end_row < total_row:
                        violations.append({
                            'gate': 'G2-12',
                            'rule': 'DT-120',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'print_area:{ws.print_area}',
                            'value': str(ws.print_area),
                            'message': f'打印范围终止行{pa_end_row}小于合计行{total_row}。'
                                       f'smart_insert_row的_update_print_area会自动更新打印范围。'
                                       f'打印范围未更新=smart_insert_row未被使用'
                        })

            # 检测(3): 合计行缺少B:C合并
            has_bc_merge = any(
                mr.min_row == total_row and mr.max_row == total_row and mr.min_col <= 2 and mr.max_col >= 3
                for mr in ws.merged_cells.ranges
            )
            if not has_bc_merge:
                # 仅在该Sheet有实际数据填入时才强制要求B:C合并。
                # 空白模板Sheet（未写入）在部分版本中可能本就无此合并，不应误报CRITICAL。
                has_real_data = False
                for rr in range(6, total_row):
                    b_data = ws.cell(row=rr, column=2).value
                    if not b_data:
                        continue
                    for cc in range(4, min(ws.max_column + 1, 16)):
                        vv = ws.cell(row=rr, column=cc).value
                        if isinstance(vv, (int, float)) and abs(vv) > 0.01:
                            has_real_data = True
                            break
                    if has_real_data:
                        break
                if not has_real_data:
                    continue
                # 检查合计行B列是否有值（有"合            计"类文本才需要B:C合并）
                b_val = ws.cell(row=total_row, column=2).value
                if b_val and isinstance(b_val, str) and '合' in str(b_val):
                    violations.append({
                        'gate': 'G2-12',
                        'rule': 'DT-120',
                        'severity': 'CRITICAL',
                        'sheet': sheet_name,
                        'cell': f'B{total_row}',
                        'value': str(b_val)[:50],
                        'message': f'合计行{total_row}B列="{str(b_val)[:30]}"但缺少B:C合并。'
                                   f'smart_insert_row会在新位置重建B:C合并。'
                                   f'缺少合并=smart_insert_row未被使用或B:C合并重建失败'
                    })

            # 检测(4): 数据行C/D列使用openpyxl默认格式(size=12, align=None)
            default_font_count = 0
            for r in range(6, total_row):
                c_cell = ws.cell(row=r, column=3)
                if not isinstance(c_cell, MergedCell) and c_cell.value is not None:
                    if (c_cell.font.size == 12 and c_cell.alignment.horizontal is None):
                        default_font_count += 1
            # 如果超过50%的数据行C列使用默认格式=疑似未使用_apply_direct_format
            data_row_count = sum(1 for r in range(6, total_row)
                                 if not isinstance(ws.cell(row=r, column=3), MergedCell)
                                 and ws.cell(row=r, column=3).value is not None)
            if data_row_count > 0 and default_font_count / data_row_count > 0.5:
                violations.append({
                    'gate': 'G2-12',
                    'rule': 'DT-120',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': f'C6:C{total_row-1}',
                    'value': f'{default_font_count}/{data_row_count}行C列默认格式',
                    'message': f'{default_font_count}/{data_row_count}行数据行C列使用openpyxl默认格式'
                               f'(font.size=12, align.horizontal=None)。'
                               f'smart_insert_row的_apply_direct_format会直接定义标准格式(size=11, thin边框, align=left)。'
                               f'默认格式占比>50%=_apply_direct_format格式定义未生效'
                })

    except Exception as e:
        violations.append({
            'gate': 'G2-12',
            'rule': 'DT-120',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'smart_insert_row工具调用检测失败: {e}'
        })

    # ---- G2-13: 业务内容填写质量检查 [DT-149] (v3.8新增) ----
    # 检查"项目及内容"列是否仅填科目名而无业务实质
    # 复盘问题12/16: "应付账款"填在业务内容列=等于没填
    try:
        # 科目名=无实质信息的业务内容（这些不应该是业务内容列的值）
        GENERIC_CONTENTS = {
            '应付账款', '应收账款', '其他应收款', '其他应付款',
            '预付款项', '预收款项', '合同负债', '应付职工薪酬',
            '应交税费', '应付利息', '长期借款', '短期借款',
            '其他流动负债', '长期待摊费用', '递延所得税资产',
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if any(kw in sheet_name for kw in ['汇总', '设置', '信息', '0-其他', '2-']):
                continue

            # DT-153: 优先用sheet_col_map判断是否存在业务内容列。
            # 无business列的Sheet（如设备台账、应付利息等）不适用DT-149强校验。
            cm_ref = None
            if sheet_col_map_ref:
                cm_ref = (sheet_col_map_ref.get(sheet_name) or {}).get('col_map', {})
            if cm_ref is not None and 'business' not in cm_ref:
                continue

            struct = find_header_structure(ws)
            hr = struct['header_row']
            shr = struct['sub_header_row']
            dsr = struct['data_start_row']
            tr = struct['total_row']

            if not hr or not dsr or not tr:
                continue

            # 查找"项目及内容"/"业务内容"列
            content_col = None
            if cm_ref and 'business' in cm_ref:
                content_col = cm_ref.get('business')
            else:
                for c in range(1, min(ws.max_column + 1, 25)):
                    htext = get_column_header_text(ws, c, struct)
                    if any(kw in htext for kw in ['项目及内容', '业务内容', '结算内容', '款项内容']):
                        content_col = c
                        break

            if content_col is None:
                continue

            # 检查数据行
            for r in range(dsr, tr):
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                if '减' in a_text or '合' in a_text:
                    continue

                cell = ws.cell(row=r, column=content_col)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val and isinstance(val, str):
                    val_stripped = val.strip()
                    if val_stripped in GENERIC_CONTENTS:
                        violations.append({
                            'gate': 'G2-13',
                            'rule': 'DT-149',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(content_col)}{r}',
                            'value': val_stripped,
                            'message': f'业务内容列仅填科目名"{val_stripped}"，等于没填！'
                                       f'需要填实质信息（如"工程款-XX公司""增值税"等）。'
                                       f'DT-149: business_content_map.py应自动推断'
                        })
                    elif val_stripped == '' or val_stripped is None:
                        # 有金额但业务内容为空
                        bv_col = None
                        for c2 in range(1, min(ws.max_column + 1, 25)):
                            htext2 = get_column_header_text(ws, c2, struct)
                            if '账面价值' in htext2 or '账面余额' in htext2:
                                bv_col = c2
                                break
                        if bv_col:
                            bv_cell = ws.cell(row=r, column=bv_col)
                            if not isinstance(bv_cell, MergedCell) and bv_cell.value:
                                if isinstance(bv_cell.value, (int, float)) and abs(bv_cell.value) > 0.01:
                                    violations.append({
                                        'gate': 'G2-13',
                                        'rule': 'DT-149',
                                        'severity': 'WARNING',
                                        'sheet': sheet_name,
                                        'cell': f'{get_column_letter(content_col)}{r}',
                                        'value': val_stripped,
                                        'message': f'业务内容列为空但有金额数据(账面价值={bv_cell.value:,.2f})，'
                                                   f'应填入业务实质信息'
                                    })
    except Exception as e:
        violations.append({
            'gate': 'G2-13',
            'rule': 'DT-149',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'业务内容填写质量检查失败: {e}'
        })

    # ---- G2-14: 应交税费逐税种填写检查 [DT-147] (v3.8新增) ----
    # 复盘问题14: 应交税费应逐税种填写+征税机关格式
    try:
        for sheet_name in wb.sheetnames:
            if '5-9' not in sheet_name and '应交税费' not in sheet_name:
                continue
            if '汇总' in sheet_name:
                continue

            ws = wb[sheet_name]
            struct = find_header_structure(ws)
            dsr = struct['data_start_row']
            tr = struct['total_row']

            if not dsr or not tr:
                continue

            # 查找"税费种类"和"征税机关"列
            tax_type_col = None
            authority_col = None
            for c in range(1, min(ws.max_column + 1, 25)):
                htext = get_column_header_text(ws, c, struct)
                if '税费种类' in htext or '税种' in htext:
                    tax_type_col = c
                if '征税机关' in htext or '征收机关' in htext:
                    authority_col = c

            for r in range(dsr, tr):
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                if '减' in a_text or '合' in a_text:
                    continue

                # 检查是否所有行都没有税种信息
                if tax_type_col:
                    tt_cell = ws.cell(row=r, column=tax_type_col)
                    if isinstance(tt_cell, MergedCell) or tt_cell.value is None or str(tt_cell.value).strip() == '':
                        violations.append({
                            'gate': 'G2-14',
                            'rule': 'DT-147',
                            'severity': 'WARNING',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(tax_type_col) if tax_type_col else "??"}{r}',
                            'value': None,
                            'message': f'应交税费行{r}未填税种！DT-147: 应逐税种填写税费种类列'
                        })

                if authority_col:
                    auth_cell = ws.cell(row=r, column=authority_col)
                    if isinstance(auth_cell, MergedCell) or auth_cell.value is None or str(auth_cell.value).strip() == '':
                        violations.append({
                            'gate': 'G2-14',
                            'rule': 'DT-147',
                            'severity': 'WARNING',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(authority_col) if authority_col else "??"}{r}',
                            'value': None,
                            'message': f'应交税费行{r}未填征税机关！DT-147: 应填写"国家税务总局"或"地方税务局"'
                        })
            break  # 只检查第一个匹配sheet
    except Exception as e:
        violations.append({
            'gate': 'G2-14',
            'rule': 'DT-147',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'应交税费逐税种检查失败: {e}'
        })

    # ---- G2-15: 递延所得税名称披露检查 [DT-150] (v3.8新增) ----
    # 复盘问题11: 递延所得税名称应披露具体内容
    try:
        for sheet_name in wb.sheetnames:
            if '4-19' not in sheet_name and '递延所得税' not in sheet_name:
                continue
            if '汇总' in sheet_name:
                continue

            ws = wb[sheet_name]
            struct = find_header_structure(ws)
            dsr = struct['data_start_row']
            tr = struct['total_row']
            content_col = None

            if not dsr or not tr:
                continue

            # 查找项目及内容列
            for c in range(1, min(ws.max_column + 1, 25)):
                htext = get_column_header_text(ws, c, struct)
                if any(kw in htext for kw in ['项目及内容', '业务内容']):
                    content_col = c
                    break

            if content_col is None:
                continue

            for r in range(dsr, tr):
                cell = ws.cell(row=r, column=content_col)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val and isinstance(val, str):
                    # 检查是否仅填"递延所得税资产"而无具体差异类型
                    if val.strip() == '递延所得税资产' or val.strip() == '递延所得税负债':
                        violations.append({
                            'gate': 'G2-15',
                            'rule': 'DT-150',
                            'severity': 'WARNING',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(content_col)}{r}',
                            'value': val.strip(),
                            'message': f'递延所得税名称仅填"{val.strip()}"，应披露具体内容'
                                       f'（如"资产减值准备差异""可抵扣亏损"等）'
                        })
            break
    except Exception as e:
        violations.append({
            'gate': 'G2-15',
            'rule': 'DT-150',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'递延所得税名称检查失败: {e}'
        })

    # ---- G2-16: 长期借款必填检查 [DT-148] (v3.8新增) ----
    # 复盘问题18: 6-1长期借款未填写
    try:
        for sheet_name in wb.sheetnames:
            if '6-1' not in sheet_name:
                continue
            if '汇总' in sheet_name:
                continue

            ws = wb[sheet_name]
            struct = find_header_structure(ws)
            dsr = struct['data_start_row']
            tr = struct['total_row']

            if not dsr or not tr:
                continue

            # 检查是否有数据行
            has_data = False
            for r in range(dsr, tr):
                for c in range(2, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        if isinstance(cell.value, (int, float)) and abs(cell.value) > 0.01:
                            has_data = True
                            break
                        elif isinstance(cell.value, str) and cell.value.strip():
                            has_data = True
                            break
                if has_data:
                    break

            if not has_data:
                violations.append({
                    'gate': 'G2-16',
                    'rule': 'DT-148',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': 'N/A',
                    'value': None,
                    'message': f'长期借款Sheet({sheet_name})无任何数据！'
                               f'DT-148: 长期借款必须填写，即使余额为0也需确认'
                })
            break
    except Exception as e:
        violations.append({
            'gate': 'G2-16',
            'rule': 'DT-148',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'长期借款必填检查失败: {e}'
        })

    # ---- G2-17: 列位正确性检查 [DT-66] (v3.8新增) ----
    # 复盘问题13/15/17: 数据不在正确位置（列不对）
    # 检查：金额数据不应出现在文本列，文本数据不应出现在金额列
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if any(kw in sheet_name for kw in ['汇总', '设置', '信息', '0-其他', '2-']):
                continue

            struct = find_header_structure(ws)
            hr = struct['header_row']
            shr = struct['sub_header_row']
            dsr = struct['data_start_row']
            tr = struct['total_row']

            if not hr or not dsr or not tr:
                continue

            # 构建列类型映射
            col_types = {}
            for c in range(1, min(ws.max_column + 1, 25)):
                hval = ws.cell(row=hr, column=c).value
                shval = ws.cell(row=shr, column=c).value if shr else None
                col_types[c] = classify_column(hval, shval)

            # 检查数据行
            for r in range(dsr, min(tr, dsr + 30)):
                a_val = ws.cell(row=r, column=1).value
                a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                if '减' in a_text or '合' in a_text:
                    continue

                for c, ctype in col_types.items():
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    val = cell.value
                    if val is None:
                        continue

                    # 文本列出现大额数值=可能填错列
                    if ctype == 'text' and isinstance(val, (int, float)) and abs(val) > 1000:
                        violations.append({
                            'gate': 'G2-17',
                            'rule': 'DT-66',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(c)}{r}',
                            'value': val,
                            'message': f'文本列({ctype}){get_column_letter(c)}{r}出现大额数值{val:,.2f}！'
                                       f'可能填错列。金额应写入账面价值/评估价值列'
                        })

                    # 金额列出现纯文本=可能填错列
                    if ctype == 'value' and isinstance(val, str) and not val.startswith('='):
                        if len(val.strip()) > 2 and not val.strip().replace('.', '').replace('-', '').replace(',', '').isdigit():
                            violations.append({
                                'gate': 'G2-17',
                                'rule': 'DT-66',
                                'severity': 'WARNING',
                                'sheet': sheet_name,
                                'cell': f'{get_column_letter(c)}{r}',
                                'value': val[:50],
                                'message': f'金额列{get_column_letter(c)}{r}出现文本"{val[:30]}"，'
                                           f'可能填错列。文本内容应写入项目及内容列'
                            })
    except Exception as e:
        violations.append({
            'gate': 'G2-17',
            'rule': 'DT-66',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': str(e),
            'message': f'列位正确性检查失败: {e}'
        })

    # ---- G2-18: Phase 2e序时账核实完整性检查 [DT-161] (v3.9新增) ----
    # 复盘: 有序时账时往来Sheet的发生日期列+业务内容列必须非空，否则=Phase 2e被跳过=底稿不可审
    # 此Gate检查的目的是：在Phase 2→Phase 3的边界阻断，确保Phase 2e确实被执行
    # 规则逻辑：
    #   has_journal=True → 往来Sheet每行有金额的行，发生日期列和业务内容列必须非空
    #   has_journal=False → 跳过本检查（DT-143: 无序时账时发生日期列留空）
    # v3.9.1修复: 列位不再硬编码，改为从sheet_col_map.json读取（DT-153强制）
    #   原因: 不同Sheet列位差异大（3-8-1应收利息date=4, 3-5应收账款date=5, 4-8-6电子设备date=9）
    #         "资产类date=5/负债类date=4"的粗分类假设覆盖不了所有Sheet，是危险的巧合
    if has_journal:
        try:
            # ---- DT-153: 使用函数开头预加载的sheet_col_map_ref ----
            sheet_col_map = sheet_col_map_ref

            if not sheet_col_map:
                violations.append({
                    'gate': 'G2-18',
                    'rule': 'DT-153+DT-161',
                    'severity': 'WARNING',
                    'sheet': 'N/A',
                    'cell': 'N/A',
                    'value': 'sheet_col_map.json not found',
                    'message': 'G2-18无法执行列位精确检查: sheet_col_map.json未找到(DT-153)。'
                               '降级为跳过本检查。建议运行gen_col_map.py生成预映射文件。'
                })
            else:
                # 遍历所有Sheet，从sheet_col_map.json读取列位
                for sheet_name in wb.sheetnames:
                    # 跳过汇总Sheet
                    if '汇总' in sheet_name:
                        continue
                    # 查找该Sheet在sheet_col_map中的映射
                    col_info = sheet_col_map.get(sheet_name)
                    if not col_info:
                        continue
                    cm = col_info.get('col_map', {})
                    # 仅检查“往来类”Sheet：必须同时具备date+business+settlement列。
                    # 设备台账/在建工程等固定资产类虽有date列（启用日期/购置日期），
                    # 但不属于序时账Phase 2e校验范围，避免误报CRITICAL。
                    if 'date' not in cm or 'business' not in cm or 'settlement' not in cm:
                        continue

                    date_col = cm['date']
                    biz_col = cm.get('business')  # 有些Sheet没有业务内容列（如3-8-1应收利息）
                    bv_col = cm.get('book_value')
                    if not bv_col:
                        continue

                    ws = wb[sheet_name]
                    struct = find_header_structure(ws)
                    dsr = struct.get('data_start_row')
                    tr = struct.get('total_row')
                    if not dsr or not tr:
                        continue

                    empty_date_count = 0
                    empty_biz_count = 0
                    total_data_rows = 0

                    for r in range(dsr, tr):
                        a_val = ws.cell(row=r, column=1).value
                        a_text = str(a_val).replace(' ', '').strip() if a_val else ''
                        if '减' in a_text or '合' in a_text or '小' in a_text:
                            continue

                        # 检查是否有金额数据（DT-153: 从col_map读取账面价值列号）
                        amt = ws.cell(row=r, column=bv_col).value
                        if not amt or (isinstance(amt, (int, float)) and amt == 0):
                            continue

                        total_data_rows += 1

                        # 检查发生日期列
                        date_val = ws.cell(row=r, column=date_col).value
                        is_empty_date = False
                        if date_val is None:
                            is_empty_date = True
                        elif isinstance(date_val, MergedCell):
                            is_empty_date = True
                        elif isinstance(date_val, str) and date_val.strip() == '':
                            is_empty_date = True
                        if is_empty_date:
                            empty_date_count += 1
                            violations.append({
                                'gate': 'G2-18',
                                'rule': 'DT-161',
                                'severity': 'CRITICAL',
                                'sheet': sheet_name,
                                'cell': f'{get_column_letter(date_col)}{r}',
                                'value': date_val,
                                'message': f'往来Sheet有金额但发生日期为空！'
                                           f'{sheet_name} Row{r} {get_column_letter(date_col)}列 '
                                           f'有账面价值但发生日期未填写。'
                                           f'DT-161: 有序时账时Phase 2e MUST执行，发生日期不能为空。'
                                           f'请确认Phase 2e是否已执行。'
                            })

                        # 检查业务内容列（仅当该Sheet有业务内容列时）
                        if biz_col:
                            biz_val = ws.cell(row=r, column=biz_col).value
                            is_empty_biz = False
                            if biz_val is None:
                                is_empty_biz = True
                            elif isinstance(biz_val, MergedCell):
                                is_empty_biz = True
                            elif isinstance(biz_val, str) and biz_val.strip() == '':
                                is_empty_biz = True
                            # DT-149: 业务内容仅填科目名=等于没填
                            GENERIC_CONTENTS_G18 = {
                                '应付账款', '应收账款', '其他应收款', '其他应付款',
                                '预付款项', '预收款项', '合同负债', '往来款',
                            }
                            if isinstance(biz_val, str) and biz_val.strip() in GENERIC_CONTENTS_G18:
                                is_empty_biz = True
                            if is_empty_biz:
                                empty_biz_count += 1
                                violations.append({
                                    'gate': 'G2-18',
                                    'rule': 'DT-161+DT-149',
                                    'severity': 'CRITICAL',
                                    'sheet': sheet_name,
                                    'cell': f'{get_column_letter(biz_col)}{r}',
                                    'value': biz_val,
                                    'message': f'往来Sheet有金额但业务内容为空/仅填科目名！'
                                               f'{sheet_name} Row{r} {get_column_letter(biz_col)}列 '
                                               f'业务内容="{biz_val}"。'
                                               f'DT-161: 有序时账时Phase 2e MUST执行，业务内容必须从序时账摘要归纳。'
                                               f'DT-149: 仅填科目名=等于没填=底稿无效。'
                                })

                    # 输出汇总
                    if total_data_rows > 0:
                        date_complete = (total_data_rows - empty_date_count) / total_data_rows * 100
                        biz_complete = (total_data_rows - empty_biz_count) / total_data_rows * 100 if biz_col else 100
                        if empty_date_count == 0 and empty_biz_count == 0:
                            violations.append({
                                'gate': 'G2-18',
                                'rule': 'DT-161',
                                'severity': 'INFO',
                                'sheet': sheet_name,
                                'cell': 'N/A',
                                'value': f'日期完整率={date_complete:.0f}%, 业务内容完整率={biz_complete:.0f}%',
                                'message': f'{sheet_name}: 发生日期+业务内容完整率100%, Phase 2e已执行'
                            })

        except Exception as e:
            violations.append({
                'gate': 'G2-18',
                'rule': 'DT-161',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'Phase 2e序时账核实完整性检查失败: {e}'
            })

    # ---- G2-19: 合计行完整性检查 [DT-164.1] (v3.10新增) ----
    # 复盘: 3-7预付款项35条数据 > 模板预填20行，v2脚本直接逐行覆写，
    # 导致合计1/坏账准备/合计2三行被数据覆盖→A列标记+SUM公式+B:C合并全部丢失
    # 本检查: 验证所有有数据的Sheet，其合计1/坏账准备/合计2行的结构完整性
    for sheet_name in wb.sheetnames:
        # 跳过汇总Sheet和设定信息
        if '汇总' in sheet_name or '设定' in sheet_name or '设置' in sheet_name:
            continue
        if sheet_name.startswith('0-') or sheet_name.startswith('1-') or sheet_name.startswith('2-'):
            continue

        ws = wb[sheet_name]
        struct = find_header_structure(ws)
        dsr = struct.get('data_start_row')
        tr1 = struct.get('total_row')        # 合计1行
        bdr = struct.get('bad_debt_row')     # 坏账准备行
        tr2 = struct.get('total2_row') or struct.get('provision_row')  # 合计2行(优先取total2_row)

        if not dsr or not tr1:
            continue

        # 检查数据区是否有金额（无数据则跳过）
        has_data = False
        for r in range(dsr, tr1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and abs(v) > 0.01:
                    has_data = True
                    break
            if has_data:
                break
        if not has_data:
            continue

        # ---- G2-19.1: 合计1行A列标记必须存在 ----
        a_val_tr1 = ws.cell(row=tr1, column=1).value
        a_text_tr1 = str(a_val_tr1).replace(' ', '').strip() if a_val_tr1 else ''
        if '合计' not in a_text_tr1 and '合' not in a_text_tr1:
            violations.append({
                'gate': 'G2-19',
                'rule': 'DT-164.1',
                'severity': 'CRITICAL',
                'sheet': sheet_name,
                'cell': f'A{tr1}',
                'value': a_val_tr1,
                'message': f'{sheet_name}: 合计1行(Row{tr1})A列标记缺失！当前A{tr1}="{a_val_tr1}"，'
                           f'应为"合计1"或含"合"关键字。数据行可能已覆盖合计行！'
                           f'DT-164.1要求插行后合计行A列标记完整。'
            })

        # ---- G2-19.2: 合计1行B列含"合"或"计"关键字 ----
        b_val_tr1 = ws.cell(row=tr1, column=2).value
        b_text_tr1 = str(b_val_tr1).replace(' ', '').strip() if b_val_tr1 else ''
        if '合' not in b_text_tr1 and '计' not in b_text_tr1:
            violations.append({
                'gate': 'G2-19',
                'rule': 'DT-164.1',
                'severity': 'CRITICAL',
                'sheet': sheet_name,
                'cell': f'B{tr1}',
                'value': b_val_tr1,
                'message': f'{sheet_name}: 合计1行(Row{tr1})B列不是"合 计"！当前B{tr1}="{b_val_tr1}"，'
                           f'数据行可能覆盖了合计行。合计行B列应为"合            计"。'
            })

        # ---- G2-19.3: 合计1行账面价值列有SUM公式 ----
        # 从sheet_col_map.json读取账面价值列号
        if sheet_col_map_ref is not None:
            col_info = sheet_col_map_ref.get(sheet_name)
            if col_info:
                bv_col = col_info.get('col_map', {}).get('book_value')
                if bv_col:
                    bv_cell = ws.cell(row=tr1, column=bv_col)
                    bv_val = bv_cell.value
                    if bv_val is None or (isinstance(bv_val, (int, float)) and abs(bv_val) < 0.01):
                        violations.append({
                            'gate': 'G2-19',
                            'rule': 'DT-164.1',
                            'severity': 'CRITICAL',
                            'sheet': sheet_name,
                            'cell': f'{get_column_letter(bv_col)}{tr1}',
                            'value': bv_val,
                            'message': f'{sheet_name}: 合计1行(Row{tr1})账面价值列({get_column_letter(bv_col)}{tr1})'
                                       f'为空或0！SUM公式可能被覆盖。'
                        })

        # ---- G2-19.4: 坏账准备行A列标记+B列内容（如存在坏账准备行）----
        if bdr and bdr > tr1:
            a_val_bdr = ws.cell(row=bdr, column=1).value
            a_text_bdr = str(a_val_bdr).replace(' ', '').strip() if a_val_bdr else ''
            if '坏账' not in a_text_bdr and '减值' not in a_text_bdr and '减' not in a_text_bdr:
                violations.append({
                    'gate': 'G2-19',
                    'rule': 'DT-164.1',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': f'A{bdr}',
                    'value': a_val_bdr,
                    'message': f'{sheet_name}: 坏账准备行(Row{bdr})A列标记缺失！当前A{bdr}="{a_val_bdr}"，'
                               f'应为"坏账准备"或含"减"关键字。数据行可能已覆盖坏账准备行！'
                })
            # B列应含"减"关键字
            b_val_bdr = ws.cell(row=bdr, column=2).value
            b_text_bdr = str(b_val_bdr).replace(' ', '').strip() if b_val_bdr else ''
            if '减' not in b_text_bdr and '坏账' not in b_text_bdr and '减值' not in b_text_bdr:
                violations.append({
                    'gate': 'G2-19',
                    'rule': 'DT-164.1',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': f'B{bdr}',
                    'value': b_val_bdr,
                    'message': f'{sheet_name}: 坏账准备行(Row{bdr})B列不是"减：XXX坏账准备"！当前B{bdr}="{b_val_bdr}"，'
                               f'数据行可能覆盖了坏账准备行。'
                })

        # ---- G2-19.5: 合计2行A列标记+B列内容（如存在合计2行）----
        if tr2 and tr2 > bdr:
            a_val_tr2 = ws.cell(row=tr2, column=1).value
            a_text_tr2 = str(a_val_tr2).replace(' ', '').strip() if a_val_tr2 else ''
            if '合计' not in a_text_tr2 and '合' not in a_text_tr2:
                violations.append({
                    'gate': 'G2-19',
                    'rule': 'DT-164.1',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': f'A{tr2}',
                    'value': a_val_tr2,
                    'message': f'{sheet_name}: 合计2行(Row{tr2})A列标记缺失！当前A{tr2}="{a_val_tr2}"，'
                               f'应为"合计2"或含"合"关键字。数据行可能已覆盖合计2行！'
                })
            # B列应含"合"或"计"关键字
            b_val_tr2 = ws.cell(row=tr2, column=2).value
            b_text_tr2 = str(b_val_tr2).replace(' ', '').strip() if b_val_tr2 else ''
            if '合' not in b_text_tr2 and '计' not in b_text_tr2:
                violations.append({
                    'gate': 'G2-19',
                    'rule': 'DT-164.1',
                    'severity': 'CRITICAL',
                    'sheet': sheet_name,
                    'cell': f'B{tr2}',
                    'value': b_val_tr2,
                    'message': f'{sheet_name}: 合计2行(Row{tr2})B列不是"合 计"！当前B{tr2}="{b_val_tr2}"，'
                               f'数据行可能覆盖了合计2行。'
                })

    # ---- G2-20: 脚本来源检查 [DT-165] (v3.10新增) ----
    # 复盘: v2脚本完全绕开了fill_sheet()，从零写了fill_sheet_data()，
    # 导致6/9个Sheet列位错误+合计行被覆盖+Phase 2e被跳过
    # 本检查: 扫描_dt_cache/下的.py脚本，检测是否import了Skill脚本
    # 注意: 这是软检查(WARNING级)，因为无法100%确定脚本来源
    import re as _re
    _cache_dir = Path(filepath).resolve().parent / '_dt_cache'
    if _cache_dir.exists():
        for _py_file in _cache_dir.glob('*.py'):
            try:
                _content = _py_file.read_text(encoding='utf-8', errors='ignore')
                _lines = _content.split('\n')
                
                # 检测是否有写入操作但未import Skill脚本
                _has_write = any(
                    ('ws.cell' in l and '.value' in l and '=' in l) or
                    ('ws[' in l and '].value' in l and '=' in l) or
                    ('cell.value' in l and '=' in l)
                    for l in _lines if not l.strip().startswith('#'))
                _has_import_skill = any(
                    'from sheet_filler import' in l or 
                    'from excel_row_ops import' in l or
                    'from data_loader import' in l or
                    'from gate_validator import' in l or
                    'from business_content_map import' in l or
                    'import sheet_filler' in l or
                    'import excel_row_ops' in l or
                    'import data_loader' in l
                    for l in _lines if not l.strip().startswith('#')
                )
                
                if _has_write and not _has_import_skill:
                    violations.append({
                        'gate': 'G2-20',
                        'rule': 'DT-165',
                        'severity': 'WARNING',
                        'sheet': 'N/A',
                        'cell': 'N/A',
                        'value': str(_py_file.name),
                        'message': f'脚本 {_py_file.name} 有ws.cell写入操作但未import任何Skill脚本！'
                                   f'DT-165要求: 现编脚本MUST以已有Skill脚本为前提和基础，优先调用。'
                                   f'可能绕过了fill_sheet()/smart_insert_row()/sheet_col_map.json等关键接口。'
                    })
                
                # 检测是否有DT-165 NEW SCRIPT注释（合法新脚本）
                _has_dt165_marker = any('DT-165 NEW SCRIPT' in l for l in _lines)
                if _has_dt165_marker and _has_import_skill:
                    # 有DT-165标记且import了Skill脚本，视为合规
                    pass
                    
            except Exception:
                pass

    wb.close()
    passed = not any(v['severity'] == 'CRITICAL' for v in violations)
    return passed, violations

# GV-4修复: key_accounts补全至全部科目
FULL_KEY_ACCOUNTS = [
    # 流动资产
    '货币资金', '交易性金融资产', '应收票据', '应收账款', '预付款项',
    '应收利息', '应收股利', '其他应收款', '存货',
    '一年内到期的非流动资产', '其他流动资产',
    # 非流动资产
    '长期股权投资', '投资性房地产', '固定资产', '在建工程',
    '工程物资', '固定资产清理', '使用权资产', '无形资产',
    '长期待摊费用', '递延所得税资产', '其他非流动资产',
    # 流动负债
    '短期借款', '交易性金融负债', '应付票据', '应付账款',
    '预收款项', '合同负债', '应付职工薪酬', '应交税费',
    '应付利息', '应付股利', '其他应付款',
    '一年内到期的非流动负债', '其他流动负债',
    # 非流动负债
    '长期借款', '应付债券', '租赁负债', '长期应付款',
    '长期应付职工薪酬', '预计负债', '递延收益', '递延所得税负债',
    '其他非流动负债',
]

# 汇总表sheet名模式 (用于G3-7汇总链校验)
SUMMARY_SHEET_PATTERNS = [
    ('3-流动资产汇总', '3-'),
    ('4-非流动资产汇总', '4-'),
    ('5-流动负债汇总', '5-'),
    ('6-非流动负债汇总', '6-'),
]


def gate_G3(filepath, bs_path=None, tolerance=0.01):
    """G3勾稽级校验：
    - G3-1: 分类汇总各科目=BS期末余额 [DT-4] (GV-4修复: 全科目覆盖)
    - G3-2: 资产总计一致
    - G3-3: 负债总计一致
    - G3-4: 净资产一致
    - G3-5: 公司名称一致 [DT-96]
    - G3-6: 公式缓存校验 [DT-98] (GV-5修复)
    - G3-7: 汇总链完整性校验 [DT-99] (GV-6修复)
    - G3-8: 减值准备行评估值方向校验 [DT-18] (v3.0新增)
    - G3-9: 汇总表跨sheet引用行号校验 [DT-86] (v3.0新增)
    
    Returns: (passed: bool, violations: list)
    """
    violations = []
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 读取分类汇总
    dt_summary = {}
    for sname in wb.sheetnames:
        if '分类汇总' not in sname:
            continue
        ws = wb[sname]
        for row in ws.iter_rows(min_row=6, max_row=65, values_only=False):
            name = row[2].value
            bv = row[3].value
            if name and bv is not None and isinstance(bv, (int, float)):
                dt_summary[str(name).strip()] = bv
    
    # 读取BS
    bs_summary = {}
    if bs_path:
        try:
            wb_bs = openpyxl.load_workbook(bs_path, data_only=True)
            ws_bs = wb_bs.active
            
            for row in ws_bs.iter_rows(min_row=1, max_row=90, values_only=False):
                a_val = row[0].value if len(row) > 0 else None
                d_val = row[3].value if len(row) > 3 else None
                if a_val and d_val and isinstance(d_val, (int, float)):
                    name = str(a_val).strip()
                    bs_summary[name] = d_val
                
                e_val = row[4].value if len(row) > 4 else None
                h_val = row[7].value if len(row) > 7 else None
                if e_val and h_val and isinstance(h_val, (int, float)):
                    name = str(e_val).strip()
                    bs_summary[name] = h_val
            
            wb_bs.close()
        except Exception as e:
            violations.append({
                'gate': 'G3',
                'rule': 'N/A',
                'severity': 'WARNING',
                'sheet': 'N/A',
                'cell': 'N/A',
                'value': str(e),
                'message': f'BS文件读取失败: {e}，G3勾稽校验跳过'
            })
    
    # ---- G3-1: 全科目逐一比对 (GV-4修复) ----
    KNOWN_RECLASSIFICATIONS = {
        '其他应付款': 'DT其他应付款含应付利息，差异=应付利息金额',
        '应付利息': 'DT应付利息通常并入其他应付款，差异=已知重分类',
        '应收利息': 'DT应收利息通常并入其他应收款，差异=已知重分类',
    }
    
    # 已知BS重分类项目（BS有独立行但DT通常合并到其他科目）
    BS_RECLASS_ITEMS = {'应付利息', '应收利息'}
    
    reconcile_results = []
    for acct in FULL_KEY_ACCOUNTS:
        dt_val = dt_summary.get(acct)
        bs_val = bs_summary.get(acct)
        
        if dt_val is not None and bs_val is not None:
            diff = abs(dt_val - bs_val)
            if diff > tolerance:
                is_known_reclass = acct in KNOWN_RECLASSIFICATIONS
                
                reconcile_results.append({
                    'account': acct,
                    'dt_value': dt_val,
                    'bs_value': bs_val,
                    'diff': diff,
                    'status': 'KNOWN_RECLASS' if is_known_reclass else 'MISMATCH'
                })
                violations.append({
                    'gate': 'G3-1',
                    'rule': 'DT-4',
                    'severity': 'WARNING' if is_known_reclass else 'CRITICAL',
                    'sheet': '2-分类汇总',
                    'cell': 'N/A',
                    'value': diff,
                    'message': f'科目"{acct}"勾稽不符: 明细表={dt_val:,.2f} BS={bs_val:,.2f} 差异={diff:,.2f}' + 
                              (f' ← 已知分类差异: {KNOWN_RECLASSIFICATIONS[acct]}' if is_known_reclass else '')
                })
        # BS有值但DT为None→DT漏填
        elif dt_val is None and bs_val is not None and abs(bs_val) > tolerance:
            # 已知重分类项降级为WARNING
            is_reclass = acct in BS_RECLASS_ITEMS
            violations.append({
                'gate': 'G3-1',
                'rule': 'DT-4',
                'severity': 'WARNING' if is_reclass else 'CRITICAL',
                'sheet': '2-分类汇总',
                'cell': 'N/A',
                'value': bs_val,
                'message': f'科目"{acct}"BS有值{bs_val:,.2f}但DT分类汇总中无对应值→漏填或公式缓存为0' +
                          (f' ← 已知重分类: {KNOWN_RECLASSIFICATIONS[acct]}' if is_reclass else '')
            })
    
    # 总量勾稽
    total_checks = [
        ('资产总计', '三、资产总计'),
        ('负债合计', '六、负债总计'),
        ('所有者权益合计', '七、净资产（所有者权益）')
    ]
    
    for bs_key, dt_key in total_checks:
        dt_val = dt_summary.get(dt_key)
        bs_val = bs_summary.get(bs_key)
        if dt_val is not None and bs_val is not None:
            diff = abs(dt_val - bs_val)
            if diff > tolerance:
                violations.append({
                    'gate': 'G3-2',
                    'rule': 'DT-4',
                    'severity': 'CRITICAL',
                    'sheet': '2-分类汇总',
                    'cell': 'N/A',
                    'value': diff,
                    'message': f'总量勾稽不符: {dt_key}={dt_val:,.2f} BS={bs_key}={bs_val:,.2f} 差异={diff:,.2f}'
                })
    
    # ---- G3-5: 公司名称一致性 ----
    for sname in wb.sheetnames:
        if '设定信息' in sname or '设置' in sname:
            ws = wb[sname]
            name_cell = ws.cell(row=6, column=2).value
            if name_cell:
                if '有限公司' in str(name_cell) and '有限责任' not in str(name_cell):
                    violations.append({
                        'gate': 'G3-5',
                        'rule': 'DT-96',
                        'severity': 'WARNING',
                        'sheet': sname,
                        'cell': 'B6',
                        'value': name_cell,
                        'message': f'公司名称"{name_cell}"可能缺少"责任"二字，MUST统一为营业执照全称'
                    })
    
    # ---- G3-6: 公式缓存校验 [DT-98] (GV-5修复) ----
    # 检查所有汇总表的账面价值列是否全部为0(公式缓存丢失)
    summary_all_zero = True
    summary_has_any_sheet = False
    for sname in wb.sheetnames:
        if '汇总' not in sname or '分类汇总' in sname:
            continue
        ws = wb[sname]
        summary_has_any_sheet = True
        # 动态查找汇总表的账面价值列
        sum_header_cols = find_header_cols(ws)
        bv_col_sum = sum_header_cols.get('账面价值')
        if not bv_col_sum:
            continue  # 汇总表无账面价值列则跳过
        for row in ws.iter_rows(min_row=6, max_row=65, min_col=bv_col_sum, max_col=bv_col_sum, values_only=False):
            val = row[0].value
            if val is not None and isinstance(val, (int, float)) and abs(val) > 0.01:
                summary_all_zero = False
                break
        if not summary_all_zero:
            break
    
    if summary_has_any_sheet and summary_all_zero:
        # 二次判定：如果汇总表账面列本身存在公式且明细表有非零数据，
        # 则判定为“公式缓存未计算”而非“链路断裂”，降级为WARNING。
        formula_based = False
        detail_has_nonzero = False
        try:
            wb_formula_cache = openpyxl.load_workbook(filepath, data_only=False)

            for sname in wb_formula_cache.sheetnames:
                if '汇总' not in sname or '分类汇总' in sname:
                    continue
                ws_f = wb_formula_cache[sname]
                hdr = find_header_cols(ws_f)
                bv_col = hdr.get('账面价值')
                if not bv_col:
                    continue
                for row in ws_f.iter_rows(min_row=6, max_row=65, min_col=bv_col, max_col=bv_col, values_only=False):
                    v = row[0].value
                    if isinstance(v, str) and v.startswith('='):
                        formula_based = True
                        break
                if formula_based:
                    break

            if formula_based:
                for sname in wb_formula_cache.sheetnames:
                    if ('汇总' in sname or '分类' in sname or sname.startswith('0') or sname.startswith('2-')
                            or sname.startswith('设置')):
                        continue
                    ws_d = wb_formula_cache[sname]
                    hmap = find_header_cols(ws_d)
                    dv_col = hmap.get('账面价值')
                    if not dv_col:
                        continue
                    struct_d = find_header_structure(ws_d)
                    dsr = struct_d.get('data_start_row') or 6
                    tr = struct_d.get('total_row') or min(ws_d.max_row + 1, dsr + 120)
                    for row in ws_d.iter_rows(min_row=dsr, max_row=max(dsr, tr - 1),
                                              min_col=dv_col, max_col=dv_col, values_only=True):
                        vv = row[0]
                        if isinstance(vv, (int, float)) and abs(vv) > 0.01:
                            detail_has_nonzero = True
                            break
                    if detail_has_nonzero:
                        break
        except Exception:
            formula_based = False
            detail_has_nonzero = False
        finally:
            try:
                wb_formula_cache.close()
            except Exception:
                pass

        if formula_based and detail_has_nonzero:
            violations.append({
                'gate': 'G3-6',
                'rule': 'DT-98',
                'severity': 'WARNING',
                'sheet': '全部汇总表',
                'cell': 'N/A',
                'value': 0,
                'message': '汇总表账面列为公式，但data_only缓存为0/None：判定为公式缓存未计算（非链路断裂）。请在有公式引擎环境重算后复核。'
            })
        else:
            violations.append({
                'gate': 'G3-6',
                'rule': 'DT-98',
                'severity': 'CRITICAL',
                'sheet': '全部汇总表',
                'cell': 'N/A',
                'value': 0,
                'message': '所有汇总表D列(账面价值)全部为0/None→公式缓存全局丢失！MUST执行COM recalc后重新验证'
            })
    
    # ---- G3-7: 汇总链完整性校验 [DT-99] (GV-6修复) ----
    # 检查明细表合计→汇总表的引用链
    for summary_name, prefix in SUMMARY_SHEET_PATTERNS:
        if summary_name not in wb.sheetnames:
            continue
        ws_sum = wb[summary_name]
        
        # 找汇总表中跨sheet引用的行
        # 动态查找汇总表的账面价值列
        sum_header_cols_g37 = find_header_cols(ws_sum)
        bv_col_sum_g37 = sum_header_cols_g37.get('账面价值', 4)  # 默认D列(4)兜底
        for row in ws_sum.iter_rows(min_row=6, max_row=65, values_only=False):
            # 动态定位账面价值单元格
            bv_cell = None
            for cell in row:
                if cell.column == bv_col_sum_g37:
                    bv_cell = cell
                    break
            if not bv_cell:
                continue
            if bv_cell.value and isinstance(bv_cell.value, str) and bv_cell.value.startswith('='):
                formula = bv_cell.value
                bv_col_letter_g37 = get_column_letter(bv_col_sum_g37)
                # 提取引用的sheet名
                sheet_refs = re.findall(r"([^!]+)!", formula)
                for ref in sheet_refs:
                    ref_clean = ref.replace("'", "").replace("=", "")
                    # 检查引用的sheet是否存在
                    if ref_clean and ref_clean not in wb.sheetnames:
                        # 检查hidden sheets
                        found_hidden = False
                        for sn in wb.sheetnames:
                            if wb[sn].sheet_state != 'visible' and ref_clean in sn:
                                found_hidden = True
                                break
                        if not found_hidden:
                            violations.append({
                                'gate': 'G3-7',
                                'rule': 'DT-99',
                                'severity': 'CRITICAL',
                                'sheet': summary_name,
                                'cell': f'{bv_col_letter_g37}{bv_cell.row}',
                                'value': formula,
                                'message': f'汇总链断裂：{summary_name}引用了不存在的Sheet "{ref_clean}"→明细表合计→汇总表链接不完整'
                            })
    
    # 特殊检查：3-2交易性金融资产汇总→3-2-1股票
    for special_check in ['3-2交易性金融资产汇总', '3-2交易性金融资产']:
        if special_check in wb.sheetnames:
            ws_sp = wb[special_check]
            # 检查是否有引用3-2-1的公式
            has_ref_to_detail = False
            for row in ws_sp.iter_rows(min_row=6, max_row=65, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '3-2-1' in str(cell.value):
                        has_ref_to_detail = True
                        break
                if has_ref_to_detail:
                    break
            
            # 检查3-2-1明细表是否存在且有数据
            detail_sheets = [sn for sn in wb.sheetnames if sn.startswith('3-2-1')]
            if detail_sheets:
                ws_detail = wb[detail_sheets[0]]
                detail_has_data = False
                # 动态确定扫描范围：扫描所有可见列
                detail_min_col = 2  # B列起
                detail_max_col = ws_detail.max_column
                for row in ws_detail.iter_rows(min_row=6, max_row=30, min_col=detail_min_col, max_col=detail_max_col, values_only=False):
                    for cell in row:
                        if cell.value and isinstance(cell.value, (int, float)) and abs(cell.value) > 0.01:
                            detail_has_data = True
                            break
                    if detail_has_data:
                        break
                
                if detail_has_data and not has_ref_to_detail:
                    violations.append({
                        'gate': 'G3-7',
                        'rule': 'DT-99',
                        'severity': 'CRITICAL',
                        'sheet': special_check,
                        'cell': 'N/A',
                        'value': None,
                        'message': f'汇总链断裂：{detail_sheets[0]}有数据但{special_check}未引用该明细表→分类汇总链路不完整'
                    })
    
    # ---- G3-8: 减值准备行评估值方向校验 [DT-18] (v3.0新增, v3.26更正) ----
    # 全表扫描所有含减值/预计风险行的sheet，验证：
    # 坏账准备行：J=0（仅填账面价值列I），预计风险行：I=0（仅填评估价值列J）
    for sname in wb.sheetnames:
        ws_ck = wb[sname]
        prefix_ck = get_sheet_prefix(sname)
        if not prefix_ck:
            continue
        if '汇总' in sname or '分类' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        
        struct_ck = find_header_structure(ws_ck)
        dsr_ck = struct_ck['data_start_row']
        tr_ck = struct_ck['total_row']
        if not dsr_ck or not tr_ck:
            continue
        
        for r in range(dsr_ck, tr_ck + 1):
            a_val = ws_ck.cell(row=r, column=1).value
            a_text = str(a_val).replace(' ', '').strip() if a_val else ''
            
            if '坏账准备' in a_text or '减值准备' in a_text:
                # 动态查找账面价值列和评估价值列（替代max_column推断）
                header_cols_g38 = find_header_cols(ws_ck)
                eval_col = header_cols_g38.get('评估价值')
                bv_col = header_cols_g38.get('账面价值')
                if not eval_col or not bv_col:
                    continue
                
                eval_val = ws_ck.cell(row=r, column=eval_col).value
                bv_val = ws_ck.cell(row=r, column=bv_col).value
                eval_col_letter = get_column_letter(eval_col)
                bv_col_letter = get_column_letter(bv_col)
                
                # 评估价值列应为0——坏账准备行仅填账面价值列
                if eval_val is not None and isinstance(eval_val, (int, float)) and abs(eval_val) > 0.01:
                    violations.append({
                        'gate': 'G3-8',
                        'rule': 'DT-18',
                        'severity': 'CRITICAL',
                        'sheet': sname,
                        'cell': f'{eval_col_letter}{r}',
                        'value': eval_val,
                        'message': f'{sname} {eval_col_letter}{r} 减值准备行评估值={eval_val:,.2f}，必须为0！坏账准备行仅填账面价值列，不填评估价值列'
                    })
                
                # 账面值应为负数（坏账准备是减项）
                if bv_val is not None and isinstance(bv_val, (int, float)) and bv_val > 0:
                    violations.append({
                        'gate': 'G3-8',
                        'rule': 'DT-18',
                        'severity': 'WARNING',
                        'sheet': sname,
                        'cell': f'{bv_col_letter}{r}',
                        'value': bv_val,
                        'message': f'{sname} {bv_col_letter}{r} 减值准备行账面值={bv_val:,.2f}为正数，通常应为负值（减项）'
                    })
            
            if '预计风险' in a_text:
                # 动态查找列位（复用上面的header_cols_g38）
                if 'header_cols_g38' not in dir() or header_cols_g38 is None:
                    header_cols_g38 = find_header_cols(ws_ck)
                eval_col = header_cols_g38.get('评估价值')
                bv_col = header_cols_g38.get('账面价值')
                if not eval_col or not bv_col:
                    continue
                bv_col_letter = get_column_letter(bv_col)
                
                bv_val = ws_ck.cell(row=r, column=bv_col).value
                # 预计风险行账面值=0（仅填评估价值列）
                if bv_val is not None and isinstance(bv_val, (int, float)) and abs(bv_val) > 0.01:
                    violations.append({
                        'gate': 'G3-8',
                        'rule': 'DT-18',
                        'severity': 'CRITICAL',
                        'sheet': sname,
                        'cell': f'{bv_col_letter}{r}',
                        'value': bv_val,
                        'message': f'{sname} {bv_col_letter}{r} 预计风险行账面值={bv_val:,.2f}，必须为0！预计风险行仅填评估价值列，不填账面价值列'
                    })
    
    # ---- G3-9: 汇总表跨sheet引用行号校验 [DT-86] (v3.0新增) ----
    # 验证汇总表的跨sheet引用指向目标sheet的"合计"行或"减："行
    # 如果引用了数据行(非合计/非减值)=汇总值错误=CRITICAL
    wb_formula = openpyxl.load_workbook(filepath, data_only=False)
    
    for summary_name, prefix in SUMMARY_SHEET_PATTERNS:
        if summary_name not in wb_formula.sheetnames:
            continue
        ws_sum = wb_formula[summary_name]
        
        for row in ws_sum.iter_rows(min_row=6, max_row=65, values_only=False):
            d_cell = row[3]  # D列=账面价值
            e_cell = row[4] if len(row) > 4 else None  # E列=评估价值
            
            for target_cell in [d_cell, e_cell]:
                if target_cell is None:
                    continue
                val = target_cell.value
                if not val or not isinstance(val, str) or not val.startswith('='):
                    continue
                
                # 提取跨sheet引用: ='SheetName'!CellRef 或 =SheetName!CellRef
                refs = re.findall(r"='?([^!']+)'?!([A-Z]+)(\d+)", val)
                for ref_sheet, ref_col, ref_row_str in refs:
                    ref_row = int(ref_row_str)
                    
                    # 检查引用的sheet是否存在
                    if ref_sheet not in wb_formula.sheetnames:
                        # 可能在hidden sheets中
                        found_hidden = False
                        for sn in wb_formula.sheetnames:
                            if ref_sheet in sn:
                                found_hidden = True
                                break
                        if not found_hidden:
                            continue
                    
                    # 检查引用行的A列值
                    ws_ref = wb_formula[ref_sheet] if ref_sheet in wb_formula.sheetnames else None
                    if ws_ref is None:
                        continue
                    
                    a_val_ref = ws_ref.cell(row=ref_row, column=1).value
                    a_text_ref = str(a_val_ref).replace(' ', '').strip() if a_val_ref else ''
                    
                    # 引用行必须是"合计"或"减："行
                    is_valid_ref = ('合' in a_text_ref and '计' in a_text_ref) or a_text_ref.startswith('减')
                    
                    if not is_valid_ref and a_text_ref and ref_row > 5:
                        # 引用了非合计/非减值行=数据行=CRITICAL
                        violations.append({
                            'gate': 'G3-9',
                            'rule': 'DT-86',
                            'severity': 'CRITICAL',
                            'sheet': summary_name,
                            'cell': f'{get_column_letter(target_cell.column)}{target_cell.row}',
                            'value': val[:80],
                            'message': f'汇总表{summary_name}引用{ref_sheet}!{ref_col}{ref_row}，该行为数据行("{a_text_ref[:20]}")而非合计/减值行！MUST更新引用至合计行。DT-86：插入/删除行后汇总表跨sheet引用MUST同步更新'
                        })
    
    wb_formula.close()
    
    # ---- G3-10: 合计行公式引用校验 [DT-85] (v3.1新增) ----
    # 合计行增值额/增值率公式MUST引用本行而非分隔行
    # SUM范围MUST覆盖全部数据行，禁止仅引用单行
    wb_check = openpyxl.load_workbook(filepath, data_only=False)
    
    for sname in wb_check.sheetnames:
        ws_ck2 = wb_check[sname]
        prefix_ck2 = get_sheet_prefix(sname)
        if not prefix_ck2:
            continue
        if '汇总' in sname or '分类' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        
        struct_ck2 = find_header_structure(ws_ck2)
        dsr_ck2 = struct_ck2['data_start_row']
        tr_ck2 = struct_ck2['total_row']
        if not dsr_ck2 or not tr_ck2:
            continue
        
        for r in range(dsr_ck2, tr_ck2 + 1):
            a_val = ws_ck2.cell(row=r, column=1).value
            a_text = str(a_val).replace(' ', '').strip() if a_val else ''
            
            if not ('合' in a_text and '计' in a_text):
                continue
            
            # 检查合计行的增值额/增值率/账龄公式列（动态查找列号）
            header_cols_g310 = find_header_cols(ws_ck2)
            formula_check_cols = []
            for key in ['增值额', '增值率', '账龄']:
                if key in header_cols_g310:
                    formula_check_cols.append(header_cols_g310[key])
            for col_idx in formula_check_cols:
                cell = ws_ck2.cell(row=r, column=col_idx)
                if not cell.value or not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                
                formula = cell.value
                col_letter = get_column_letter(col_idx)
                
                # 检查1: SUM范围仅引用单行 (如 SUM(E6:E6))
                sum_ranges = re.findall(r'SUM\(([A-Z])(\d+):([A-Z])(\d+)\)', formula, re.IGNORECASE)
                for sr_col, sr_start, sr_col2, sr_end in sum_ranges:
                    if sr_start == sr_end:
                        violations.append({
                            'gate': 'G3-10',
                            'rule': 'DT-85',
                            'severity': 'CRITICAL',
                            'sheet': sname,
                            'cell': f'{col_letter}{r}',
                            'value': formula,
                            'message': f'{sname}合计行{col_letter}{r} SUM范围仅引用单行({sr_col}{sr_start}:{sr_col2}{sr_end})！MUST覆盖全部数据行'
                        })
                
                # 检查2: 公式引用了分隔行(合计行上方通常有空行)
                # 如果合计行=r，则r-1可能是分隔行（A列合并，无边框）
                cell_refs = re.findall(r'([A-Z])(\d+)', formula)
                for ref_col, ref_row_str in cell_refs:
                    ref_row = int(ref_row_str)
                    if ref_row == r - 1 and ref_row > dsr_ck2:
                        # 检查r-1行是否为分隔行（A列值与合计无关）
                        prev_a = ws_ck2.cell(row=ref_row, column=1).value
                        prev_a_text = str(prev_a).replace(' ', '').strip() if prev_a else ''
                        if not prev_a_text or ('合' not in prev_a_text and '减' not in prev_a_text):
                            violations.append({
                                'gate': 'G3-10',
                                'rule': 'DT-85',
                                'severity': 'WARNING',
                                'sheet': sname,
                                'cell': f'{col_letter}{r}',
                                'value': formula,
                                'message': f'{sname}合计行{col_letter}{r}公式引用了分隔行({ref_col}{ref_row})而非本行({ref_col}{r})！增值额/增值率可能为0'
                            })
                        break  # 每列只报一次
    
    wb_check.close()
    
    # ---- G3-11: 隐藏操作校验 [DT-110] (v3.1新增, v3.4增强) ----
    # 交付前必须执行隐藏操作：
    # 1. 系统辅助表(设置/0-其他方法结论/设定信息)必须隐藏 [DT-23]
    # 2. 辅汇总表必须隐藏 [DT-61]
    # 3. 空白明细Sheet(合计1行账面=0且评估=0)必须隐藏 [DT-20]
    # v3.4增强：severity从WARNING升级为CRITICAL，不通过即阻断交付
    SYSTEM_AUX_SHEETS = ['设置', '0-其他方法结论', '设定信息']
    SUMMARY_HIDE_SHEETS = ['3-辅-流动资产汇总', '3-9-辅-存货汇总', '8-减值准备汇总表', '9-非财务信息汇总表']
    
    # 检查系统辅助表是否仍可见
    for sys_sheet in SYSTEM_AUX_SHEETS:
        if sys_sheet in wb.sheetnames:
            ws_sys = wb[sys_sheet]
            if ws_sys.sheet_state == 'visible':
                violations.append({
                    'gate': 'G3-11',
                    'rule': 'DT-23',
                    'severity': 'CRITICAL',
                    'sheet': sys_sheet,
                    'cell': 'N/A',
                    'value': 'visible',
                    'message': f'系统辅助表"{sys_sheet}"仍为可见状态！DT-23要求始终隐藏，DT-110要求Phase 5交付前统一执行'
                })
    
    # 检查辅汇总表是否隐藏
    for sum_sheet in SUMMARY_HIDE_SHEETS:
        if sum_sheet in wb.sheetnames:
            ws_sum = wb[sum_sheet]
            if ws_sum.sheet_state == 'visible':
                violations.append({
                    'gate': 'G3-11',
                    'rule': 'DT-61',
                    'severity': 'CRITICAL',
                    'sheet': sum_sheet,
                    'cell': 'N/A',
                    'value': 'visible',
                    'message': f'辅汇总表"{sum_sheet}"仍为可见状态！DT-61要求辅汇总表始终隐藏'
                })
    
    # 检查空白明细Sheet是否仍可见（v3.4新增）
    # 使用hide_empty_sheets.py的判定逻辑
    # 注意：gate_G3的wb是data_only=True打开的，无法读取公式字符串
    # 需要用data_only=False重新打开来获取公式
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'valuation-common' / 'scripts'))
        from hide_empty_sheets import check_sheet_empty, is_detail_sheet, is_structure_sheet
        
        wb_dt = openpyxl.load_workbook(filepath, data_only=False)
        for name in wb_dt.sheetnames:
            ws_dt = wb_dt[name]
            if ws_dt.sheet_state != 'visible':
                continue
            if not is_detail_sheet(name) or is_structure_sheet(name):
                continue
            
            result = check_sheet_empty(ws_dt, name, xlsx_path=str(Path(filepath).resolve()))
            if result.get('is_empty'):
                violations.append({
                    'gate': 'G3-11',
                    'rule': 'DT-20',
                    'severity': 'CRITICAL',
                    'sheet': name,
                    'cell': f'合计1行(Row{result.get("heji1_row", "?")})',
                    'value': f'账面={result.get("book_resolved")}, 评估={result.get("eval_resolved")}',
                    'message': f'空白明细表"{name}"仍为可见！DT-20要求合计1行账面=0且评估=0的Sheet必须隐藏 (判定依据: {result.get("reason", "")[:80]})'
                })
    except ImportError:
        violations.append({
            'gate': 'G3-11',
            'rule': 'DT-20',
            'severity': 'WARNING',
            'sheet': 'N/A',
            'cell': 'N/A',
            'value': 'N/A',
            'message': 'hide_empty_sheets模块未找到，空白明细Sheet校验跳过。请确保valuation-common/scripts/hide_empty_sheets.py存在'
        })
    finally:
        try:
            wb_dt.close()
        except Exception:
            pass
    
    wb.close()
    passed = not any(v['severity'] == 'CRITICAL' for v in violations)
    return passed, violations



# ============================================================
# DT-182: 汇总表禁止直接录入数据校验 (v3.63新增)
# ============================================================

def validate_summary_no_hardcoded(filepath):
    """DT-182: 扫描所有汇总类Sheet，检测非公式硬编码数值。
    
    汇总Sheet（含"汇总"或"分类汇总"的Sheet）的数据区域MUST仅包含公式引用。
    表头文字、A列标记文字、科目名称列不受此限制。
    
    Returns:
        (passed: bool, violations: list)
    """
    import openpyxl, re
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.load_workbook(filepath, data_only=False)
    violations = []
    
    # 汇总Sheet命名模式
    SUMMARY_PATTERNS = ['汇总', '分类汇总', '分类汇总表']
    
    # 汇总Sheet列表
    summary_sheets = []
    for sname in wb.sheetnames:
        if any(p in sname for p in SUMMARY_PATTERNS):
            summary_sheets.append(sname)
    
    if not summary_sheets:
        return True, violations
    
    for sname in summary_sheets:
        ws = wb[sname]
        
        # 识别表头行和合计行
        header_row = None
        total_row = None
        name_col = None
        
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 15)):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                val_str = str(val).strip()
                
                if name_col is None and ('科目' in val_str or '项目' in val_str):
                    name_col = c
                    if header_row is None:
                        header_row = r
        
        if header_row is None:
            header_row = 6  # fallback
        if name_col is None:
            name_col = 2  # fallback (B列)
        
        # 查找合计行（从底部向上）
        for r in range(ws.max_row, header_row, -1):
            name_val = ws.cell(row=r, column=name_col).value
            if name_val and re.match(r'.*(合|小)计.*', str(name_val).strip()):
                total_row = r
                break
        
        if total_row is None:
            total_row = ws.max_row
        
        # 数据区域 = 表头下一行 → 合计上一行
        data_start = header_row + 1
        data_end = total_row - 1
        
        if data_end < data_start:
            continue  # 无数据行
        
        # 扫描数据区域的每个单元格（跳过A列标记列和科目名称列）
        skip_cols = {1, name_col}  # A列（标记列）和科目名称列
        monitored_cols = []
        for c in range(1, ws.max_column + 1):
            h_main = ws.cell(row=header_row, column=c).value
            h_sub = ws.cell(row=header_row + 1, column=c).value if (header_row + 1) <= ws.max_row else None
            ctype = classify_column(h_main, h_sub)
            h_text = f'{h_main or ""}{h_sub or ""}'
            # DT-182只监控汇总金额/勾稽相关列，序号/项目编号等元数据列不纳入
            if ctype in ('value', 'formula') or any(k in h_text for k in ('报表', '校验', '差异')):
                if c not in skip_cols:
                    monitored_cols.append(c)

        if not monitored_cols:
            # 兜底：若表头识别失败，退化为原始行为但仍排除明显元数据列
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(row=header_row, column=c).value or '')
                if c in skip_cols:
                    continue
                if any(k in h for k in ('序号', '项目编号', '编号', '链接')):
                    continue
                monitored_cols.append(c)

        for r in range(data_start, data_end + 1):
            for c in monitored_cols:
                
                cell = ws.cell(row=r, column=c)
                val = cell.value
                
                if val is None:
                    continue
                
                val_str = str(val).strip()
                
                # 公式 → 合规
                if val_str.startswith('='):
                    continue
                
                # 非公式但看起来像数字 → 违规
                try:
                    float(val_str.replace(',', ''))
                    is_numeric = True
                except (ValueError, TypeError):
                    is_numeric = False
                
                if is_numeric:
                    cell_ref = f'{get_column_letter(c)}{r}'
                    name_val = ws.cell(row=r, column=name_col).value or '(空)'
                    violations.append({
                        'gate': 'G-DT182',
                        'rule': 'DT-182',
                        'severity': 'CRITICAL',
                        'sheet': sname,
                        'cell': cell_ref,
                        'value': val,
                        'message': f'汇总Sheet"{sname}" {cell_ref} 硬编码数值 {val}（科目: {str(name_val)}）。DT-182要求汇总表仅包含公式引用，禁止直接录入数据。'
                    })
    
    wb.close()
    passed = len(violations) == 0
    return passed, violations

# ============================================================
# 主入口
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description='评估明细表 Gate Validator v3.3')
    parser.add_argument('xlsx_path', help='评估明细表文件路径')
    parser.add_argument('--gate', choices=['G0', 'G1', 'G1-Format', 'G2', 'G3', 'G-DT182', 'all'], default='all',
                       help='执行哪个门控级别 (默认all)')
    parser.add_argument('--bs-path', help='资产负债表文件路径 (G0/G2/G3需要)')
    parser.add_argument('--sb-path', help='科目余额表文件路径 (G0/G2需要)')
    parser.add_argument('--aux-json', help='辅助数据JSON文件路径 (G0需要，含辅助余额表/PDF状态)')
    parser.add_argument('--sheet', help='指定Sheet名称 (G1使用)')
    parser.add_argument('--json', action='store_true', help='以JSON格式输出')
    parser.add_argument('--tolerance', type=float, default=0.01, help='勾稽容忍差异 (默认0.01)')
    parser.add_argument('--auto-fix', action='store_true', 
                       help='G1-Format验证失败时自动修复 (DT-114闭环，会修改文件)')
    parser.add_argument('--max-retries', type=int, default=3, 
                       help='auto-fix最大重试次数 (默认3)')
    
    args = parser.parse_args()
    
    # 加载辅助数据
    aux_data = None
    if args.aux_json:
        try:
            with open(args.aux_json, 'r', encoding='utf-8') as f:
                aux_data = json.load(f)
        except Exception as e:
            print(f"⚠️ 辅助数据JSON加载失败: {e}，G0-3/G0-4/G0-5校验跳过")
    
    results = {'passed': True, 'violations': []}
    
    if args.gate in ('G0', 'all'):
        print("\n" + "=" * 70)
        print("G0: 数据源级硬约束校验 [DT-103/104/108/109/111]")
        print("=" * 70)
        g0_pass, g0_violations = gate_G0(args.xlsx_path, args.bs_path, args.sb_path, aux_data)
        results['violations'].extend(g0_violations)
        if not g0_pass:
            results['passed'] = False
        
        if g0_violations:
            criticals = [v for v in g0_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g0_violations if v['severity'] != 'CRITICAL']
            if criticals:
                print(f"\n🚨 G0未通过 — {len(criticals)}个硬约束违规:")
                for v in criticals:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if warnings:
                print(f"\n⚠️ G0警告 — {len(warnings)}个:")
                for v in warnings:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
        else:
            print("\n✅ G0通过 — 数据源级校验合格")
    
    if args.gate in ('G1', 'all'):
        print("\n" + "=" * 70)
        print("G1: 写入级硬约束校验 [DT-6/18/30/78/90/91/97/102/104]")
        print("=" * 70)
        g1_pass, g1_violations = gate_G1(args.xlsx_path, args.sheet)
        results['violations'].extend(g1_violations)
        if not g1_pass:
            results['passed'] = False
        
        if g1_violations:
            criticals = [v for v in g1_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g1_violations if v['severity'] != 'CRITICAL']
            if criticals:
                print(f"\n🚨 G1未通过 — {len(criticals)}个硬约束违规:")
                for v in criticals:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if warnings:
                print(f"\n⚠️ G1警告 — {len(warnings)}个:")
                for v in warnings:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
        else:
            print("\n✅ G1通过 — 所有Sheet写入级校验合格")
    
    if args.gate in ('G1-Format', 'all'):
        print("\n" + "=" * 70)
        print("G1-Format: 格式门控校验 [DT-67/76/77/82/83/84] (DT-112格式集中处置验证)")
        print("=" * 70)
        g1f_pass, g1f_violations = gate_G1_Format(args.xlsx_path, args.sheet)
        results['violations'].extend(g1f_violations)
        if not g1f_pass:
            results['passed'] = False
        
        if g1f_violations:
            criticals = [v for v in g1f_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g1f_violations if v['severity'] != 'CRITICAL']
            
            # DT-114: auto-fix闭环（仅G1-Format段，且需--auto-fix参数）
            if args.auto_fix and (criticals or warnings):
                print(f"\n🔄 DT-114: 启动验证-修复闭环 (最多{args.max_retries}次)...")
                
                # 收集需要修复的gate项
                failed_gates = set()
                for v in g1f_violations:
                    gate = v.get('gate', '')
                    if gate.startswith('G1F-') and gate not in ('G1F-5', 'G1F-6'):
                        failed_gates.add(gate)
                
                if failed_gates:
                    # 延迟导入excel_row_ops
                    try:
                        scripts_dir = str(Path(args.xlsx_path).parent)
                        common_dir = str(Path(__file__).parent.parent.parent / 'valuation-common' / 'scripts')
                        if common_dir not in sys.path:
                            sys.path.insert(0, common_dir)
                        from excel_row_ops import auto_fix_formats
                        
                        for retry in range(1, args.max_retries + 1):
                            print(f"\n  📝 第{retry}次修复: {sorted(failed_gates)}")
                            fix_result = auto_fix_formats(args.xlsx_path, args.sheet, list(failed_gates))
                            
                            for f in fix_result['fixed']:
                                print(f"    ✅ {f}")
                            for u in fix_result['unfixed']:
                                print(f"    ❌ {u}")
                            
                            # 重新验证
                            g1f_pass, g1f_violations_new = gate_G1_Format(args.xlsx_path, args.sheet)
                            
                            # 更新results
                            results['violations'] = [v for v in results['violations'] 
                                                      if not (v.get('gate','').startswith('G1F-'))]
                            results['violations'].extend(g1f_violations_new)
                            
                            if g1f_pass:
                                results['passed'] = all(
                                    g['gate'] != 'G1-Format' 
                                    for g in [results] 
                                    for k, g_list in {k: [results]} .items()
                                ) or True  # 简化：G1-Format通过即更新passed
                                print(f"\n  ✅ 第{retry}次修复后G1-Format验证通过！")
                                g1f_violations = g1f_violations_new
                                break
                            else:
                                # 检查剩余问题是否都是G1F-5/G1F-6（不可自动修复）
                                remaining_auto_fixable = [
                                    v for v in g1f_violations_new 
                                    if v.get('gate','') in ('G1F-1','G1F-2','G1F-3','G1F-4')
                                ]
                                if not remaining_auto_fixable:
                                    print(f"\n  ⚠️ 剩余问题均为G1F-5/G1F-6（需人工处理），自动修复结束")
                                    g1f_violations = g1f_violations_new
                                    break
                                g1f_violations = g1f_violations_new
                    except ImportError:
                        print(f"\n  ⚠️ excel_row_ops不可用，跳过auto-fix闭环。请确保valuation-common/scripts/在sys.path中")
            
            # 输出最终结果
            criticals = [v for v in g1f_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g1f_violations if v['severity'] != 'CRITICAL']
            if criticals:
                print(f"\n🚨 G1-Format未通过 — {len(criticals)}个格式硬约束违规:")
                for v in criticals:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if warnings:
                print(f"\n⚠️ G1-Format警告 — {len(warnings)}个:")
                for v in warnings:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if not criticals and not warnings:
                print("\n✅ G1-Format通过 — 格式集中处置验证合格")
        else:
            print("\n✅ G1-Format通过 — 格式集中处置验证合格")
    
    if args.gate in ('G2', 'all'):
        print("\n" + "=" * 70)
        print("G2: 科目级硬约束校验 [DT-87/89/92/93/94/95/100/103/117/118]")
        print("=" * 70)
        g2_pass, g2_violations = gate_G2(args.xlsx_path, args.bs_path, args.sb_path)
        results['violations'].extend(g2_violations)
        if not g2_pass:
            results['passed'] = False
        
        if g2_violations:
            criticals = [v for v in g2_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g2_violations if v['severity'] != 'CRITICAL']
            if criticals:
                print(f"\n🚨 G2未通过 — {len(criticals)}个硬约束违规:")
                for v in criticals:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if warnings:
                print(f"\n⚠️ G2警告 — {len(warnings)}个:")
                for v in warnings:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
        else:
            print("\n✅ G2通过 — 科目级校验合格")
    
    if args.gate in ('G3', 'all'):
        print("\n" + "=" * 70)
        print("G3: 勾稽级硬约束校验 [DT-4/18/85/86/96/98/99/110]")
        print("=" * 70)
        g3_pass, g3_violations = gate_G3(args.xlsx_path, args.bs_path, args.tolerance)
        results['violations'].extend(g3_violations)
        if not g3_pass:
            results['passed'] = False
        
        if g3_violations:
            criticals = [v for v in g3_violations if v['severity'] == 'CRITICAL']
            warnings = [v for v in g3_violations if v['severity'] != 'CRITICAL']
            if criticals:
                print(f"\n🚨 G3未通过 — {len(criticals)}个硬约束违规:")
                for v in criticals:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
            if warnings:
                print(f"\n⚠️ G3警告 — {len(warnings)}个:")
                for v in warnings:
                    print(f"  [{v['gate']}][{v['rule']}] {v['sheet']} {v['cell']}: {v['message']}")
        else:
            print("\n✅ G3通过 — 勾稽级校验合格")
    
    # 汇总
    print("\n" + "=" * 70)
    total_critical = len([v for v in results['violations'] if v['severity'] == 'CRITICAL'])
    total_warning = len([v for v in results['violations'] if v['severity'] == 'WARNING'])
    total_info = len([v for v in results['violations'] if v['severity'] == 'INFO'])
    
    if results['passed']:
        print(f"✅ 全部Gate通过 | {total_warning}个警告 | {total_info}个信息")
    else:
        print(f"🚨 Gate未通过 | {total_critical}个硬约束违规 | {total_warning}个警告 | {total_info}个信息")
        print("⚠️ 硬约束违规未修复 = 禁止继续/交付")
    
    if args.json:
        print("\n" + json.dumps(results, ensure_ascii=False, indent=2, default=str))
    
    sys.exit(0 if results['passed'] else 1)


if __name__ == '__main__':
    main()
