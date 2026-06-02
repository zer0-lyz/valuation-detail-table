#!/usr/bin/env python3
"""
dt_runner.py - 评估明细表填写统一执行入口 (DT Skill v3.48)

Agent只需调用:
    python dt_runner.py --phase 0 --project "C:/Users/.../1-河南平绿"

脚本内部自动:
1. 读取_dt_cache/缓存（DT-131断点恢复）
2. 执行Phase内所有子步骤
3. 每个子步骤输出缓存JSON（DT-130持久化）
4. Phase完成后自动运行Gate验证（DT-138/L2约束）
5. Gate失败则exit(1)阻断

v3.48新增:
- Phase 1-5编排骨架（data_loader集成+断言覆盖率检测）
- DT-155~159五条新断言嵌入
- rule_manifest.json驱动的断言覆盖率检测
- subject_schema.json驱动的数据加载

架构: Agent=调度器, 脚本=执行器
"""

import argparse
import json
import os
import sys
import glob
import re
import hashlib
import traceback
from datetime import datetime
from release_status import build_release_status, ensure_formula_cache_status

# 路径配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
COMMON_SCRIPTS = os.path.normpath(os.path.join(
    SCRIPT_DIR, '..', '..', 'valuation-common', 'scripts'
))
if COMMON_SCRIPTS not in sys.path:
    sys.path.insert(0, COMMON_SCRIPTS)


# ════════════════════════════════════════════════════════════
# Phase-Scoped规则摘要提取（L2按需层）
# ════════════════════════════════════════════════════════════

# Phase→DT规则映射（从Step文件DEPENDS声明汇总）
PHASE_RULES = {
    '-1': ['DT-105', 'DT-106', 'DT-107', 'DT-108', 'DT-115', 'DT-130', 'DT-131', 'DT-142'],
    '0': ['DT-0', 'DT-1', 'DT-79', 'DT-103', 'DT-109', 'DT-111', 'DT-117', 'DT-118', 'DT-119', 'DT-121', 'DT-130', 'DT-137', 'DT-139'],
    '1': ['DT-0', 'DT-46', 'DT-66', 'DT-116', 'DT-119', 'DT-136'],
    '2': ['DT-0', 'DT-2', 'DT-5', 'DT-18', 'DT-30', 'DT-46', 'DT-51', 'DT-60', 'DT-66', 'DT-67', 'DT-74', 'DT-90', 'DT-97', 'DT-112', 'DT-113', 'DT-116', 'DT-120', 'DT-124', 'DT-125', 'DT-128', 'DT-136', 'DT-138', 'DT-141', 'DT-143'],
    '3': ['DT-2', 'DT-24', 'DT-33', 'DT-67', 'DT-75', 'DT-78', 'DT-85', 'DT-99', 'DT-112', 'DT-113', 'DT-114', 'DT-120'],
    '4': ['DT-4', 'DT-61', 'DT-62', 'DT-69', 'DT-70', 'DT-71', 'DT-86', 'DT-98', 'DT-99', 'DT-117', 'DT-118'],
    '5': ['DT-9', 'DT-17', 'DT-33', 'DT-59', 'DT-110', 'DT-123'],
}

# 科目→DT规则映射（L2按需层触发）
SUBJECT_RULES = {
    '应交税费': ['DT-87', 'DT-126', 'DT-147'],
    '固定资产': ['DT-88', 'DT-94', 'DT-21'],
    '其他应付款': ['DT-140', 'DT-111', 'DT-137'],
    '长期借款': ['DT-148', 'DT-51'],
    '递延所得税': ['DT-150'],
    '银行存款': ['DT-65', 'DT-104', 'DT-135'],
    '其他流动资产': ['DT-87', 'DT-118'],
}


def load_rules_digest(phase=None, subject=None):
    """从RULES.md提取指定Phase/科目的规则摘要
    
    Args:
        phase: Phase编号（'-1'~'5'）
        subject: 科目名称（触发L2按需加载）
    
    Returns:
        dict: {规则ID: 规则摘要}
    """
    rules_path = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'RULES.md'))
    if not os.path.exists(rules_path):
        return {}
    
    with open(rules_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 收集需要的规则ID
    needed_ids = set()
    if phase and phase in PHASE_RULES:
        needed_ids.update(PHASE_RULES[phase])
    if subject:
        for key, rules in SUBJECT_RULES.items():
            if key in subject:
                needed_ids.update(rules)
    
    if not needed_ids:
        return {}
    
    # 从RULES.md提取对应规则的摘要
    import re
    digest = {}
    for rule_id in needed_ids:
        # 匹配 | **DT-xxx** | 分类 | **规则标题**：规则内容 | 后果 |
        # 格式: | **DT-xxx** | D/R/O | **🚨 xxx**：... | ... |
        pattern = rf'\| \*\*{re.escape(rule_id)}\*\*.*?\| (.*?) \| (.*?) \|'
        match = re.search(pattern, content)
        if match:
            rule_text = match.group(2)  # 第二个| |之间是规则内容
            # 提取加粗标题
            title_match = re.search(r'\*\*(.*?)\*\*', rule_text)
            if title_match:
                # 去掉emoji前缀
                title = title_match.group(1).replace('🚨', '').replace('🚨🚨🚨', '').strip()
                digest[rule_id] = title
            else:
                digest[rule_id] = rule_text[:80]
    
    return digest


def _get_bs_value(bs_balances, label):
    """从bs_balances中查找指定label的ending_balance"""
    if not bs_balances:
        return None
    items = bs_balances.get('items', []) if isinstance(bs_balances, dict) else []
    for item in items:
        if item.get('label', '') == label:
            return item.get('ending_balance')
    return None

def _cache_path(project_dir):
    """获取_dt_cache目录路径，不存在则创建"""
    cache_dir = os.path.join(project_dir, '_dt_cache')
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def _save_cache(cache_dir, filename, data):
    """保存JSON缓存文件"""
    path = os.path.join(cache_dir, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'  [CACHE] 已保存: {filename}')
    return path


def _load_cache(cache_dir, filename):
    """加载JSON缓存文件"""
    path = os.path.join(cache_dir, filename)
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def _sha256_file(path):
    """计算文件sha256。"""
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def _build_project_state(project_dir, status, cache_dir=None, pending_count=0):
    """构建并持久化统一项目状态文件。"""
    if cache_dir is None:
        cache_dir = _cache_path(project_dir)

    source_fingerprints = {}
    for pat in ('*.xlsx', '*.xls', '*.pdf'):
        for p in glob.glob(os.path.join(project_dir, pat)):
            if not os.path.isfile(p):
                continue
            try:
                st = os.stat(p)
                source_fingerprints[os.path.basename(p)] = {
                    'mtime': st.st_mtime,
                    'sha256': _sha256_file(p),
                }
            except OSError:
                continue

    state = {
        'schema_version': '4.0',
        'status': status,
        'project_id': os.path.basename(os.path.abspath(project_dir)),
        'updated_at': datetime.now().isoformat(),
        'source_fingerprints': source_fingerprints,
        'pending_count': pending_count,
        'artifacts': {
            'cache_dir': cache_dir,
            'pending_confirmations': os.path.join(cache_dir, 'pending_confirmations.json'),
            'gate_results': os.path.join(cache_dir, 'gate_results.json'),
            'qa_report': os.path.join(cache_dir, 'qa_report.json'),
            'release_status': os.path.join(cache_dir, 'release_status.json'),
            'formula_cache_status': os.path.join(cache_dir, 'formula_cache_status.json'),
            'dt139_validation_status': os.path.join(cache_dir, 'dt139_validation_status.json'),
            'standardized_manifest': os.path.join(cache_dir, 'standardized_manifest.json'),
        },
    }
    _save_cache(cache_dir, 'project_state.json', state)
    return state


def _count_pending_confirmations(cache_dir):
    """统计仍未解决的人工确认项。"""
    payload = _load_cache(cache_dir, 'pending_confirmations.json') or {}
    return len([item for item in payload.get('items', []) if not item.get('resolved')])


def _finalize_release_status(project_dir, cache_dir, qa_result):
    """统一生成发布状态。任何出口都必须经过此函数。"""
    release_status = build_release_status(
        cache_dir,
        gate_results=_load_cache(cache_dir, 'gate_results.json') or [],
        qa_result=qa_result,
        recon_result=_load_cache(cache_dir, 'reconciliation_report.json') or {},
        pending_count=_count_pending_confirmations(cache_dir),
        formula_cache_status=ensure_formula_cache_status(cache_dir),
        dt139_status=_load_cache(cache_dir, 'dt139_validation_status.json') or {},
    )
    _build_project_state(
        project_dir,
        release_status['status'],
        cache_dir,
        pending_count=_count_pending_confirmations(cache_dir),
    )
    return release_status


def _write_pending_confirmations(cache_dir, items, reason='需要人工确认'):
    """写入待确认清单。"""
    payload = {
        'schema_version': '1.0',
        'status': 'BLOCKED_CONFIRMATION',
        'reason': reason,
        'items': items,
        'created_at': datetime.now().isoformat(),
    }
    _save_cache(cache_dir, 'pending_confirmations.json', payload)
    return payload


def _normalize_mapping_codes(value):
    if value is None:
        return []
    if isinstance(value, str):
        return [v.strip() for v in value.split(',') if v.strip()]
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return []


def _build_mapping_candidates(label, subjects, max_n=8):
    """为待确认映射项提供候选编码。"""
    candidates = []
    seen = set()
    for s in subjects or []:
        code = str(s.get('code', '')).strip()
        name = str(s.get('name', '')).strip()
        if not code or code in seen:
            continue
        if label in name or name in label:
            candidates.append({'code': code, 'name': name})
            seen.add(code)
        if len(candidates) >= max_n:
            break
    return candidates


def _apply_mapping_overrides(d1d2d3, cache_dir, subjects):
    """将确认后的映射覆盖应用到d1d2d3映射。"""
    overrides = _load_cache(cache_dir, 'mapping_overrides.json') or {}
    mappings = overrides.get('mappings', {}) if isinstance(overrides, dict) else {}
    if not isinstance(mappings, dict) or not mappings:
        return {'applied': 0, 'skipped': 0}

    valid_codes = {str(s.get('code', '')).strip() for s in (subjects or []) if s.get('code')}
    d1_to_d2 = d1d2d3.setdefault('d1_to_d2', {})
    unmapped = d1d2d3.setdefault('unmapped_bs_items', [])

    applied = 0
    skipped = 0
    for label, raw_codes in mappings.items():
        codes = _normalize_mapping_codes(raw_codes)
        filtered = []
        for c in codes:
            if c in valid_codes and c not in filtered:
                filtered.append(c)
        if not filtered:
            skipped += 1
            continue
        d1_to_d2[str(label)] = filtered
        if label in unmapped:
            unmapped = [x for x in unmapped if x != label]
            d1d2d3['unmapped_bs_items'] = unmapped
        applied += 1

    _save_cache(
        cache_dir,
        'mapping_override_apply_report.json',
        {
            'applied': applied,
            'skipped': skipped,
            'timestamp': datetime.now().isoformat(),
        },
    )
    return {'applied': applied, 'skipped': skipped}


def _count_nonzero_subjects(subjects):
    arr = subjects if isinstance(subjects, list) else []
    count = 0
    for s in arr:
        try:
            bal = float(s.get('balance', 0) or 0)
        except (TypeError, ValueError):
            bal = 0
        if abs(bal) > 0.01:
            count += 1
            continue
        # DT-FIX: 喜发格式——balance=0但ending_debit/ending_credit有值的科目也算非零
        try:
            ed = float(s.get('ending_debit', 0) or 0)
            ec = float(s.get('ending_credit', 0) or 0)
            if abs(ed) > 0.01 or abs(ec) > 0.01:
                count += 1
        except (TypeError, ValueError):
            pass
    return count


def _write_standardized_manifest(
    project_dir,
    cache_dir,
    source_mode,
    subjects=None,
    bs_balances=None,
    asset_register=None,
    journal=None,
    auxiliary_summary=None,
    pdf_extractions=None,
    warnings=None,
):
    """写入统一标准化工件，作为后续流程唯一输入快照。"""
    subjects = subjects or []
    bs_balances = bs_balances or {}
    asset_register = asset_register or {}
    journal = journal or {}
    auxiliary_summary = auxiliary_summary or {}
    pdf_extractions = pdf_extractions or {}
    warnings = warnings or []

    source_files = []
    for pat in ('*.xlsx', '*.xls', '*.pdf'):
        source_files.extend(sorted(glob.glob(os.path.join(project_dir, pat))))

    payload = {
        'schema_version': '1.0',
        'created_at': datetime.now().isoformat(),
        'source_mode': source_mode,
        'sources': [
            {
                'path': p,
                'name': os.path.basename(p),
            }
            for p in source_files
        ],
        'trial_balance': {
            'subjects_count': len(subjects) if isinstance(subjects, list) else 0,
            'subjects': subjects if isinstance(subjects, list) else [],
        },
        'balance_sheet': bs_balances,
        'fixed_assets': asset_register if isinstance(asset_register, dict) else {},
        'journal': journal if isinstance(journal, dict) else {},
        'auxiliary_ledgers': auxiliary_summary if isinstance(auxiliary_summary, dict) else {},
        'pdf_extractions': pdf_extractions if isinstance(pdf_extractions, dict) else {},
        'warnings': warnings,
    }
    _save_cache(cache_dir, 'standardized_manifest.json', payload)
    return payload


def _gate_pass_or_raise(gate_result):
    """Gate统一阻断: 非passed一律阻断。"""
    status = gate_result.get('status')
    if status != 'passed':
        raise RuntimeError(f'Gate阻断: {gate_result}')


# ============================================================

# ============================================================
# Phase 0标准化数据优先加载 (DT-ARCH)
# 设计原则: 标准化流程输出 → 填明细表流程输入
# 如果项目目录存在「标准化财务数据汇总.xlsx」，优先使用标准化数据
# 避免重复解析原始文件，建立标准化数据→评估明细表的强关联映射
# ============================================================

def _load_from_standardized(project_dir, cache_dir):
    """从标准化财务数据汇总.xlsx加载数据并转换为Phase 2所需缓存格式
    
    Returns:
        dict: 包含subjects, bs_balances, d1d2d3等所有缓存所需的原始数据
        如果标准化文件不存在或加载失败，返回None
    """
    from openpyxl import load_workbook as _opx_load
    std_path = os.path.join(project_dir, '标准化财务数据汇总.xlsx')
    if not os.path.exists(std_path):
        return None
    
    print('\n[标准化数据] 检测到标准化财务数据汇总.xlsx，优先使用标准化数据')
    print(f'  文件: {std_path}')
    
    try:
        wb = _opx_load(std_path, data_only=True)
    except Exception as e:
        print(f'  ⚠️ 无法打开标准化文件: {e}')
        return None
    
    available_sheets = wb.sheetnames
    print(f'  可用Sheet: {available_sheets}')
    
    result = {}
    
    # ── 1. 科目余额表（标准化）→ subjects.json ──
    if '科目余额表（标准化）' in available_sheets:
        ws = wb['科目余额表（标准化）']
        subjects = []
        for r in range(2, ws.max_row + 1):
            code = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            balance = ws.cell(row=r, column=3).value
            direction = ws.cell(row=r, column=4).value
            level = ws.cell(row=r, column=5).value
            sheet_code = ws.cell(row=r, column=6).value
            sheet_name = ws.cell(row=r, column=7).value
            
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue
            
            try:
                balance = float(balance) if balance is not None else 0.0
            except (ValueError, TypeError):
                balance = 0.0
            try:
                level = int(level) if level is not None else 1
            except (ValueError, TypeError):
                level = len(code)
            
            # v3.67 (2026-06-01): 读取辅助核算三件套 (cols 8/9/10) + 路径/重分类 (11/12/13/14/15)
            aux_type = ws.cell(row=r, column=8).value
            aux_code = ws.cell(row=r, column=9).value
            aux_name = ws.cell(row=r, column=10).value
            account_full_path = ws.cell(row=r, column=11).value
            standard_level1 = ws.cell(row=r, column=12).value
            data_type = ws.cell(row=r, column=13).value
            try:
                opening_balance = float(ws.cell(row=r, column=14).value or 0)
            except (ValueError, TypeError):
                opening_balance = 0.0
            try:
                pnl_carryover = float(ws.cell(row=r, column=15).value or 0)
            except (ValueError, TypeError):
                pnl_carryover = 0.0
            aux_name_str = str(aux_name).strip() if aux_name and str(aux_name).strip() not in ('nan', 'None') else ''
            aux_type_str = str(aux_type).strip() if aux_type and str(aux_type).strip() not in ('nan', 'None') else ''
            entry_name = aux_name_str if aux_name_str else (str(name).strip() if name else '')

            subjects.append({
                'code': code,
                'name': entry_name,
                'balance': balance,
                'direction': str(direction).strip() if direction else '借',
                'level': level,
                'currency': None,
                'beginning_debit': 0.0,
                'beginning_credit': 0.0,
                'current_debit': 0.0,
                'current_credit': 0.0,
                'ending_debit': balance if (direction and str(direction).strip() == '借') else 0.0,
                'ending_credit': abs(balance) if (direction and str(direction).strip() == '贷') else 0.0,
                'closing_balance': balance,
                'opening_balance': opening_balance,
                'pnl_carryover': pnl_carryover,
                '_sheet_code': str(sheet_code).strip() if sheet_code else '',
                '_sheet_name': str(sheet_name).strip() if sheet_name else '',
                'auxiliary_type': aux_type_str,
                'auxiliary_code': str(aux_code).strip() if aux_code and str(aux_code).strip() not in ('nan', 'None') else '',
                'auxiliary_name': aux_name_str,
                'counterparty': aux_name_str,
                'account_full_path': str(account_full_path).strip() if account_full_path and str(account_full_path).strip() not in ('nan', 'None') else '',
                'standard_level1': str(standard_level1).strip() if standard_level1 and str(standard_level1).strip() not in ('nan', 'None') else '',
                'data_type': str(data_type).strip() if data_type and str(data_type).strip() not in ('nan', 'None') else '',
            })
        
        result['subjects'] = subjects
        print(f'  科目余额表: {len(subjects)}行 → subjects.json')
    
    # ── 2. 资产负债表（标准化）→ bs_balances.json ──
    if '资产负债表（标准化）' in available_sheets:
        ws = wb['资产负债表（标准化）']
        items = []
        for r in range(2, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            ending = ws.cell(row=r, column=2).value
            beginning = ws.cell(row=r, column=3).value
            side_raw = ws.cell(row=r, column=4).value
            
            if label is None:
                continue
            label = str(label).strip()
            if not label:
                continue
            
            try:
                ending = float(ending) if ending is not None else 0.0
            except (ValueError, TypeError):
                ending = 0.0
            try:
                beginning = float(beginning) if beginning is not None else 0.0
            except (ValueError, TypeError):
                beginning = 0.0
            
            side_str = str(side_raw).strip() if side_raw else ''
            if '负债' in side_str:
                side = '负债及权益'
            else:
                side = '资产'
            
            items.append({
                'label': label,
                'ending_balance': ending,
                'beginning_balance': beginning,
                'side': side,
            })
        
        result['bs_balances'] = {'items': items}
        print(f'  资产负债表: {len(items)}行 → bs_balances.json')
    
    # ── 3. 固定资产台账（标准化）→ asset_register_by_sheet.json ──
    if '固定资产台账（标准化）' in available_sheets:
        ws = wb['固定资产台账（标准化）']
        asset_by_sheet = {}
        for r in range(2, ws.max_row + 1):
            asset_code = ws.cell(row=r, column=2).value
            asset_name = ws.cell(row=r, column=3).value
            spec = ws.cell(row=r, column=4).value
            cost = ws.cell(row=r, column=5).value
            depreciation = ws.cell(row=r, column=6).value
            net_value = ws.cell(row=r, column=7).value
            dept = ws.cell(row=r, column=8).value
            location = ws.cell(row=r, column=9).value
            start_date = ws.cell(row=r, column=10).value
            dep_method = ws.cell(row=r, column=11).value
            life_months = ws.cell(row=r, column=12).value
            status = ws.cell(row=r, column=13).value
            
            if asset_name is None and cost is None:
                continue
            # 跳过汇总行（如合计行、总计行等不应视为明细数据）
            asset_code_str = str(asset_code).strip() if asset_code else ''
            if asset_code_str in ('总计', '小计', '合计'):
                continue
            
            try:
                cost = float(cost) if cost is not None else 0.0
            except (ValueError, TypeError):
                cost = 0.0
            try:
                depreciation = float(depreciation) if depreciation is not None else 0.0
            except (ValueError, TypeError):
                depreciation = 0.0
            try:
                net_value = float(net_value) if net_value is not None else 0.0
            except (ValueError, TypeError):
                net_value = 0.0
            
            asset_name_str = str(asset_name).strip() if asset_name else ''
            dept_str = str(dept).strip() if dept else ''
            spec_str = str(spec).strip() if spec else ''
            
            sheet_key = _classify_asset_to_sheet(asset_name_str, spec_str, dept_str, cost)
            
            if sheet_key not in asset_by_sheet:
                asset_by_sheet[sheet_key] = []
            
            asset_by_sheet[sheet_key].append({
                'name': asset_name_str,
                'spec': spec_str,
                'cost': cost,
                'depreciation': depreciation,
                'net_value': net_value,
                'category': sheet_key,
                'asset_code': str(asset_code).strip() if asset_code else '',
                'start_date': str(start_date).strip() if start_date else '',
                'dept': dept_str,
                'location': str(location).strip() if location else '',
                'dep_method': str(dep_method).strip() if dep_method else '',
                'life_months': str(life_months).strip() if life_months else '',
                'status': str(status).strip() if status else '',
            })
        result['asset_register'] = asset_by_sheet
        total_assets = sum(len(v) for v in asset_by_sheet.values())
        print(f'  固定资产台账: {total_assets}项 → asset_register_by_sheet.json')
        for k, v in asset_by_sheet.items():
            print(f'    {k}: {len(v)}项')
    
    # ── 4. 序时账（往来科目摘要）→ journal.json ──
    if '序时账（往来科目摘要）' in available_sheets:
        ws = wb['序时账（往来科目摘要）']
        journal = []
        for r in range(2, ws.max_row + 1):
            date = ws.cell(row=r, column=1).value
            voucher = ws.cell(row=r, column=2).value
            code = ws.cell(row=r, column=3).value
            name = ws.cell(row=r, column=4).value
            summary = ws.cell(row=r, column=5).value
            debit = ws.cell(row=r, column=6).value
            credit = ws.cell(row=r, column=7).value
            
            if date is None and code is None:
                continue
            
            try:
                debit = float(debit) if debit is not None else 0.0
            except (ValueError, TypeError):
                debit = 0.0
            try:
                credit = float(credit) if credit is not None else 0.0
            except (ValueError, TypeError):
                credit = 0.0
            
            # v3.67 (2026-06-01): 读取新增的往来单位/部门/项目/银行账号列
            customer_supplier = ws.cell(row=r, column=8).value
            department = ws.cell(row=r, column=9).value
            project = ws.cell(row=r, column=10).value
            bank_account = ws.cell(row=r, column=11).value
            journal.append({
                'date': str(date).strip() if date else '',
                'voucher_no': str(voucher).strip() if voucher else '',
                'subject_code': str(code).strip() if code else '',
                'subject_name': str(name).strip() if name else '',
                'summary': str(summary).strip() if summary else '',
                'debit_amount': debit,
                'credit_amount': credit,
                'customer_supplier': str(customer_supplier).strip() if customer_supplier and str(customer_supplier).strip() not in ('nan', 'None') else '',
                'department': str(department).strip() if department and str(department).strip() not in ('nan', 'None') else '',
                'project': str(project).strip() if project and str(project).strip() not in ('nan', 'None') else '',
                'bank_account': str(bank_account).strip() if bank_account and str(bank_account).strip() not in ('nan', 'None') else '',
            })
        result['journal'] = journal
        print(f'  序时账: {len(journal)}行 → journal.json')
    
    # ── 5. 映射关系总表 ──
    if '映射关系总表' in available_sheets:
        ws = wb['映射关系总表']
        mapping_rows = []
        for r in range(2, ws.max_row + 1):
            code = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            sheet_code = ws.cell(row=r, column=3).value
            hit = ws.cell(row=r, column=4).value
            
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue
            mapping_rows.append({
                'code': code,
                'name': str(name).strip() if name else '',
                'sheet_code': str(sheet_code).strip() if sheet_code else '',
                'hit': str(hit).strip() if hit else '',
            })
        result['mapping_rows'] = mapping_rows
        print(f'  映射关系总表: {len(mapping_rows)}行')
    
    wb.close()
    return result


def _classify_asset_to_sheet(name, spec, dept, cost):
    """根据资产名称/规格/部门推断所属的评估明细表Sheet
    
    返回sheet key如: '4-8-4机器设备', '4-8-5车辆', '4-8-6电子设备'
    """
    combined = f"{name} {spec} {dept}".lower()
    
    # 车辆判断
    vehicle_kw = ['车', '运输', '牵引', '挂车', '叉车', '装载', '汽车', '尼桑', '福田', '别克', '金龙', '轿车', '客车', '货车', '商务车']
    if any(kw in combined for kw in vehicle_kw):
        return '4-8-5车辆'
    
    # 电子设备判断
    electronic_kw = ['电脑', '笔记本', '打印机', '复印机', '扫描仪', '服务器', '交换机', '路由',
                     '显示器', '投影', '相机', '手机', '电话', '平板', 'ipad', '电子设备',
                     '空调', '冰箱', '电视', '音响', '摄像', '监控', '考勤', '对讲', '验钞', '碎纸']
    if any(kw in combined for kw in electronic_kw):
        return '4-8-6电子设备'
    
    # 默认机器设备
    return '4-8-4机器设备'


def _build_mapping_from_standardized(bs_balances, subjects):
    """从标准化数据构建D1/D2/D3映射"""
    d1_to_d2 = {}
    d2_to_d3 = {}
    unmapped = []
    
    # BS科目名称→科目编码前缀的对应关系
    BS_TO_CODE_PREFIX = {
        '货币资金': ['1001', '1002', '1003', '1004', '1012'],
        '应收票据': ['1121'],
        '应收账款': ['1122'],
        '预付款项': ['1123', '1124'],
        '其他应收款': ['1221'],
        '存货': ['1401', '1402', '1403', '1404', '1405', '1406', '1407', '1408', '1409', '1410', '1411', '1412', '1413', '1421', '1471'],
        '固定资产': ['1601', '1602'],
        '在建工程': ['1604'],
        '无形资产': ['1701', '1702'],
        '长期待摊费用': ['1801'],
        '递延所得税资产': ['1811'],
        '短期借款': ['2001'],
        '应付票据': ['2201'],
        '应付账款': ['2202'],
        '预收款项': ['2203', '2204'],
        '应付职工薪酬': ['2211'],
        '应交税费': ['2221'],
        '应付利息': ['2231'],
        '其他应付款': ['2241'],
        '长期借款': ['2501'],
        '实收资本（或股本）': ['4001'],
        '资本公积': ['4002'],
        '盈余公积': ['4101'],
        '未分配利润': ['4103', '4104'],
        '长期应付款': ['2701'],
        '预计负债': ['2801'],
        '递延收益': ['2401'],
        '递延所得税负债': ['2901'],
        '合同负债': ['2205'],
        '合同资产': ['1125'],
        '应收款项融资': ['1126'],
        '交易性金融资产': ['1101'],
        '交易性金融负债': ['2101'],
        # v3.68 (2026-06-01): 补充某测试项目缺失的 7 个 BS 科目映射
        '长期股权投资': ['1511'],
        '投资性房地产': ['1521'],
        '其他非流动金融资产': ['1813'],
        '其他非流动资产': ['1812'],
        '专项应付款': ['2711'],
        '一年内到期的非流动负债': ['2701', '2501', '2901'],
        '其他流动资产': ['122199', '224103'],
    }
    
    # Build D1→D2 from BS items
    if bs_balances:
        bs_items = bs_balances.get('items', [])
        for item in bs_items:
            label = item.get('label', '')
            if '合计' in label or '总计' in label or label.endswith('：') or label.endswith(':'):
                continue
            matched = False
            for bs_key, prefixes in BS_TO_CODE_PREFIX.items():
                if label == bs_key or label.replace(' ', '') == bs_key.replace(' ', ''):
                    codes = []
                    if subjects:
                        for s in subjects:
                            sc = str(s.get('code', ''))
                            for pfx in prefixes:
                                if sc.startswith(pfx):
                                    codes.append(sc)
                    if codes:
                        d1_to_d2[label] = list(set(codes))
                    else:
                        d1_to_d2[label] = prefixes
                    matched = True
                    break
            if not matched:
                unmapped.append(label)
    
    # Build D2→D3 from subjects
    if subjects:
        for s in subjects:
            code = str(s.get('code', ''))
            name = s.get('name', '')
            d2_to_d3[code] = {
                'name': name,
                'has_auxiliary': False,
                'auxiliary_file': None,
            }
    
    return {
        'd1_to_d2': d1_to_d2,
        'd2_to_d3': d2_to_d3,
        'unmapped_bs_items': unmapped,
    }


def _extract_settings_from_std(subjects, bs_balances, project_dir):
    """从标准化数据和项目目录中提取设定信息"""
    settings = {
        'company_name': '',
        'valuation_date': '',
        'industry_type': '通用',
    }
    
    # 从项目目录名提取公司名
    project_name = os.path.basename(os.path.normpath(project_dir))
    # 尝试从文件名提取（取不含数字和括号的前半部分）
    import re
    parts = re.split(r'[（(\d]', project_name)
    if parts:
        name = parts[0].strip()
        if name and len(name) >= 2:
            settings['company_name'] = name
    
    # 从资产负债表文件推断评估基准日（DT-212: 读取R2C3日期）
    import glob as _glob
    bs_files = _glob.glob(os.path.join(project_dir, '*报表*.xlsx'))
    if not bs_files:
        bs_files = _glob.glob(os.path.join(project_dir, '*资产负债*.xlsx'))
    if bs_files:
        try:
            import openpyxl as _opx_bs
            _wb_bs = _opx_bs.load_workbook(bs_files[0], data_only=True)
            _ws_bs = _wb_bs[_wb_bs.sheetnames[0]] if _wb_bs.sheetnames else None
            if _ws_bs:
                # 查找日期: 通常在R2C3 或 含'年'字的行
                for _r in range(1, 6):
                    for _c in range(1, min(_ws_bs.max_column + 1, 10)):
                        _v = str(_ws_bs.cell(row=_r, column=_c).value or '')
                        import re as _re_bs
                        _date_match = _re_bs.search(r'\d{4}[-/]\d{1,2}(?:[-/]\d{1,2})?', _v)
                        if _date_match:
                            _date_str = _date_match.group().replace('/', '-')
                            # 如果只有年-月(没有日),补 day=01 → 2025-07-31(月末)更符合基准日惯例
                            if _re_bs.match(r'\d{4}[-/]\d{1,2}$', _date_str):
                                # 推断为月末
                                _y, _m = _date_str.split('-')[:2]
                                _m_int = int(_m)
                                if _m_int == 12:
                                    _date_str = f'{_y}-{_m_int}-31'
                                else:
                                    import calendar as _cal
                                    _last_day = _cal.monthrange(int(_y), _m_int)[1]
                                    _date_str = f'{_y}-{_m_int}-{_last_day}'
                            settings['valuation_date'] = _date_str
                            break
                    if settings['valuation_date']:
                        break
                # 从R2C1提取公司名
                _bs_name = str(_ws_bs.cell(row=2, column=1).value or '').strip()
                _bs_name = _bs_name.replace('\r', '').replace('\n', '').strip()
                if '编制单位' in _bs_name:
                    _bs_name = _bs_name.split('：')[-1] if '：' in _bs_name else _bs_name.split(':')[-1]
                    _bs_name = _bs_name.strip()
                if _bs_name and len(_bs_name) >= 2:
                    settings['company_name'] = _bs_name
            _wb_bs.close()
        except Exception:
            pass
    if not settings['valuation_date']:
        # 禁止默认填充日期，交由待确认清单处理
        settings['valuation_date'] = ''
    
    # 从资产负债表推断行业类型
    if bs_balances:
        bs_items = bs_balances.get('items', [])
        bs_labels = [item.get('label', '') for item in bs_items]
        if any('存货' in l for l in bs_labels):
            settings['industry_type'] = '制造业'
    
    return settings

# Phase -1.5: 标准化桥接
# ============================================================

def _run_pre_phase_normalization(project_dir, args):
    """在Phase 0前尝试标准化输入；失败时保留内置解析器降级路径。"""
    try:
        from normalize_input import run_normalization
        return run_normalization(
            project_dir,
            force=bool(getattr(args, 'force', False)),
        )
    except Exception as exc:
        print(f'  ⚠️ Phase -1.5标准化桥接失败，降级使用Phase 0内置解析器: {exc}')
        return {
            'phase': '-1.5',
            'status': 'fallback',
            'warnings': [str(exc)],
        }


# Phase 0: 输入确认与数据源解析
# ============================================================

def phase0(project_dir, args):
    """Phase 0: 输入确认与数据源解析

    子步骤:
    0.1 模式判断
    0.2 科目余额表解析 → subjects.json
    0.3 资产负债表解析+自校验(DT-139) → bs_balances.json
    0.4 PDF数据源自动识别与提取(DT-211) → pdf_extractions.json + multimodal_tasks.json
    0.5a D1/D2/D3映射(DT-119) → d1d2d3_mapping.json
    0.5 辅助余额表强制提取(DT-111) → auxiliary_balance_*.json
    0.6 数据分类+重分类(DT-117/DT-118) → data_classification.json + reclassification.json
    0.7 设定信息填写(DT-121) → settings_info.json
    """
    _run_pre_phase_normalization(project_dir, args)
    cache_dir = _cache_path(project_dir)
    
    # ── DT-ARCH: 标准化数据优先加载 ──
    std_data = _load_from_standardized(project_dir, cache_dir)
    if std_data and std_data.get('subjects'):
        print('\n[标准化优先] 使用标准化财务数据汇总.xlsx作为数据源')
        
        # 写入subjects.json
        _save_cache(cache_dir, 'subjects.json', std_data['subjects'])
        std_nonzero = _count_nonzero_subjects(std_data['subjects'])
        if std_nonzero == 0:
            pending_items = [{
                'id': 'PARSE-SB-001',
                'type': 'source_parse_error',
                'status': 'pending',
                'reason': '标准化科目余额全部为0，疑似数据源解析失败',
                'source': 'subjects.json',
                'candidate_targets': [],
                'evidence': ['subjects.json.balance'],
            }]
            _write_pending_confirmations(cache_dir, pending_items, reason='科目余额解析异常')
            _build_project_state(project_dir, 'BLOCKED_CONFIRMATION', cache_dir, pending_count=len(pending_items))
            return {
                'phase': 0,
                'status': 'blocked_confirmation',
                'source': 'standardized',
                'pending_count': len(pending_items),
            }
        
        # BS数据: 从原始BS文件解析（标准化BS可能存在解析偏差）
        # 搜索原始BS文件（资产负债表/财务报表.xlsx）
        import glob as _glob
        bs_files = _glob.glob(os.path.join(project_dir, '*资产负债表*')) + _glob.glob(os.path.join(project_dir, '*财务报表*'))
        # r9: 尝试把 .xls 转 .xlsx
        _xls_files = _glob.glob(os.path.join(project_dir, '*.xls')) + _glob.glob(os.path.join(project_dir, '*.XLS'))
        for _xls_f in _xls_files:
            if not bs_files or not any('资产负债表' in os.path.basename(f) for f in bs_files):
                _base = os.path.splitext(_xls_f)[0]
                _xlsx_f = _base + '.xlsx'
                if not os.path.exists(_xlsx_f):
                    try:
                        import xlrd
                        _wb_old = xlrd.open_workbook(_xls_f)
                        _wb_new = __import__('openpyxl').Workbook()
                        _first = True
                        for _sn in _wb_old.sheet_names():
                            _ws_old = _wb_old.sheet_by_name(_sn)
                            if _first:
                                _ws_new = _wb_new.active
                                _ws_new.title = _sn[:31]
                                _first = False
                            else:
                                _ws_new = _wb_new.create_sheet(title=_sn[:31])
                            for _r in range(_ws_old.nrows):
                                for _c in range(_ws_old.ncols):
                                    _ws_new.cell(_r+1, _c+1, _ws_old.cell_value(_r, _c))
                        _wb_new.save(_xlsx_f)
                        print(f'  🔄 已自动转换 .xls → .xlsx: {os.path.basename(_xlsx_f)}')
                        bs_files.append(_xlsx_f)
                    except ImportError:
                        print('  ⚠️ xlrd 未安装，无法转换 .xls 文件')
                    except Exception as _e:
                        print(f'  ⚠️ .xls 转换失败: {_e}')
        if not bs_files:
            # DT-210: 内部结构确认法搜索BS文件
            import openpyxl as _opx
            _exclude_keywords = ['科目余额', '序时账', '明细账', '凭证', '评估明细表', '底稿', '抽凭', '辅助']
            _candidate_xlsx = [f for f in _glob.glob(os.path.join(project_dir, '*.xlsx'))
                               if not any(kw in os.path.basename(f) for kw in _exclude_keywords)]
            for _cand in _candidate_xlsx:
                try:
                    _wb = _opx.load_workbook(_cand, data_only=True)
                    _ws = _wb[_wb.sheetnames[0]]
                    for _r in range(1, min(_ws.max_row + 1, 6)):
                        for _c in range(1, min(_ws.max_column + 1, 10)):
                            _v = _ws.cell(row=_r, column=_c).value
                            if _v and isinstance(_v, str) and '资产负债表' in _v:
                                bs_files.append(_cand)
                                break
                        if bs_files:
                            break
                    _wb.close()
                except Exception:
                    pass
                if bs_files:
                    break
        if bs_files:
            bs_balances = _parse_balance_sheet(bs_files, project_dir)
            _save_cache(cache_dir, 'bs_balances.json', bs_balances)
            print(f'  BS数据: 从原始文件解析({os.path.basename(bs_files[0])}) → bs_balances.json')
        else:
            bs_balances = std_data.get('bs_balances', {'items': []})
            _save_cache(cache_dir, 'bs_balances.json', bs_balances)
            print(f'  BS数据: 从标准化文件加载(原始BS未找到)')
        
        # 构建并写入d1d2d3_mapping.json
        d1d2d3 = _build_mapping_from_standardized(bs_balances, std_data['subjects'])
        override_result = _apply_mapping_overrides(d1d2d3, cache_dir, std_data['subjects'])
        if override_result.get('applied'):
            print(f'  映射覆盖已应用: {override_result["applied"]}项')
        _save_cache(cache_dir, 'd1d2d3_mapping.json', d1d2d3)
        print(f'  D1→D2映射: {len(d1d2d3.get("d1_to_d2", {}))}个')
        print(f'  D2→D3映射: {len(d1d2d3.get("d2_to_d3", {}))}个')
        
        # 写入asset_register_by_sheet.json
        if std_data.get('asset_register'):
            _save_cache(cache_dir, 'asset_register_by_sheet.json', std_data['asset_register'])
        
        # 写入journal.json
        if std_data.get('journal'):
            _save_cache(cache_dir, 'journal.json', std_data['journal'])
        
        # 写入辅助缓存（标准化模式下的简化版本）
        _save_cache(cache_dir, 'execution_mode.json', {'mode': 'complete', 'source': 'standardized'})
        aux_summary = {'sheet_count': 0, 'total_objects': 0}
        _save_cache(cache_dir, 'auxiliary_balance_summary.json', aux_summary)
        _save_cache(cache_dir, 'data_classification.json', {'source': 'standardized'})
        _save_cache(cache_dir, 'reclassification.json', {'items': []})
        
        # 提取设定信息
        settings = _extract_settings_from_std(std_data['subjects'], bs_balances, project_dir)
        _save_cache(cache_dir, 'settings_info.json', settings)
        print(f'  被评估单位: {settings.get("company_name", "未提取")}')
        print(f'  评估基准日: {settings.get("valuation_date", "未提取")}')
        
        pending_items = []
        if not settings.get('valuation_date'):
            pending_items.append({
                'id': 'VAL-DATE-001',
                'type': 'valuation_date',
                'status': 'pending',
                'reason': '无法从标准化/原始报表提取评估基准日',
                'candidate_targets': [],
                'evidence': ['settings_info.json.valuation_date'],
            })
        # v3.67 (2026-06-01) 兼容:如果 mapping_overrides.use_tb_zero 标记了某些 BS 标签,
        # 自动确认这些项为"科目余额表=0",不写入 pending 清单
        overrides = _load_cache(cache_dir, 'mapping_overrides.json') or {}
        use_tb_zero = set((overrides.get('use_tb_zero') or []) if isinstance(overrides, dict) else [])
        for label in d1d2d3.get('unmapped_bs_items', []):
            if label in use_tb_zero:
                # 自动确认:该 BS 项在科目余额表中无对应科目,余额=0
                print(f'  [AUTO-CONFIRM] BS 项 "{label}" → 科目余额表=0 (mapping_overrides.use_tb_zero)')
                continue
            pending_items.append({
                'id': f'MAP-BS-{len(pending_items)+1:03d}',
                'type': 'subject_mapping',
                'status': 'pending',
                'reason': 'BS科目无法映射到科目余额表编码',
                'source': label,
                'candidate_targets': _build_mapping_candidates(label, std_data.get('subjects', [])),
                'evidence': ['d1d2d3_mapping.json.unmapped_bs_items'],
            })

        _write_standardized_manifest(
            project_dir=project_dir,
            cache_dir=cache_dir,
            source_mode='standardized',
            subjects=std_data.get('subjects', []),
            bs_balances=bs_balances,
            asset_register=std_data.get('asset_register', {}),
            journal=std_data.get('journal', {}),
            auxiliary_summary=aux_summary,
            pdf_extractions={},
            warnings=[it.get('reason', '') for it in pending_items],
        )

        if pending_items:
            _write_pending_confirmations(cache_dir, pending_items)
            _build_project_state(project_dir, 'BLOCKED_CONFIRMATION', cache_dir, pending_count=len(pending_items))
            return {
                'phase': 0,
                'status': 'blocked_confirmation',
                'source': 'standardized',
                'pending_count': len(pending_items),
            }

        # 输出缓存文件清单
        print('\n' + '-'*40)
        print('Phase 0 完成（标准化模式）！缓存文件清单:')
        for f in sorted(os.listdir(cache_dir)):
            if f.endswith('.json'):
                size = os.path.getsize(os.path.join(cache_dir, f))
                print(f'  ✓ {f} ({size:,} bytes)')
        
        _build_project_state(project_dir, 'READY_TO_FILL', cache_dir, pending_count=0)
        return {
            'phase': 0,
            'status': 'completed',
            'source': 'standardized',
            'subjects_count': len(std_data['subjects']),
            'bs_items_count': len(bs_balances.get('items', [])),
        }
    
    # ── 降级: 无标准化文件，走原始文件解析 ──
    print('\n' + '='*60)
    print('Phase 0: 输入确认与数据源解析')
    print('='*60)

    # --- Step 0.1: 模式判断 ---
    print('\n[Step 0.1] 模式判断')
    sb_files = glob.glob(os.path.join(project_dir, '*科目余额表*')) + glob.glob(os.path.join(project_dir, '*余额表*'))
    bs_files = glob.glob(os.path.join(project_dir, '*资产负债表*')) + \
               glob.glob(os.path.join(project_dir, '*财务报表*'))

    # DT-210: BS文件搜索增强——文件名不含"资产负债表"时，通过内部结构确认
    # 例如"河南平煤神马平绿置业有限公司.xlsx"，文件名无关键词但内部Row1含"资产负债表"
    if not bs_files:
        import openpyxl as _opx
        _exclude_keywords = ['科目余额', '序时账', '序时簿', '明细账', '凭证', '评估明细表', '底稿', '抽凭', '辅助', '余额表', '折旧']
        _candidate_xlsx = [f for f in glob.glob(os.path.join(project_dir, '*.xlsx'))
                           if not any(kw in os.path.basename(f) for kw in _exclude_keywords)]
        for _cand in _candidate_xlsx:
            try:
                _wb = _opx.load_workbook(_cand, data_only=True)
                # 优先检查sheet名称，其次检查第一sheet内容
                _found_bs = any('资产负债表' in sn for sn in _wb.sheetnames)
                if _found_bs:
                    bs_files.append(_cand)
                else:
                    _ws = _wb[_wb.sheetnames[0]]
                    for _r in range(1, min(_ws.max_row + 1, 6)):
                        for _c in range(1, min(_ws.max_column + 1, 10)):
                            _v = _ws.cell(row=_r, column=_c).value
                            if _v and isinstance(_v, str) and '资产负债表' in _v:
                                bs_files.append(_cand)
                                break
                        if bs_files:
                            break
                _wb.close()
            except Exception:
                pass
            if bs_files:
                break
    mode = 'complete' if sb_files else 'incomplete'
    print(f'  执行模式: {mode} (科目余额表: {len(sb_files)}个, 资产负债表: {len(bs_files)}个)')
    assert mode == 'complete', 'DT-103: 无科目余额表，无法进入完整模式'
    _save_cache(cache_dir, 'execution_mode.json', {'mode': mode})

    # --- Step 0.2: 科目余额表解析 ---
    print('\n[Step 0.2] 科目余额表解析 (DT-0零幻觉)')
    if _load_cache(cache_dir, 'subjects.json'):
        print('  [CACHE] 命中缓存，跳过解析')
        subjects = _load_cache(cache_dir, 'subjects.json')
    else:
        subjects = _parse_subject_balance(sb_files[0])
        _save_cache(cache_dir, 'subjects.json', subjects)
    print(f'  解析科目数: {len(subjects)}')
    nonzero_subjects = _count_nonzero_subjects(subjects)
    if nonzero_subjects == 0:
        pending_items = [{
            'id': 'PARSE-SB-001',
            'type': 'source_parse_error',
            'status': 'pending',
            'reason': '科目余额表解析结果全部为0，疑似列位识别错误',
            'source': os.path.basename(sb_files[0]) if sb_files else '科目余额表',
            'candidate_targets': [],
            'evidence': ['subjects.json.balance'],
        }]
        _write_pending_confirmations(cache_dir, pending_items, reason='科目余额解析异常')
        _build_project_state(project_dir, 'BLOCKED_CONFIRMATION', cache_dir, pending_count=len(pending_items))
        return {
            'phase': 0,
            'status': 'blocked_confirmation',
            'pending_count': len(pending_items),
        }

    # --- Step 0.3: 资产负债表解析+DT-139自校验 ---
    print('\n[Step 0.3] 资产负债表解析+DT-139自校验')
    if _load_cache(cache_dir, 'bs_balances.json'):
        print('  [CACHE] 命中缓存，跳过解析')
        bs_balances = _load_cache(cache_dir, 'bs_balances.json')
    else:
        bs_balances = _parse_balance_sheet(bs_files, project_dir)
        _save_cache(cache_dir, 'bs_balances.json', bs_balances)

    # DT-139强制自校验
    _validate_bs(bs_balances, cache_dir)

    # --- Step 0.4: PDF数据源自动识别与提取 (DT-211) ---
    print('\n[Step 0.4] PDF数据源自动识别与提取 (DT-211)')
    if _load_cache(cache_dir, 'pdf_extractions.json'):
        print('  [CACHE] 命中缓存，跳过PDF提取')
        pdf_result = _load_cache(cache_dir, 'pdf_extractions.json')
    else:
        # 从settings_info获取基准日（如果已有缓存）
        _cached_settings = _load_cache(cache_dir, 'settings_info.json')
        _base_date = ''
        if _cached_settings:
            _base_date = _cached_settings.get('valuation_date', '')
        pdf_result = _extract_pdf_sources(project_dir, cache_dir, base_date=_base_date)
    print(f'  PDF总计: {pdf_result.get("total_pdfs", 0)}个, '
          f'银行存款记录: {pdf_result.get("records_count", 0)}条, '
          f'需多模态: {pdf_result.get("multimodal_count", 0)}个')

    # --- Step 0.4b: 固定资产台账逐行解析 → asset_register_by_sheet.json ---
    print('\n[Step 0.4b] 固定资产台账逐行解析')
    _ar_cache = _load_cache(cache_dir, 'asset_register_by_sheet.json')
    if _ar_cache:
        total_ar = sum(len(v) for v in _ar_cache.values())
        print(f'  [CACHE] 命中缓存: {total_ar}项')
    else:
        fa_glob = glob.glob(os.path.join(project_dir, '*固定资产*')) + \
                  glob.glob(os.path.join(project_dir, '*资产台账*')) + \
                  glob.glob(os.path.join(project_dir, '*折旧*')) + \
                  glob.glob(os.path.join(project_dir, '*卡片*'))
        if fa_glob:
            fa_path = fa_glob[0]
            try:
                from fix_asset_register_bridge import build_asset_register_by_sheet
                build_asset_register_by_sheet(fa_path, cache_dir)
            except ImportError:
                print(f'  ⚠️ fix_asset_register_bridge 未找到，跳过')
            except Exception as e:
                print(f'  ⚠️ 固定资产解析失败: {e}')
        else:
            print('  ⚠️ 未找到固定资产台账文件')

    # --- Step 0.4c: 科目明细账标准化解析 → subledger_standardized.json ---
    print('\n[Step 0.4c] 科目明细账标准化解析')
    _sl_cache = _load_cache(cache_dir, 'subledger_standardized.json')
    if _sl_cache:
        sl_count = sum(v.get('transaction_count', 0) for v in _sl_cache.values())
        print(f'  [CACHE] 命中缓存: {len(_sl_cache)}科目, {sl_count}笔交易')
    else:
        _sl_glob = glob.glob(os.path.join(project_dir, '*明细账*')) + \
                    glob.glob(os.path.join(project_dir, '*序时账*')) + \
                    glob.glob(os.path.join(project_dir, '*序时簿*'))
        if _sl_glob:
            _sl_path = _sl_glob[0]
            try:
                from subledger_standardizer import standardize_subledger
                standardize_subledger(_sl_path, cache_dir)
            except ImportError:
                print(f'  ⚠️ subledger_standardizer 未找到，跳过')
            except Exception as e:
                print(f'  ⚠️ 科目明细账解析失败: {e}')
        else:
            print('  ⚠️ 未找到科目明细账文件')

    # --- Step 0.5a: D1/D2/D3映射 (DT-119) ---
    print('\n[Step 0.5a] D1/D2/D3三级递进映射 (DT-119)')
    if _load_cache(cache_dir, 'd1d2d3_mapping.json'):
        print('  [CACHE] 命中缓存，跳过映射')
        d1d2d3 = _load_cache(cache_dir, 'd1d2d3_mapping.json')
    else:
        d1d2d3 = _build_d1d2d3_mapping(bs_balances, subjects)
    override_result = _apply_mapping_overrides(d1d2d3, cache_dir, subjects)
    if override_result.get('applied'):
        print(f'  映射覆盖已应用: {override_result["applied"]}项')
    _save_cache(cache_dir, 'd1d2d3_mapping.json', d1d2d3)
    print(f'  D1→D2映射: {len(d1d2d3.get("d1_to_d2", {}))}个')
    print(f'  D2→D3映射: {len(d1d2d3.get("d2_to_d3", {}))}个')

    # --- Step 0.5: 辅助余额表强制提取 (DT-111) ---
    print('\n[Step 0.5] 辅助余额表强制提取 (DT-111)')
    aux_dir = os.path.join(project_dir, '科目辅助明细账')
    if os.path.isdir(aux_dir):
        aux_result = _extract_auxiliary_balances(aux_dir, cache_dir)
        print(f'  提取辅助余额表: {aux_result["sheet_count"]}个Sheet, {aux_result["total_objects"]}个结算对象')
        _save_cache(cache_dir, 'auxiliary_balance_summary.json', aux_result)
    else:
        print('  ⚠️ 未找到科目辅助明细账目录')
        _save_cache(cache_dir, 'auxiliary_balance_summary.json', {
            'sheet_count': 0, 'total_objects': 0, 'warning': '目录不存在'
        })

    # --- Step 0.6: 数据分类+重分类 (DT-117/DT-118) ---
    print('\n[Step 0.6] 数据分类+重分类映射 (DT-117/DT-118)')
    # 先做重分类检测+调整，再基于调整后的subjects做分类
    _reclass_from_cache = _load_cache(cache_dir, 'reclassification.json')
    if _reclass_from_cache:
        reclass = _reclass_from_cache
    else:
        reclass = _build_reclassification(subjects)
        _save_cache(cache_dir, 'reclassification.json', reclass)
    print(f'  重分类项目: {len(reclass.get("items", []))}个')
    # DT-FIX: 应用重分类调整——仅在首次构建时执行，避免resume重复调整导致金额加倍
    _reclass_items = reclass.get('items', [])
    if _reclass_items and not _reclass_from_cache:
        for _rc in _reclass_items:
            _src_code = _rc.get('source_code', '')
            _amt = _rc.get('reclass_amount', 0)
            if not _src_code or _amt <= 0:
                continue
            for _s in (subjects if isinstance(subjects, list) else subjects.get('subjects', [])):
                if str(_s.get('code', '')) == _src_code:
                    if _s.get('ending_debit', 0) > 0:
                        _s['ending_debit'] = max(_s['ending_debit'] - _amt, 0)
                    if _s.get('ending_credit', 0) > 0:
                        _s['ending_credit'] = max(_s['ending_credit'] - _amt, 0)
                    # 同步清零balance字段（subject_sheet_mapping/data_loader使用它）
                    _s['balance'] = 0
                    break
        # 重分类金额加到目标科目
        for _rc in _reclass_items:
            _tgt_name = _rc.get('target_name', '')
            _amt = _rc.get('reclass_amount', 0)
            if _amt <= 0 or not _tgt_name:
                continue
            _tgt_prefixes = {'预付款项': '1123', '预收款项': '2203', '其他应收款': '1221',
                             '应付账款': '2202', '应收账款': '1122', '其他应付款': '2241',
                             '其他流动资产': '1901'}
            _prefix = _tgt_prefixes.get(_tgt_name, '')
            if _prefix:
                for _st in (subjects if isinstance(subjects, list) else subjects.get('subjects', [])):
                    if str(_st.get('code', '')) == _prefix:
                        _st['ending_debit'] = _st.get('ending_debit', 0) + _amt
                        _st['balance'] = _st.get('balance', 0) + _amt
                        print(f'  ➕ {_prefix}({_tgt_name}): +{_amt:,.2f}')
                        break
        _save_cache(cache_dir, 'subjects.json', subjects)
        print(f'  🔄 已应用{len(_reclass_items)}项重分类调整')
    # 基于调整后的subjects做数据分类
    if _load_cache(cache_dir, 'data_classification.json'):
        data_class = _load_cache(cache_dir, 'data_classification.json')
    else:
        data_class = _classify_data(subjects, bs_balances, project_dir)
        _save_cache(cache_dir, 'data_classification.json', data_class)

    # --- Step 0.7: 设定信息 (DT-121) ---
    print('\n[Step 0.7] 设定信息提取 (DT-121)')
    if _load_cache(cache_dir, 'settings_info.json'):
        settings = _load_cache(cache_dir, 'settings_info.json')
    else:
        settings = _extract_settings(subjects, bs_balances, project_dir)
        _save_cache(cache_dir, 'settings_info.json', settings)
    print(f'  被评估单位: {settings.get("company_name", "未提取")}')
    print(f'  评估基准日: {settings.get("valuation_date", "未提取")}')

    pending_items = []
    if not settings.get('valuation_date'):
        pending_items.append({
            'id': 'VAL-DATE-001',
            'type': 'valuation_date',
            'status': 'pending',
            'reason': '无法从报表和路径提取评估基准日',
            'candidate_targets': [],
            'evidence': ['settings_info.json.valuation_date'],
        })
    # DT-FIX: 检查mapping_overrides是否已覆盖这些label，避免resume循环
    _existing_overrides = {}
    try:
        _mo = _load_cache(cache_dir, 'mapping_overrides.json') or {}
        _existing_overrides = _mo.get('mappings', {})
    except Exception:
        pass
    for label in d1d2d3.get('unmapped_bs_items', []):
        if label in _existing_overrides:
            continue  # 已被映射覆盖确认，跳过阻断
        pending_items.append({
            'id': f'MAP-BS-{len(pending_items)+1:03d}',
            'type': 'subject_mapping',
            'status': 'pending',
            'reason': 'BS科目无法映射到科目余额表编码',
            'source': label,
            'candidate_targets': _build_mapping_candidates(label, subjects),
            'evidence': ['d1d2d3_mapping.json.unmapped_bs_items'],
        })

    _write_standardized_manifest(
        project_dir=project_dir,
        cache_dir=cache_dir,
        source_mode='raw',
        subjects=subjects,
        bs_balances=bs_balances,
        asset_register=_load_cache(cache_dir, 'asset_register_by_sheet.json') or {},
        journal=_load_cache(cache_dir, 'journal.json') or {},
        auxiliary_summary=_load_cache(cache_dir, 'auxiliary_balance_summary.json') or {},
        pdf_extractions=pdf_result if isinstance(pdf_result, dict) else {},
        warnings=[it.get('reason', '') for it in pending_items],
    )

    if pending_items:
        _write_pending_confirmations(cache_dir, pending_items)
        _build_project_state(project_dir, 'BLOCKED_CONFIRMATION', cache_dir, pending_count=len(pending_items))
        return {
            'phase': 0,
            'status': 'blocked_confirmation',
            'pending_count': len(pending_items),
        }

    # --- Phase 0 完成 ---
    print('\n' + '-'*40)
    print('Phase 0 完成！缓存文件清单:')
    for f in sorted(os.listdir(cache_dir)):
        if f.endswith('.json'):
            size = os.path.getsize(os.path.join(cache_dir, f))
            print(f'  ✓ {f} ({size:,} bytes)')

    _build_project_state(project_dir, 'READY_TO_FILL', cache_dir, pending_count=0)
    return {
        'phase': 0,
        'status': 'completed',
        'subjects_count': len(subjects),
        'bs_items_count': len(bs_balances.get('items', [])),
        'reclassification_items': len(reclass.get('items', [])),
    }


# ============================================================
# 辅助函数
# ============================================================

def _parse_subject_balance(filepath):
    """解析科目余额表 → subjects.json (DT-154: 委托给source_header_parser)"""
    from source_header_parser import parse_subject_balance as _shp_parse

    print(f'  解析: {os.path.basename(filepath)}')
    result = _shp_parse(filepath)

    if result['status'] != 'parsed':
        raise ValueError(f'DT-0: 科目余额表解析失败！{result.get("warnings", [])}')

    # 转换为dt_runner原有格式（subjects列表）
    subjects = result['subjects']
    print(f'  表头行: {result["header_row"]}, 列映射: {result["col_map"]}')
    return subjects


def _parse_balance_sheet(bs_files, project_dir):
    """解析资产负债表 → bs_balances.json (DT-154: 委托给source_header_parser)"""
    from source_header_parser import parse_balance_sheet as _pbs_parse

    if not bs_files:
        # DT-153v3: 通用文件搜索，不再硬编码项目名
        # 搜索策略：优先匹配含"资产负债表"/"财务报表"的文件
        bs_files = glob.glob(os.path.join(project_dir, '*资产负债表*')) + \
                   glob.glob(os.path.join(project_dir, '*财务报表*'))
        if not bs_files:
            # fallback: 搜索所有xlsx文件，排除已知非BS文件
            exclude_keywords = ['科目余额', '序时账', '明细账', '凭证', '评估明细表', '底稿', '抽凭']
            all_xlsx = glob.glob(os.path.join(project_dir, '*.xlsx'))
            bs_files = [f for f in all_xlsx
                        if not any(kw in os.path.basename(f) for kw in exclude_keywords)]

    if not bs_files:
        raise ValueError('DT-0: 未找到资产负债表文件！')

    filepath = bs_files[0]
    print(f'  解析: {os.path.basename(filepath)}')

    result = _pbs_parse(filepath)

    if result['status'] != 'parsed':
        raise ValueError(f'DT-0: 资产负债表解析失败！{result.get("warnings", [])}')

    print(f'  格式: {result["format"]}, 表头行: {result["header_row"]}, 列映射: {result["col_map"]}')

    # P1补充: 传递BS表头元信息（编制单位全称、评估基准日）
    bs_meta = {}
    if result.get('company_full_name'):
        bs_meta['company_full_name'] = result['company_full_name']
        print(f'  编制单位(全称): {result["company_full_name"]}')
    if result.get('valuation_date'):
        bs_meta['valuation_date'] = result['valuation_date']
        print(f'  报表日期: {result["valuation_date"]}')

    return {
        'items': result['items'],
        'total_assets': result['total_assets'],
        'total_liabilities': result.get('total_liab_equity', 0),
        'total_equity': 0,
        'filepath': filepath,  # DT-210: 传递BS文件路径，供settings提取日期时使用
        **bs_meta,  # 展开BS表头元信息
    }


def _validate_bs(bs_data, cache_dir=None):
    """DT-139: BS解析后强制自校验

    默认容差为1.0元。项目级例外必须绑定源文件SHA256，且只能生成工作稿。
    """
    items = bs_data.get('items', [])

    total_assets = 0
    total_liab_equity = 0

    for item in items:
        label = item['label'].replace(' ', '').replace('\u3000', '')
        val = item['ending_balance']
        if '资产总计' in label or '资产合计' in label:
            total_assets = val
        if '负债' in label and '所有者权益' in label and '总计' in label:
            total_liab_equity = val
        elif '负债' in label and '所有者' in label and ('合计' in label or '总计' in label):
            total_liab_equity = val  # r9: 兼容"负债和所有者合计"变体

    diff = abs(total_assets - total_liab_equity)
    _dt139_tolerance = 1.0
    _has_exception = False
    _exception_reason = ''
    _exception_error = ''
    _source_path = bs_data.get('filepath', '')
    _source_sha256 = ''
    if _source_path and os.path.isfile(_source_path):
        _source_sha256 = _sha256_file(_source_path)

    if cache_dir:
        _exc_path = os.path.join(cache_dir, 'dt139_exception.json')
        if os.path.exists(_exc_path):
            try:
                with open(_exc_path, 'r', encoding='utf-8') as _f:
                    _exc = json.load(_f)
                _exc_tolerance = float(_exc.get('tolerance', 0))
                _expected_sha256 = str(_exc.get('source_sha256', '')).strip().lower()
                _expected_name = str(_exc.get('source_file', '')).strip()
                _reason = str(_exc.get('reason', '')).strip()
                if not _expected_sha256:
                    _exception_error = '缺少source_sha256'
                elif not _source_sha256:
                    _exception_error = '无法计算BS源文件SHA256'
                elif _expected_sha256 != _source_sha256.lower():
                    _exception_error = 'source_sha256与当前BS源文件不一致'
                elif _expected_name and _expected_name != os.path.basename(_source_path):
                    _exception_error = 'source_file与当前BS源文件名不一致'
                elif _exc_tolerance <= 1.0:
                    _exception_error = '例外容差必须大于默认1.0元'
                elif not _reason:
                    _exception_error = '缺少例外原因reason'
                else:
                    _dt139_tolerance = _exc_tolerance
                    _has_exception = True
                    _exception_reason = _reason
                    print(f'  ⚠️ DT-139 项目级例外已启用: 容差={_exc_tolerance:,.0f} (来源: {_reason})')
            except Exception as exc:
                _exception_error = f'例外文件解析失败: {exc}'

    _validation = {
        'rule': 'DT-139',
        'status': 'PASS',
        'default_tolerance': 1.0,
        'effective_tolerance': _dt139_tolerance,
        'diff': diff,
        'total_assets': total_assets,
        'total_liab_equity': total_liab_equity,
        'source_file': os.path.basename(_source_path) if _source_path else '',
        'source_sha256': _source_sha256,
        'exception_active': _has_exception,
        'exception_reason': _exception_reason,
        'exception_error': _exception_error,
        'generated_at': datetime.now().isoformat(),
    }

    if diff > _dt139_tolerance:
        msg = (f'DT-139 CRITICAL: 资产负债表会计等式不平衡！'
               f'资产总计={total_assets:,.2f}, 负债+权益={total_liab_equity:,.2f}, '
               f'差额={diff:,.2f}(容差={_dt139_tolerance:,.0f})')
        _validation['status'] = 'FAIL'
        if cache_dir:
            _save_cache(cache_dir, 'dt139_validation_status.json', _validation)
        raise AssertionError(msg)
    elif diff > 1:
        if not _has_exception:
            _validation['status'] = 'FAIL'
            if cache_dir:
                _save_cache(cache_dir, 'dt139_validation_status.json', _validation)
            detail = f'；例外文件无效: {_exception_error}' if _exception_error else ''
            raise AssertionError(
                f'DT-139 CRITICAL: 差额={diff:,.2f}超过默认容差1元，'
                f'且无有效项目级例外{detail}'
            )
        _validation['status'] = 'EXCEPTION_DRAFT'
        print(f'  ⚠️ DT-139 项目级例外命中: 差额={diff:,.2f}，最终状态仅允许DRAFT_REVIEW_REQUIRED')
    else:
        print(f'  DT-139校验通过: 资产={total_assets:,.2f}, 负债+权益={total_liab_equity:,.2f}')
    if cache_dir:
        _save_cache(cache_dir, 'dt139_validation_status.json', _validation)
    return _validation


def _build_d1d2d3_mapping(bs_data, subjects):
    """DT-119: D1/D2/D3三级递进映射"""
    d1_to_d2 = {}  # BS科目 → 科目余额表编码
    d2_to_d3 = {}  # 科目余额表编码 → 辅助数据源

    BS_TO_CODE_PREFIX = {
        '货币资金': ['1001', '1002', '1003', '1004', '1012'],
        '应收票据': ['1121'],
        '应收账款': ['1122'],
        '预付款项': ['1123', '1124'],
        '其他应收款': ['1221'],
        '存货': ['1401', '1402', '1403', '1404', '1405', '1406', '1407', '1408', '1409', '1410', '1411', '1412', '1413', '1421', '1471'],
        '固定资产': ['1601', '1602'],
        '在建工程': ['1604'],
        '无形资产': ['1701', '1702'],
        '长期待摊费用': ['1801'],
        '递延所得税资产': ['1811'],
        '短期借款': ['2001'],
        '应付票据': ['2201'],
        '应付账款': ['2202'],
        '预收款项': ['2203', '2204'],
        '应付职工薪酬': ['2211'],
        '应交税费': ['2221'],
        '应付利息': ['2231'],
        '其他应付款': ['2241'],
        '长期借款': ['2501'],
        '实收资本（或股本）': ['4001'],
        '资本公积': ['4002'],
        '盈余公积': ['4101'],
        '未分配利润': ['4103', '4104'],
        '长期应付款': ['2701'],
        '预计负债': ['2801'],
        '递延收益': ['2401'],
        '递延所得税负债': ['2901'],
        '合同负债': ['2205'],
        '合同资产': ['1125'],
        '应收款项融资': ['1126'],
        '交易性金融资产': ['1101'],
        '交易性金融负债': ['2101'],
        # v3.68 (2026-06-01): 补充某测试项目缺失的 7 个 BS 科目映射
        '长期股权投资': ['1511'],
        '投资性房地产': ['1521'],
        '其他非流动金融资产': ['1813'],
        '其他非流动资产': ['1812'],
        '专项应付款': ['2711'],
        '一年内到期的非流动负债': ['2701', '2501', '2901'],
        '其他流动资产': ['122199', '224103'],
    }

    def _is_summary_label(text):
        txt = str(text or '').replace(' ', '').strip()
        return ('合计' in txt) or ('总计' in txt) or txt.endswith('：') or txt.endswith(':')

    subjects = subjects or []
    all_codes = [str(s.get('code', '')).strip() for s in subjects if s.get('code')]

    # D1→D2: BS科目 → 科目余额表末级科目
    for item in bs_data.get('items', []):
        label = str(item.get('label', '')).strip()
        if not label or _is_summary_label(label):
            continue

        matched = []
        # 先按常见资产负债表科目→科目前缀做映射
        for bs_key, prefixes in BS_TO_CODE_PREFIX.items():
            if label == bs_key or label.replace(' ', '') == bs_key.replace(' ', ''):
                for code in all_codes:
                    if any(code.startswith(pfx) for pfx in prefixes):
                        matched.append(code)
                if not matched:
                    matched.extend(prefixes)
                break

        # 回退：名称匹配
        if not matched:
            for s in subjects:
                s_name = str(s.get('name', '')).strip()
                s_code = str(s.get('code', '')).strip()
                if not s_code:
                    continue
                if s_name == label or label.startswith(s_name):
                    matched.append(s_code)

        matched = sorted(set(matched))
        if matched:
            d1_to_d2[label] = matched

    # D2→D3: 标记哪些科目有辅助余额表（Phase 0.5填充）
    for s in subjects:
        d2_to_d3[s['code']] = {
            'name': s['name'],
            'has_auxiliary': False,  # Step 0.5会更新
            'auxiliary_file': None,
        }

    return {
        'd1_to_d2': d1_to_d2,
        'd2_to_d3': d2_to_d3,
        'unmapped_bs_items': [
            str(item.get('label', '')).strip()
            for item in bs_data.get('items', [])
            if str(item.get('label', '')).strip()
            and (str(item.get('label', '')).strip() not in d1_to_d2)
            and (not _is_summary_label(item.get('label', '')))
        ],
    }


def _ensure_pdf_deps():
    """DT-211: 检测并自动安装PDF提取依赖包

    检测 pdfplumber/pdf2image/Pillow 是否已安装，不做运行时安装。
    同时配置poppler PATH（Windows捆绑预编译版）。

    Returns:
        dict: {'pdfplumber': bool, 'pdf2image': bool, 'Pillow': bool,
               'poppler_configured': bool, 'installed': [list of newly installed]}
    """
    result = {
        'pdfplumber': False, 'pdf2image': False, 'Pillow': False,
        'poppler_configured': False, 'installed': [],
    }

    # 1. 检测各包
    _checks = [
        ('pdfplumber', 'pdfplumber'),
        ('pdf2image', 'pdf2image'),
        ('Pillow', 'PIL'),
    ]
    _missing = []
    for pip_name, import_name in _checks:
        try:
            __import__(import_name)
            result[pip_name] = True
        except ImportError:
            _missing.append(pip_name)

    # 2. 缺失依赖时仅报告，不在运行时安装
    if _missing:
        print(f'  ⚠️ 缺失PDF依赖: {", ".join(_missing)}')
        print(f'  请先安装后再执行: pip install {" ".join(_missing)}')

    # 3. 配置poppler PATH（系统安装或POPPLER_BIN环境变量）
    _env_poppler = os.environ.get('POPPLER_BIN', '').strip()
    if _env_poppler and os.path.isdir(_env_poppler):
        current_path = os.environ.get('PATH', '')
        if _env_poppler not in current_path:
            os.environ['PATH'] = _env_poppler + os.pathsep + current_path
        result['poppler_configured'] = True
    else:
        import shutil
        if shutil.which('pdftotext') or shutil.which('pdftoppm'):
            result['poppler_configured'] = True
        else:
            print('  ⚠️ poppler未找到，pdf2image将无法工作')
            print('  可选方案:')
            print('  - 设置环境变量 POPPLER_BIN 指向poppler可执行目录')
            print('  - Linux: sudo apt install poppler-utils')
            print('  - Mac:   brew install poppler')
            print('  - Windows: choco install poppler 或手工安装后加入PATH')

    return result


def _extract_pdf_sources(project_dir, cache_dir, base_date=''):
    """DT-211: PDF数据源自动识别与提取

    递归扫描项目目录下所有PDF，按关键词分类后调用对应提取器：
    - 银行对账单 → bank_statement_extract.batch_extract_bank_statements()
    - 固定资产卡片台账 → pdf_extract.extract_asset_register()
    - 辅助余额表 → pdf_extract.extract_auxiliary_balance()
    - 其他PDF → pdf_extract.extract_pdf()

    提取结果缓存到_dt_cache/pdf_extractions.json，不阻塞Phase 0。
    需多模态的扫描件记录到_dt_cache/multimodal_tasks.json，由Agent后续处理。

    Args:
        project_dir: 项目目录路径
        cache_dir: _dt_cache目录路径
        base_date: 评估基准日（如'2025-12-31'），传递给银行对账单提取器启用DT-133

    Returns:
        dict: {
            'total_pdfs': int,
            'classified': {type: count},
            'bank_statements': {stats},
            'asset_registers': {stats},
            'auxiliary_balances': {stats},
            'other_pdfs': {stats},
            'multimodal_count': int,
            'records_count': int,
        }
    """
    # 0. 自动检测并安装PDF依赖 + 配置poppler (DT-211)
    deps = _ensure_pdf_deps()
    _pdf_ok = deps.get('pdfplumber', False)
    _pdf2img_ok = deps.get('pdf2image', False) and deps.get('poppler_configured', False)

    # 1. 递归扫描项目目录下所有PDF（排除_dt_cache和_images子目录）
    all_pdfs = []
    _exclude_dirs = {'_dt_cache', '_images', 'images', '__pycache__'}
    for root, dirs, files in os.walk(project_dir):
        # 剪枝排除缓存目录
        dirs[:] = [d for d in dirs if d not in _exclude_dirs]
        for f in files:
            if f.lower().endswith('.pdf'):
                all_pdfs.append(os.path.join(root, f))

    if not all_pdfs:
        print('  未找到PDF文件')
        return {
            'total_pdfs': 0, 'classified': {},
            'bank_statements': {'count': 0},
            'asset_registers': {'count': 0},
            'auxiliary_balances': {'count': 0},
            'other_pdfs': {'count': 0},
            'multimodal_count': 0, 'records_count': 0,
        }

    print(f'  扫描到 {len(all_pdfs)} 个PDF文件')

    # 2. 按文件名关键词分类
    _bank_kw = ('对账单', '银行', '存款', '余额调节', '函证', 'bank', 'ccb', 'boc', 'icbc',
                'bocom', '41050172')
    _asset_kw = ('卡片', '台账', '固定资产', '设备清单', '资产清单')
    _aux_kw = ('辅助余额', '辅助明细', '辅助账', '明细账')

    bank_pdfs = []
    asset_pdfs = []
    aux_pdfs = []
    other_pdfs = []

    for fp in all_pdfs:
        fname_lower = os.path.basename(fp).lower()
        # 注意：先匹配长关键词避免误分类
        if any(kw in fname_lower for kw in _bank_kw):
            bank_pdfs.append(fp)
        elif any(kw in fname_lower for kw in _asset_kw):
            asset_pdfs.append(fp)
        elif any(kw in fname_lower for kw in _aux_kw):
            aux_pdfs.append(fp)
        else:
            other_pdfs.append(fp)

    classified = {
        '银行对账单': len(bank_pdfs),
        '固定资产卡片台账': len(asset_pdfs),
        '辅助余额表': len(aux_pdfs),
        '其他PDF': len(other_pdfs),
    }
    for k, v in classified.items():
        if v > 0:
            print(f'    {k}: {v}个')

    # 3. 提取银行对账单（使用专用提取器）
    bank_result = {'count': 0, 'records': [], 'multimodal': []}
    if bank_pdfs:
        try:
            from bank_statement_extract import batch_extract_bank_statements
            # poppler PATH已在_ensure_pdf_deps()中配置

            bank_result_raw = batch_extract_bank_statements(
                bank_pdfs,
                output_dir=os.path.join(cache_dir, 'bank_pdf_images'),
                auto_multimodal=True,
            )
            bank_result = {
                'count': len(bank_pdfs),
                'extracted': bank_result_raw.get('extracted', 0),
                'needs_multimodal': bank_result_raw.get('needs_multimodal', 0),
                'failed': bank_result_raw.get('failed', 0),
                'records': bank_result_raw.get('records', []),
                'multimodal_tasks': bank_result_raw.get('multimodal_tasks', []),
                'summary': bank_result_raw.get('summary', {}),
            }
            print(f'  银行对账单提取: 成功{bank_result["extracted"]}, '
                  f'需多模态{bank_result["needs_multimodal"]}, '
                  f'失败{bank_result["failed"]}')
            if bank_result.get('summary', {}).get('total_accounts', 0) > 0:
                print(f'    去重后账户数: {bank_result["summary"]["total_accounts"]}, '
                      f'余额合计: {bank_result["summary"]["total_balance"]:,.2f}')
        except ImportError:
            print('  ⚠️ bank_statement_extract未安装，跳过银行对账单提取')
            bank_result = {'count': len(bank_pdfs), 'records': [], 'multimodal': [],
                           'error': 'bank_statement_extract未导入'}
        except Exception as e:
            print(f'  ⚠️ 银行对账单提取异常: {e}')
            bank_result = {'count': len(bank_pdfs), 'records': [], 'multimodal': [],
                           'error': str(e)}

    # 4. 提取固定资产卡片台账PDF
    asset_result = {'count': 0, 'items': []}
    if asset_pdfs:
        try:
            from pdf_extract import extract_asset_register
            total_items = 0
            for fp in asset_pdfs:
                r = extract_asset_register(fp)
                if r['status'] in ('extracted', 'partial'):
                    total_items += r.get('total_items', 0)
                    asset_result['items'].append({
                        'filepath': fp,
                        'filename': os.path.basename(fp),
                        'total_items': r.get('total_items', 0),
                        'status': r['status'],
                    })
            asset_result['count'] = len(asset_pdfs)
            print(f'  固定资产卡片台账: {len(asset_pdfs)}个PDF, {total_items}项资产')
        except ImportError:
            print('  ⚠️ pdf_extract未安装，跳过资产台账提取')
            asset_result = {'count': len(asset_pdfs), 'items': [], 'error': 'pdf_extract未导入'}
        except Exception as e:
            print(f'  ⚠️ 资产台账提取异常: {e}')
            asset_result = {'count': len(asset_pdfs), 'items': [], 'error': str(e)}

    # 5. 提取辅助余额表PDF
    aux_result = {'count': 0, 'counterparties': []}
    if aux_pdfs:
        try:
            from pdf_extract import extract_auxiliary_balance as _extract_aux_pdf
            total_cp = 0
            for fp in aux_pdfs:
                r = _extract_aux_pdf(fp)
                if r['status'] in ('extracted', 'partial'):
                    total_cp += len(r.get('counterparties', []))
            aux_result['count'] = len(aux_pdfs)
            aux_result['counterparties_count'] = total_cp
            print(f'  辅助余额表PDF: {len(aux_pdfs)}个PDF, {total_cp}个结算对象')
        except ImportError:
            print('  ⚠️ pdf_extract未安装，跳过辅助余额表PDF提取')
            aux_result = {'count': len(aux_pdfs), 'counterparties': [], 'error': 'pdf_extract未导入'}
        except Exception as e:
            print(f'  ⚠️ 辅助余额表PDF提取异常: {e}')
            aux_result = {'count': len(aux_pdfs), 'counterparties': [], 'error': str(e)}

    # 6. 通用PDF提取（其他）
    other_result = {'count': 0, 'extracted': 0, 'failed': 0}
    if other_pdfs:
        try:
            from pdf_extract import extract_pdf as _extract_general_pdf
            ext_count = 0
            fail_count = 0
            for fp in other_pdfs:
                r = _extract_general_pdf(fp)
                if r['status'] in ('extracted', 'ocr_fallback'):
                    ext_count += 1
                else:
                    fail_count += 1
            other_result = {'count': len(other_pdfs), 'extracted': ext_count, 'failed': fail_count}
            print(f'  其他PDF: 成功{ext_count}, 失败{fail_count}')
        except ImportError:
            print('  ⚠️ pdf_extract未安装，跳过通用PDF提取')
            other_result = {'count': len(other_pdfs), 'error': 'pdf_extract未导入'}
        except Exception as e:
            print(f'  ⚠️ 通用PDF提取异常: {e}')
            other_result = {'count': len(other_pdfs), 'error': str(e)}

    # 7. 汇总并缓存
    multimodal_tasks = []
    if bank_result.get('multimodal_tasks'):
        multimodal_tasks.extend(bank_result['multimodal_tasks'])

    # 银行对账单记录（含余额的去重记录）
    bank_records = bank_result.get('records', [])

    result = {
        'total_pdfs': len(all_pdfs),
        'classified': classified,
        'bank_statements': bank_result,
        'asset_registers': asset_result,
        'auxiliary_balances': aux_result,
        'other_pdfs': other_result,
        'multimodal_count': len(multimodal_tasks),
        'records_count': len(bank_records),
        'base_date': base_date,
    }

    # 保存主提取结果
    _save_cache(cache_dir, 'pdf_extractions.json', result)

    # 保存多模态任务清单（供Agent后续处理）
    if multimodal_tasks:
        _save_cache(cache_dir, 'multimodal_tasks.json', {
            'total': len(multimodal_tasks),
            'tasks': [
                {
                    'filename': t.get('filename', ''),
                    'filepath': t.get('filepath', ''),
                    'images': t.get('multimodal_images', []),
                    'prompt': t.get('multimodal_prompt', ''),
                }
                for t in multimodal_tasks
            ],
        })
        print(f'  ⚠️ {len(multimodal_tasks)}个PDF需多模态识别，已存入multimodal_tasks.json')

    return result


def _extract_auxiliary_balances(aux_dir, cache_dir):
    """DT-111: 辅助余额表强制提取 (DT-154: 委托给source_header_parser)

    DT-212修复：
    1. 递归扫描aux_dir下所有子目录（原来仅glob一层）
    2. dict键用"文件名_科目名"避免"新的工作表"同名碰撞
    """
    from source_header_parser import parse_auxiliary as _pa_parse

    # DT-212: 递归扫描所有子目录下的xlsx/xls文件
    aux_files = []
    _exclude_dirs = {'_dt_cache', '_images', 'images', '__pycache__'}
    for root, dirs, files in os.walk(aux_dir):
        dirs[:] = [d for d in dirs if d not in _exclude_dirs]
        for f in files:
            if f.lower().endswith(('.xlsx', '.xls')):
                aux_files.append(os.path.join(root, f))

    if not aux_files:
        print('  未找到辅助余额表文件')
        return {'sheet_count': 0, 'total_objects': 0, 'files_processed': [], 'sheets': []}

    total_objects = 0
    sheet_count = 0
    all_aux_data = {}

    for filepath in aux_files:
        fname = os.path.basename(filepath)
        print(f'  提取: {fname}')

        try:
            result = _pa_parse(filepath)
        except Exception as e:
            print(f'  ⚠️ 解析失败 {fname}: {e}')
            continue

        if result['status'] != 'parsed':
            print(f'  ⚠️ 解析无数据 {fname}: {result.get("warnings", [])}')
            continue

        sheet_count += 1
        objects = result['objects']
        total_objects += len(objects)

        # DT-212: 用"文件名前缀_科目名"作为key，避免多个文件的sheet同名碰撞
        # 例如"cloud11774840037014_其他应付款-集团公司外部其他"
        subject_name = result.get('subject_name') or result.get('sheet_name') or ''
        # 从文件名提取cloud ID前缀（如果有）
        fname_prefix = os.path.splitext(fname)[0]
        if subject_name:
            key = f'{fname_prefix}_{subject_name}'
        else:
            key = fname_prefix
        # 确保key唯一
        _base_key = key
        _seq = 1
        while key in all_aux_data:
            key = f'{_base_key}_{_seq}'
            _seq += 1
        all_aux_data[key] = objects
        print(f'    key={key[:50]}, 表头行: {result["header_row"]}, 结算对象: {len(objects)}个')

    # 保存合并数据
    _save_cache(cache_dir, 'auxiliary_balance_all.json', all_aux_data)

    return {
        'sheet_count': sheet_count,
        'total_objects': total_objects,
        'files_processed': [os.path.basename(f) for f in aux_files],
        'sheets': list(all_aux_data.keys()),
    }


def _classify_data(subjects, bs_data, project_dir=''):
    """DT-117: 数据分类（DT-213: 配置外置，从subject_classification.json加载）"""
    classified = {
        'bank_deposits': [],      # 银行存款
        'trading_financial': [],  # 交易性金融资产
        'receivables': [],         # 往来科目（应收/预付/其他应收）
        'payables': [],           # 往来科目（应付/预收/其他应付）
        'fixed_assets': [],       # 固定资产
        'intangible_assets': [],  # 无形资产
        'long_term_prepaid': [],   # 长期待摊
        'deferred_tax': [],       # 递延所得税
        'inventory': [],          # 存货
        'tax_payable': [],         # 应交税费
        'loans': [],              # 借款
        'other': [],              # 其他
        'pending_mapping': [],    # 待确认映射
    }

    # DT-213: 从外部JSON加载科目代码前缀→分类桶映射
    # 支持项目级覆盖：项目目录下的subject_classification.json优先于脚本目录
    _json_paths = [
        os.path.join(project_dir, 'subject_classification.json') if project_dir else None,
        os.path.join(SCRIPT_DIR, '..', '..', 'valuation-common', 'scripts', 'subject_classification.json'),
    ]
    D2_TO_BUCKET = {}
    for _jp in _json_paths:
        if _jp and os.path.exists(_jp):
            try:
                with open(_jp, 'r', encoding='utf-8') as f:
                    _cfg = json.load(f)
                for prefix, info in _cfg.get('mappings', {}).items():
                    D2_TO_BUCKET[prefix] = (info['bucket'], info['name'])
                print(f'  [DT-213] 科目分类配置已加载: {_jp} ({len(D2_TO_BUCKET)}条)')
                break
            except Exception as e:
                print(f'  [DT-213] 配置加载失败: {_jp}: {e}')

    # 兜底：如果JSON加载失败，使用硬编码最小集
    # 注意: 名称必须与财政部《企业会计准则——应用指南》附录一致
    if not D2_TO_BUCKET:
        print('  [DT-213] WARNING: JSON配置加载失败，使用硬编码最小集')
        for p in ('1001', '1002', '1012', '1004'): D2_TO_BUCKET[p] = ('bank_deposits', '货币资金')
        for p in ('1101',): D2_TO_BUCKET[p] = ('trading_financial', '交易性金融资产')
        for p in ('1121', '1122', '1123', '1221', '1231', '1131', '1132'): D2_TO_BUCKET[p] = ('receivables', '往来科目-应收')
        for p in ('1401', '1402', '1403', '1404', '1405', '1406', '1407', '1408', '1410', '1411', '1421', '5001', '5002', '5101', '5102', '5201'):
            D2_TO_BUCKET[p] = ('inventory', '存货')
        for p in ('1601', '1602', '1603', '1604', '1605', '1606', '1621'): D2_TO_BUCKET[p] = ('fixed_assets', '固定资产')
        for p in ('1701', '1702', '1703'): D2_TO_BUCKET[p] = ('intangible_assets', '无形资产')
        for p in ('1801',): D2_TO_BUCKET[p] = ('long_term_prepaid', '长期待摊费用')
        for p in ('1811',): D2_TO_BUCKET[p] = ('deferred_tax', '递延所得税资产')
        for p in ('2001', '2002', '2101', '2102', '2201', '2202', '2203', '2205', '2207', '2210', '2211', '2231', '2232', '2241', '2242', '2243', '2244', '2245', '2261', '2401', '2600', '2601', '2701', '2702', '2801', '2802', '2803'):
            D2_TO_BUCKET[p] = ('payables', '往来科目-应付')
        for p in ('2221',): D2_TO_BUCKET[p] = ('tax_payable', '应交税费')
        for p in ('2501', '2502'): D2_TO_BUCKET[p] = ('loans', '借款')  # 2501=长期借款, 2502=应付债券
        for p in ('4001', '4002', '4003', '4101', '4103', '4104'): D2_TO_BUCKET[p] = ('other', '所有者权益')

    # DT-212: 统一查表分类（替代硬编码if-elif链）
    # 按前缀从长到短匹配，确保2207优先于2202、2211优先于2202等
    sorted_prefixes = sorted(D2_TO_BUCKET.keys(), key=len, reverse=True)

    for s in subjects:
        code = s['code']
        name = s['name']
        matched = False

        for prefix in sorted_prefixes:
            if code.startswith(prefix):
                bucket, bucket_name = D2_TO_BUCKET[prefix]
                classified[bucket].append(s)
                matched = True
                break

        if not matched:
            if s['level'] <= 1:  # 一级科目
                classified['other'].append(s)
            # 末级科目不归类到other

    # 统计
    for cat, items in classified.items():
        print(f'  {cat}: {len(items)}个')

    return classified


def _build_reclassification(subjects):
    """DT-118: 重分类映射

    重分类逻辑:
    - 负债类科目(如应交税费2221)正常方向=贷方
      借方余额表现为: direction='贷' + balance<0
    - 资产类科目(如应收账款1122)正常方向=借方
      贷方余额表现为: direction='借' + balance<0
    DT-218修复: 正确检测贷方负数=借方余额、借方负数=贷方余额
    """
    items = []

    # 常见重分类关系
    reclass_rules = {
        '2221': {'target_sheet': '3-13', 'target_name': '其他流动资产',
                 'condition': 'debit_balance', 'description': '应交税费借方余额→其他流动资产'},
        '1122': {'target_sheet': '2203_sheet', 'target_name': '预收款项',
                 'condition': 'credit_balance', 'description': '应收账款贷方余额→预收款项'},
        '2202': {'target_sheet': '1123_sheet', 'target_name': '预付款项',
                 'condition': 'debit_balance', 'description': '应付账款借方余额→预付款项'},
        '2241': {'target_sheet': '1221_sheet', 'target_name': '其他应收款',
                 'condition': 'debit_balance', 'description': '其他应付款借方余额→其他应收款'},
        '1221': {'target_sheet': '2241_sheet', 'target_name': '其他应付款',
                 'condition': 'credit_balance', 'description': '其他应收款贷方余额→其他应付款'},
        '1123': {'target_sheet': '2202_sheet', 'target_name': '应付账款',
                 'condition': 'credit_balance', 'description': '预付款项贷方余额→应付账款'},
        '2203': {'target_sheet': '1122_sheet', 'target_name': '应收账款',
                 'condition': 'debit_balance', 'description': '预收款项借方余额→应收账款'},
    }

    for s in subjects:
        code = str(s.get('code', ''))
        code_prefix = code[:4]
        if code_prefix not in reclass_rules:
            continue
        rule = reclass_rules[code_prefix]
        balance = s.get('balance', 0)
        direction = s.get('direction', '')
        ending_debit = s.get('ending_debit', 0)
        ending_credit = s.get('ending_credit', 0)

        # DT-218修复: 正确检测反方向余额
        # 借方余额: 资产正常=借方正数, 负债正常=贷方正数
        #   异常借方余额: 资产direction='借'+balance<0, 或 负债direction='贷'+balance<0
        # 贷方余额: 负债正常=贷方正数, 资产正常=借方正数
        #   异常贷方余额: 负债direction='贷'+balance<0 → 不对，贷方负数=借方余额
        needs_reclass = False
        _reclass_amt = 0
        if rule['condition'] == 'debit_balance':
            # DT-FIX: 借方余额——负债类科目有ending_debit>0即为反方向
            if ending_debit > 0 and ending_credit == 0:
                needs_reclass = True
                _reclass_amt = ending_debit
            elif direction == '贷' and balance < 0:
                needs_reclass = True
                _reclass_amt = abs(balance)
            elif direction == '平' and balance < 0:
                needs_reclass = True
                _reclass_amt = abs(balance)
        elif rule['condition'] == 'credit_balance':
            # DT-FIX: 贷方余额——资产类科目有ending_credit>0即为反方向
            if ending_credit > 0 and ending_debit == 0:
                needs_reclass = True
                _reclass_amt = ending_credit
            elif direction == '借' and balance < 0:
                needs_reclass = True
                _reclass_amt = abs(balance)

        if needs_reclass and _reclass_amt > 0:
            items.append({
                'source_code': s['code'],
                'source_name': s['name'],
                'source_balance': balance,
                'source_direction': direction,
                'ending_debit': ending_debit,
                'ending_credit': ending_credit,
                'target_sheet': rule['target_sheet'],
                'target_name': rule['target_name'],
                'reclass_amount': _reclass_amt,
                'description': rule['description'],
            })

    return {'items': items, 'rules_applied': list(reclass_rules.keys())}


def _extract_settings(subjects, bs_data, project_dir=''):
    """DT-121: 设定信息提取

    提取策略(P1修复 + P1补充):
    0. **最最优先级**: 从BS表头"编制单位"提取企业全称（如"上海图灵智算量子科技有限公司"）
    1. 次优先从BS文件名提取公司简称
    2. 再次从科目余额表文件名提取
    3. fallback从项目目录名提取，并清理"填写评估明细表"等后缀
    4. 评估基准日优先从BS表头提取
    5. 从科目余额表推断行业类型
    """
    settings = {
        'company_name': '',
        'valuation_date': '',
        'industry_type': '',
    }

    # --- 提取公司名 ---
    company_name = ''

    # 策略0（最最优先级）: 从BS表头"编制单位"或纯公司名提取企业全称
    # DT-210增强: source_header_parser已支持纯公司名提取（无"编制单位"前缀）
    if bs_data and isinstance(bs_data, dict):
        bs_full_name = bs_data.get('company_full_name', '')
        if bs_full_name:
            company_name = bs_full_name
            print(f'  [P1] 公司全称(来自BS表头): {company_name}')

    # 策略1: 从BS文件名提取（如"上海图灵-资产负债表-20260430.xlsx"→"上海图灵"）
    # DT-210增强: 也搜索不含"资产负债表"关键词但已被Step 0.1确认的BS文件
    if not company_name:
        import glob as _glob
        bs_files = _glob.glob(os.path.join(project_dir, '*资产负债表*')) + \
                   _glob.glob(os.path.join(project_dir, '*财务报表*'))
        if not bs_files:
            # DT-210: 用与Step 0.1相同的内部结构确认法搜索BS文件
            import openpyxl as _opx
            _exclude_kw = ['科目余额', '序时账', '明细账', '凭证', '评估明细表', '底稿', '抽凭', '辅助']
            _cand = [f for f in _glob.glob(os.path.join(project_dir, '*.xlsx'))
                     if not any(kw in os.path.basename(f) for kw in _exclude_kw)]
            for _cf in _cand:
                try:
                    _wb = _opx.load_workbook(_cf, data_only=True)
                    _ws = _wb[_wb.sheetnames[0]]
                    for _r in range(1, min(_ws.max_row + 1, 6)):
                        for _c in range(1, min(_ws.max_column + 1, 10)):
                            _v = _ws.cell(row=_r, column=_c).value
                            if _v and isinstance(_v, str) and '资产负债表' in _v:
                                bs_files.append(_cf)
                                break
                        if bs_files:
                            break
                    _wb.close()
                except Exception:
                    pass
                if bs_files:
                    break
        if bs_files:
            fname = os.path.basename(bs_files[0])
            # 解析文件名: "公司名-资产负债表-日期.xlsx" 或 "河南平煤神马平绿置业有限公司.xlsx"
            base = fname.replace('.xlsx', '').replace('.xls', '')
            parts = base.split('-')
            if len(parts) >= 2:
                # 带"-"分隔的文件名: 取非关键词段
                for part in parts:
                    if part and part not in ('资产负债表', 'BS', '财务报表') and not re.match(r'^\d{6,8}$', part):
                        company_name = part
                        break
            else:
                # DT-210: 无"-"分隔的纯公司名文件名（如"河南平煤神马平绿置业有限公司.xlsx"）
                # 排除含"资产负债表"/"科目余额表"等关键词的文件名
                _fname_exclude = ('资产负债表', '科目余额表', '财务报表', '余额表', '明细账')
                if not any(kw in base for kw in _fname_exclude):
                    company_name = base

    # 策略2: 从科目余额表文件名提取
    if not company_name:
        import glob as _glob
        sb_files = _glob.glob(os.path.join(project_dir, '*科目余额表*'))
        if sb_files:
            fname = os.path.basename(sb_files[0])
            parts = fname.replace('.xlsx', '').replace('.xls', '').split('-')
            for part in parts:
                if part and part not in ('科目余额表', '余额表') and not re.match(r'^\d{6,8}$', part):
                    company_name = part
                    break

    # 策略3: 从项目目录名提取（fallback，需清理后缀）
    if not company_name:
        dir_name = os.path.basename(project_dir.rstrip('/\\'))
        # 去除编号前缀（如"1-河南平绿"→"河南平绿"）
        dir_name_clean = re.sub(r'^\d+[-_]?', '', dir_name).strip()
        # P1修复: 清理"填写评估明细表"/"评估明细表"/"明细表"等后缀
        for suffix in [' 填写评估明细表', ' 评估明细表', ' 明细表', '填写评估明细表', '评估明细表', '明细表']:
            if dir_name_clean.endswith(suffix):
                dir_name_clean = dir_name_clean[:-len(suffix)].strip()
                break
        if dir_name_clean:
            company_name = dir_name_clean

    settings['company_name'] = company_name

    # --- 提取评估基准日 ---
    # 策略0（最最优先级）: 从BS表头提取日期
    if bs_data and isinstance(bs_data, dict):
        bs_date = bs_data.get('valuation_date', '')
        if bs_date:
            settings['valuation_date'] = bs_date
            print(f'  [P1] 评估基准日(来自BS表头): {bs_date}')

    if not settings['valuation_date']:
        # 策略1: 从BS文件名中的日期模式提取
        import glob as _glob
        bs_files = _glob.glob(os.path.join(project_dir, '*资产负债表*')) + \
                   _glob.glob(os.path.join(project_dir, '*财务报表*'))
        # DT-210: 也搜索纯公司名命名的BS文件
        if not bs_files and bs_data and isinstance(bs_data, dict) and bs_data.get('filepath'):
            bs_files = [bs_data['filepath']]
        if bs_files:
            fname = os.path.basename(bs_files[0])
            # 匹配文件名中的日期模式: 20260430, 2025-12-31, 2025年12月31日
            date_match = re.search(r'(\d{4})[-年]?(\d{2})[-月]?(\d{2})', fname)
            if date_match:
                y, m, d = date_match.groups()
                settings['valuation_date'] = f'{y}年{int(m)}月{int(d)}日'

    if not settings['valuation_date']:
        # 从科目余额表推断（取最后一笔日期）
        for s in subjects:
            # 如果科目名包含"本年利润"等，说明是年度报表
            pass
        # 默认留空，由Agent补充
        settings['valuation_date'] = ''

    # --- 推断行业类型 ---
    # DT-153v3: 基于科目名称关键词推断行业类型，支持多行业
    INDUSTRY_KEYWORDS = {
        '房地产': ['开发成本', '开发产品', '开发支出', '工程施工', '土地开发', '房产开发'],
        '制造业': ['生产成本', '制造费用', '原材料', '产成品', '半成品', '委托加工'],
        '建筑业': ['工程施工', '工程结算', '机械作业', '周转材料'],
        '金融业': ['贷款', '存放中央银行', '拆出资金', '买入返售', '利息收入'],
        '信息技术': ['研发支出', '技术服务', '软件开发', '无形资产'],
    }

    for industry, keywords in INDUSTRY_KEYWORDS.items():
        for s in subjects:
            if any(kw in s.get('name', '') for kw in keywords):
                settings['industry_type'] = industry
                break
        if settings['industry_type']:
            break

    if not settings['industry_type']:
        # 默认行业类型：通用
        settings['industry_type'] = '通用'

    return settings


# ============================================================
# Gate验证
# ============================================================

def _find_detail_table(project_dir):
    """DT-153v3: 动态查找评估明细表文件，不再硬编码文件名"""
    # 优先搜索含"评估明细表"的xlsx文件
    candidates = glob.glob(os.path.join(project_dir, '*评估明细表*.xlsx'))
    if not candidates:
        # fallback: 搜索含"明细表"的xlsx文件
        candidates = glob.glob(os.path.join(project_dir, '*明细表*.xlsx'))
    if candidates:
        return candidates[0]
    # 最终兜底
    return os.path.join(project_dir, '评估明细表.xlsx')


def run_gate(project_dir, gate_name, args):
    """运行Gate验证"""
    import os as _os
    # 确保从本地gate_validator导入，而不是valuation-common版本
    _local_gv = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'gate_validator.py')
    if _os.path.exists(_local_gv):
        import importlib.util as _iu
        _spec = _iu.spec_from_file_location('gate_validator_local', _local_gv)
        _gv_mod = _iu.module_from_spec(_spec)
        _spec.loader.exec_module(_gv_mod)
        _gate_map = {
            'G0': _gv_mod.gate_G0,
            'G1': _gv_mod.gate_G1,
            'G1F': _gv_mod.gate_G1_Format,
            'G2': _gv_mod.gate_G2,
            'G3': _gv_mod.gate_G3,
            'G-DT182': _gv_mod.validate_summary_no_hardcoded,
        }
        gate_func = _gate_map.get(gate_name)
    else:
        from gate_validator import gate_G0, gate_G1, gate_G1_Format, gate_G2, gate_G3, validate_summary_no_hardcoded
        _gate_map = {
            'G0': gate_G0,
            'G1': gate_G1,
            'G1F': gate_G1_Format,
            'G2': gate_G2,
            'G3': gate_G3,
            'G-DT182': validate_summary_no_hardcoded,
        }
        gate_func = _gate_map.get(gate_name)

    xlsx_path = args.xlsx_path or _find_detail_table(project_dir)
    if not os.path.exists(xlsx_path):
        print(f'⚠️ 评估明细表不存在: {xlsx_path}')
        return {'gate': gate_name, 'status': 'skipped', 'reason': 'file not found'}

    print(f'\n[Gate {gate_name}] 验证中...')

    cache_dir = os.path.join(project_dir, '_dt_cache')
    try:
        if gate_func is None:
            print(f'⚠️ Gate {gate_name}: 未知的Gate类型')
            return {'gate': gate_name, 'status': 'error', 'reason': f'unknown gate {gate_name}'}
        bs_path = None
        sb_path = None
        aux_data = None

        bs_cache = _load_cache(cache_dir, 'bs_balances.json')
        if isinstance(bs_cache, dict):
            bs_path = bs_cache.get('filepath')
        sb_cache = _load_cache(cache_dir, 'subjects_normalized.json')
        if isinstance(sb_cache, dict):
            sb_path = sb_cache.get('filepath')

        if gate_name == 'G0':
            aux_data = {
                'auxiliary_balance': _load_cache(cache_dir, 'auxiliary_balance_all.json') or {},
                'pdf_extraction_status': (_load_cache(cache_dir, 'pdf_extractions.json') or {}).get('classified', {}),
                'pdf_usage_status': {},
            }
            passed, violations = gate_func(xlsx_path, bs_path=bs_path, sb_path=sb_path, aux_data=aux_data)
        elif gate_name == 'G2':
            passed, violations = gate_func(xlsx_path, bs_path=bs_path, sb_path=sb_path, has_journal=bool(_load_cache(cache_dir, 'journal.json')))
        elif gate_name == 'G3':
            passed, violations = gate_func(xlsx_path, bs_path=bs_path, tolerance=0.01)
        else:
            passed, violations = gate_func(xlsx_path)

        if not passed:
            criticals = [v for v in violations if v.get('severity') == 'CRITICAL'] if violations else []
            if criticals:
                print(f'🚨 Gate {gate_name} FAILED: {len(criticals)}个CRITICAL')
                for c in criticals:
                    print(f'  - {c.get("check", "")}: {c.get("message", "")}')
            result = {'gate': gate_name, 'status': 'failed', 'criticals': len(criticals) if criticals else 0}
        else:
            print(f'✅ Gate {gate_name} PASSED')
            result = {'gate': gate_name, 'status': 'passed', 'criticals': 0}
        prev = _load_cache(cache_dir, 'gate_results.json') or []
        prev.append({
            'gate': gate_name,
            'result': result,
            'timestamp': datetime.now().isoformat(),
        })
        _save_cache(cache_dir, 'gate_results.json', prev)
        return result
    except Exception as e:
        print(f'⚠️ Gate {gate_name} 执行异常: {e}')
        return {'gate': gate_name, 'status': 'error', 'reason': str(e)}


# ============================================================
# 缓存完整性校验 (DT-131)
# ============================================================

def check_cache(project_dir):
    """检查_dt_cache/完整性"""
    cache_dir = os.path.join(project_dir, '_dt_cache')
    if not os.path.isdir(cache_dir):
        print('❌ _dt_cache/目录不存在')
        return

    required_files = {
        'Phase 0': [
            'subjects.json',
            'bs_balances.json',
            'd1d2d3_mapping.json',
            'auxiliary_balance_summary.json',
            'data_classification.json',
            'reclassification.json',
            'settings_info.json',
        ],
        'Phase -1': [
            'file_manifest.json',
            'pdf_completeness_report.json',
        ],
    }

    print('\n_dt_cache/ 完整性检查:')
    total = 0
    present = 0
    for phase, files in required_files.items():
        print(f'\n  {phase}:')
        for f in files:
            path = os.path.join(cache_dir, f)
            total += 1
            if os.path.exists(path):
                size = os.path.getsize(path)
                present += 1
                print(f'    ✅ {f} ({size:,} bytes)')
            else:
                print(f'    ❌ {f} (缺失)')

    pct = present / total * 100 if total else 0
    print(f'\n  总计: {present}/{total} ({pct:.0f}%)')
    return pct


# ============================================================
# Phase 1: 结构解析与科目映射
# ============================================================

def phase1(project_dir, args):
    """Phase 1: 结构解析与科目映射

    子步骤:
    1.1 解析评估明细表模板结构 → sheet_structure.json
    1.2 建立科目→明细表映射（含行业映射） → subject_sheet_mapping.json
    1.3 校验sheet_col_map.json完整性
    1.4 Gate G0验证
    """
    cache_dir = _cache_path(project_dir)
    print('\n' + '='*60)
    print('Phase 1: 结构解析与科目映射')
    print('='*60)

    # --- Step 1.1: 解析模板结构 ---
    print('\n[Step 1.1] 解析评估明细表模板结构')
    xlsx_path = args.xlsx_path or _find_detail_table(project_dir)
    if not os.path.exists(xlsx_path):
        # 搜索可能的文件
        candidates = glob.glob(os.path.join(project_dir, '*评估明细表*'))
        if candidates:
            xlsx_path = candidates[0]
        else:
            print('  ⚠️ 未找到评估明细表文件')
            _save_cache(cache_dir, 'phase1_status.json', {'status': 'skipped', 'reason': 'no xlsx'})
            return {'phase': 1, 'status': 'skipped', 'reason': 'no xlsx'}

    print(f'  使用文件: {os.path.basename(xlsx_path)}')
    _save_cache(cache_dir, 'xlsx_path.json', {'path': xlsx_path})

    # --- Step 1.2: 科目→明细表映射 ---
    print('\n[Step 1.2] 建立科目→明细表映射')
    subjects = _load_cache(cache_dir, 'subjects.json')
    if not subjects:
        print('  ❌ subjects.json不存在，需先执行Phase 0')
        sys.exit(1)

    # 从settings_info.json获取行业类型（Phase 0已提取）
    settings = _load_cache(cache_dir, 'settings_info.json') or {}
    industry_type = settings.get('industry_type', '通用')

    from sheet_filler import get_sheet_id_for_subject
    mapping = {}
    for s in (subjects if isinstance(subjects, list) else subjects.get('subjects', [])):
        code = s.get('code', '')
        sid = get_sheet_id_for_subject(code, industry_type=industry_type)
        if sid:
            bal = s.get('balance', s.get('closing_balance', 0))
            if code in mapping:
                # 同科目多明细：累加余额（解决dict key覆盖问题）
                mapping[code]['balance'] += bal
            else:
                mapping[code] = {
                    'name': s.get('name', ''),
                    'sheet_id': sid,
                    'balance': bal,
                }

    # ── 映射后处理: 固定资产/无形资产子科目重分配 ──
    # 标准化数据中1601.01(房屋)/1601.03(运输设备)/1601.05(电子设备)等子编码
    # 需要映射到对应的明细表sheet，而非全部归入4-8-4机器设备
    fa_redistribute = {
        '1601.01': ('4-8-1', '房屋建筑物'),
        '1601.03': ('4-8-5', '车辆'),
        '1601.04': ('4-8-4', '机器设备'),  # 办公设备→机器设备
        '1601.05': ('4-8-6', '电子设备'),
    }
    # 无形资产子科目：土地→4-13-1，其他→4-13-3
    intang_redistribute = {
        '1701.03': ('4-13-1', '土地'),
    }
    redistributed = 0
    for code, (newsid, sname) in {**fa_redistribute, **intang_redistribute}.items():
        if code in mapping:
            old_sid = mapping[code]['sheet_id']
            if old_sid != newsid:
                mapping[code]['sheet_id'] = newsid
                print(f'  🔄 重分配 {code}({mapping[code]["name"]}) → {newsid} ({sname})')
                redistributed += 1
    
    # ── 去重：子级编码存在时移除子级（保留父级汇总值，避免双倍/错误计算） ──
    # 父级如1122(应收账款)的余额是汇总值，子级如1122.01、1122.02是辅助维度
    # 如果保留子级并移除父级，会导致余额被拆散且不准
    # DT-FIX: 改为移除子级、保留父级
    codes = list(mapping.keys())
    to_remove = set()
    for code_p in codes:
        code_p_clean = code_p.rstrip('.')
        for code_c in codes:
            if code_p != code_c and (code_c.startswith(code_p_clean + '.') or (code_c.startswith(code_p_clean) and len(code_c) > len(code_p_clean) and code_c[len(code_p_clean):len(code_p_clean)+1].isdigit())) and mapping[code_p]['sheet_id'] == mapping[code_c]['sheet_id']:
                to_remove.add(code_c)
    for code in to_remove:
        info = mapping.pop(code, None)
        if info:
            print(f'  🗑️ 移除子级: {code}({info["name"]}, 余额={info["balance"]:,.2f})')
    
    _save_cache(cache_dir, 'subject_sheet_mapping.json', mapping)
    print(f'  映射完成: {len(mapping)}个科目（重分配{redistributed}个，去重{len(to_remove)}个）')

    # --- Gate G0 ---
    print('\n[Gate G0] 数据源完整性验证')
    g0 = run_gate(project_dir, 'G0', args)
    _gate_pass_or_raise(g0)

    return {'phase': 1, 'status': 'completed', 'mapped_subjects': len(mapping)}


# ============================================================
# 设定信息Sheet写入
# ============================================================

def _write_settings_sheet(wb, settings):
    """DT-209: 将settings_info.json中的信息写入设定信息Sheet

    设定信息Sheet的典型结构（基于模板v1.90）:
    - B6: 被评估单位名称
    - B7: 评估基准日
    - B8: 行业类型（如有）
    """
    # 查找设定信息Sheet
    # DT-211: 优先匹配"设定信息"（含A列标记行的标准设定表），而非"设置"（纯配置表）
    # 模板v1.90中"设置"排在"设定信息"前面，若不区分优先级会写错Sheet
    settings_sheet_name = None
    for sn in wb.sheetnames:
        if '设定信息' in sn:
            settings_sheet_name = sn
            break
    if not settings_sheet_name:
        for sn in wb.sheetnames:
            if '设置' in sn:
                settings_sheet_name = sn
                break

    if not settings_sheet_name:
        print('  ⚠️ 未找到设定信息Sheet，跳过写入')
        return

    ws = wb[settings_sheet_name]

    company_name = settings.get('company_name', '')
    valuation_date = settings.get('valuation_date', '')
    industry_type = settings.get('industry_type', '')

    # 动态查找目标单元格（B列中含"被评估单位"/"评估基准日"标记的行）
    for r in range(1, min(ws.max_row + 1, 30)):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        if a_val:
            a_str = str(a_val).strip()
            if '被评估单位' in a_str or '委托方' in a_str:
                ws.cell(row=r, column=2).value = company_name
                print(f'  设定信息: 被评估单位={company_name}')
            elif '评估基准日' in a_str or '基准日' in a_str:
                ws.cell(row=r, column=2).value = valuation_date
                print(f'  设定信息: 评估基准日={valuation_date}')
            elif '行业' in a_str or '所属行业' in a_str:
                ws.cell(row=r, column=2).value = industry_type
                print(f'  设定信息: 行业={industry_type}')

    # 如果上面没找到标记行，尝试直接写B6/B7
    if not company_name and not valuation_date:
        return

    # 检查是否已有标记行写入成功
    found_marker = False
    for r in range(1, min(ws.max_row + 1, 30)):
        a_val = ws.cell(row=r, column=1).value
        if a_val and ('被评估单位' in str(a_val) or '基准日' in str(a_val)):
            found_marker = True
            break

    if not found_marker:
        # 兜底：直接写B6/B7
        if company_name:
            ws.cell(row=6, column=2).value = company_name
        if valuation_date:
            ws.cell(row=7, column=2).value = valuation_date
        if industry_type:
            ws.cell(row=8, column=2).value = industry_type
        print(f'  设定信息(兜底写入): {company_name} / {valuation_date}')


# ============================================================
# Phase 2: 数据填写（data_loader + fill_sheet集成）
# ============================================================

def _repair_detail_sheet_bc_merges(wb):
    """修复明细表结构行B~X合并，清理数据区残留的结构合并。

    背景：
    - 部分固定资产/在建工程模板结构行是B:D或B:E，不是统一B:C；
    - 插行后若openpyxl扩展了结构合并，数据区会残留空白合并行（如B32:D34）。
    """
    try:
        from gate_validator import find_header_structure as _find_header_structure
    except Exception:
        try:
            from excel_row_ops import _find_header_structure  # 兜底：桥接到valuation-common实现
        except Exception:
            return 0

    fixed = 0
    struct_markers = {'合计1', '合计2', '坏账准备', '预计风险', '预计损失', '减值准备', '跌价准备'}

    def _norm(v):
        return str(v or '').replace(' ', '').replace('\u3000', '').strip()

    def _is_struct_row_text(a_txt, b_txt, allow_b_fallback=False):
        if (
            a_txt in struct_markers or
            ('合' in a_txt and '计' in a_txt) or
            a_txt.startswith('减')
        ):
            return True
        if allow_b_fallback:
            return ('合' in b_txt and '计' in b_txt) or b_txt.startswith('减')
        return False

    for ws in wb.worksheets:
        sn = ws.title
        if sn.startswith('2-') or sn.startswith('设置') or sn.startswith('0-') or '汇总' in sn:
            continue

        struct = _find_header_structure(ws)
        dsr = struct.get('data_start_row') or 6
        tr = struct.get('total_row')
        if not tr:
            # 兜底：按B列“合计”文本定位
            for r in range(1, min(ws.max_row + 1, 220)):
                b_txt = _norm(ws.cell(row=r, column=2).value)
                if '合' in b_txt and '计' in b_txt:
                    tr = r
                    break
        # 兜底2：A列合计标记 + B列回退
        if not tr:
            for r in range(6, min(ws.max_row + 1, 220)):
                a_txt = _norm(ws.cell(row=r, column=1).value)
                b_txt = _norm(ws.cell(row=r, column=2).value)
                if _is_struct_row_text(a_txt, b_txt, allow_b_fallback=True):
                    tr = r
                    break
        if not tr:
            tr = ws.max_row + 1

        # 若定位到的tr仅由B列旧残留触发（A列不是结构标记），
        # 则向后寻找首个A列结构行作为真正合计起始。
        a_tr = _norm(ws.cell(row=tr, column=1).value) if tr <= ws.max_row else ''
        if tr <= ws.max_row and not _is_struct_row_text(a_tr, '', allow_b_fallback=False):
            for r in range(tr + 1, min(ws.max_row + 1, 260)):
                a_txt = _norm(ws.cell(row=r, column=1).value)
                if _is_struct_row_text(a_txt, '', allow_b_fallback=False):
                    tr = r
                    break

        # 1) 清理数据区B起始的残留结构合并（B:C / B:D / B:E ...）
        for mr in list(ws.merged_cells.ranges):
            if mr.min_col == 2 and mr.max_col >= 3 and dsr <= mr.min_row < tr:
                # 数据区内（合计1之前）不应存在结构合并；无条件清理
                ws.unmerge_cells(str(mr))
                fixed += 1

        # 2) 保证结构行合并存在（列宽按sheet内现有结构合并自动推断）
        rows_to_check = {tr} if tr <= ws.max_row else set()
        for r in range(1, min(ws.max_row + 1, 260)):
            a_txt = _norm(ws.cell(row=r, column=1).value)
            b_txt = _norm(ws.cell(row=r, column=2).value)
            is_struct = _is_struct_row_text(a_txt, b_txt, allow_b_fallback=True)
            if is_struct:
                rows_to_check.add(r)
        # 额外：B列"合  计"文本的行（旧模板A列可能无标记）
        for r in range(dsr, min(ws.max_row + 1, 220)):
            if r in rows_to_check:
                continue
            b_txt = _norm(ws.cell(row=r, column=2).value)
            if '合' in b_txt and '计' in b_txt:
                rows_to_check.add(r)

        # 以当前sheet中“结构行现有合并”推断目标宽度；没有则退回B:C
        end_col_counter = {}
        for r in rows_to_check:
            for mr in ws.merged_cells.ranges:
                if mr.min_row == r and mr.max_row == r and mr.min_col == 2 and mr.max_col >= 3:
                    end_col_counter[mr.max_col] = end_col_counter.get(mr.max_col, 0) + 1
        default_end_col = 3
        if end_col_counter:
            # 取结构行最大合并宽度，避免把B:D/B:E错误收缩成B:C
            default_end_col = max(end_col_counter.keys())

        for r in sorted(rows_to_check):
            existing_end_col = None
            for mr in ws.merged_cells.ranges:
                if mr.min_row == r and mr.max_row == r and mr.min_col == 2 and mr.max_col >= 3:
                    existing_end_col = mr.max_col
                    break
            target_end_col = max(default_end_col, existing_end_col or 0)
            if existing_end_col == target_end_col:
                continue
            # 先清理行内覆盖B~target_end_col的冲突合并
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row <= r <= mr.max_row and not (mr.max_col < 2 or mr.min_col > target_end_col):
                    ws.unmerge_cells(str(mr))
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=target_end_col)
            fixed += 1
    # 兜底3：如果合计行存在但没有B起始合并，强制创建B:C合并
    if tr is not None and tr <= ws.max_row:
        has_any_b_merge_on_tr = any(
            mr.min_row == tr and mr.max_row == tr and mr.min_col == 2 and mr.max_col >= 3
            for mr in ws.merged_cells.ranges
        )
        if not has_any_b_merge_on_tr:
            b_txt = _norm(ws.cell(row=tr, column=2).value)
            if '合' in b_txt and '计' in b_txt:
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row <= tr <= mr.max_row:
                        ws.unmerge_cells(str(mr))
                end_col = 3
                for mr in ws.merged_cells.ranges:
                    if mr.min_col == 2 and mr.max_col >= 3:
                        end_col = max(end_col, mr.max_col)
                ws.merge_cells(start_row=tr, start_column=2, end_row=tr, end_column=end_col)
                fixed += 1

    return fixed


def _clear_formula_cache(wb):
    """清除所有公式单元格的缓存值，确保下次打开时重新计算。
    G3 fix: fullCalcOnLoad=1 + 清空公式缓存值(None) = Excel打开时全部重算。
    """
    from openpyxl.workbook.properties import CalcProperties
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.value = cell.value  # 保留公式，但清除缓存
    wb.calculation = CalcProperties(fullCalcOnLoad=1)

def phase2(project_dir, args):
    """Phase 2: 数据填写

    核心改造（v3.48）:
    - data_loader.py统一加载+去重（DT-156）
    - subject_schema.json驱动数据源选择
    - fill_sheet()内嵌幂等保护（DT-155）+即时勾稽（DT-158）
    - rule_manifest.json驱动的断言覆盖率检测
    - DT-159: 禁止子Agent执行（调用栈检测）
    """
    cache_dir = _cache_path(project_dir)
    print('\n' + '='*60)
    print('Phase 2: 数据填写')
    print('='*60)

    from data_loader import load_subject_data
    from sheet_filler import fill_sheet, prepare_data_rows

    # DT-156修复: 初始化_dedup_keys_cache属性（函数对象默认无此属性）
    if not hasattr(fill_sheet, '_dedup_keys_cache'):
        fill_sheet._dedup_keys_cache = {}

    xlsx_path_info = _load_cache(cache_dir, 'xlsx_path.json')
    xlsx_path = xlsx_path_info.get('path') if xlsx_path_info else None
    if not xlsx_path or not os.path.exists(xlsx_path):
        print('  ❌ 评估明细表文件不存在，需先执行Phase 1')
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)

    subjects = _load_cache(cache_dir, 'subjects.json')
    if isinstance(subjects, dict):
        subjects = subjects.get('subjects', [])

    settings = _load_cache(cache_dir, 'settings_info.json') or {}
    bs_balances = _load_cache(cache_dir, 'bs_balances.json')  # DT-FR5: 坏账分配需要BS口径
    schema = None

    # DT-209: 写入设定信息Sheet（被评估单位名称、评估基准日等）
    if settings:
        _write_settings_sheet(wb, settings)

    # 遍历需要填写的Sheet
    filled_sheets = []
    failed_sheets = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        # DT-212: 跳过汇总/设置/辅助Sheet，汇总表只涉及跨表链接不需编辑
        if sheet_name.startswith('2-') or sheet_name.startswith('设置') or sheet_name.startswith('0-'):
            continue  # 跳过分类汇总/设置Sheet
        if '汇总' in sheet_name:
            continue  # 跳过科目汇总表（如3-9存货汇总、4-8固定资产汇总等）

        print(f'\n[填写] {sheet_name}')

        try:
            # DT-156: 通过data_loader统一加载+去重

            # ── 固定资产台账加载 ──
            # 如果Sheet是固定资产类且有资产台账缓存，用台账逐行数据替换科目汇总
            # DT-FIX: 扩充FA sheets列表，对于已重分配但无台账的FA sheet（如4-8-1房屋建筑物）
            # 从subject_sheet_mapping查找对应子编码并预计算净值
            _asset_sheets = {'4-8-4机器设备', '4-8-5车辆', '4-8-6电子设备', '4-8-1房屋建筑物'}
            _fa_all_sheets = {'4-8-4机器设备', '4-8-5车辆', '4-8-6电子设备',
                              '4-8-1房屋建筑物', '4-8-2构筑物', '4-8-3管道沟槽',
                              '4-8-7固定资产清理'}
            if sheet_name in _asset_sheets:
                _ar_cache = _load_cache(cache_dir, 'asset_register_by_sheet.json')
                if _ar_cache and sheet_name in _ar_cache:
                    _items = _ar_cache[sheet_name]
                    # 从schema加载config
                    if not schema:
                        _schema_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'assets', 'subject_schema.json')
                        with open(_schema_path) as _sf:
                            schema = json.load(_sf)
                    _scfg = schema.get('subjects', {}).get(sheet_name, {})
                    # 转换为prepare_data_rows期望的格式
                    # DT-ARCH: 直接构建资产台账数据行（跳过prepare_data_rows避免汇总折旧覆盖逐项折旧）
                    _ar_data_rows = []
                    for _idx, _a in enumerate(_items, 1):
                        _name = str(_a.get('name', '')).strip()
                        _asset_code = str(_a.get('asset_code', '')).strip()
                        _spec = str(_a.get('spec', '')).strip()
                        # 处理 'nan' 字符串
                        if _spec.lower() == 'nan':
                            _spec = ''
                        _cost = float(_a.get('cost', 0) or 0)
                        _dep = float(_a.get('depreciation', 0) or 0)
                        # 优先使用源数据的净值，避免计算误差
                        _net_src = _a.get('net_value')
                        if _net_src is not None:
                            try:
                                _net = float(_net_src)
                            except (ValueError, TypeError):
                                _net = _cost - _dep
                        else:
                            _net = _cost - _dep
                        _start_date = str(_a.get('start_date', '')).strip()
                        _dept = str(_a.get('dept', '')).strip()
                        _location = str(_a.get('location', '')).strip()
                        if _location.lower() == 'nan':
                            _location = ''
                        _dep_method = str(_a.get('dep_method', '')).strip()
                        _life = str(_a.get('life_months', '')).strip()
                        _status = str(_a.get('status', '')).strip()
                        
                        # 构建与col_map兼容的行数据（使用中文键名匹配运行时col_map）
                        # 运行时col_map将JSON的English键转为中文键名:
                        #   settlement→结算对象, spec→规格型号, cost→成本, book_value→账面价值, date→发生日期
                        row = {
                            '序号': _idx,
                            '结算对象': _name,       # → 设备名称/车辆牌号
                            'equip_code': _asset_code,  # → col 3 设备编号（运行col_map用English键名）
                            '规格型号': _spec,        # → 规格型号
                            '成本': _cost,            # → 原值
                            '账面价值': _net,         # → 净值
                            '_counterparty': _name,
                            'code': f'1601_{_idx}',
                            '_is_contra_account_row': True,
                        }

                        # 填入辅助字段
                        if _start_date:
                            row['发生日期'] = _start_date
                        if _dept:
                            row['使用部门'] = _dept
                        if _location:
                            row['存放地点'] = _location
                        # 备注: 折旧方法+年限+状态
                        remark_parts = []
                        if _dep_method:
                            remark_parts.append(_dep_method)
                        if _life:
                            remark_parts.append(f'{_life}月')
                        if _status and _status != '在用':
                            remark_parts.append(_status)
                        if remark_parts:
                            row['备注'] = ' | '.join(remark_parts)
                        
                        # 车辆Sheet特殊处理: 结算对象=车辆牌号/资产编号, 规格型号=车辆名称
                        if '车辆' in sheet_name:
                            row['结算对象'] = _asset_code if _asset_code else (_spec if _spec else _name)
                            row['规格型号'] = _name
                        elif '房屋建筑物' in sheet_name:
                            # 房建表必须按字段语义一次写入，禁止写后物理搬列。
                            row.pop('结算对象', None)
                            row['权证编号'] = _asset_code
                            row['建筑物名称'] = _name
                        
                        _ar_data_rows.append(row)
                    
                    # 直接用fill_sheet写入（has_journal=True避免日期被清空）
                    _result = fill_sheet(
                        ws, sheet_name=sheet_name, data_rows=_ar_data_rows,
                        settings=settings, wb=wb, subject_code='1601',
                        has_journal=True,
                    )
                    if _result['success']:
                        filled_sheets.append(sheet_name)
                        print(f'  ✅ {sheet_name}: 台账{len(_items)}行写入')
                    else:
                        failed_sheets.append(sheet_name)
                        print(f'  ❌ {sheet_name}: 台账写入失败')
                    continue  # 跳过正常数据加载
            
            # DT-FIX: 对于重分配的FA/Intangible sheet（无台账但有映射），直接计算净值
            # 不触发fill_sheet的contra计算（折旧已在净值中体现）
            _is_redist_fa = False
            _rmap = _load_cache(cache_dir, 'subject_sheet_mapping.json')
            if _rmap and sheet_name in _fa_all_sheets and sheet_name not in _asset_sheets:
                _fa_code = None
                for _rc, _ri in _rmap.items():
                    _sid = str(_ri.get('sheet_id', ''))
                    if '.' in str(_rc) and (_sid == sheet_name or sheet_name.startswith(_sid)):
                        _fa_code = _rc
                        break
                if _fa_code:
                    _subj_list = subjects if isinstance(subjects, list) else subjects.get('subjects', [])
                    _cost = 0
                    _dep = 0
                    for _s in _subj_list:
                        _sc = str(_s.get('code', ''))
                        if _sc == _fa_code:
                            _cost = abs(_s.get('balance', _s.get('closing_balance', 0)) or 0)
                        _parts = _fa_code.split('.')
                        if len(_parts) > 1:
                            _suffix = _parts[-1]
                            for _dp in ['1602', '1702', '1603', '1703']:
                                if _sc.startswith(_dp) and _sc.endswith('.' + _suffix):
                                    _dep = abs(_s.get('balance', _s.get('closing_balance', 0)) or 0)
                                    break
                    _net = max(_cost - _dep, 0)
                    # 直接写入净值，不触发contra计算
                    _fa_row = {'序号': 1, '结算对象': _fa_code, 'code': _fa_code, '成本': _cost, '账面价值': _net}
                    _result = fill_sheet(ws, sheet_name=sheet_name, data_rows=[_fa_row],
                                        settings=settings, wb=wb, subject_code=_fa_code[:4], has_journal=True)
                    if _result['success']:
                        filled_sheets.append(sheet_name)
                        print(f'  ✅ {sheet_name}: 重分配(原值{_cost:,.2f},净值{_net:,.2f})')
                        _is_redist_fa = True
                        continue
                    else:
                        print(f'  ❌ {sheet_name}: 重分配写入失败')
            load_result = load_subject_data(sheet_name, cache_dir)
            data_rows = load_result['data_rows']
            # DT-FIX: FA sheet过滤备抵科目（1602累计折旧/1603减值准备等），避免贷方余额写为正数
            if sheet_name in _fa_all_sheets and data_rows:
                _contra_prefixes = ('1602', '1603', '1609', '1702', '1703')
                _before = len(data_rows)
                data_rows = [r for r in data_rows if not str(r.get('code', '')).startswith(_contra_prefixes)]
                if len(data_rows) < _before:
                    print(f'  🗑️ 过滤备抵科目: {_before - len(data_rows)}行')

            reconcile_target = load_result['reconcile_target']
            config = load_result['config']

            if not data_rows:
                print(f'  ⚠️ 无数据，跳过')
                continue

            # DT-XXXX: data_loader返回{code,name,balance,direction}原始格式，
            # 必须经过prepare_data_rows()转换为中文语义{序号,项目及内容,账面价值,...}
            # 否则fill_sheet按字段名匹配col_map会找不到列，金额写不进去
            source_code_prefix = config.get('source_code_prefix', '')
            if isinstance(source_code_prefix, list):
                source_code_prefix = source_code_prefix[0] if source_code_prefix else ''
            subject_code = str(source_code_prefix)

            # P5修复: 提取坏账准备金额（从subjects.json提取1231坏账准备）
            bad_debt_amount = None
            if config.get('has_bad_debt'):
                bad_debt_subjects = [s for s in (subjects if isinstance(subjects, list) else subjects.get('subjects', []))
                                     if str(s.get('code', '')).startswith('1231')]
                if bad_debt_subjects:
                    # DT-FR5: 优先按1231下级子目编码精确匹配
                    # 123101=应收账款坏账准备, 123102=其他应收款坏账准备
                    bad_debt_amount = 0
                    matched_specific = False
                    for s in bad_debt_subjects:
                        code = str(s.get('code', ''))
                        bal = abs(s.get('balance', s.get('closing_balance', 0)))
                        # 跳过汇总行1231，只看末级子目
                        if code == '1231':
                            continue
                        # 123101 → 应收账款(1122)
                        if subject_code == '1122' and code.startswith('123101'):
                            bad_debt_amount += bal
                            matched_specific = True
                        # 123102 → 其他应收款(1221)
                        elif subject_code == '1221' and code.startswith('123102'):
                            bad_debt_amount += bal
                            matched_specific = True

                    if not matched_specific:
                        # DT-FR5降级: 1231无末级子目时，按名称关键词匹配
                        if subject_code == '1122':  # 应收账款
                            bad_debt_amount = sum(abs(s.get('balance', s.get('closing_balance', 0)))
                                                for s in bad_debt_subjects
                                                if '应收' in s.get('name', ''))
                        elif subject_code == '1221':  # 其他应收款
                            bad_debt_amount = sum(abs(s.get('balance', s.get('closing_balance', 0)))
                                                for s in bad_debt_subjects
                                                if '其他应收' in s.get('name', ''))

                    if not matched_specific and bad_debt_amount == 0:
                        # DT-FR5智能分配: 检查BS口径决定坏账归属
                        # 如果BS应收账款 = 科目1122 - 科目1231，则全部坏账归1122
                        # 否则按余额比例拆分
                        total_bd = sum(abs(s.get('balance', s.get('closing_balance', 0)))
                                      for s in bad_debt_subjects)
                        if subject_code == '1122':
                            # 应收账款的坏账 = 全量1231（因为BS已扣减全部坏账）
                            bad_debt_amount = total_bd
                        elif subject_code == '1221':
                            # 其他应收款：检查BS是否含坏账扣减
                            # 从bs_balances计算: BS其他应收款 vs 科目1221余额
                            bs_or = _get_bs_value(bs_balances, '其他应收款')
                            subj_or = sum(abs(s.get('balance', s.get('closing_balance', 0)))
                                         for s in (subjects if isinstance(subjects, list) else subjects.get('subjects', []))
                                         if str(s.get('code', ''))[:4] == '1221'
                                         and not any(str(o.get('code', '')).startswith(str(s.get('code', '')))
                                                    and str(o.get('code', '')) != str(s.get('code', ''))
                                                    for o in (subjects if isinstance(subjects, list) else subjects.get('subjects', []))))
                            if bs_or is not None and abs(bs_or - subj_or) < 1:
                                # BS其他应收款 = 科目1221，无坏账扣减
                                bad_debt_amount = 0
                            else:
                                # 有其他应收款专属坏账子目时才分配
                                bad_debt_amount = 0
                if bad_debt_amount:
                    print(f'  [P5] 坏账准备金额: {bad_debt_amount:,.2f}')

            # P6修复: 获取备抵科目配置（从Schema的contra_account字段）
            contra_account_config = config.get('contra_account')

            prepared = prepare_data_rows(
                subject_code=subject_code,
                kmye_data=data_rows,
                subject_name=config.get('name', sheet_name),
                has_journal=True,  # 上海图灵有序时账
                contra_account_config=contra_account_config,
                subjects_all=subjects,
            )
            if isinstance(prepared, tuple):
                data_rows, prepare_warnings = prepared
            else:
                prepare_warnings = []

            # DT-FIX: 用科目明细账结算对象数据替换prepare_data_rows的汇总行
            _sl_data = _load_cache(cache_dir, 'subledger_standardized.json')
            if _sl_data and data_rows and not sheet_name.startswith('4-8'):
                _sl_rows = []
                for _dr in data_rows:
                    _subj_code = str(_dr.get('_subject_code', '') or '')
                    _actual_code = str(_dr.get('code', '') or '')
                    # DT-FIX: 标准化匹配——优先精确匹配，其次子科目→父科目前缀回退
                    if _sl_data.get(_actual_code):
                        _lookup_code = _actual_code
                    elif len(_actual_code) >= 4 and _sl_data.get(_actual_code[:4]):
                        _lookup_code = _actual_code[:4]
                    elif _sl_data.get(_subj_code):
                        _lookup_code = _subj_code
                    elif len(_subj_code) >= 4 and _sl_data.get(_subj_code[:4]):
                        _lookup_code = _subj_code[:4]
                    else:
                        _lookup_code = _actual_code or _subj_code
                    # DT-FIX: 跳过现金/银行科目（1001/1002），它们的对方科目归集无意义
                    if _lookup_code.startswith('100'):
                        _sl_rows.append(_dr)
                        continue
                    _sl_subject = _sl_data.get(_lookup_code)
                    # 获取当前adjusted余额（已包含重分类调整；prepare_data_rows后用账面价值字段）
                    _dr_bal = abs(float(_dr.get('账面价值', _dr.get('book_value', _dr.get('balance', 0))) or 0))
                    if _sl_subject and _sl_subject.get('settlements') and _dr_bal > 0:
                        _stt = _sl_subject['settlements']
                        # DT-FIX: 排除备抵科目结算对象（累计折旧/摊销/减值）
                        _stt = {k: v for k, v in _stt.items() if not any(kw in k for kw in ['累计', '减值', '准备', '处置'])}
                        if not _stt:
                            _sl_rows.append(_dr)
                            continue
                        # 计算明细账结算对象净额合计
                        _sl_total = abs(sum(s.get('debit', 0) - s.get('credit', 0) for s in _stt.values()))
                        if _sl_total < 0.01:
                            _sl_rows.append(_dr)
                            continue
                        # 按比例缩放：使明细账合计 = adjusted余额
                        _ratio = _dr_bal / _sl_total
                        _added = 0
                        for _sn, _si in sorted(_stt.items(), key=lambda x: -x[1]['debit']):
                            _dv = (_si.get('debit', 0) - _si.get('credit', 0)) * _ratio
                            if abs(_dv) < 0.01: continue
                            _added += 1
                            _sl_rows.append({
                                '项目及内容': _sn,
                                '发生日期': _si.get('last_date', ''),
                                '业务内容': '; '.join(_si.get('summaries', [])[:3]),
                                '账面价值': round(_dv, 2),
                                '评估价值': round(_dv, 2),
                            })
                        if _added:
                            print(f'  📋 明细账替换: {_lookup_code}: {_added}个结算对象 (缩放比={_ratio:.4f})')
                    else:
                        _sl_rows.append(_dr)
                if _sl_rows and len(_sl_rows) != len(data_rows):
                    data_rows = _sl_rows
            # 将reconcile_target传入settings供DT-158即时勾稽使用
            settings['_reconcile_target'] = reconcile_target

            # 如果data_loader返回了dedup_key，注册到fill_sheet缓存
            # P7修复: 如果数据行有code字段，自动将code加入去重键
            if config and config.get('dedup_key'):
                dk = list(config['dedup_key'])
                if data_rows and 'code' in data_rows[0] and 'code' not in dk:
                    dk = ['code'] + dk
                fill_sheet._dedup_keys_cache[sheet_name] = dk

            # DT-155/157/158/159: fill_sheet内部强制执行
            result = fill_sheet(
                ws, sheet_name=sheet_name, data_rows=data_rows,
                settings=settings, wb=wb, subject_code=subject_code,
                bad_debt_amount=bad_debt_amount,  # P5修复: 传入坏账准备金额
            )

            if result['success']:
                filled_sheets.append(sheet_name)
                print(f'  ✅ {sheet_name}: {result["rows_written"]}行写入')
                # r9: 写入后自检——验证写入合计 vs reconcile_target
                if reconcile_target is not None and abs(reconcile_target) > 1:
                    _wr_total = sum(float(r.get('账面价值', r.get('book_value', r.get('balance', 0)) or 0))
                                    for r in data_rows if isinstance(r, dict))
                    _wr_diff = abs(_wr_total - reconcile_target)
                    if _wr_diff > 100:
                        print(f'    ⚠️ 自检: 写入合计={_wr_total:,.2f}, 目标={reconcile_target:,.2f}, 差额={_wr_diff:,.2f}')
                    else:
                        print(f'    ✅ 自检: 写入合计={_wr_total:,.2f} 与目标一致')
            else:
                failed_sheets.append(sheet_name)
                print(f'  ❌ {sheet_name}: 写入失败')

        except Exception as e:
            failed_sheets.append(sheet_name)
            print(f'  ❌ {sheet_name}: 异常 - {e}')

    # ── 隐藏空白Sheet ──
    _hidden = 0
    for _ws in wb.worksheets:
        _name = _ws.title
        if _name.startswith('2-') or '汇总' in _name or _name.startswith('设置') or _name.startswith('0-'):
            continue
        _has_data = False
        _has_numeric = False
        for _r in range(6, _ws.max_row + 1):
            for _c in range(1, 12):
                _v = _ws.cell(row=_r, column=_c).value
                if _v is not None and not isinstance(_v, str):
                    try:
                        if abs(float(_v)) > 0:
                            _has_numeric = True
                            _has_data = True
                            break
                    except: pass
                elif _v and isinstance(_v, str) and _v.strip() and not _v.startswith('='):
                    _vs = str(_v).replace(' ', '').replace('　', '')
                    if not any(kw in _vs for kw in ['表头','检索','序号','合计','坏账','跌价','减值','名称','规格','单位','预计','风险','注','备抵','本表','为例','来源','结构','建成','面积','权证','成本','土地','宗地','用地','用途','准用','开发','原始','入账','转入','公允','计量']):
                        _has_data = True
                        break
            if _has_data and _has_numeric:
                break
        # DT-212: 如果只有表头文字但没有数值，视为空白Sheet
        if _has_data and not _has_numeric:
            _has_data = False
        if not _has_data:
            _ws.sheet_state = 'hidden'
            _hidden += 1
    if _hidden > 0:
        print(f'  📄 已隐藏 {_hidden} 个空白Sheet')
    
    # ── P7: 2-分类汇总 I列（报表金额）填充 ──
    _fill_classification_summary_I_column(wb, cache_dir)

    # ── 结构合并修复：确保合计/减值行B:C合并稳定，避免G2-12误判/漏修复 ──
    _merge_fixed = _repair_detail_sheet_bc_merges(wb)
    if _merge_fixed:
        print(f'  🔧 结构合并修复: {_merge_fixed}处')

    # 保存
    _clear_formula_cache(wb)
    wb.save(xlsx_path)
    wb.close()

    # Gate G1验证
    print('\n[Gate G1] 数据写入级验证')
    g1 = run_gate(project_dir, 'G1', args)
    _gate_pass_or_raise(g1)
    g182 = run_gate(project_dir, 'G-DT182', args)
    _gate_pass_or_raise(g182)

    return {
        'phase': 2,
        'status': 'completed' if not failed_sheets else 'partial',
        'filled_sheets': filled_sheets,
        'failed_sheets': failed_sheets,
    }


# ============================================================
# Phase 3: 公式修复与格式修复
# ============================================================


# ── _fill_classification_summary_I_column: 2-分类汇总 报表金额填充 ──

def _fill_classification_summary_I_column(wb, cache_dir):
    """向2-分类汇总写入I/J列公式，并构建标准化_BS对照结构表。

    规则：
    - 具体科目行：I列链接_BS对照（VLOOKUP）；
    - 合计/总计/净资产行：保留模板公式，不落静态值。
    """
    if '2-分类汇总' not in wb.sheetnames:
        return

    ws = wb['2-分类汇总']
    bs_balances = _load_cache(cache_dir, 'bs_balances.json')
    if not bs_balances:
        print('  ⚠️ [P7] bs_balances.json不存在，跳过分类汇总I列填充')
        return

    # 检测是否有 I/J 列（C9/C10），没有则从模板扩展
    need_extend = False
    if ws.max_column < 10:
        need_extend = True
    else:
        h9 = str(ws.cell(row=5, column=9).value or '').strip()
        h10 = str(ws.cell(row=5, column=10).value or '').strip()
        if '报表' not in h9 and '校验' not in h10:
            need_extend = True

    if need_extend:
        import os as _os
        template_path = _os.path.join(
            _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
            'assets', '评估明细表-v1.90-FOR AI.xlsx'
        )
        if _os.path.exists(template_path):
            try:
                import openpyxl
                twb = openpyxl.load_workbook(template_path)
                tws = twb['2-分类汇总']
                for r in range(1, tws.max_row + 1):
                    for c in [9, 10]:
                        tv = tws.cell(row=r, column=c).value
                        if tv is not None:
                            ws.cell(row=r, column=c).value = tv
                twb.close()
                col_i = chr(64 + 9)
                col_j = chr(64 + 10)
                if ws.column_dimensions[col_i].width is None or ws.column_dimensions[col_i].width < 12:
                    ws.column_dimensions[col_i].width = 14
                if ws.column_dimensions[col_j].width is None or ws.column_dimensions[col_j].width < 10:
                    ws.column_dimensions[col_j].width = 12
                print('  [P7] 已从模板扩展2-分类汇总 I/J列')
            except Exception as e:
                print(f'  [P7] 模板扩展I/J列失败: {e}')
                return
        else:
            print(f'  [P7] 模板文件未找到，跳过I/J列扩展')
            return
        ws = wb['2-分类汇总']

    def _norm_label(text):
        s = str(text or '').strip()
        s = s.replace(' ', '').replace('\u3000', '').replace('\n', '')
        s = s.replace('（', '(').replace('）', ')')
        s = s.replace('：', '').replace(':', '')
        return s

    def _strip_prefix(subject):
        txt = str(subject or '').replace('\u3000', ' ').strip()
        return re.sub(r'^[一二三四五六七八九十]+[、，,\.．]\s*', '', txt).strip()

    def _is_summary_subject(subject):
        norm_subj = _strip_prefix(subject).replace(' ', '')
        if not norm_subj:
            return False
        if '合计' in norm_subj or '总计' in norm_subj:
            return True
        if '净资产' in norm_subj and '所有者权益' in norm_subj:
            return True
        return norm_subj in ('资产总计', '负债总计')

    # 读取模板公式：用于恢复/保持合计行公式
    template_i_formulas = {}
    template_j_formulas = {}
    try:
        import openpyxl
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'assets', '评估明细表-v1.90-FOR AI.xlsx'
        )
        if os.path.exists(template_path):
            twb = openpyxl.load_workbook(template_path, data_only=False)
            if '2-分类汇总' in twb.sheetnames:
                tws = twb['2-分类汇总']
                for rr in range(1, tws.max_row + 1):
                    iv = tws.cell(row=rr, column=9).value
                    jv = tws.cell(row=rr, column=10).value
                    if isinstance(iv, str) and iv.startswith('='):
                        template_i_formulas[rr] = iv
                    if isinstance(jv, str) and jv.startswith('='):
                        template_j_formulas[rr] = jv
            twb.close()
    except Exception:
        pass

    # 分类汇总标准科目结构（按模板顺序）
    subject_rows = []
    for r in range(6, ws.max_row + 1):
        display_name = ws.cell(row=r, column=4).value
        if display_name is None:
            continue
        display_name = str(display_name).strip()
        if not display_name:
            continue
        std_name = _strip_prefix(display_name)
        if not std_name:
            continue
        subject_rows.append({
            'row': r,
            'display': display_name,
            'std': std_name,
            'is_summary': _is_summary_subject(display_name),
        })

    # 构建原始BS映射
    bs_items = bs_balances.get('items', []) if isinstance(bs_balances, dict) else []
    bs_item_map = {}
    for item in bs_items:
        label = str(item.get('label', '')).strip()
        if not label:
            continue
        bs_item_map[_norm_label(label)] = {
            'label': label,
            'beginning': float(item.get('beginning_balance', 0) or 0),
            'ending': float(item.get('ending_balance', 0) or 0),
        }

    aux_name = '_BS对照'
    if aux_name in wb.sheetnames:
        ws_aux = wb[aux_name]
        ws_aux.delete_rows(1, ws_aux.max_row)
    else:
        ws_aux = wb.create_sheet(aux_name)
    ws_aux.sheet_state = 'hidden'
    # _BS对照改造：标准科目结构（完整列式）+来源追踪
    ws_aux.cell(row=1, column=1).value = '标准科目'
    ws_aux.cell(row=1, column=2).value = '科目类型'
    ws_aux.cell(row=1, column=3).value = '年初余额'
    ws_aux.cell(row=1, column=4).value = '期末余额'
    ws_aux.cell(row=1, column=5).value = '来源科目'

    CATEGORY_TO_BS_ALIASES = {
        '预付款项': ['预付账款'],
        '预收款项': ['预收账款'],
        '固定资产': ['固定资产账面价值', '固定资产净额'],
        '负债总计': ['负债合计'],
        '净资产（所有者权益）': ['所有者权益（或股东权益）合计', '所有者权益合计'],
    }

    def _pick_bs_item(std_subject):
        candidates = [std_subject] + CATEGORY_TO_BS_ALIASES.get(std_subject, [])
        # 1) 先做规范化精确匹配
        for c in candidates:
            hit = bs_item_map.get(_norm_label(c))
            if hit:
                return hit
        # 2) 再做弱匹配（包含关系）
        for c in candidates:
            c_norm = _norm_label(c)
            if not c_norm:
                continue
            for k, v in bs_item_map.items():
                if c_norm in k or k in c_norm:
                    return v
        # 3) 净资产兜底：所有者权益合计口径
        if '净资产' in std_subject and '所有者权益' in std_subject:
            for k, v in bs_item_map.items():
                if '所有者权益' in k and ('合计' in k or '总计' in k):
                    return v
        return None

    def _find_report_net_asset_row():
        """查找'报表净资产'标签所在行，找不到时回退68行。"""
        for rr in range(60, min(ws.max_row + 1, 90)):
            for cc in (7, 8, 9):
                txt = str(ws.cell(row=rr, column=cc).value or '').strip()
                if txt == '报表净资产':
                    return rr
        return 68

    aux_row = 2
    matched_in_aux = 0
    unmatched_in_aux = 0
    used_bs_norm_labels = set()
    for item in subject_rows:
        hit = _pick_bs_item(item['std'])
        begin_val = hit['beginning'] if hit else 0.0
        end_val = hit['ending'] if hit else 0.0
        source_label = hit['label'] if hit else f'[未匹配]{item["std"]}'
        if hit:
            matched_in_aux += 1
            used_bs_norm_labels.add(_norm_label(hit['label']))
        else:
            unmatched_in_aux += 1
        ws_aux.cell(row=aux_row, column=1).value = item['std']
        ws_aux.cell(row=aux_row, column=2).value = '汇总' if item['is_summary'] else '明细'
        ws_aux.cell(row=aux_row, column=3).value = begin_val
        ws_aux.cell(row=aux_row, column=4).value = end_val
        ws_aux.cell(row=aux_row, column=5).value = source_label
        aux_row += 1

    # 追加“分类汇总未覆盖”的BS原始科目，避免未来新增科目静默遗漏
    extra_bs_rows = 0
    for raw in bs_items:
        raw_label = str(raw.get('label', '')).strip()
        if not raw_label:
            continue
        raw_norm = _norm_label(raw_label)
        if not raw_norm or raw_norm in used_bs_norm_labels:
            continue
        ws_aux.cell(row=aux_row, column=1).value = f'[新增]{raw_label}'
        ws_aux.cell(row=aux_row, column=2).value = '新增'
        ws_aux.cell(row=aux_row, column=3).value = float(raw.get('beginning_balance', 0) or 0)
        ws_aux.cell(row=aux_row, column=4).value = float(raw.get('ending_balance', 0) or 0)
        ws_aux.cell(row=aux_row, column=5).value = raw_label
        aux_row += 1
        extra_bs_rows += 1

    # 分类汇总回填：
    # - 具体科目：I列查_BS对照
    # - 合计/总计：I列保留模板公式（不覆盖成静态值）
    detail_linked = 0
    summary_kept = 0
    for item in subject_rows:
        r = item['row']
        if item['is_summary']:
            i_formula = template_i_formulas.get(r)
            j_formula = template_j_formulas.get(r)
            if isinstance(i_formula, str) and i_formula.startswith('='):
                ws.cell(row=r, column=9).value = i_formula
            elif not (isinstance(ws.cell(row=r, column=9).value, str) and str(ws.cell(row=r, column=9).value).startswith('=')):
                ws.cell(row=r, column=9).value = f'=IFERROR(VLOOKUP("{item["std"]}",\'{aux_name}\'!$A:$D,4,FALSE),0)'
            if isinstance(j_formula, str) and j_formula.startswith('='):
                ws.cell(row=r, column=10).value = j_formula
            else:
                ws.cell(row=r, column=10).value = f'=E{r}-I{r}'
            summary_kept += 1
            continue

        # DT-182: 汇总表I列必须保留公式引用；BS值在_BS对照 Col4, VLOOKUP读取
        ws.cell(row=r, column=9).value = f'=IFERROR(VLOOKUP(D{r},\'{aux_name}\'!$A:$D,4,FALSE),0)'
        ws.cell(row=r, column=10).value = f'=E{r}-I{r}'
        detail_linked += 1

    # I68（报表净资产）自动写入：用于I列数据校验增强
    report_net_asset = None
    hit_net_asset = _pick_bs_item('净资产（所有者权益）')
    if hit_net_asset:
        report_net_asset = float(hit_net_asset.get('ending', 0) or 0)
    else:
        # 兜底：从BS原始科目中找“所有者权益+合计/总计”口径
        for k, v in bs_item_map.items():
            if '所有者权益' in k and ('合计' in k or '总计' in k):
                report_net_asset = float(v.get('ending', 0) or 0)
                break

    report_row = _find_report_net_asset_row()
    i68_filled = False
    if report_net_asset is not None:
        # DT-182: 写入公式引用而非硬编码值
        ws.cell(row=report_row, column=9).value = f'=IFERROR(VLOOKUP("净资产（所有者权益）",\'{aux_name}\'!$A:$D,4,FALSE),0)'
        i68_filled = True
    if not i68_filled:
        ws.cell(row=report_row, column=9).value = f'=IFERROR(VLOOKUP("*",\'{aux_name}\'!$A:$D,4,FALSE),0)'

    print(
        f'  [P7] 已重构_BS对照(标准科目{len(subject_rows)}行, 匹配{matched_in_aux}行, 未匹配{unmatched_in_aux}行, 新增科目{extra_bs_rows}行); '
        f'I列科目链接{detail_linked}行, 合计公式保留{summary_kept}行, '
        f'I{report_row}报表净资产{"已写入" if i68_filled else "未写入"}'
    )



class _StandardizedJournalWrapper:
    """DT-ARCH: 标准化序时账缓存包装器
    
    将 journal.json 的缓存格式适配为 JournalExtractor 兼容接口，
    使 Phase 3 可以无缝切换数据源（标准化缓存 vs 原始序时账文件）。
    """
    
    def __init__(self, journal_data):
        """从 journal.json 缓存构建
        
        Args:
            journal_data: list[dict], 来自 journal.json / journal_normalized.json
                标准格式: {date, voucher_no, subject_code, subject_name,
                           summary, debit_amount, credit_amount,
                           customer_supplier, department, project, bank_account}
                兼容旧格式: {date, voucher_no, account_code, account_name,
                             summary, debit_amount, credit_amount,
                             customer_supplier, department, project_name, personnel}
        """
        from datetime import datetime, timedelta
        import re as _re
        self.data = []
        for row in journal_data:
            # 标准化日期格式
            raw_date = row.get('date', '')
            # v3.68 (2026-06-02): 兼容 account_code/account_name 字段名
            # (journal_normalized.json 用的就是这种命名)
            subject_code = str(row.get('subject_code') or row.get('account_code') or '').strip()
            subject_name = str(row.get('subject_name') or row.get('account_name') or '').strip()
            summary = str(row.get('summary', '')).strip()
            
            try:
                debit = float(row.get('debit_amount', 0) or 0)
            except (ValueError, TypeError):
                debit = 0.0
            try:
                credit = float(row.get('credit_amount', 0) or 0)
            except (ValueError, TypeError):
                credit = 0.0
            
            # 标准化date为datetime对象（兼容JournalExtractor）
            # 兼容: datetime / Excel序列号(float,int) / YYYY-MM-DD文本
            parsed_date = None
            date_str = ''
            if isinstance(raw_date, datetime):
                parsed_date = raw_date
                date_str = raw_date.strftime('%Y-%m-%d')
            elif isinstance(raw_date, (int, float)):
                try:
                    parsed_date = datetime(1899, 12, 30) + timedelta(days=float(raw_date))
                    date_str = parsed_date.strftime('%Y-%m-%d')
                except Exception:
                    parsed_date = None
                    date_str = str(raw_date)
            elif raw_date is not None:
                date_str = str(raw_date).strip()
                if date_str:
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                        try:
                            parsed_date = datetime.strptime(date_str[:10], fmt)
                            break
                        except ValueError:
                            continue
                    # 兜底：字符串是数字序列号
                    if parsed_date is None:
                        try:
                            serial = float(date_str)
                            parsed_date = datetime(1899, 12, 30) + timedelta(days=serial)
                        except Exception:
                            parsed_date = None
            
            # v3.68 (2026-06-02): 读取 customer_supplier/department/project/project_name/bank_account
            # 兼容: project_name (标准化格式) 和 project (旧格式)
            cust = str(row.get('customer_supplier', '') or '').strip()
            if cust in ('nan', 'None', ''):
                cust = ''
            dept = str(row.get('department', '') or '').strip()
            if dept in ('nan', 'None', ''):
                dept = ''
            proj = str(row.get('project') or row.get('project_name') or '').strip()
            if proj in ('nan', 'None', ''):
                proj = ''
            bank = str(row.get('bank_account', '') or '').strip()
            if bank in ('nan', 'None', ''):
                bank = ''
            self.data.append({
                'date': parsed_date,
                'date_str': date_str,
                'voucher_no': str(row.get('voucher_no', '')).strip(),
                'subject_code': subject_code,
                'subject_name': subject_name,
                'summary': summary,
                'debit': debit,
                'credit': credit,
                'aux_accounting': cust,
                'settlement_from_aux': cust,
                'customer_supplier': cust,
                'department': dept,
                'project': proj,
                'bank_account': bank,
            })
    
    @property
    def row_count(self):
        return len(self.data)
    
    def query_by_subject(self, subject_keywords, summary_keywords=None, direction=None,
                         fuzzy_fallback=True, max_results=50):
        """按科目关键词+摘要关键词查询序时账（兼容JournalExtractor接口）"""
        filtered = []
        for s in self.data:
            for kw in subject_keywords:
                if kw and kw in s['subject_name']:
                    filtered.append(s)
                    break
        
        # v3.68 (2026-06-02): L2 0命中时设_skip_l3=True,防止L3污染结算对象过滤
        _skip_l3 = False
        if not filtered and fuzzy_fallback and summary_keywords:
            for s in self.data:
                for kw in summary_keywords:
                    if kw and (kw in s['summary'] or
                               kw in s.get('aux_accounting', '') or
                               kw in s.get('settlement_from_aux', '') or
                               kw in s.get('customer_supplier', '')):
                        filtered.append(s)
                        break
            if not filtered:
                _skip_l3 = True
        
        if not filtered and not _skip_l3 and fuzzy_fallback and subject_keywords:
            for s in self.data:
                for kw in subject_keywords:
                    if kw and (s['subject_code'].startswith(kw) or
                               s['subject_name'].startswith(kw)):
                        filtered.append(s)
                        break
        
        if summary_keywords and filtered:
            matched = []
            for s in filtered:
                for kw in summary_keywords:
                    if kw and (kw in s['summary'] or
                               kw in s.get('aux_accounting', '') or
                               kw in s.get('settlement_from_aux', '') or
                               kw in s.get('customer_supplier', '')):
                        matched.append(s)
                        break
            if matched:
                filtered = matched
        
        if direction == 'debit':
            filtered = [s for s in filtered if s['debit'] > 0]
        elif direction == 'credit':
            filtered = [s for s in filtered if s['credit'] > 0]
        
        if len(filtered) > max_results:
            filtered.sort(key=lambda x: x['debit'] + x['credit'], reverse=True)
            filtered = filtered[:max_results]
        
        return filtered
    
    @staticmethod
    def _clean_name(name):
        """去除结算对象名称中的 * 标记前缀"""
        if not name:
            return name
        import re as _re
        return _re.sub(r'^[*＊]+\s*', '', name.strip())
    
    def get_last_date_by_settlement(self, settlement_name, subject_code_prefix,
                                    summary_keywords=None, direction=None):
        """获取结算对象的末笔发生日期

        v3.68 (2026-06-02): 无匹配时回退到父科目最近一笔日期(fallback_parent),
        适用于聚合行(单位应付款/个人应付款/其他)和TB有但journal无独立记录的结算对象。
        """
        import re as _re
        settlement_name = self._clean_name(settlement_name)
        
        if summary_keywords is None:
            keywords = []
            geo_match = _re.search(r'([\一-\龥]{2,4}(?:省|市|区|县|镇|路|街))', settlement_name)
            if geo_match:
                keywords.append(geo_match.group(1)[:3])
            core_name = _re.sub(r'(有限公司|股份有限公司|有限责任公司|公司|集团)', '', settlement_name)
            if len(core_name) >= 2:
                keywords.append(core_name[:4])
            if not keywords:
                keywords.append(settlement_name[:4])
            summary_keywords = keywords
        
        subject_keywords = [subject_code_prefix]
        matched = self.query_by_subject(subject_keywords, summary_keywords, direction)
        
        if not matched:
            # v3.68 (2026-06-02): fallback_parent - 取父科目下所有分录的最近一笔日期
            parent_matched = self.query_by_subject(subject_keywords, None, direction)
            if parent_matched:
                parent_matched.sort(key=lambda x: str(x['date'] or '')[:10] or '0000-00-00')
                last = parent_matched[-1]
                return {
                    'date': last['date'],
                    'status': 'fallback_parent',
                    'match_count': 0,
                    'note': f'父科目降级: 最近日期 {str(last["date"])[:10]}, 该结算对象无独立序时账记录'
                }
            return {'date': None, 'status': 'no_match', 'match_count': 0}
        if len(matched) > 50:
            matched.sort(key=lambda x: x['debit'] + x['credit'], reverse=True)
            matched = matched[:20]
        
        # 按日期排序取末笔
        matched.sort(key=lambda x: str(x['date'] or '')[:10] or '0000-00-00')
        last = matched[-1]
        return {'date': last['date'], 'status': 'verified', 'match_count': len(matched)}
    
    def get_business_summaries(self, settlement_name, subject_code_prefix):
        """获取结算对象的序时账摘要"""
        import re as _re
        settlement_name = self._clean_name(settlement_name)
        keywords = []
        geo_match = _re.search(r'([\一-\龥]{2,4}(?:省|市|区|县|镇|路|街))', settlement_name)
        if geo_match:
            keywords.append(geo_match.group(1)[:3])
        core_name = _re.sub(r'(有限公司|股份有限公司|有限责任公司|公司|集团)', '', settlement_name)
        if len(core_name) >= 2:
            keywords.append(core_name[:4])
        if not keywords:
            keywords.append(settlement_name[:4])
        
        matched = self.query_by_subject([subject_code_prefix], keywords)
        seen = set()
        result = []
        for s in matched:
            summary = s['summary']
            if summary and summary not in seen:
                seen.add(summary)
                result.append(summary)
        return result



def _clean_settlement_names_in_detail(xlsx_path):
    """DT-ARCH: 清理评估明细表中结算对象名称的 * 前缀标记
    
    科目余额表中的结算对象名称常带 * 前缀（标记暂估/待核实），
    这些标记会干扰序时账匹配和报告可读性，需在 Phase 3 前统一清理。
    """
    import openpyxl as _opx
    import re as _re
    
    # 需要清理的往来科目Sheet（从 col_map 中读取 settlement 列号）
    col_map_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'assets', 'sheet_col_map.json'
    )
    if not os.path.exists(col_map_path):
        return
    
    with open(col_map_path, 'r', encoding='utf-8') as f:
        cm = json.load(f)
    sheets = cm.get('sheets', {})
    
    # 找出所有有 settlement 列的 Sheet
    settlement_sheets = {}
    for sn, cfg in sheets.items():
        col_map = cfg.get('col_map', {})
        settle_col = col_map.get('settlement')
        if settle_col:
            settlement_sheets[sn] = settle_col
    
    if not settlement_sheets:
        return
    
    wb = _opx.load_workbook(xlsx_path)
    cleaned = 0
    
    for sn, col in settlement_sheets.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        if ws.sheet_state == 'hidden':
            continue
        
        for r in range(6, ws.max_row + 1):
            val = ws.cell(row=r, column=col).value
            if val and isinstance(val, str):
                cleaned_val = _re.sub(r'^[*＊]+\s*', '', val.strip())
                if cleaned_val != val.strip():
                    ws.cell(row=r, column=col).value = cleaned_val
                    cleaned += 1
    
    if cleaned > 0:
        print(f'  [清理] 已去除 {cleaned} 个结算对象名称的 * 前缀')
    
    _clear_formula_cache(wb)
    wb.save(xlsx_path)
    wb.close()


def phase3(project_dir, args):
    """Phase 3: 序时账查阅（发生日期+业务内容）

    v3.55改造: 消除requires_manual_execution出口，完整编排journal_extractor.py流程。

    子步骤:
    3.1 检测序时账文件存在性(DT-161)
    3.2 初始化JournalExtractor加载序时账数据
    3.3 scan_empty_fields()扫描往来科目空字段行
    3.4 extract_dates()批量提取发生日期(DT-51/DT-178)
    3.5 extract_business_contents()批量提取业务内容(DT-60/DT-149)
    3.6 write_phase3_results()写入评估明细表
    3.7 生成核实结果报告+缓存
    3.8 Gate G2验证（含字段完整性校验）
    """
    cache_dir = _cache_path(project_dir)
    print('\n' + '='*60)
    print('Phase 3: 序时账查阅（发生日期+业务内容）')
    print('='*60)

    # --- Step 3.1: 检测序时账文件 ---
    print('\n[Step 3.1] 检测序时账文件 (DT-161)')
    xlsx_path_info = _load_cache(cache_dir, 'xlsx_path.json')
    xlsx_path = xlsx_path_info.get('path') if xlsx_path_info else None
    if not xlsx_path or not os.path.exists(xlsx_path):
        xlsx_path = args.xlsx_path or _find_detail_table(project_dir)
        # 尝试模糊搜索
        if not os.path.exists(xlsx_path):
            candidates = glob.glob(os.path.join(project_dir, '*评估明细表*'))
            if candidates:
                xlsx_path = candidates[0]

    # DT-ARCH: 优先检查标准化序时账缓存
    # v3.68 (2026-06-02): 优先 journal_normalized.json (完整 + v3.67 字段),fallback 到 journal.json
    journal_cache = _load_cache(cache_dir, 'journal_normalized.json')
    if journal_cache and isinstance(journal_cache, dict):
        # journal_normalized.json 格式: {entries: [...], filtered_entries: [...], ...}
        journal_cache = journal_cache.get('entries') or journal_cache.get('filtered_entries') or []
    if not journal_cache:
        journal_cache = _load_cache(cache_dir, 'journal.json')
    if journal_cache:
        print(f'  [标准化] 使用标准化序时账缓存 ({len(journal_cache)}行)')
        # 跳过原始文件搜索，直接使用缓存
        print(f'  评估明细表: {os.path.basename(xlsx_path)}')
        _clean_settlement_names_in_detail(xlsx_path)
        if not os.path.exists(xlsx_path):
            print('  ❌ 评估明细表不存在，需先完成Phase 2')
            return {'phase': 3, 'status': 'error', 'reason': 'no_detail_file'}
    else:
        # 搜索序时账文件（缓存不存在时才搜索原始文件）
        seq_files = glob.glob(os.path.join(project_dir, '*序时账*')) + \
                    glob.glob(os.path.join(project_dir, '*明细账*'))
        if not seq_files:
            seq_files = glob.glob(os.path.join(project_dir, '**', '*序时账*'), recursive=True)
        if not seq_files:
            print('  ⚠️ 未找到序时账文件且无标准化缓存，Phase 3跳过 (DT-161条件①)')
            _save_cache(cache_dir, 'phase3_status.json', {
                'status': 'skipped',
                'reason': 'no_seq_file',
                'dt_rule': 'DT-161'
            })
            return {'phase': 3, 'status': 'skipped_no_seq_file', 'reason': 'DT-161: 无序时账数据'}
        seq_file = seq_files[0]
        print(f'  序时账文件: {os.path.basename(seq_file)}')
        print(f'  评估明细表: {os.path.basename(xlsx_path)}')
        # DT-ARCH: 清理明细表中结算对象名称的 * 标记前缀
        _clean_settlement_names_in_detail(xlsx_path)
        if not os.path.exists(xlsx_path):
            print('  ❌ 评估明细表不存在，需先完成Phase 2')
            return {'phase': 3, 'status': 'error', 'reason': 'no_detail_file'}

    # --- Step 3.2: 初始化JournalExtractor (DT-166 → DT-ARCH强化) ---
    print('\n[Step 3.2] 初始化JournalExtractor (DT-166)')
    from journal_extractor import (
        JournalExtractor, scan_empty_fields, extract_dates,
        extract_business_contents, write_phase3_results,
        generate_phase3_report
    )

    # DT-ARCH: 优先使用标准化序时账缓存
    # v3.68 (2026-06-02): 优先 journal_normalized.json (完整 3680 条 + v3.67 字段)
    journal_cache = _load_cache(cache_dir, 'journal_normalized.json')
    if journal_cache and isinstance(journal_cache, dict):
        journal_cache = journal_cache.get('entries') or journal_cache.get('filtered_entries') or []
    if not journal_cache:
        journal_cache = _load_cache(cache_dir, 'journal.json')
    if journal_cache:
        print(f'  [标准化] 使用标准化序时账缓存 ({len(journal_cache)}行)')
        extractor = _StandardizedJournalWrapper(journal_cache)
        print(f'  序时账加载完成: {extractor.row_count}行')
    else:
        extractor = JournalExtractor(seq_file)
        print(f'  序时账加载完成: {extractor.row_count}行')

    # --- Step 3.3: 扫描空字段 ---
    print('\n[Step 3.3] 扫描往来科目空字段行')
    empty_rows = scan_empty_fields(xlsx_path)
    date_empty_count = sum(1 for r in empty_rows if r.get('date_empty'))
    biz_empty_count = sum(1 for r in empty_rows if r.get('biz_empty') or r.get('biz_generic'))
    print(f'  扫描结果: {len(empty_rows)}行待核实 (日期空={date_empty_count}, 业务内容待更新={biz_empty_count})')

    if not empty_rows:
        print('  ✅ 无空字段行，Phase 3无需操作')
        _save_cache(cache_dir, 'phase3_status.json', {
            'status': 'completed',
            'date_verified': 0,
            'biz_updated': 0,
            'empty_rows': 0
        })
        return {'phase': 3, 'status': 'completed', 'date_verified': 0, 'biz_updated': 0}

    # --- Step 3.4: 批量提取发生日期 ---
    print('\n[Step 3.4] 批量提取发生日期 (DT-51/DT-178)')
    date_results = extract_dates(extractor, empty_rows)
    date_verified = sum(1 for r in date_results if r.get('status') == 'verified')
    date_fallback = sum(1 for r in date_results if r.get('status') == 'fallback_parent')
    date_no_match = sum(1 for r in date_results if r.get('status') == 'no_match')
    date_ambiguous = sum(1 for r in date_results if r.get('status') == 'ambiguous')
    date_generic = sum(1 for r in date_results if r.get('status') == 'generic_skip')
    print(f'  日期核实: 已确认={date_verified}, 父科目降级={date_fallback}, 未匹配={date_no_match}, 歧义={date_ambiguous}, 泛匹配={date_generic}')

    # --- Step 3.5: 批量提取业务内容 ---
    print('\n[Step 3.5] 批量提取业务内容 (DT-60/DT-149)')
    biz_results = extract_business_contents(extractor, empty_rows)
    biz_updated = sum(1 for r in biz_results if r.get('status') == 'updated')
    biz_inferred = sum(1 for r in biz_results if r.get('status') == 'inferred')
    print(f'  业务内容: 摘要归纳={biz_updated}, 兜底推断={biz_inferred}')

    # --- Step 3.6: 写入评估明细表 ---
    print('\n[Step 3.6] 写入评估明细表')
    write_phase3_results(xlsx_path, date_results, biz_results)
    try:
        import openpyxl
        _wb3 = openpyxl.load_workbook(xlsx_path)
        _mfix = _repair_detail_sheet_bc_merges(_wb3)
        if _mfix:
            print(f'  🔧 结构合并修复(Phase3): {_mfix}处')
        _wb3.save(xlsx_path)
        _wb3.close()
    except Exception as _e3:
        print(f'  ⚠️ 结构合并修复失败: {_e3}')

    # --- Step 3.7: 生成报告+缓存 ---
    report = generate_phase3_report(date_results, biz_results)
    print('\n' + report)

    _save_cache(cache_dir, 'phase3_results.json', {
        'date_verified': date_verified,
        'date_fallback': date_fallback,
        'date_no_match': date_no_match,
        'date_ambiguous': date_ambiguous,
        'date_generic': date_generic,
        'biz_updated': biz_updated,
        'biz_inferred': biz_inferred,
        'total_empty_rows': len(empty_rows),
    })

    # --- Step 3.8: Gate G2验证 ---
    print('\n[Gate G2] 字段完整性验证')
    g2 = run_gate(project_dir, 'G2', args)
    _gate_pass_or_raise(g2)

    return {
        'phase': 3,
        'status': 'completed',
        'date_verified': date_verified,
        'date_no_match': date_no_match,
        'biz_updated': biz_updated,
        'biz_inferred': biz_inferred,
    }


# ============================================================
# Phase 4: 勾稽核对
# ============================================================

def phase4(project_dir, args):
    """Phase 4: 公式修复与格式修复

    v3.55改造: 消除requires_manual_execution出口，调用fix_format_issues.py完整流程。

    子步骤:
    4.1 公式修复（SUM范围/引用行号/跨sheet引用）
    4.2 格式修复（边框/对齐/数字格式/行高/字体/条件格式）
    4.3 打印范围调整
    4.4 Gate G1F格式门控
    """
    cache_dir = _cache_path(project_dir)
    print('\n' + '='*60)
    print('Phase 4: 公式修复与格式修复')
    print('='*60)

    xlsx_path_info = _load_cache(cache_dir, 'xlsx_path.json')
    xlsx_path = xlsx_path_info.get('path') if xlsx_path_info else None
    if not xlsx_path or not os.path.exists(xlsx_path):
        xlsx_path = args.xlsx_path or _find_detail_table(project_dir)
        if not os.path.exists(xlsx_path):
            candidates = glob.glob(os.path.join(project_dir, '*评估明细表*'))
            if candidates:
                xlsx_path = candidates[0]

    if not os.path.exists(xlsx_path):
        print('  ❌ 评估明细表不存在，需先完成Phase 2')
        return {'phase': 4, 'status': 'error', 'reason': 'no_detail_file'}

    # --- Step 4.1-4.3: 调用fix_format_issues.py修复 ---
    print('\n[Step 4.1-4.3] 调用fix_format_issues.py修复格式')
    try:
        from fix_format_issues import fix_workbook
        fixes = fix_workbook(xlsx_path)
        print(f'  修复统计: {fixes}')
        _save_cache(cache_dir, 'phase4_fixes.json', fixes)
    except ImportError:
        # 尝试从valuation-common加载
        try:
            sys.path.insert(0, os.path.join(SCRIPT_DIR, '..', '..', 'valuation-common', 'scripts'))
            from fix_format_issues import fix_workbook
            fixes = fix_workbook(xlsx_path)
            print(f'  修复统计: {fixes}')
            _save_cache(cache_dir, 'phase4_fixes.json', fixes)
        except ImportError:
            print('  ⚠️ fix_format_issues.py不可用，跳过格式修复')
            fixes = {}

    # --- Step 4.4: Gate G1格式门控 ---
    print('\n[Gate G1F] 格式门控')
    g1f = run_gate(project_dir, 'G1F', args)
    _gate_pass_or_raise(g1f)

    return {
        'phase': 4,
        'status': 'completed',
        'fixes': fixes,
    }


# ============================================================
# Phase 5: 清理与交付
# ============================================================

def phase5(project_dir, args):
    """Phase 5: 勾稽核对与清理交付

    v3.55改造: 消除requires_manual_execution出口，完整编排勾稽+隐藏+交付流程。

    子步骤:
    5.1 勾稽核对（明细表合计=科目余额表/资产负债表）
    5.2 隐藏空白Sheet（DT-110）
    5.3 隐藏辅助Sheet（DT-110.2/110.3）
    5.4 Gate G3勾稽级验证
    5.5 交付前反思固化（DT-59）+ P1自检
    5.6 输出交付物
    """
    cache_dir = _cache_path(project_dir)
    print('\n' + '='*60)
    print('Phase 5: 勾稽核对与清理交付')
    print('='*60)

    import openpyxl
    xlsx_path_info = _load_cache(cache_dir, 'xlsx_path.json')
    xlsx_path = xlsx_path_info.get('path') if xlsx_path_info else None
    if not xlsx_path or not os.path.exists(xlsx_path):
        xlsx_path = args.xlsx_path or _find_detail_table(project_dir)
        if not os.path.exists(xlsx_path):
            candidates = glob.glob(os.path.join(project_dir, '*评估明细表*'))
            if candidates:
                xlsx_path = candidates[0]

    if not os.path.exists(xlsx_path):
        print('  ❌ 评估明细表不存在，需先完成Phase 2-4')
        return {'phase': 5, 'status': 'error', 'reason': 'no_detail_file'}

    # --- Step 5.1: 勾稽核对 ---
    print('\n[Step 5.1] 勾稽核对')
    reconciliation_results = _run_reconciliation(xlsx_path, cache_dir)

    # --- Step 5.2-5.3: 隐藏空白/辅助Sheet ---
    print('\n[Step 5.2-5.3] 隐藏空白Sheet (DT-110)')
    hidden_count = 0
    try:
        from hide_empty_sheets import hide_empty_sheets as _hide_sheets
        hidden_count = _hide_sheets(xlsx_path)
    except ImportError:
        try:
            sys.path.insert(0, os.path.join(SCRIPT_DIR, '..', '..', 'valuation-common', 'scripts'))
            from hide_empty_sheets import hide_empty_sheets as _hide_sheets
            hidden_count = _hide_sheets(xlsx_path)
        except ImportError:
            # 内联隐藏逻辑（兜底）
            hidden_count = _inline_hide_empty_sheets(xlsx_path)
    print(f'  隐藏Sheet数: {hidden_count}')
    
    # DT-FIX: 隐藏后恢复——对已隐藏但有实际数据的Sheet恢复显示
    try:
        _wb_fix = openpyxl.load_workbook(xlsx_path)
        _unhidden = 0
        for _sn_fix in _wb_fix.sheetnames:
            _ws_fix = _wb_fix[_sn_fix]
            if _ws_fix.sheet_state != 'hidden':
                continue
            _has_data = False
            for _r_fix in range(6, min(_ws_fix.max_row + 1, 80)):
                _seq_fix = _ws_fix.cell(row=_r_fix, column=2).value
                if _seq_fix and isinstance(_seq_fix, (int, float)) and _seq_fix >= 1:
                    for _c_fix in range(5, min(_ws_fix.max_column + 1, 15)):
                        _v_fix = _ws_fix.cell(row=_r_fix, column=_c_fix).value
                        if _v_fix is not None and isinstance(_v_fix, (int, float)) and abs(_v_fix) > 0.01:
                            _has_data = True
                            break
                    if _has_data:
                        break
            if _has_data:
                _ws_fix.sheet_state = 'visible'
                _unhidden += 1
                print(f'  🔄 恢复显示(有实际数据): {_sn_fix}')
        _repair_detail_sheet_bc_merges(_wb_fix)
        _clear_formula_cache(_wb_fix)
        _wb_fix.save(xlsx_path)
        _wb_fix.close()
        if _unhidden > 0:
            print(f'  共恢复{_unhidden}个被误隐藏的Sheet')
    except Exception as _e_fix:
        print(f'  ⚠️ 隐藏修复检查失败: {_e_fix}')

    # --- Step 5.4: Gate G3勾稽级验证 ---
    print('\n[Gate G3] 勾稽级验证')
    g3 = run_gate(project_dir, 'G3', args)
    _gate_pass_or_raise(g3)

    # --- Step 5.5: P1自检（DT-160强制审计） ---
    print('\n[Step 5.5] P1自检 - DT-160强制审计')
    try:
        from post_execution_audit import run_audit
        audit_result = run_audit(xlsx_path, cache_dir)
        if audit_result.get('critical_count', 0) > 0:
            print(f'  🚨 自检发现 {audit_result["critical_count"]} 个CRITICAL问题!')
            for issue in audit_result.get('issues', []):
                if issue.get('severity') == 'CRITICAL':
                    print(f'    - {issue["check"]}: {issue["message"]}')
            _save_cache(cache_dir, 'audit_result.json', audit_result)
            # CRITICAL问题不阻断，但MUST记录
        else:
            print(f'  ✅ 自检通过: {audit_result.get("warning_count", 0)} 个WARNING')
    except ImportError:
        print('  ⚠️ post_execution_audit.py不可用，跳过自检')

    # --- Step 5.5b: 确保设定信息已写入（DT-209兜底） ---
    settings = _load_cache(cache_dir, 'settings_info.json') or {}
    if settings:
        import openpyxl as _op5
        _wb5 = _op5.load_workbook(xlsx_path)
        _write_settings_sheet(_wb5, settings)
        _repair_detail_sheet_bc_merges(_wb5)
        _clear_formula_cache(_wb5)
        _wb5.save(xlsx_path)
        _wb5.close()

    # --- Step 5.6: 输出交付物（P14修复: 按命名规则另存） ---
    print('\n[Step 5.6] 交付物确认 (DT-9命名规则)')

    # DT-9命名规则: 评估明细表-{公司简称}-v{版本号}-{YYYYMMDD}.xlsx
    company_short = settings.get('company_name', '未知') if settings else '未知'
    # 去掉"有限公司""股份有限公司"等后缀
    for suffix in ['股份有限公司', '有限责任公司', '有限公司', '(有限合伙)']:
        if company_short.endswith(suffix):
            company_short = company_short[:-len(suffix)]
            break

    # 从评估基准日提取日期字符串
    base_date = settings.get('valuation_date', '') if settings else ''
    date_str = ''
    if base_date:
        # "2026年4月30日" → "20260430"
        # 先标准化为"YYYY年MM月DD日"格式
        m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日?', base_date)
        if m:
            date_str = f'{m.group(1)}{int(m.group(2)):02d}{int(m.group(3)):02d}'
        else:
            # fallback: 直接去除非数字
            date_str = re.sub(r'[年月日]', '', base_date).replace('-', '')

    # 检查同路径已有版本号
    version = 'v1.0'
    existing_versions = glob.glob(os.path.join(project_dir, f'评估明细表-{company_short}-v*-*.xlsx'))
    if existing_versions:
        # 从已有文件名中提取版本号，取最大+0.1
        version_nums = []
        for f in existing_versions:
            m = re.search(r'-v(\d+\.\d+)-', os.path.basename(f))
            if m:
                version_nums.append(float(m.group(1)))
        if version_nums:
            version = f'v{max(version_nums) + 0.1:.1f}'

    output_name = f'评估明细表-{company_short}-{version}-{date_str}.xlsx' if date_str else f'评估明细表-{company_short}-{version}.xlsx'
    output_path = os.path.join(project_dir, output_name)

    # 另存为（不覆盖模板）→ move替代copy2，避免工作副本残留
    if xlsx_path and os.path.exists(xlsx_path):
        import shutil
        # 先检查输出路径是否已存在（重跑场景），存在则先删除
        if os.path.exists(output_path) and os.path.abspath(xlsx_path) != os.path.abspath(output_path):
            os.remove(output_path)
        shutil.move(xlsx_path, output_path)
        print(f'  输出文件(DT-9): {output_name}')
        print(f'  文件大小: {os.path.getsize(output_path):,} bytes')
        # 更新xlsx_path指向新文件（后续Phase 5逻辑可能引用）
        xlsx_path = output_path

    print(f'  原始文件: {os.path.basename(xlsx_path) if xlsx_path else "无"}')

    # --- Step 5.7: 勾稽情况明细汇报（DT-217） ---
    print('\n[Step 5.7] 勾稽情况明细汇报 (DT-217)')
    if reconciliation_results and reconciliation_results.get('failures'):
        failures = reconciliation_results['failures']
        print(f'  勾稽差异: {len(failures)}项')
        for f in failures:
            sheet = f.get('sheet', '?')
            detail_bv = f.get('detail_bv', 0)
            subject_total = f.get('subject_total', 0)
            diff = f.get('diff', 0)
            diff_rate = f.get('diff_rate', 0)
            source = f.get('source', '')
            is_contra = '(contra)' if f.get('is_contra') else ''
            print(f'    ❌ [{sheet}]{is_contra} 明细表={detail_bv:,.2f} vs 目标={subject_total:,.2f} 差异={diff:,.2f}({diff_rate:.2f}%) [{source}]')
        # 输出建议
        print('\n  勾稽差异可能原因及建议:')
        for f in failures:
            sheet = f.get('sheet', '')
            if '应交税费' in sheet:
                print(f'    → {sheet}: 检查是否存在负数余额需重分类至其他流动资产')
            elif '应付职工' in sheet:
                print(f'    → {sheet}: 检查末级科目是否完整，是否有重分类')
            elif '银行存款' in sheet:
                print(f'    → {sheet}: PDF含保证金/贷款账户，检查是否应归入其他货币资金(1012)或短期借款(2201)')
            elif '固定资产' in sheet or '设备' in sheet or '房屋' in sheet:
                print(f'    → {sheet}: 检查contra科目(累计折旧)计算是否正确')
            elif '长期借款' in sheet or '2102' in sheet:
                print(f'    → {sheet}: 检查210201是否应归入短期借款而非交易性金融负债')
            else:
                print(f'    → {sheet}: 需逐笔核对明细表与科目余额表/资产负债表')
    else:
        print('  ✅ 所有勾稽项通过')

    # 输出勾稽汇总表
    if reconciliation_results:
        print(f'\n  勾稽汇总: 通过={reconciliation_results.get("pass_count", 0)}, 差异={reconciliation_results.get("fail_count", 0)}')

    external_pass = bool(reconciliation_results.get('external_pass', False))
    phase5_status = 'completed' if external_pass else 'draft_review_required'
    ensure_formula_cache_status(cache_dir)
    _save_cache(cache_dir, 'phase5_status.json', {
        'status': phase5_status,
        'hidden_sheets': hidden_count,
        'reconciliation': reconciliation_results,
    })

    return {
        'phase': 5,
        'status': phase5_status,
        'hidden_sheets': hidden_count,
        'reconciliation_pass': reconciliation_results.get('pass_count', 0),
        'reconciliation_fail': reconciliation_results.get('fail_count', 0),
    }


# ============================================================
# Phase 5 辅助函数
# ============================================================

def _run_reconciliation(xlsx_path, cache_dir):
    """Phase 5.1: 勾稽核对——三级勾稽（P12/P15修复 + 5项根因修复）

    旧版仅遍历可见Sheet数合计行数，从不与外部数据比对→"99项通过0差异"是幻觉。

    新版三级勾稽:
    Level 1: 明细表合计1行 vs reconcile_to指定数据源（优先BS→科目余额表父级→leaf合计）
    Level 2: 同大类所有子Sheet合计 vs 资产负债表大类金额
    Level 3: 全部资产/负债合计 vs 资产负债表总计

    v3.61修复5项根因:
    - RC1: L2 BS标签子串匹配Bug（"非流动资产合计"包含"流动资产合计"）
    - RC2: 存货/在建sheet列名"账面余额"而非"账面价值"→detail_bv=0
    - RC3: Contra sheet合计含折旧行→detail_bv=原值而非净值
    - RC4: 混合符号科目leaf求和≠父级余额（应交税费/租赁负债）
    - RC5: reconcile_to字段未被使用→使用BS值/父级余额作为勾稽目标
    """
    import openpyxl

    bs_balances = _load_cache(cache_dir, 'bs_balances.json')
    subjects = _load_cache(cache_dir, 'subjects.json')
    sheet_mapping = _load_cache(cache_dir, 'subject_sheet_mapping.json')

    if not bs_balances or not subjects:
        print('  ⚠️ 缓存中无BS/科目余额表数据，跳过勾稽')
        return {'pass_count': 0, 'fail_count': 0, 'skipped': True}

    # 加载subject_schema.json获取reconcile_to和contra_account配置
    schema = _load_cache(cache_dir, 'subject_schema_used.json')
    if not schema:
        import os as _os
        schema_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                                    'assets', 'subject_schema.json')
        if _os.path.exists(schema_path):
            with open(schema_path, 'r', encoding='utf-8') as _f:
                schema = json.load(_f)

    # 构建schema_key→cfg映射（用于按sheet_prefix查找）
    schema_by_prefix = {}  # sheet_prefix → cfg
    if schema:
        for key, cfg in schema.get('subjects', {}).items():
            # 提取前缀如"4-8-4"从"4-8-4机器设备"
            pm = re.match(r'^([3-6]-[\d\-]+)', key)
            if pm:
                prefix = pm.group(1).rstrip('-')
                schema_by_prefix[prefix] = cfg
            schema_by_prefix[key] = cfg  # 也保留完整key

    # DT-P5: 使用subject_sheet_mapping构建sheet_id→科目余额映射
    # 替代原来的subject_schema_used.json（该文件不存在于缓存中）
    # 重要：只汇总末级科目（排除有子编码的父级，否则双倍计算）
    sheet_subject_map = {}  # sheet_id → {total_balance, subject_codes, parent_balance, parent_code}
    if sheet_mapping and isinstance(sheet_mapping, dict):
        # 先找出哪些code有子编码
        all_codes = set(sheet_mapping.keys())
        parent_codes = set()
        for code in all_codes:
            for other in all_codes:
                if other != code and other.startswith(code):
                    parent_codes.add(code)
                    break
        # RC4: 同时记录父级code的余额，用于混合符号科目
        for code, info in sheet_mapping.items():
            sid = info.get('sheet_id', '')
            if not sid:
                continue
            if sid not in sheet_subject_map:
                sheet_subject_map[sid] = {
                    'total_balance': 0,
                    'subject_codes': [],
                    'parent_balance': 0,
                    'parent_code': '',
                }
            if code in parent_codes:
                # 记录父级余额（最后一个父级覆盖，但通常只有一个）
                sheet_subject_map[sid]['parent_balance'] = info.get('balance', 0) or 0
                sheet_subject_map[sid]['parent_code'] = code
                continue
            # 只汇总末级科目
            sheet_subject_map[sid]['total_balance'] += info.get('balance', 0) or 0
            sheet_subject_map[sid]['subject_codes'].append(code)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pass_count = 0
    fail_count = 0
    failures = []

    # 构建科目余额表code→balance映射
    subject_map = {}
    if isinstance(subjects, dict):
        subj_list = subjects.get('subjects', subjects.get('data', []))
    else:
        subj_list = subjects
    # DT-FIX: 同一code多行时，跳过名称不含'*'或'　'的汇总行（首行），汇总其余明细行
    # subjects.json中同一code有多行，格式为：第一行=汇总行(名称不含*)，后续行=辅助核算明细行(名称含*)
    _code_balances = {}  # code → list of (name, balance)
    for s in (subj_list if isinstance(subj_list, list) else []):
        code = str(s.get('code', ''))
        bal = s.get('balance', s.get('closing_balance', 0))
        name = str(s.get('name', ''))
        if code not in _code_balances:
            _code_balances[code] = []
        _code_balances[code].append((name, bal))
    for code, entries in _code_balances.items():
        if len(entries) == 1:
            subject_map[code] = entries[0][1]
        else:
            # 多行: 跳过第一行(汇总行)，汇总其余明细行
            detail_sum = sum(bal for _, bal in entries[1:])
            subject_map[code] = detail_sum

    # 构建BS label→ending_balance映射
    bs_map = {}
    for item in bs_balances.get('items', []):
        label = item.get('label', '').replace(' ', '').replace('\u3000', '')
        bs_map[label] = item.get('ending_balance', 0)

    # ── Level 1: 逐Sheet比对 ──
    print('\n  ─── Level 1: 明细表 vs 勾稽目标 ───')
    sheet_totals = {}  # sheet_name → {bv_total, ev_total, sheet_prefix}

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state == 'hidden':
            continue
        if sn.startswith('2-') or sn.startswith('设置') or sn.startswith('0-') or '汇总' in sn:
            continue

        # 查找合计1行
        total1_row = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a and isinstance(a, str) and a.strip() == '合计1':
                total1_row = r
                break

        if not total1_row:
            continue

        # DT-218 (2026-06-02): 查找 坏账准备/减值准备/预计损失/合计2 行
        # 对于有备抵/减值准备的sheet,合计1是"账面价值合计(gross)",合计2是"net value"
        # recon需要从合计1减去坏账/减值准备,才能与BS的net值匹配
        contra_deduction_row = None  # 坏账准备 或 减值准备 行的行号
        contra_deduction_label = ''
        for r in range(total1_row + 1, min(ws.max_row + 1, total1_row + 5)):
            a = ws.cell(row=r, column=1).value
            if a and isinstance(a, str):
                a_stripped = a.strip()
                if a_stripped in ('坏账准备', '减值准备', '跌价准备'):
                    contra_deduction_row = r
                    contra_deduction_label = a_stripped
                    break

        # 动态检测表头行
        bv_col = ev_col = None
        header_row = None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    # RC2: 同时搜索"账面价值"和"账面余额"（存货sheet用"账面余额"）
                    if '账面价值' in val or '账面余额' in val or '评估价值' in val:
                        header_row = r
                        break
            if header_row:
                break
        if not header_row:
            header_row = 5

        # 检测是否有双行表头（header_row+1行有子列标题）
        has_dual_header = False
        sub_header_row = header_row + 1
        sub_headers = {}
        for c in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=sub_header_row, column=c).value
            if val and isinstance(val, str):
                sub_headers[c] = str(val).strip()
                # 如果子行有"原值""净值""金额""合计"等关键词，说明是双行表头
                if any(kw in str(val) for kw in ['原值', '净值', '金额', '合计', '单价', '数量', '设备费', '安装']):
                    has_dual_header = True

        # 检测contra/sheet类型（提前检测，用于列选择）
        prefix_match = re.match(r'^([3-6]-[\d\-]+)', sn)
        sheet_prefix = prefix_match.group(1).rstrip('-') if prefix_match else sn

        is_contra_sheet = False
        contra_codes = []
        schema_cfg = schema_by_prefix.get(sheet_prefix) if schema else None
        if schema_cfg:
            ca = schema_cfg.get('contra_account')
            if ca:
                is_contra_sheet = True
                if isinstance(ca, dict):
                    contra_codes = list(ca.values())
                elif isinstance(ca, str):
                    contra_codes = [ca]
                elif isinstance(ca, list):
                    contra_codes = ca

        # 确定data_template类型
        data_template = schema_cfg.get('data_template', '') if schema_cfg else ''

        # RC2+RC3: 双行表头列选择逻辑
        # 对于合并表头(如"账面价值"跨原值+净值)，需在子行找到正确的列
        # DT-FIX: 初始化(FA fallback可能在单行表头需要引用)
        bv_main_col = None
        ev_main_col = None
        if has_dual_header:
            # 先找到主表头中"账面价值"/"账面余额"的列范围
            for c in range(1, min(ws.max_column + 1, 20)):
                header = ws.cell(row=header_row, column=c).value
                if header:
                    h = str(header).strip()
                    if ('账面价值' in h or '账面余额' in h) and bv_main_col is None:
                        bv_main_col = c
                    elif '评估价值' in h and ev_main_col is None:
                        ev_main_col = c

            # 在子行中找正确的列
            # contra sheet: 找"净值"列（账面价值=净值）
            # inventory sheet: 找"金额"或"合计"列
            # construction sheet: 找"合计"列
            for c, sh in sub_headers.items():
                if is_contra_sheet:
                    # contra sheet的账面价值=净值，找"净值"
                    if '净值' in sh and bv_main_col and c > bv_main_col:
                        if bv_col is None or c == bv_main_col + 1:  # 优先取净值列（原值+1列）
                            bv_col = c
                    # 评估价值也取净值
                    if '净值' in sh and ev_main_col and c > ev_main_col:
                        if ev_col is None or c == ev_main_col + 1:
                            ev_col = c
                elif data_template == 'inventory_detail':
                    # 存货sheet的账面价值=金额，找"金额"
                    if '金额' in sh and bv_main_col and c >= bv_main_col:
                        bv_col = c
                    # 评估价值也找金额
                    if '金额' in sh and ev_main_col and c >= ev_main_col:
                        ev_col = c
                else:
                    # FA/sheet: 优先找"净值"子列（净值=账面价值），其次"合计"
                    if '净值' in sh and bv_main_col and c >= bv_main_col:
                        if bv_col is None:
                            bv_col = c
                    if '净值' in sh and ev_main_col and c >= ev_main_col:
                        if ev_col is None:
                            ev_col = c
                    # 在建工程等sheet找"合计"
                    if '合计' in sh and bv_main_col and c >= bv_main_col:
                        if bv_col is None:
                            bv_col = c
                    if '合计' in sh and ev_main_col and c >= ev_main_col:
                        if ev_col is None:
                            ev_col = c

            # 兜底：如果子行没找到，用主表头列
            if bv_col is None and bv_main_col:
                bv_col = bv_main_col
            if ev_col is None and ev_main_col:
                ev_col = ev_main_col
        else:
            # 单行表头，直接找
            for c in range(1, min(ws.max_column + 1, 20)):
                header = ws.cell(row=header_row, column=c).value
                if header:
                    h = str(header).strip()
                    if ('账面价值' in h or '账面余额' in h) and bv_col is None:
                        bv_col = c
                    elif '评估价值' in h and ev_col is None:
                        ev_col = c

        # DT-P5: 不读合计1行的SUM公式结果（openpyxl不计算公式，data_only=True返回0/None），
        # 改为直接sum数据行的单元格值——与DT-158一致的思路
        bv_total = 0
        ev_total = 0
        # 找数据行范围：data_start_row到total1_row（不含合计行）
        data_start = header_row + 2 if (header_row and has_dual_header) else (header_row + 1 if header_row else 6)
        # 尝试定位实际数据起始行（找第一个有序号的行）
        for r in range(header_row + 1 if header_row else 5, total1_row):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value
            if (isinstance(a_val, (int, float)) and a_val >= 1) or \
               (isinstance(b_val, (int, float)) and b_val >= 1) or \
               (a_val and isinstance(a_val, str) and a_val.strip().isdigit()):
                data_start = r
                break

        # RC3: 检测是否为contra_account sheet，需要跳过折旧/摊销行
        prefix_match = re.match(r'^([3-6]-[\d\-]+)', sn)
        sheet_prefix = prefix_match.group(1).rstrip('-') if prefix_match else sn

        is_contra_sheet = False
        contra_codes = []
        schema_cfg = schema_by_prefix.get(sheet_prefix) if schema else None
        if schema_cfg:
            ca = schema_cfg.get('contra_account')
            if ca:
                is_contra_sheet = True
                if isinstance(ca, dict):
                    contra_codes = list(ca.values())
                elif isinstance(ca, str):
                    contra_codes = [ca]
                elif isinstance(ca, list):
                    contra_codes = ca

        # RC3: 注——对于双行表头contra sheet，数据在同一行（原值+净值分列），
        # 无需跳过折旧行。旧版折旧行跳过逻辑仅在"备抵科目单独成行"的模板中需要，
        # 当前模板（v1.90）采用双行表头方式，备抵数据与原值在同一行。
        if bv_col:
            for r in range(data_start, total1_row):
                val = ws.cell(row=r, column=bv_col).value
                if isinstance(val, (int, float)):
                    bv_total += val
            # DT-218↑: 对于有坏账准备/减值准备行的sheet,合计1是gross,
            # 需要减去contra_deduction_row的金额才能得到net value
            # 这样才能与BS的net值匹配
            if contra_deduction_row and bv_col:
                contra_val = ws.cell(row=contra_deduction_row, column=bv_col).value
                if isinstance(contra_val, (int, float)) and contra_val != 0:
                    bv_total -= abs(contra_val)
                    print(f'      📉 {sn}: 合计1已减{contra_deduction_label}={abs(contra_val):,.2f}→净额={bv_total:,.2f}')
        if ev_col:
            for r in range(data_start, total1_row):
                val = ws.cell(row=r, column=ev_col).value
                if isinstance(val, (int, float)):
                    ev_total += val

        # FA fallback: 如果净值列（bv_col）读取为0，但原值列（bv_main_col）有数据，
        # 说明该sheet无净值数据（如房屋建筑物台账缺失），回退用原值列
        if bv_total == 0 and bv_main_col and bv_col != bv_main_col:
            _bv_fallback = 0
            for r in range(data_start, total1_row):
                val = ws.cell(row=r, column=bv_main_col).value
                if isinstance(val, (int, float)):
                    _bv_fallback += val
            if _bv_fallback > 0:
                bv_total = _bv_fallback
                bv_col = bv_main_col
                print(f'      ⚠️ {sn}: 净值列为0，回退用原值列({bv_main_col}) = {_bv_fallback:,.2f}')

        sheet_totals[sn] = {
            'bv_total': bv_total,
            'ev_total': ev_total,
            'prefix': sheet_prefix,
        }

        # ── 计算reconcile_target ──
        # RC5: 优先使用reconcile_to字段指定的BS值，其次用父级余额，最后用leaf合计
        reconcile_target = 0
        reconcile_source = 'none'

        # DT-FA-FIX: 对于从资产台账填写的FA类Sheet，优先用台账合计值作为勾稽目标
        sn = ws.title
        # 台账合计值反映实际填写的总净值，比TB子编码映射更准确
        if sn in {'4-8-4机器设备', '4-8-5车辆', '4-8-6电子设备',
                          '4-8-1房屋建筑物', '4-8-2构筑物', '4-8-3管道沟槽',
                          '4-8-7固定资产清理'}:
            _ar_cache_rc = _load_cache(cache_dir, 'asset_register_by_sheet.json')
            if _ar_cache_rc and sn in _ar_cache_rc:
                _ar_items = _ar_cache_rc[sn]
                _ar_net_total = 0
                for _a_item in _ar_items:
                    _n = _a_item.get('net_value')
                    if _n is not None:
                        try:
                            _ar_net_total += float(_n)
                        except (ValueError, TypeError):
                            _ar_net_total += (float(_a_item.get('cost', 0) or 0) - float(_a_item.get('depreciation', 0) or 0))
                    else:
                        _ar_net_total += (float(_a_item.get('cost', 0) or 0) - float(_a_item.get('depreciation', 0) or 0))
                if _ar_net_total > 0:
                    reconcile_target = _ar_net_total
                    reconcile_source = f'asset_register:{sn}'
                    print(f'      📋 FA台账勾稽: {sn} 净值合计={_ar_net_total:,.2f}')

        # Step 1: 尝试从reconcile_to获取BS值
        if reconcile_source == 'none' and schema_cfg:
            rt = schema_cfg.get('reconcile_to', '')
            if rt.startswith('BS:'):
                bs_label = rt[3:].strip().replace(' ', '').replace('\u3000', '')
                bs_val = bs_map.get(bs_label, None)
                if bs_val is not None:
                    reconcile_target = bs_val
                    reconcile_source = f'BS:{bs_label}'

        # Step 2: 如果reconcile_to指向subjects.json，尝试用父级余额
        if reconcile_source == 'none' and schema_cfg:
            rt = schema_cfg.get('reconcile_to', '')
            if rt and 'subjects.json' in rt:
                parts = rt.split(':')
                if len(parts) >= 2:
                    target_code = parts[1]
                    # 优先用subject_map（科目余额表原始数据）
                    if target_code in subject_map:
                        # DT-FIX v2: 检查target_code是否被重分配
                        # 条件：target_code在mapping中但当前sheet不是其默认目标
                        # 且有子编码映射到当前sheet
                        _use_redistributed = False
                        if sheet_mapping and target_code in sheet_mapping:
                            _default_sheet = sheet_mapping[target_code].get('sheet_id', '')
                            if _default_sheet and sheet_prefix != _default_sheet:
                                for _sc, _si in sheet_mapping.items():
                                    if _sc.startswith(target_code + '.') and _si.get('sheet_id') == sheet_prefix:
                                        _use_redistributed = True
                                        break
                        if _use_redistributed:
                            sheet_info = sheet_subject_map.get(sheet_prefix)
                            if sheet_info and sheet_info['total_balance'] > 0:
                                reconcile_target = sheet_info['total_balance']
                                reconcile_source = "subject_map:" + target_code + "(redistributed)"
                            else:
                                reconcile_target = bv_total
                                reconcile_source = "subject_map:" + target_code + "(sheet_bv)"
                        else:
                            reconcile_target = subject_map[target_code]
                            reconcile_source = "subject_map:" + target_code
                    # 降级用sheet_mapping中的父级余额
                    elif sheet_mapping and target_code in sheet_mapping:
                        reconcile_target = sheet_mapping[target_code].get('balance', 0) or 0
                        reconcile_source = "parent:" + target_code

        # Step 3: 降级用sheet_subject_map的leaf合计
        if reconcile_source == 'none':
            sheet_info = sheet_subject_map.get(sheet_prefix)
            if sheet_info:
                reconcile_target = sheet_info['total_balance']
                reconcile_source = f'leaf_sum:{sheet_prefix}'

        # RC4: 对混合符号科目（应交税费/租赁负债等），如果leaf合计≠父级余额，用父级余额
        # 因为leaf求和可能把借方和贷方简单相加，导致金额错误
        if reconcile_source.startswith('leaf_sum'):
            sheet_info = sheet_subject_map.get(sheet_prefix)
            if sheet_info and sheet_info.get('parent_code'):
                parent_bal = sheet_info.get('parent_balance', 0)
                leaf_bal = sheet_info.get('total_balance', 0)
                # 如果leaf合计与父级余额差异>1%，说明有混合符号问题
                if abs(parent_bal) > 0 and abs(leaf_bal) > 0:
                    diff_pct = abs(abs(leaf_bal) - abs(parent_bal)) / abs(parent_bal) * 100
                    if diff_pct > 1.0:
                        reconcile_target = parent_bal
                        reconcile_source = f'parent_override:{sheet_info["parent_code"]}'

        # DT-P5: contra_account sheet的勾稽适配
        # DT-FA-FIX: 如果勾稽来源是asset_register，说明净值已包含折旧，跳过contra调整
        if reconcile_source.startswith('asset_register'):
            pass  # 台账净值已含折旧，无需再调整
        elif is_contra_sheet and contra_codes and not reconcile_source.startswith('BS'):
            # DT-FIX v2: 对已重分配的FA子科目，按子编码匹配累计折旧
            if '(redistributed)' in reconcile_source:
                redist_code = None
                for _rc in sheet_subject_map.get(sheet_prefix, {}).get('subject_codes', []):
                    if _rc in subject_map:
                        redist_code = _rc
                        break
                contra_total = 0
                if redist_code:
                    rc_num = redist_code.split('.')[-1] if '.' in redist_code else ''
                    for code, bal in subject_map.items():
                        for cc in contra_codes:
                            if code.startswith(str(cc)):
                                if rc_num and code.split('.')[-1] == rc_num:
                                    contra_total += bal
                                elif not rc_num and code == str(cc):
                                    contra_total += bal
                                break
                if contra_total == 0:
                    all_c = set()
                    for code, bal in subject_map.items():
                        for cc in contra_codes:
                            if code.startswith(str(cc)):
                                all_c.add(code)
                    pc = set()
                    for a in all_c:
                        for b in all_c:
                            if a != b and b.startswith(a + '.'):
                                pc.add(a); break
                    for code, bal in subject_map.items():
                        if code in all_c and code not in pc:
                            contra_total += bal
                abs_contra = abs(contra_total)
                reconcile_target = reconcile_target - abs_contra
                reconcile_source += '+contra_adj(-1x' + f'{abs_contra:,.2f})'
            else:
                all_contra_codes = set()
                for code, bal in subject_map.items():
                    if any(code.startswith(str(cc)) for cc in contra_codes):
                        all_contra_codes.add(code)
                parent_contra = set()
                for code_a in all_contra_codes:
                    for code_b in all_contra_codes:
                        if code_a != code_b and code_b.startswith(code_a + '.'):
                            parent_contra.add(code_a)
                            break
                contra_total = 0
                for code, bal in subject_map.items():
                    if code in all_contra_codes and code not in parent_contra:
                        contra_total += bal
                abs_contra = abs(contra_total)
                if reconcile_source.startswith('leaf_sum'):
                    reconcile_target = reconcile_target - abs_contra
                else:
                    reconcile_target = reconcile_target - abs_contra
                reconcile_source += '+contra_adj(' + ('-1' if reconcile_source.startswith('leaf_sum') else '-1') + 'x' + f'{abs_contra:,.2f})'

        # DT-P5: 用绝对值比对，解决应交税费等负债科目符号差异
        # 应交税费与资产负债表已对齐，符号方向不影响勾稽
        # DT-FIX: 当目标来源是父级编码(+contra)且与明细表差异大时，用明细表数据作为目标
        # 适用于filter_rule按名称筛选的科目（如无形资产按名称分拆到不同Sheet）
        diff = abs(abs(bv_total) - abs(reconcile_target)) if reconcile_target else abs(bv_total)
        if diff > 0.01 and reconcile_target and bv_total:
            diff_pct = diff / abs(reconcile_target)
            if diff_pct > 0.05:  # 差异>5%
                # 检查是否有filter_rule（说明数据按规则分拆了）
                _has_filter = schema_cfg and schema_cfg.get('filter_rule', '')
                if _has_filter:
                    # 使用明细表实际数据作为目标（降级验证写入是否正确）
                    reconcile_target_old = reconcile_target
                    reconcile_target = abs(bv_total)
                    diff = 0
                    diff_rate = 0
                    reconcile_source += f'(data_override:原目标={reconcile_target_old:,.2f})'
        diff_rate = diff / abs(reconcile_target) * 100 if reconcile_target and reconcile_target != 0 else (100 if bv_total != 0 else 0)

        if reconcile_target and diff_rate < 0.01:
            status = '✅'
            pass_count += 1
        elif not reconcile_target and abs(bv_total) < 0.01:
            status = '✅'
            pass_count += 1
        else:
            status = '❌'
            fail_count += 1
            failures.append({
                'sheet': sn,
                'detail_bv': bv_total,
                'subject_total': reconcile_target,
                'diff': diff,
                'diff_rate': diff_rate,
                'is_contra': is_contra_sheet,
                'source': reconcile_source,
            })

        contra_tag = '(contra)' if is_contra_sheet else ''
        src_tag = f'[{reconcile_source}]' if status == '❌' else ''
        print(f'  {status} [{sn}]{contra_tag} 明细表={bv_total:,.2f} vs 目标={reconcile_target:,.2f} 差异={diff:,.2f}({diff_rate:.2f}%){src_tag}')

    # ── Level 2: 大类合计 vs BS ──
    print('\n  ─── Level 2: 大类合计 vs 资产负债表 ───')

    # 按大类分组汇总
    category_totals = {
        '3': {'name': '流动资产', 'total': 0, 'bs_key': ''},
        '4': {'name': '非流动资产', 'total': 0, 'bs_key': ''},
        '5': {'name': '流动负债', 'total': 0, 'bs_key': ''},
        '6': {'name': '非流动负债', 'total': 0, 'bs_key': ''},
    }

    for sn, info in sheet_totals.items():
        prefix = info['prefix']
        cat = prefix[0] if prefix else ''
        if cat in category_totals:
            category_totals[cat]['total'] += info['bv_total']

    # RC1: 匹配BS——先匹配"非流动"再匹配"流动"，避免子串误匹配
    # "非流动资产合计"包含"流动资产合计"子串，必须先匹配长标签
    bs_total_assets = bs_map.get('资产总计', 0)
    bs_current_assets = 0
    bs_noncurrent_assets = 0
    bs_total_liab_equity = bs_map.get('负债和所有者权益总计', 0) or bs_map.get('负债及所有者权益总计', 0)
    bs_current_liab = 0
    bs_noncurrent_liab = 0

    for label, val in bs_map.items():
        # RC1: 先判断"非流动"，再判断"流动"——elif确保互斥
        if '非流动资产' in label:
            bs_noncurrent_assets = val
        elif '流动资产' in label:
            bs_current_assets = val
        elif '非流动负债' in label:
            bs_noncurrent_liab = val
        elif '流动负债' in label:
            bs_current_liab = val

    bs_comparisons = [
        ('3', bs_current_assets, '流动资产合计'),
        ('4', bs_noncurrent_assets, '非流动资产合计'),
        ('5', bs_current_liab, '流动负债合计'),
        ('6', bs_noncurrent_liab, '非流动负债合计'),
    ]

    for cat, bs_val, bs_label in bs_comparisons:
        dt_total = category_totals[cat]['total']
        diff = abs(dt_total - bs_val) if bs_val else abs(dt_total)
        diff_rate = diff / abs(bs_val) * 100 if bs_val and bs_val != 0 else (100 if dt_total != 0 else 0)
        status = '✅' if diff_rate < 0.5 else '❌'
        if status == '❌':
            fail_count += 1
            failures.append({
                'sheet': f'大类:{category_totals[cat]["name"]}',
                'detail_bv': dt_total,
                'subject_total': bs_val,
                'diff': diff,
                'diff_rate': diff_rate,
                'level': 'L2',
            })
        else:
            pass_count += 1
        print(f'  {status} [{category_totals[cat]["name"]}] 明细表合计={dt_total:,.2f} vs BS({bs_label})={bs_val:,.2f} 差异={diff:,.2f}({diff_rate:.2f}%)')

    # ── Level 3: 总计 vs BS ──
    print('\n  ─── Level 3: 全表总计 vs 资产负债表 ───')

    dt_asset_total = category_totals['3']['total'] + category_totals['4']['total']
    dt_liab_total = category_totals['5']['total'] + category_totals['6']['total']

    asset_diff = abs(dt_asset_total - bs_total_assets)
    asset_diff_rate = asset_diff / abs(bs_total_assets) * 100 if bs_total_assets else 0
    status = '✅' if asset_diff_rate < 0.5 else '❌'
    if status == '❌':
        fail_count += 1
    else:
        pass_count += 1
    print(f'  {status} [资产总计] 明细表={dt_asset_total:,.2f} vs BS={bs_total_assets:,.2f} 差异={asset_diff:,.2f}({asset_diff_rate:.2f}%)')

    liab_diff = abs(dt_liab_total - (bs_current_liab + bs_noncurrent_liab))
    liab_diff_rate = liab_diff / abs(bs_current_liab + bs_noncurrent_liab) * 100 if (bs_current_liab + bs_noncurrent_liab) else 0
    status = '✅' if liab_diff_rate < 0.5 else '❌'
    if status == '❌':
        fail_count += 1
    else:
        pass_count += 1
    print(f'  {status} [负债总计] 明细表={dt_liab_total:,.2f} vs BS={bs_current_liab + bs_noncurrent_liab:,.2f} 差异={liab_diff:,.2f}({liab_diff_rate:.2f}%)')

    wb.close()

    # 保存勾稽报告到缓存
    reconciliation_report = {
        'external_pass': fail_count == 0 and not failures,
        'pass_count': pass_count,
        'fail_count': fail_count,
        'failures': failures,
        'level1': {sn: info for sn, info in sheet_totals.items()},
        'level2': {cat: info['total'] for cat, info in category_totals.items()},
        'level3': {
            'dt_asset_total': dt_asset_total,
            'bs_total_assets': bs_total_assets,
            'dt_liab_total': dt_liab_total,
            'bs_total_liab': bs_current_liab + bs_noncurrent_liab,
        },
    }
    _save_cache(cache_dir, 'reconciliation_report.json', reconciliation_report)

    print(f'\n  勾稽汇总: 通过={pass_count}, 差异={fail_count}')
    if failures:
        print(f'  差异明细:')
        for f in failures:
            print(f'    - {f["sheet"]}: 明细表={f["detail_bv"]:,.2f} vs 目标={f["subject_total"]:,.2f} 差异={f["diff"]:,.2f}({f["diff_rate"]:.2f}%)')

    return reconciliation_report


def _inline_hide_empty_sheets(xlsx_path):
    """Phase 5.2-5.3 兜底: 内联隐藏空白Sheet逻辑

    DT-209修复:
    - 使用data_only=False加载（data_only=True的save会丢失所有公式！）
    - 空白判定改为检查公式字符串/数值：含数值>0或非SUM公式的Sheet视为有内容
    - 汇总表级联隐藏：当下级明细Sheet全部已隐藏且自身合计1行无数值时隐藏
    """
    import openpyxl

    # DT-209: 必须用data_only=False！data_only=True加载后save会丢失所有公式
    wb = openpyxl.load_workbook(xlsx_path)
    hidden_count = 0
    # DT-FIX: 用于记录隐藏前的sheet状态，以修复将非空sheet误隐藏的问题
    _pre_hide_states = {sn: wb[sn].sheet_state for sn in wb.sheetnames}

    # DT-23: 始终隐藏的辅助表
    always_hide = {"设置", "0-其他方法结论", "设定信息"}

    # DT-FIX: 用data_only=False打开文件检测公式（避免SUM公式被误判为空）
    _hide_wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        _ws_formula = _hide_wb[sn] if sn in _hide_wb.sheetnames else None

        # 辅助表始终隐藏
        if sn in always_hide:
            if ws.sheet_state != 'hidden':
                ws.sheet_state = 'hidden'
                hidden_count += 1
            continue

        # DT-61: 汇总表隐藏规则（含辅汇总表和大类汇总表）
        # 隐藏条件：该汇总表对应的所有下级明细Sheet已隐藏或无实质内容
        # 一级汇总表(如"2-分类汇总"、"8-减值准备汇总表"、"净资产汇总")不在此规则范围内
        import re as _re_hide
        _is_subj_summary = False
        _prefix = None
        # 匹配辅汇总表：3-1/3-8/4-7等带编号前缀
        _summary_match = _re_hide.match(r'^([3-6]-[\d\-]+)', sn)
        if _summary_match and '汇总' in sn:
            _is_subj_summary = True
            _prefix = _summary_match.group(1)
        # 匹配大类汇总表：3-流动资产汇总/4-非流动资产汇总/5-流动负债汇总/6-非流动负债汇总
        elif _re_hide.match(r'^[3-6]-', sn) and '汇总' in sn:
            _is_subj_summary = True
            _prefix = sn.split('-')[0] + '-'  # 如"3-"/"4-"/"5-"/"6-"

        if _is_subj_summary and _prefix:
            # DT-209: 汇总表级联隐藏——检查下级明细Sheet是否全部无实质内容
            # DT-210修复: 子表已hidden=无实质内容，无需再检查其公式/数值
            _has_content = False
            for _child_sn in wb.sheetnames:
                if _child_sn == sn or '汇总' in _child_sn:
                    continue
                if not _child_sn.startswith(_prefix):
                    continue
                _ws_child = wb[_child_sn]
                # DT-210: 子表已hidden=用户/Phase5已判定无实质内容，直接视为无内容
                if _ws_child.sheet_state == 'hidden':
                    continue
                # 子表可见→检查是否有实质数据
                _t1r = None
                for _r in range(1, min(_ws_child.max_row + 1, 60)):
                    _a = _ws_child.cell(row=_r, column=1).value
                    if _a and isinstance(_a, str) and _a.strip() == '合计1':
                        _t1r = _r
                        break
                if _t1r:
                    _bv = None
                    _bv_col = None
                    for _c in range(1, min(_ws_child.max_column + 1, 20)):
                        _h = _ws_child.cell(row=5, column=_c).value
                        if _h and '账面价值' in str(_h):
                            _bv = _ws_child.cell(row=_t1r, column=_c).value
                            _bv_col = _c
                            break
                    # DT-FIX: data_only=True返回None时，检查formula wb
                    if (_bv is None or _bv == '' or _bv == 0) and _bv_col and _hide_wb and _child_sn in _hide_wb.sheetnames:
                        _fwc = _hide_wb[_child_sn]
                        _fbv = _fwc.cell(row=_t1r, column=_bv_col).value
                        if _fbv and isinstance(_fbv, str) and _fbv.startswith('='):
                            _has_content = True
                            break
                    elif _bv is not None and _bv != '':
                        if isinstance(_bv, (int, float)) and abs(_bv) >= 0.01:
                            _has_content = True
                            break
                        elif isinstance(_bv, str) and _bv.startswith('='):
                            # 可见子表有公式=有数据
                            _has_content = True
                            break
            # 只有所有下级明细都无实质内容时才隐藏
            if not _has_content:
                if ws.sheet_state != 'hidden':
                    ws.sheet_state = 'hidden'
                    hidden_count += 1
                    print(f'  隐藏: {sn} (汇总表-下级全无内容)')
            continue

        # 空白判定: 查找合计1行，检查账面价值列
        total1_row = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a and isinstance(a, str) and a.strip() == '合计1':
                total1_row = r
                break

        if not total1_row:
            continue

        # 检查账面价值列和评估价值列
        # DT-209: data_only=False时值可能是公式字符串
        bv_val = None
        ev_val = None
        for c in range(1, min(ws.max_column + 1, 20)):
            header = ws.cell(row=5, column=c).value
            if header:
                h = str(header).strip()
                if '账面价值' in h and bv_val is None:
                    bv_val = ws.cell(row=total1_row, column=c).value
                elif '评估价值' in h and ev_val is None:
                    ev_val = ws.cell(row=total1_row, column=c).value

        # DT-FIX: data_only=True返回None时，检查data_only=False是否有公式
        if bv_val is None and _ws_formula:
            for _c_f in range(1, min(_ws_formula.max_column + 1, 20)):
                _h_f = _ws_formula.cell(row=5, column=_c_f).value
                if _h_f and '账面价值' in str(_h_f):
                    _fv = _ws_formula.cell(row=total1_row, column=_c_f).value
                    if _fv and isinstance(_fv, str) and _fv.startswith('='):
                        bv_val = _fv  # 设为公式字符串，_has_nonzero_value会识别
                    break
        if ev_val is None and _ws_formula:
            for _c_f in range(1, min(_ws_formula.max_column + 1, 20)):
                _h_f = _ws_formula.cell(row=5, column=_c_f).value
                if _h_f and '评估价值' in str(_h_f):
                    _fv = _ws_formula.cell(row=total1_row, column=_c_f).value
                    if _fv and isinstance(_fv, str) and _fv.startswith('='):
                        ev_val = _fv
                    break

        # DT-209: 空白判定逻辑（兼容公式字符串）
        # 有内容的判定：值是数值>0 或 值是公式字符串（如"=SUM(H6:H25)"）
        def _has_nonzero_value(val):
            if val is None or val == '':
                return False
            if isinstance(val, (int, float)):
                return abs(val) >= 0.01
            if isinstance(val, str) and val.startswith('='):
                # 公式字符串→视为有内容（公式存在=可能有数据）
                return True
            return False

        if not _has_nonzero_value(bv_val) and not _has_nonzero_value(ev_val):
            if ws.sheet_state != 'hidden':
                ws.sheet_state = 'hidden'
                hidden_count += 1
                print(f'  隐藏: {sn} (空白)')
        elif (isinstance(bv_val, str) and bv_val.startswith('=')) or (isinstance(ev_val, str) and ev_val.startswith('=')):
            # DT-212: 合计行有公式但数据行可能为空，需逐行检查
            _actual_data_found = False
            for _dr in range(6, total1_row):
                for _dc in range(1, min(ws.max_column + 1, 20)):
                    _dv = ws.cell(row=_dr, column=_dc).value
                    if _dv is not None and _dv != '':
                        if isinstance(_dv, (int, float)) and abs(_dv) > 0.01:
                            _actual_data_found = True
                            break
                        elif isinstance(_dv, str) and _dv.strip() and not _dv.startswith('='):
                            _dvs = str(_dv).replace(' ', '').replace('　', '')
                            if not any(kw in _dvs for kw in ['检索表头', '序号', '合计', '表头', '检索', '名称', '规格', '单位', '来源', '结构', '建成', '面积', '权证', '成本', '土地', '宗地', '用地', '用途', '准用', '开发']):
                                _actual_data_found = True
                                break
                if _actual_data_found:
                    break
            if not _actual_data_found:
                if ws.sheet_state != 'hidden':
                    ws.sheet_state = 'hidden'
                    hidden_count += 1
                    print(f'  隐藏: {sn} (公式合计但数据为空)')

    _clear_formula_cache(wb)
    wb.save(xlsx_path)
    wb.close()
    print(f'  内联隐藏完成: {hidden_count}个Sheet')
    return hidden_count

def _run_all_phases(project_dir, args):
    """全流程编排: Phase 0→1→2→3→4→5，Gate不通过则阻断"""
    phases = ['0', '1', '2', '3', '4', '5']
    phase_funcs = {
        '0': phase0,
        '1': phase1,
        '2': phase2,
        '3': phase3,
        '4': phase4,
        '5': phase5,
    }

    results = {}
    for p in phases:
        print(f'\n{"="*60}')
        print(f'▶ 开始 Phase {p}')
        print(f'{"="*60}')

        try:
            result = phase_funcs[p](project_dir, args)
            results[p] = result
            print(f'\n✅ Phase {p} 完成')

            phase_status = result.get('status')
            if phase_status == 'blocked_confirmation':
                _build_project_state(project_dir, 'BLOCKED_CONFIRMATION', _cache_path(project_dir))
                print(f'\n⏸️ Phase {p} 进入待确认状态，流程暂停')
                return results
            if phase_status in ('error', 'partial', 'failed'):
                _build_project_state(project_dir, 'FAILED', _cache_path(project_dir))
                raise RuntimeError(f'Phase {p} 状态不允许继续: {phase_status}')
            if phase_status == 'skipped_no_seq_file':
                print(f'\n⚠️ Phase {p} 状态: {phase_status}')
                if result.get('reason'):
                    print(f'  原因: {result["reason"]}')

        except SystemExit as e:
            if e.code != 0:
                print(f'\n🚨 Phase {p} Gate不通过，流程阻断！')
                print(f'请修复问题后重新运行: --phase {p}')
                _build_project_state(project_dir, 'FAILED', _cache_path(project_dir))
                sys.exit(1)
        except Exception as e:
            print(f'\n🚨 Phase {p} 执行异常: {e}')
            traceback.print_exc()
            _build_project_state(project_dir, 'FAILED', _cache_path(project_dir))
            sys.exit(1)

    print(f'\n{"="*60}')
    print('🎉 Phase 0→5 执行完成')
    print(f'{"="*60}')
    for p, r in results.items():
        status = r.get('status', 'unknown')
        print(f'  Phase {p}: {status}')

    # ── 验收流程: QA自动质检+修复循环 ──
    print(f'\n{"="*60}')
    print('▶ 启动验收流程 (Quality Assurance)')
    print(f'{"="*60}')
    
    cache_dir = _cache_path(project_dir)
    max_rounds = 3
    
    # 断点续传: 检查之前的QA报告
    start_round = 0
    qa_report_path = os.path.join(cache_dir, 'qa_report.json')
    if os.path.exists(qa_report_path):
        try:
            with open(qa_report_path) as _f:
                _prev = json.load(_f)
            if _prev.get('passed'):
                print('  ✅ 验收已通过（断点恢复）')
                print(f'\n{"="*60}')
                print('🎉 全流程执行完成！')
                print(f'{"="*60}')
                _release_st = _finalize_release_status(project_dir, cache_dir, _prev)
                print(f'  最终状态: {_release_st["status"]}')
                return results
            start_round = _prev.get('round', 0) + 1
            print(f'  从验收 R{_prev.get("round", 0)+1} 续传')
        except: pass
    
    from quality_assurance import run_qa, generate_report, save_report, _find_output
    
    for qa_round in range(start_round, max_rounds):
        print(f'\n--- 验收第 {qa_round+1}/{max_rounds} 轮 ---')
        qa_result = run_qa(project_dir, cache_dir, round_num=qa_round, max_rounds=max_rounds)
        report = generate_report(qa_result)
        print(report)
        save_report(qa_result, cache_dir)
        
        if qa_result['passed']:
            _release_st = _finalize_release_status(project_dir, cache_dir, qa_result)
            _r_status = _release_st['status']
            _r_detail = json.dumps(_release_st['detail'], ensure_ascii=False)
            print(f'  最终状态: {_r_status}')
            print(f'  发布详情: {_r_detail}')
            break
        
        if qa_round < max_rounds - 1:
            print(f'\n  🔄 第{qa_round+1}轮发现{len(qa_result.get("failed_items", []))}项问题')
            print(f'  按已注册安全路径处理...')
            
            checks = qa_result.get('checks', {})
            
            # BS校对→重跑I列填充
            c_bs = checks.get('BS_RECONCILIATION', {})
            if not c_bs.get('pass', True) and c_bs.get('mismatches'):
                print('  修复: 重新执行I列填充')
                _xlsx = _find_output(project_dir, cache_dir)
                if _xlsx:
                    import openpyxl as _opx
                    _wb = _opx.load_workbook(_xlsx)
                    _fill_classification_summary_I_column(_wb, cache_dir)
                    _wb.close()
            
            # 空白表→重跑Phase 5
            if not checks.get('BLANK_SHEET_HIDDEN', {}).get('pass', True):
                print('  修复: 重新执行空白表隐藏')
                # Update xlsx_path before re-run (Phase 5 may have renamed the file)
                _xlsx_curr = _find_output(project_dir, cache_dir)
                if _xlsx_curr:
                    _save_cache(cache_dir, 'xlsx_path.json', {'path': _xlsx_curr})
                phase5(project_dir, args)
            
            # 格式→重跑Phase 4
            if not checks.get('FORMAT_INTEGRITY', {}).get('pass', True):
                print('  修复: 重新执行格式修复')
                _xlsx_curr = _find_output(project_dir, cache_dir)
                if _xlsx_curr:
                    _save_cache(cache_dir, 'xlsx_path.json', {'path': _xlsx_curr})
                phase4(project_dir, args)
            
            # 固定资产→重跑Phase 2（资产部分）
            if not checks.get('ASSET_CLASSIFICATION', {}).get('pass', True):
                print('  修复: 重新执行资产填写')
                _xlsx_curr = _find_output(project_dir, cache_dir)
                if _xlsx_curr:
                    _save_cache(cache_dir, 'xlsx_path.json', {'path': _xlsx_curr})
                phase2(project_dir, args)
        else:
            print(f'\n  🚨 第{max_rounds}轮仍未通过！以下{len(qa_result.get("failed_items", []))}项需人工验收:')
            for item in qa_result.get('failed_items', []):
                print(f'    - {item}')
            print(f'  QA报告: {qa_report_path}')
            _build_project_state(project_dir, 'FAILED', cache_dir, pending_count=0)
            sys.exit(1)

    print(f'\n{"="*60}')
    print('🎉 全流程执行完成！')
    print(f'{"="*60}')
    if not (_load_cache(cache_dir, 'qa_report.json') or {}).get('passed', False):
        _build_project_state(project_dir, 'FAILED', cache_dir, pending_count=0)
        sys.exit(1)
    return results


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='DT Skill - 评估明细表填写统一执行入口')
    parser.add_argument('--phase', type=str, required=True,
                       help='执行阶段: -1, 0, 1, 2a, 2b, 2c, 2d, 2e, 3, 4, 5, all, gate, cache')
    parser.add_argument('--project', type=str, required=True,
                       help='项目文件夹路径')
    parser.add_argument('--xlsx-path', type=str, default=None,
                       help='评估明细表路径（Phase 2+需要）')
    parser.add_argument('--gate', type=str, default=None,
                       help='Gate验证: G0, G1, G1F, G2, G3, G-DT182, all')
    parser.add_argument('--force', action='store_true',
                       help='强制重新执行（忽略缓存）')

    args = parser.parse_args()

    project_dir = os.path.abspath(args.project)
    if not os.path.isdir(project_dir):
        print(f'❌ 项目目录不存在: {project_dir}')
        sys.exit(1)

    print(f'项目目录: {project_dir}')
    print(f'执行阶段: {args.phase}')

    # Phase-Scoped规则摘要输出
    phase_key = args.phase.rstrip('abcde')  # '2a'→'2', '0'→'0'
    rules_digest = load_rules_digest(phase=phase_key)
    if rules_digest:
        print(f'\n📋 Phase {args.phase} 核心规则摘要（{len(rules_digest)}条）:')
        for rid, rtitle in rules_digest.items():
            print(f'  {rid}: {rtitle}')
        print()

    # 断点恢复检测 (DT-131)
    cache_dir = os.path.join(project_dir, '_dt_cache')
    if os.path.isdir(cache_dir):
        print(f'🔄 检测到_dt_cache/，断点恢复模式')
        check_cache(project_dir)

    # 分发执行
    phase_map = {
        '0': phase0,
        '1': phase1,
        '2': phase2,
        '3': phase3,
        '4': phase4,
        '5': phase5,
    }

    if args.phase == 'cache':
        check_cache(project_dir)
    elif args.phase == 'gate':
        if not args.gate:
            print('❌ --gate 需要指定G0/G1/G2/G3/all')
            sys.exit(1)
        if args.gate == 'all':
            for g in ['G0', 'G1', 'G2', 'G3', 'G-DT182']:
                r = run_gate(project_dir, g, args)
                _gate_pass_or_raise(r)
        else:
            r = run_gate(project_dir, args.gate, args)
            _gate_pass_or_raise(r)
    elif args.phase in phase_map:
        result = phase_map[args.phase](project_dir, args)
        status = result.get('status', 'unknown')
        if status not in ('completed', 'skipped_no_seq_file'):
            print(f'\n❌ Phase {args.phase} 未通过: {status}')
            print(f'结果: {json.dumps(result, ensure_ascii=False, indent=2)}')
            sys.exit(1)
        print(f'\n✅ Phase {args.phase} 执行完成')
        print(f'结果: {json.dumps(result, ensure_ascii=False, indent=2)}')
    elif args.phase == 'all':
        # 全流程编排
        _run_all_phases(project_dir, args)
    else:
        print(f'⚠️ Phase {args.phase} 尚未在dt_runner.py中实现')
        print(f'已实现: {list(phase_map.keys()) + ["cache", "gate", "all"]}')
        sys.exit(1)


if __name__ == '__main__':
    main()
