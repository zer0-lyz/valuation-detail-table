"""
评估明细表 — Phase 2 每Sheet填写后即时验证脚本

设计原则：
- Phase 2每填完一个Sheet，MUST立即调用本脚本验证
- 任一R类(红线)检查未通过 → 禁止填写下一个Sheet，必须先修复
- W类(警告)检查未通过 → 可继续但MUST在最终交付前修复
- 本脚本不调用save，不修改文件，纯只读验证

调用方式：
    python validate_sheet_after_fill.py <xlsx_path> <sheet_name>

返回：
    exit 0 = 全部通过
    exit 1 = 有红线问题，必须修复
    exit 2 = 有警告，建议修复
"""

import sys
import re
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from datetime import datetime

# 从valuation-common导入共享工具函数
sys.path.insert(0, os.path.expanduser('~/.codex/skills/valuation-detail-table/valuation-common/scripts'))
from shared_utils import get_sheet_prefix
from sheet_col_finder import find_header_cols, get_formula_cols, get_amount_cols


# ============================================================
# 列位动态查找（替代原KNOWN_SHEET_STRUCTURES硬编码字典）
# 原字典50+个列号全部硬编码，不同sheet列位差异极大，必须动态识别。
# ============================================================

def _build_sheet_structure(ws):
    """根据表头动态构建sheet结构信息（替代KNOWN_SHEET_STRUCTURES硬编码）。
    
    返回与原KNOWN_SHEET_STRUCTURES兼容的字典格式：
    - formula_cols: 含公式的列号列表（增值额+增值率+账龄）
    - date_cols: 日期列号列表
    - content_cols: 业务内容列号列表
    - cols: 总列数
    """
    header_cols = find_header_cols(ws)
    
    formula_cols = get_formula_cols(ws, header_cols)
    
    date_cols = []
    for key in ['发生日期', '购置日期', '启用日期', '到期日']:
        if key in header_cols:
            date_cols.append(header_cols[key])
    
    content_cols = []
    for key in ['业务内容']:
        if key in header_cols:
            content_cols.append(header_cols[key])
    
    cols = 0
    for c in range(1, min(ws.max_column + 1, 25)):
        cl = get_column_letter(c)
        if ws.column_dimensions[cl].hidden:
            continue
        cell5 = ws.cell(row=5, column=c)
        cell6 = ws.cell(row=6, column=c)
        if (not isinstance(cell5, MergedCell) and cell5.value) or \
           (not isinstance(cell6, MergedCell) and cell6.value):
            cols = c
    
    return {
        'cols': cols,
        'formula_cols': formula_cols,
        'date_cols': date_cols,
        'content_cols': content_cols,
    }


def find_key_rows(ws):
    """找到表头行、合计行、数据行范围
    
    多行表头处理：某些Sheet（如4-8-x固定资产子表）有Row5主标题+Row6子标题，
    数据行从Row7开始。本函数通过检测Row6是否为子标题行来智能判断数据起始行。
    """
    header_row = None
    sub_header_row = None
    total_row = None
    total_rows = []
    first_data_row = None
    last_data_row = None

    for r in range(1, min(ws.max_row + 1, 300)):
        a_val = ws.cell(row=r, column=1).value
        if a_val and str(a_val).strip() == '序号':
            header_row = r
            # 检查下一行是否为子标题行（如4-8-x的Row6含"原值""净值"等）
            if r + 1 <= ws.max_row:
                next_vals = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    cv = ws.cell(row=r + 1, column=c).value
                    if cv and not isinstance(ws.cell(row=r + 1, column=c), MergedCell):
                        next_vals.append(str(cv).strip())
                next_text = ''.join(next_vals)
                if any(kw in next_text for kw in ['原值', '净值', '成新率', '增值额', '增值率']):
                    sub_header_row = r + 1
        
        if a_val and '合' in str(a_val).replace(' ', ''):
            # 排除标题行（如"合同负债评估明细表"含"合"字但不是合计行）
            # 标题行特征：行号<表头行，且含"评估明细表"/"汇总"等词
            a_text_full = str(a_val).replace(' ', '').strip()
            if '评估明细表' in a_text_full or '汇总表' in a_text_full:
                continue
            # 只在表头行之后的行才算合计行
            if header_row and r < header_row:
                continue
            total_rows.append(r)
            total_row = r
            
        if header_row:
            data_start_candidate = (sub_header_row or header_row) + 1
            if r >= data_start_candidate and a_val is not None:
                # 排除子标题行
                if sub_header_row and r == sub_header_row:
                    continue
                b_val = ws.cell(row=r, column=2).value
                if b_val and str(b_val).strip() and '合' not in str(a_val).replace(' ', ''):
                    if first_data_row is None:
                        first_data_row = r
                    last_data_row = r

    return header_row, total_row, total_rows, first_data_row, last_data_row




def validate_sheet(filepath, sheet_name):
    """验证单个Sheet的数据质量"""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' 不存在于文件中")
        wb.close()
        return False

    ws = wb[sheet_name]
    issues_r = []  # 红线问题
    issues_w = []  # 警告

    header_row, total_row, total_rows, first_data_row, last_data_row = find_key_rows(ws)
    # 获取sub_header_row（find_key_rows中通过ws.cell直接检测）
    sub_header_row = None
    if header_row and header_row + 1 <= ws.max_row:
        next_vals = []
        for c in range(1, min(ws.max_column + 1, 20)):
            cv = ws.cell(row=header_row + 1, column=c).value
            if cv and not isinstance(ws.cell(row=header_row + 1, column=c), MergedCell):
                next_vals.append(str(cv).strip())
        next_text = ''.join(next_vals)
        if any(kw in next_text for kw in ['原值', '净值', '成新率', '增值额', '增值率']):
            sub_header_row = header_row + 1
    
    prefix = get_sheet_prefix(sheet_name)
    known_struct = _build_sheet_structure(ws)  # 动态构建，替代KNOWN_SHEET_STRUCTURES硬编码

    if not header_row:
        issues_w.append(f"[{sheet_name}] 未找到表头行(序号行)，跳过验证")
        wb.close()
        return True

    data_start = (sub_header_row + 1) if sub_header_row else (header_row + 1)
    data_end = (total_row or ws.max_row) - 1
    if total_row and data_end > total_row:
        data_end = total_row - 1
    
    # 过滤掉减值/小计行和子标题行
    data_rows = []
    SUB_HEADER_KEYWORDS = {'原值', '净值', '成新率%', '增值额', '增值率%', '%'}
    for r in range(data_start, data_end + 1):
        # 跳过子标题行
        if sub_header_row and r == sub_header_row:
            continue
        a_val = ws.cell(row=r, column=1).value
        a_text = str(a_val).replace(' ', '').strip() if a_val else ''
        if '减' in a_text or ('小' in a_text and '计' in a_text):
            continue
        # 检查是否有实际数据（排除只有子标题文本的行）
        has_data = False
        for c in range(1, min(ws.max_column + 1, 19)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is not None and not (isinstance(v, str) and v.strip() in SUB_HEADER_KEYWORDS):
                has_data = True
                break
        if has_data:
            data_rows.append(r)

    # ============================================================
    # 检查1 [R]: 公式列未被覆写 (DT-67)
    # ============================================================
    if known_struct and known_struct['formula_cols']:
        for r in data_rows:
            for fc in known_struct['formula_cols']:
                cell = ws.cell(row=r, column=fc)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is not None and not (isinstance(val, str) and val.startswith('=')):
                    col_letter = get_column_letter(fc)
                    issues_r.append(
                        f"[DT-67][{sheet_name}] {col_letter}{r} 公式列被数值覆写: value={val} "
                        f"(应为公式，如'=L{r}-J{r}'或'=IF(...)')"
                    )

    # ============================================================
    # 检查2 [R]: 发生日期列不含文字 / 业务内容列不含日期序列号 (DT-46)
    # ============================================================
    if known_struct:
        for dc in known_struct.get('date_cols', []):
            for r in data_rows:
                cell = ws.cell(row=r, column=dc)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is not None:
                    if isinstance(val, str) and len(val) > 5 and not any(d in val for d in '0123456789'):
                        col_letter = get_column_letter(dc)
                        issues_r.append(
                            f"[DT-46][{sheet_name}] {col_letter}{r} 发生日期列填入文字: '{val[:30]}'"
                        )
        for cc in known_struct.get('content_cols', []):
            for r in data_rows:
                cell = ws.cell(row=r, column=cc)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if isinstance(val, (int, float)) and val > 40000:
                    col_letter = get_column_letter(cc)
                    issues_r.append(
                        f"[DT-46][{sheet_name}] {col_letter}{r} 业务内容列填入日期序列号: {val}"
                    )

    # ============================================================
    # 检查3 [R]: 列位校验 — 读取实际表头与已知结构对比 (DT-66)
    # ============================================================
    if header_row and known_struct:
        actual_cols = 0
        for c in range(1, 30):
            cell = ws.cell(row=header_row, column=c)
            # 检查header_row+1（子标题行）是否有合并单元格延伸
            if not isinstance(cell, MergedCell) and cell.value is not None:
                actual_cols = c
            # 也检查子标题行
            sub_cell = ws.cell(row=header_row + 1, column=c) if header_row + 1 <= ws.max_row else None
            if sub_cell and not isinstance(sub_cell, MergedCell) and sub_cell.value is not None:
                actual_cols = max(actual_cols, c)
        
        # 列数差异过大（超过2列）= 列映射可能错误
        if actual_cols > 0 and known_struct['cols'] > 0:
            diff = abs(actual_cols - known_struct['cols'])
            if diff > 2:
                issues_w.append(
                    f"[DT-66][{sheet_name}] 实际列数({actual_cols})与已知结构({known_struct['cols']})差异>{2}，"
                    f"请确认列位映射是否正确"
                )

    # ============================================================
    # 检查4 [R]: 数据行无"待补充"/"待核对"占位符 (DT-0延伸)
    # ============================================================
    for r in data_rows:
        for c in range(1, min(ws.max_column + 1, 19)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if isinstance(val, str) and ('待补充' in val or '待核对' in val):
                col_letter = get_column_letter(c)
                issues_w.append(
                    f"[DT-0][{sheet_name}] {col_letter}{r} 残留占位符: '{val}'"
                )

    # ============================================================
    # 检查5 [R]: 合计行唯一性 (DT-35)
    # 正常模式："合计"(小计) → "减：xxx" → "合计"(净合计) 三行，
    # 即两个"合计"行中间夹一个"减"行是标准结构，不算重复。
    # 只有连续两个不含"减"间隔的"合计"行才算真正重复。
    # ============================================================
    pure_total_rows = []  # 不含"减"的合计行
    for tr in total_rows:
        a_val = ws.cell(row=tr, column=1).value
        a_text = str(a_val).replace(' ', '').strip() if a_val else ''
        if '减' in a_text:
            continue
        if '合' in a_text:
            pure_total_rows.append(tr)
    # 检查是否有两个相邻的"合计"行中间没有"减"行间隔
    duplicate_totals = []
    for i in range(len(pure_total_rows) - 1):
        r1 = pure_total_rows[i]
        r2 = pure_total_rows[i + 1]
        # 检查中间是否有"减"行
        has_jian = False
        for r in range(r1 + 1, r2):
            a_val = ws.cell(row=r, column=1).value
            if a_val and '减' in str(a_val):
                has_jian = True
                break
        if not has_jian:
            duplicate_totals.append((r1, r2))
    if duplicate_totals:
        for r1, r2 in duplicate_totals:
            issues_r.append(
                f"[DT-35][{sheet_name}] 行{r1}和行{r2}都是合计行且中间无减值行间隔，汇总值可能翻倍"
            )

    # ============================================================
    # 检查6 [W]: 数据行与合计行之间无空白无格式行 (DT-44)
    # ============================================================
    if last_data_row and total_row and total_row > last_data_row + 1:
        for r in range(last_data_row + 1, total_row):
            a_val = ws.cell(row=r, column=1).value
            if a_val and '减' in str(a_val):
                continue  # 减值行是正常的
            # 检查是否完全空白且无格式
            all_none = all(ws.cell(row=r, column=c).value is None for c in range(1, min(ws.max_column + 1, 19)))
            if all_none:
                has_border = False
                for c in range(1, min(ws.max_column + 1, 19)):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell) and cell.border and cell.border.top and cell.border.top.style:
                        has_border = True
                        break
                if not has_border:
                    issues_w.append(
                        f"[DT-44][{sheet_name}] 行{r}为空白无格式行，位于数据行与合计行之间"
                    )

    # ============================================================
    # 检查7 [W]: 关键数值列格式 (DT-3)
    # ============================================================
    if known_struct:
        money_cols = []
        for c in range(1, known_struct['cols'] + 1):
            if header_row:
                hval = ws.cell(row=header_row, column=c).value
                if hval and any(kw in str(hval) for kw in ['价值', '原值', '净值', '余额', '金额', '增值额']):
                    money_cols.append(c)
        
        for r in data_rows[:5]:  # 抽查前5行
            for mc in money_cols:
                cell = ws.cell(row=r, column=mc)
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)):
                    nf = cell.number_format
                    if nf == 'General' or not nf:
                        col_letter = get_column_letter(mc)
                        issues_w.append(
                            f"[DT-3][{sheet_name}] {col_letter}{r} 金额列数字格式为General，应为#,##0.00"
                        )

    # ============================================================
    # 检查8 [R]: 增值额/增值率列格式 (DT-76)
    # 模板中增值额和增值率列都使用财务会计格式 #,##0.00_);[Red]\-#,##0.00_);_(* ""_)
    # 注意：增值率公式乘以100（如=K/I*100），所以格式与增值额相同，非0.00%
    # 检查标准：formula_cols中的列MUST不是General格式
    # ============================================================
    if known_struct and known_struct['formula_cols']:
        for fc in known_struct['formula_cols']:
            # 抽查前3行数据
            for r in data_rows[:3]:
                cell = ws.cell(row=r, column=fc)
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                # 只检查有值的单元格
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    continue  # 非公式行跳过（已在DT-67检查）
                nf = cell.number_format
                col_letter = get_column_letter(fc)
                # 格式为General或空=红线
                if nf == 'General' or not nf:
                    issues_r.append(
                        f"[DT-76][{sheet_name}] {col_letter}{r} 公式列格式为General，应为财务会计格式"
                    )

    # ============================================================
    # 检查9 [R/W]: 行高统一性 (DT-77)
    # 同一Sheet内所有数据行行高MUST基本一致
    # 以众数（出现次数最多的行高）为参考，少量行差异为警告，大量差异为红线
    # 不同Sheet的模板默认行高可能不同（16.5/17.25/18.0/18.75等），不强制统一值
    # ============================================================
    if data_rows:
        # 收集所有数据行行高
        height_counts = {}
        for r in data_rows:
            h = ws.row_dimensions[r].height
            if h is not None:
                h_rounded = round(h, 1)
                height_counts[h_rounded] = height_counts.get(h_rounded, 0) + 1
        
        if height_counts:
            # 取众数作为参考行高
            ref_height = max(height_counts, key=height_counts.get)
            total_counted = sum(height_counts.values())
            ref_count = height_counts[ref_height]
            
            # 如果众数占比<60%，说明行高非常混乱
            if ref_count / total_counted < 0.6:
                issues_r.append(
                    f"[DT-77][{sheet_name}] 行高严重不一致: {height_counts}，应以最常见行高{ref_height}统一"
                )
            # 检查非众数行高是否为模板预设（如3-1-2前3行18.75后16.5）
            # 如果非众数行高占>20%且>=5行，报告警告
            elif ref_count / total_counted < 0.8 and (total_counted - ref_count) >= 5:
                issues_w.append(
                    f"[DT-77][{sheet_name}] 部分行高不一致: {height_counts}，建议检查是否为模板预设"
                )

    # ============================================================
    # 检查10 [W]: 合计行下方无残留格式 (DT-78②)
    # ============================================================
    if total_row:
        for r in range(total_row + 1, min(total_row + 6, ws.max_row + 1)):
            has_border = False
            has_value = False
            for c in range(1, min(ws.max_column + 1, 19)):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    has_value = True
                if cell.border and cell.border.top and cell.border.top.style:
                    has_border = True
                if cell.border and cell.border.bottom and cell.border.bottom.style:
                    has_border = True
            if has_border and not has_value:
                issues_w.append(
                    f"[DT-78][{sheet_name}] 行{r}(合计行下方)有边框但无数据，应清除残留格式"
                )

    # ============================================================
    # 检查11 [R]: 数据行首行无空白跳过 (DT-82①)
    # 数据MUST从第一个数据行开始填写，禁止首行空白跳过
    # 检查方法：data_start行（表头/子表头的下一行）是否有实际数据
    # ============================================================
    if data_start and total_row:
        # 第一个数据行应该是 data_start
        first_row = ws.cell(row=data_start, column=1)
        first_row_b = ws.cell(row=data_start, column=2)
        # 如果第一行的A列和B列都为空，但后续行有数据→首行被跳过
        a_val = first_row.value if not isinstance(first_row, MergedCell) else None
        b_val = first_row_b.value if not isinstance(first_row_b, MergedCell) else None
        if a_val is None and b_val is None and data_rows and data_rows[0] > data_start:
            issues_r.append(
                f"[DT-82-1][{sheet_name}] 行{data_start}为首行空白(A/B列均空),数据从行{data_rows[0]}开始填写,"
                f"违反数据从第一个数据行开始规则"
            )

    # ============================================================
    # 检查12 [W]: 空白数据行格式完整性 (DT-82②)
    # 数据行区域内空白行应有thin边框和公式
    # ============================================================
    if data_start and total_row:
        for r in range(data_start, total_row):
            # 跳过有数据的行和减值/小计行
            a_val = ws.cell(row=r, column=1).value
            a_text = str(a_val).replace(' ', '').strip() if a_val else ''
            if '减' in a_text or ('小' in a_text and '计' in a_text):
                continue
            # 检查是否为空白行
            all_none = True
            for c in range(1, min(ws.max_column + 1, 19)):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    all_none = False
                    break
            if all_none:
                # 空白行应有thin边框
                has_border = False
                for c in range(1, min(ws.max_column + 1, 19)):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.border and cell.border.top and cell.border.top.style == 'thin':
                        has_border = True
                        break
                if not has_border:
                    issues_w.append(
                        f"[DT-82-2][{sheet_name}] 行{r}为数据区空白行但无边框,应保留thin边框和公式"
                    )

    # ============================================================
    # 检查13 [R]: 结构行A列居中对齐 (DT-84)
    # 合计/减值/小计行的A列MUST为center对齐
    # ============================================================
    if data_start and total_row:
        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            if a_val is None:
                continue
            a_text = str(a_val).replace(' ', '').strip()
            is_struct = ('合' in a_text and '计' in a_text) or a_text.startswith('减')
            if is_struct:
                align = ws.cell(row=r, column=1).alignment.horizontal
                if align != 'center':
                    issues_r.append(
                        f"[DT-84][{sheet_name}] A{r}为结构行(合计/减值)但alignment={align},应为center"
                    )

    # ============================================================
    # 检查14 [R]: 合计行公式MUST引用本行而非前一行 (DT-85)
    # 增值额/增值率公式必须引用合计行自身的账面价值和评估价值
    # 禁止引用合计行上方分隔行的空值
    # ============================================================
    if known_struct and known_struct['formula_cols'] and total_row:
        for tr in total_rows:
            a_val = ws.cell(row=tr, column=1).value
            a_text = str(a_val).replace(' ', '').strip() if a_val else ''
            if '减' in a_text:
                continue  # 减值行的公式逻辑不同，跳过
            # 检查合计行的公式列是否引用了本行
            for fc in known_struct['formula_cols']:
                cell = ws.cell(row=tr, column=fc)
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                val = str(cell.value)
                if not val.startswith('='):
                    continue
                col_letter = get_column_letter(fc)
                # 增值额公式应含本行号（如 =F27-E27 或 =I{tr}-G{tr}）
                # 增值率公式应含本行号（如 =IF(E27=0,"",G27/E27*100)）
                if str(tr) not in val:
                    issues_r.append(
                        f"[DT-85][{sheet_name}] {col_letter}{tr} 合计行公式未引用本行号{tr}: {val[:60]}"
                    )

    wb.close()

    # ============================================================
    # 输出结果
    # ============================================================
    sheet_label = f"[{sheet_name}]"
    
    if issues_r:
        print(f"\n🚨 {sheet_label} 即时验证未通过 — {len(issues_r)}个红线问题:")
        for i in issues_r:
            print(f"  {i}")
        if issues_w:
            print(f"\n⚠️ {sheet_label} 另有{len(issues_w)}个警告:")
            for i in issues_w:
                print(f"  {i}")
        return False
    
    if issues_w:
        print(f"\n⚠️ {sheet_label} 即时验证通过但有{len(issues_w)}个警告:")
        for i in issues_w:
            print(f"  {i}")
        return True
    
    print(f"\n✅ {sheet_label} 即时验证通过，所有检查项合格")
    return True


def validate_all_sheets(filepath):
    """验证所有明细表Sheet（Phase 4.5用途）"""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    all_pass = True
    any_warning = False
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.sheet_state == 'hidden':
            continue
        # 跳过汇总表和系统表
        if '汇总' in sname or sname.startswith('0') or sname.startswith('2-') or sname == '设置' or sname == '设定信息':
            continue
        
        prefix = get_sheet_prefix(sname)
        if not prefix:
            continue
        
        result = validate_sheet(filepath, sname)
        if not result:
            all_pass = False
    
    wb.close()
    return all_pass


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python validate_sheet_after_fill.py <xlsx_path> [sheet_name]")
        print("  指定sheet_name: 验证单个Sheet")
        print("  不指定: 验证所有明细表Sheet")
        sys.exit(2)
    
    filepath = sys.argv[1]
    
    if len(sys.argv) >= 3:
        sheet_name = sys.argv[2]
        passed = validate_sheet(filepath, sheet_name)
    else:
        passed = validate_all_sheets(filepath)
    
    if passed:
        sys.exit(0)
    else:
        sys.exit(1)
