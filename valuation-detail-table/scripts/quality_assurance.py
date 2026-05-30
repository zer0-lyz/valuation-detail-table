#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
quality_assurance.py — 验收流程：填写完成后自动质检

在 Phase 2→3→4→5 全流程完成后执行，检查6个维度：
  1. BS_RECONCILIATION  — 分类汇总 I列 vs BS差异 < 0.5%
  2. FIELD_COMPLETENESS  — 结算对象/设备编号/发生日期/金额是否完整
  3. TOTAL_CROSS_CHECK   — 汇总表合计 vs 明细合计行是否吻合
  4. BLANK_SHEET_HIDDEN  — 所有空表是否被隐藏
  5. FORMAT_INTEGRITY    — 序号连续、SUM公式覆盖正确
  6. ASSET_CLASSIFICATION — 固定资产原值合计 vs BS

失败时自动触发修复+重检，最多3轮，超限转人工。
"""

import json, os, sys, glob, re, logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
except ImportError:
    openpyxl = None

# ============================================================
# QA 检查项定义
# ============================================================

CHECK_BS_RECONCILIATION = 'BS_RECONCILIATION'   # 分类汇总 I列 vs BS
CHECK_FIELD_COMPLETENESS = 'FIELD_COMPLETENESS'   # 字段内容齐全
CHECK_TOTAL_CROSS = 'TOTAL_CROSS_CHECK'           # 汇总交叉校验
CHECK_BLANK_HIDDEN = 'BLANK_SHEET_HIDDEN'         # 空白表隐藏
CHECK_FORMAT = 'FORMAT_INTEGRITY'                 # 格式完整性
CHECK_ASSET = 'ASSET_CLASSIFICATION'              # 固定资产分类

QA_LEVELS = {
    CHECK_BS_RECONCILIATION: {'name': '报表校对', 'tolerance': 0.005},  # 0.5%
    CHECK_FIELD_COMPLETENESS: {'name': '字段齐全', 'tolerance': 0.05},  # 5%缺失容忍
    CHECK_TOTAL_CROSS: {'name': '汇总校验', 'tolerance': 0.005},
    CHECK_BLANK_HIDDEN: {'name': '空白表隐藏', 'tolerance': 1.0},      # 1个未隐藏=OK
    CHECK_FORMAT: {'name': '格式完整性', 'tolerance': 0.05},
    CHECK_ASSET: {'name': '固定资产分类', 'tolerance': 0.01},
}

# ============================================================
# 核心函数
# ============================================================

def run_qa(project_dir, cache_dir=None, round_num=0, max_rounds=3):
    """执行完整的验收流程

    Args:
        project_dir: 项目文件夹路径
        cache_dir: 缓存目录（默认 project_dir/_dt_cache）
        round_num: 当前轮次
        max_rounds: 最大自动重试轮次

    Returns:
        dict: {
            'passed': bool,          # 全部通过？
            'round': int,            # 当前轮次
            'checks': {check_id: {pass, detail, score}},
            'summary': str,
            'failed_items': [str],   # 失败项清单
            'recommend_manual': bool # 建议人工介入
        }
    """
    if cache_dir is None:
        cache_dir = os.path.join(project_dir, '_dt_cache')
    
    result = {
        'passed': True,
        'round': round_num,
        'max_rounds': max_rounds,
        'checks': {},
        'summary': '',
        'failed_items': [],
        'recommend_manual': False,
        'timestamp': datetime.now().isoformat(),
    }
    
    # 找输出文件
    xlsx_path = _find_output(project_dir, cache_dir)
    if not xlsx_path:
        result['passed'] = False
        result['summary'] = '❌ 未找到评估明细表输出文件'
        result['recommend_manual'] = True
        return result
    
    # 加载BS数据
    bs_data = _load_bs_data(cache_dir)
    if not bs_data:
        result['passed'] = False
        result['summary'] = '❌ 未找到bs_balances.json'
        result['recommend_manual'] = True
        return result
    
    wb = openpyxl.load_workbook(xlsx_path)
    
    # 逐项检查
    checks_results = {}
    errors = []
    
    # 1. BS校对
    c1 = _check_bs_reconciliation(wb, bs_data)
    checks_results[CHECK_BS_RECONCILIATION] = c1
    if not c1['pass']:
        errors.append(f"报表校对: {c1['detail']}")
    
    # 2. 字段齐全
    c2 = _check_field_completeness(wb)
    checks_results[CHECK_FIELD_COMPLETENESS] = c2
    if not c2['pass']:
        errors.append(f"字段齐全: {c2['detail']}")
    
    # 3. 汇总交叉校验
    c3 = _check_total_cross(wb)
    checks_results[CHECK_TOTAL_CROSS] = c3
    if not c3['pass']:
        errors.append(f"汇总校验: {c3['detail']}")
    
    # 4. 空白表隐藏
    c4 = _check_blank_hidden(wb)
    checks_results[CHECK_BLANK_HIDDEN] = c4
    if not c4['pass']:
        errors.append(f"空白表隐藏: {c4['detail']}")
    
    # 5. 格式完整性
    c5 = _check_format_integrity(wb)
    checks_results[CHECK_FORMAT] = c5
    if not c5['pass']:
        errors.append(f"格式完整性: {c5['detail']}")
    
    # 6. 固定资产分类
    c6 = _check_asset_classification(wb, bs_data)
    checks_results[CHECK_ASSET] = c6
    if not c6['pass']:
        errors.append(f"固定资产: {c6['detail']}")
    
    result['checks'] = checks_results
    result['failed_items'] = errors
    
    # 计算总分
    passed_count = sum(1 for c in checks_results.values() if c['pass'])
    total_count = len(checks_results)
    result['score'] = f"{passed_count}/{total_count}"
    
    if errors:
        result['passed'] = False
        if round_num >= max_rounds - 1:
            result['recommend_manual'] = True
            result['summary'] = f"⚠️ 第{round_num+1}轮仍有{len(errors)}项未通过，建议人工验收"
        else:
            result['summary'] = f"🔄 第{round_num+1}轮发现{len(errors)}项问题，将自动修复并重检"
    else:
        result['summary'] = f"✅ 全部{passed_count}/{total_count}项检查通过！"
    
    wb.close()
    return result


def auto_fix(project_dir, checks, cache_dir=None):
    """根据QA失败项执行自动修复

    Returns:
        dict: {check_id: {fixed: bool, action: str}}
    """
    fixes = {}
    
    # TODO: 实现自动修复逻辑
    # 目前框架只记录需要修复的项，实际修复调用已有Phase函数
    
    for check_id, check_result in checks.items():
        if check_result['pass']:
            continue
        fixes[check_id] = {'fixed': False, 'action': f'需要人工处理: {check_id}'}
    
    return fixes


# ============================================================
# 内部检查函数
# ============================================================

def _find_output(project_dir, cache_dir):
    """查找评估明细表输出文件"""
    # 先从cache读取
    xlsx_cache = os.path.join(cache_dir, 'xlsx_path.json')
    if os.path.exists(xlsx_cache):
        with open(xlsx_cache) as f:
            info = json.load(f)
        path = info.get('path', '')
        if path and os.path.exists(path):
            return path
    
    # 从项目目录搜索
    candidates = glob.glob(os.path.join(project_dir, '*评估明细表*.xlsx'))
    if candidates:
        return sorted(candidates, key=os.path.getmtime)[-1]
    return None


def _load_bs_data(cache_dir):
    """加载资产负债表数据"""
    path = os.path.join(cache_dir, 'bs_balances.json')
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return json.load(f)


def _check_bs_reconciliation(wb, bs_data):
    """检查1: 分类汇总 I列 vs BS"""
    if '2-分类汇总' not in wb.sheetnames:
        return {'pass': False, 'detail': '2-分类汇总 sheet不存在', 'score': 0}
    
    ws = wb['2-分类汇总']
    bs_items = {str(item.get('label', '')).replace(' ', '').replace('\u3000', ''): item.get('ending_balance', 0)
                for item in bs_data.get('items', []) if item.get('ending_balance')}
    
    mismatches = []
    total_items = 0
    matching_items = 0
    
    for r in range(6, ws.max_row + 1):
        name = ws.cell(row=r, column=4).value  # C4=科目名称
        i_val = ws.cell(row=r, column=9).value  # I=报表金额
        
        if not name or str(name).strip() in ('', 'None'):
            continue
        name_s = str(name).strip()
        if name_s.startswith(('一', '二', '三', '四', '五', '六', '七')):
            continue
        
        name_key = name_s.replace(' ', '').replace('\u3000', '')
        
        if name_key in bs_items:
            total_items += 1
            bs_val = bs_items[name_key]
            if isinstance(i_val, (int, float)):
                if abs(bs_val) < 0.01 and abs(i_val) < 0.01:
                    matching_items += 1
                else:
                    diff_pct = abs(bs_val - i_val) / max(abs(bs_val), 0.01)
                    if diff_pct > 0.005:  # 0.5%
                        mismatches.append(f"{name_s}: BS={bs_val:,.2f}, I列={i_val:,.2f}, diff={diff_pct:.2%}")
                    else:
                        matching_items += 1
    
    score = matching_items / max(total_items, 1)
    passed = score >= 0.95 and len(mismatches) == 0
    
    detail = f"{matching_items}/{total_items}匹配"
    if mismatches:
        detail += f", {len(mismatches)}项差异: {'; '.join(mismatches[:5])}"
    
    return {'pass': passed, 'detail': detail, 'score': score, 'mismatches': mismatches}


def _check_field_completeness(wb):
    """检查2: 字段内容齐全"""
    issues = []
    
    # 需要检查的关键字段 per sheet
    checks = {
        '3-1-2银行存款': [('结算对象(开户银行)', 3, 'str'), ('账面价值', 8, 'num')],
        '3-5应收账款': [('结算对象', 3, 'str'), ('账面价值', 10, 'num')],
        '4-8-4机器设备': [('设备名称', 4, 'str'), ('设备编号', 3, 'str'), ('账面价值', 11, 'num')],
        '4-8-5车辆': [('车辆牌号', 3, 'str'), ('车辆名称', 4, 'str'), ('账面价值', 11, 'num')],
        '4-8-6电子设备': [('设备名称', 4, 'str'), ('设备编号', 3, 'str'), ('账面价值', 11, 'num')],
        '5-5应付账款': [('结算对象', 3, 'str'), ('账面价值', 10, 'num')],
        '5-10-3其他应付款': [('结算对象', 3, 'str'), ('账面价值', 9, 'num')],
    }
    
    for sheet_name, fields in checks.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for field_name, col, ftype in fields:
            for r in range(6, min(ws.max_row + 1, 200)):
                seq = ws.cell(row=r, column=2).value
                if not seq or not isinstance(seq, (int, float)) or seq < 1:
                    continue
                # Skip 合计/小计/坏账准备行
                a_val = str(ws.cell(row=r, column=1).value or '').strip()
                if a_val in ('合计1', '合计2', '坏账准备') or '合' in str(seq) or '计' in str(seq):
                    continue
                val = ws.cell(row=r, column=col).value
                # For str type fields, '0' is a valid value (e.g. asset code '0' = no code)
                if val is None or str(val).strip() in ('', 'None'):
                    issues.append(f"{sheet_name}: R{r}C{col}({field_name})为空")
                elif ftype == 'num' and isinstance(val, (int, float)) and abs(val) < 0.01:
                    issues.append(f"{sheet_name}: R{r}C{col}({field_name})为零")
                    issues.append(f"{sheet_name}: R{r}C{col}({field_name})为空")
    
    passed = len(issues) <= max(3, len(checks) * 2 * 0.05)  # 容忍少量缺失
    detail = f"{len(issues)}个字段为空" if issues else "全部完整"
    
    return {'pass': passed, 'detail': detail, 'issues': issues[:20]}


def _check_total_cross(wb):
    """检查3: 汇总交叉校验"""
    issues = []
    
    # 检查2-分类汇总的合计行
    if '2-分类汇总' not in wb.sheetnames:
        return {'pass': False, 'detail': '2-分类汇总不存在', 'score': 0}
    
    ws = wb['2-分类汇总']
    
    # 检查E列(账面价值)合计行
    for r in [6, 20, 39, 40, 54, 64, 65]:
        e_val = ws.cell(row=r, column=5).value
        if e_val is None:
            continue
        e_str = str(e_val)
        if not e_str.startswith('=') and not e_str.startswith('SUM'):
            issues.append(f"R{r} E列非公式: {e_str}")
    
    # 检查各汇总sheet
    summary_sheets = [sn for sn in wb.sheetnames if '汇总' in sn and not sn.startswith('2-')]
    for sn in summary_sheets:
        ws_sum = wb[sn]
        # 检查合计行是否有SUM公式
        for r in range(1, min(ws_sum.max_row + 1, 20)):
            a_val = str(ws_sum.cell(row=r, column=1).value or '')
            if '合计' in a_val and '1' in a_val:
                # 找到合计1行，检查各列是否为公式
                for c in range(5, min(ws_sum.max_column + 1, 15)):
                    v = ws_sum.cell(row=r, column=c).value
                    if v is not None and not str(v).startswith('='):
                        issues.append(f"{sn} R{r}C{c}非公式: {str(v)[:20]}")
                break
    
    passed = len(issues) == 0
    detail = f"{len(issues)}个异常" if issues else "全部合计行公式正确"
    
    return {'pass': passed, 'detail': detail, 'issues': issues[:10]}


def _check_blank_hidden(wb):
    """检查4: 空白表隐藏"""
    visible_empty = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state == 'hidden':
            continue
        if sn.startswith(('2-', '0-')) or '汇总' in sn or sn in ('设置', '目录', '公式数据表', '设定信息'):
            continue
        
        # 检查是否有任何金额数据
        has_data = False
        for r in range(6, min(ws.max_row + 1, 200)):
            for c in range(5, 15):
                v = ws.cell(row=r, column=c).value
                if v is not None and isinstance(v, (int, float)) and abs(v) > 0.01:
                    has_data = True
                    break
            if has_data:
                break
        
        if not has_data:
            visible_empty.append(sn)
    
    passed = len(visible_empty) == 0
    detail = f"{len(visible_empty)}个可见空表: {', '.join(visible_empty[:5])}" if visible_empty else "全部空表已隐藏"
    
    return {'pass': passed, 'detail': detail, 'visible_empty': visible_empty}


def _check_format_integrity(wb):
    """检查5: 格式完整性"""
    issues = []
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.sheet_state == 'hidden':
            continue
        if sn.startswith(('2-', '0-')) or '汇总' in sn or sn in ('设置', '目录', '公式数据表', '设定信息'):
            continue
        
        # 检查序号连续性
        last_seq = 0
        seq_gaps = 0
        for r in range(6, min(ws.max_row + 1, 200)):
            seq = ws.cell(row=r, column=2).value
            if seq and isinstance(seq, (int, float)) and seq >= 1:
                if seq != last_seq + 1:
                    seq_gaps += 1
                last_seq = seq
        
        if seq_gaps > 2:
            issues.append(f"{sn}: 序号不连续({seq_gaps}处)")
    
    passed = len(issues) <= 2
    detail = f"{len(issues)}个格式问题" if issues else "格式正常"
    
    return {'pass': passed, 'detail': detail, 'issues': issues[:10]}


def _check_asset_classification(wb, bs_data):
    """检查6: 固定资产分类合计 vs BS"""
    # 从2-分类汇总D列(科目名称)+I列(报表金额，硬编码)读取
    if '2-分类汇总' in wb.sheetnames:
        ws_sum = wb['2-分类汇总']
        fa_total = 0
        for r in range(1, min(ws_sum.max_row + 1, 100)):
            label_d = str(ws_sum.cell(row=r, column=4).value or '')  # D列=科目名称
            if '固定资产' in label_d:
                i_val = ws_sum.cell(row=r, column=9).value  # I列=报表金额(硬编码)
                if isinstance(i_val, (int, float)):
                    fa_total += i_val
                break  # 只取第一行"固定资产"
    else:
        fa_total = 0
    
    # BS固定资产
    bs_fa = next((i.get('ending_balance', 0) for i in bs_data.get('items', [])
                  if i.get('label', '') == '固定资产'), 0)
    
    if bs_fa == 0:
        bs_fa = next((i.get('ending_balance', 0) for i in bs_data.get('items', [])
                      if '固定资产' in i.get('label', '') and '累计' not in i.get('label', '')
                      and '在建' not in i.get('label', '')), 0)
    
    if bs_fa and fa_total:
        diff_pct = abs(fa_total - bs_fa) / bs_fa
        passed = diff_pct < 0.01
        detail = f"明细表(I列)={fa_total:,.2f}, BS={bs_fa:,.2f}, 差异={diff_pct:.2%}"
    elif fa_total == 0 and bs_fa == 0:
        passed = True
        detail = "无固定资产"
    elif fa_total == 0 and bs_fa > 0:
        passed = False
        detail = f"明细表I列固定资产=0.00, BS={bs_fa:,.2f}, 请检查2-分类汇总I列数据"
    else:
        passed = False
        detail = f"明细表I列={fa_total:,.2f}, BS={bs_fa:,.2f}, 差异={abs(fa_total-bs_fa):,.2f}"
    
    return {'pass': passed, 'detail': detail, 'score': 1 - abs(fa_total - bs_fa) / max(bs_fa, 0.01) if bs_fa else 1}


def generate_report(qa_result):
    """生成可读的QA报告字符串"""
    lines = [
        "=" * 60,
        f"📋 验收报告 — 第{qa_result['round']+1}轮",
        "=" * 60,
        f"状态: {'✅ 通过' if qa_result['passed'] else '❌ 未通过'}",
        f"得分: {qa_result.get('score', '?')}",
    ]
    
    for check_id, check_result in qa_result['checks'].items():
        check_name = QA_LEVELS.get(check_id, {}).get('name', check_id)
        status = '✅' if check_result['pass'] else '❌'
        lines.append(f"  {status} {check_name}: {check_result['detail']}")
    
    if qa_result['failed_items']:
        lines.append(f"\n⚠️ 失败项 ({len(qa_result['failed_items'])}):")
        for item in qa_result['failed_items']:
            lines.append(f"  - {item}")
    
    if qa_result['recommend_manual']:
        lines.append("\n🚨 建议人工验收")
    
    lines.append(f"\n时间: {qa_result['timestamp']}")
    lines.append("=" * 60)
    
    return '\n'.join(lines)


def save_report(qa_result, cache_dir):
    """保存QA报告到文件"""
    report_path = os.path.join(cache_dir, 'qa_report.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(qa_result, f, ensure_ascii=False, indent=2, default=str)
    return report_path


# ============================================================
# CLI入口
# ============================================================

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='质量验收')
    parser.add_argument('--project', required=True, help='项目文件夹')
    parser.add_argument('--round', type=int, default=0, help='轮次')
    parser.add_argument('--save', action='store_true', help='保存报告')
    args = parser.parse_args()
    
    cache = os.path.join(args.project, '_dt_cache')
    result = run_qa(args.project, cache, round_num=args.round)
    report = generate_report(result)
    print(report)
    
    if args.save:
        path = save_report(result, cache)
        print(f"\n报告已保存: {path}")
