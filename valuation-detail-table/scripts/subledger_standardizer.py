#!/usr/bin/env python3
"""
subledger_standardizer.py — 科目明细账标准化处理器

将多sheet科目明细账（每个科目一个sheet）解析为标准化工件，
按对方科目归类结算对象，供Phase 2/3替代科目余额表汇总数据使用。

输入: 科目明细账202202-202303.xlsx (134 sheets)
输出: _dt_cache/subledger_standardized.json

标准化格式:
{
  "1001": {
    "code": "1001",
    "name": "库存现金",
    "transactions": [
      {"date": "2023-03-31", "voucher": "记-0007", "summary": "零售",
       "counterparty": "22210106", "debit": 9900.0, "credit": 0.0, "direction": "借", "balance": 9900.0}
    ],
    "settlements": {
      "恩施市人民政府": {"debit": 5898480.0, "credit": 0.0, "last_date": "2024-01-15", "summaries": ["货款","工程款"]}
    }
  }
}
"""

import json
import os
import re
import openpyxl
from datetime import datetime
from collections import defaultdict


def _safe_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        try: return float(val.replace(',', '').strip())
        except: return 0.0
    return 0.0


def _extract_code_from_sheetname(sheet_name):
    """从sheet名提取科目编码，如 '1001 库存现金' → '1001'"""
    m = re.match(r'^(\d+)', str(sheet_name).strip())
    return m.group(1) if m else ''


def standardize_subledger(filepath: str, cache_dir: str) -> dict:
    """标准化科目明细账 → 按科目编码索引的结构化数据"""
    print(f'[SUBLEDGER] 正在解析: {os.path.basename(filepath)}')
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {}
    total_sheets = len(wb.sheetnames)
    parsed = 0
    skipped = 0

    for sn in wb.sheetnames:
        code = _extract_code_from_sheetname(sn)
        if not code:
            skipped += 1
            continue

        ws = wb[sn]

        # 找表头行（Row 7 是标准明细账簿格式）
        header_row = None
        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if '日期' in row_vals and '借方' in row_vals and '贷方' in row_vals:
                header_row = r
                break

        if header_row is None:
            skipped += 1
            continue

        # 列索引
        col_map = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(header_row, c).value or '').strip()
            if h == '日期': col_map['date'] = c
            elif h == '凭证字号': col_map['voucher'] = c
            elif h == '摘要': col_map['summary'] = c
            elif h == '对方科目': col_map['counterparty'] = c
            elif h == '借方': col_map['debit'] = c
            elif h == '贷方': col_map['credit'] = c
            elif h == '方向': col_map['direction'] = c
            elif h == '余额': col_map['balance'] = c

        if 'date' not in col_map or 'debit' not in col_map:
            skipped += 1
            continue

        # 提取科目名称（从sheet名去掉编码部分）
        name = str(sn).strip()
        if name.startswith(code):
            name = name[len(code):].strip()

        transactions = []
        current_date = None

        for r in range(header_row + 1, ws.max_row + 1):
            date_val = ws.cell(r, col_map['date']).value
            summary = str(ws.cell(r, col_map['summary']).value or '').strip()

            # 跳过汇总行
            if summary in ('期初余额', '本月合计', '本年累计', '') and not date_val:
                continue
            if '本月合计' in summary or '本年累计' in summary:
                continue

            # 日期解析
            dt = None
            if date_val:
                if isinstance(date_val, datetime):
                    dt = date_val
                elif isinstance(date_val, str):
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                        try:
                            dt = datetime.strptime(date_val.strip(), fmt)
                            break
                        except: pass
                if dt:
                    current_date = dt

            debit = _safe_float(ws.cell(r, col_map['debit']).value) if 'debit' in col_map else 0.0
            credit = _safe_float(ws.cell(r, col_map['credit']).value) if 'credit' in col_map else 0.0

            if debit == 0 and credit == 0:
                continue

            counterparty = str(ws.cell(r, col_map['counterparty']).value or '').strip() if 'counterparty' in col_map else ''
            direction = str(ws.cell(r, col_map['direction']).value or '').strip() if 'direction' in col_map else ''
            balance = _safe_float(ws.cell(r, col_map['balance']).value) if 'balance' in col_map else 0.0

            transactions.append({
                'date': dt.strftime('%Y-%m-%d') if dt else '',
                'voucher': str(ws.cell(r, col_map['voucher']).value or '').strip() if 'voucher' in col_map else '',
                'summary': summary,
                'counterparty': counterparty,
                'debit': debit,
                'credit': credit,
                'direction': direction,
                'balance': balance,
            })

        if transactions:
            # 按对方科目归类结算对象
            settlements = defaultdict(lambda: {'debit': 0.0, 'credit': 0.0, 'last_date': '', 'summaries': []})
            for t in transactions:
                # 从对方科目提取结算对象名称（取逗号前的第一个名称，去掉编码前缀）
                cp = t['counterparty']
                if cp:
                    # 对方科目格式: "22210106 销项税额,500100 主营业务收入" → 取"销项税额"
                    parts = cp.split(',')
                    for p in parts:
                        clean = re.sub(r'^[\d\s]+', '', p).strip()
                        if clean:
                            settlements[clean]['debit'] += t['debit']
                            settlements[clean]['credit'] += t['credit']
                            if t['date'] and t['date'] > settlements[clean]['last_date']:
                                settlements[clean]['last_date'] = t['date']
                            if t['summary'] and t['summary'] not in settlements[clean]['summaries']:
                                settlements[clean]['summaries'].append(t['summary'])

            result[code] = {
                'code': code,
                'name': name,
                'sheet_name': sn,
                'transaction_count': len(transactions),
                'settlement_count': len(settlements),
                'transactions': transactions,
                'settlements': dict(settlements),
            }
            parsed += 1

    wb.close()

    # 保存缓存
    cache_path = os.path.join(cache_dir, 'subledger_standardized.json')
    os.makedirs(cache_dir, exist_ok=True)
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    total_transactions = sum(v['transaction_count'] for v in result.values())
    total_settlements = sum(v['settlement_count'] for v in result.values())
    print(f'[SUBLEDGER] ✅ 标准化完成: {parsed}/{total_sheets} sheets解析')
    print(f'[SUBLEDGER]    科目数: {len(result)}, 交易笔数: {total_transactions}, 结算对象: {total_settlements}')

    return result


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 3:
        print('用法: python subledger_standardizer.py <科目明细账路径> <_dt_cache目录>')
        sys.exit(1)
    standardize_subledger(sys.argv[1], sys.argv[2])
