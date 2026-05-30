#!/usr/bin/env python3
"""修复明细表第3轮问题：
1) 增值额/增值率/账龄公式列格式应与模板行一致（固化标准格式）
2) 部分科目备注列右侧存在打印区域（应截止到备注列）
3) 坏账准备等减值类科目贷方金额→明细表填正数（合计2公式=合计1-减值）
4) 条件格式(ISFORMULA)范围未扩展→增值额/增值率/账龄公式列无浅灰底色

v1.1 - 2026-05-25: 增加条件格式范围修复
"""
import sys
import re
import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, numbers, PatternFill
from openpyxl.workbook.properties import CalcProperties

# 使用公共模块的find_header_cols（替代本地实现，消除硬编码列号）
from sheet_col_finder import find_header_cols as _find_header_cols_public
from sheet_col_finder import find_data_start_row as find_data_start_row_local


# ========== 固化格式标准 ==========
# 基于模板行(Row6)的实际格式提取
FONT_TNR = 'Times New Roman'
FONT_CN = '宋体'
FONT_SIZE = 11

# 增值额/增值率 列格式标准
FMT_VALUE = '#,##0.00_);[Red]\\-#,##0.00_);_(* ""_)'  # 会计格式
FMT_RATE = '#,##0.00_);[Red]\\-#,##0.00_);_(* ""_)'    # 会计格式（同增值额）
FMT_AGE = 'General'  # 账龄：常规格式
FMT_AMOUNT = '#,##0.00'  # 金额格式
FMT_SEQ = '0'  # 序号：整数
FMT_DATE = 'yyyy"年"mm"月"'  # 发生日期

font_tnr = Font(name=FONT_TNR, size=FONT_SIZE)
font_cn = Font(name=FONT_CN, size=FONT_SIZE)
align_right = Alignment(horizontal='right', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

THIN_SIDE = Side(style='thin')
STANDARD_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# 减值类行标识（A列文本包含这些关键词的行）
BAD_DEBT_MARKERS = {'坏账准备', '预计风险', '计提跌价准备', '减值准备', '跌价准备', '预计损失'}


def find_struct(ws):
    """找到三行结构：合计1、减值行、合计2"""
    total1 = total2 = None
    bad_debts = []  # 可能有多个减值行
    for r in range(1, min(ws.max_row + 1, 60)):
        a = ws.cell(row=r, column=1).value
        if a and isinstance(a, str):
            t = a.replace(' ', '').strip()
            if t == '合计1':
                total1 = r
            elif t == '合计2':
                total2 = r
            elif t in BAD_DEBT_MARKERS:
                bad_debts.append((r, t))
    return total1, bad_debts, total2


def find_header_cols(ws):
    """动态查找表头列号映射（委托给sheet_col_finder公共模块）。
    
    不同sheet中列号差异巨大（如3-1-1增值额=I，4-8-1增值额=O），
    绝不能硬编码列号。
    """
    return _find_header_cols_public(ws)


def find_remark_col(ws):
    """找到备注列的列号"""
    for c in range(1, min(ws.max_column + 1, 25)):
        h = ws.cell(row=5, column=c).value
        if h and '备注' in str(h):
            return c
    return None


def find_last_data_col(ws):
    """找到最后一个有数据的列（用于确定打印范围右边界）"""
    # 优先取备注列
    remark_col = find_remark_col(ws)
    if remark_col:
        return remark_col
    
    # 无备注列时，取最后一个有表头的非隐藏列
    last_col = 0
    for c in range(1, min(ws.max_column + 1, 25)):
        h = ws.cell(row=5, column=c).value
        cl = get_column_letter(c)
        hidden = ws.column_dimensions[cl].hidden
        if h and not hidden:
            last_col = c
    return last_col if last_col > 0 else 14  # 默认N列


def fix_workbook(filepath):
    """主修复函数"""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    stats = {
        'formula_format': 0,  # 公式列格式修正
        'print_area': 0,      # 打印范围修正
        'bad_debt_sign': 0,    # 减值行金额符号修正
        'cond_fmt_range': 0,   # 条件格式范围修正
    }
    details = {
        'formula_format': [],
        'print_area': [],
        'bad_debt_sign': [],
        'cond_fmt_range': [],
    }
    
    for sname in wb.sheetnames:
        # 跳过汇总表和辅助sheet
        if '汇总' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        if sname in ('目录', '公式数据表', '设置', '设定信息'):
            continue
        
        ws = wb[sname]
        total1, bad_debts, total2 = find_struct(ws)
        if not total1:
            continue
        
        # ========== 1. 公式列格式标准化 ==========
        # 基于模板行(Row6)的固化格式标准
        # 找到数据区
        data_start = find_data_start_row_local(ws)  # 动态查找数据起始行
        data_end = total1 - 1  # 合计1上方
        
        # 逐列检查表头语义，确定格式标准
        col_formats = {}  # col -> (font, alignment, number_format)
        for c in range(1, min(ws.max_column + 1, 25)):
            h = ws.cell(row=5, column=c).value
            h_text = str(h) if h else ''
            
            if '增值额' in h_text:
                col_formats[c] = (font_tnr, align_right, FMT_VALUE)
            elif '增值率' in h_text:
                col_formats[c] = (font_tnr, align_right, FMT_RATE)
            elif '账龄' in h_text:
                col_formats[c] = (font_tnr, align_center, FMT_AGE)
            elif '账面价值' in h_text or '评估价值' in h_text:
                col_formats[c] = (font_tnr, align_right, FMT_AMOUNT)
            elif '序号' in h_text:
                col_formats[c] = (font_tnr, align_center, FMT_SEQ)
            elif '发生日期' in h_text:
                col_formats[c] = (font_cn, align_center, FMT_DATE)
        
        # 应用格式到所有数据行和结构行
        all_rows = list(range(data_start, data_end + 1))
        if total2:
            all_rows.extend(range(total1, total2 + 1))
        else:
            all_rows.append(total1)
        
        for r in all_rows:
            for c, (fmt_font, fmt_align, fmt_num) in col_formats.items():
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                
                changed = False
                # 字体
                if cell.font.name != fmt_font.name or cell.font.size != fmt_font.size:
                    cell.font = fmt_font
                    changed = True
                # 对齐
                if cell.alignment.horizontal != fmt_align.horizontal:
                    cell.alignment = fmt_align
                    changed = True
                # 数字格式（仅对有公式或数值的单元格应用）
                has_formula = isinstance(cell.value, str) and cell.value.startswith('=')
                has_number = isinstance(cell.value, (int, float))
                if (has_formula or has_number) and cell.number_format != fmt_num:
                    cell.number_format = fmt_num
                    changed = True
                
                if changed:
                    stats['formula_format'] += 1
                    if len(details['formula_format']) < 20:
                        details['formula_format'].append(
                            f'[{sname}] {get_column_letter(c)}{r}: 格式标准化'
                        )
        
        # ========== 2. 打印范围修正：右至备注列 ==========
        last_data_col = find_last_data_col(ws)
        end_row = total2 if total2 else total1
        
        if last_data_col and end_row:
            new_pa = f"B1:{get_column_letter(last_data_col)}{end_row}"
            old_pa = str(ws.print_area)
            
            if old_pa != new_pa:
                # 检查旧打印范围是否超出备注列
                remark_col = find_remark_col(ws)
                m = re.search(r':([A-Z]+)(\d+)', old_pa)
                if m:
                    old_last_col = column_index_from_string(m.group(1))
                    if remark_col and old_last_col > remark_col:
                        ws.print_area = new_pa
                        stats['print_area'] += 1
                        details['print_area'].append(
                            f'[{sname}] {old_pa} → {new_pa} (备注列={get_column_letter(remark_col)})'
                        )
                    elif not remark_col and old_last_col > last_data_col:
                        # 无备注列但打印范围超出最后数据列
                        ws.print_area = new_pa
                        stats['print_area'] += 1
                        details['print_area'].append(
                            f'[{sname}] {old_pa} → {new_pa} (最后数据列={get_column_letter(last_data_col)})'
                        )
        
        # ========== 3. 减值行金额符号修正：负数→正数 ==========
        # 规则：坏账准备等减值类科目在科目余额表中为贷方余额
        # 明细表公式：合计2 = 合计1 - 减值 - 预计风险
        # 所以明细表中减值行应填正数（绝对值），公式会自动作差
        
        # 动态查找账面价值和评估价值列号（不同sheet列号不同！）
        header_cols = find_header_cols(ws)
        amount_cols_for_sign = []
        if '账面价值' in header_cols:
            amount_cols_for_sign.append(header_cols['账面价值'])
        if '评估价值' in header_cols:
            amount_cols_for_sign.append(header_cols['评估价值'])
        
        for bad_row, bad_label in bad_debts:
            for c in amount_cols_for_sign:
                cell = ws.cell(row=bad_row, column=c)
                val = cell.value
                
                if isinstance(val, (int, float)) and val < 0:
                    # 取绝对值
                    cell.value = abs(val)
                    stats['bad_debt_sign'] += 1
                    cl = get_column_letter(c)
                    details['bad_debt_sign'].append(
                        f'[{sname}] {cl}{bad_row}({bad_label}): {val} → {abs(val)}'
                    )
        
        # ========== 4. 条件格式(ISFORMULA)范围扩展 ==========
        # 模板中ISFORMULA条件格式标记公式单元格为浅灰底色
        # insert_rows后条件格式范围不自动扩展，新行公式单元格无底色
        end_row = total2 if total2 else total1
        remark_col = find_remark_col(ws)
        
        if remark_col and end_row:
            remark_letter = get_column_letter(remark_col)
            data_start_row = 7  # 默认数据起始行
            # 检查A列标记确定数据起始行
            for r in range(1, min(ws.max_row + 1, 10)):
                a_val = ws.cell(row=r, column=1).value
                if a_val and isinstance(a_val, str):
                    if a_val.strip() in ('检索表头2', '检索表头'):
                        data_start_row = r + 1
                        break
            
            try:
                # 收集现有条件格式
                cf_list = []
                for cf in ws.conditional_formatting:
                    sqref_str = str(cf.sqref) if cf.sqref else ''
                    rules_info = []
                    for rule in cf.rules:
                        rules_info.append({
                            'type': rule.type,
                            'dxfId': rule.dxfId,
                            'formula': rule.formula,
                            'priority': rule.priority,
                        })
                    cf_list.append({
                        'sqref': sqref_str,
                        'rules': rules_info,
                    })
                
                if not cf_list:
                    continue
                
                # 重建条件格式，扩展范围
                new_cf = openpyxl.formatting.formatting.ConditionalFormattingList()
                for cf_info in cf_list:
                    old_sqref = cf_info['sqref']
                    
                    # 替换范围中的行号
                    def replace_row_in_ref(match):
                        col1 = match.group(1)
                        row1 = int(match.group(2))
                        col2 = match.group(3)
                        row2 = int(match.group(4))
                        # 如果终止行号 < end_row且起始行 >= data_start_row，扩展
                        if row2 < end_row and row1 >= data_start_row:
                            return f'{col1}{row1}:{col2}{end_row}'
                        return match.group(0)
                    
                    new_sqref = re.sub(
                        r'([A-Z]+)(\d+):([A-Z]+)(\d+)',
                        replace_row_in_ref,
                        old_sqref
                    )
                    
                    if new_sqref != old_sqref:
                        stats['cond_fmt_range'] += 1
                        details['cond_fmt_range'].append(
                            f'[{sname}] {old_sqref} → {new_sqref}'
                        )
                    
                    # 使用add方法重建
                    from openpyxl.formatting.rule import Rule
                    for ri in cf_info['rules']:
                        new_rule = Rule(
                            type=ri['type'],
                            dxfId=ri['dxfId'],
                            formula=ri['formula'],
                            priority=ri['priority'],
                        )
                        new_cf.add(new_sqref, new_rule)
                
                # 替换整个条件格式对象
                ws.conditional_formatting = new_cf
            except Exception as e:
                details['cond_fmt_range'].append(
                    f'[{sname}] 条件格式更新失败: {e}'
                )
    
    # 保存
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=0)
    wb.save(filepath)
    wb.close()
    
    # 输出统计
    print(f'===== 修复完成 =====')
    print(f'1. 公式列格式标准化: {stats["formula_format"]}项')
    for d in details['formula_format'][:10]:
        print(f'   {d}')
    if len(details['formula_format']) > 10:
        print(f'   ... 共{len(details["formula_format"])}项')
    
    print(f'2. 打印范围修正: {stats["print_area"]}项')
    for d in details['print_area']:
        print(f'   {d}')
    
    print(f'3. 减值行金额符号修正: {stats["bad_debt_sign"]}项')
    for d in details['bad_debt_sign']:
        print(f'   {d}')
    
    print(f'4. 条件格式范围修正: {stats["cond_fmt_range"]}项')
    for d in details['cond_fmt_range'][:10]:
        print(f'   {d}')
    if len(details['cond_fmt_range']) > 10:
        print(f'   ... 共{len(details["cond_fmt_range"])}项')
    
    return stats


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python fix_round3.py <xlsx_file>')
        sys.exit(1)
    
    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f'文件不存在: {filepath}')
        sys.exit(1)
    
    fix_workbook(filepath)
