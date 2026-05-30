"""
fix_summary_refs.py — 修复汇总表跨Sheet引用行号偏移 (v2.0)

问题根因（DT-182b）：
  smart_insert_row() 在明细表的数据区（表头~合计之间）插入行后，
  合计行位置下移。但汇总表中的跨Sheet公式引用（如 '4-8-4机器设备'!K34）
  是硬编码的，openpyxl 不会自动更新它们。
  
  结果：汇总表引用的行号不再指向正确的合计行，导致汇总数据错误。

v2.0 核心改进（2026-05-29）：
  1. fix_intra_sheet_total2_formulas: 不再使用硬编码行号映射(26/27/28, 32/33/34)，
     改为动态扫描公式中的行号引用，自动匹配到正确的合计行位置
  2. fix_all_summary_refs_batch: 增加自引用检测，避免修改自身公式导致循环引用
  3. fix_summary_sheet_refs: 增加边界校验，跳过自身引用

使用：
  from fix_summary_refs import fix_summary_sheet_refs
  
  # 在 fill_sheet() 之后调用
  fix_summary_sheet_refs(wb, filled_sheet_name, original_total_row)

v1.0 (2026-05-29): 初始版本
v2.0 (2026-05-29): 修复循环引用bug，消除硬编码行号
"""

import re
from openpyxl.utils import column_index_from_string, get_column_letter


def find_total_rows(ws):
    """找到Sheet中的合计行位置
    
    通过A列标记定位：合计1 / 坏账准备 / 合计2
    
    Returns:
        dict: {'total1': int, 'bad_debt': int|None, 'total2': int|None, 'provision': int|None}
    """
    result = {'total1': None, 'bad_debt': None, 'total2': None, 'provision': None}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        av = str(a).strip()
        if '合计1' in av:
            result['total1'] = r
        elif '坏账准备' in av or '减值准备' in av:
            result['bad_debt'] = r
        elif '合计2' in av:
            result['total2'] = r
        elif '预计损失' in av or '预计风险' in av:
            result['provision'] = r
    return result


def is_summary_sheet(sheet_name):
    """判断是否为汇总类Sheet（需要检查引用正确性）"""
    keywords = ['汇总', '分类汇总']
    return any(k in sheet_name for k in keywords)


def get_sheet_dependencies(wb, summary_sheet_name):
    """扫描汇总Sheet中的所有跨Sheet引用，返回 {detail_sheet: [(col, old_row, formula_str)]}"""
    ws = wb[summary_sheet_name]
    deps = {}
    pattern = r"'([^']+)'!([A-Z]+)\$?(\d+)|(?<!\w)([A-Za-z0-9_\u4e00-\u9fff-]+)!([A-Z]+)\$?(\d+)"
    
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str) or '!' not in v:
                continue
            
            matches = re.finditer(pattern, v)
            for m in matches:
                if m.group(1):
                    sn = m.group(1)
                    col_l = m.group(2)
                    row_n = int(m.group(3))
                else:
                    sn = m.group(4)
                    col_l = m.group(5)
                    row_n = int(m.group(6))
                
                if sn not in deps:
                    deps[sn] = []
                deps[sn].append({
                    'summary_sheet': summary_sheet_name,
                    'summary_row': r,
                    'summary_col': c,
                    'col_letter': col_l,
                    'old_row': row_n,
                    'full_formula': v,
                })
    
    return deps


def fix_summary_sheet_refs(wb, filled_sheet_name, original_total2_row=None):
    """修复所有汇总表中指向 filled_sheet_name 的跨Sheet引用 (v2.0)
    
    v2.0改进：跳过自身引用，增加边界校验，防止循环引用。
    """
    if filled_sheet_name not in wb.sheetnames:
        return {'fixed': 0, 'refs_checked': 0, 'details': [], 'error': f'Sheet {filled_sheet_name} not found'}
    
    ws_filled = wb[filled_sheet_name]
    totals = find_total_rows(ws_filled)
    actual_total2 = totals.get('total2')
    actual_total1 = totals.get('total1')
    
    if not actual_total2 and not actual_total1:
        return {'fixed': 0, 'refs_checked': 0, 'details': [], 
                'error': f'在 {filled_sheet_name} 中未找到合计行（A列标记）'}
    
    target_row = actual_total2 if actual_total2 else actual_total1
    if actual_total2:
        target_row_name = '合计2'
    else:
        target_row_name = '合计1'
    
    total_fixed = 0
    details = []
    
    for sn in wb.sheetnames:
        if not is_summary_sheet(sn):
            continue
        # v2.0: 跳过指向自身的汇总表，防止自引用
        if sn == filled_sheet_name:
            continue
        
        ws = wb[sn]
        sheet_fixed = 0
        
        pattern = r"'([^']+)'!\$?([A-Z]+)\$?(\d+)"
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=r, column=c).value
                if not v or not isinstance(v, str) or '!' not in v:
                    continue
                
                old_formula = v
                matches = re.finditer(pattern, v)
                
                for m in matches:
                    ref_sn = m.group(1)
                    if ref_sn != filled_sheet_name:
                        continue
                    col_l = m.group(2)
                    ref_row = int(m.group(3))
                    
                    if ref_row == target_row:
                        continue  # 已指向正确行
                    
                    row_diff = abs(ref_row - target_row)
                    if row_diff < 3:
                        continue  # 偏移很小，跳过
                    if row_diff > 50:
                        continue  # 偏移过大，不是同一张表
                    
                    old_ref = "'" + ref_sn + "'!" + col_l + str(ref_row)
                    new_ref = "'" + ref_sn + "'!" + col_l + str(target_row)
                    
                    # v2.0: 检查新引用是否已存在（防止重复修复）
                    if new_ref in old_formula:
                        continue
                    
                    v = v.replace(old_ref, new_ref)
                
                if v != old_formula:
                    # v2.0: 自引用检测 - 新公式不能引用汇总表自身的当前单元格
                    updated_col_l = get_column_letter(c)
                    self_ref = f"'{sn}'!{updated_col_l}{r}"
                    if self_ref not in v:  # 确保没有自引用
                        ws.cell(row=r, column=c).value = v
                        sheet_fixed += 1
                        details.append({
                            'sheet': sn,
                            'row': r,
                            'col': c,
                            'old': old_formula,
                            'new': v,
                        })
        
        if sheet_fixed > 0:
            total_fixed += sheet_fixed
    
    return {'fixed': total_fixed, 'refs_checked': total_fixed, 'details': details}


MONEY_COLS = set('EFGHIJKLMNOP')


def _is_detail(sn):
    """判断是否为明细表（排除汇总表、结构表）"""
    if any(k in sn for k in ['汇总', '设定信息', '公式数据表', '_BS对照', '设置', '0-其他方法', '目录']):
        return False
    return True


def fix_all_summary_refs_batch(wb):
    """批量修复所有汇总表引用 (v2.0 - 增加自引用检测和边界校验)"""
    total_fixed = 0
    per_sheet = {}

    for sn in wb.sheetnames:
        if not any(k in sn for k in ['汇总', '分类汇总']):
            continue

        ws = wb[sn]
        sheet_fixed = 0

        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=r, column=c).value
                if not v or not isinstance(v, str) or '!' not in v:
                    continue

                formula = v
                pattern = r"'([^']+)'!([A-Z]+)(\d+)"
                matches = list(re.finditer(pattern, formula))

                for m in matches:
                    detail_sn = m.group(1)
                    col_l = m.group(2)
                    ref_row = int(m.group(3))

                    # v2.0: 跳过自引用
                    if detail_sn == sn:
                        continue
                    if detail_sn not in wb.sheetnames or not _is_detail(detail_sn):
                        continue
                    if col_l not in MONEY_COLS:
                        continue

                    totals = find_total_rows(wb[detail_sn])
                    target_row = totals.get('total2') or totals.get('total1')
                    if not target_row:
                        continue
                    
                    # v2.0: 边界校验
                    if ref_row == target_row:
                        continue
                    row_diff = abs(ref_row - target_row)
                    if row_diff < 3:
                        continue
                    if row_diff > 50:
                        continue
                    
                    old_ref = "'" + detail_sn + "'!" + col_l + str(ref_row)
                    new_ref = "'" + detail_sn + "'!" + col_l + str(target_row)
                    
                    if new_ref in formula:
                        continue
                    
                    if old_ref in formula:
                        formula = formula.replace(old_ref, new_ref)
                        sheet_fixed += 1

                if formula != v:
                    # v2.0: 自引用检测
                    current_col = get_column_letter(c)
                    self_ref = f"'{sn}'!{current_col}{r}"
                    if self_ref in formula:
                        continue
                    ws.cell(row=r, column=c).value = formula

        if sheet_fixed > 0:
            per_sheet[sn] = sheet_fixed
            total_fixed += sheet_fixed

    return {'total_fixed': total_fixed, 'per_sheet': per_sheet}


def _find_row_refs_in_formula(formula):
    """在公式中找到所有的行号引用（排除函数名）
    
    Returns:
        list of int: 公式中引用的行号
    """
    refs = set()
    # 匹配: 大写字母1-3位+数字（可能是行号）
    for m in re.finditer(r'([A-Z]{1,3})(\d+)', formula):
        col = m.group(1)
        row = int(m.group(2))
        # 排除Excel函数名
        if col in ('SUM', 'IF', 'ROW', 'COL', 'MAX', 'MIN', 'AVG',
                   'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'OFFSET',
                   'INDIRECT', 'ADDRESS', 'ROWS', 'COLUMNS', 'ABS',
                   'ROUND', 'INT', 'MOD', 'SUMIF', 'COUNT', 'COUNTA',
                   'SUBTOTAL', 'TEXT', 'VALUE', 'DATE', 'TIME',
                   'YEAR', 'MONTH', 'DAY', 'NOW', 'TODAY'):
            continue
        # 排除明显不是行号的（以0开头且不止一位数字）
        row_str = str(row)
        if row_str.startswith('0') and len(row_str) > 1:
            continue
        # 排除超大行号（Excel最大行1048576）
        if row > 1048576 or row < 1:
            continue
        refs.add(row)
    return sorted(refs)


def fix_intra_sheet_total2_formulas(wb):
    """修复明细表内合计2行引用旧行号的公式 (v2.0 - 动态检测)
    
    v2.0改进（2026-05-29）：
      不再使用硬编码行号映射(26/27/28, 32/33/34)，
      改为动态扫描合计行公式中的行号引用，匹配到最近的标记行位置。
    
    思路：
      每行有A列标记（合计1、坏账准备、合计2）。
      smart_insert_row插行后，非SUM公式中的硬编码行号可能指向旧位置。
      扫描这些公式中的数字引用，如果指向的不是当前标记行，
      且距离某个标记行较近(≈插行偏移量)，则修正为该标记行的新位置。
    """
    results = {}
    
    for sn in wb.sheetnames:
        if any(k in sn for k in ['汇总', '辅-', '分类汇总',
                                  '设定信息', '公式数据表', '_BS对照',
                                  '设置', '0-其他方法', '目录']):
            continue
        
        ws = wb[sn]
        totals = find_total_rows(ws)
        t1 = totals.get('total1')
        t2 = totals.get('total2')
        bd = totals.get('bad_debt') or totals.get('provision')
        
        if not t2:
            continue
        
        fixed = 0
        
        # 修复行：合计2、坏账准备（如有）、合计1（如有）
        rows_to_fix = [t2]
        if bd:
            rows_to_fix.append(bd)
        if t1:
            rows_to_fix.append(t1)
        
        # 当前正确的标记行位置
        marker_rows = {r: r for r in [t1, bd, t2] if r is not None}
        
        for fix_row in rows_to_fix:
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=fix_row, column=c).value
                if not v or not isinstance(v, str):
                    continue
                
                # v2.0: SUM公式会自动适应行数变化，跳过
                if v.strip().upper().startswith('=SUM('):
                    continue
                
                old = v
                
                # 找到公式中所有的行号引用（同Sheet引用）
                ref_rows = _find_row_refs_in_formula(v)
                
                for ref_row in ref_rows:
                    if ref_row == fix_row:
                        continue  # 自引用
                    if ref_row in marker_rows:
                        continue  # 已指向正确位置
                    
                    # 找最近的标记行
                    best_match = None
                    best_dist = float('inf')
                    for m_row in [t1, bd, t2]:
                        if m_row is None:
                            continue
                        dist = abs(ref_row - m_row)
                        if dist < best_dist:
                            best_dist = dist
                            best_match = m_row
                    
                    # 如果距离在合理范围内（排除完全无关的行号）
                    if best_match is not None and 0 < best_dist < 10 and ref_row != best_match:
                        # 替换该行号的所有出现（配合列字母）
                        for col_match in set(re.finditer(r'([A-Z]{1,3})' + str(ref_row) + r'(?![0-9\.])', v)):
                            col_letter = col_match.group(1)
                            if col_letter in ('SUM', 'IF', 'ROW', 'COL', 'MAX', 'MIN', 'AVG',
                                             'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'OFFSET',
                                             'INDIRECT', 'ADDRESS', 'ROWS', 'COLUMNS', 'ABS'):
                                continue
                            old_ref_str = col_letter + str(ref_row)
                            new_ref_str = col_letter + str(best_match)
                            v = v.replace(old_ref_str, new_ref_str)
                
                if v != old:
                    # v2.0: 最终自引用检查
                    col_letter = get_column_letter(c)
                    self_ref = col_letter + str(fix_row)
                    if v.count(self_ref) > old.count(self_ref):
                        # 替换后自引用变多了，跳过
                        continue
                    ws.cell(row=fix_row, column=c).value = v
                    fixed += 1
        
        if fixed > 0:
            results[sn] = fixed
    
    return results
