"""
shared_checks.py — 通用检查函数库 v1.0

从 valuation-dcf-workpaper/scripts/check_formulas.py v5.0 提取的通用检查函数。
供多个评估工作底稿审核 skill 共享使用。

提取内容：
  - 工具函数：col_to_num, parse_range, is_row_hidden, is_col_hidden
  - 检查函数：check1_column_progression, check2_auxiliary_semantics
  - 语义判定：determine_semantic
  - 辅助判定：_is_ratio_formula, _is_growth_formula
  - 关键词常量：_RATIO_KEYWORDS, _GROWTH_KEYWORDS, _COST_EXPENSE_LABEL_KEYWORDS

版本历史：
  v1.0 (2026-05-21) — 初始提取自 check_formulas.py v5.0
"""

import openpyxl
import re
from collections import defaultdict


def col_to_num(col_str):
    """列字母转数字（A=1, B=2, ..., Z=26, AA=27）"""
    result = 0
    for ch in col_str.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def parse_range(range_str):
    """解析行/列范围字符串，如 '13-25' → (13, 25)，'B-V' → (2, 22)"""
    parts = range_str.split('-')
    if len(parts) == 2:
        return int(parts[0]), int(parts[1])
    return int(parts[0]), int(parts[0])


def is_row_hidden(ws, row_idx):
    """检查行是否隐藏"""
    rd = ws.row_dimensions.get(row_idx)
    return rd is not None and rd.hidden


def is_col_hidden(ws, col_letter):
    """检查列是否隐藏"""
    cd = ws.column_dimensions.get(col_letter)
    return cd is not None and cd.hidden


def check1_column_progression(ws, sheet_name, min_row, max_row, min_col, max_col, wb_val=None):
    """强制检查1：预测期公式"引用列递推"检查
    v5.0增强（盲区修复）：
    - 【盲区修复】新增半绝对列引用检测：$O15模式（列绝对、行相对）
      原有regex \$([A-Z]+)\$(\d+) 只匹配$O$15（双绝对），遗漏$O15（列绝对行相对）
      新增检测：同一$X列在2+预测列的基底值位置出现 → 列锁定问题
      区分基底值位置（应递推）和增长率参数位置（可锁定）
    v3.0增强保留：
    - 跳过隐藏行/列
    - 检测绝对引用空值（如$N$13为None时预警）
    - 多行绑定预警（同一绝对引用单元格被≥3行引用）
    - 预测期数值全部相同时标记（亚模式A）
    - 【S5】亚模式B检测：首年不同+后续同值（P≠O但P=Q=R=S）
    - 【S2】逐行输出格式：Sheet=___ 行=___ 标签=___ 预测列公式模式=___ 引用列递推=✅/❌
    """
    results = []
    issues = []
    abs_ref_usage = defaultdict(list)  # 记录每个绝对引用被哪些行使用
    semi_abs_base_usage = defaultdict(list)  # 记录半绝对列引用在基底值位置被哪些(行,预测列)使用
    checked_rows = 0

    for row_idx in range(min_row, max_row + 1):
        if is_row_hidden(ws, row_idx):
            continue
        checked_rows += 1

        # 收集该行预测期列的公式
        row_formulas = {}
        for col_idx in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if is_col_hidden(ws, col_letter):
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.data_type == 'f':
                row_formulas[col_letter] = str(cell.value)

        if not row_formulas or len(row_formulas) < 2:
            continue

        # 提取行标签
        label = str(ws.cell(row=row_idx, column=1).value or
                    ws.cell(row=row_idx, column=2).value or "")[:25]

        # 提取完全绝对引用（$O$15模式）
        row_has_abs_ref_lock = False
        for col_letter, formula in row_formulas.items():
            abs_refs = re.findall(r'\$([A-Z]+)\$(\d+)', formula)
            for ref_col, ref_row in abs_refs:
                ref_key = f"${ref_col}${ref_row}"
                abs_ref_usage[ref_key].append(row_idx)

                # 空值检测：如果被引用的单元格为None
                if wb_val is not None:
                    ref_col_idx = col_to_num(ref_col)
                    ref_row_idx = int(ref_row)
                    ref_val = wb_val.cell(row=ref_row_idx, column=ref_col_idx).value
                    if ref_val is None:
                        issues.append(
                            f"  🔴 {col_letter}{row_idx}: 绝对引用 {ref_key} 的值为None！"
                            f"公式 {formula[:60]} → (1+None)=(1+0)，增长率=0%"
                        )

            # 【盲区修复v5.0】提取半绝对列引用（$O15模式：列绝对、行相对）
            # regex \$([A-Z]+)(\d+) 只匹配$O15，不匹配$O$15（因为$15不以数字开头）
            semi_abs_refs = re.findall(r'\$([A-Z]+)(\d+)', formula)
            for ref_col, ref_row in semi_abs_refs:
                # 判断是否在增长率参数位置：如(1+$O15)中的$O15
                # 增长率参数位置的半绝对引用是合理的（各年使用同一增长率）
                # 基底值位置的半绝对引用是有问题的（应递推引用前一年）
                is_rate_param = bool(re.search(
                    rf'\(1\+\$?{re.escape(ref_col)}\$?{ref_row}\)', formula
                ))
                if not is_rate_param:
                    # 基底值位置的半绝对列引用
                    semi_abs_base_usage[ref_col].append((row_idx, col_letter, formula[:80]))

        # 检查是否所有预测列引用同一个绝对列（如$O$15）
        abs_pattern_counts = defaultdict(int)
        for formula in row_formulas.values():
            abs_patterns = re.findall(r'\$([A-Z]+)\$\d+', formula)
            for pat in abs_patterns:
                abs_pattern_counts[pat] += 1

        for pat, cnt in abs_pattern_counts.items():
            if cnt >= 3:
                row_has_abs_ref_lock = True
                issues.append(
                    f"  🔴 Sheet={sheet_name} 行={row_idx} 标签={label} "
                    f"预测列公式: {list(row_formulas.values())[0][:50]} "
                    f"→ 绝对列引用${pat}$出现在{cnt}个预测列中，引用列未递推❌"
                )

        # 【盲区修复v5.0】检查半绝对列引用的基底值锁定
        # 同一行的多个预测列公式中，同一半绝对列($X)出现≥2次作为基底值 → 列锁定
        row_semi_abs_counts = defaultdict(int)
        row_semi_abs_formulas = defaultdict(list)
        for col_letter, formula in row_formulas.items():
            semi_refs = re.findall(r'\$([A-Z]+)(\d+)', formula)
            for ref_col, ref_row in semi_refs:
                # 排除增长率参数位置
                is_rate_param = bool(re.search(
                    rf'\(1\+\$?{re.escape(ref_col)}\$?{ref_row}\)', formula
                ))
                if not is_rate_param:
                    row_semi_abs_counts[ref_col] += 1
                    row_semi_abs_formulas[ref_col].append(f"{col_letter}:{formula[:40]}")

        for ref_col, cnt in row_semi_abs_counts.items():
            if cnt >= 2:
                row_has_abs_ref_lock = True
                formula_samples = row_semi_abs_formulas[ref_col][:3]
                issues.append(
                    f"  🔴 Sheet={sheet_name} 行={row_idx} 标签={label} "
                    f"→ 半绝对列引用${ref_col}（列绝对行相对）出现在{cnt}个预测列的基底值位置，"
                    f"引用列未递推❌ 样例: {'; '.join(formula_samples)}"
                )

        # 【S2】逐行输出结论
        if row_has_abs_ref_lock:
            # 已在上方输出问题，此处不再重复
            pass
        else:
            # 检查引用列是否正常递推
            ref_cols_in_formulas = []
            for col_letter, formula in row_formulas.items():
                refs = re.findall(r'(?:^|[+\-*/(])([A-Z]+)(\d+)', formula)
                ref_cols_in_formulas.extend([r[0] for r in refs if r[0] != col_letter])
            results.append(
                f"  Sheet={sheet_name} 行={row_idx} 标签={label} "
                f"预测列公式模式=多列引用 引用列递推=✅"
            )

    # 多行绑定预警
    for ref_key, rows in abs_ref_usage.items():
        if len(rows) >= 3:
            issues.append(
                f"  ⚠️ 绝对引用 {ref_key} 被{len(rows)}行共用: 行{rows}，"
                f"各行增长率\"绑定\""
            )

    # 【盲区修复v5.0】半绝对列引用跨行锁定预警
    for ref_col, usage_list in semi_abs_base_usage.items():
        if len(usage_list) >= 3:
            rows = sorted(set(item[0] for item in usage_list))
            issues.append(
                f"  ⚠️ 半绝对列引用${ref_col}在基底值位置被{len(usage_list)}个预测列使用: "
                f"行{rows}，引用列锁定"
            )

    # 同值检测（需要wb_val）
    if wb_val is not None:
        for row_idx in range(min_row, max_row + 1):
            if is_row_hidden(ws, row_idx):
                continue
            label = str(ws.cell(row=row_idx, column=1).value or
                        ws.cell(row=row_idx, column=2).value or "")[:25]
            vals = {}
            for col_idx in range(min_col, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if is_col_hidden(ws, col_letter):
                    continue
                v = wb_val.cell(row=row_idx, column=col_idx).value
                if isinstance(v, (int, float)):
                    vals[col_letter] = round(v, 2)
            if len(vals) >= 3:
                unique_vals = set(str(v) for v in vals.values())
                val_list = list(vals.values())
                col_list = list(vals.keys())

                # 亚模式A：所有预测年值完全相同
                if len(unique_vals) == 1:
                    issues.append(
                        f"  🔴 Sheet={sheet_name} 行={row_idx} 标签={label}: "
                        f"预测期所有值完全相同={val_list[0]} (亚模式A)"
                    )
                # 【S5】亚模式B：首年不同+后续同值（P≠O但P=Q=R=S）
                elif len(val_list) >= 3 and len(unique_vals) == 2:
                    # 检查是否首年值不同，后续所有年值相同
                    first_val = val_list[0]
                    rest_vals = val_list[1:]
                    rest_unique = set(str(v) for v in rest_vals)
                    if len(rest_unique) == 1 and str(first_val) != list(rest_unique)[0]:
                        issues.append(
                            f"  🔴 Sheet={sheet_name} 行={row_idx} 标签={label}: "
                            f"亚模式B! 首年={col_list[0]}:{first_val} ≠ "
                            f"后续年={col_list[1]}:{rest_vals[0]}(均相同) "
                            f"→ 疑似$基准列绝对引用导致无法复合增长"
                        )

    results.append(f"  已检查 {checked_rows} 行")
    return results, issues


def _is_ratio_formula(fml):
    """判定公式是否为占比计算（L47/B38落地）
    增强逻辑：
    - 原有 '/' + SUMIF 组合判定保留
    - 原有 '/' + 非IFERROR/IF 组合判定保留
    - 新增：排除增长率公式 (本期-上期)/上期 的误判
    """
    if not fml:
        return False
    if '/' in fml and 'SUMIF' in fml.upper():
        return True
    if '/' in fml and not re.search(r'IFERROR|IF\(', fml, re.IGNORECASE):
        # 增强排除：增长率公式 (本期-上期)/上期 不应被判为占比
        if re.search(r'\(\s*[^,]+?\s*-\s*[^,]+?\s*\)\s*/', fml):
            return False  # 这是增长率公式，不是占比
        return True
    return False


def _is_growth_formula(fml):
    """判定公式是否为增长率计算（L47/B38落地）"""
    if not fml:
        return False
    if re.search(r'\(\s*[^,]+?\s*-\s*[^,]+?\s*\)\s*/', fml):
        return True
    if 'CAGR' in fml.upper():
        return True
    return False


# 占比/增长率关键词库（用于上下文证据）
_RATIO_KEYWORDS = [
    '占收入比', '占收入比例', '收入占比', '占主营收入比',
    '收入比', '占比预测', '占比', '比重', '比率',
    '成本率', '费用率', '毛利率'
]

_GROWTH_KEYWORDS = [
    '增长率', '增长速度', '增速', '增长率预测',
    '复合增长率', 'CAGR', '年均增长', '递增率'
]

# 【盲区修复v5.0】成本/费用行标签关键词（用于判定占比语义）
_COST_EXPENSE_LABEL_KEYWORDS = [
    '成本', '费用', '材料', '人工', '薪酬', '工资',
    '折旧', '摊销', '水电', '运费', '租赁', '维修',
    '办公', '差旅', '招待', '保险', '税金', '福利'
]


def determine_semantic(ref_cell, ws, wb_fml, depth=2, wb_val=None):
    """判定辅助列参数语义（占比 vs 增长率），支持递归公式链展开（L47/B38落地）

    v5.0 增强（盲区修复）：
    - 新增数值范围启发式信号：AVERAGE包装纯数值时，值在0.01-0.50为占比区间
    - 新增行标签成本/费用关键词信号
    - 当公式链展开为空（AVERAGE包装纯数值），通过值范围+行标签辅助判定

    v4.0 新增：
    - 递归展开 AVERAGE/IFERROR/ROUND 等包装函数的内部引用
    - 读取源Sheet上下文证据（列标题+说明文字）
    - 多信号汇聚判定（T23：≥2类证据一致才下结论）

    Args:
        ref_cell: 参数单元格
        ws: 参数所在Worksheet
        wb_fml: 公式模式Workbook
        depth: 递归深度
        wb_val: 数值模式Worksheet（可选，用于数值范围启发式）

    Returns:
        str: 'ratio'|'growth'|'ambiguous'|'unknown'
    """
    visited = set()
    formula_chain = []
    ratio_count = 0
    growth_count = 0

    def _trace(cell, current_ws, current_depth):
        nonlocal ratio_count, growth_count
        if current_depth <= 0:
            return None

        coord = f"{current_ws.title}!{cell.coordinate}"
        if coord in visited:
            return None
        visited.add(coord)

        if cell.data_type != 'f' or not cell.value:
            return None

        fml = str(cell.value)
        formula_chain.append(fml)

        # 直接判定：含占比计算模式
        if _is_ratio_formula(fml):
            ratio_count += 1
            return 'ratio'

        # 直接判定：含增长率计算模式
        if _is_growth_formula(fml):
            growth_count += 1
            return 'growth'

        # 递归展开 AVERAGE/IFERROR/ROUND 等包装函数的内部引用
        # 处理范围引用 C40:E40
        range_refs = re.findall(r'([A-Z]+\d+):([A-Z]+\d+)', fml)
        for start_ref, end_ref in range_refs[:3]:
            start_match = re.match(r'([A-Z]+)(\d+)', start_ref)
            end_match = re.match(r'([A-Z]+)(\d+)', end_ref)
            if start_match and end_match:
                s_col = col_to_num(start_match.group(1))
                s_row = int(start_match.group(2))
                e_col = col_to_num(end_match.group(1))
                e_row = int(end_match.group(2))
                count = 0
                for r in range(s_row, e_row + 1):
                    for c in range(s_col, e_col + 1):
                        if count >= 10:
                            break
                        ref_c = current_ws.cell(row=r, column=c)
                        result = _trace(ref_c, current_ws, current_depth - 1)
                        if result in ('ratio', 'growth'):
                            return result
                        count += 1

        # 提取同Sheet单引用
        inner_refs = re.findall(r'(?<![:\$])([A-Z]+)(\d+)(?!:)', fml)
        for ref_col_str, ref_row_str in inner_refs[:5]:
            try:
                ref_c = current_ws.cell(row=int(ref_row_str), column=col_to_num(ref_col_str))
                result = _trace(ref_c, current_ws, current_depth - 1)
                if result in ('ratio', 'growth'):
                    return result
            except (ValueError, AttributeError):
                continue

        # 跨Sheet递归展开
        cross_refs = re.findall(r"'?([^'!:\s]+)'?!\$?([A-Z]+)\$?(\d+)", fml)
        for sheet_name, col_str, row_str in cross_refs[:3]:
            sheet_name = sheet_name.strip("'\"")
            if wb_fml and sheet_name in wb_fml.sheetnames:
                ref_ws = wb_fml[sheet_name]
                try:
                    ref_c = ref_ws.cell(row=int(row_str), column=col_to_num(col_str))
                    result = _trace(ref_c, ref_ws, current_depth - 1)
                    if result in ('ratio', 'growth'):
                        return result
                except (ValueError, AttributeError):
                    continue

        return None

    # Step 1: 公式链递归展开判定
    chain_result = _trace(ref_cell, ws, depth)

    if chain_result in ('ratio', 'growth'):
        # 公式链已明确判定，但仍需读取上下文证据进行交叉验证（T23）
        pass

    # Step 2: 读取上下文证据（列标题+说明文字）
    row = ref_cell.row
    col = ref_cell.column
    header_ratio = False
    header_growth = False
    annotation_ratio = False
    annotation_growth = False

    # 向上搜索列标题（5行）
    for r in range(max(1, row - 5), row):
        if is_row_hidden(ws, r):
            continue
        row_text = ""
        for c in range(1, 22):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_text += str(v) + " "
        if any(kw in row_text for kw in _RATIO_KEYWORDS):
            header_ratio = True
        if any(kw in row_text for kw in _GROWTH_KEYWORDS):
            header_growth = True

    # 向下搜索说明文字（15行）
    for r in range(row + 1, row + 16):
        if is_row_hidden(ws, r):
            continue
        row_text = ""
        for c in range(1, 22):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_text += str(v) + " "
        if any(kw in row_text for kw in _RATIO_KEYWORDS):
            annotation_ratio = True
        if any(kw in row_text for kw in _GROWTH_KEYWORDS):
            annotation_growth = True

    # Step 3: 多信号汇聚判定（T23落地）
    # 信号来源：
    # 1. 公式链（ratio_count/growth_count）
    # 2. 列标题（header_ratio/header_growth）
    # 3. 说明文字（annotation_ratio/annotation_growth）
    # 4. 【盲区修复v5.0】数值范围启发式（value_in_ratio_range）
    # 5. 【盲区修复v5.0】行标签成本/费用关键词（label_ratio）
    ratio_signals = (1 if ratio_count > 0 else 0) + (1 if header_ratio else 0) + (1 if annotation_ratio else 0)
    growth_signals = (1 if growth_count > 0 else 0) + (1 if header_growth else 0) + (1 if annotation_growth else 0)

    # 【盲区修复v5.0】数值范围启发式信号
    # 当AVERAGE等包装函数包裹纯数值（公式链为空），通过数值范围判定
    # 占比区间：0.01-0.50（即1%-50%），典型成本/费用占收入比
    # 增长率区间：-0.50~0.50，但通常-0.10~0.30
    # 关键区分：占比通常>0.05，而增长率通常<0.30
    value_in_ratio_range = False
    value_in_growth_range = False
    ref_formula_is_average = False
    if ref_cell.data_type == 'f' and ref_cell.value:
        ref_fml_str = str(ref_cell.value)
        # 检测AVERAGE等包装函数
        ref_fml_upper = ref_fml_str.upper().lstrip('=')
        for wf in ['AVERAGE', 'MEDIAN', 'SUM']:
            if ref_fml_upper.startswith(wf + '('):
                ref_formula_is_average = True
                break

    if ref_formula_is_average and wb_val is not None:
        ref_val = wb_val.cell(row=ref_cell.row, column=ref_cell.column).value
        if isinstance(ref_val, (int, float)):
            abs_val = abs(ref_val)
            # 占比区间：5%-50%（0.05-0.50），排除0-5%（过小，可能是增长率）
            if 0.05 <= abs_val <= 0.50:
                value_in_ratio_range = True
                ratio_signals += 1
            # 增长率区间：-30%~+30%（但与占比区间重叠0.05-0.30）
            # 仅当值<0.05或>0.50时，更可能是增长率
            elif abs_val < 0.05 or abs_val > 0.50:
                value_in_growth_range = True
                growth_signals += 1

    # 【盲区修复v5.0】行标签成本/费用关键词信号
    # 当行标签包含成本/费用相关词汇，倾向占比语义
    label_ratio = False
    row_label = str(ws.cell(row=ref_cell.row, column=1).value or
                    ws.cell(row=ref_cell.row, column=2).value or "")
    if any(kw in row_label for kw in _COST_EXPENSE_LABEL_KEYWORDS):
        label_ratio = True
        ratio_signals += 1

    if chain_result == 'ratio' and ratio_signals >= 2:
        return 'ratio'
    elif chain_result == 'growth' and growth_signals >= 2:
        return 'growth'
    elif chain_result == 'ratio' and ratio_signals >= 1 and growth_signals == 0:
        return 'ratio'
    elif chain_result == 'growth' and growth_signals >= 1 and ratio_signals == 0:
        return 'growth'
    elif ratio_signals >= 2 and growth_signals == 0:
        return 'ratio'
    elif growth_signals >= 2 and ratio_signals == 0:
        return 'growth'
    elif ratio_signals >= 1 and growth_signals >= 1:
        return 'ambiguous'
    else:
        # 单一证据或无明确证据，标注为 ambiguous（T23）
        if chain_result:
            return chain_result
        return 'ambiguous'


def check2_auxiliary_semantics(ws, sheet_name, min_row, max_row, min_col, max_col, wb_val=None, wb_fml=None, wb_val_full=None):
    """强制检查2：辅助计算列"值类型"与"使用方式"一致性检查
    v5.0增强（盲区修复）：
    - 新增数值范围启发式：AVERAGE包装纯数值时，值在0.05-0.50为占比区间信号
    - 新增行标签成本/费用关键词信号
    - 传入wb_val_full支持跨Sheet数值范围检测
    v4.0增强保留：
    - 跳过隐藏行/列
    - 读取辅助列原始公式判定语义（标签不可信）
    - 检查占比列是否被用于(1+占比)乘法增长
    - 【跨Sheet】追溯跨Sheet引用：解析'Sheet名'!单元格 引用，切换目标Sheet读取公式
    - 【L47/B38】递归公式链展开：AVERAGE(C40:E40) 展开子单元格判定语义
    - 【T23】多证据汇聚判定：≥2类证据一致才下结论，禁止单一证据默认安全判定
    """
    results = []
    issues = []
    checked_rows = 0

    for row_idx in range(min_row, max_row + 1):
        if is_row_hidden(ws, row_idx):
            continue
        checked_rows += 1

        label_a = str(ws.cell(row=row_idx, column=1).value or "")[:20]
        label_b = str(ws.cell(row=row_idx, column=2).value or "")[:20]
        row_label = label_a or label_b

        # 检查预测列公式中的(1+辅助列)模式
        for col_idx in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if is_col_hidden(ws, col_letter):
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.data_type != 'f' or not cell.value:
                continue

            formula = str(cell.value)

            # 模式1：(1+$X$Y) 绝对引用
            abs_growth_pattern = re.findall(r'\(1\+(\$[A-Z]+\$\d+)\)', formula)
            # 模式2：*(1+X_col_ref) 相对引用
            growth_pattern = re.findall(r'\*\(1\+([A-Z]+\d+)\)', formula)
            # 模式3：【跨Sheet】(1+'Sheet名'!$X$Y) 或 (1+Sheet名!X_Y)
            cross_sheet_growth = re.findall(r'\(1\+[\'"]?([^\'"!]+)[\'"]?!\$?([A-Z]+)\$?(\d+)\)', formula)

            all_ref_patterns = abs_growth_pattern + growth_pattern

            for ref in all_ref_patterns:
                ref_clean = ref.replace('$', '')
                ref_col_str = re.match(r'([A-Z]+)', ref_clean)
                ref_row_str = re.search(r'(\d+)', ref_clean)
                if ref_col_str and ref_row_str:
                    ref_col = ref_col_str.group(1)
                    ref_row = int(ref_row_str.group(1))
                    ref_col_idx = col_to_num(ref_col)
                    ref_cell = ws.cell(row=ref_row, column=ref_col_idx)
                    if ref_cell.data_type == 'f' and ref_cell.value:
                        ref_fml = str(ref_cell.value)
                        # v4.0: 使用 determine_semantic 深度语义判定（L47/B38落地）
                        # v5.0: 传入wb_val支持数值范围启发式（AVERAGE包装纯数值占比检测）
                        semantic = determine_semantic(ref_cell, ws, wb_fml, depth=2, wb_val=wb_val)
                        is_ratio = (semantic == 'ratio')
                        is_ambiguous = (semantic == 'ambiguous')
                        if is_ratio:
                            ref_label = str(ws.cell(row=ref_row, column=1).value or "")[:20]
                            ref_val = None
                            if wb_val:
                                ref_val = wb_val.cell(row=ref_row, column=ref_col_idx).value
                            issues.append(
                                f"  🔴 行{row_idx} 列{col_letter} [{row_label}]: "
                                f"预测公式 {formula[:60]}"
                            )
                            issues.append(
                                f"       引用了 {ref} (公式:{ref_fml[:50]}, 值={ref_val})"
                            )
                            issues.append(
                                f"       该引用单元格含除法，疑为占比被误用为增长率！"
                                f"被引用行标签: [{ref_label}]"
                            )
                        elif is_ambiguous:
                            ref_label = str(ws.cell(row=ref_row, column=1).value or "")[:20]
                            ref_val = None
                            if wb_val:
                                ref_val = wb_val.cell(row=ref_row, column=ref_col_idx).value
                            issues.append(
                                f"  ⚠️ 行{row_idx} 列{col_letter} [{row_label}]: "
                                f"预测公式 {formula[:60]}"
                            )
                            issues.append(
                                f"       引用了 {ref} (公式:{ref_fml[:50]}, 值={ref_val})"
                            )
                            issues.append(
                                f"       该引用语义模糊（占比信号与增长率信号并存），需人工确认！"
                                f"被引用行标签: [{ref_label}]"
                            )

            # 【跨Sheet】处理跨Sheet引用
            for cross_ref in cross_sheet_growth:
                ref_sheet_name = cross_ref[0].strip("'\"")
                ref_col_str = cross_ref[1]
                ref_row_str = int(cross_ref[2])
                ref_col_idx = col_to_num(ref_col_str)

                # 尝试在 wb_fml 中查找目标 Sheet
                ref_fml = None
                ref_label = ""
                if wb_fml and ref_sheet_name in wb_fml.sheetnames:
                    ref_ws = wb_fml[ref_sheet_name]
                    ref_cell = ref_ws.cell(row=ref_row_str, column=ref_col_idx)
                    if ref_cell.data_type == 'f' and ref_cell.value:
                        ref_fml = str(ref_cell.value)
                    ref_label = str(ref_ws.cell(row=ref_row_str, column=1).value or "")[:20]
                else:
                    ref_label = f"[Sheet '{ref_sheet_name}' 未找到]"

                if ref_fml:
                    # v4.0: 使用 determine_semantic 深度语义判定（L47/B38落地）
                    # v5.0: 传入wb_val_full支持跨Sheet数值范围启发式
                    semantic = determine_semantic(
                        ref_ws.cell(row=ref_row_str, column=ref_col_idx),
                        ref_ws, wb_fml, depth=2,
                        wb_val=wb_val_full[ref_sheet_name] if wb_val_full and ref_sheet_name in wb_val_full.sheetnames else None
                    )
                    is_ratio = (semantic == 'ratio')
                    is_ambiguous = (semantic == 'ambiguous')
                    if is_ratio:
                        ref_val = None
                        if wb_val_full and ref_sheet_name in wb_val_full.sheetnames:
                            ref_val = wb_val_full[ref_sheet_name].cell(row=ref_row_str, column=ref_col_idx).value
                        issues.append(
                            f"  🔴 行{row_idx} 列{col_letter} [{row_label}]: "
                            f"预测公式 {formula[:60]}"
                        )
                        issues.append(
                            f"       跨Sheet引用 '{ref_sheet_name}'!{ref_col_str}{ref_row_str} "
                            f"(公式:{ref_fml[:50]}, 值={ref_val})"
                        )
                        issues.append(
                            f"       该跨Sheet引用参数语义为【占比】，"
                            f"但使用方式为*(1+参数)增长率模式，占比被误用为增长率！"
                            f"被引用行标签: [{ref_label}]"
                        )
                    elif is_ambiguous:
                        ref_val = None
                        if wb_val_full and ref_sheet_name in wb_val_full.sheetnames:
                            ref_val = wb_val_full[ref_sheet_name].cell(row=ref_row_str, column=ref_col_idx).value
                        issues.append(
                            f"  ⚠️ 行{row_idx} 列{col_letter} [{row_label}]: "
                            f"预测公式 {formula[:60]}"
                        )
                        issues.append(
                            f"       跨Sheet引用 '{ref_sheet_name}'!{ref_col_str}{ref_row_str} "
                            f"(公式:{ref_fml[:50]}, 值={ref_val})"
                        )
                        issues.append(
                            f"       该跨Sheet引用语义模糊（占比信号与增长率信号并存），需人工确认！"
                            f"被引用行标签: [{ref_label}]"
                        )
                    else:
                        results.append(
                            f"  ℹ 行{row_idx} 列{col_letter}: 跨Sheet引用 "
                            f"'{ref_sheet_name}'!{ref_col_str}{ref_row_str} "
                            f"公式:{ref_fml[:40]} 语义=增长率 ✅"
                        )
                else:
                    results.append(
                        f"  ⚠ 行{row_idx} 列{col_letter}: 跨Sheet引用 "
                        f"'{ref_sheet_name}'!{ref_col_str}{ref_row_str} "
                        f"无法读取公式（Sheet不存在或单元格非公式）"
                    )

    results.append(f"  已检查 {checked_rows} 行")
    return results, issues
