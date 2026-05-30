"""
shared_utils.py — 评估Skill通用工具函数库 v1.0

集中存放各估值Skill重复使用的工具函数，避免同一函数在多个Skill中维护多份副本。

依赖:
  - re

包含函数:
  - get_sheet_prefix(sheet_name): 从sheet名提取前缀编码
  - compact_process_table(ws): 过程表空白行彻底清除与合计行上移

使用:
  import sys, os
  sys.path.insert(0, os.path.expanduser('~/.workbuddy/skills/valuation-common/scripts'))
  from shared_utils import get_sheet_prefix, compact_process_table

变更记录:
  - v1.0 (2026-05-23): 初始版本，从DT/AB Skill提取get_sheet_prefix和compact_process_table
"""

import re


# ─────────────────────────────────────────────────────
# get_sheet_prefix — 从sheet名提取前缀编码
# ─────────────────────────────────────────────────────
def get_sheet_prefix(sheet_name):
    """
    从sheet名提取前缀编码，如 '3-5应收账款' -> '3-5', '4-1-1房屋' -> '4-1-1'
    
    来源: 原DT Skill gate_validator.py:150 / validate_sheet_after_fill.py:149 的重复定义
    
    参数:
      sheet_name (str): Sheet名称，如 '3-5应收账款'
    返回:
      str | None: 前缀编码字符串，无法匹配则返回None
    """
    m = re.match(r'^([\d\-]+)', sheet_name)
    return m.group(1) if m else None


# ─────────────────────────────────────────────────────
# compact_process_table — 过程表空白行彻底清除与合计行上移
# ─────────────────────────────────────────────────────
def compact_process_table(ws):
    """
    彻底清除过程表空白行：移动合计行紧接数据行，重写所有公式
    
    来源: 原AB Skill compact_process_table.py:33 的独立实现
    
    参数:
      ws: openpyxl.worksheet.Worksheet 对象（过程表sheet）
    返回:
      int: 删除的空行数（0=无需清理）
    
    ⚠️ Phase 4 强制步骤，每个过程表必须调用
    ⚠️ 合计行公式必须重写，不可保留旧公式引用旧行号
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    from copy import copy

    # 1. 识别关键行
    header_row = total_row = first_data_row = last_data_row = None
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        if a_val and str(a_val).strip() == '序号':
            header_row = r
        if a_val and '合' in str(a_val):
            total_row = r
        if header_row and r > header_row and b_val and str(b_val).strip() and '合' not in str(b_val):
            if first_data_row is None:
                first_data_row = r
            last_data_row = r

    if not total_row or not header_row:
        return 0  # 无法处理

    data_start = first_data_row or (header_row + 1)
    data_end = last_data_row or header_row
    blank_count = total_row - data_end - 1

    if blank_count <= 0:
        return 0  # 无需清理

    # 2. 保存合计行完整数据
    total_row_data = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=c)
        if not isinstance(cell, MergedCell):
            total_row_data[c] = {
                'value': cell.value,
                'number_format': cell.number_format,
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
            }
    merges_to_restore = [str(mr) for mr in ws.merged_cells.ranges
                         if mr.min_row == total_row and mr.max_row == total_row]

    # 3. 清除旧合计行及空行区域
    for r in range(data_end + 1, total_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= data_end + 1 and mr.max_row <= total_row:
            ws.unmerge_cells(str(mr))

    # 4. 写入合计行到新位置
    new_total_row = data_end + 1
    for c, data in total_row_data.items():
        cell = ws.cell(row=new_total_row, column=c)
        cell.value = data['value']
        cell.number_format = data['number_format']
        cell.font = data['font']
        cell.fill = data['fill']
        cell.border = data['border']
        cell.alignment = data['alignment']

    # 恢复合并单元格
    for merge_str in merges_to_restore:
        parts = merge_str.split(':')
        col_start = ''.join(c for c in parts[0] if c.isalpha())
        col_end = ''.join(c for c in parts[1] if c.isalpha()) if len(parts) > 1 else col_start
        ws.merge_cells(f"{col_start}{new_total_row}:{col_end}{new_total_row}")

    # 5. 重写合计行公式（根据表头列含义）
    col_meanings = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)
        if not isinstance(cell, MergedCell) and cell.value:
            col_meanings[c] = str(cell.value).strip()

    bv_col = ev_col = va_col = vr_col = None
    for c, m in col_meanings.items():
        if '账面价值' in m: bv_col = c
        elif '评估价值' in m: ev_col = c
        elif '增值额' in m: va_col = c
        elif '增值率' in m: vr_col = c

    if bv_col:
        bv = get_column_letter(bv_col)
        ws.cell(row=new_total_row, column=bv_col).value = f"=ROUND(SUM({bv}{data_start}:{bv}{data_end}),2)"
    if ev_col:
        ev = get_column_letter(ev_col)
        ws.cell(row=new_total_row, column=ev_col).value = f"=ROUND(SUM({ev}{data_start}:{ev}{data_end}),2)"
    if va_col and bv_col and ev_col:
        va = get_column_letter(va_col)
        bv = get_column_letter(bv_col)
        ev = get_column_letter(ev_col)
        ws.cell(row=new_total_row, column=va_col).value = f"={ev}{new_total_row}-{bv}{new_total_row}"
    if vr_col and va_col and bv_col:
        vr = get_column_letter(vr_col)
        va = get_column_letter(va_col)
        bv = get_column_letter(bv_col)
        ws.cell(row=new_total_row, column=vr_col).value = f'=IF({bv}{new_total_row}=0,"",{va}{new_total_row}/{bv}{new_total_row}*100)'

    # 清除引用旧行号的非核心公式
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=new_total_row, column=c)
        if not isinstance(cell, MergedCell) and isinstance(cell.value, str) and cell.value.startswith('='):
            if c in [bv_col, ev_col, va_col, vr_col]:
                continue
            refs = re.findall(r'[A-Z]+(\d+)', cell.value)
            if any(int(r) > new_total_row for r in refs):
                cell.value = None

    # 6. 清除新合计行以下内容
    for r in range(new_total_row + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row > new_total_row:
            ws.unmerge_cells(str(mr))

    # 7. 设置打印区域和页面适配
    last_visible_col = 0
    for r in range(1, new_total_row + 1):
        for c in range(1, ws.max_column + 1):
            cl = get_column_letter(c)
            if not ws.column_dimensions[cl].hidden:
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell) and cell.value is not None:
                    last_visible_col = max(last_visible_col, c)
    ws.print_area = f"A1:{get_column_letter(last_visible_col)}{new_total_row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'
    import openpyxl.worksheet.properties
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    return blank_count  # 返回删除的空行数
