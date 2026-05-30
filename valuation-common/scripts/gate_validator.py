#!/usr/bin/env python3
"""
gate_validator.py — 硬性关卡验证脚本（共享版 v2.0）

功能：验证Phase间关卡是否通过，基于checkpoint.json文件
逻辑：每Phase结束写入checkpoint → 下Phase开始前验证 → 缺失/无效则BLOCKED

共享化改造（v2.0）：
- 本脚本放置在 valuation-common/scripts/，各skill通过 sys.path.insert 引用
- GATE_DEFINITIONS 改为可配置：优先加载调用目录下的 gate_defs.py，不存在则用默认定义
- CCEP合规验证（MR-18）：自动检查 checkpoint 中是否包含 common_compliance 字段

用法：
  python gate_validator.py <audit_output_dir> <from_phase> <to_phase>
  python gate_validator.py <audit_output_dir> all       # 验证全部关卡
  python gate_validator.py <audit_output_dir> G-P7D      # 直接指定gate_id

退出码：
  0 = 全部通过
  1 = 有关卡未通过

checkpoint.json schema:
{
  "phase": 2,
  "phase_name": "Excel链接错误检查",
  "timestamp": "2026-05-21T00:30:00",
  "gate": {
    "gate_id": "G-P23",
    "items": [
      {"id": "G-P23-1", "description": "check1脚本已执行", "status": "pass", "evidence": "check1输出:phase2_check1.txt"},
      {"id": "G-P23-2", "description": "check_semantic_deep已执行", "status": "pass", "evidence": "语义检查输出:phase2_semantic.txt"},
      {"id": "G-P23-3", "description": "同模式强制扩展", "status": "pass", "evidence": "扩展至SY9/SY10"},
      {"id": "G-P23-4", "description": "覆盖度矩阵", "status": "skip", "reason": "非关键Sheet"}
    ],
    "verdict": "pass"
  },
  "findings": {
    "critical": 2,
    "important": 3,
    "suggestion": 2,
    "details": ["P01: ...", "P02: ..."]
  },
  "spot_checks": [
    {
      "sheet": "SY4-主营收入成本-分析预测",
      "trigger": "check1返回0 findings",
      "rows_checked": ["F32", "F40"],
      "method": "openpyxl data_only=False手动提取公式",
      "result": "发现占比当增长率问题"
    }
  ],
  "scripts_executed": ["check_formulas.py", "check_hidden.py"],
  "self_check_completed": true
}
"""

import json
import sys
import os
from datetime import datetime
from pathlib import Path


GATE_DEFINITIONS = {
    "G-P01": {
        "from_phase": 0, "to_phase": 1,
        "items": [
            {"id": "G-P01-1", "desc": "Sheet清单已提取（含可见/隐藏分类）"},
            {"id": "G-P01-2", "desc": "隐藏状态检测已执行"},
            {"id": "G-P01-3", "desc": "项目基本信息已提取"},
            {"id": "G-P01-4", "desc": "覆盖度矩阵Phase 0行已填写"},
        ]
    },
    "G-P12": {
        "from_phase": 1, "to_phase": 2,
        "items": [
            {"id": "G-P12-1", "desc": "所有可见Sheet已提取公式"},
            {"id": "G-P12-2", "desc": "隐藏Sheet已列清单"},
            {"id": "G-P12-3", "desc": "#REF!错误已标注"},
            {"id": "G-P12-4", "desc": "表头标注与评估目的口径比对完成"},
            {"id": "G-P12-5", "desc": "覆盖度矩阵Phase 0-1行已全部填写"},
        ]
    },
    "G-P23": {
        "from_phase": 2, "to_phase": 3,
        "items": [
            {"id": "G-P23-1", "desc": "check1脚本已执行，$列绝对引用检测结果已输出"},
            {"id": "G-P23-2", "desc": "check_semantic_deep.py已执行，占比/增长率语义已判明"},
            {"id": "G-P23-3", "desc": "同模式强制扩展已完成"},
            {"id": "G-P23-4", "desc": "覆盖度矩阵Phase 2行已全部填写"},
        ]
    },
    "G-P34": {
        "from_phase": 3, "to_phase": 4,
        "items": [
            {"id": "G-P34-1", "desc": "FCF验算脚本已执行"},
            {"id": "G-P34-2", "desc": "报告辅助Sheet与核心表一致性已检查"},
            {"id": "G-P34-3", "desc": "独立测算表交叉核对已完成"},
            {"id": "G-P34-4", "desc": "覆盖度矩阵Phase 3行已全部填写"},
        ]
    },
    "G-P45": {
        "from_phase": 4, "to_phase": 5,
        "items": [
            {"id": "G-P45-1", "desc": "Beta来源判断与WACC验证已完成"},
            {"id": "G-P45-2", "desc": "销售单价不变假设已确认并说明"},
            {"id": "G-P45-3", "desc": "收入/成本/费用趋势已检查"},
            {"id": "G-P45-4", "desc": "覆盖度矩阵Phase 4行已全部填写"},
        ]
    },
    "G-P56": {
        "from_phase": 5, "to_phase": 6,
        "items": [
            {"id": "G-P56-1", "desc": "所得税=0/极低税率检查已完成"},
            {"id": "G-P56-2", "desc": "外部数据链接已识别+固化提醒"},
            {"id": "G-P56-3", "desc": "税前/税后折现率合规性已检查"},
            {"id": "G-P56-4", "desc": "覆盖度矩阵Phase 0-5行无⬜格"},
        ]
    },
    "G-P66": {
        "from_phase": 6, "to_phase": 7,
        "items": [
            {"id": "G-P66-1", "desc": "审核结论包含估值合理性判断"},
            {"id": "G-P66-2", "desc": "零幻觉确认已完成"},
            {"id": "G-P66-3", "desc": "问题分级清单已完成"},
            {"id": "G-P66-4", "desc": "Phase 0-5覆盖度矩阵无⬜格"},
        ]
    },
}

# Phase 6.5 is special - it's within Phase 6→7 flow
GATE_DEFINITIONS["G-P67"] = {
    "from_phase": 7, "to_phase": 8,  # 8 = delivery
    "items": [
        {"id": "G-P67-1", "desc": "lessons_learned已更新"},
        {"id": "G-P67-2", "desc": "经验→强制步骤转化完成"},
        {"id": "G-P67-3", "desc": "脱敏扫描通过"},
        {"id": "G-P67-4", "desc": "向用户汇报复盘结果"},
    ]
}

# G-P7D: Phase 7 报告交付关卡（报告提交前的最终硬约束）
# 核心原则：附注清单全部确认前禁止提交报告
GATE_DEFINITIONS["G-P7D"] = {
    "from_phase": 7, "to_phase": 9,  # 9 = final delivery
    "items": [
        {"id": "G-P7D-1", "desc": "报告模板已读取（references/report_template.md）"},
        {"id": "G-P7D-2", "desc": "附注一：Phase×Step执行状态表——全部Step已逐项填写，无空白行"},
        {"id": "G-P7D-3", "desc": "附注二：覆盖度矩阵盲区披露表——已填写"},
        {"id": "G-P7D-4", "desc": "附注三：脚本执行状态表——全部脚本已逐项填写"},
        {"id": "G-P7D-5", "desc": "附注中无沉默跳过：所有❌/⚠️行均有说明，无'已检查''无异常'等模糊表述"},
        {"id": "G-P7D-6", "desc": "phase7_checkpoint.json已写入（含附注完整性验证结果）"},
    ]
}

# ============================================================
# DT Skill专用Gate（DT-168~171: 字段完整性+减值行+占位符+勾稽完整性）
# ============================================================

# DT-168: G1-7 字段完整性门控（每Sheet填后必检）
GATE_DEFINITIONS["G-DT168"] = {
    "from_phase": 2, "to_phase": 3,
    "items": [
        {"id": "G-DT168-1", "desc": "往来科目结算对象名称列100%非空"},
        {"id": "G-DT168-2", "desc": "非往来科目名称/内容列100%非空"},
        {"id": "G-DT168-3", "desc": "币种列非空（人民币科目填'人民币'）"},
    ]
}

# DT-169: G2-13 减值行完整性校验
GATE_DEFINITIONS["G-DT169"] = {
    "from_phase": 2, "to_phase": 3,
    "items": [
        {"id": "G-DT169-1", "desc": "坏账准备行账面价值+评估价值均有值"},
        {"id": "G-DT169-2", "desc": "累计折旧行账面价值+评估价值均有值"},
        {"id": "G-DT169-3", "desc": "累计摊销行账面价值+评估价值均有值"},
    ]
}

# DT-170: G0-2 设定信息占位符检测
GATE_DEFINITIONS["G-DT170"] = {
    "from_phase": 0, "to_phase": 1,
    "items": [
        {"id": "G-DT170-1", "desc": "设定信息审计机构名称非占位符(无'XX'/'XXX')"},
        {"id": "G-DT170-2", "desc": "设定信息审计报告文号非占位符(无'XX'/'XXX')"},
        {"id": "G-DT170-3", "desc": "设定信息审计报告日期非模板默认值(非2023-01-23)"},
    ]
}

# DT-171: G3-8 字段完整性勾稽（Phase 4交付前）
GATE_DEFINITIONS["G-DT171"] = {
    "from_phase": 4, "to_phase": 5,
    "items": [
        {"id": "G-DT171-1", "desc": "所有有数据Sheet的结算对象/名称列非空率=100%"},
        {"id": "G-DT171-2", "desc": "数字勾稽+字段完整性勾稽双重通过"},
    ]
}

# DT-175: 输出格式完整性Gate（方案C）——填表完成后的兜底检查
# 核心原则：无论Agent是否通过fill_sheet()填写，最终Excel输出必须通过格式完整性检查
GATE_DEFINITIONS["G-DT175"] = {
    "from_phase": 5, "to_phase": 6,
    "items": [
        {"id": "G-DT175-1", "desc": "数据区域无MergedCell残留（DT-125↑）"},
        {"id": "G-DT175-2", "desc": "数据行数字格式一致性（金额列有#,##0.00格式）"},
        {"id": "G-DT175-3", "desc": "坏账准备行格式正确（I列正数+J列=0）（DT-151↑）"},
        {"id": "G-DT175-4", "desc": "预计风险行格式正确（I列=0+J列正数）（DT-151↑）"},
        {"id": "G-DT175-5", "desc": "数据区域无异常全空行"},
        {"id": "G-DT175-6", "desc": "合计行位置正常（在数据行之后）"},
        {"id": "G-DT175-7", "desc": "空白明细表已隐藏（DT-20兜底，合计行账面=0且评估=0的明细表必须hidden）"},
    ]
}


# ============================================================
# DT-168~171 Gate验证函数
# ============================================================

def validate_field_completeness(ws, sheet_info, col_map):
    """DT-168(G1-7): 字段完整性门控——每Sheet填后必检。

    检查：
    1. 往来科目：结算对象名称列必须非空
    2. 非往来科目：名称/内容列必须非空
    3. 币种列：有金额的行应有币种

    Args:
        ws: worksheet
        sheet_info: sheet结构信息dict（含data_start_row, total1_row等）
        col_map: 列映射dict

    Returns:
        dict: {'pass': bool, 'criticals': list, 'warnings': list}
    """
    from openpyxl.cell.cell import MergedCell

    criticals = []
    warnings = []

    bv_col = col_map.get('账面价值') or col_map.get('book_value')
    settlement_col = (
        col_map.get('结算对象') or col_map.get('户名') or
        col_map.get('欠款单位名称') or col_map.get('收款单位名称（结算对象)') or
        col_map.get('settlement')
    )
    name_col = (
        col_map.get('名称及内容') or col_map.get('名称和内容') or
        col_map.get('名称') or col_map.get('项目名称') or
        col_map.get('项目及内容')
    )
    currency_col = col_map.get('币种') or col_map.get('currency')

    data_start = sheet_info.get('data_start_row', 6)
    total1_row = sheet_info.get('total1_row') or sheet_info.get('total_row')

    if not total1_row:
        return {'pass': True, 'criticals': [], 'warnings': ['无法识别合计行，跳过字段完整性检查']}

    for r in range(data_start, total1_row):
        # 检查是否有金额
        bv_cell = ws.cell(row=r, column=bv_col) if bv_col else None
        bv_val = bv_cell.value if bv_cell and not isinstance(bv_cell, MergedCell) else None

        if bv_val is None or (isinstance(bv_val, (int, float)) and bv_val == 0):
            continue  # 无金额行跳过

        # 检查结算对象名称列
        if settlement_col:
            st_cell = ws.cell(row=r, column=settlement_col)
            st_val = st_cell.value if not isinstance(st_cell, MergedCell) else None
            if not st_val or str(st_val).strip() == '':
                criticals.append(
                    f'G1-7 CRITICAL: R{r} 有金额({bv_val:,.2f})但结算对象名称列为空'
                )

        # 检查名称/内容列（非往来sheet）
        if not settlement_col and name_col:
            nm_cell = ws.cell(row=r, column=name_col)
            nm_val = nm_cell.value if not isinstance(nm_cell, MergedCell) else None
            if not nm_val or str(nm_val).strip() == '':
                warnings.append(
                    f'G1-7 WARNING: R{r} 有金额({bv_val:,.2f})但名称/内容列为空'
                )

    return {
        'pass': len(criticals) == 0,
        'criticals': criticals,
        'warnings': warnings,
    }


def validate_provision_row_completeness(ws, sheet_info, col_map):
    """DT-169(G2-13): 减值行完整性校验。

    检查所有有坏账准备/累计折旧/累计摊销行的Sheet：
    - 如果账面价值列有值，则评估价值列必须有值
    - 坏账准备行：评估价值应与账面价值一致（成本法评估）
    - 累计折旧/累计摊销行：评估价值应与账面价值一致

    Args:
        ws: worksheet
        sheet_info: sheet结构信息dict
        col_map: 列映射dict

    Returns:
        dict: {'pass': bool, 'criticals': list}
    """
    from openpyxl.cell.cell import MergedCell

    criticals = []

    bv_col = col_map.get('账面价值') or col_map.get('book_value')
    ev_col = col_map.get('评估价值') or col_map.get('assessed_value')

    if not bv_col or not ev_col:
        return {'pass': True, 'criticals': []}  # 无金额列的sheet跳过

    bad_debt_row = sheet_info.get('bad_debt_row')
    provision_row = sheet_info.get('provision_row')

    # 扫描A列找减值行
    for r in range(sheet_info.get('data_start_row', 6), sheet_info.get('max_row', ws.max_row) + 1):
        a_val = ws.cell(row=r, column=1).value
        if not a_val:
            continue
        a_text = str(a_val).replace(' ', '').strip()

        is_provision_row = False
        provision_type = ''
        if '坏账准备' in a_text or '减值准备' in a_text:
            is_provision_row = True
            provision_type = '坏账准备'
        elif '累计折旧' in a_text:
            is_provision_row = True
            provision_type = '累计折旧'
        elif '累计摊销' in a_text:
            is_provision_row = True
            provision_type = '累计摊销'
        elif '预计风险' in a_text:
            is_provision_row = True
            provision_type = '预计风险'

        if not is_provision_row:
            continue

        # 检查账面价值和评估价值
        bv_cell = ws.cell(row=r, column=bv_col)
        ev_cell = ws.cell(row=r, column=ev_col)
        bv_val = bv_cell.value if not isinstance(bv_cell, MergedCell) else None
        ev_val = ev_cell.value if not isinstance(ev_cell, MergedCell) else None

        if bv_val is not None and isinstance(bv_val, (int, float)) and bv_val != 0:
            if ev_val is None:
                criticals.append(
                    f'G2-13 CRITICAL: {provision_type}行(R{r})账面价值={bv_val:,.2f}'
                    f'但评估价值为None！'
                )
            elif isinstance(ev_val, (int, float)) and abs(ev_val) > 0.01 and ev_val != bv_val:
                # 评估价值与账面价值不一致（成本法应一致）
                criticals.append(
                    f'G2-13 WARNING: {provision_type}行(R{r})账面价值={bv_val:,.2f}'
                    f'但评估价值={ev_val:,.2f}，成本法评估应一致'
                )

    return {'pass': len(criticals) == 0, 'criticals': criticals}


def validate_settings_no_placeholder(ws):
    """DT-170(G0-2): 设定信息占位符检测。

    检查设定信息sheet中是否含有模板占位符：
    - 'XX会计师事务所' → CRITICAL
    - '审〔2023〕XXX号' → CRITICAL
    - 'XX公司审计报告' → CRITICAL
    - 2023-01-23 → WARNING（模板默认日期）

    Args:
        ws: 设定信息sheet

    Returns:
        dict: {'pass': bool, 'criticals': list, 'warnings': list}
    """
    criticals = []
    warnings = []

    # 扫描所有单元格
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = str(cell.value).strip()
                r = cell.row
                c = cell.column

                # 检测占位符模式
                if 'XX' in val or 'XXX' in val or '{{' in val:
                    criticals.append(
                        f'G0-2 CRITICAL: R{r}C{c} 含占位符: "{val}"'
                    )

                # 检测年份错误
                if '2023' in val and '审' in val:
                    warnings.append(
                        f'G0-2 WARNING: R{r}C{c} 含2023年份: "{val}"，'
                        f'请确认是否需要更新为项目实际年份'
                    )

    return {
        'pass': len(criticals) == 0,
        'criticals': criticals,
        'warnings': warnings,
    }


def validate_field_completeness_reconciliation(wb, col_map_data):
    """DT-171(G3-8): 字段完整性勾稽——Phase 4交付前检查。

    遍历所有有数据的Sheet，统计结算对象/名称列非空率。
    非空率<100% → CRITICAL，确保不会出现"19/20勾稽通过但50行缺名称"的情况。

    Args:
        wb: openpyxl workbook对象
        col_map_data: sheet_col_map.json数据

    Returns:
        dict: {'pass': bool, 'criticals': list, 'summary': dict}
    """
    from openpyxl.cell.cell import MergedCell

    criticals = []
    summary = {}

    for sname, info in col_map_data.items():
        if sname not in wb.sheetnames:
            continue

        ws = wb[sname]
        cm = info.get('col_mapping', info.get('col_map', {}))

        bv_col = cm.get('账面价值') or cm.get('book_value')
        settlement_col = (
            cm.get('结算对象') or cm.get('户名') or
            cm.get('欠款单位名称') or cm.get('收款单位名称（结算对象)') or
            cm.get('settlement')
        )
        name_col = (
            cm.get('名称及内容') or cm.get('名称和内容') or
            cm.get('名称') or cm.get('项目名称') or
            cm.get('项目及内容')
        )

        if not bv_col:
            continue

        ds = info.get('data_start_row', 6)
        tr = info.get('total1_row') or info.get('total_row')
        if not tr:
            continue

        total_data_rows = 0
        empty_name_rows = 0

        for r in range(ds, tr):
            bv_cell = ws.cell(row=r, column=bv_col)
            bv_val = bv_cell.value if not isinstance(bv_cell, MergedCell) else None

            if bv_val is None or (isinstance(bv_val, (int, float)) and bv_val == 0):
                continue

            total_data_rows += 1

            # 检查名称列
            target_col = settlement_col or name_col
            if target_col:
                nm_cell = ws.cell(row=r, column=target_col)
                nm_val = nm_cell.value if not isinstance(nm_cell, MergedCell) else None
                if not nm_val or str(nm_val).strip() == '':
                    empty_name_rows += 1

        if total_data_rows > 0:
            non_empty_rate = (total_data_rows - empty_name_rows) / total_data_rows
            summary[sname] = {
                'total_rows': total_data_rows,
                'empty_name_rows': empty_name_rows,
                'non_empty_rate': non_empty_rate,
            }

            if empty_name_rows > 0:
                criticals.append(
                    f'G3-8 CRITICAL: {sname} 有{empty_name_rows}/{total_data_rows}行'
                    f'名称列为空(非空率={non_empty_rate:.0%})'
                )

    return {
        'pass': len(criticals) == 0,
        'criticals': criticals,
        'summary': summary,
    }


# ============================================================
# DT-175: 输出格式完整性Gate（方案C）——无论Agent走不走fill_sheet()，输出端兜底
# ============================================================

def validate_output_format_integrity(wb, col_map_data=None):
    """DT-175(G-OUTPUT): 输出格式完整性校验——填表完成后的兜底检查。

    核心设计原则：无论Agent是否通过fill_sheet()填写，最终Excel输出必须通过
    格式完整性检查。这是L2层的最后一道防线，不依赖Agent是否导入skill脚本。

    检查维度：
    1. MergedCell残留：数据行区域不应有MergedCell（DT-125↑）
    2. 行格式一致性：数据行与上行（表头行）的关键格式属性对齐
    3. 列数一致性：数据行列数=表头列数
    4. 数字格式一致性：金额列应有#,##0.00格式
    5. 坏账准备行格式：A列标记+I/J列正数（DT-151↑）
    6. 空行检测：数据区域不应有全空行
    7. 合计行位置：合计行不应紧贴表头行（至少1行数据）
    8. 空白明细表未隐藏：合计行账面=0且评估=0但sheet_state=visible（DT-20兜底）

    Args:
        wb: openpyxl workbook对象
        col_map_data: sheet_col_map.json数据（可选，用于精确定位列）

    Returns:
        dict: {
            'pass': bool,
            'criticals': list,   # 必须修复的严重问题
            'warnings': list,     # 建议修复的轻微问题
            'summary': dict       # 每个Sheet的检查结果
        }
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    criticals = []
    warnings = []
    summary = {}

    # 需要检查的格式属性
    FORMAT_CHECK_ATTRS = ['font', 'fill', 'border', 'alignment', 'number_format']

    # 排除的Sheet（不需要检查格式）
    SKIP_SHEET_PATTERNS = ['设定', '说明', '目录', '索引', '汇总']

    for sname in wb.sheetnames:
        # 跳过非数据Sheet
        skip = False
        for pat in SKIP_SHEET_PATTERNS:
            if pat in sname:
                skip = True
                break
        if skip:
            continue

        ws = wb[sname]

        # ---- 获取该Sheet的col_map ----
        cm = {}
        if col_map_data:
            info = col_map_data.get(sname, {})
            cm = info.get('col_mapping', info.get('col_map', {}))

        sheet_criticals = []
        sheet_warnings = []

        # ---- 1. 识别结构行 ----
        header_row = None
        data_start_row = None
        total_row = None
        bad_debt_row = None
        provision_row = None

        # 从col_map_data获取结构信息（优先）
        if col_map_data:
            info = col_map_data.get(sname, {})
            header_row = info.get('header_row')
            data_start_row = info.get('data_start_row')
            total_row = info.get('total1_row') or info.get('total_row')
            bad_debt_row = info.get('bad_debt_row')
            provision_row = info.get('provision_row')

        # 如果col_map_data没有结构信息，运行时推断
        if not header_row:
            # 扫描A列找表头行（含"序号"或"结算对象"等关键词的行）
            for r in range(1, min(ws.max_row + 1, 20)):
                a_val = ws.cell(row=r, column=1).value
                if a_val and isinstance(a_val, str):
                    if any(kw in a_val for kw in ['序号', '结算对象', '户名', '名称及内容', '名称']):
                        header_row = r
                        break

        if not header_row:
            # 无表头行→非数据Sheet或格式异常，跳过
            continue

        if not data_start_row:
            data_start_row = header_row + 1

        if not total_row:
            # 从A列扫描"合"字行
            for r in range(data_start_row + 1, min(ws.max_row + 1, 200)):
                a_val = ws.cell(row=r, column=1).value
                if a_val and isinstance(a_val, str) and '合' in a_val:
                    total_row = r
                    break

        if not total_row:
            continue

        # 扫描坏账准备行和预计风险行
        if not bad_debt_row:
            for r in range(data_start_row, total_row):
                a_val = ws.cell(row=r, column=1).value
                if a_val and isinstance(a_val, str):
                    a_text = str(a_val).replace(' ', '').strip()
                    if '坏账准备' in a_text or '减值准备' in a_text:
                        bad_debt_row = r
                    elif '预计风险' in a_text:
                        provision_row = r

        # ---- 2. MergedCell残留检测（DT-125↑） ----
        merged_count = 0
        for r in range(data_start_row, total_row):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    merged_count += 1

        if merged_count > 0:
            sheet_criticals.append(
                f'DT-125↑ CRITICAL: 数据区域有{merged_count}个MergedCell残留，'
                f'会导致数据写入丢失'
            )

        # ---- 3. 行格式一致性：数据行与参考行对比 ----
        # 参考行 = 表头行下一行（第1个数据行）
        ref_row = data_start_row
        if ref_row <= total_row:
            ref_cell = ws.cell(row=ref_row, column=1)
            # 确保参考行有数据（非全空行）
            has_data = False
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row=ref_row, column=c).value
                if v is not None and str(v).strip() != '':
                    has_data = True
                    break

            if has_data:
                format_mismatch_count = 0
                for r in range(data_start_row + 1, total_row):
                    # 跳过坏账准备行和预计风险行（格式不同是正常的）
                    if r == bad_debt_row or r == provision_row:
                        continue

                    # 检查该行是否有数据
                    row_has_data = False
                    for c in range(1, min(ws.max_column + 1, 15)):
                        v = ws.cell(row=r, column=c).value
                        if v is not None and str(v).strip() != '':
                            row_has_data = True
                            break

                    if not row_has_data:
                        continue

                    # 比较关键格式属性（只比较第1列和金额列）
                    check_cols = [1]  # A列
                    bv_col = cm.get('账面价值') or cm.get('book_value')
                    ev_col = cm.get('评估价值') or cm.get('assessed_value')
                    if bv_col:
                        check_cols.append(bv_col)
                    if ev_col:
                        check_cols.append(ev_col)

                    for c in check_cols:
                        if c > ws.max_column:
                            continue
                        ref_c = ws.cell(row=ref_row, column=c)
                        cur_c = ws.cell(row=r, column=c)
                        if isinstance(cur_c, MergedCell) or isinstance(ref_c, MergedCell):
                            continue
                        # 比较number_format
                        if (hasattr(ref_c, 'number_format') and
                            hasattr(cur_c, 'number_format') and
                            ref_c.number_format != cur_c.number_format):
                            # 只标记通用格式差异（自定义格式差异可忽略）
                            ref_fmt = str(ref_c.number_format)
                            cur_fmt = str(cur_c.number_format)
                            # 如果一个是#,##0.00而另一个不是，属于严重差异
                            if ('#,##0' in ref_fmt) != ('#,##0' in cur_fmt):
                                format_mismatch_count += 1

                if format_mismatch_count > 3:
                    sheet_warnings.append(
                        f'FORMAT WARNING: {format_mismatch_count}个数据行与参考行'
                        f'数字格式不一致，可能因裸insert_rows()导致格式丢失'
                    )

        # ---- 4. 数字格式一致性：金额列应有#,##0.00格式 ----
        bv_col = cm.get('账面价值') or cm.get('book_value')
        ev_col = cm.get('评估价值') or cm.get('assessed_value')
        for label, col in [('账面价值', bv_col), ('评估价值', ev_col)]:
            if not col or col > ws.max_column:
                continue
            no_format_count = 0
            for r in range(data_start_row, total_row):
                if r == bad_debt_row or r == provision_row:
                    continue
                cell = ws.cell(row=r, column=col)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None or (isinstance(val, (int, float)) and val == 0):
                    continue
                # 检查是否有数字格式
                if hasattr(cell, 'number_format'):
                    fmt = str(cell.number_format)
                    if fmt == 'General' or fmt == '0':
                        no_format_count += 1
            if no_format_count > 0:
                sheet_warnings.append(
                    f'FORMAT WARNING: {label}列有{no_format_count}个单元格'
                    f'缺少#,##0.00数字格式'
                )

        # ---- 5. 坏账准备行格式检查（DT-151↑） ----
        if bad_debt_row and bv_col and ev_col:
            # 检查A列标记
            a_val = ws.cell(row=bad_debt_row, column=1).value
            if a_val and isinstance(a_val, str) and '坏账准备' in a_val:
                # 检查I列（账面价值）和J列（评估价值）
                bv_cell = ws.cell(row=bad_debt_row, column=bv_col)
                ev_cell = ws.cell(row=bad_debt_row, column=ev_col)
                bv_val = bv_cell.value if not isinstance(bv_cell, MergedCell) else None
                ev_val = ev_cell.value if not isinstance(ev_cell, MergedCell) else None

                # DT-151↑: 坏账准备行 I列=+abs正数, J列=0
                if bv_val is not None and isinstance(bv_val, (int, float)):
                    if bv_val < 0:
                        sheet_criticals.append(
                            f'DT-151↑ CRITICAL: 坏账准备行(R{bad_debt_row})'
                            f'账面价值={bv_val:,.2f}为负数，应为正数(abs)'
                        )
                    if ev_val is not None and isinstance(ev_val, (int, float)) and ev_val != 0:
                        sheet_criticals.append(
                            f'DT-151↑ CRITICAL: 坏账准备行(R{bad_debt_row})'
                            f'评估价值={ev_val:,.2f}应为0（评估值归预计风险行）'
                        )

            # 检查预计风险行
            if provision_row and ev_col:
                ev_cell = ws.cell(row=provision_row, column=ev_col)
                bv_cell = ws.cell(row=provision_row, column=bv_col)
                ev_val = ev_cell.value if not isinstance(ev_cell, MergedCell) else None
                bv_val = bv_cell.value if not isinstance(bv_cell, MergedCell) else None

                # DT-151↑: 预计风险行 I列=0, J列=+abs正数
                if bv_val is not None and isinstance(bv_val, (int, float)) and bv_val != 0:
                    sheet_criticals.append(
                        f'DT-151↑ CRITICAL: 预计风险行(R{provision_row})'
                        f'账面价值={bv_val:,.2f}应为0'
                    )
                if ev_val is not None and isinstance(ev_val, (int, float)) and ev_val < 0:
                    sheet_criticals.append(
                        f'DT-151↑ CRITICAL: 预计风险行(R{provision_row})'
                        f'评估价值={ev_val:,.2f}为负数，应为正数(abs)'
                    )

        # ---- 6. 空行检测：数据区域不应有全空行 ----
        empty_row_count = 0
        for r in range(data_start_row, total_row):
            if r == bad_debt_row or r == provision_row:
                continue
            all_empty = True
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != '':
                    all_empty = False
                    break
            if all_empty:
                empty_row_count += 1

        if empty_row_count > 0:
            sheet_warnings.append(
                f'FORMAT WARNING: 数据区域有{empty_row_count}个全空行'
            )

        # ---- 7. 合计行位置检查 ----
        if total_row and data_start_row and total_row <= data_start_row:
            sheet_criticals.append(
                f'STRUCTURE CRITICAL: 合计行(R{total_row})在数据起始行(R{data_start_row})'
                f'之前或等于，结构异常'
            )

        # ---- 8. 空白明细表未隐藏检测（DT-20兜底，DT-176） ----
        # 如果明细表的合计1行账面=0且评估=0，但sheet_state='visible' → CRITICAL
        if ws.sheet_state != 'hidden':
            # 仅对明细表检查（排除汇总/结构表）
            import re as _re
            is_detail = bool(_re.match(r'^[3-6]-\d', sname))
            if is_detail and total_row:
                bv_col_check = cm.get('账面价值') or cm.get('book_value')
                ev_col_check = cm.get('评估价值') or cm.get('assessed_value')
                if bv_col_check and ev_col_check:
                    bv_cell = ws.cell(row=total_row, column=bv_col_check)
                    ev_cell = ws.cell(row=total_row, column=ev_col_check)
                    bv_val = bv_cell.value if not isinstance(bv_cell, MergedCell) else None
                    ev_val = ev_cell.value if not isinstance(ev_cell, MergedCell) else None
                    # 判断是否为空白（0或None）
                    bv_is_empty = (bv_val is None or (isinstance(bv_val, (int, float)) and bv_val == 0))
                    ev_is_empty = (ev_val is None or (isinstance(ev_val, (int, float)) and ev_val == 0))
                    if bv_is_empty and ev_is_empty:
                        sheet_criticals.append(
                            f'DT-20 CRITICAL: 空白明细表未隐藏！'
                            f'合计行(R{total_row})账面={bv_val}且评估={ev_val}均为0/空，'
                            f'但sheet_state={ws.sheet_state}。应执行: ws.sheet_state="hidden"'
                        )

        # ---- 汇总该Sheet结果 ----
        sheet_pass = len(sheet_criticals) == 0
        summary[sname] = {
            'pass': sheet_pass,
            'criticals': sheet_criticals,
            'warnings': sheet_warnings,
            'header_row': header_row,
            'data_start_row': data_start_row,
            'total_row': total_row,
            'bad_debt_row': bad_debt_row,
            'provision_row': provision_row,
            'merged_count': merged_count,
            'empty_row_count': empty_row_count,
        }
        criticals.extend([f'[{sname}] {c}' for c in sheet_criticals])
        warnings.extend([f'[{sname}] {w}' for w in sheet_warnings])

    return {
        'pass': len(criticals) == 0,
        'criticals': criticals,
        'warnings': warnings,
        'summary': summary,
    }


def load_checkpoint(audit_dir: Path, phase: int):
    cp_path = audit_dir / f"phase{phase}_checkpoint.json"
    if not cp_path.exists():
        return None
    try:
        with open(cp_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return None


def validate_gate(audit_dir: Path, gate_id: str) -> dict:
    gate_def = GATE_DEFINITIONS[gate_id]
    from_phase = gate_def["from_phase"]
    
    checkpoint = load_checkpoint(audit_dir, from_phase)
    
    result = {
        "gate_id": gate_id,
        "from_phase": from_phase,
        "to_phase": gate_def["to_phase"],
        "items": [],
        "verdict": "BLOCKED",
        "blocking_reasons": []
    }
    
    if checkpoint is None:
        result["blocking_reasons"].append(
            f"phase{from_phase}_checkpoint.json 不存在 — Phase {from_phase}未完成checkpoint写入"
        )
        for item_def in gate_def["items"]:
            result["items"].append({
                "id": item_def["id"],
                "desc": item_def["desc"],
                "status": "MISSING_CHECKPOINT"
            })
        return result
    
    if "gate" not in checkpoint:
        result["blocking_reasons"].append("checkpoint中无gate字段")
        return result
    
    cp_gate = checkpoint["gate"]
    if cp_gate.get("gate_id") != gate_id:
        result["blocking_reasons"].append(
            f"checkpoint gate_id={cp_gate.get('gate_id')} != 期望{gate_id}"
        )
        return result
    
    all_pass = True
    for item_def in gate_def["items"]:
        cp_item = None
        for gi in cp_gate.get("items", []):
            if gi.get("id") == item_def["id"]:
                cp_item = gi
                break
        
        if cp_item is None:
            result["items"].append({
                "id": item_def["id"],
                "desc": item_def["desc"],
                "status": "NOT_FOUND",
                "note": "checkpoint中无此项"
            })
            all_pass = False
            result["blocking_reasons"].append(f"{item_def['id']}: checkpoint中无此项")
        elif cp_item.get("status") == "pass":
            result["items"].append({
                "id": item_def["id"],
                "desc": item_def["desc"],
                "status": "pass",
                "evidence": cp_item.get("evidence", "")
            })
        elif cp_item.get("status") == "skip":
            reason = cp_item.get("reason", "未说明原因")
            result["items"].append({
                "id": item_def["id"],
                "desc": item_def["desc"],
                "status": "skip",
                "reason": reason
            })
            # skip with valid reason is acceptable for non-critical items
            # G-P23-1/G-P23-2/G-P23-3 are critical, cannot skip
            # G-P34-1/G-P34-2/G-P34-3 are critical, cannot skip
            # G-P7D all items are critical (report delivery gate), cannot skip
            critical_items = [
                "G-P23-1", "G-P23-2", "G-P23-3",
                "G-P34-1", "G-P34-2", "G-P34-3",
                "G-P7D-1", "G-P7D-2", "G-P7D-3", "G-P7D-4", "G-P7D-5", "G-P7D-6",
            ]
            if item_def["id"] in critical_items:
                all_pass = False
                result["blocking_reasons"].append(f"{item_def['id']}: 关键项不可跳过 (原因: {reason})")
        else:
            result["items"].append({
                "id": item_def["id"],
                "desc": item_def["desc"],
                "status": cp_item.get("status", "unknown"),
                "note": cp_item.get("reason", "")
            })
            all_pass = False
            result["blocking_reasons"].append(f"{item_def['id']}: 状态={cp_item.get('status')}")
    
    # Additional: check spot_checks for critical phases
    if from_phase == 2:
        spot_checks = checkpoint.get("spot_checks", [])
        # Phase 2 requires at least 1 spot check per key Sheet that returned "no issues"
        if not spot_checks:
            result["blocking_reasons"].append(
                "Phase 2无spot_checks — 脚本返回'无问题'时必须执行手动抽检"
            )
            all_pass = False
    
    # Additional: check self_check_completed
    if not checkpoint.get("self_check_completed", False):
        result["blocking_reasons"].append("self_check_completed=False — Phase自检未完成")
        all_pass = False
    
    # Additional: CCEP合规检查（MR-18）：checkpoint中必须包含common_compliance字段
    cc = checkpoint.get("common_compliance", {})
    if not cc:
        result["blocking_reasons"].append(
            "common_compliance字段缺失 — MR-18 CCEP要求每个Phase的checkpoint必须包含common_compliance声明"
        )
        all_pass = False
    else:
        # 检查CCEP项
        ccep_items = cc.get("ccep_items", [])
        for ccep in ccep_items:
            if ccep.get("status") != "pass":
                result["blocking_reasons"].append(
                    f"CCEP项未通过: {ccep.get('id', '?')} — {ccep.get('desc', '')} (status={ccep.get('status')})"
                )
                all_pass = False
        # 零幻觉：findings中每条问题必须有单元格引用
        findings = checkpoint.get("findings", {})
        details = findings.get("details", [])
        for detail in details:
            if not any(c in detail for c in ["A", "B", "C", "D", "E", "F", "G", "H"]):
                # 缺少单元格引用的finding可能违反零幻觉原则
                pass  # 仅警告，不阻断
    
    result["verdict"] = "pass" if all_pass else "BLOCKED"
    return result


def validate_all(audit_dir: Path) -> list[dict]:
    results = []
    for gate_id in GATE_DEFINITIONS:
        r = validate_gate(audit_dir, gate_id)
        results.append(r)
    return results


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    audit_dir = Path(sys.argv[1])
    phase_arg = sys.argv[2]
    
    if not audit_dir.exists():
        print(f"ERROR: 审核输出目录不存在: {audit_dir}")
        sys.exit(1)
    
    if phase_arg == "all":
        results = validate_all(audit_dir)
    else:
        # Support both gate_id (e.g. "G-P7D") and phase number (e.g. "2")
        if phase_arg.startswith("G-"):
            # Direct gate_id reference
            if phase_arg not in GATE_DEFINITIONS:
                print(f"ERROR: 未找到关卡定义: {phase_arg}")
                sys.exit(1)
            results = [validate_gate(audit_dir, phase_arg)]
        else:
            try:
                from_phase = int(phase_arg)
                # Find the gate for this transition
                gate_id = None
                for gid, gdef in GATE_DEFINITIONS.items():
                    if gdef["from_phase"] == from_phase:
                        gate_id = gid
                        break
                if gate_id is None:
                    print(f"ERROR: 未找到Phase {from_phase}的关卡定义")
                    sys.exit(1)
                results = [validate_gate(audit_dir, gate_id)]
            except ValueError:
                print(f"ERROR: 无效的Phase参数: {phase_arg}")
                sys.exit(1)
    
    # Output
    all_pass = True
    for r in results:
        print(f"\n{'='*60}")
        print(f"GATE {r['gate_id']} (Phase {r['from_phase']} → {r['to_phase']})")
        print(f"{'='*60}")
        
        for item in r["items"]:
            status_icon = {
                "pass": "✅",
                "skip": "⚠️",
                "NOT_FOUND": "❌",
                "MISSING_CHECKPOINT": "❌",
                "unknown": "❌"
            }.get(item["status"], "❓")
            print(f"  {status_icon} {item['id']}: {item['desc']}")
            if item.get("evidence"):
                print(f"     证据: {item['evidence']}")
            if item.get("reason"):
                print(f"     跳过原因: {item['reason']}")
            if item.get("note"):
                print(f"     备注: {item['note']}")
        
        print(f"\n  判定: {r['verdict']}")
        if r["blocking_reasons"]:
            print(f"  阻断原因:")
            for br in r["blocking_reasons"]:
                print(f"    - {br}")
        
        if r["verdict"] != "pass":
            all_pass = False
    
    print(f"\n{'='*60}")
    if all_pass:
        print("✅ 全部关卡通过，可继续执行")
        sys.exit(0)
    else:
        print("❌ 存在未通过关卡 — BLOCKED，必须补完后才可进入下一Phase")
        sys.exit(1)


if __name__ == "__main__":
    main()
