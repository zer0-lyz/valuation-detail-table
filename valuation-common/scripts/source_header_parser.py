# -*- coding: utf-8 -*-
"""
source_header_parser.py — 统一表头识别引擎 (DT-154)

解决问题：
  估值Skill体系中，不同公司的财务文件（科目余额表、资产负债表、序时账、
  辅助明细账等）格式千差万别，表头列名、列数、列序完全不同。

  之前：3套各自实现的表头识别逻辑
    - pdf_extract.py._find_col()  → PDF端关键词子串匹配
    - dt_runner.py._parse_xxx()   → Excel端精确匹配+fallback
    - vouching_extract/           → Pandas header=0

  现在：1个统一引擎
    - 统一关键词词典 SOURCE_HEADER_KEYWORDS（7种数据源类型）
    - 统一匹配引擎 find_col_by_keywords()（升级版_find_col）
    - 数据源类型自动检测 detect_source_type()
    - 7个parse_xxx()统一解析函数 + auto_parse()统一入口
    - 标准化输出格式

设计原则：
  - 关键词子串匹配（非精确匹配），适配不同系统导出格式
  - 多关键词候选列表，按优先级排序
  - 支持Excel(.xlsx/.xls)和PDF两种输入
  - 输出标准化字段映射，下游脚本无需关心原始列名

覆盖数据源类型：
  1. subject_balance  — 科目余额表
  2. balance_sheet   — 资产负债表
  3. journal         — 序时账/凭证一览表
  4. auxiliary       — 辅助明细账/辅助余额表
  5. bank_statement  — 银行对账单
  6. asset_register  — 固定资产卡片台账/收发存明细表
  7. income_statement— 利润表/利润及利润分配表

v1.0 (2026-05-24): 初始版本
  - 合并pdf_extract.py._find_col()和dt_runner.py中的所有关键词
  - 7种数据源parse函数
  - auto_parse()统一入口
  - detect_source_type()自动检测
"""

import os
import sys
import re
from typing import Optional, List, Dict, Any, Tuple


# ============================================================
# 统一关键词词典
# ============================================================

SOURCE_HEADER_KEYWORDS = {
    # ----------------------------------------------------------
    # 1. 科目余额表 (subject_balance)
    # ----------------------------------------------------------
    'subject_balance': {
        # 检测标识：表头中必须包含的标志性列名（用于定位表头行）
# 注意：'发生额及余额表'格式使用'编码'/'名称'/'期初余额'/'本期发生'/'期末余额'
        '_detect': ['科目编码', '科目代码', '科目号', '科目编号', '编码'],
        # 字段映射：标准字段名 → 关键词候选列表（按优先级）
        # 注意：双行表头合并后会产生如"期初余额借方""期末余额贷方"等组合列名
        'code':            ['科目编码', '科目代码', '科目号', '科目编号', '编码'],
        'name':            ['科目名称', '名称'],
        'direction':       ['余额方向', '方向', '借贷方向'],
        'beginning_debit': ['期初余额借方', '年初借方', '期初借方余额', '期初借方发生额', '期初借方', '期初借方金额', '期初余额金额', '余额金额'],
        'beginning_credit':['期初余额贷方', '年初贷方', '期初贷方余额', '期初贷方发生额', '期初贷方', '期初贷方金额', '期初余额金额', '余额金额'],
        'current_debit':   ['本期发生借方', '本期借方', '本期借方发生额', '借方发生', '本期借方发生'],
        'current_credit':  ['本期发生贷方', '本期贷方', '本期贷方发生额', '贷方发生', '本期贷方发生'],
        'ending_debit':    ['期末余额借方', '期末借方', '期末借方余额', '借方余额', '期末借方发生额'],
        'ending_credit':   ['期末余额贷方', '期末贷方', '期末贷方余额', '贷方余额', '期末贷方发生额'],
        'balance':         ['期末余额', '年末余额'],  # 仅精确匹配单列余额；双行表头下ending_debit/credit优先
        'level':           ['级次', '科目级次', '级别'],
    },

    # ----------------------------------------------------------
    # 2. 资产负债表 (balance_sheet)
    # ----------------------------------------------------------
    'balance_sheet': {
        '_detect': ['资产', '负债', '所有者权益', '资产负债表', '项目', '行次'],
        # DT-212: 增加"项目""行次"——8列BS中R3="项 目"+C2="行次"是表头标志行
        # 左栏（资产侧）
        'asset_label':     ['项目', '资产', '项目名称'],
        'asset_line_no':   ['行次', '行号'],
        'asset_beginning': ['年初余额', '年初数', '期初余额', '年初', '期初'],
        'asset_ending':    ['期末余额', '期末数', '年末余额', '期末', '年末'],
        # 右栏（负债及权益侧）—— 一般与左栏同列名但偏移4列
        'liab_label':      ['项目', '负债', '项目名称'],
        'liab_line_no':    ['行次', '行号'],
        'liab_beginning':  ['年初余额', '年初数', '期初余额', '年初', '期初'],
        'liab_ending':     ['期末余额', '期末数', '年末余额', '期末', '年末'],
    },

    # ----------------------------------------------------------
    # 3. 序时账/凭证一览表 (journal)
    # ----------------------------------------------------------
    'journal': {
        '_detect': ['凭证', '摘要', '借贷', '序时账', '明细账', '日记账'],
        'date':            ['日期', '记账日期', '凭证日期', '发生日期', '业务日期'],
        'voucher_no':      ['凭证号', '凭证编号', '凭证字号', '凭证字', '凭证'],
        'summary':         ['摘要', '内容', '业务摘要', '说明'],
        'code':            ['科目编码', '科目代码', '科目号', '科目编号', '编码'],
        'name':            ['科目名称', '科目', '名称'],
        'debit':           ['借方', '借方金额', '借方发生额', '借'],
        'credit':          ['贷方', '贷方金额', '贷方发生额', '贷'],
        'balance':         ['余额', '余额方向', '当前余额'],
        'direction':       ['方向', '借贷方向', '借/贷'],
        'counterpart':     ['对方科目', '对应科目', '对方', '对应'],
        'aux_accounting':  ['辅助核算', '辅助', '核算项目', '辅助项'],
        'settlement':      ['结算对象', '结算', '客商', '往来单位', '对方单位'],
        'quantity':        ['数量', '数量单位'],
        'unit':            ['单位', '计量单位'],
    },

    # ----------------------------------------------------------
    # 4. 辅助明细账/辅助余额表 (auxiliary)
    # ----------------------------------------------------------
    'auxiliary': {
        '_detect': ['辅助', '辅助余额', '辅助明细', '科目辅助', '辅助核算',
                    '供应商', '客户', '结算对象', '往来单位'],
        'code':            ['科目编码', '科目代码', '编码', '编号'],
        'name':            ['科目名称', '科目', '名称'],
        'settlement':      ['供应商档案名称', '结算对象', '对象名称', '客户名称', '供应商名称', '户名',
                           '往来单位', '客商', '对方名称', '结算', '辅助名称'],
        'debit':           ['期末余额借方', '期末借方', '借方余额', '期末借方余额',
                           '本期借方本币', '本期借方', '借方发生额'],
        'credit':          ['期末余额贷方', '期末贷方', '贷方余额', '期末贷方余额',
                           '本期贷方本币', '本期贷方', '贷方发生额'],
        'balance':         ['期末余额本币', '期末余额', '余额', '年末余额'],
        'direction':       ['方向', '余额方向', '借贷方向'],
        'quantity':        ['数量'],
        'unit':            ['单位'],
        'business':        ['摘要', '业务内容', '款项内容', '款项', '内容', '备注', '说明'],
        'beginning_debit': ['期初余额借方', '期初借方', '期初借方余额'],
        'beginning_credit':['期初余额贷方', '期初贷方', '期初贷方余额'],
    },

    # ----------------------------------------------------------
    # 5. 银行对账单 (bank_statement)
    # ----------------------------------------------------------
    'bank_statement': {
        '_detect': ['银行', '账号', '对账单', '余额', 'Bank'],
        'date':            ['日期', '交易日期', '记账日期', '发生日期'],
        'summary':        ['摘要', '说明', '交易摘要', '备注'],
        'debit':          ['借方', '支出', '借方金额', '借方发生额', '付出'],
        'credit':         ['贷方', '收入', '贷方金额', '贷方发生额', '存入'],
        'balance':        ['余额', '账户余额', '当前余额', '结余'],
        'counterpart':    ['对方', '对方户名', '对方账号', '交易对方'],
        'voucher_no':     ['凭证号', '凭证编号', '流水号'],
    },

    # ----------------------------------------------------------
    # 6. 固定资产卡片台账/收发存明细表 (asset_register)
    # ----------------------------------------------------------
    'asset_register': {
        '_detect': ['资产', '设备', '原值', '净值', '卡片', '台账', '收发存'],
        'name':            ['资产名称', '设备名称', '项目名称', '名称', '品名'],
        'category':        ['类别', '分类', '资产类别', '设备类别', '设备类型'],
        'original_value':  ['原值', '账面原值', '入账价值', '购置价值', '期初原值'],
        'net_value':       ['净值', '账面净值', '净额', '期末净值'],
        'accum_depreciation': ['累计折旧', '折旧', '已提折旧', '累计折旧额'],
        'quantity':        ['数量', '台数', '台/套', '数量(台)', '台'],
        'depreciation':    ['本期折旧', '本期折旧额', '月折旧额'],
        'residual_rate':  ['残值率', '净残值率'],
        'useful_life':    ['使用年限', '年限', '预计使用年限', '折旧年限'],
        'purchase_date':  ['入账日期', '购置日期', '开始使用日期', '启用日期'],
        'location':       ['存放地点', '使用部门', '部门', '位置'],
    },

    # ----------------------------------------------------------
    # 7. 利润表 (income_statement)
    # ----------------------------------------------------------
    'income_statement': {
        '_detect': ['利润', '收入', '费用', '利润表', '损益'],
        'label':           ['项目', '项目名称', '科目'],
        'line_no':         ['行次', '行号'],
        'current_period':  ['本期金额', '本期', '本月金额', '本月数', '本期数'],
        'year_accumulated': ['本年累计金额', '本年累计', '年初至本期', '累计金额', '上期金额'],
        'last_year':       ['上年金额', '上年', '去年同期', '上年同期'],
    },
}


# ============================================================
# 统一匹配引擎
# ============================================================

def find_col_by_keywords(headers: List[str], keywords: List[str],
                         case_sensitive: bool = False,
                         prefer_exact: bool = True) -> Optional[int]:
    """在表头列表中查找包含关键词的列索引（升级版_find_col）。

    与原_find_col的区别：
    1. 去除全角/半角空格+全角数字归一化后再匹配
    2. 支持大小写敏感开关
    3. 默认优先精确匹配（prefer_exact=True时，先找完全相同的列名，再找子串包含）
    4. 返回0-based列索引

    Args:
        headers: 表头文字列表（可以是0-based索引对应的列表）
        keywords: 关键词候选列表（按优先级排序，匹配第一个命中的）
        case_sensitive: 是否区分大小写
        prefer_exact: 优先精确匹配（先找完全相同的列名，再找子串包含）

    Returns:
        int: 0-based列索引，未找到返回None
    """
    if prefer_exact:
        # 第一轮：精确匹配
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_clean = str(h).replace(' ', '').replace('\u3000', '').strip()
            h_clean = _normalize_fullwidth(h_clean)
            if not case_sensitive:
                h_clean = h_clean.lower()
            for kw in keywords:
                kw_clean = kw.replace(' ', '').strip()
                if not case_sensitive:
                    kw_clean = kw_clean.lower()
                if h_clean == kw_clean:
                    return i

    # 第二轮：子串包含匹配
    for i, h in enumerate(headers):
        if h is None:
            continue
        # 归一化：去空格、全角→半角
        h_clean = str(h).replace(' ', '').replace('\u3000', '').strip()
        # 全角数字→半角
        h_clean = _normalize_fullwidth(h_clean)
        if not case_sensitive:
            h_clean = h_clean.lower()
        for kw in keywords:
            kw_clean = kw.replace(' ', '').strip()
            if not case_sensitive:
                kw_clean = kw_clean.lower()
            if kw_clean in h_clean:
                return i
    return None


def _normalize_fullwidth(s: str) -> str:
    """全角字符→半角（数字+字母+括号等）"""
    result = []
    for ch in s:
        code = ord(ch)
        # 全角数字/字母: 0xFF01~0xFF5E → 0x21~0x7E
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        # 全角空格
        elif code == 0x3000:
            result.append(' ')
        else:
            result.append(ch)
    return ''.join(result)


def find_all_cols(headers: List[str], source_type: str) -> Dict[str, int]:
    """一次性查找某数据源类型的所有标准字段映射。

    Args:
        headers: 表头文字列表
        source_type: 数据源类型（SOURCE_HEADER_KEYWORDS的key）

    Returns:
        dict: {标准字段名: 0-based列索引}
    """
    if source_type not in SOURCE_HEADER_KEYWORDS:
        return {}

    field_map = {}
    keywords_dict = SOURCE_HEADER_KEYWORDS[source_type]

    for field_name, kw_list in keywords_dict.items():
        if field_name.startswith('_'):
            continue  # 跳过_detect等元信息
        col_idx = find_col_by_keywords(headers, kw_list)
        if col_idx is not None:
            field_map[field_name] = col_idx

    return field_map


# ============================================================
# 数据源类型自动检测
# ============================================================

def detect_source_type(headers: List[str], filename: str = '') -> Optional[str]:
    """根据表头内容和文件名自动检测数据源类型。

    Args:
        headers: 表头文字列表
        filename: 文件名（辅助判断）

    Returns:
        str: 数据源类型key，未识别返回None
    """
    filename_clean = os.path.basename(filename) if filename else ''

    # 1. 先用文件名快速判断
    fname_hints = {
        'subject_balance': ['科目余额', '余额表'],
        'balance_sheet':   ['资产负债', '财务报表'],
        'journal':         ['序时账', '凭证', '日记账', '明细账'],
        'auxiliary':       ['辅助', '辅助余额', '辅助明细'],
        'bank_statement':  ['对账单', '银行'],
        'asset_register':  ['卡片', '台账', '收发存', '固定资产'],
        'income_statement':['利润表', '损益表', '利润及利润分配'],
    }

    for stype, hints in fname_hints.items():
        for hint in hints:
            if hint in filename_clean:
                # 文件名命中，再用_detect验证表头
                detect_kw = SOURCE_HEADER_KEYWORDS[stype].get('_detect', [])
                for kw in detect_kw:
                    if find_col_by_keywords(headers, [kw]) is not None:
                        return stype
                # 文件名强匹配但表头无_detect关键词，仍返回（可能表头格式特殊）
                return stype

    # 2. 用_detect关键词检测
    best_match = None
    best_score = 0

    for stype, kw_dict in SOURCE_HEADER_KEYWORDS.items():
        detect_kw = kw_dict.get('_detect', [])
        score = 0
        for kw in detect_kw:
            if find_col_by_keywords(headers, [kw]) is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_match = stype

    return best_match if best_score > 0 else None


# ============================================================
# 表头行定位
# ============================================================

def locate_header_row(ws, source_type: str = None,
                      max_scan_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    """在Excel Worksheet中定位表头行并提取列映射。

    支持双行表头（如科目余额表：Row4=主表头"期末余额", Row5=子表头"借方/贷方"）
    自动合并双行表头信息。

    Args:
        ws: openpyxl Worksheet对象
        source_type: 数据源类型（None则自动检测）
        max_scan_rows: 最大扫描行数

    Returns:
        (header_row, col_map): 表头行号(1-based), {标准字段名: 列号(1-based)}
        未找到返回 (0, {})
    """
    for row in range(1, min(ws.max_row + 1, max_scan_rows + 1)):
        # 收集本行所有单元格值
        row_values = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            row_values.append(str(val).strip() if val else '')

        # 检测类型（如果未指定）
        detected_type = source_type
        if detected_type is None:
            # 传入filename辅助检测（当表头不含典型关键词时，日期可无此参数）
            _fname = getattr(ws, 'title', '') if hasattr(ws, 'title') else ''
            detected_type = detect_source_type(row_values, filename=_fname)
            if detected_type is None:
                continue

        # 用_detect关键词验证表头行
        detect_kw = SOURCE_HEADER_KEYWORDS.get(detected_type, {}).get('_detect', [])
        header_confirmed = False
        for kw in detect_kw:
            if find_col_by_keywords(row_values, [kw]) is not None:
                header_confirmed = True
                break

        if not header_confirmed:
            continue

        # 表头行验证：必须有>=3个非空列（排除标题行如"辅助余额表"只有1个单元格）
        non_empty_count = sum(1 for v in row_values if v.strip())
        if non_empty_count < 3:
            continue

        # ════════════════════════════════════════════════════
        # 双行表头检测（两种模式）
        # ════════════════════════════════════════════════════
        # 子表头关键词
        _sub_kw = ['借方', '贷方', '借', '贷', '数量', '单价', '金额']

        # 模式A：当前行本身包含子表头关键词 → 当前行是子表头行
        # （常见于'发生额及余额表'格式：Row5=主表头［期初余额/本期发生/期末余额］,
        #  Row6=子表头［编码/名称/借方/贷方］，扫描到Row6命中'编码'但Row6本身有'借方''贷方'）
        _current_has_sub = any(
            find_col_by_keywords(row_values, [kw]) is not None
            for kw in ['借方', '贷方', '借', '贷']
        )
        if _current_has_sub and row > 1:
            # 获取上一行作为主表头
            prev_row_values = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row - 1, col).value
                prev_row_values.append(str(val).strip() if val else '')

            _merged = _merge_double_header(prev_row_values, row_values, row, ws)
            col_map = {}
            keywords_dict = SOURCE_HEADER_KEYWORDS.get(detected_type, {})
            for field_name, kw_list in keywords_dict.items():
                if field_name.startswith('_'):
                    continue
                col_idx = find_col_by_keywords(_merged, kw_list)
                if col_idx is not None:
                    col_map[field_name] = col_idx + 1  # 转1-based
            # 后处理
            if 'ending_debit' in col_map and 'ending_credit' in col_map:
                col_map.pop('balance', None)
            if 'beginning_debit' in col_map and 'beginning_credit' in col_map:
                col_map.pop('beginning_balance', None)
            return row, col_map

        # 模式B：下一行包含子表头关键词 → 当前行是主表头行，下一行是子表头行
        next_row_values = []
        if row + 1 <= ws.max_row:
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row + 1, col).value
                next_row_values.append(str(val).strip() if val else '')

        is_double_header = any(
            find_col_by_keywords(next_row_values, [kw]) is not None
            for kw in _sub_kw
        )

        if is_double_header:
            # 合并双行表头：主表头+子表头 → 组合列名
            merged_values = _merge_double_header(row_values, next_row_values, row, ws)

            # DT-FIX: 三层表头（喜发格式：Row7主+Row8子+Row9数量/金额）
            _row2 = row + 2
            if _row2 <= ws.max_row:
                _row9_vals = []
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(_row2, col).value
                    _row9_vals.append(str(val).strip() if val else '')
                _has_triple = any(
                    find_col_by_keywords(_row9_vals, [kw]) is not None
                    for kw in ['数量', '金额']
                )
                if _has_triple:
                    merged_values = _merge_double_header(merged_values, _row9_vals, _row2, ws)

            col_map = {}
            keywords_dict = SOURCE_HEADER_KEYWORDS.get(detected_type, {})
            for field_name, kw_list in keywords_dict.items():
                if field_name.startswith('_'):
                    continue
                col_idx = find_col_by_keywords(merged_values, kw_list)
                if col_idx is not None:
                    col_map[field_name] = col_idx + 1  # 转1-based
            # 金额列优先：对金额相关字段，如果匹配到'数量'列但旁边有'金额'列，切换到金额列
            _finance_fields = ['current_debit', 'current_credit', 'beginning_debit', 'beginning_credit', 'balance']
            for _ff in _finance_fields:
                _fc = col_map.get(_ff)
                if _fc and _fc < len(merged_values):
                    _header_text = merged_values[_fc - 1] if _fc > 0 else ''
                    if '数量' in _header_text and _fc < len(merged_values):
                        _next_text = merged_values[_fc] if _fc < len(merged_values) else ''
                        if '金额' in _next_text:
                            col_map[_ff] = _fc + 1
            # 后处理
            if 'ending_debit' in col_map and 'ending_credit' in col_map:
                col_map.pop('balance', None)
            if 'beginning_debit' in col_map and 'beginning_credit' in col_map:
                col_map.pop('beginning_balance', None)
            return row, col_map
        else:
            # 单行表头
            col_map = {}
            keywords_dict = SOURCE_HEADER_KEYWORDS.get(detected_type, {})
            for field_name, kw_list in keywords_dict.items():
                if field_name.startswith('_'):
                    continue
                col_idx = find_col_by_keywords(row_values, kw_list)
                if col_idx is not None:
                    col_map[field_name] = col_idx + 1  # 转1-based
            return row, col_map

    return 0, {}


def _merge_double_header(main_headers: List[str], sub_headers: List[str],
                         main_row: int, ws) -> List[str]:
    """合并双行表头：主表头+子表头 → 组合列名。

    例如：
      主表头: ['科目编码', '科目名称', '期初余额', '', '本期发生', '', ...]
      子表头: ['', '', '借方', '贷方', '借方', '贷方', ...]
      合并后: ['科目编码', '科目名称', '期初余额借方', '期初余额贷方', '本期发生借方', '本期发生贷方', ...]

    原则：
    - 如果子表头为空，保留主表头
    - 如果主表头为空但子表头有值，用子表头
    - 如果都有值，合并为主表头+子表头
    """
    merged = []
    current_main = ''  # 当前主表头值（用于填充子表头）

    for i in range(max(len(main_headers), len(sub_headers))):
        main_val = main_headers[i].strip() if i < len(main_headers) else ''
        sub_val = sub_headers[i].strip() if i < len(sub_headers) else ''

        # 更新当前主表头
        if main_val:
            current_main = main_val

        if main_val and sub_val:
            # 两行都有值：合并
            merged.append(f'{main_val}{sub_val}')
        elif main_val and not sub_val:
            # 仅主表头有值
            merged.append(main_val)
        elif not main_val and sub_val:
            # 仅子表头有值：继承上方主表头
            merged.append(f'{current_main}{sub_val}')
        else:
            merged.append('')

    return merged


# ============================================================
# 通用解析函数
# ============================================================

def _safe_float(val) -> float:
    """安全转浮点数"""
    if val is None or val == '':
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(',', '').replace('，', '').replace('￥', '').replace('¥', '')
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_subject_balance(filepath: str) -> Dict[str, Any]:
    """解析科目余额表 → 标准化输出

    Args:
        filepath: Excel文件路径

    Returns:
        dict: {
            'source_type': 'subject_balance',
            'filepath': str,
            'status': 'parsed' | 'failed',
            'header_row': int,
            'col_map': dict,
            'subjects': [{
                'code': str, 'name': str, 'balance': float,
                'direction': str, 'level': int,
                'beginning_debit': float, 'beginning_credit': float,
                'current_debit': float, 'current_credit': float,
                'ending_debit': float, 'ending_credit': float,
            }],
            'warnings': list,
        }
    """
    import openpyxl

    result = {
        'source_type': 'subject_balance',
        'filepath': filepath,
        'status': 'failed',
        'header_row': 0,
        'col_map': {},
        'subjects': [],
        'warnings': [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # 定位表头行+列映射
        header_row, col_map = locate_header_row(ws, 'subject_balance')
        if header_row == 0:
            result['warnings'].append('无法识别科目余额表表头行')
            wb.close()
            return result

        result['header_row'] = header_row
        result['col_map'] = col_map

        code_col = col_map.get('code', 1)
        name_col = col_map.get('name', 2)
        balance_col = col_map.get('balance')
        direction_col = col_map.get('direction')
        ending_debit_col = col_map.get('ending_debit')
        ending_credit_col = col_map.get('ending_credit')
        beginning_debit_col = col_map.get('beginning_debit')
        beginning_credit_col = col_map.get('beginning_credit')
        current_debit_col = col_map.get('current_debit')
        current_credit_col = col_map.get('current_credit')
        level_col = col_map.get('level')

        for row in range(header_row + 1, ws.max_row + 1):
            code = ws.cell(row, code_col).value
            name = ws.cell(row, name_col).value
            if not code or not name:
                continue
            code = str(code).strip()
            name = str(name).strip()
            if not code or not name:
                continue
            # 跳过合计行
            if '合计' in name or '小计' in name:
                continue

            # 余额逻辑：优先期末余额单列，其次借方/贷方分列
            balance = 0.0
            direction = ''

            _balance_src_col = balance_col
            # DT-FIX: 阳江晶步格式——期末余额/原币(Col17,空)+/本币(Col18,有值)
            # 如果balance_col指向原币列且下一列是"本币"，自动切换
            if balance_col and balance_col < ws.max_column:
                _next_hdr = str(ws.cell(header_row, balance_col + 1).value or ws.cell(header_row + 1, balance_col + 1).value or '').strip()
                _cur_sub = str(ws.cell(header_row + 1, balance_col).value or '').strip()
                if _cur_sub == '原币' and ('本币' in _next_hdr or '本币' in _next_hdr):
                    _balance_src_col = balance_col + 1

            if _balance_src_col:
                balance = _safe_float(ws.cell(row, _balance_src_col).value)
            elif ending_debit_col and ending_credit_col:
                d = _safe_float(ws.cell(row, ending_debit_col).value)
                c = _safe_float(ws.cell(row, ending_credit_col).value)
                if d and not c:
                    balance = d
                    direction = '借'
                elif c and not d:
                    balance = c
                    direction = '贷'
                elif d and c:
                    balance = d - c
                    direction = '借' if d >= c else '贷'

            # 余额方向列
            if direction_col:
                d = ws.cell(row, direction_col).value
                if d:
                    direction = str(d).strip()

            # DT-FIX: 阳江晶步格式——原币/本币子表头，自动切换到本币列
            _cur_debit_col = current_debit_col
            _cur_credit_col = current_credit_col
            if current_debit_col and current_debit_col < ws.max_column:
                if str(ws.cell(header_row + 1, current_debit_col).value or '').strip() == '原币':
                    _next = str(ws.cell(header_row + 1, current_debit_col + 1).value or '').strip()
                    if _next == '本币':
                        _cur_debit_col = current_debit_col + 1
            if current_credit_col and current_credit_col < ws.max_column:
                if str(ws.cell(header_row + 1, current_credit_col).value or '').strip() == '原币':
                    _next = str(ws.cell(header_row + 1, current_credit_col + 1).value or '').strip()
                    if _next == '本币':
                        _cur_credit_col = current_credit_col + 1

            # DT-FIX: 期初借贷列冲突（喜发等：beginning_debit=beginning_credit→同一列）
            _bd_col = beginning_debit_col
            _bc_col = beginning_credit_col
            if _bd_col and _bc_col and _bd_col == _bc_col:
                _same_begin_val = _safe_float(ws.cell(row, _bd_col).value)
                if direction == '借':
                    _bd_col, _bc_col = _bd_col, None
                elif direction == '贷':
                    _bd_col, _bc_col = None, _bc_col
                else:
                    _bd_col, _bc_col = None, None

            _ed = _safe_float(ws.cell(row, ending_debit_col).value) if ending_debit_col else 0.0
            _ec = _safe_float(ws.cell(row, ending_credit_col).value) if ending_credit_col else 0.0
            # DT-FIX: 无期末借贷列但有方向+余额→按方向拆分
            if not ending_debit_col and not ending_credit_col:
                if balance != 0:
                    if direction == '借':
                        _ed = balance
                    elif direction == '贷':
                        _ec = balance
                    else:
                        _ed = balance
                elif _cur_debit_col or _cur_credit_col or _bd_col:
                    # DT-FIX: 从期初+本期发生计算（喜发等无期末余额列格式）
                    _beg = _safe_float(ws.cell(row, _bd_col).value) if _bd_col else 0.0
                    _cd = _safe_float(ws.cell(row, _cur_debit_col).value) if _cur_debit_col else 0.0
                    _cc = _safe_float(ws.cell(row, _cur_credit_col).value) if _cur_credit_col else 0.0
                    if _beg or _cd or _cc:
                        _net = _beg + _cd - _cc
                        if direction == '借':
                            _ed = max(_net, 0)
                        elif direction == '贷':
                            _ec = max(-_net, 0)
                        else:
                            _ed = max(_net, 0) if _net >= 0 else 0
                            _ec = max(-_net, 0) if _net < 0 else 0

            # DT-FIX: 喜发格式——balance=0但ed/ec有值时，balance取ed或ec的值
            if balance == 0 and (_ed or _ec):
                balance = _ed if _ed else _ec
            subjects_item = {
                'code': code,
                'name': name,
                'balance': balance,
                'direction': direction,
                'level': int(len(code) // 2) if len(code) <= 8 else 1,
                'beginning_debit': _safe_float(ws.cell(row, _bd_col).value) if _bd_col else 0.0,
                'beginning_credit': _safe_float(ws.cell(row, _bc_col).value) if _bc_col else 0.0,
                'current_debit': _safe_float(ws.cell(row, _cur_debit_col).value) if _cur_debit_col else 0.0,
                'current_credit': _safe_float(ws.cell(row, _cur_credit_col).value) if _cur_credit_col else 0.0,
                'ending_debit': _ed,
                'ending_credit': _ec,
            }

            if level_col:
                lv = ws.cell(row, level_col).value
                if lv:
                    subjects_item['level'] = int(lv)

            result['subjects'].append(subjects_item)

        result['status'] = 'parsed'
        wb.close()

    except Exception as e:
        result['warnings'].append(f'解析异常: {str(e)}')

    return result


def parse_balance_sheet(filepath: str) -> Dict[str, Any]:
    """解析资产负债表 → 标准化输出

    适配两种常见格式：
    1. 左右分栏（资产1-4列 + 负债及权益5-8列）
    2. 上下排列

    Args:
        filepath: Excel文件路径

    Returns:
        dict: {
            'source_type': 'balance_sheet',
            'filepath': str,
            'status': 'parsed' | 'failed',
            'format': 'side_by_side' | 'stacked',
            'header_row': int,
            'col_map': dict,
            'items': [{
                'label': str, 'side': '资产'|'负债及权益',
                'beginning_balance': float, 'ending_balance': float,
            }],
            'total_assets': float,
            'total_liab_equity': float,
            'warnings': list,
        }
    """
    import openpyxl

    result = {
        'source_type': 'balance_sheet',
        'filepath': filepath,
        'status': 'failed',
        'format': 'unknown',
        'header_row': 0,
        'col_map': {},
        'items': [],
        'total_assets': 0.0,
        'total_liab_equity': 0.0,
        'warnings': [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # 查找资产负债表Sheet
        ws = None
        for name in wb.sheetnames:
            if '资产负债表' in name:
                ws = wb[name]
                break
        if ws is None:
            # DT-FIX: 按内容搜索——扫描所有sheet查找含"资产负债表"文本的sheet
            for name in wb.sheetnames:
                _tws = wb[name]
                for r in range(1, min(_tws.max_row + 1, 5)):
                    for c in range(1, min(_tws.max_column + 1, 10)):
                        _v = str(_tws.cell(row=r, column=c).value or '')
                        if '资产负债表' in _v or '会企01' in _v:
                            ws = _tws
                            break
                    if ws:
                        break
                if ws:
                    break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # ── P1补充+DT-210: 提取表头元信息（编制单位/公司全称/评估基准日） ──
        # 资产负债表表头常见格式：
        #   格式A: Row1="资产负债表", Row2="编制单位：上海图灵量子科技有限公司" + "2026年04月30日" + "单位：元"
        #   格式B: Row1="资产负债表"(D列), Row2 A列="河南平煤神马平绿置业有限公司" + D列="2025年12月" + H列="单位：元"
        #   格式C: Row1="资产负债表", Row2=公司名, Row3="编制单位：XXX"
        company_full_name = ''
        valuation_date_from_bs = ''

        # DT-213: 先找"资产负债表"标题行位置，确定标题下方哪些是公司名候选行
        # 扫描范围：前15行全列宽（原硬编码7行8列，某些BS标题行偏下会遗漏）
        bs_title_row = 0
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 15)):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str) and '资产负债表' in val.strip():
                    bs_title_row = r
                    break
            if bs_title_row:
                break

        # DT-213: 公司名/日期搜索范围——以bs_title_row为基准向下3行
        _meta_scan_end = min(ws.max_row + 1, (bs_title_row + 4) if bs_title_row else 10)
        for r in range(1, _meta_scan_end):
            for c in range(1, min(ws.max_column + 1, 15)):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    val_stripped = val.strip()

                    # ── 提取公司全称 ──
                    # 策略A: 含"编制单位"关键词 → 提取冒号后内容
                    if '编制单位' in val_stripped:
                        name_part = re.sub(r'^编制单位[：:]\s*', '', val_stripped).strip()
                        if name_part:
                            company_full_name = name_part

                    # DT-FIX: 策略A2: "单位：公司名"格式（新华体育等，无"编制"前缀）
                    if not company_full_name and '单位' in val_stripped and '元' not in val_stripped:
                        name_part = re.sub(r'^单位[：:]\s*', '', val_stripped).strip()
                        # 只有含'公司'/'企业'/'厂'等实体词才确认是公司名
                        if name_part and any(kw in name_part for kw in ['公司', '企业', '厂', '集团', '有限']):
                            company_full_name = name_part

                    # DT-210 策略B: 无"编制单位"前缀的纯公司名
                    # 判定条件：位于"资产负债表"标题行下方1-2行、A列(c=1)、
                    # 不含"项目/行次/列/单位/金额"等表头关键词、长度≥4（排除"单位：元"等短文本）
                    if not company_full_name and bs_title_row and c == 1:
                        row_offset = r - bs_title_row
                        if 1 <= row_offset <= 2:
                            _header_keywords = ('项目', '行次', '列', '单位', '金额', '资产', '负债', '权益', '附注')
                            _is_header = any(kw in val_stripped for kw in _header_keywords)
                            if not _is_header and len(val_stripped) >= 4:
                                # 进一步排除纯数字/日期/单位行
                                if not re.match(r'^[\d年月日\s]+$', val_stripped):
                                    company_full_name = val_stripped

                    # ── 提取评估基准日 ──
                    if not valuation_date_from_bs:
                        # 优先匹配 YYYY-MM-DD 格式（新华体育等）
                        date_match_iso = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', val_stripped)
                        if date_match_iso:
                            y, m, d = date_match_iso.groups()
                            valuation_date_from_bs = f'{y}年{int(m)}月{int(d)}日'
                        else:
                            # 次优匹配"年月日"完整格式
                            date_match = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', val_stripped)
                            if date_match:
                                y, m, d = date_match.groups()
                                valuation_date_from_bs = f'{y}年{int(m)}月{int(d)}日'
                            else:
                                # DT-210: 降级匹配"年月"格式（无"日"，如"2025年12月"）
                                date_match_ym = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月(?!.*日)', val_stripped)
                                if date_match_ym:
                                    y, m = date_match_ym.groups()
                                    # 默认月末：用calendar取当月最后一天
                                    try:
                                        import calendar
                                        last_day = calendar.monthrange(int(y), int(m))[1]
                                        valuation_date_from_bs = f'{y}年{int(m)}月{last_day}日'
                                    except (ValueError, calendar.IllegalMonthError):
                                        valuation_date_from_bs = f'{y}年{int(m)}月31日'

        if company_full_name:
            result['company_full_name'] = company_full_name
        if valuation_date_from_bs:
            result['valuation_date'] = valuation_date_from_bs

        # 定位表头行
        header_row, col_map = locate_header_row(ws, 'balance_sheet')

        # DT-213: 如果自动检测失败，用全列扫描替代固定列1/5
        if header_row == 0:
            for row in range(1, min(ws.max_row + 1, 10)):
                for c in range(1, min(ws.max_column + 1, 10)):
                    val = ws.cell(row, c).value
                    if val and ('资产' in str(val) or '项目' in str(val)):
                        header_row = row
                        break
                if header_row:
                    break

        if header_row == 0:
            result['warnings'].append('无法识别资产负债表表头行')
            wb.close()
            return result

        result['header_row'] = header_row

        # DT-153v3+DT-212+DT-FIX: 判断格式——动态检测右栏是否有独立表头
        # 从第4列开始扫描（新华体育格式：Col4=负债和所有者权益，Col1-3为资产侧）
        # DT-212: 去除空格后匹配，因为BS表头常为"项            目"含大量空格
        right_header = None
        for c in range(4, min(ws.max_column + 1, 15)):
            val = ws.cell(header_row, c).value
            if val:
                val_compact = str(val).replace(' ', '').replace('\u3000', '')
                if '项目' in val_compact or '负债' in val_compact or '资产' in val_compact or '所有者' in val_compact:
                    right_header = val
                    break
        is_side_by_side = right_header is not None and str(right_header).strip() != ''

        result['format'] = 'side_by_side' if is_side_by_side else 'stacked'

        # 数据起始行
        data_start = header_row + 1
        # 跳过列次行（含"行次""行号""列次"等）
        # DT-212: 增加"列次"检测——8列BS中R4="列            次"行需跳过
        # 注意：单元格值可能含大量空格（如"列            次"），需去除空格后匹配
        for r in range(data_start, min(data_start + 3, ws.max_row + 1)):
            v = ws.cell(r, 1).value
            if v:
                v_compact = str(v).replace(' ', '').replace('\u3000', '')
                if '行次' in v_compact or '行号' in v_compact or '列次' in v_compact:
                    data_start = r + 1
                    break

        if is_side_by_side:
            # 左右分栏格式
            # 用col_map或默认列位
            label_left_col = col_map.get('asset_label', 1)
            beginning_left_col = col_map.get('asset_beginning', 3)
            ending_left_col = col_map.get('asset_ending', 4)

            # DT-213: 右栏列位——跨项目自适应推断
            # 原逻辑：col_map值<5时硬编码兜底5/7/8，对6列/10列BS完全错位。
            # 新逻辑：
            # 1. 计算左栏各列的间距（beginning_col - label_col, ending_col - label_col）
            # 2. 推断右栏label_col：从表头行动态扫描右栏"项目"关键词
            # 3. 右栏beginning/ending = 右栏label + 左栏间距
            # 4. col_map值有效时（>= label_left_col + 间距）优先使用

            # Step 1: 计算左栏列间距
            _left_span_begin = beginning_left_col - label_left_col  # 通常2
            _left_span_end = ending_left_col - label_left_col       # 通常3

            # Step 2: 动态推断右栏label列
            # 策略A: 从表头行扫描右栏"项目"关键词（含紧凑字符串匹配）
            _right_label_candidate = None
            for c in range(ending_left_col + 1, min(ws.max_column + 1, 20)):
                val = ws.cell(header_row, c).value
                if val:
                    v_compact = str(val).replace(' ', '').replace('\u3000', '')
                    if '项目' in v_compact or '负债' in v_compact or '所有者' in v_compact:
                        _right_label_candidate = c
                        break

            # 策略B: 如果col_map的liab_label有效且>左栏最大列，使用它
            _col_map_liab_label = col_map.get('liab_label', 0)
            if _col_map_liab_label > ending_left_col:
                _right_label_candidate = _right_label_candidate or _col_map_liab_label

            # 策略C: 兜底——假设左右栏等宽，右栏起始=左栏最大列+1
            if _right_label_candidate is None:
                _right_label_candidate = ending_left_col + 1

            label_right_col = _right_label_candidate

            # Step 3: 用间距推断右栏beginning/ending列
            beginning_right_col = label_right_col + _left_span_begin
            ending_right_col = label_right_col + _left_span_end

            # Step 4: col_map值若>ending_left_col（有效映射），优先使用
            _col_map_liab_begin = col_map.get('liab_beginning', 0)
            _col_map_liab_end = col_map.get('liab_ending', 0)
            if _col_map_liab_begin > ending_left_col:
                beginning_right_col = _col_map_liab_begin
            if _col_map_liab_end > ending_left_col:
                ending_right_col = _col_map_liab_end

            # 动态扫描右栏列位（从右栏label列开始，避免扫到左栏列）
            _right_beginning = None
            _right_ending = None
            for c in range(max(5, label_right_col), min(ws.max_column + 1, 20)):
                val = ws.cell(header_row, c).value
                if val:
                    v = str(val).strip()
                    if v in ('年初余额', '年初数', '期初余额', '年初') and _right_beginning is None:
                        _right_beginning = c
                    elif v in ('期末余额', '期末数', '期末', '年末余额') and _right_ending is None:
                        _right_ending = c

            if _right_beginning and _right_ending:
                # 动态检测成功
                beginning_right_col = _right_beginning
                ending_right_col = _right_ending
            else:
                # 动态检测失败，用偏移量修正
                if beginning_right_col <= ending_left_col:
                    # 推断偏移量：右栏label_col - 左栏label_col
                    _offset = label_right_col - label_left_col
                    if _offset <= 0:
                        _offset = 4  # 标准偏移兜底
                    beginning_right_col = beginning_left_col + _offset
                    print(f'  [DT-153v3] 负债侧年初余额列推断: 左栏{beginning_left_col}+偏移{_offset}={beginning_right_col}')

                if ending_right_col <= ending_left_col:
                    _offset = label_right_col - label_left_col
                    if _offset <= 0:
                        _offset = 4
                    ending_right_col = ending_left_col + _offset
                    print(f'  [DT-153v3] 负债侧期末余额列推断: 左栏{ending_left_col}+偏移{_offset}={ending_right_col}')

            # DT-213: 验证右栏列合理性——右栏列必须大于左栏最大列
            if beginning_right_col <= ending_left_col:
                beginning_right_col = label_right_col + _left_span_begin
            if ending_right_col <= ending_left_col:
                ending_right_col = label_right_col + _left_span_end

            for row in range(data_start, ws.max_row + 1):
                # 左侧：资产
                label_left = ws.cell(row, label_left_col).value
                if label_left:
                    label_left = str(label_left).strip()
                    if label_left and not label_left.startswith('注') and '行次' not in label_left:
                        val_begin = _safe_float(ws.cell(row, beginning_left_col).value)
                        val_end = _safe_float(ws.cell(row, ending_left_col).value)
                        if val_end != 0 or val_begin != 0:
                            result['items'].append({
                                'label': label_left,
                                'beginning_balance': val_begin,
                                'ending_balance': val_end,
                                'side': '资产',
                            })

                # 右侧：负债及权益
                label_right = ws.cell(row, label_right_col).value
                if label_right:
                    label_right = str(label_right).strip()
                    if label_right and not label_right.startswith('注') and '行次' not in label_right:
                        val_begin = _safe_float(ws.cell(row, beginning_right_col).value)
                        val_end = _safe_float(ws.cell(row, ending_right_col).value)
                        if val_end != 0 or val_begin != 0:
                            result['items'].append({
                                'label': label_right,
                                'beginning_balance': val_begin,
                                'ending_balance': val_end,
                                'side': '负债及权益',
                            })
        else:
            # 上下排列格式：逐行扫描，按科目名判断侧
            current_side = '资产'
            for row in range(data_start, ws.max_row + 1):
                label = ws.cell(row, 1).value
                if not label:
                    continue
                label = str(label).strip()
                if not label or label.startswith('注'):
                    continue

                # 判断侧切换
                if '负债' in label and '所有者权益' in label and ('合计' in label or '总计' in label):
                    current_side = '负债及权益'
                    continue
                if '负债及所有者权益' in label:
                    current_side = '负债及权益'
                    continue

                # DT-213: 上下排列格式——使用col_map动态定位列，而非硬编码列3/4
                val_begin = _safe_float(ws.cell(row, col_map.get('asset_beginning', 3)).value)
                val_end = _safe_float(ws.cell(row, col_map.get('asset_ending', 4)).value)
                if val_end != 0 or val_begin != 0:
                    result['items'].append({
                        'label': label,
                        'beginning_balance': val_begin,
                        'ending_balance': val_end,
                        'side': current_side,
                    })

        # 提取合计项
        for item in result['items']:
            label = item['label'].replace(' ', '').replace('\u3000', '')
            if '资产总计' in label or '资产合计' in label:
                result['total_assets'] = item['ending_balance']
            if '负债' in label and '所有者权益' in label and '总计' in label:
                result['total_liab_equity'] = item['ending_balance']

        result['status'] = 'parsed'
        wb.close()

    except Exception as e:
        result['warnings'].append(f'解析异常: {str(e)}')

    return result


def parse_auxiliary(filepath: str) -> Dict[str, Any]:
    """解析辅助明细账/辅助余额表 → 标准化输出

    Args:
        filepath: Excel文件路径

    Returns:
        dict: {
            'source_type': 'auxiliary',
            'filepath': str,
            'status': 'parsed' | 'failed',
            'header_row': int,
            'col_map': dict,
            'sheet_name': str,
            'subject_name': str,
            'objects': [{
                'code': str, 'name': str, 'settlement': str,
                'debit': float, 'credit': float, 'balance': float,
                'direction': str, 'business': str,
            }],
            'warnings': list,
        }
    """
    import openpyxl

    result = {
        'source_type': 'auxiliary',
        'filepath': filepath,
        'status': 'failed',
        'header_row': 0,
        'col_map': {},
        'sheet_name': '',
        'subject_name': '',
        'objects': [],
        'warnings': [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]

            # 定位表头行
            header_row, col_map = locate_header_row(ws, 'auxiliary')
            if header_row == 0:
                # 辅助明细账可能表头检测条件较宽，尝试更宽松的检测
                for row in range(1, min(ws.max_row + 1, 10)):
                    row_values = []
                    for col in range(1, ws.max_column + 1):
                        val = ws.cell(row, col).value
                        row_values.append(str(val).strip() if val else '')
                    # 只要找到"名称"或"编码"或"余额"就认为是表头
                    if find_col_by_keywords(row_values, ['名称', '编码', '余额', '结算']):
                        header_row = row
                        # 重新提取列映射
                        col_map = find_all_cols(row_values, 'auxiliary')
                        # 转1-based
                        col_map = {k: v + 1 for k, v in col_map.items()}
                        break

            if header_row == 0:
                continue

            result['header_row'] = header_row
            result['col_map'] = col_map
            result['sheet_name'] = ws_name

            code_col = col_map.get('code')
            name_col = col_map.get('name')
            settlement_col = col_map.get('settlement')
            debit_col = col_map.get('debit')
            credit_col = col_map.get('credit')
            balance_col = col_map.get('balance')
            direction_col = col_map.get('direction')
            business_col = col_map.get('business')

            # 尝试从Sheet名提取科目名
            if not result['subject_name']:
                # 常见格式: "1122应收账款" 或 "应收账款"
                m = re.match(r'^(\d+)?\s*(.+)$', ws_name)
                if m:
                    result['subject_name'] = m.group(2).strip()

            for row in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row, name_col).value if name_col else None
                settlement = ws.cell(row, settlement_col).value if settlement_col else None

                # 至少有一个标识（名称或结算对象）
                if not name and not settlement:
                    continue

                name_str = str(name).strip() if name else ''
                settlement_str = str(settlement).strip() if settlement else ''

                # 跳过合计行
                if '合计' in name_str or '小计' in name_str:
                    continue
                if '合计' in settlement_str or '小计' in settlement_str:
                    continue

                code = str(ws.cell(row, code_col).value).strip() if code_col and ws.cell(row, code_col).value else ''
                debit = _safe_float(ws.cell(row, debit_col).value) if debit_col else 0.0
                credit = _safe_float(ws.cell(row, credit_col).value) if credit_col else 0.0
                balance = _safe_float(ws.cell(row, balance_col).value) if balance_col else 0.0
                direction = str(ws.cell(row, direction_col).value).strip() if direction_col and ws.cell(row, direction_col).value else ''
                business = str(ws.cell(row, business_col).value).strip() if business_col and ws.cell(row, business_col).value else ''

                # 如果没有balance但借方/贷方都有，计算
                if balance == 0 and (debit != 0 or credit != 0):
                    balance = debit - credit
                    if not direction:
                        direction = '借' if debit >= credit else '贷'

                obj = {
                    'code': code,
                    'name': name_str,
                    'settlement': settlement_str,
                    'debit': debit,
                    'credit': credit,
                    'balance': balance,
                    'direction': direction,
                    'business': business,
                }

                # 过滤空对象（名称和结算对象都为空）
                if name_str or settlement_str:
                    result['objects'].append(obj)

            # 只处理第一个有效Sheet
            if result['objects']:
                break

        result['status'] = 'parsed' if result['objects'] else 'failed'
        wb.close()

    except Exception as e:
        result['warnings'].append(f'解析异常: {str(e)}')

    return result


def parse_journal(filepath: str) -> Dict[str, Any]:
    """解析序时账/凭证一览表 → 标准化输出

    Args:
        filepath: Excel文件路径

    Returns:
        dict: {
            'source_type': 'journal',
            'filepath': str,
            'status': 'parsed' | 'failed',
            'header_row': int,
            'col_map': dict,
            'entries': [{
                'date': str, 'voucher_no': str, 'summary': str,
                'debit': float, 'credit': float, 'balance': float,
                'direction': str, 'counterpart': str, 'settlement': str,
            }],
            'warnings': list,
        }
    """
    import openpyxl

    result = {
        'source_type': 'journal',
        'filepath': filepath,
        'status': 'failed',
        'header_row': 0,
        'col_map': {},
        'entries': [],
        'warnings': [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        header_row, col_map = locate_header_row(ws, 'journal')
        if header_row == 0:
            result['warnings'].append('无法识别序时账表头行')
            wb.close()
            return result

        result['header_row'] = header_row
        result['col_map'] = col_map

        date_col = col_map.get('date')
        voucher_no_col = col_map.get('voucher_no')
        summary_col = col_map.get('summary')
        debit_col = col_map.get('debit')
        credit_col = col_map.get('credit')
        balance_col = col_map.get('balance')
        direction_col = col_map.get('direction')
        counterpart_col = col_map.get('counterpart')
        settlement_col = col_map.get('settlement')

        for row in range(header_row + 1, ws.max_row + 1):
            # 至少有摘要或借贷金额
            summary_val = ws.cell(row, summary_col).value if summary_col else None
            debit_val = ws.cell(row, debit_col).value if debit_col else None
            credit_val = ws.cell(row, credit_col).value if credit_col else None

            if not summary_val and not debit_val and not credit_val:
                continue

            summary_str = str(summary_val).strip() if summary_val else ''
            # 跳过合计行
            if '合计' in summary_str or '小计' in summary_str or '承前页' in summary_str or '过次页' in summary_str:
                continue

            entry = {
                'date': str(ws.cell(row, date_col).value).strip() if date_col and ws.cell(row, date_col).value else '',
                'voucher_no': str(ws.cell(row, voucher_no_col).value).strip() if voucher_no_col and ws.cell(row, voucher_no_col).value else '',
                'summary': summary_str,
                'debit': _safe_float(debit_val) if debit_val else 0.0,
                'credit': _safe_float(credit_val) if credit_val else 0.0,
                'balance': _safe_float(ws.cell(row, balance_col).value) if balance_col and ws.cell(row, balance_col).value else 0.0,
                'direction': str(ws.cell(row, direction_col).value).strip() if direction_col and ws.cell(row, direction_col).value else '',
                'counterpart': str(ws.cell(row, counterpart_col).value).strip() if counterpart_col and ws.cell(row, counterpart_col).value else '',
                'settlement': str(ws.cell(row, settlement_col).value).strip() if settlement_col and ws.cell(row, settlement_col).value else '',
            }

            result['entries'].append(entry)

        result['status'] = 'parsed' if result['entries'] else 'failed'
        wb.close()

    except Exception as e:
        result['warnings'].append(f'解析异常: {str(e)}')

    return result


def parse_income_statement(filepath: str) -> Dict[str, Any]:
    """解析利润表 → 标准化输出

    Args:
        filepath: Excel文件路径

    Returns:
        dict: {
            'source_type': 'income_statement',
            'filepath': str,
            'status': 'parsed' | 'failed',
            'header_row': int,
            'col_map': dict,
            'items': [{
                'label': str, 'current_period': float,
                'year_accumulated': float, 'last_year': float,
            }],
            'warnings': list,
        }
    """
    import openpyxl

    result = {
        'source_type': 'income_statement',
        'filepath': filepath,
        'status': 'failed',
        'header_row': 0,
        'col_map': {},
        'items': [],
        'warnings': [],
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        # 查找利润表Sheet
        ws = None
        for name in wb.sheetnames:
            if '利润' in name or '损益' in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        header_row, col_map = locate_header_row(ws, 'income_statement')
        if header_row == 0:
            result['warnings'].append('无法识别利润表表头行')
            wb.close()
            return result

        result['header_row'] = header_row
        result['col_map'] = col_map

        label_col = col_map.get('label', 1)
        current_col = col_map.get('current_period')
        accumulated_col = col_map.get('year_accumulated')
        last_year_col = col_map.get('last_year')

        data_start = header_row + 1
        for r in range(data_start, min(data_start + 3, ws.max_row + 1)):
            v = ws.cell(r, label_col).value
            if v and ('行次' in str(v) or '行号' in str(v)):
                data_start = r + 1
                break

        for row in range(data_start, ws.max_row + 1):
            label = ws.cell(row, label_col).value
            if not label:
                continue
            label = str(label).strip()
            if not label or label.startswith('注'):
                continue

            item = {
                'label': label,
                'current_period': _safe_float(ws.cell(row, current_col).value) if current_col else 0.0,
                'year_accumulated': _safe_float(ws.cell(row, accumulated_col).value) if accumulated_col else 0.0,
                'last_year': _safe_float(ws.cell(row, last_year_col).value) if last_year_col else 0.0,
            }

            if any(v != 0.0 for k, v in item.items() if k != 'label'):
                result['items'].append(item)

        result['status'] = 'parsed' if result['items'] else 'failed'
        wb.close()

    except Exception as e:
        result['warnings'].append(f'解析异常: {str(e)}')

    return result


# ============================================================
# 统一入口
# ============================================================

def auto_parse(filepath: str, source_type: str = None) -> Dict[str, Any]:
    """自动检测数据源类型并解析。

    Args:
        filepath: 文件路径（支持.xlsx/.xls/.pdf）
        source_type: 指定数据源类型（None则自动检测）

    Returns:
        dict: 解析结果（格式取决于数据源类型）
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in ('.xlsx', '.xls'):
        # Excel文件
        if source_type == 'subject_balance':
            return parse_subject_balance(filepath)
        elif source_type == 'balance_sheet':
            return parse_balance_sheet(filepath)
        elif source_type == 'auxiliary':
            return parse_auxiliary(filepath)
        elif source_type == 'journal':
            return parse_journal(filepath)
        elif source_type == 'income_statement':
            return parse_income_statement(filepath)
        else:
            # 自动检测：先扫描第一个Sheet的表头
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            row_values = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(1, col).value
                row_values.append(str(val).strip() if val else '')
            detected = detect_source_type(row_values, filepath)
            wb.close()

            if detected:
                return auto_parse(filepath, detected)
            else:
                return {
                    'source_type': 'unknown',
                    'filepath': filepath,
                    'status': 'failed',
                    'warnings': ['无法自动识别数据源类型，请指定source_type参数'],
                }

    elif ext == '.pdf':
        # PDF文件：委托给pdf_extract.py
        from pdf_extract import extract_pdf, extract_bank_statement, extract_asset_register, extract_auxiliary_balance

        if source_type == 'bank_statement':
            return extract_bank_statement(filepath)
        elif source_type == 'asset_register':
            return extract_asset_register(filepath)
        elif source_type == 'auxiliary':
            return extract_auxiliary_balance(filepath)
        else:
            # 通用PDF提取
            return extract_pdf(filepath)
    else:
        return {
            'source_type': 'unsupported',
            'filepath': filepath,
            'status': 'failed',
            'warnings': [f'不支持的文件格式: {ext}'],
        }


# ============================================================
# CLI入口
# ============================================================

if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='统一表头识别引擎 (DT-154)')
    parser.add_argument('path', help='文件路径')
    parser.add_argument('--type', choices=[
        'subject_balance', 'balance_sheet', 'journal',
        'auxiliary', 'bank_statement', 'asset_register',
        'income_statement', 'auto'
    ], default='auto', help='数据源类型（auto=自动检测）')
    parser.add_argument('--detect-only', action='store_true',
                       help='仅检测数据源类型，不解析')

    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f'❌ 文件不存在: {args.path}')
        sys.exit(1)

    if args.detect_only:
        import openpyxl
        ext = os.path.splitext(args.path)[1].lower()
        if ext in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(args.path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            row_values = []
            for col in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(1, col).value
                row_values.append(str(val).strip() if val else '')
            detected = detect_source_type(row_values, args.path)
            wb.close()
            print(f'检测类型: {detected or "未识别"}')
            print(f'表头内容: {row_values[:10]}')
        else:
            print(f'仅支持Excel文件的detect-only模式')
        sys.exit(0)

    source_type = args.type if args.type != 'auto' else None
    result = auto_parse(args.path, source_type)

    # 输出摘要
    print(f'数据源类型: {result.get("source_type", "unknown")}')
    print(f'解析状态: {result.get("status", "unknown")}')
    print(f'表头行号: {result.get("header_row", 0)}')
    print(f'列映射: {result.get("col_map", {})}')

    if result.get('subjects'):
        print(f'科目数: {len(result["subjects"])}')
    if result.get('objects'):
        print(f'结算对象数: {len(result["objects"])}')
    if result.get('entries'):
        print(f'分录数: {len(result["entries"])}')
    if result.get('items'):
        print(f'项目数: {len(result["items"])}')

    if result.get('warnings'):
        print(f'警告: {result["warnings"]}')
