"""修复评估明细表8类格式问题（Phase 4b补充）.
问题清单:
1. 打印范围从A列开始→应从B列开始（A列是检索表头辅助列）
2. 序号列格式：宋体+General→应为TNR+0（整数格式）
3. C26-C28格式：12pt+无边框→应为11pt+thin
4. 账龄/增值额/增值率公式未下拉至合计1上方
5. O-T列和A26-A43多余格式：_apply_direct_format对1~max_col全覆盖thin边框
6. 【v1.1新增】序号列空值自动填写1,2,3...
7. 【v1.1新增】条件格式下拉至数据行末尾
8. 【v1.1新增】SUM公式范围自适应修复（合计行SUM覆盖实际数据行）
"""
import sys
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side
from sheet_col_finder import find_header_cols, find_data_start_row, find_last_print_col as _find_last_print_col

# ========== 常量 ==========
FONT_TNR = 'Times New Roman'
FONT_CN = '宋体'
FONT_SIZE = 11
THIN_SIDE = Side(style='thin')
STANDARD_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER = Border()
FMT_AMOUNT = '#,##0.00'
# DT-212: 增值额/增值率会计格式——0值显示空白而非0/0%
# 模板使用会计格式: #,##0.00_);[Red]\-#,##0.00_);_(* ""_)
# 正数右对齐+右空格、负数红色+右对齐、零值显示空白
FMT_INCREMENT = '#,##0.00_);[Red]\\-#,##0.00_);_(* ""_)'
FMT_INCREMENT_RATE = '#,##0.00_);[Red]\\-#,##0.00_);_(* ""_)'
FMT_SEQ = '0'
FMT_GENERAL = 'General'

font_tnr = Font(name=FONT_TNR, size=FONT_SIZE)
font_cn = Font(name=FONT_CN, size=FONT_SIZE)
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')


def find_struct(ws):
    """识别合计1/坏账/合计2行号（A列标记 + B列文字联合识别）.
    
    DT-202: A列标记可能被数据序号覆盖，需同时扫描B列文字：
    - B列含"合  计"（含空格）→ 合计1或合计2
    - B列含"减：xxx坏账准备"/"减：xxx预计风险" → 坏账准备行
    - B列含"减：xxx减值准备"/"减：xxx计提跌价" → 减值行
    """
    total1 = total2 = bad_debt = None
    # DT-202: 增加B列文字识别
    b_total1_candidates = []  # 可能有多个"合计"，按顺序分配
    
    for r in range(1, min(ws.max_row + 1, 60)):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        
        # 优先A列标记
        if a and isinstance(a, str):
            t = a.replace(' ', '').strip()
            if t == '合计1' or (t == '合计' and not total1):
                total1 = r
                continue
            elif t == '合计2':
                total2 = r
                continue
            elif t in ('坏账准备', '预计风险', '预计损失', '计提跌价准备', '减值准备', '跌价准备'):
                bad_debt = r
                continue
        
        # DT-202: A列无标记时，扫描B列文字
        if b and isinstance(b, str):
            b_text = b.replace(' ', '').strip()
            if '合' in b_text and '计' in b_text:
                if '2' in b_text:
                    if not total2:
                        total2 = r
                else:
                    b_total1_candidates.append(r)
            elif any(kw in b_text for kw in ['坏账准备', '预计风险', '预计损失', '计提跌价', '减值准备', '跌价准备']):
                if not bad_debt:
                    bad_debt = r
    
    # 如果A列没找到合计1，用B列候选分配
    if not total1 and b_total1_candidates:
        if total2 and len(b_total1_candidates) >= 2:
            # 有合计2时，B列第一个"合计"为合计1
            total1 = b_total1_candidates[0]
        elif len(b_total1_candidates) >= 1:
            # 只有一个合计行且无合计2 → 这是合计1（单行结构）
            total1 = b_total1_candidates[0]
            # 如果有两个以上合计行且无合计2标记 → 第一个=合计1，第二个=合计2
            if len(b_total1_candidates) >= 2:
                total1 = b_total1_candidates[0]
                total2 = b_total1_candidates[-1]
                # 中间行如果没识别为bad_debt，找减值行
                if not bad_debt:
                    for mid_r in b_total1_candidates[1:-1]:
                        if mid_r != total2:
                            bad_debt = mid_r

    # DT-202补充：A列找到合计1但没找到合计2时，从B列候选中分配合计2
    if not total2 and b_total1_candidates:
        # 找B列候选中行号大于total1的合计行
        if total1:
            for cand_r in b_total1_candidates:
                if cand_r > total1 and cand_r != total1:
                    total2 = cand_r
                    break
    
    return total1, bad_debt, total2


def find_last_print_col(ws):
    """找到打印范围的右边界列（委托给sheet_col_finder公共模块）。"""
    return _find_last_print_col(ws)


def fix_workbook(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    fixes = {
        'print_area': 0,
        'seq_format': 0,
        'cell_format': 0,
        'formula_fill': 0,
        'border_cleanup': 0,
        'seq_autofill': 0,        # v1.1新增
        'cond_format_fill': 0,     # v1.1新增
        'sum_range_fix': 0,        # v1.1新增
    }

    for sname in wb.sheetnames:
        if '汇总' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        ws = wb[sname]
        total1, bad_debt, total2 = find_struct(ws)
        if not total1:
            continue

        data_start = find_data_start_row(ws)  # 动态查找数据起始行
        data_end = total1 - 1
        last_print_col = find_last_print_col(ws)
        header_cols = find_header_cols(ws)  # 动态列位映射

        # ===== 1. 打印范围修正 =====
        end_row = total2 if total2 else total1
        new_print = f"B1:{get_column_letter(last_print_col)}{end_row}"
        old_print = str(ws.print_area)
        if old_print != new_print:
            ws.print_area = new_print
            fixes['print_area'] += 1
            print(f"  [{sname}] 打印范围: {old_print} -> {new_print}")

        # ===== 2-3-5: 逐行逐列修复格式 =====
        struct_rows = {total1}
        if bad_debt:
            struct_rows.add(bad_debt)
        if total2:
            struct_rows.add(total2)

        # DT-153v3: 动态检测表头行（不再固定Row 5）
        header_row = None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 25)):
                val = ws.cell(row=r, column=c).value
                if val and ('账面价值' in str(val) or '评估价值' in str(val) or '结算对象' in str(val)):
                    header_row = r
                    break
            if header_row:
                break
        if not header_row:
            header_row = 5  # fallback

        for r in range(data_start, (total2 or total1) + 1):
            for c in range(1, min(ws.max_column + 1, 25)):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue

                in_print_area = (2 <= c <= last_print_col)
                header = ws.cell(row=header_row, column=c).value
                header_text = str(header) if header else ''

                is_seq = '序号' in header_text
                is_increment = '增值额' in header_text  # DT-212: 增值额用会计格式
                is_amount = any(k in header_text for k in
                               ['账面价值', '评估价值', '外币', '汇率', '减值', '原值', '余额'])
                is_pct = '增值率' in header_text  # DT-208: 增值率列用金额格式(与增值额一致)
                is_left_text = any(k in header_text for k in
                                   ['结算对象', '户名', '项目及内容', '设备名称', '备注', '存放地点',
                                    '规格型号', '生产厂家', '合同编号', '税费种类', '征税机关'])
                is_center_text = any(k in header_text for k in
                                    ['业务内容', '发生日期', '结算内容'])
                is_date = '发生日期' in header_text or '日期' in header_text
                is_age = '账龄' in header_text

                # --- 2. 序号列格式修正 ---
                if is_seq and in_print_area:
                    if cell.font.name != FONT_TNR or cell.number_format != FMT_SEQ:
                        cell.font = font_tnr
                        cell.alignment = align_center
                        cell.number_format = FMT_SEQ
                        fixes['seq_format'] += 1

                # --- 3. 其他列格式修正 ---
                elif is_left_text and in_print_area:
                    if cell.font.name != FONT_CN or cell.font.size != FONT_SIZE:
                        cell.font = font_cn
                        cell.alignment = align_left
                        cell.number_format = FMT_GENERAL
                        fixes['cell_format'] += 1
                elif is_center_text and in_print_area:
                    # DT-206: 发生日期列用Times New Roman字体，其他居中文本用宋体
                    if is_date:
                        if cell.font.name != FONT_TNR or cell.font.size != FONT_SIZE:
                            cell.font = font_tnr
                            cell.alignment = align_center
                            cell.number_format = 'yyyy"年"mm"月"'
                            fixes['cell_format'] += 1
                    else:
                        if cell.font.name != FONT_CN or cell.font.size != FONT_SIZE:
                            cell.font = font_cn
                            cell.alignment = align_center
                            cell.number_format = FMT_GENERAL
                            fixes['cell_format'] += 1
                elif is_amount and in_print_area:
                    # DT-213: 加number_format检查——公式列font对但format=General时也要修
                    needs_fix = (cell.font.name != FONT_TNR or cell.font.size != FONT_SIZE
                                 or (isinstance(cell.value, (int, float)) and cell.number_format != FMT_AMOUNT))
                    if needs_fix:
                        cell.font = font_tnr
                        cell.alignment = align_right
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = FMT_AMOUNT
                        elif cell.value is None or cell.value == '':
                            cell.number_format = FMT_GENERAL
                        fixes['cell_format'] += 1
                elif is_increment and in_print_area:
                    # DT-212: 增值额列用会计格式，0值显示空白
                    # DT-213: 加number_format检查——font对了但format=General时也要修
                    if cell.font.name != FONT_TNR or cell.font.size != FONT_SIZE or cell.number_format != FMT_INCREMENT:
                        cell.font = font_tnr
                        cell.alignment = align_right
                        cell.number_format = FMT_INCREMENT
                        fixes['cell_format'] += 1
                elif is_pct and in_print_area:
                    # DT-212: 增值率列用会计金额格式(与增值额一致)，0值显示空白
                    # 2026-05-26修复：公式=L6/J6*100产出已乘100的数值(如15.5)，
                    # 用0.00%格式会再乘100→显示1550%，改为金额格式后15.5显示为15.50
                    if cell.font.name != FONT_TNR or cell.number_format != FMT_INCREMENT_RATE:
                        cell.font = font_tnr
                        cell.alignment = align_right
                        cell.number_format = FMT_INCREMENT_RATE
                        fixes['cell_format'] += 1
                elif is_age and in_print_area:
                    if cell.font.name != FONT_TNR:
                        cell.font = font_tnr
                        cell.alignment = align_center
                        fixes['cell_format'] += 1

                # --- 5. 边框修正 ---
                if in_print_area:
                    if cell.border.left.style != 'thin' or cell.border.right.style != 'thin':
                        cell.border = STANDARD_BORDER
                        fixes['border_cleanup'] += 1
                else:
                    # 打印范围外(A列, O-T列)不应有thin边框
                    if cell.border.left.style == 'thin' or cell.border.right.style == 'thin':
                        cell.border = NO_BORDER
                        fixes['border_cleanup'] += 1

                # DT-153v3: 名称列12pt→11pt修正（不再固定C列）
                if is_left_text and cell.font.size and abs(cell.font.size - 12) < 0.5:
                    cell.font = font_cn
                    fixes['cell_format'] += 1
                    cell.border = STANDARD_BORDER
                    fixes['cell_format'] += 1

        # ===== 4. 公式下拉 =====
        ref_row = data_start

        # 4a. 账龄（动态查找列号）
        age_col = header_cols.get('账龄')
        if age_col:
            ref_formula_age = ws.cell(row=ref_row, column=age_col).value
            if isinstance(ref_formula_age, str) and ref_formula_age.startswith('='):
                for r in range(data_start + 1, total1):
                    cell = ws.cell(row=r, column=age_col)
                    if not isinstance(cell, MergedCell) and (cell.value is None or cell.value == ''):
                        new_formula = re.sub(
                            r'([A-Z])(\d+)',
                            lambda m: f'{m.group(1)}{r}' if int(m.group(2)) == ref_row else m.group(0),
                            ref_formula_age
                        )
                        cell.value = new_formula
                        fixes['formula_fill'] += 1

        # 4b. 增值额（动态查找列号）—— DT-203: 扩展覆盖结构行
        va_col = header_cols.get('增值额')
        if va_col:
            ref_formula_va = ws.cell(row=ref_row, column=va_col).value
            if isinstance(ref_formula_va, str) and ref_formula_va.startswith('='):
                # DT-203: 覆盖data_start+1到total2（含结构行），不再仅到total1-1
                end_row = total2 if total2 else total1
                for r in range(data_start + 1, end_row + 1):
                    cell = ws.cell(row=r, column=va_col)
                    if isinstance(cell, MergedCell):
                        continue
                    # 结构行的增值额公式特殊处理：
                    # 合计1行: 增值额=评估价值-账面价值（基于自身SUM结果）
                    # 坏账准备行: 增值额=评估价值-账面价值（基于自身值）
                    # 合计2行: 增值额=评估价值-账面价值（基于ROUND差值结果）
                    # 统一公式: =评估价值列{r}-账面价值列{r}（每行指向自身）
                    new_formula = re.sub(
                        r'([A-Z])(\d+)',
                        lambda m: f'{m.group(1)}{r}' if int(m.group(2)) == ref_row else m.group(0),
                        ref_formula_va
                    )
                    # 对已有公式的结构行也强制修正（修复行号锁死问题）
                    if r in struct_rows:
                        if cell.value != new_formula:
                            cell.value = new_formula
                            fixes['formula_fill'] += 1
                    elif cell.value is None:
                        cell.value = new_formula
                        fixes['formula_fill'] += 1

        # 4c. 增值率（动态查找列号）—— DT-203: 扩展覆盖结构行
        vr_col = header_cols.get('增值率')
        if vr_col:
            ref_formula_vr = ws.cell(row=ref_row, column=vr_col).value
            if isinstance(ref_formula_vr, str) and ref_formula_vr.startswith('='):
                # DT-203: 覆盖data_start+1到total2（含结构行）
                end_row = total2 if total2 else total1
                for r in range(data_start + 1, end_row + 1):
                    cell = ws.cell(row=r, column=vr_col)
                    if isinstance(cell, MergedCell):
                        continue
                    new_formula = re.sub(
                        r'([A-Z])(\d+)',
                        lambda m: f'{m.group(1)}{r}' if int(m.group(2)) == ref_row else m.group(0),
                        ref_formula_vr
                    )
                    # 对已有公式的结构行也强制修正
                    if r in struct_rows:
                        if cell.value != new_formula:
                            cell.value = new_formula
                            fixes['formula_fill'] += 1
                    elif cell.value is None:
                        cell.value = new_formula
                        fixes['formula_fill'] += 1

        # ===== 6. 序号列空值自动填写（v1.3：DT-201空行跳过 + DT-204 A列冗余序号清除）=====
        # DT-201: 只对有数据行(C/D列非空)填序号，空行跳过不填
        # DT-204: A列是模板辅助标记列（"检索表头"/"合计1"等），数据行A列必须为空
        #         旧修复脚本(fix_7_issues.py)曾将序号写入A列→与B列序号重复+覆盖结构行标记
        #         现清除数据行A列的冗余序号，只保留模板原始标记
        seq_col = header_cols.get('序号')
        if seq_col:
            seq_num = 1
            for r in range(data_start, total1):
                # DT-201: 只对有数据的行填序号，空行跳过
                # 判断依据：C列(结算对象/项目及内容)或D列(业务内容)非空
                has_data = False
                for check_col_name in ['结算对象', '项目及内容', '业务内容']:
                    check_col = header_cols.get(check_col_name)
                    if check_col:
                        cv = ws.cell(row=r, column=check_col).value
                        if cv is not None and cv != '':
                            has_data = True
                            break
                # 兜底：检查C列和D列
                if not has_data:
                    for fallback_col in [3, 4]:
                        cv = ws.cell(row=r, column=fallback_col).value
                        if cv is not None and cv != '':
                            has_data = True
                            break

                # DT-204: 清除数据行A列冗余序号
                # A列在模板中仅用于结构标记（"检索表头"/"合计1"等），数据行A列应为空
                # 旧脚本曾将序号1,2,3...写入A列，与B列序号完全重复
                a_cell = ws.cell(row=r, column=1)
                if not isinstance(a_cell, MergedCell):
                    a_val = a_cell.value
                    if a_val is not None and a_val != '':
                        # 数据行区域不应有A列内容（结构行已由struct_rows处理）
                        a_cell.value = None
                        fixes['seq_autofill'] += 1

                cell = ws.cell(row=r, column=seq_col)
                if isinstance(cell, MergedCell):
                    continue

                if not has_data:
                    # DT-201: 空行清除序号（如果之前被误填）
                    if cell.value is not None and cell.value != '':
                        cell.value = None
                        fixes['seq_autofill'] += 1
                    continue

                val = cell.value
                if val is None or val == '':
                    cell.value = seq_num
                    cell.number_format = FMT_SEQ
                    cell.font = font_tnr
                    cell.alignment = align_center
                    fixes['seq_autofill'] += 1
                elif isinstance(val, (int, float)):
                    seq_num = int(val) + 1  # 从已有序号继续
                    continue
                seq_num += 1

        # ===== 6b. DT-204: 结构行A列标记恢复 =====
        # 旧脚本(fix_7_issues.py)可能将序号写入结构行A列→覆盖了模板原始标记
        # 恢复规则：结构行A列为空/数字时，根据B列文字推断并恢复A列标记
        # 重要：合计行需根据位置(total1/total2)区分"合计1"/"合计2"，不能仅靠B列"合  计"
        _STRUCT_A_MAP_DEDUCT = {
            # B列关键词 → A列标记（减值行专用）
            '坏账准备': '坏账准备',
            '预计风险': '预计风险',
            '预计损失': '预计损失',
            '减值准备': '减值准备',
            '计提跌价': '计提跌价准备',
            '跌价准备': '跌价准备',
        }
        struct_row_list = [total1]
        if bad_debt:
            struct_row_list.append(bad_debt)
        if total2:
            struct_row_list.append(total2)
        # 也检查total1和total2之间的行（可能有预计损失行）
        if total1 and total2:
            for r in range(total1 + 1, total2):
                if r not in struct_row_list:
                    b_val = ws.cell(row=r, column=2).value
                    if b_val and isinstance(b_val, str):
                        b_text = b_val.replace(' ', '').strip()
                        if any(kw in b_text for kw in ['坏账准备', '预计风险', '预计损失',
                                                         '减值准备', '计提跌价', '跌价准备']):
                            struct_row_list.append(r)

        for r in struct_row_list:
            a_cell = ws.cell(row=r, column=1)
            if isinstance(a_cell, MergedCell):
                continue
            a_val = a_cell.value
            # A列已有正确的字符串标记→跳过
            if a_val and isinstance(a_val, str) and a_val.replace(' ', '').strip():
                continue
            # A列为空或为数字→需要恢复
            b_val = ws.cell(row=r, column=2).value
            if not b_val or not isinstance(b_val, str):
                continue
            b_text = b_val.replace(' ', '').strip()

            # 根据行位置推断合计行标记（优先级最高）
            restored = None
            if r == total1 and '合' in b_text and '计' in b_text:
                restored = '合计1'
            elif r == total2 and '合' in b_text and '计' in b_text:
                restored = '合计2'
            else:
                # 减值行：用B列关键词推断
                for kw, marker in _STRUCT_A_MAP_DEDUCT.items():
                    if kw in b_text:
                        restored = marker
                        break
            if restored:
                a_cell.value = restored
                fixes['seq_autofill'] += 1

        # ===== 7. 条件格式下拉（v1.1新增）=====
        # 从参考行(data_start)复制条件格式到所有数据行
        if ws.conditional_formatting:
            new_cf_rules = []
            for cf in ws.conditional_formatting:
                cf_range = str(cf)
                # 检查是否只覆盖了第1行数据
                if str(data_start) in cf_range and str(total1 - 1) not in cf_range:
                    # 扩展范围到全部数据行
                    # DT-153v2: 修正条件格式范围替换逻辑
                    # 原逻辑用replace替换行号，可能错误替换列号中的数字
                    # 新逻辑: 精确替换行号部分
                    import re as _re
                    # 解析原范围的行列
                    # 格式如: $D$7:$D$7 或 D7:D7
                    range_match = _re.match(
                        r'(\$?)([A-Z]+)(\$?)(\d+):(\$?)([A-Z]+)(\$?)(\d+)',
                        cf_range
                    )
                    if range_match:
                        col_start = range_match.group(2)
                        col_end = range_match.group(6)
                        new_range = f'{col_start}{data_start}:{col_end}{total1 - 1}'
                    else:
                        # fallback: 简单替换
                        new_range = cf_range.replace(
                            str(data_start), f'{data_start}:{total1 - 1}'
                        ).replace('$', '')

                    # 创建新的条件格式规则（深拷贝避免修改原规则）
                    for rule in cf.rules:
                        from openpyxl.formatting.rule import CellIsRule, FormulaRule
                        from copy import copy, deepcopy
                        try:
                            new_rule = deepcopy(rule)
                        except Exception:
                            new_rule = copy(rule)
                        new_cf_rules.append((new_range, new_rule))
                        fixes['cond_format_fill'] += 1
            # 应用新的条件格式规则
            for cf_range, rule in new_cf_rules:
                try:
                    ws.conditional_formatting.add(cf_range, rule)
                except Exception as e:
                    # 条件格式添加失败不阻断
                    print(f'  [WARN] 条件格式添加失败: {e}')

        # ===== 8. SUM公式范围自适应修复（v1.1新增）=====
        # 检查合计1行的SUM公式是否覆盖了实际数据行范围
        for c in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(row=total1, column=c)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if isinstance(val, str) and 'SUM' in val.upper():
                # 解析SUM范围
                match = re.search(r'SUM\(([A-Z])(\d+):([A-Z])(\d+)\)', val, re.IGNORECASE)
                if match:
                    col_letter = match.group(1)
                    start_row = int(match.group(2))
                    end_row = int(match.group(4))
                    # 修正：起始行应为data_start，结束行应为total1-1
                    if start_row != data_start or end_row != total1 - 1:
                        new_formula = f'=SUM({col_letter}{data_start}:{col_letter}{total1 - 1})'
                        cell.value = new_formula
                        fixes['sum_range_fix'] += 1

        # 坏账准备行/合计2行的SUM也检查
        if bad_debt and total2:
            for target_row in [bad_debt, total2]:
                if target_row:
                    for c in range(1, min(ws.max_column + 1, 25)):
                        cell = ws.cell(row=target_row, column=c)
                        if isinstance(cell, MergedCell):
                            continue
                        val = cell.value
                        if isinstance(val, str) and 'SUM' in val.upper():
                            match = re.search(r'SUM\(([A-Z])(\d+):([A-Z])(\d+)\)', val, re.IGNORECASE)
                            if match:
                                col_letter = match.group(1)
                                start_row = int(match.group(2))
                                end_row = int(match.group(4))
                                # 合计2行应覆盖合计1到合计2之间
                                if target_row == total2 and bad_debt:
                                    expected_start = total1
                                    expected_end = target_row - 1
                                    if start_row != expected_start or end_row != expected_end:
                                        new_formula = f'=SUM({col_letter}{expected_start}:{col_letter}{expected_end})'
                                        cell.value = new_formula
                                        fixes['sum_range_fix'] += 1

    # 保存
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=0)
    wb.save(filepath)
    wb.close()

    print(f"\n===== 修复统计 =====")
    print(f"打印范围修正: {fixes['print_area']}个sheet")
    print(f"序号格式修正: {fixes['seq_format']}个单元格")
    print(f"其他格式修正: {fixes['cell_format']}个单元格")
    print(f"公式下拉: {fixes['formula_fill']}个单元格")
    print(f"边框清理: {fixes['border_cleanup']}个单元格")
    print(f"序号自动填写: {fixes['seq_autofill']}个单元格")
    print(f"条件格式下拉: {fixes['cond_format_fill']}个规则")
    print(f"SUM范围修复: {fixes['sum_range_fix']}个公式")
    return fixes


if __name__ == '__main__':
    # DT-153v3: 不再硬编码项目特定文件名
    fp = sys.argv[1] if len(sys.argv) > 1 else None
    if not fp:
        # 自动查找当前目录下含"评估明细表"的xlsx文件
        import glob as _glob
        candidates = _glob.glob('*评估明细表*.xlsx')
        fp = candidates[0] if candidates else None
    if not fp:
        print('用法: python fix_format_issues.py <评估明细表.xlsx>')
        sys.exit(1)
    fix_workbook(fp)
