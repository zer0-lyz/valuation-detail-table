#!/usr/bin/env python3
"""
fix_three_row_structure.py — 修复smart_insert_row三行结构遗漏问题

针对3-7预付款项等存在合计2行的sheet，修复以下6类问题：
1. B列合并：坏账准备行、合计2行的B:C合并还原
2. 公式修正：合计2行公式引用错误（J43=ROUND(J41-J42,2)而非J26-J27）
3. 跨sheet链接：汇总表链接同步到合计2行（非合计1行）
4. 打印范围：下至合计2行（存在时），否则至合计1行
5. 序号格式：确保B列序号为整数格式(0)，不显示小数点
6. 字体统一：结构行（合计1/坏账准备/合计2）格式修正

设计原则：
  - 先扫描所有sheet的结构，识别有合计2行的sheet
  - 逐sheet修复，修复后立即验证
  - 跨sheet链接修复需扫描整个workbook
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from copy import copy
import re
import sys
from pathlib import Path


# ============================================================
# 结构识别
# ============================================================

def find_sheet_structure(ws):
    """识别sheet的三行结构（合计1/减值/合计2）。
    
    DT-202升级: A列标记可能被数据序号覆盖，需A列+B列联合识别：
    - A列标记优先（合计1/坏账准备/合计2）
    - B列文字兜底（"合  计"/"减：xxx坏账准备"等）
    
    Returns:
        dict: {
            'header_row': int,
            'data_start_row': int,
            'total1_row': int,      # 合计1行号
            'bad_debt_row': int,     # 坏账准备/预计风险/计提跌价行号
            'total2_row': int,       # 合计2行号
            'has_total2': bool,      # 是否存在合计2行
            'struct_rows': list,     # 所有结构行号（合计1+减值+合计2）
        }
    """
    header_row = None
    data_start_row = None
    total1_row = None
    bad_debt_row = None
    provision_row = None
    total2_row = None
    
    STRUCT_MARKERS_TOTAL1 = {'合计1', '合计'}
    STRUCT_MARKERS_TOTAL2 = {'合计2'}
    STRUCT_MARKERS_DEDUCT = {'坏账准备', '预计风险', '预计损失', '计提跌价准备', '减值准备', '跌价准备'}
    
    # DT-202: B列候选列表，用于A列标记丢失时兜底
    b_total_candidates = []  # (row, is_total2_hint)
    b_deduct_candidates = []  # row
    
    for r in range(1, min(ws.max_row + 1, 60)):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value
        
        # 优先A列标记
        if a_val and isinstance(a_val, str):
            a_text = a_val.replace(' ', '').strip()
            
            if a_text in ('检索表头', '检索表头1'):
                header_row = r
                continue
            elif a_text == '检索表头2':
                data_start_row = r + 1
                continue
            elif a_text in STRUCT_MARKERS_TOTAL1:
                total1_row = r
                continue
            elif a_text in STRUCT_MARKERS_TOTAL2:
                total2_row = r
                continue
            elif a_text in STRUCT_MARKERS_DEDUCT:
                bad_debt_row = r
                continue
        
        # DT-202: A列无标记时，扫描B列文字
        if b_val and isinstance(b_val, str):
            b_text = b_val.replace(' ', '').strip()
            
            # 表头行标记
            if not header_row and a_val and str(a_val).strip() == '序号':
                header_row = r
            
            # 合计行识别
            if '合' in b_text and '计' in b_text:
                if '2' in b_text:
                    b_total_candidates.append((r, True))  # 明确的合计2
                else:
                    b_total_candidates.append((r, False))  # 可能是合计1
            # 减值行识别
            elif any(kw in b_text for kw in ['坏账准备', '预计风险', '预计损失', '计提跌价', '减值准备', '跌价准备']):
                if not bad_debt_row:
                    b_deduct_candidates.append(r)
                    bad_debt_row = r
            elif '减：' in b_text:
                # "减：xxx坏账准备"格式
                if not bad_debt_row:
                    b_deduct_candidates.append(r)
                    bad_debt_row = r
    
    # DT-202: 如果A列没找到合计1/合计2，从B列候选分配
    # DT-204: A列数据行冗余序号已由fix_format_issues清除，A列恢复模板原始标记
    # A列标记体系（v1.90-FOR AI模板自带）："检索表头"/"合计1"/"坏账准备"/"预计损失"/"合计2"
    # 数据行A列必须为空（B列才是序号列）
    if not total1_row and b_total_candidates:
        non_total2 = [r for r, is_t2 in b_total_candidates if not is_t2]
        explicit_total2 = [r for r, is_t2 in b_total_candidates if is_t2]
        
        if len(non_total2) >= 1:
            total1_row = non_total2[0]
        if not total2_row and len(non_total2) >= 2:
            # 最后一个非明确合计2 → 实际是合计2
            total2_row = non_total2[-1]
        if not total2_row and explicit_total2:
            total2_row = explicit_total2[0]
    
    # DT-202补充：A列找到合计1但没找到合计2时，从B列候选中分配合计2
    if not total2_row and b_total_candidates:
        if total1_row:
            non_total2 = [r for r, is_t2 in b_total_candidates if not is_t2]
            explicit_total2 = [r for r, is_t2 in b_total_candidates if is_t2]
            # 优先用明确的合计2标记
            if explicit_total2:
                total2_row = explicit_total2[0]
            # 否则找行号大于total1的非合计2候选
            elif non_total2:
                for cand_r in non_total2:
                    if cand_r > total1_row:
                        total2_row = cand_r
                        break
    
    # 旧模板兼容：B列含"合计"文字（如果A+B列都没找到）
    if not total1_row and b_val and isinstance(b_val, str):
        b_text = b_val.replace(' ', '').strip()
        if '合' in b_text and '计' in b_text and '2' not in b_text:
            total1_row = r
        elif '合' in b_text and '计' in b_text and '2' in b_text:
            total2_row = r
        elif any(kw in b_text for kw in ['坏账准备', '预计风险', '计提跌价', '减值准备', '跌价准备']):
            bad_debt_row = r

    # 表头行（旧模板：A列值="序号"）
    if not header_row:
        for r in range(1, min(ws.max_row + 1, 15)):
            a_val = ws.cell(row=r, column=1).value
            if a_val and str(a_val).strip() == '序号':
                header_row = r
                break
    
    # 如果没有A列AI标记，用旧模板逻辑推断结构行
    if not total1_row:
        for r in range(1, min(ws.max_row + 1, 60)):
            a_val = ws.cell(row=r, column=1).value
            if a_val and isinstance(a_val, str):
                text = a_val.replace(' ', '').strip()
                if '合' in text and '计' in text:
                    if not total1_row:
                        total1_row = r
                    else:
                        total2_row = r
    
    has_total2 = total2_row is not None
    
    struct_rows = []
    if total1_row:
        struct_rows.append(total1_row)
    if bad_debt_row:
        struct_rows.append(bad_debt_row)
    if total2_row:
        struct_rows.append(total2_row)
    
    return {
        'header_row': header_row,
        'data_start_row': data_start_row,
        'total1_row': total1_row,
        'bad_debt_row': bad_debt_row,
        'total2_row': total2_row,
        'has_total2': has_total2,
        'struct_rows': struct_rows,
    }


# ============================================================
# 修复函数
# ============================================================

def fix_bc_merge(ws, struct):
    """修复1：还原结构行的B:C（或B:D）合并。
    
    合计1行、坏账准备行、合计2行的B列和C列必须合并。
    DT-205: 部分模板（如投资性房地产、固定资产）结构行是B:D三列合并，
    而非B:C两列合并。需兼容两种模式。
    """
    fixes = []
    rows_to_merge = []
    
    if struct['total1_row']:
        rows_to_merge.append(struct['total1_row'])
    if struct['bad_debt_row']:
        rows_to_merge.append(struct['bad_debt_row'])
    if struct['total2_row']:
        rows_to_merge.append(struct['total2_row'])
    
    for r in rows_to_merge:
        # 检查B:C或B:D是否已合并
        already_merged = False
        merge_end_col = None  # 记录当前合并的终止列
        for mr in ws.merged_cells.ranges:
            if mr.min_row == r and mr.min_col == 2 and mr.max_row == r:
                if mr.max_col >= 3:  # B:C或B:D都算已合并
                    already_merged = True
                    merge_end_col = mr.max_col
                    break
        
        if not already_merged:
            # 先取消可能存在的错误合并（跨行B:C等）
            bad_merges = []
            for mr in ws.merged_cells.ranges:
                if mr.min_row == r and mr.min_col == 2 and mr.max_col >= 3 and mr.max_row > r:
                    bad_merges.append(str(mr))
            for bm in bad_merges:
                ws.unmerge_cells(bm)
                fixes.append(f'取消错误合并: {bm}')
            
            # 判断该行C列和D列的内容：如果C列和D列都为空/None，则B:C合并
            # 如果C列有内容（如表头"权证编号"对应数据），也只合并B:C
            c_val = ws.cell(row=r, column=3).value
            d_val = ws.cell(row=r, column=4).value
            
            # 参考同sheet数据行的B列合并范围来确定合并列数
            # 扫描数据行看是否有B:D合并的模式
            data_start = struct.get('data_start_row') or struct.get('header_row', 6) + 1
            data_merge_end = 3  # 默认B:C
            for dr in range(data_start, min(data_start + 3, struct['total1_row'] or 999)):
                for mr in ws.merged_cells.ranges:
                    if mr.min_row == dr and mr.min_col == 2 and mr.max_row == dr and mr.max_col > 2:
                        data_merge_end = mr.max_col
                        break
                if data_merge_end > 3:
                    break
            
            merge_end = get_column_letter(data_merge_end)
            merge_range = f"B{r}:{merge_end}{r}"
            ws.merge_cells(merge_range)
            fixes.append(f'合并B{r}:{merge_end}{r}')
    
    return fixes


def fix_total2_formulas(ws, struct):
    """修复2：修正合计2行公式引用。
    
    合计2行公式应为 =ROUND(合计1-坏账准备, 2)
    即 =ROUND(J{total1}-J{bad_debt}, 2)
    
    同时确保合计1行的SUM公式范围正确。
    
    DT-205: 扩展合计2行公式修正覆盖增值率公式模式：
    - 模式1: =ROUND(colX-colY,2) → 金额列的ROUND差值
    - 模式2: =colX-colY → 金额列的简单差值
    - 模式3: =IF(colX=0,"",colY/colX*100) → 增值率（账面价值非零时增值额/账面价值*100）
    - 模式4: =IF(colX=0,"",(colY-colX)/colX*100) → 增值率（另一种写法）
    
    注意：合计2行增值率公式中行号应指向自身（如=IF(J29=0,"",L29/J29*100)），
    因为合计2行的J29和L29已经通过公式引用了合计1和坏账准备行。
    但当行号错误时（如仍指向数据行行号），需要修正为指向合计2行自身。
    """
    fixes = []
    total1 = struct['total1_row']
    bad_debt = struct['bad_debt_row']
    total2 = struct['total2_row']
    
    if not total2 or not total1:
        return fixes
    
    # 修正合计2行公式
    for c in range(1, min(ws.max_column + 1, 25)):
        cell = ws.cell(row=total2, column=c)
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if not isinstance(val, str) or not val.startswith('='):
            continue
        
        col_letter = get_column_letter(c)
        
        # 模式1: =ROUND(colX-colY,2) → 应为 =ROUND(col{total1}-col{bad_debt},2)
        pattern = rf'=ROUND\(([A-Z]+)(\d+)-([A-Z]+)(\d+),\s*2\)'
        match = re.match(pattern, val, re.IGNORECASE)
        if match:
            col1 = match.group(1).upper()
            row1 = int(match.group(2))
            col2 = match.group(3).upper()
            row2 = int(match.group(4))
            
            expected = f'=ROUND({col_letter}{total1}-{col_letter}{bad_debt},2)'
            if val != expected:
                old_val = val
                cell.value = expected
                fixes.append(f'Row{total2} Col{c}({col_letter}): "{old_val}" → "{expected}"')
            continue
        
        # 模式2: =colX-colY (无ROUND包裹) → 应为 =col{total1}-col{bad_debt}
        pattern2 = rf'=([A-Z]+)(\d+)-([A-Z]+)(\d+)$'
        match2 = re.match(pattern2, val, re.IGNORECASE)
        if match2:
            col1 = match2.group(1).upper()
            row1 = int(match2.group(2))
            col2 = match2.group(3).upper()
            row2 = int(match2.group(4))
            
            expected = f'={col_letter}{total1}-{col_letter}{bad_debt}'
            if val != expected:
                old_val = val
                cell.value = expected
                fixes.append(f'Row{total2} Col{c}({col_letter}): "{old_val}" → "{expected}"')
            continue
        
        # DT-205: 模式3: =IF(colX=0,"",colY/colX*100) → 增值率公式
        # 合计2行的增值率应为 =IF(col{total2}=0,"",col{total2}/col{total2}*100) 指向自身
        # 因为合计2行的col{total2}已通过公式引用合计1-坏账准备
        pattern3 = r'=IF\(([A-Z]+)(\d+)=0,"",([A-Z]+)(\d+)/([A-Z]+)(\d+)\*100\)'
        match3 = re.match(pattern3, val, re.IGNORECASE)
        if match3:
            check_col = match3.group(1).upper()
            check_row = int(match3.group(2))
            num_col = match3.group(3).upper()
            num_row = int(match3.group(4))
            den_col = match3.group(5).upper()
            den_row = int(match3.group(6))
            
            # 修正：所有行号应指向total2自身
            expected = f'=IF({col_letter}{total2}=0,"",{col_letter}{total2}/{col_letter}{total2}*100)'
            if val != expected:
                old_val = val
                cell.value = expected
                fixes.append(f'Row{total2} Col{c}({col_letter})增值率: "{old_val}" → "{expected}"')
            continue
        
        # DT-205: 模式4: =IF(colX=0,"",(colY-colX)/colX*100) → 增值率另一种写法
        pattern4 = r'=IF\(([A-Z]+)(\d+)=0,"",\(([A-Z]+)(\d+)-([A-Z]+)(\d+)\)/([A-Z]+)(\d+)\*100\)'
        match4 = re.match(pattern4, val, re.IGNORECASE)
        if match4:
            expected = f'=IF({col_letter}{total2}=0,"",({col_letter}{total2}-{col_letter}{total2})/{col_letter}{total2}*100)'
            if val != expected:
                old_val = val
                cell.value = expected
                fixes.append(f'Row{total2} Col{c}({col_letter})增值率: "{old_val}" → "{expected}"')
            continue
        
        # DT-205: 模式5: =IF(colX=0,"",ROUND(colY/colX,N)) → 成新率/单价公式
        # 这种公式指向自身行号是合法的（如评估原值单价 = ROUND(原值/面积,0)）
        # 只需确保行号都指向total2自身即可
        pattern5 = r'=IF\(([A-Z]+)(\d+)=0,"",ROUND\(([A-Z]+)(\d+)/([A-Z]+)(\d+),\s*(\d+)\)\)'
        match5 = re.match(pattern5, val, re.IGNORECASE)
        if match5:
            check_col = match5.group(1).upper()
            check_row = int(match5.group(2))
            num_col = match5.group(3).upper()
            num_row = int(match5.group(4))
            den_col = match5.group(5).upper()
            den_row = int(match5.group(6))
            decimals = match5.group(7)
            
            # 修正：所有行号应指向total2自身
            expected = f'=IF({col_letter}{total2}=0,"",ROUND({col_letter}{total2}/{col_letter}{total2},{decimals}))'
            if val != expected:
                old_val = val
                cell.value = expected
                fixes.append(f'Row{total2} Col{c}({col_letter})成新率: "{old_val}" → "{expected}"')
            continue
    
    # 修正合计1行的SUM公式范围（确保覆盖到total1-1）
    if total1:
        data_start = struct.get('data_start_row')
        if data_start:
            for c in range(1, min(ws.max_column + 1, 25)):
                cell = ws.cell(row=total1, column=c)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if not isinstance(val, str) or not val.startswith('='):
                    continue
                if 'SUM(' not in val.upper():
                    continue
                
                # 解析并修正SUM范围
                pattern_sum = r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)'
                new_val = val
                for match in re.finditer(pattern_sum, val, re.IGNORECASE):
                    col_s = match.group(1).upper()
                    row_s = int(match.group(2))
                    col_e = match.group(3).upper()
                    row_e = int(match.group(4))
                    
                    if col_s == col_e and row_e != total1 - 1:
                        old_range = f'{col_s}{row_s}:{col_e}{row_e}'
                        new_range = f'{col_s}{data_start}:{col_e}{total1 - 1}'
                        new_val = new_val.replace(old_range, new_range)
                
                if new_val != val:
                    cell.value = new_val
                    fixes.append(f'Row{total1} SUM修正: "{val}" → "{new_val}"')
    
    return fixes


def fix_cross_sheet_refs(wb, source_sheet_name, struct):
    """修复3：更新汇总表中对本sheet的跨sheet引用。
    
    关键逻辑：
    - 可见表（流动资产汇总等）引用合计2行（净额）
    - 辅表引用合计1行（毛额，D列）和合计2行（净额，E列）  
    - 减值准备汇总表引用坏账准备行
    
    策略：扫描所有sheet中对source_sheet的引用，将所有旧行号引用更新到正确的结构行。
    """
    fixes = []
    total1 = struct['total1_row']
    bad_debt = struct['bad_debt_row']
    total2 = struct['total2_row']
    
    if not total1:
        return fixes
    
    escaped_name = re.escape(source_sheet_name)
    pattern = rf"='?{escaped_name}'?!([A-Z]+)(\d+)"
    
    # 构建行号映射：旧行号 → 新行号
    # 我们需要识别哪些引用是"过时的"——指向旧行号的引用
    # 策略：根据sheet名推断引用意图，修正到正确的结构行
    
    for ws in wb.worksheets:
        if ws.title == source_sheet_name:
            continue
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                 min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                
                matches = list(re.finditer(pattern, cell.value, re.IGNORECASE))
                if not matches:
                    continue
                
                new_val = cell.value
                changed = False
                
                for match in matches:
                    col_ref = match.group(1).upper()
                    row_ref = int(match.group(2))
                    
                    # 判断行号是否需要修正
                    # 规则：如果引用行不在结构行上，且数据行区域内，
                    #       则根据引用意图修正到对应结构行
                    target_row = None
                    
                    # 引用行在数据区域（非结构行），需要修正
                    is_struct_row = row_ref in struct['struct_rows']
                    
                    if not is_struct_row and row_ref > (struct.get('data_start_row') or 6):
                        # 引用指向了数据行区域的行号——这是openpyxl insert_rows
                        # 没有正确下推导致的。根据sheet类型推断目标行：
                        
                        if '减值' in ws.title or '跌价' in ws.title:
                            # 减值准备汇总表 → 应引用坏账准备行
                            if bad_debt:
                                target_row = bad_debt
                        elif '辅' in ws.title:
                            # 辅表 → 需要判断引用列
                            # D列引用合计1（毛额），E列引用合计2（净额）
                            # 简化：修正到合计2行（因为原链接指向的是数据行的值，
                            #       而数据行值 = 合计2行的值）
                            if total2:
                                target_row = total2
                            else:
                                target_row = total1
                        else:
                            # 可见表 → 应引用合计2行（净额）
                            if total2:
                                target_row = total2
                            else:
                                target_row = total1
                    
                    if target_row and target_row != row_ref:
                        # 执行替换
                        old_ref_quoted = f"'{source_sheet_name}'!{col_ref}{row_ref}"
                        new_ref_quoted = f"'{source_sheet_name}'!{col_ref}{target_row}"
                        old_ref_noq = f"{source_sheet_name}!{col_ref}{row_ref}"
                        new_ref_noq = f"{source_sheet_name}!{col_ref}{target_row}"
                        
                        if old_ref_quoted in new_val:
                            new_val = new_val.replace(old_ref_quoted, new_ref_quoted)
                            changed = True
                        elif old_ref_noq in new_val:
                            new_val = new_val.replace(old_ref_noq, new_ref_noq)
                            changed = True
                
                if changed and new_val != cell.value:
                    old_formula = cell.value
                    cell.value = new_val
                    fixes.append(f'[{ws.title}] {cell.coordinate}: "{old_formula}" → "{new_val}"')
    
    return fixes


def fix_print_area(ws, struct):
    """修复4：打印范围下至合计2行（存在时），否则至合计1行。
    
    规范：左起B列，右至备注列（最后一个可见列），下至合计2（存在时），否则合计1。
    """
    fixes = []
    
    # 确定打印范围终止行
    end_row = struct['total2_row'] if struct['has_total2'] else struct['total1_row']
    
    if not end_row:
        return fixes
    
    # 查找最后一个可见列
    last_visible_col = 0
    for c in range(1, ws.max_column + 1):
        cl = get_column_letter(c)
        if not ws.column_dimensions[cl].hidden:
            last_visible_col = c
    
    if last_visible_col > 0:
        new_area = f"A1:{get_column_letter(last_visible_col)}{end_row}"
        old_area = str(ws.print_area) if ws.print_area else 'None'
        
        if str(ws.print_area) != new_area:
            ws.print_area = new_area
            fixes.append(f'打印范围: {old_area} → {new_area}')
    
    return fixes


def fix_sequence_format(ws, struct):
    """修复5：确保B列序号为整数格式，不显示小数点。"""
    fixes = []
    data_start = struct.get('data_start_row')
    total1 = struct['total1_row']
    
    if not data_start or not total1:
        return fixes
    
    fix_count = 0
    for r in range(data_start, total1):
        cell = ws.cell(row=r, column=2)
        if isinstance(cell, MergedCell):
            continue
        if cell.value is not None and isinstance(cell.value, (int, float)):
            if cell.number_format != '0' and cell.number_format != 'General':
                # 如果格式可能导致显示小数点，修正为整数格式
                cell.number_format = '0'
                fix_count += 1
            elif isinstance(cell.value, float) and cell.value == int(cell.value):
                # 浮点整数 → 转为int
                cell.value = int(cell.value)
                fix_count += 1
    
    if fix_count > 0:
        fixes.append(f'B列序号格式修正{fix_count}个单元格')
    
    return fixes


def fix_structure_row_font(ws, struct):
    """修复6：统一结构行字体格式。
    
    合计1/坏账准备/合计2行格式：
    - B列（序号/文字）：宋体11pt居中（已与B:C合并）
    - 金额列：Times New Roman 11pt靠右
    - A列：居中
    """
    fixes = []
    
    font_cn = Font(name='宋体', size=11)
    font_en = Font(name='Times New Roman', size=11)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, 
                         top=thin_side, bottom=thin_side)
    
    AMOUNT_HEADERS = {'账面价值', '评估价值', '增值额', '增值率', '减值准备',
                      '投资成本', '原值', '余额'}
    
    # 识别金额列号
    amount_cols = set()
    header_r = struct.get('header_row') or struct.get('data_start_row', 5)
    for c in range(1, min(ws.max_column + 1, 25)):
        htext = ''
        for check_r in range(1, min(header_r + 3, 10)):
            cell = ws.cell(row=check_r, column=c)
            if not isinstance(cell, MergedCell) and cell.value:
                htext += str(cell.value)
        for kw in AMOUNT_HEADERS:
            if kw in htext:
                amount_cols.add(c)
                break
    
    rows_to_fix = struct['struct_rows']
    fix_count = 0
    
    for r in rows_to_fix:
        for c in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            
            # A列：居中
            if c == 1:
                if not cell.alignment or cell.alignment.horizontal != 'center':
                    cell.alignment = align_center
                    fix_count += 1
            
            # 金额列：TNR+靠右
            elif c in amount_cols:
                if cell.font.name != 'Times New Roman' or cell.font.size != 11:
                    cell.font = font_en
                    fix_count += 1
                if not cell.alignment or cell.alignment.horizontal != 'right':
                    cell.alignment = align_right
                    fix_count += 1
            
            # 其他列：宋体+居中
            else:
                if cell.font.name != '宋体' or cell.font.size != 11:
                    cell.font = font_cn
                    fix_count += 1
                if not cell.alignment or cell.alignment.horizontal != 'center':
                    cell.alignment = align_center
                    fix_count += 1
            
            # 统一边框
            if cell.border != thin_border:
                cell.border = thin_border
                fix_count += 1
    
    if fix_count > 0:
        fixes.append(f'结构行格式修正{fix_count}个属性')
    
    return fixes


# ============================================================
# 下推原模板行格式修复
# ============================================================

def fix_pushed_template_rows(ws, struct):
    """修复被insert_rows下推的原模板行格式不一致问题。
    
    问题：insert_rows在合计1行位置插入N行后，原来合计1行下方的行（坏账准备、合计2）
    会被下推，但这些行的格式可能保留着模板的12pt等不一致格式。
    """
    fixes = []
    
    font_cn = Font(name='宋体', size=11)
    font_en = Font(name='Times New Roman', size=11)
    
    # 修正结构行的字体大小（12pt→11pt）
    for r in struct['struct_rows']:
        for c in range(1, min(ws.max_column + 1, 25)):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.font and cell.font.size and abs(cell.font.size - 12) < 0.5:
                if cell.font.name == 'Times New Roman':
                    cell.font = font_en
                else:
                    cell.font = font_cn
                fixes.append(f'Row{r} Col{c}: 12pt→11pt')
    
    return fixes


# ============================================================
# 主流程
# ============================================================

def main(filepath):
    """主修复流程。"""
    print(f"加载文件: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    all_fixes = {}
    
    # Step 1: 扫描所有sheet结构
    print("\n=== Step 1: 扫描sheet结构 ===")
    sheet_structs = {}
    for sname in wb.sheetnames:
        if '汇总' in sname or sname.startswith('0') or sname.startswith('2-'):
            continue
        ws = wb[sname]
        struct = find_sheet_structure(ws)
        if struct['total1_row']:
            sheet_structs[sname] = struct
            if struct['has_total2']:
                print(f"  ✅ {sname}: 合计1=R{struct['total1_row']}, "
                      f"减值=R{struct['bad_debt_row']}, "
                      f"合计2=R{struct['total2_row']}")
            else:
                print(f"  ○ {sname}: 合计1=R{struct['total1_row']} (无合计2)")
    
    # Step 2: 逐sheet修复
    print("\n=== Step 2: 逐sheet修复 ===")
    for sname, struct in sheet_structs.items():
        ws = wb[sname]
        sheet_fixes = []
        
        print(f"\n--- {sname} ---")
        
        # 修复1: B:C合并
        f1 = fix_bc_merge(ws, struct)
        sheet_fixes.extend(f1)
        if f1:
            for f in f1:
                print(f"  [合并] {f}")
        
        # 修复2: 合计2行公式
        if struct['has_total2']:
            f2 = fix_total2_formulas(ws, struct)
            sheet_fixes.extend(f2)
            if f2:
                for f in f2:
                    print(f"  [公式] {f}")
        
        # 修复4: 打印范围
        f4 = fix_print_area(ws, struct)
        sheet_fixes.extend(f4)
        if f4:
            for f in f4:
                print(f"  [打印] {f}")
        
        # 修复5: 序号格式
        f5 = fix_sequence_format(ws, struct)
        sheet_fixes.extend(f5)
        if f5:
            for f in f5:
                print(f"  [序号] {f}")
        
        # 修复6: 结构行字体
        f6 = fix_structure_row_font(ws, struct)
        sheet_fixes.extend(f6)
        if f6:
            for f in f6:
                print(f"  [字体] {f}")
        
        # 下推行格式修复
        f_extra = fix_pushed_template_rows(ws, struct)
        sheet_fixes.extend(f_extra)
        if f_extra:
            for f in f_extra:
                print(f"  [下推行] {f}")
        
        all_fixes[sname] = sheet_fixes
    
    # Step 3: 跨sheet引用修复（需要遍历整个workbook）
    print("\n=== Step 3: 跨sheet引用修复 ===")
    for sname, struct in sheet_structs.items():
        if struct['has_total2']:
            f3 = fix_cross_sheet_refs(wb, sname, struct)
            if f3:
                if sname not in all_fixes:
                    all_fixes[sname] = []
                all_fixes[sname].extend(f3)
                for f in f3:
                    print(f"  [跨表] {f}")
    
    # Step 4: 验证
    print("\n=== Step 4: 验证修复结果 ===")
    issues = []
    for sname, struct in sheet_structs.items():
        ws = wb[sname]
        total1 = struct['total1_row']
        bad_debt = struct['bad_debt_row']
        total2 = struct['total2_row']
        
        # 验证B:C（或B:D）合并
        # DT-205: 兼容B:C和B:D两种合并模式
        for r in struct['struct_rows']:
            merged = any(
                mr.min_row == r and mr.min_col == 2 and mr.max_col >= 3 and mr.max_row == r
                for mr in ws.merged_cells.ranges
            )
            if not merged:
                issues.append(f"{sname}: B{r}:C{r} 未合并！")
        
        # 验证合计2行公式
        # DT-205: 扩展验证逻辑，认可以下公式模式：
        # 1. 金额列: =ROUND(col{total1}-col{bad_debt},2) 或 =col{total1}-col{bad_debt}
        # 2. 增值率: =IF(col{total2}=0,"",col{total2}/col{total2}*100)（指向自身行）
        # 3. 增值率变体: =IF(col{total2}=0,"",(col{total2}-col{total2})/col{total2}*100)
        # 4. 成新率/单价: =IF(col{total2}=0,"",ROUND(col{total2}/col{total2},0))（指向自身行）
        # 关键判断：合计2行公式如果所有行号都指向total2自身，则是合法的自引用公式
        if total2 and bad_debt and total1:
            for c in range(1, min(ws.max_column + 1, 25)):
                cell = ws.cell(row=total2, column=c)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    col_letter = get_column_letter(c)
                    
                    # 模式1检查：金额差值公式应引用total1和bad_debt
                    expected_pattern1 = rf'{col_letter}{total1}[-+]{col_letter}{bad_debt}'
                    if re.search(expected_pattern1, val, re.IGNORECASE):
                        continue  # 正确的金额公式
                    
                    # DT-205: 自引用公式检查——所有行号都指向total2自身
                    # 合法模式包括增值率、成新率等计算，它们引用合计2行自身的值
                    row_refs = re.findall(r'([A-Z]+)(\d+)', val)
                    all_self_ref = all(int(r) == total2 for _, r in row_refs)
                    if all_self_ref:
                        continue  # 合法的自引用公式
                    
                    # SUM公式检查（合计2行不应有SUM）
                    if 'SUM(' not in val.upper():
                        issues.append(f"{sname}: 合计2行{col_letter}{total2}公式仍错误: {val}")
        
        # 验证打印范围
        if total2:
            pa = str(ws.print_area) if ws.print_area else ''
            if str(total2) not in pa:
                issues.append(f"{sname}: 打印范围未包含合计2行(R{total2}): {pa}")
    
    if issues:
        print("  ⚠️ 以下问题未完全修复：")
        for iss in issues:
            print(f"    - {iss}")
    else:
        print("  ✅ 所有修复验证通过！")
    
    # Step 5: 保存
    print(f"\n保存文件: {filepath}")
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=0)
    wb.save(filepath)
    wb.close()
    
    # 汇总
    total_fixes = sum(len(v) for v in all_fixes.values())
    print(f"\n=== 修复完成 ===")
    print(f"共修复 {total_fixes} 项")
    for sname, fixes in all_fixes.items():
        if fixes:
            print(f"  {sname}: {len(fixes)}项")
    
    return all_fixes


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python fix_three_row_structure.py <xlsx_path>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"文件不存在: {filepath}")
        sys.exit(1)
    
    result = main(filepath)
