"""sheet_col_finder.py — 明细表列位动态查找公共模块

不同sheet的列位差异极大（如账面价值从E到N，增值额从G到P），绝不能硬编码。
本模块通过扫描Row5/Row6表头文字，按语义关键词动态匹配列号。

设计原则：
  - 零配置：运行时从模板表头自动识别，无需预定义映射
  - 兼容所有模板：12列往来类、8列合同资产、18列固定资产、9列负债等
  - 处理合并单元格：部分sheet的Row5/Row6是跨行合并表头
  - 返回统一接口：所有调用方使用语义键名（如'账面价值'）获取列号

使用：
  from sheet_col_finder import find_header_cols, get_col, SheetColFinder
  cols = find_header_cols(ws)
  bv_col = cols.get('账面价值')  # 返回int列号或None
  bv_col = get_col(ws, '账面价值')  # 带缓存+兜底

版本：v1.0 (2026-05-25)
"""

from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


# ========== 语义关键词定义 ==========
# key=语义名(调用方使用), value=表头中需要匹配的关键词列表（优先级从高到低）
SEMANTIC_KEYWORDS = {
    # 核心金额列
    '账面价值': ['账面价值'],
    '评估价值': ['评估价值'],
    '增值额': ['增值额'],
    '增值率': ['增值率'],
    
    # 账龄
    '账龄': ['账龄'],
    
    # 日期列
    '发生日期': ['发生日期'],
    '购置日期': ['购置日期'],
    '启用日期': ['启用日期'],
    '到期日': ['到期日'],
    
    # 文本描述列
    '备注': ['备注'],
    '业务内容': ['业务内容', '结算内容'],
    '结算对象': ['结算对象', '户名', '项目及内容'],
    '序号': ['序号'],
    
    # 币种/外币
    '币种': ['币种'],
    '外币金额': ['外币金额'],
    '汇率': ['汇率'],
    
    # 减值/风险
    '减值准备': ['减值准备', '跌价准备', '坏账准备'],
    '预计风险': ['预计风险'],
    
    # 固定资产专有列
    '设备编号': ['设备编号'],
    '设备名称': ['设备名称'],
    '规格型号': ['规格型号'],
    '生产厂家': ['生产厂家'],
    '计量单位': ['计量单位', '单位'],
    '数量': ['数量'],
    '账面原值': ['账面原值'],
    '账面净值': ['账面净值'],
    '评估原值': ['评估原值'],
    '评估净值': ['评估净值'],
    '成新率': ['成新率'],
    '增值额原值': ['增值额', '原值'],  # 固定资产有两列增值额
    
    # 负债专有列
    '年利率': ['年利率'],
    '费用类别': ['费用类别'],
    '税费项目': ['税费项目', '税种'],
    '税率': ['税率'],
    '征税机关': ['征税机关'],
    
    # 银行存款专有
    '开户银行': ['开户银行', '放款银行'],
    '账号': ['账号'],
    '账户余额': ['账户余额'],
    '核实余额': ['核实余额'],
    '差异': ['差异'],
    
    # 长期借款专有
    '借款合同编号': ['借款合同编号'],
    '借款金额': ['借款金额'],
    '外币评估': ['外币评估'],
    '本币': ['本币'],
    
    # 无形资产专有
    '无形资产名称': ['名称'],
    
    # 长期待摊专有
    '摊销期限': ['摊销期限', '摊销期'],
}

# DT-153v3: 动态检测表头行（不再固定Row 5/6）
# 策略：扫描前15行，找到含"账面价值""评估价值""结算对象"等关键词的行即为表头行
def _detect_header_rows(ws, max_scan=15):
    """动态检测表头行号列表"""
    header_keywords = ['账面价值', '评估价值', '结算对象', '序号', '科目名称', '项目']
    header_rows = []
    for r in range(1, min(ws.max_row + 1, max_scan)):
        for c in range(1, min(ws.max_column + 1, 25)):
            val = ws.cell(row=r, column=c).value
            if val and any(kw in str(val) for kw in header_keywords):
                if r not in header_rows:
                    header_rows.append(r)
                break
        if len(header_rows) >= 2:
            break
    return header_rows if header_rows else [5, 6]  # fallback

# 最大扫描列数（防止读到模板右侧的辅助数据）
MAX_SCAN_COL = 25

# 表头行范围（v1.90-FOR AI模板：Row5=简单表头/Row6=数据起始，Row5-6=双行表头）
HEADER_ROWS = [5, 6]


def find_header_cols(ws):
    """动态查找表头列号映射。
    
    扫描Row5和Row6的表头文字，返回列语义→列号的映射。
    不同sheet中列号差异巨大（如3-1-1增值额=I，4-8-1增值额=O），
    绝不能硬编码列号。
    
    匹配规则：
    - 对每个语义，按关键词列表优先级依次匹配
    - 同一语义只取第一个匹配的列（避免重复）
    - 合并单元格只取左上角（MergedCell跳过）
    - 隐藏列不纳入结果
    
    Args:
        ws: openpyxl worksheet对象
        
    Returns:
        dict: {语义名: 列号(int)}，例如 {'账面价值': 10, '评估价值': 11, ...}
    """
    result = {}
    hidden_cols = set()
    
    # 收集隐藏列
    for c in range(1, MAX_SCAN_COL + 1):
        cl = get_column_letter(c)
        if ws.column_dimensions[cl].hidden:
            hidden_cols.add(c)
    
    # DT-153v3: 动态检测表头行
    header_rows = _detect_header_rows(ws)

    for c in range(1, min(ws.max_column + 1, MAX_SCAN_COL + 1)):
        if c in hidden_cols:
            continue
        for r in header_rows:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            h = cell.value
            if not h or not isinstance(h, str):
                continue
            h_text = str(h).strip()
            if not h_text:
                continue
            for semantic, keywords in SEMANTIC_KEYWORDS.items():
                if semantic in result:
                    continue  # 已匹配，跳过
                for kw in keywords:
                    if kw in h_text:
                        result[semantic] = c
                        break
    
    return result


def get_formula_cols(ws, header_cols=None):
    """获取含公式的列号列表（增值额+增值率+账龄）。
    
    用于smart_insert_row的sum_cols参数——这些列需要公式下拉但不参与SUM。
    
    Args:
        ws: openpyxl worksheet对象
        header_cols: 可选，已缓存的header_cols结果
        
    Returns:
        list[int]: 含公式的列号列表
    """
    if header_cols is None:
        header_cols = find_header_cols(ws)
    cols = []
    for key in ['增值额', '增值率', '账龄']:
        if key in header_cols:
            cols.append(header_cols[key])
    return cols


def get_amount_cols(ws, header_cols=None):
    """获取金额类列号列表（账面价值+评估价值）。
    
    用于减值准备行金额修正等场景。
    
    Args:
        ws: openpyxl worksheet对象
        header_cols: 可选，已缓存的header_cols结果
        
    Returns:
        list[int]: 金额列号列表
    """
    if header_cols is None:
        header_cols = find_header_cols(ws)
    cols = []
    for key in ['账面价值', '评估价值']:
        if key in header_cols:
            cols.append(header_cols[key])
    return cols


def get_col(ws, semantic, header_cols=None):
    """获取指定语义的列号，带友好错误提示。
    
    Args:
        ws: openpyxl worksheet对象
        semantic: 语义名（如'账面价值'）
        header_cols: 可选，已缓存的header_cols结果
        
    Returns:
        int or None: 列号
    """
    if header_cols is None:
        header_cols = find_header_cols(ws)
    return header_cols.get(semantic)


def find_data_start_row(ws):
    """DT-153v3: 动态查找数据起始行。

    扫描策略（按优先级）：
    1. 查找A列"检索表头"标记行
    2. 查找表头行（含"账面价值"等关键词）之后的第一个空行或数值行
    3. 兜底：6

    Args:
        ws: openpyxl worksheet对象

    Returns:
        int: 数据起始行
    """
    # 策略1: 查找"检索表头"标记
    for r in range(1, 15):
        a_val = ws.cell(row=r, column=1).value
        if a_val and '检索表头' in str(a_val):
            next_a = ws.cell(row=r + 1, column=1).value
            if next_a and '检索表头' in str(next_a):
                return r + 2
            return r + 1

    # 策略2: 查找表头行后数据起始
    header_rows = _detect_header_rows(ws)
    if header_rows:
        last_header = max(header_rows)
        for r in range(last_header + 1, min(last_header + 5, ws.max_row + 1)):
            # 跳过行次行
            val = ws.cell(row=r, column=1).value
            if val and ('行次' in str(val) or '行号' in str(val)):
                continue
            return r

    return 6  # 最终兜底


def find_last_print_col(ws):
    """找到打印范围的右边界列。
    
    优先级：备注列 > 最后一个有表头的可见列
    不硬编码N(14)，不同sheet的备注列位置不同。
    
    Args:
        ws: openpyxl worksheet对象
        
    Returns:
        int: 右边界列号
    """
    # 优先查找备注列
    header_cols = find_header_cols(ws)
    if '备注' in header_cols:
        return header_cols['备注']
    
    # 无备注列：取最后一个有表头且可见的列
    last_col = 0
    for c in range(1, min(ws.max_column + 1, MAX_SCAN_COL + 1)):
        cl = get_column_letter(c)
        if ws.column_dimensions[cl].hidden:
            continue
        for r in HEADER_ROWS:
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell) and cell.value:
                last_col = c
                break
    
    return last_col if last_col > 0 else None


def classify_sheet(prefix):
    """根据sheet前缀判断科目分类。
    
    Args:
        prefix: sheet前缀（如'3-5', '4-8-1', '5-7'）
        
    Returns:
        str: 'asset' | 'liability' | 'equity' | 'unknown'
    """
    if not prefix:
        return 'unknown'
    # 资产类：3-x, 4-x
    if prefix.startswith('3-') or prefix.startswith('4-'):
        return 'asset'
    # 负债类：5-x, 6-x
    if prefix.startswith('5-') or prefix.startswith('6-'):
        return 'liability'
    # 所有者权益：7-x
    if prefix.startswith('7-'):
        return 'equity'
    return 'unknown'


class SheetColFinder:
    """带缓存的列位查找器，避免同一worksheet重复扫描。
    
    用法：
        finder = SheetColFinder()
        cols = finder.get(ws)        # 首次扫描并缓存
        bv = cols.get('账面价值')    # 后续直接读缓存
        bv2 = finder.get_col(ws, '账面价值')  # 便捷方法
    """
    
    def __init__(self):
        self._cache = {}  # ws_title → header_cols dict
    
    def get(self, ws):
        """获取指定worksheet的列位映射（带缓存）。"""
        title = ws.title
        if title not in self._cache:
            self._cache[title] = find_header_cols(ws)
        return self._cache[title]
    
    def get_col(self, ws, semantic):
        """获取指定语义的列号（带缓存）。"""
        return self.get(ws).get(semantic)
    
    def get_formula_cols(self, ws):
        """获取公式列号列表（带缓存）。"""
        return get_formula_cols(ws, self.get(ws))
    
    def get_amount_cols(self, ws):
        """获取金额列号列表（带缓存）。"""
        return get_amount_cols(ws, self.get(ws))
    
    def get_last_print_col(self, ws):
        """获取打印右边界列号（不缓存，因为逻辑不同）。"""
        return find_last_print_col(ws)
    
    def invalidate(self, ws_title=None):
        """清除缓存。"""
        if ws_title:
            self._cache.pop(ws_title, None)
        else:
            self._cache.clear()
