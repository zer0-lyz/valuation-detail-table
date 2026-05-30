#!/usr/bin/env python3
"""
格式校验脚本 — 以原始模板为标准，逐格对比当前文件的格式完整性。

用法:
    python3 validate_format.py <当前文件.xlsx> <模板文件.xlsx>
    python3 validate_format.py <当前文件.xlsx> <模板文件.xlsx> --sheets 4-8-4,4-8-5  (限定Sheet)
    python3 validate_format.py <当前文件.xlsx> <模板文件.xlsx> --json            (JSON输出)
    python3 validate_format.py <当前文件.xlsx> <模板文件.xlsx> --fix              (自动修复)

校验内容:
  1. 数据行字体名称/大小/加粗 (对比模板R7)
  2. 四边边框 (thin)
  3. 数字格式 (金额列/日期列)
  4. 对齐方式 (水平/垂直)
  5. 序号连续性
  6. 首行序号是否为1
  7. 打印区域是否覆盖至合计行
  8. 合并单元格完整性 (合计行合并范围)
"""

import argparse, sys, json, re
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Sheet类型识别
def sheet_type(sn):
    if sn.startswith('4-8-'): return 'FA明细'
    if sn.startswith(('3-5','3-7','3-8-','5-5','5-10-','4-9-','4-13-')): return '往来科目'
    if '汇总' in sn: return '汇总表'
    return '其他'

def load_workbook_safe(path):
    try:
        return load_workbook(path)
    except Exception as e:
        print(f"❌ 无法打开文件: {path}")
        print(f"   {e}")
        sys.exit(1)

def find_data_range(ws):
    """Find data rows: from 合计1向上找"""
    data_start = 7  # 模板R5=表头, R6=子表头, R7=数据开始
    data_end = ws.max_row
    sum1 = bad = sum2 = None
    
    for r in range(1, ws.max_row + 1):
        v1 = str(ws.cell(r, 1).value or '')
        if '合计1' in v1: sum1 = r
        elif '坏账' in v1: bad = r
        elif '合计2' in v1: sum2 = r
    
    if sum1:
        data_end = sum1 - 1
    if sum2 and sum2 > (sum1 or 0):
        data_end = data_end  # keep sum1-1
    
    return data_start, data_end, sum1, bad, sum2

def get_row_type_label(ws, r):
    """Get label for what this row is"""
    v1 = str(ws.cell(r, 1).value or '')
    if '合计1' in v1: return '合计1'
    if '合计2' in v1: return '合计2'
    if '坏账' in v1: return '坏账准备'
    if '预计风险' in v1 or '预计损失' in v1: return '预计风险'
    return None

def check_sheet(wb_cur, wb_tpl, sn, options):
    """Check format of one sheet, return list of issues."""
    if sn not in wb_cur:
        return [f"Sheet '{sn}' 在当前文件中不存在"]
    if sn not in wb_tpl:
        return [f"Sheet '{sn}' 在模板文件中不存在 — 跳过模板对比"]
    
    ws = wb_cur[sn]
    ws_tpl = wb_tpl[sn]
    issues = []
    s_type = sheet_type(sn)
    
    data_start, data_end, sum1, bad, sum2 = find_data_range(ws)
    if data_end < data_start:
        return [f"数据行范围异常: R{data_start}-R{data_end}"]
    
    # ── 检查1: 首行序号 ──
    if options.check_seq:
        seq = ws.cell(data_start, 2).value
        if seq is None or seq == '':
            issues.append(f"[序号] R{data_start}(首行) 序号为空")
        elif s_type in ('FA明细', '往来科目'):
            try:
                if int(seq) != 1:
                    issues.append(f"[序号] R{data_start}(首行) 序号={seq}, 应为1")
            except:
                issues.append(f"[序号] R{data_start} 序号无法解析: {seq}")
    
    # ── 检查2: 序号连续性 ──
    if options.check_seq:
        last_seq = 0
        for r in range(data_start, min(data_end + 1, data_start + 500)):
            s = ws.cell(r, 2).value
            if s is not None:
                try:
                    seq_val = int(s)
                    if seq_val != last_seq + 1:
                        issues.append(f"[序号] R{r} 序号={seq_val}(不连续, 期望{last_seq+1})")
                    last_seq = seq_val
                except:
                    pass
    
    # ── 检查3: 字体 ──
    if options.check_font and ws_tpl:
        ref_row = min(data_start, ws_tpl.max_row)
        for c in options.cols:
            try:
                tc = ws_tpl.cell(ref_row, c)
                cc = ws.cell(data_start, c)
                
                # Font name
                tn = tc.font.name
                cn = cc.font.name
                if tn and cn and tn != cn:
                    issues.append(f"[字体] R{data_start}C{c}: font={cn}(期望{tn})")
                
                # Font size
                ts = tc.font.size
                cs = cc.font.size
                if ts and cs and abs(ts - cs) > 0.5:
                    issues.append(f"[字号] R{data_start}C{c}: size={cs}(期望{ts})")
                
                # Bold
                if tc.font.bold != cc.font.bold:
                    issues.append(f"[加粗] R{data_start}C{c}: bold={cc.font.bold}(期望{tc.font.bold})")
            except Exception as e:
                pass
    
    # ── 检查4: 边框 ──
    if options.check_border:
        for r in range(data_start, data_end + 1):
            for c in range(1, 15):  # Check A-N columns
                try:
                    cell = ws.cell(r, c)
                    b = cell.border
                    for side in ['top', 'bottom', 'left', 'right']:
                        s = getattr(b, side)
                        if not s or s.style != 'thin':
                            issues.append(f"[边框] R{r}C{c} {side}={s.style if s else 'None'}")
                            break
                except:
                    pass
            if len(issues) > options.max_issues:
                break
    
    # ── 检查5: 数字格式 (金额列/日期列) ──
    if options.check_fmt and ws_tpl:
        ref_row = min(data_start, ws_tpl.max_row)
        for c in options.amount_cols + options.date_cols:
            try:
                tfmt = ws_tpl.cell(ref_row, c).number_format
                cfmt = ws.cell(data_start, c).number_format
                
                if tfmt and cfmt and tfmt != cfmt:
                    # Check if both are essentially the same (e.g. currency formats)
                    if not options.fmt_loose or tfmt.replace('_', '') != cfmt.replace('_', ''):
                        if c in options.amount_cols:
                            issues.append(f"[格式] R{data_start}C{c}(金额): {cfmt}(期望{tfmt})")
                        else:
                            issues.append(f"[格式] R{data_start}C{c}(日期): {cfmt}(期望{tfmt})")
            except:
                pass
    
    # ── 检查6: 对齐 ──
    if options.check_align and ws_tpl:
        ref_row = min(data_start, ws_tpl.max_row)
        for c in options.cols:
            try:
                ta = ws_tpl.cell(ref_row, c).alignment
                ca = ws.cell(data_start, c).alignment
                if ta.horizontal != ca.horizontal:
                    issues.append(f"[对齐] R{data_start}C{c}: h={ca.horizontal}(期望{ta.horizontal})")
                if ta.vertical != ca.vertical:
                    issues.append(f"[对齐] R{data_start}C{c}: v={ca.vertical}(期望{ta.vertical})")
            except:
                pass
    
    # ── 检查7: 打印区域 ──
    if options.check_print:
        pa = ws.print_area
        if pa:
            m = re.search(r'\$(\d+)$', pa)
            if m:
                pa_row = int(m.group(1))
                target = sum2 or sum1 or data_end
                if pa_row < target:
                    issues.append(f"[打印] print_area仅到R{pa_row}, 应有R{target}")
        elif s_type in ('FA明细', '往来科目'):
            issues.append(f"[打印] 未设置打印区域")
    
    # ── 检查8: 合并单元格 ──
    if options.check_merge and ws_tpl:
        # Check 合计行 merged cells match template
        template_merges = {}
        for mc in ws_tpl.merged_cells.ranges:
            if mc.min_row >= (ws_tpl.max_row - 5):  # Last few rows
                for r in range(mc.min_row, mc.max_row + 1):
                    row_type = get_row_type_label(ws_tpl, r)
                    if row_type:
                        template_merges[row_type] = mc
        
        current_merges = {}
        for mc in ws.merged_cells.ranges:
            for r in range(mc.min_row, mc.max_row + 1):
                row_type = get_row_type_label(ws, r)
                if row_type:
                    current_merges[row_type] = mc
        
        for row_type, tm in template_merges.items():
            cm = current_merges.get(row_type)
            if not cm:
                issues.append(f"[合并] {row_type}行缺少合并单元格")
                continue
            # Compare column range only (rows naturally shift when data is added)
            t_cols2 = f'{chr(64+tm.min_col)}:{chr(64+tm.max_col)}'
            c_cols2 = f'{chr(64+cm.min_col)}:{chr(64+cm.max_col)}'
            if t_cols2 != c_cols2:
                issues.append(f"[合并] {row_type}行合并列应为{t_cols2}, 实际{c_cols2}")
    
    return issues

def fix_sheet(wb_cur, wb_tpl, sn):
    """Auto-fix format issues using template as reference."""
    if sn not in wb_cur or sn not in wb_tpl:
        return 0
    
    ws = wb_cur[sn]
    ws_tpl = wb_tpl[sn]
    fixes = 0
    
    data_start, data_end, sum1, bad, sum2 = find_data_range(ws)
    if data_end < data_start:
        return 0
    
    ref_row = min(data_start, ws_tpl.max_row)
    
    for r in range(data_start, data_end + 1):
        for c in range(1, 20):
            try:
                cell = ws.cell(r, c)
                tc = ws_tpl.cell(ref_row, c)
                
                # Copy font from template
                cell.font = Font(
                    name=tc.font.name or 'Times New Roman',
                    size=tc.font.size or 11,
                    bold=tc.font.bold or False,
                    italic=tc.font.italic or False,
                    color=tc.font.color,
                )
                
                # Copy border
                cell.border = THIN_BORDER
                
                # Copy number format
                cell.number_format = tc.number_format
                
                # Copy alignment
                cell.alignment = Alignment(
                    horizontal=tc.alignment.horizontal or 'center',
                    vertical=tc.alignment.vertical or 'center',
                    wrap_text=tc.alignment.wrap_text or False,
                )
                
                fixes += 1
            except:
                pass
    
    # Fix print area
    target = sum2 or sum1 or data_end
    ws.print_area = f"'{sn}'!$B$1:$S${target}"
    
    # Fix merged cells for summary rows
    for label, r in [('合计1', sum1), ('坏账准备', bad), ('合计2', sum2)]:
        if not r:
            continue
        # Find matching template merge
        for mc in ws_tpl.merged_cells.ranges:
            row_type = get_row_type_label(ws_tpl, mc.min_row)
            if row_type == label:
                # Remove any existing merges at this row
                for existing in list(ws.merged_cells.ranges):
                    if existing.min_row == r:
                        ws.unmerge_cells(str(existing))
                end_col = mc.max_col
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=end_col)
                break
    
    return fixes


def main():
    parser = argparse.ArgumentParser(description='评估明细表格式校验工具')
    parser.add_argument('current', help='当前文件 (.xlsx)')
    parser.add_argument('template', help='模板文件 (.xlsx)')
    parser.add_argument('--sheets', help='限定检查的Sheet (逗号分隔)')
    parser.add_argument('--json', action='store_true', help='JSON格式输出')
    parser.add_argument('--fix', action='store_true', help='自动修复格式问题')
    parser.add_argument('--max-issues', type=int, default=50, help='每Sheet最大问题数')
    parser.add_argument('--fmt-loose', action='store_true', default=True, help='宽松的数字格式对比')
    parser.add_argument('--cols', type=str, default='3,4,5,11,12',
                       help='检查的列(逗号分隔, 用字母或数字)')
    
    args = parser.parse_args()
    
    # Parse column spec
    col_specs = set()
    for s in args.cols.split(','):
        s = s.strip().lower()
        if s.startswith('c') and s[1:].isdigit():
            col_specs.add(int(s[1:]))
        elif s.isdigit():
            col_specs.add(int(s))
        elif len(s) == 1 and s.isalpha():
            col_specs.add(ord(s) - 96)
    
    check_options = argparse.Namespace(
        check_seq=True,
        check_font=True,
        check_border=True,
        check_fmt=True,
        check_align=True,
        check_print=True,
        check_merge=True,
        max_issues=args.max_issues,
        fmt_loose=args.fmt_loose,
        cols=list(col_specs),
        amount_cols=[11, 12, 13, 15],
        date_cols=[9, 10],
    )
    
    wb_cur = load_workbook_safe(args.current)
    wb_tpl = load_workbook_safe(args.template)
    
    # Determine which sheets to check
    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(',')]
    else:
        # Auto-detect: all sheets that exist in both files
        sheet_names = [s for s in wb_cur.sheetnames if s in wb_tpl]
    
    all_results = {}
    total_issues = 0
    
    if args.fix:
        print(f"🔧 自动修复模式: {args.current} (模板: {args.template})\n")
    
    for sn in sheet_names:
        if args.fix:
            fix_count = fix_sheet(wb_cur, wb_tpl, sn)
            status = f"✅ 修复{fix_count}个单元格" if fix_count > 0 else "⏭️ 无需修复"
            print(f"  {status}: {sn}")
            continue
        
        issues = check_sheet(wb_cur, wb_tpl, sn, check_options)
        all_results[sn] = issues
        total_issues += len(issues)
        
        if args.json:
            continue
        
        if issues:
            print(f"❌ {sn} ({len(issues)}个问题):")
            for iss in issues[:args.max_issues]:
                print(f"    {iss}")
            if len(issues) > args.max_issues:
                print(f"    ... 还有{len(issues) - args.max_issues}个问题未显示")
        else:
            print(f"✅ {sn}: 格式完整")
    
    if args.fix:
        wb_cur.save(args.current)
        print(f"\n💾 已保存: {args.current}")
        return
    
    if args.json:
        print(json.dumps(all_results, ensure_ascii=False, indent=2))
        return
    
    print(f"\n{'='*50}")
    if total_issues == 0:
        print(f"🏆 全部 {len(sheet_names)} 个Sheet格式完整!")
    else:
        print(f"📋 共 {len(sheet_names)} 个Sheet, {total_issues} 个问题")
        print(f"💡 使用 --fix 参数自动修复")


if __name__ == '__main__':
    main()
